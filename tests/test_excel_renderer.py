# tests/test_excel_renderer.py
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.monthly_platform.excel_renderer import render_bundle_to_excel
from bundle_factory import make_test_bundle


def test_render_produces_all_expected_sheets(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    assert "Summary" in wb.sheetnames
    assert "Pipeline Open FY26" in wb.sheetnames
    assert "Won Lost FY26" in wb.sheetnames
    assert "Pipeline Inspection" in wb.sheetnames
    assert "Activity Volume" in wb.sheetnames


def test_summary_is_first_sheet(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    assert wb.sheetnames[0] == "Summary"


def test_pipeline_headers_match(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    ws = wb["Pipeline Open FY26"]
    headers = [ws.cell(1, c).value for c in range(1, 23)]
    assert headers[0] == "Account"
    assert headers[1] == "Opportunity"
    assert headers[6] == "ARR Unweighted (EUR)"
    assert headers[7] == "ARR Weighted (EUR)"
    assert headers[21] == "Competitor"


def test_pipeline_data_row(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    ws = wb["Pipeline Open FY26"]
    assert ws.cell(2, 1).value == "Acme Corp"
    assert ws.cell(2, 7).value == 500000.0
    assert ws.cell(2, 20).value == "Yes"


def test_freeze_panes_set(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    ws = wb["Pipeline Open FY26"]
    assert ws.freeze_panes == "A2"


def test_eur_formatting_applied(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    ws = wb["Pipeline Open FY26"]
    assert ws.cell(2, 7).number_format == "#,##0"


def test_empty_datasets_produce_no_rows(tmp_path):
    bundle = make_test_bundle()
    out = tmp_path / "test.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))
    ws = wb["Renewals FY26"]
    assert ws.cell(1, 1).value is not None  # headers exist
    assert ws.cell(2, 1).value is None  # no data rows


def test_workbook_parity_with_legacy_contract(tmp_path):
    """Verify rendered workbook matches the legacy sheet/column contract."""
    import dataclasses

    from scripts.monthly_platform.models import (
        ApprovalDeal,
        CommitItem,
        MovementEvent,
        RenewalDeal,
        StageEvent,
        WonLostDeal,
    )

    bundle = make_test_bundle()
    bundle = dataclasses.replace(
        bundle,
        datasets=dataclasses.replace(
            bundle.datasets,
            won_lost=[
                WonLostDeal(
                    account="Won Co",
                    opportunity="Won Deal",
                    owner="Jane",
                    stage="8 - Won",
                    close_date="2026-03-01",
                    arr_unweighted=300000,
                    deal_type="Land",
                    industry="Insurance",
                    sales_region="APAC",
                    reason_won_lost="",
                    competitor="",
                    created_date="2025-10-01",
                    currency="EUR",
                    age_days=152,
                    quarter="Q1 2026",
                ),
            ],
            renewals=[
                RenewalDeal(
                    account="R Co",
                    opportunity="Renewal",
                    owner="Jane",
                    stage="5 - Preferred",
                    close_date="2026-06-01",
                    acv_unweighted=125000,
                    deal_type="Renewal",
                    quarter="Q2 2026",
                    probability=80,
                    comments="",
                ),
            ],
            approvals=[
                ApprovalDeal(
                    account="Acme Corp",
                    opportunity="Big Deal",
                    owner="Jane",
                    stage="4 - Shortlisted",
                    close_date="2026-06-30",
                    arr_unweighted=500000,
                    status="Approved 2026",
                    approval_date="2026-03-01",
                    next_step="Demo",
                    quarter="Q2 2026",
                    lead_scope="Core",
                ),
            ],
            commit_items=[
                CommitItem(
                    account="Acme Corp",
                    opportunity="Big Deal",
                    owner="Jane",
                    forecast_category="Commit",
                    arr_weighted=250000,
                    arr_unweighted=500000,
                    close_date="2026-06-30",
                    period="Q2 2026",
                    stage="4 - Shortlisted",
                ),
            ],
            movement_prior=[
                MovementEvent(
                    account="Acme Corp",
                    opportunity="Big Deal",
                    owner="Jane",
                    stage="4 - Shortlisted",
                    movement_type="Q1 Slipped",
                    old_close="2026-03-15",
                    new_close="2026-06-30",
                    changed_on="2026-03-20",
                    arr_unweighted=500000,
                ),
            ],
            stage_events=[
                StageEvent(
                    opportunity_id="006x",
                    opportunity="Big Deal",
                    account="Acme Corp",
                    owner="Jane",
                    current_stage="4 - Shortlisted",
                    old_value="3 - Engagement",
                    new_value="4 - Shortlisted",
                    created_date="2026-03-10",
                    arr_unweighted=500000,
                    is_closed=False,
                    is_won=False,
                ),
            ],
        ),
        dataset_counts={
            "pipeline_open": 1,
            "won_lost": 1,
            "renewals": 1,
            "approvals": 1,
            "pi_current": 1,
            "pi_forward": 0,
            "activity": 1,
            "commit_items": 1,
            "stage_events": 1,
            "forecast_category_events": 0,
            "close_date_events": 0,
            "movement_prior": 1,
            "movement_current": 0,
            "snapshot_trend": 0,
        },
    )

    out = tmp_path / "parity.xlsx"
    render_bundle_to_excel(bundle, out)
    wb = load_workbook(str(out))

    expected_sheets = [
        "Summary",
        "Pipeline Open FY26",
        "Won Lost FY26",
        "Commercial Approval",
        "Renewals FY26",
        "Pipeline Inspection",
        "Activity Volume",
        "Commit Items",
        "Q1 Movement",
        "Q2 Movement",
        "Stage History",
        "Forecast Category History",
    ]
    for name in expected_sheets:
        assert name in wb.sheetnames, f"Missing sheet: {name}"

    legacy_headers = {
        "Pipeline Open FY26": [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Forecast Category",
            "Close Date",
            "ARR Unweighted (EUR)",
            "ARR Weighted (EUR)",
            "Probability %",
            "Push Count",
            "Type",
            "Lead Scope",
            "Industry",
            "Tier",
            "Sales Region",
            "Created",
            "Last Activity",
            "Next Step",
            "Last Modified",
            "Approved",
            "Approval Date",
            "Competitor",
        ],
        "Won Lost FY26": [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            "ARR Unweighted (EUR)",
            "Type",
            "Reason",
            "Lost To Competitor",
            "Industry",
            "Sales Region",
            "Created",
        ],
        "Commercial Approval": [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            "ARR Unweighted (EUR)",
            "Status",
            "Approval Date",
            "Next Step",
            "Lead Scope",
        ],
        "Renewals FY26": [
            "Close Date",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "ACV Unweighted (EUR)",
            "Probability %",
            "Comments",
        ],
        "Pipeline Inspection": [
            "Opportunity",
            "Owner",
            "Stage",
            "Forecast Category",
            "ARR Weighted (native ccy)",
            "Currency",
            "Close Date",
            "Push Count",
            "Score",
            "Priority",
        ],
        "Activity Volume": [
            "Account",
            "Opportunity",
            "Owner",
            "Tasks 90d",
            "Events 90d",
            "Total Touches 90d",
            "Last Activity",
            "Flag",
        ],
        "Commit Items": [
            "Account",
            "Opportunity",
            "Owner",
            "Forecast Category",
            "Forecast ARR Wtd (EUR)",
            "ARR Unwtd (EUR)",
            "Close Date",
            "Period",
            "Stage",
        ],
        "Q1 Movement": [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Movement",
            "Old Close",
            "New Close",
            "Changed On",
            "ARR Unweighted (EUR)",
        ],
        "Stage History": [
            "Account",
            "Opportunity",
            "Owner",
            "Stage (live)",
            "From Stage",
            "To Stage",
            "Changed On",
            "ARR Unweighted (EUR)",
        ],
    }

    for sheet_name, expected_hdrs in legacy_headers.items():
        ws = wb[sheet_name]
        actual = [ws.cell(1, c).value for c in range(1, len(expected_hdrs) + 1)]
        assert actual == expected_hdrs, f"{sheet_name} headers mismatch: {actual}"

    ws = wb["Won Lost FY26"]
    assert ws.cell(2, 1).value == "Won Co"
    assert ws.cell(2, 6).value == 300000

    ws = wb["Renewals FY26"]
    assert ws.cell(2, 6).value == 125000

    ws = wb["Q1 Movement"]
    assert ws.cell(2, 5).value == "Q1 Slipped"
    assert ws.cell(2, 9).value == 500000

    ws = wb["Stage History"]
    assert ws.cell(2, 5).value == "3 - Engagement"
    assert ws.cell(2, 6).value == "4 - Shortlisted"

    ws = wb["Summary"]
    assert ws.cell(1, 1).value == "Jesper Tyrer (APAC)"
    kpi_labels = [ws.cell(r, 1).value for r in range(7, 25) if ws.cell(r, 1).value]
    assert "Open Pipeline Unweighted (stages 1-6)" in kpi_labels
    assert "Won Deal Count" in kpi_labels
