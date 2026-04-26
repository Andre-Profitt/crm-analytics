from pathlib import Path

from openpyxl import Workbook

from scripts import build_deck_from_excel as deck_builder
from scripts.build_deck_from_excel import (
    _is_pending_approval_status as deck_pending_status,
    _resolve_runtime_period_context,
)
from scripts.monthly_platform.policy import (
    is_active_forecast_category,
    summarize_approval_rows,
)
from scripts.validate_tie_out import excel_metrics, resolve_runtime_scopes


def _append_sheet_row(ws, headers, values):
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for idx, header in enumerate(headers, start=1):
            ws.cell(1, idx).value = header
    ws.append(values)


def test_excel_metrics_excludes_omitted_and_counts_pending_approval(tmp_path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    pipeline = wb.create_sheet("Pipeline Open FY26")
    pipeline_headers = [
        "Type",
        "Close Date",
        "Forecast Category",
        "ARR Unweighted (EUR)",
    ]
    _append_sheet_row(
        pipeline,
        pipeline_headers,
        ["Land", "2026-04-15", "Pipeline", 1_500_000],
    )
    _append_sheet_row(
        pipeline,
        pipeline_headers,
        ["Land", "2026-05-01", "Omitted", 500_000],
    )
    _append_sheet_row(
        pipeline,
        pipeline_headers,
        ["Land", "2026-05-10", "", 200_000],
    )
    _append_sheet_row(
        pipeline,
        pipeline_headers,
        ["Renewal", "2026-05-01", "Pipeline", 900_000],
    )

    won_lost = wb.create_sheet("Won Lost FY26")
    won_lost_headers = ["Type", "Stage", "Close Date", "ARR Unweighted (EUR)"]
    _append_sheet_row(
        won_lost,
        won_lost_headers,
        ["Land", "8 - Won", "2026-03-15", 750_000],
    )
    _append_sheet_row(
        won_lost,
        won_lost_headers,
        ["Land", "0 - Lost", "2026-02-10", 250_000],
    )

    approvals = wb.create_sheet("Commercial Approval")
    approval_headers = ["Status", "Close Date"]
    _append_sheet_row(approvals, approval_headers, ["Approved 2026", "2026-04-20"])
    _append_sheet_row(approvals, approval_headers, ["Pending Approval", "2026-04-20"])
    _append_sheet_row(approvals, approval_headers, ["Missing Stage 3+", "2026-04-20"])

    renewals = wb.create_sheet("Renewals FY26")
    renewal_headers = ["Close Date", "ACV Unweighted (EUR)"]
    _append_sheet_row(renewals, renewal_headers, ["2026-05-01", 100_000])

    path = tmp_path / "sample.xlsx"
    wb.save(path)

    metrics = excel_metrics(path)

    assert metrics["open_land_deals"] == 1
    assert metrics["open_land_arr"] == 1_500_000
    assert metrics["q1_land_wins"] == 1
    assert metrics["q1_land_lost"] == 1
    assert metrics["approved_2026"] == 1
    assert metrics["conditionally_approved"] == 1
    assert metrics["missing_stage3"] == 1


def test_pending_approval_helper_matches_legacy_pending_labels() -> None:
    assert deck_pending_status("Pending Approval") is True
    assert deck_pending_status("Conditionally approved (pending)") is True
    assert deck_pending_status("Approved 2026") is False


def test_legacy_builder_infers_report_date_from_workbook_folder() -> None:
    period = _resolve_runtime_period_context(
        workbook_path=Path("/tmp/director_live_workbooks/2026-04-22/jesper-tyrer.xlsx")
    )

    assert period["run_date"] == "2026-04-22"
    assert period["prior"]["label"] == "Q1"
    assert period["current"]["label"] == "Q2"
    assert period["forward"]["label"] == "Q3"


def test_legacy_builder_historical_trending_contract_preserves_prior_year() -> None:
    previous = dict(deck_builder.FQ)
    deck_builder.FQ = deck_builder._resolve_runtime_period_context(as_of_date="2027-01-15")
    try:
        contract = deck_builder._historical_trending_contract()
    finally:
        deck_builder.FQ = previous

    assert contract.retrospective_title == "Q4 2026"
    assert contract.current_title == "Q1 2027"


def test_excel_metrics_follow_explicit_runtime_scopes(tmp_path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    pipeline = wb.create_sheet("Pipeline Open FY26")
    _append_sheet_row(
        pipeline,
        ["Type", "Close Date", "Forecast Category", "ARR Unweighted (EUR)"],
        ["Land", "2026-10-20", "Pipeline", 2_000_000],
    )

    won_lost = wb.create_sheet("Won Lost FY26")
    won_lost_headers = ["Type", "Stage", "Close Date", "ARR Unweighted (EUR)"]
    _append_sheet_row(
        won_lost,
        won_lost_headers,
        ["Land", "8 - Won", "2026-08-15", 700_000],
    )
    _append_sheet_row(
        won_lost,
        won_lost_headers,
        ["Land", "0 - Lost", "2026-09-10", 300_000],
    )

    approvals = wb.create_sheet("Commercial Approval")
    _append_sheet_row(approvals, ["Status", "Close Date"], ["Approved 2026", "2026-10-20"])

    renewals = wb.create_sheet("Renewals FY26")
    _append_sheet_row(
        renewals,
        ["Close Date", "ACV Unweighted (EUR)"],
        ["2026-11-01", 125_000],
    )

    path = tmp_path / "q4-sample.xlsx"
    wb.save(path)

    scopes = resolve_runtime_scopes("2026-10-15")
    metrics = excel_metrics(path, scopes=scopes)

    assert metrics["open_land_deals"] == 1
    assert metrics["q1_land_wins"] == 1
    assert metrics["q1_land_lost"] == 1
    assert metrics["q2_renewals"] == 1
    assert metrics["q2_renewals_acv"] == 125_000


def test_shared_policy_excludes_blank_forecast_category_and_summarizes_approvals() -> None:
    assert is_active_forecast_category("Pipeline") is True
    assert is_active_forecast_category("Omitted") is False
    assert is_active_forecast_category("") is False

    summary = summarize_approval_rows(
        [
            {"Status": "Approved 2026"},
            {"Status": "Pending Approval"},
            {"Status": "Missing (Stage 3+)"},
        ]
    )
    assert summary == {
        "approved_2026": 1,
        "conditionally_approved": 1,
        "missing_stage3": 1,
    }
