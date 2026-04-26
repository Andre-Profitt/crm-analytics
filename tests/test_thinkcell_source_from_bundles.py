from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from scripts.monthly_platform.thinkcell_source import (
    build_thinkcell_source_from_manifest,
)

REPO_ROOT = Path(__file__).resolve().parents[1]


def test_build_thinkcell_source_from_source_backed_manifest(tmp_path: Path) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_dir = tmp_path / "thinkcell"

    result = build_thinkcell_source_from_manifest(
        manifest_path=manifest_path,
        output_dir=output_dir,
        template_path="template.pptx",
    )

    assert result["summary"] == {
        "director_count": 1,
        "source_backed_dataset_count": 4,
        "source_backed_row_count": 10,
        "coverage_row_count": 6,
        "metric_count": 4,
    }

    workbook = load_workbook(result["workbook_path"], data_only=True)
    assert workbook.sheetnames == [
        "Control",
        "Source Counts",
        "Coverage",
        "Metric Store",
        "PPTTC Map",
    ]
    assert "TC_Control" in workbook["Control"].tables
    assert "TC_SourceCounts" in workbook["Source Counts"].tables
    assert "TC_Coverage" in workbook["Coverage"].tables
    assert "TC_MetricStore" in workbook["Metric Store"].tables
    assert "TC_PpttcMap" in workbook["PPTTC Map"].tables

    count_rows = _worksheet_dicts(workbook["Source Counts"])
    assert [row["dataset"] for row in count_rows] == [
        "pi_current",
        "pi_forward",
        "pipeline_open",
        "snapshot_trend",
    ]
    assert {
        row["metric_id"] for row in count_rows
    } == {
        "source_backed.apac.jesper-tyrer.pi_current.row_count",
        "source_backed.apac.jesper-tyrer.pi_forward.row_count",
        "source_backed.apac.jesper-tyrer.pipeline_open.row_count",
        "source_backed.apac.jesper-tyrer.snapshot_trend.row_count",
    }
    assert {row["dataset"] for row in count_rows}.isdisjoint({"won_lost", "activity"})

    coverage_rows = _worksheet_dicts(workbook["Coverage"])
    won_lost = next(row for row in coverage_rows if row["dataset"] == "won_lost")
    assert won_lost["policy"] == "optional_empty"
    assert won_lost["covered"] is True
    pipeline_open = next(
        row for row in coverage_rows if row["dataset"] == "pipeline_open"
    )
    assert pipeline_open["policy"] == "source_backed"
    assert pipeline_open["required_for_publish"] is True
    assert pipeline_open["source_contract_present"] is True

    metric_rows = _worksheet_dicts(workbook["Metric Store"])
    assert metric_rows[0]["scope"] == "source_backed_director_bundle"
    assert metric_rows[0]["unit"] == "rows"

    ppttc = json.loads(Path(result["ppttc_path"]).read_text(encoding="utf-8"))
    assert ppttc[0]["template"] == "template.pptx"
    assert ppttc[0]["metadata"]["requires_think_cell_to_generate"] is False
    data_by_name = {entry["name"]: entry for entry in ppttc[0]["data"]}
    assert data_by_name["ThinkCellSourceStatus"]["table"] == [[{"string": "ok"}]]
    metric_table = data_by_name["MetricStoreTable"]["table"]
    assert [cell["string"] for cell in metric_table[0]][:4] == [
        "metric_id",
        "scope",
        "label",
        "value",
    ]
    assert [
        row[0]["string"] for row in metric_table[1:]
    ] == [
        "source_backed.apac.jesper-tyrer.pi_current.row_count",
        "source_backed.apac.jesper-tyrer.pi_forward.row_count",
        "source_backed.apac.jesper-tyrer.pipeline_open.row_count",
        "source_backed.apac.jesper-tyrer.snapshot_trend.row_count",
    ]


def test_thinkcell_source_cli_accepts_manifest(tmp_path: Path) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_dir = tmp_path / "cli-output"

    completed = subprocess.run(
        [
            sys.executable,
            str(REPO_ROOT / "scripts/build_thinkcell_source_from_bundles.py"),
            "--manifest",
            str(manifest_path),
            "--output-dir",
            str(output_dir),
            "--template",
            "template.pptx",
            "--json",
        ],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )

    payload = json.loads(completed.stdout)
    assert payload["status"] == "ok"
    assert payload["summary"]["metric_count"] == 4
    assert Path(payload["workbook_path"]).exists()
    assert Path(payload["ppttc_path"]).exists()


def _write_manifest(tmp_path: Path) -> Path:
    bundle_path = tmp_path / "apac.json"
    bundle_path.write_text(json.dumps(_bundle_payload()) + "\n", encoding="utf-8")
    manifest = {
        "schema_version": "monthly_platform.director_bundle_manifest.v1",
        "snapshot_date": "2026-04-30",
        "source_run_id": "source-run",
        "status": "ok",
        "source_bundle_manifest_path": "source_bundle_manifest.json",
        "output_dir": str(tmp_path),
        "bundle_paths": [str(bundle_path)],
        "summary": {
            "bundle_count": 1,
            "source_bundle_count": 1,
            "directors": [
                {
                    "director": "Jesper Tyrer",
                    "territory": "APAC",
                    "dataset_counts": _dataset_counts(),
                    "dataset_coverage": {
                        "schema_version": "monthly_platform.director_bundle_contract.v1",
                        "source_backed": [
                            "pi_current",
                            "pi_forward",
                            "pipeline_open",
                            "snapshot_trend",
                        ],
                        "optional_empty": ["activity", "won_lost"],
                        "publish_required": [
                            "pi_current",
                            "pi_forward",
                            "pipeline_open",
                            "snapshot_trend",
                        ],
                        "dataset_counts": _dataset_counts(),
                        "source_requirement_ids": [
                            "sd_historical_trending",
                            "sd_pipeline_inspection",
                            "sd_pipeline_open",
                        ],
                    },
                }
            ],
        },
        "findings": [],
    }
    manifest_path = tmp_path / "director_bundle_manifest.json"
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    return manifest_path


def _bundle_payload() -> dict:
    return {
        "schema_version": "1",
        "snapshot_date": "2026-04-30",
        "director": "Jesper Tyrer",
        "territory": "APAC",
        "corp_ccy": "EUR",
        "extract_timestamp": "2026-04-30T09:30:00Z",
        "source_contract": {
            "sf_org": "simcorp.my.salesforce.com",
            "api_version": "v66.0",
            "territory_soql_where": "source_bundle:APAC",
            "extract_timestamp": "2026-04-30T09:30:00Z",
            "sources": {
                "source_bundle": {
                    "source_type": "source_bundle",
                    "source_id": "source-run",
                    "query_label": "APAC:source_bundle",
                    "row_count": 10,
                    "duration_ms": 0,
                },
                "pipeline_open": {
                    "source_type": "salesforce_list_view",
                    "source_id": "src-pipeline-open",
                    "query_label": "APAC:pipeline_open",
                    "row_count": 2,
                    "duration_ms": 0,
                },
                "pi_current": {
                    "source_type": "salesforce_list_view",
                    "source_id": "src-pi-current",
                    "query_label": "APAC:pi_current",
                    "row_count": 3,
                    "duration_ms": 0,
                },
                "pi_forward": {
                    "source_type": "salesforce_list_view",
                    "source_id": "src-pi-forward",
                    "query_label": "APAC:pi_forward",
                    "row_count": 1,
                    "duration_ms": 0,
                },
                "snapshot_trend": {
                    "source_type": "salesforce_report",
                    "source_id": "src-snapshot-trend",
                    "query_label": "APAC:historical_trending",
                    "row_count": 4,
                    "duration_ms": 0,
                },
            },
        },
        "dataset_counts": _dataset_counts(),
        "datasets": {},
    }


def _dataset_counts() -> dict[str, int]:
    return {
        "pipeline_open": 2,
        "won_lost": 0,
        "pi_current": 3,
        "pi_forward": 1,
        "activity": 0,
        "snapshot_trend": 4,
    }


def _worksheet_dicts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]
