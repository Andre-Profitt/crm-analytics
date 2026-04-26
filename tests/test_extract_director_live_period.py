import json
from pathlib import Path

from openpyxl import load_workbook

import scripts.extract_director_live as extract_live
from scripts.extract_director_live import (
    _build_pipeline_inspection_rows,
    _filter_pi_records_to_territory_scope,
    _load_forward_quarter_pi_audit_fallback,
    _quarter_label,
    _resolve_forward_quarter_pi_source,
    _runtime_period,
    _write_run_audit,
)


def _make_opportunity(
    *,
    opp_id: str,
    name: str,
    stage: str,
    close_date: str,
    approval_status: str,
    approved: bool = False,
    approval_date: str = "",
    submitted: bool = False,
    submitted_date: str = "",
) -> dict:
    return {
        "Id": opp_id,
        "Name": name,
        "Type": "Land",
        "StageName": stage,
        "CloseDate": close_date,
        "Approval_Status__c": approval_status,
        "Stage_20_Approval__c": approved,
        "Stage_20_Approval_Date__c": approval_date,
        "Submit_for_Stage_20_Review__c": submitted,
        "Submit_for_Stage_20_Review_Date__c": submitted_date,
        "APTS_Opportunity_ARR__c": 100000,
        "APTS_Forecast_ARR__c": 50000,
        "ForecastCategoryName": "Commit",
        "Probability": 50,
        "PushCount": 0,
        "Lead_Scope__c": "Core",
        "Sales_Region__c": "APAC",
        "CreatedDate": "2026-01-01T00:00:00.000+0000",
        "LastActivityDate": "2026-04-01",
        "LastModifiedDate": "2026-04-10T00:00:00.000+0000",
        "NextStep": "Advance",
        "Lost_to_Competitor__c": "",
        "Account": {
            "Name": "Acme",
            "Industry": "Asset Management",
            "Tier_Calculation__c": "Tier 1",
        },
        "Owner": {"Name": "Jesper Tyrer"},
    }


def test_runtime_period_uses_snapshot_year_for_fiscal_filters() -> None:
    period = _runtime_period("2027-02-10")

    assert period["analysis_year"] == 2027
    assert period["fy_label"] == "FY27"
    assert (
        period["fy_close_filter"]
        == "AND CloseDate >= 2027-01-01 AND CloseDate <= 2027-12-31"
    )
    assert period["q1_start"] == "2027-01-01"
    assert period["q2_end"] == "2027-06-30"
    assert period["q3_start"] == "2027-07-01"
    assert period["forward_quarter_label"] == "Q2"
    assert period["forward_start"] == "2027-04-01"
    assert period["forward_end"] == "2027-06-30"


def test_quarter_label_is_dynamic_for_reporting_year() -> None:
    assert _quarter_label("2027-05-11", 2027) == "Q2 2027"
    assert _quarter_label("2027-11-01", 2027) == "Q4 2027"
    assert _quarter_label("2026-05-11", 2027) == ""


def test_resolve_forward_quarter_pi_source_uses_runtime_quarter() -> None:
    source = _resolve_forward_quarter_pi_source(
        {
            "forward_quarter_pi_list_views": {
                "Q3": {
                    "list_view_id": "00BTb00000LEdNBMA1",
                    "list_view_label": "PI ARR Forecast APAC Q3 2026 Land",
                }
            }
        },
        _runtime_period("2026-04-22"),
    )

    assert source == {
        "list_view_id": "00BTb00000LEdNBMA1",
        "list_view_label": "PI ARR Forecast APAC Q3 2026 Land",
        "quarter_label": "Q3",
        "quarter_title": "Q3 2026",
        "start_date": "2026-07-01",
        "end_date": "2026-09-30",
    }


def test_resolve_forward_quarter_pi_source_can_use_audit_fallback() -> None:
    source = _resolve_forward_quarter_pi_source(
        {"forward_quarter_pi_list_views": {}},
        _runtime_period("2026-08-10"),
        audit_fallback={
            "list_view_id": "00BTb00000Q4APAC",
            "list_view_label": "PI ARR Forecast APAC Q4 2026 Land",
        },
    )

    assert source == {
        "list_view_id": "00BTb00000Q4APAC",
        "list_view_label": "PI ARR Forecast APAC Q4 2026 Land",
        "quarter_label": "Q4",
        "quarter_title": "Q4 2026",
        "start_date": "2026-10-01",
        "end_date": "2026-12-31",
    }


def test_load_forward_quarter_pi_audit_fallback_reads_ok_candidate_entries(
    tmp_path: Path, monkeypatch
) -> None:
    audit_root = tmp_path / "output" / "source_contract_audit" / "2026-08-10"
    audit_root.mkdir(parents=True, exist_ok=True)
    (audit_root / "source_contract_audit.json").write_text(
        json.dumps(
            {
                "candidate_forward_quarter": {
                    "quarter_label": "Q4",
                    "pi_list_views": [
                        {
                            "territory": "APAC",
                            "list_view_id": "00BTb00000Q4APAC",
                            "list_view_label": "PI ARR Forecast APAC Q4 2026 Land",
                            "status": "ok",
                        },
                        {
                            "territory": "Canada",
                            "list_view_id": "00BTb00000Q4CA",
                            "list_view_label": "PI ARR Forecast Canada Q4 2026 Land",
                            "status": "failed",
                        },
                    ],
                }
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        extract_live,
        "SOURCE_CONTRACT_AUDIT_ROOT",
        tmp_path / "output" / "source_contract_audit",
    )

    assert _load_forward_quarter_pi_audit_fallback(
        "2026-08-10",
        quarter_label="Q4",
    ) == {
        "APAC": {
            "list_view_id": "00BTb00000Q4APAC",
            "list_view_label": "PI ARR Forecast APAC Q4 2026 Land",
        }
    }


def test_build_pipeline_inspection_rows_can_filter_to_forward_quarter() -> None:
    rows = _build_pipeline_inspection_rows(
        [
            {
                "fields": {
                    "Name": {"value": "APAC Q2 Land"},
                    "Owner": {"value": {"fields": {"Name": {"value": "Jesper Tyrer"}}}},
                    "StageName": {"value": "4 - Shortlisted"},
                    "ForecastCategoryName": {"value": "Commit"},
                    "APTS_Forecast_ARR__c": {"value": 100000},
                    "CloseDate": {"value": "2026-06-20"},
                    "PushCount": {"value": 1},
                    "OpportunityScore": {"value": {"fields": {"Score": {"value": 55}}}},
                    "IsPriorityRecord": {"value": False},
                    "IsClosed": {"value": False},
                    "Type": {"value": "Land"},
                }
            },
            {
                "fields": {
                    "Name": {"value": "APAC Q3 Land"},
                    "Owner": {"value": {"fields": {"Name": {"value": "Jesper Tyrer"}}}},
                    "StageName": {"value": "5 - Preferred"},
                    "ForecastCategoryName": {"value": "Commit"},
                    "APTS_Forecast_ARR__c": {"value": 250000},
                    "CloseDate": {"value": "2026-07-18"},
                    "PushCount": {"value": 2},
                    "OpportunityScore": {"value": {"fields": {"Score": {"value": 71}}}},
                    "IsPriorityRecord": {"value": True},
                    "IsClosed": {"value": False},
                    "Type": {"value": "Land"},
                }
            },
            {
                "fields": {
                    "Name": {"value": "APAC Q4 Land"},
                    "Owner": {"value": {"fields": {"Name": {"value": "Jesper Tyrer"}}}},
                    "StageName": {"value": "3 - Engagement"},
                    "ForecastCategoryName": {"value": "Best Case"},
                    "APTS_Forecast_ARR__c": {"value": 500000},
                    "CloseDate": {"value": "2026-10-01"},
                    "PushCount": {"value": 0},
                    "OpportunityScore": {"value": {"fields": {"Score": {"value": 63}}}},
                    "IsPriorityRecord": {"value": False},
                    "IsClosed": {"value": False},
                    "Type": {"value": "Land"},
                }
            },
        ],
        analysis_year=2026,
        close_start="2026-07-01",
        close_end="2026-09-30",
    )

    assert rows == [
        [
            "APAC Q3 Land",
            "Jesper Tyrer",
            "5 - Preferred",
            "Commit",
            250000,
            "EUR",
            "2026-07-18",
            2,
            71,
            "Yes",
        ]
    ]


def test_filter_pi_records_to_territory_scope_uses_soql_where(monkeypatch) -> None:
    records = [
        {"id": "006KEEP", "fields": {"Name": {"value": "Keep"}}},
        {"id": "006DROP", "fields": {"Name": {"value": "Drop"}}},
    ]
    seen_queries = []

    def fake_run_soql(session, instance_url, query, label=""):
        seen_queries.append((query, label))
        assert "Account.Industry IN ('Asset Management','Wealth Management')" in query
        return [{"Id": "006KEEP"}]

    monkeypatch.setattr(extract_live, "run_soql", fake_run_soql)

    assert _filter_pi_records_to_territory_scope(
        object(),
        "https://example.my.salesforce.com",
        records,
        "Account.Industry IN ('Asset Management','Wealth Management')",
        label="NA Asset Management:pi",
    ) == [records[0]]
    assert seen_queries[0][1] == "NA Asset Management:pi:territory_scope"


def test_write_run_audit_emits_json_and_summary(tmp_path: Path) -> None:
    _write_run_audit(
        tmp_path,
        {
            "run_date": "2026-08-10",
            "status": "failed",
            "scope": "all",
            "territories_requested": ["APAC", "Canada"],
            "processed": [
                {
                    "territory": "APAC",
                    "director": "Jesper Tyrer",
                    "workbook_path": "output/director_live_workbooks/2026-08-10/jesper-tyrer.xlsx",
                    "counts": {
                        "pipeline_open": 12,
                        "pipeline_inspection": 12,
                    },
                    "forward_quarter_pi": {
                        "status": "audit_fallback",
                        "quarter_title": "Q4 2026",
                        "deal_count": 3,
                    },
                }
            ],
            "failures": [
                {
                    "territory": "Canada",
                    "error_type": "RuntimeError",
                    "message": "source missing",
                }
            ],
            "query_telemetry_totals": {
                "queries": 9,
                "rows": 123,
                "duration_ms": 4567,
            },
        },
    )

    audit_json = tmp_path / "director_live_extract_audit.json"
    summary_md = tmp_path / "summary.md"

    assert audit_json.exists()
    assert summary_md.exists()
    assert '"status": "failed"' in audit_json.read_text(encoding="utf-8")
    summary = summary_md.read_text(encoding="utf-8")
    assert "Jesper Tyrer" in summary
    assert "Canada" in summary
    assert "Q4 2026" in summary


def test_extract_territory_tracks_actionable_commercial_approval_rows(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        extract_live,
        "TERRITORIES",
        {
            "APAC": {
                "director": "Jesper Tyrer",
                "soql_where": "Owner.Name = 'Jesper Tyrer'",
                "pi_list_view_id": "00BTb00000Ksa4bMAB",
                "forward_quarter_pi_list_views": {},
            }
        },
    )
    monkeypatch.setattr(
        extract_live,
        "_runtime_period",
        lambda _: {
            "analysis_year": 2026,
            "fy_label": "FY26",
            "fy_close_filter": "",
            "q1_start": "2026-01-01",
            "q1_end": "2026-03-31",
            "q2_start": "2026-04-01",
            "q2_end": "2026-06-30",
            "q3_start": "2026-07-01",
            "snapshot_date": "2026-04-22",
            "forward_quarter_label": "Q3",
            "forward_quarter_title": "Q3 2026",
        },
    )
    monkeypatch.setattr(
        extract_live,
        "_load_forward_quarter_pi_audit_fallback",
        lambda *args, **kwargs: {},
    )
    monkeypatch.setattr(
        extract_live,
        "_resolve_forward_quarter_pi_source",
        lambda *args, **kwargs: {},
    )
    monkeypatch.setattr(extract_live, "fetch_pi", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        extract_live, "BUNDLE_OUTPUT_ROOT", tmp_path / "director_bundles"
    )

    pipeline = [
        _make_opportunity(
            opp_id="001",
            name="Approved Current",
            stage="4 - Shortlisted",
            close_date="2026-05-01",
            approval_status="Approved",
            approved=True,
            approval_date="2026-02-10",
        ),
        _make_opportunity(
            opp_id="002",
            name="Approved Prior",
            stage="4 - Shortlisted",
            close_date="2026-05-02",
            approval_status="Approved",
            approved=True,
            approval_date="2025-12-20",
        ),
        _make_opportunity(
            opp_id="003",
            name="Pending",
            stage="4 - Shortlisted",
            close_date="2026-05-03",
            approval_status="Needs Approval",
            submitted=True,
            submitted_date="2026-03-15",
        ),
        _make_opportunity(
            opp_id="004",
            name="Missing",
            stage="3 - Engagement",
            close_date="2026-05-04",
            approval_status="Pending",
        ),
        _make_opportunity(
            opp_id="005",
            name="Exempt",
            stage="3 - Engagement",
            close_date="2026-05-05",
            approval_status="No Approval Necessary",
        ),
    ]
    renewals = [
        {
            "Account": {"Name": "Acme"},
            "Name": "Renewal A",
            "Owner": {"Name": "Jesper Tyrer"},
            "StageName": "4 - Shortlisted",
            "CloseDate": "2026-06-01",
            "Amount": 250000,
            "Probability": 80,
        },
        {
            "Account": {"Name": "Beta"},
            "Name": "Renewal B",
            "Owner": {"Name": "Jesper Tyrer"},
            "StageName": "5 - Preferred",
            "CloseDate": "2026-06-15",
            "Amount": 125000,
            "Probability": 60,
        },
    ]

    def fake_run_soql(session, instance_url, query, label):
        if label == "APAC:all_fy_deals":
            return pipeline
        if label == "APAC:renewals":
            return renewals
        if label in {"APAC:tasks_90d", "APAC:events_90d", "APAC:field_history"}:
            return []
        return []

    monkeypatch.setattr(extract_live, "run_soql", fake_run_soql)

    output_path = tmp_path / "jesper-tyrer.xlsx"
    result = extract_live.extract_territory(
        "APAC",
        "2026-04-22",
        output_path,
        session=object(),
        instance_url="https://example.my.salesforce.com",
    )

    assert result["counts"]["commercial_approval_land"] == 5
    assert result["counts"]["commercial_approval_sheet_rows"] == 4
    assert result["counts"]["approved_current_year"] == 1
    assert result["counts"]["approved_prior_year"] == 1
    assert result["counts"]["pending_approval"] == 1
    assert result["counts"]["missing_approval"] == 1

    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert wb["Commercial Approval"].max_row - 1 == 4
        summary_rows = list(wb["Summary"].iter_rows(values_only=True))
        assert ("Commercial Approval", 4, "SOQL — open Land, FY26") in summary_rows
    finally:
        wb.close()


def test_extract_territory_writes_json_bundle(tmp_path: Path, monkeypatch) -> None:
    from scripts.monthly_platform.models import DirectorBundle

    monkeypatch.setattr(
        extract_live,
        "TERRITORIES",
        {
            "APAC": {
                "director": "Jesper Tyrer",
                "soql_where": "Account_Unit_Group__c = 'SC Asia'",
                "pi_list_view_id": "00BTb00000Ksa4bMAB",
                "forward_quarter_pi_list_views": {},
            }
        },
    )
    monkeypatch.setattr(
        extract_live, "BUNDLE_OUTPUT_ROOT", tmp_path / "director_bundles"
    )

    def fake_run_soql(session, instance_url, query, label=""):
        if "all_fy_deals" in label:
            return [
                _make_opportunity(
                    opp_id="006TEST",
                    name="Test Opp",
                    stage="3 - Engagement",
                    close_date="2026-06-15",
                    approval_status="",
                    approved=False,
                )
            ]
        return []

    monkeypatch.setattr(extract_live, "run_soql", fake_run_soql)
    monkeypatch.setattr(extract_live, "fetch_pi", lambda *a, **kw: [])

    workbook_path = tmp_path / "2026-04-22" / "jesper-tyrer.xlsx"
    result = extract_live.extract_territory(
        "APAC",
        "2026-04-22",
        workbook_path,
        session=object(),
        instance_url="https://test.my.salesforce.com",
    )

    bundle_path = tmp_path / "director_bundles" / "2026-04-22" / "jesper-tyrer.json"
    assert bundle_path.exists()

    bundle = DirectorBundle.from_json(bundle_path.read_text())
    assert bundle.director == "Jesper Tyrer"
    assert bundle.territory == "APAC"
    assert bundle.schema_version == "1"
    assert len(bundle.datasets.pipeline_open) == 1
    assert bundle.datasets.pipeline_open[0].opportunity == "Test Opp"
    assert bundle.dataset_counts["pipeline_open"] == 1

    assert workbook_path.exists()
    assert result["territory"] == "APAC"
    assert "bundle_path" in result


def test_write_run_manifest(tmp_path: Path) -> None:
    from scripts.extract_director_live import _write_run_manifest

    processed = [
        {
            "territory": "APAC",
            "director": "Jesper Tyrer",
            "workbook_path": "output/jesper-tyrer.xlsx",
            "bundle_path": "output/jesper-tyrer.json",
            "counts": {"pipeline_open": 12, "won_lost": 5},
        },
    ]
    manifest_path = tmp_path / "manifest.json"
    _write_run_manifest(
        manifest_path,
        processed=processed,
        failures=[],
        durations={"APAC": 3.2},
        snapshot_date="2026-04-22",
        started_at="2026-04-22T09:30:00Z",
        finished_at="2026-04-22T09:32:45Z",
        query_telemetry_totals={"queries": 76, "rows": 24539, "duration_ms": 21900},
    )

    assert manifest_path.exists()
    data = json.loads(manifest_path.read_text())
    assert data["schema_version"] == "1"
    assert data["run_date"] == "2026-04-22"
    assert len(data["directors"]) == 1
    assert data["directors"][0]["name"] == "Jesper Tyrer"
    assert data["directors"][0]["status"] == "ok"
    assert data["directors"][0]["duration_seconds"] == 3.2
    assert data["failures"] == []
