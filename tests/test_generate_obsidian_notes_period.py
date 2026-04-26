import json
from pathlib import Path

from openpyxl import Workbook

from scripts import generate_obsidian_notes as notes


def _set_row(ws, row_index: int, values: list[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=column_index).value = value


def _set_cell_row(ws, row_index: int, mapping: dict[int, object]) -> None:
    for column_index, value in mapping.items():
        ws.cell(row=row_index, column=column_index).value = value


def test_read_analytics_workbook_uses_runtime_historical_trending_contract(
    tmp_path: Path,
) -> None:
    sharepoint = tmp_path / "sharepoint"
    sharepoint.mkdir(parents=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    velocity = wb.create_sheet("Pipeline Velocity")
    _set_row(velocity, 1, ["Q3 2026 (Historical), ARR by Snapshot (EUR)"])
    _set_row(velocity, 2, ["Director", "2026-07-01", "2026-09-30"])
    _set_row(velocity, 3, ["Jesper Tyrer", 100.0, 50.0])
    _set_row(velocity, 5, ["Q4 2026 (Current), ARR by Snapshot (EUR)"])
    _set_row(velocity, 6, ["Director", "2026-10-01", "2026-10-15"])
    _set_row(velocity, 7, ["Jesper Tyrer", 80.0, 90.0])

    trend = wb.create_sheet("Q3 Trend Consolidated")
    _set_cell_row(trend, 3, {1: "APAC", 21: 100.0, 22: 0.0, 25: "Won"})
    _set_cell_row(trend, 4, {1: "APAC", 21: 0.0, 22: 50.0, 25: "Added"})

    wb.save(sharepoint / "FY26 Pipeline Review, All Territories.xlsx")

    previous_sharepoint = notes.SHAREPOINT
    previous_runtime = dict(notes.RUNTIME_PERIOD)
    notes.SHAREPOINT = sharepoint
    notes._configure_runtime_period(as_of_date="2026-10-15")
    try:
        findings = notes._read_analytics_workbook()
    finally:
        notes.SHAREPOINT = previous_sharepoint
        notes.RUNTIME_PERIOD = previous_runtime

    assert findings["velocity_retrospective"] == {
        "dates": ["2026-07-01", "2026-09-30"],
        "directors": [("Jesper Tyrer", [100.0, 50.0])],
    }
    assert findings["velocity_current"] == {
        "dates": ["2026-10-01", "2026-10-15"],
        "directors": [("Jesper Tyrer", [80.0, 90.0])],
    }
    assert findings["variance_totals"] == {
        "initial": 100.0,
        "final": 50.0,
        "delta": -50.0,
        "Won": 100.0,
        "Lost": 0.0,
        "Added": 50.0,
        "RevisedUp": 0.0,
        "RevisedDown": 0.0,
    }


def test_director_stats_uses_current_quarter_for_renewals(tmp_path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    renewals = wb.create_sheet("Renewals FY26")
    _set_row(renewals, 1, ["Close Date"])
    _set_row(renewals, 2, ["2026-11-01"])
    _set_row(renewals, 3, ["2026-08-01"])

    path = tmp_path / "director.xlsx"
    wb.save(path)

    previous_runtime = dict(notes.RUNTIME_PERIOD)
    notes._configure_runtime_period(as_of_date="2026-10-15")
    try:
        stats = notes._director_stats("Jesper Tyrer", "APAC", path)
    finally:
        notes.RUNTIME_PERIOD = previous_runtime

    assert stats["renewals_q2"] == 1


def test_write_monthly_director_uses_runtime_velocity_and_risk_labels(
    tmp_path: Path,
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    q3_sheet = wb.create_sheet("Q3 Snapshot Trend")
    _set_row(q3_sheet, 2, ["Account", "ARR 2026-07-01", "ARR 2026-09-30"])
    _set_row(q3_sheet, 3, ["Acme", 100.0, 50.0])

    q4_sheet = wb.create_sheet("Q4 Snapshot Trend")
    _set_row(q4_sheet, 2, ["Account", "ARR 2026-10-01", "ARR 2026-10-15"])
    _set_row(q4_sheet, 3, ["Acme", 80.0, 90.0])

    wb_path = tmp_path / "jesper-tyrer.xlsx"
    wb.save(wb_path)

    month_dir = tmp_path / "Monthly" / "2026-10"
    month_dir.mkdir(parents=True)
    stats = {
        "director": "Jesper Tyrer",
        "territory": "APAC",
        "open_land_deals": 0,
        "open_land_arr_unwtd": 0.0,
        "open_land_arr_wtd": 0.0,
        "q1_won_count": 0,
        "q1_won_arr": 0.0,
        "q1_lost_count": 0,
        "q1_lost_arr": 0.0,
        "approved_2026": 0,
        "conditionally_approved": 0,
        "missing_approval": 0,
        "renewals_q2": 2,
        "q1_slip_events": 0,
        "top_land_deals": [],
        "owner_push_concentration": [],
    }
    findings = {
        "deal_risk": [
            {
                "director": "Jesper Tyrer",
                "close_date": "2026-11-01",
                "score": 77,
                "account": "Acme",
                "opportunity": "Alpha",
                "stage": "3 - Engagement",
                "arr": 125000.0,
                "reason_codes": "STALE",
                "proof": "No activity 70d",
            }
        ]
    }

    previous_runtime = dict(notes.RUNTIME_PERIOD)
    notes._configure_runtime_period(as_of_date="2026-10-15")
    try:
        path = notes.write_monthly_director(
            month_dir,
            stats,
            "2026-10-15",
            wb_path=wb_path,
            findings=findings,
        )
    finally:
        notes.RUNTIME_PERIOD = previous_runtime

    text = path.read_text()
    assert "- Q4 renewals due: 2." in text
    assert "- Q3 Land outcome: 0 wins (EUR 0), 0 losses (EUR 0)." in text
    assert "- Q3 2026: EUR 100 on 2026-07-01 -> EUR 50 on 2026-09-30" in text
    assert "- Q4 2026: EUR 80 on 2026-10-01 -> EUR 90 on 2026-10-15" in text
    assert "## Top Q4 deals at risk" in text
    assert "Composite risk score, this director, Q4 2026 close dates." in text


def test_update_snapshot_history_persists_runtime_retrospective_metadata(
    tmp_path: Path,
) -> None:
    previous_vault = notes.VAULT
    previous_runtime = dict(notes.RUNTIME_PERIOD)
    notes.VAULT = tmp_path
    notes._configure_runtime_period(as_of_date="2026-10-15")
    try:
        path = notes._update_snapshot_history(
            "2026-10-15",
            [
                {
                    "director": "Jesper Tyrer",
                    "territory": "APAC",
                    "open_land_deals": 4,
                    "open_land_arr_unwtd": 400.0,
                    "open_land_arr_wtd": 250.0,
                    "open_land_q1q2_deals": 2,
                    "open_land_q1q2_arr_unwtd": 200.0,
                    "open_land_q1q2_arr_wtd": 125.0,
                    "q1_won_count": 3,
                    "q1_won_arr": 300.0,
                    "q1_lost_count": 1,
                    "q1_lost_arr": 100.0,
                    "approved_2026": 2,
                    "conditionally_approved": 1,
                    "missing_approval": 0,
                    "renewals_q2": 1,
                }
            ],
        )
    finally:
        notes.VAULT = previous_vault
        notes.RUNTIME_PERIOD = previous_runtime

    history = json.loads(path.read_text())
    snapshot = history["snapshots"][0]
    director = snapshot["directors"]["Jesper Tyrer"]

    assert snapshot["retrospective_quarter_label"] == "Q3"
    assert snapshot["retrospective_quarter_title"] == "Q3 2026"
    assert director["retrospective_land_label"] == "Q3"
    assert director["retrospective_land_title"] == "Q3 2026"
    assert director["retrospective_land_won_count"] == 3
    assert director["retrospective_land_won_arr"] == 300.0
    assert director["q1_won_count"] == 3
