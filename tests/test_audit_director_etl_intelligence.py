import dataclasses
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from bundle_factory import make_test_bundle
from scripts.audit_director_etl_intelligence import build_etl_intelligence_audit
from scripts.monthly_platform.excel_renderer import render_bundle_to_excel
from scripts.monthly_platform.models import CloseDateEvent, ForecastEvent, StageEvent


def test_audit_flags_json_dataset_not_rendered_in_workbook(tmp_path):
    bundle = make_test_bundle()
    bundle = dataclasses.replace(
        bundle,
        datasets=dataclasses.replace(
            bundle.datasets,
            close_date_events=[
                CloseDateEvent(
                    opportunity_id="006x",
                    opportunity="Big Deal",
                    account="Acme Corp",
                    owner="Jane Smith",
                    current_stage="3 - Engagement",
                    old_value="2026-06-30",
                    new_value="2026-09-30",
                    created_date="2026-04-01",
                    arr_unweighted=500000,
                    is_closed=False,
                )
            ],
        ),
        dataset_counts={**bundle.dataset_counts, "close_date_events": 1},
    )
    workbook = tmp_path / "director.xlsx"
    bundle_json = tmp_path / "bundle.json"
    render_bundle_to_excel(bundle, workbook)
    wb = load_workbook(workbook)
    del wb["Close Date History"]
    wb.save(workbook)
    bundle_json.write_text(bundle.to_json(), encoding="utf-8")

    audit = build_etl_intelligence_audit(
        bundle_path=bundle_json,
        workbook_path=workbook,
    )

    assert audit["summary"]["high_gap_count"] >= 1
    assert any(
        gap["type"] in {"dataset_not_rendered", "sheet_missing"}
        and gap["dataset"] == "close_date_events"
        for gap in audit["coverage_gaps"]
    )
    assert any(
        rec["priority"] == "P0" and "Close Date History" in rec["recommendation"]
        for rec in audit["recommendations"]
    )


def test_audit_builds_joined_deal_risk_index(tmp_path):
    bundle = make_test_bundle()
    bundle = dataclasses.replace(
        bundle,
        datasets=dataclasses.replace(
            bundle.datasets,
            stage_events=[
                StageEvent(
                    opportunity_id="006x",
                    opportunity="Big Deal",
                    account="Acme Corp",
                    owner="Jane Smith",
                    current_stage="3 - Engagement",
                    old_value="2 - Discovery",
                    new_value="3 - Engagement",
                    created_date="2026-04-01",
                    arr_unweighted=500000,
                    is_closed=False,
                    is_won=False,
                )
            ]
            * 4,
            forecast_category_events=[
                ForecastEvent(
                    opportunity_id="006x",
                    opportunity="Big Deal",
                    account="Acme Corp",
                    owner="Jane Smith",
                    current_stage="3 - Engagement",
                    old_value="Pipeline",
                    new_value="Best Case",
                    created_date="2026-04-01",
                    arr_unweighted=500000,
                )
            ]
            * 2,
        ),
        dataset_counts={
            **bundle.dataset_counts,
            "stage_events": 4,
            "forecast_category_events": 2,
        },
    )
    workbook = tmp_path / "director.xlsx"
    bundle_json = tmp_path / "bundle.json"
    render_bundle_to_excel(bundle, workbook)
    bundle_json.write_text(bundle.to_json(), encoding="utf-8")

    audit = build_etl_intelligence_audit(
        bundle_path=bundle_json,
        workbook_path=workbook,
    )

    risk = audit["analytics"]["deal_risk_index"]
    assert risk[0]["opportunity"] == "Big Deal"
    assert "stage changes" in " ".join(risk[0]["risk_reasons"])
    assert "forecast-category changes" in " ".join(risk[0]["risk_reasons"])


def test_audit_catches_workbook_row_mismatch(tmp_path):
    bundle = make_test_bundle()
    workbook = tmp_path / "director.xlsx"
    bundle_json = tmp_path / "bundle.json"
    render_bundle_to_excel(bundle, workbook)
    wb = load_workbook(workbook)
    ws = wb["Pipeline Open FY26"]
    ws.delete_rows(2)
    wb.save(workbook)
    bundle_json.write_text(bundle.to_json(), encoding="utf-8")

    audit = build_etl_intelligence_audit(
        bundle_path=bundle_json,
        workbook_path=workbook,
    )

    assert any(
        gap["type"] == "row_count_mismatch" and gap["dataset"] == "pipeline_open"
        for gap in audit["coverage_gaps"]
    )
