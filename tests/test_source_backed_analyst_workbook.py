from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.monthly_platform.analyst_workbook import (
    SHEET_NAMES,
    build_source_backed_analyst_workbook,
)
from scripts.monthly_platform.director_bundle_builder import (
    build_director_bundle_from_source_bundle,
)
from scripts.monthly_platform.models import DirectorBundle
from scripts.monthly_platform.source_bundles import (
    HistoricalTrendRows,
    PipelineDisplayDecision,
    PipelineInspectionRow,
    PipelineOpenRow,
    TerritorySourceBundle,
)


def _source_bundle() -> TerritorySourceBundle:
    return TerritorySourceBundle(
        snapshot_date="2026-04-30",
        source_run_id="source-run",
        generated_at="2026-04-30T09:30:00Z",
        territory="APAC",
        director="Jesper Tyrer",
        period_context={},
        source_extract_ids=["src-current", "src-forward", "src-history"],
        pipeline_open=[
            PipelineOpenRow(
                opportunity="Big Deal",
                account="Acme",
                owner="Owner",
                stage="3 - Engagement",
                forecast_category="Commit",
                close_date="2026-06-30",
                arr_unweighted=500.0,
                arr_weighted=250.0,
                probability=50.0,
                deal_type="Land",
                sales_region="APAC",
                created_date="2026-01-01",
                last_modified_date="2026-04-01",
                currency="EUR",
                quarter="Q2 2026",
                source_extract_id="src-pipeline-open",
            )
        ],
        historical_trending=HistoricalTrendRows(
            prior_quarter=[
                {
                    "Opportunity Name": "Big Deal",
                    "Account Name": "Acme",
                    "Close Date": "2026-06-30",
                    "Forecast ARR (converted) (2026-04-01)": "EUR 100,00",
                    "Forecast ARR (converted) (2026-04-07)": "EUR 250,00",
                    "Stage (2026-04-01)": "2 - Discovery",
                    "Stage (2026-04-07)": "3 - Engagement",
                }
            ]
        ),
        pi_current=[
            PipelineInspectionRow(
                opportunity="Big Deal",
                account="Acme",
                owner="Owner",
                stage="3 - Engagement",
                forecast_category="Commit",
                arr_weighted=250.0,
                currency="EUR",
                close_date="2026-06-30",
                push_count=1,
                score=75,
                priority=True,
                source_extract_id="src-current",
                active_in_period=True,
            )
        ],
        pi_forward=[
            PipelineInspectionRow(
                opportunity="Forward Deal",
                owner="Owner",
                stage="2 - Discovery",
                forecast_category="Pipeline",
                arr_weighted=100.0,
                currency="EUR",
                close_date="2026-09-30",
                source_extract_id="src-forward",
                active_in_period=True,
            )
        ],
        pipeline_display_decision=PipelineDisplayDecision(
            display_period_role="current_quarter",
            display_quarter_label="Q2",
            display_quarter_title="Q2 2026",
            display_reason="current_quarter",
            requires_forward_quarter_fallback=False,
            current_quarter_active_deals=1,
            current_quarter_active_arr=250.0,
            forward_quarter_active_deals=1,
            forward_quarter_active_arr=100.0,
        ),
    )


def _write_manifest(tmp_path: Path) -> Path:
    source_bundle = _source_bundle()
    director_bundle = build_director_bundle_from_source_bundle(source_bundle)
    bundle_path = tmp_path / "apac.json"
    bundle_path.write_text(director_bundle.to_json() + "\n", encoding="utf-8")
    manifest = {
        "schema_version": "monthly_platform.director_bundle_manifest.v1",
        "snapshot_date": "2026-04-30",
        "source_run_id": "source-run",
        "status": "ok",
        "generated_at": "2026-04-30T09:30:00Z",
        "source_bundle_manifest_path": "source/source_bundle_manifest.json",
        "output_dir": str(tmp_path),
        "bundle_paths": [str(bundle_path)],
        "summary": {
            "bundle_count": 1,
            "source_bundle_count": 1,
            "directors": [
                {
                    "director": director_bundle.director,
                    "territory": director_bundle.territory,
                    "dataset_counts": director_bundle.dataset_counts,
                    "dataset_coverage": {
                        "schema_version": "monthly_platform.director_bundle_contract.v1",
                        "source_backed": [
                            "pi_current",
                            "pi_forward",
                            "pipeline_open",
                            "snapshot_trend",
                        ],
                        "optional_empty": [
                            "activity",
                            "approvals",
                            "close_date_events",
                            "commit_items",
                            "forecast_category_events",
                            "movement_current",
                            "movement_prior",
                            "renewals",
                            "stage_events",
                            "won_lost",
                        ],
                        "publish_required": [
                            "pi_current",
                            "pi_forward",
                            "pipeline_open",
                            "snapshot_trend",
                        ],
                        "dataset_counts": director_bundle.dataset_counts,
                    },
                }
            ],
            "unsupported_datasets_empty": [
                "activity",
                "approvals",
                "close_date_events",
                "commit_items",
                "forecast_category_events",
                "movement_current",
                "movement_prior",
                "renewals",
                "stage_events",
                "won_lost",
            ],
        },
        "findings": [],
    }
    manifest_path = tmp_path / "director_bundle_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return manifest_path


def test_build_source_backed_analyst_workbook_sheets_and_row_counts(
    tmp_path: Path,
) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_path = tmp_path / "analyst.xlsx"

    result = build_source_backed_analyst_workbook(
        manifest_path=manifest_path,
        output_path=output_path,
    )

    assert result.workbook_path == output_path
    assert result.bundle_count == 1
    assert result.sheet_row_counts == {
        "Executive Summary": 1,
        "Source Coverage": 14,
        "Metric Store": 22,
        "Deal Exceptions": 3,
        "Region Narrative Inputs": 10,
        "Analyst Notes Seed": 4,
    }

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == SHEET_NAMES
    assert workbook["Executive Summary"].max_row == 2
    assert workbook["Source Coverage"].max_row == 15
    assert workbook["Metric Store"].max_row == 23
    assert workbook["Deal Exceptions"].max_row == 4
    assert workbook["Region Narrative Inputs"].max_row == 11
    assert workbook["Analyst Notes Seed"].max_row == 5


def test_workbook_metrics_are_source_bundle_derived(tmp_path: Path) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_path = tmp_path / "analyst.xlsx"

    build_source_backed_analyst_workbook(
        manifest_path=manifest_path,
        output_path=output_path,
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    metric_rows = list(workbook["Metric Store"].iter_rows(values_only=True))
    metric_by_key = {row[4]: row[5] for row in metric_rows[1:]}

    assert metric_by_key["dataset_count.pipeline_open"] == 1
    assert metric_by_key["dataset_count.snapshot_trend"] == 2
    assert metric_by_key["pipeline_open_arr_weighted"] == 250
    assert metric_by_key["pi_current_priority_count"] == 1

    restored = DirectorBundle.from_json((tmp_path / "apac.json").read_text())
    assert (
        metric_by_key["pi_forward_arr_weighted"]
        == restored.datasets.pi_forward[0].arr_weighted
    )


def test_workbook_has_deterministic_executive_summary_and_notes(
    tmp_path: Path,
) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_path = tmp_path / "analyst.xlsx"

    build_source_backed_analyst_workbook(
        manifest_path=manifest_path,
        output_path=output_path,
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    summary = workbook["Executive Summary"]
    summary_row = next(summary.iter_rows(min_row=2, max_row=2, values_only=True))
    assert summary_row[4] == "Ready with deal review flags"
    assert summary_row[5] == (
        "4 source-backed datasets; 10 optional-empty datasets; 0 unknown policies."
    )
    assert summary_row[6] == "1 open pipeline deal; EUR 250 weighted ARR."
    assert summary_row[8] == (
        "3 review flags: 1 close-date movement, 1 execution hygiene, "
        "1 Pipeline Inspection."
    )
    assert summary_row[10] == (
        "Deterministic source-backed summary; no generated judgment or unsupported claim."
    )

    notes = {
        row[4]: row[5]
        for row in workbook["Analyst Notes Seed"].iter_rows(
            min_row=2,
            values_only=True,
        )
    }
    assert notes["coverage"] == (
        "Coverage is explicit: 4 source-backed datasets, 10 optional-empty datasets, "
        "and no inferred sources."
    )
    assert notes["pipeline_open"] == (
        "Open pipeline shows 1 deal and EUR 250 weighted ARR."
    )


def test_deal_exceptions_use_human_review_categories(tmp_path: Path) -> None:
    manifest_path = _write_manifest(tmp_path)
    output_path = tmp_path / "analyst.xlsx"

    build_source_backed_analyst_workbook(
        manifest_path=manifest_path,
        output_path=output_path,
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    rows = list(workbook["Deal Exceptions"].iter_rows(values_only=True))
    assert rows[0][13:19] == (
        "Exception Category",
        "Exception Type",
        "Severity",
        "Evidence Value",
        "Deterministic Rule",
        "Analyst Action",
    )
    categories = {row[13] for row in rows[1:]}
    actions = {row[18] for row in rows[1:]}
    assert categories == {
        "Close-date movement",
        "Execution hygiene",
        "Pipeline Inspection",
    }
    assert actions == {
        "Ask owner for the next customer-facing action and due date.",
        "Check close-date confidence and next confirmed milestone.",
        "Review PI priority rationale before leadership readout.",
    }
