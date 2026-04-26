import json
import sys
from pathlib import Path

from openpyxl import Workbook

from scripts import validate_sharepoint_analysis_contract as contract_script


def _build_workbook(path: Path, required_sheets: list[str], minimum_sheets: int) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name in required_sheets:
        ws = wb.create_sheet(name)
        ws.append(["header"])
        ws.append([1])
    filler_index = 1
    while len(wb.sheetnames) < minimum_sheets:
        ws = wb.create_sheet(f"Filler {filler_index}")
        ws.append(["header"])
        ws.append([filler_index])
        filler_index += 1
    wb.save(path)


def _create_valid_sharepoint_outputs(root: Path, *, report_date: str = "2026-04-22") -> None:
    contract = contract_script._historical_contract(report_date)
    master_required = list(contract_script.MASTER_REQUIRED_SHEETS_BASE) + [
        contract.retrospective_consolidated_sheet,
        contract.current_consolidated_sheet,
    ]
    _build_workbook(
        root / contract_script.MASTER_WORKBOOK,
        master_required,
        contract_script.MIN_SHEET_COUNTS["master"],
    )
    for _, filename in contract_script.REGIONAL_WORKBOOKS:
        _build_workbook(
            root / filename,
            list(contract_script.REGIONAL_REQUIRED_SHEETS),
            contract_script.MIN_SHEET_COUNTS["regional"],
        )
    _build_workbook(
        root / contract_script.DASHBOARD_WORKBOOK,
        list(contract_script.DASHBOARD_REQUIRED_SHEETS),
        contract_script.MIN_SHEET_COUNTS["dashboard"],
    )


def test_main_validates_sharepoint_analysis_contract(
    tmp_path: Path, monkeypatch
) -> None:
    sharepoint_root = tmp_path / "sharepoint"
    sharepoint_root.mkdir(parents=True, exist_ok=True)
    _create_valid_sharepoint_outputs(sharepoint_root)

    monkeypatch.setattr(contract_script, "SHAREPOINT_ROOT", sharepoint_root)
    monkeypatch.setattr(
        contract_script,
        "OUTPUT_ROOT",
        tmp_path / "output" / "sharepoint_analysis_contract",
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "validate_sharepoint_analysis_contract.py",
            "--date",
            "2026-04-22",
            "--sharepoint-root",
            str(sharepoint_root),
        ],
    )

    assert contract_script.main() == 0
    audit_path = (
        tmp_path
        / "output"
        / "sharepoint_analysis_contract"
        / "2026-04-22"
        / "sharepoint_analysis_contract_audit.json"
    )
    payload = json.loads(audit_path.read_text(encoding="utf-8"))
    assert payload["status"] == "ok"
    assert len(payload["validated"]) == 11
    assert not payload["failures"]


def test_main_fails_when_dashboard_sheet_is_missing(
    tmp_path: Path, monkeypatch
) -> None:
    sharepoint_root = tmp_path / "sharepoint"
    sharepoint_root.mkdir(parents=True, exist_ok=True)
    _create_valid_sharepoint_outputs(sharepoint_root)

    dashboard_path = sharepoint_root / contract_script.DASHBOARD_WORKBOOK
    wb = Workbook()
    wb.remove(wb.active)
    for name in contract_script.DASHBOARD_REQUIRED_SHEETS:
        if name == "PI Summary":
            continue
        ws = wb.create_sheet(name)
        ws.append(["header"])
        ws.append([1])
    filler_index = 1
    while len(wb.sheetnames) < contract_script.MIN_SHEET_COUNTS["dashboard"]:
        ws = wb.create_sheet(f"Filler {filler_index}")
        ws.append(["header"])
        ws.append([filler_index])
        filler_index += 1
    wb.save(dashboard_path)

    monkeypatch.setattr(contract_script, "SHAREPOINT_ROOT", sharepoint_root)
    monkeypatch.setattr(
        contract_script,
        "OUTPUT_ROOT",
        tmp_path / "output" / "sharepoint_analysis_contract",
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "validate_sharepoint_analysis_contract.py",
            "--date",
            "2026-04-22",
            "--sharepoint-root",
            str(sharepoint_root),
        ],
    )

    assert contract_script.main() == 1
    audit_path = (
        tmp_path
        / "output"
        / "sharepoint_analysis_contract"
        / "2026-04-22"
        / "sharepoint_analysis_contract_audit.json"
    )
    payload = json.loads(audit_path.read_text(encoding="utf-8"))
    assert payload["status"] == "failed"
    assert {
        "workbook_id": "dashboard_q1",
        "workbook_type": "dashboard",
        "territory": None,
        "issue": "missing_required_sheets",
        "message": "PI Summary",
    } in payload["failures"]
