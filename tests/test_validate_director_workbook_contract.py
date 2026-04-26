import json
import sys
from pathlib import Path

from openpyxl import Workbook

from scripts import validate_director_workbook_contract as contract_script


def _add_standard_sheet(wb: Workbook, name: str, row_count: int) -> None:
    ws = wb.create_sheet(name)
    ws.append(["header"])
    for idx in range(row_count):
        ws.append([idx + 1])


def _add_historical_sheet(wb: Workbook, name: str, row_count: int) -> None:
    ws = wb.create_sheet(name)
    ws["A1"] = "Historical Trending"
    ws["A2"] = "header"
    for idx in range(row_count):
        ws.cell(row=idx + 3, column=1, value=idx + 1)


def test_expected_live_metric_count_uses_actionable_approval_rows() -> None:
    assert contract_script._expected_live_metric_count(
        {
            "commercial_approval_land": 34,
            "approved_current_year": 2,
            "approved_prior_year": 4,
            "pending_approval": 2,
            "missing_approval": 0,
        },
        "commercial_approval_land",
    ) == 8


def test_main_validates_live_and_historical_contract(
    tmp_path: Path, monkeypatch
) -> None:
    workbooks_dir = tmp_path / "workbooks" / "2026-04-22"
    workbooks_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbooks_dir / "jesper-tyrer.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _add_standard_sheet(wb, "Summary", 27)
    _add_standard_sheet(wb, "Pipeline Open FY26", 2)
    _add_standard_sheet(wb, "Won Lost FY26", 3)
    _add_standard_sheet(wb, "Commercial Approval", 4)
    _add_standard_sheet(wb, "Renewals FY26", 5)
    _add_standard_sheet(wb, "Pipeline Inspection", 6)
    _add_standard_sheet(wb, "Pipeline Inspection Forward", 7)
    _add_standard_sheet(wb, "Activity Volume", 8)
    _add_standard_sheet(wb, "Commit Items", 9)
    _add_standard_sheet(wb, "Q1 Movement", 10)
    _add_standard_sheet(wb, "Q2 Movement", 11)
    _add_standard_sheet(wb, "Stage History", 12)
    _add_standard_sheet(wb, "Forecast Category History", 13)
    _add_historical_sheet(wb, "Q1 Snapshot Trend", 14)
    _add_historical_sheet(wb, "Q2 Snapshot Trend", 15)
    wb.save(workbook_path)

    live_audit_dir = tmp_path / "output" / "director_live_extract" / "2026-04-22"
    live_audit_dir.mkdir(parents=True, exist_ok=True)
    (live_audit_dir / "director_live_extract_audit.json").write_text(
        json.dumps(
            {
                "processed": [
                    {
                        "workbook_path": str(workbook_path),
                            "counts": {
                                "pipeline_open": 2,
                                "won_lost": 3,
                                "commercial_approval_land": 4,
                                "commercial_approval_sheet_rows": 4,
                                "renewals": 5,
                                "pipeline_inspection": 6,
                                "pipeline_inspection_forward": 7,
                            "activity_volume_rows": 8,
                            "commit_items": 9,
                            "q1_movement": 10,
                            "q2_movement": 11,
                            "stage_history_events": 12,
                            "forecast_category_history_events": 13,
                        },
                        "forward_quarter_pi": {
                            "status": "configured",
                        },
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    historical_audit_dir = (
        tmp_path / "output" / "historical_trending_extract" / "2026-04-22"
    )
    historical_audit_dir.mkdir(parents=True, exist_ok=True)
    (historical_audit_dir / "historical_trending_extract_audit.json").write_text(
        json.dumps(
            {
                "processed": [
                    {
                        "slug": "jesper-tyrer",
                        "workbook_path": str(workbook_path),
                        "sheets": [
                            {"sheet_name": "Q1 Snapshot Trend", "row_count": 14},
                            {"sheet_name": "Q2 Snapshot Trend", "row_count": 15},
                        ],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(contract_script, "LIVE_AUDIT_ROOT", tmp_path / "output" / "director_live_extract")
    monkeypatch.setattr(
        contract_script,
        "HISTORICAL_AUDIT_ROOT",
        tmp_path / "output" / "historical_trending_extract",
    )
    monkeypatch.setattr(
        contract_script,
        "OUTPUT_ROOT",
        tmp_path / "output" / "director_workbook_contract",
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "validate_director_workbook_contract.py",
            "--snapshot-date",
            "2026-04-22",
            "--workbooks-dir",
            str(workbooks_dir),
            "--director",
            "jesper-tyrer",
            "--require-historical",
        ],
    )

    assert contract_script.main() == 0
    summary = (
        tmp_path
        / "output"
        / "director_workbook_contract"
        / "2026-04-22"
        / "summary.md"
    )
    assert summary.exists()
    assert "jesper-tyrer" in summary.read_text(encoding="utf-8")


def test_main_fails_when_forward_pi_sheet_is_missing(
    tmp_path: Path, monkeypatch
) -> None:
    workbooks_dir = tmp_path / "workbooks" / "2026-04-22"
    workbooks_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbooks_dir / "jesper-tyrer.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for name in [
        "Summary",
        "Pipeline Open FY26",
        "Won Lost FY26",
        "Commercial Approval",
        "Renewals FY26",
        "Pipeline Inspection",
        "Activity Volume",
        "Commit Items",
        "Q1 Movement",
        "Q2 Movement",
        "Stage History",
        "Forecast Category History",
    ]:
        _add_standard_sheet(wb, name, 0)
    wb.save(workbook_path)

    live_audit_dir = tmp_path / "output" / "director_live_extract" / "2026-04-22"
    live_audit_dir.mkdir(parents=True, exist_ok=True)
    (live_audit_dir / "director_live_extract_audit.json").write_text(
        json.dumps(
            {
                "processed": [
                    {
                        "workbook_path": str(workbook_path),
                        "counts": {
                            "pipeline_open": 0,
                            "won_lost": 0,
                            "commercial_approval_land": 0,
                            "renewals": 0,
                            "pipeline_inspection": 0,
                            "pipeline_inspection_forward": 3,
                            "activity_volume_rows": 0,
                            "commit_items": 0,
                            "q1_movement": 0,
                            "q2_movement": 0,
                            "stage_history_events": 0,
                            "forecast_category_history_events": 0,
                        },
                        "forward_quarter_pi": {
                            "status": "configured",
                        },
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(contract_script, "LIVE_AUDIT_ROOT", tmp_path / "output" / "director_live_extract")
    monkeypatch.setattr(contract_script, "OUTPUT_ROOT", tmp_path / "output" / "director_workbook_contract")
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "validate_director_workbook_contract.py",
            "--snapshot-date",
            "2026-04-22",
            "--workbooks-dir",
            str(workbooks_dir),
            "--director",
            "jesper-tyrer",
        ],
    )

    assert contract_script.main() == 1
    payload = json.loads(
        (
            tmp_path
            / "output"
            / "director_workbook_contract"
            / "2026-04-22"
            / "director_workbook_contract_audit.json"
        ).read_text(encoding="utf-8")
    )
    assert payload["status"] == "failed"
    assert any(
        item.get("issue") == "missing_sheet"
        and item.get("message") == "Pipeline Inspection Forward"
        for item in payload["failures"]
    )
