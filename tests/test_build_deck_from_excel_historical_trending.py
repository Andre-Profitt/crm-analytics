from pathlib import Path

from openpyxl import Workbook

from scripts import build_deck_from_excel as deck


def _set_row(ws, row_index: int, values: list[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=column_index).value = value


def test_read_director_analytics_uses_runtime_trend_sheet_names(
    tmp_path: Path,
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    q3_trend = wb.create_sheet("Q3 Trend Consolidated")
    _set_row(q3_trend, 2, ["Territory", "Initial ARR", "Final ARR", "Bucket"])
    _set_row(q3_trend, 3, ["APAC", 100.0, 0.0, "Won"])
    _set_row(q3_trend, 4, ["APAC", 0.0, 50.0, "Added"])

    variance = wb.create_sheet("Forecast Variance")
    _set_row(variance, 5, ["Jesper Tyrer", "APAC"])

    velocity = wb.create_sheet("Pipeline Velocity")
    _set_row(velocity, 1, ["Q3 2026 (Historical), ARR by Snapshot (EUR)"])
    _set_row(velocity, 2, ["Director", "2026-07-01", "2026-09-30"])
    _set_row(velocity, 3, ["Jesper Tyrer", 100.0, 50.0])
    _set_row(velocity, 5, ["Q4 2026 (Current), ARR by Snapshot (EUR)"])
    _set_row(velocity, 6, ["Director", "2026-10-01", "2026-10-15"])
    _set_row(velocity, 7, ["Jesper Tyrer", 80.0, 90.0])

    path = tmp_path / "analytics.xlsx"
    wb.save(path)

    previous = dict(deck.FQ)
    deck.FQ = deck._resolve_runtime_period_context(as_of_date="2026-10-15")
    try:
        analytics = deck.read_director_analytics(path, "Jesper Tyrer")
    finally:
        deck.FQ = previous

    assert analytics["variance"] == {
        "initial": 100.0,
        "final": 50.0,
        "net": -50.0,
        "won": 100.0,
        "lost": 0.0,
        "added": 50.0,
        "up": 0.0,
        "down": 0.0,
    }
    assert analytics["velocity_retrospective"] == {
        "dates": ["2026-07-01", "2026-09-30"],
        "series": [100.0, 50.0],
    }
    assert analytics["velocity_current"] == {
        "dates": ["2026-10-01", "2026-10-15"],
        "series": [80.0, 90.0],
    }
