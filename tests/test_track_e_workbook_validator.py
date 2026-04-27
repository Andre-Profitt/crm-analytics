"""Track E — E5 tests for the real workbook validator (E2).

The positive control runs against the live APAC anchor at
``~/Downloads/jesper-tyrer-2026-04-20.xlsx``. Skipped if missing.

Negative controls synthesize a small in-memory workbook so we don't
need a fixture .xlsx file in the repo. We assert that the validator
correctly surfaces:
  - missing_sheet
  - missing_required_column
  - snapshot_role_unresolved (no headers match pattern)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from scripts.monthly_platform import director_workbook_contract
from scripts.validate_track_e_workbook import validate_workbook


APAC_ANCHOR = Path("/Users/test/Downloads/jesper-tyrer-2026-04-20.xlsx")


@pytest.mark.skipif(
    not APAC_ANCHOR.exists(),
    reason="Live APAC anchor not present; run with the workbook in ~/Downloads/",
)
def test_apac_anchor_passes():
    report = validate_workbook(APAC_ANCHOR)
    assert report["status"] == "pass", report["findings"]
    assert report["sheet_count_present"] == 13
    # All 9 declared snapshot roles resolve (7 pattern + 2 runtime).
    resolved = [r for r in report["resolved_snapshot_roles"] if r["status"] == "pass"]
    assert len(resolved) == 9


def _build_minimal_workbook(
    tmp_path: Path,
    *,
    drop_sheet: str | None = None,
    drop_column: tuple[str, str] | None = None,
    blank_snapshot_columns: bool = False,
) -> Path:
    """Build a minimal workbook from the contract sheet list. Each sheet
    gets its declared required_columns as the only header row, so we
    can test missing-sheet / missing-column / unresolved-role paths."""
    contract = director_workbook_contract.load()
    wb = openpyxl.Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)
    for sheet in contract.raw["sheets"]:
        name = sheet["name"]
        if name == drop_sheet:
            continue
        ws = wb.create_sheet(name)
        header_row = sheet.get("header_row", 1)
        cols = list(sheet.get("required_columns", []) or [])
        if drop_column is not None and drop_column[0] == name:
            cols = [c for c in cols if c != drop_column[1]]
        for i, c in enumerate(cols, start=1):
            ws.cell(header_row, i, c)
        # Add an empty earlier row for trend sheets that use header_row=2.
        if header_row == 2:
            ws.cell(1, 1, "Historical Trending placeholder")
        # Trend sheets need ARR YYYY-MM-DD columns to satisfy snapshot
        # roles unless we deliberately blank them.
        if (
            name in ("Q1 Snapshot Trend", "Q2 Snapshot Trend")
            and not blank_snapshot_columns
        ):
            base_col = ws.max_column + 1
            for offset, date in enumerate(["2026-01-01", "2026-03-31"]):
                ws.cell(header_row, base_col + offset, f"ARR {date}")
            ws.cell(header_row, base_col + 2, "ARR Change 2026-03-31")
            ws.cell(header_row, base_col + 3, "StageName_ 2026-01-01")
            ws.cell(header_row, base_col + 4, "StageName_ 2026-03-31")
    out = tmp_path / "synth.xlsx"
    wb.save(out)
    return out


def test_missing_sheet_is_blocked(tmp_path):
    path = _build_minimal_workbook(tmp_path, drop_sheet="Pipeline Open FY26")
    report = validate_workbook(path)
    assert report["status"] == "fail"
    assert any(f["code"] == "missing_sheet" for f in report["findings"])


def test_missing_required_column_is_blocked(tmp_path):
    path = _build_minimal_workbook(
        tmp_path, drop_column=("Pipeline Open FY26", "ARR Unweighted (EUR)")
    )
    report = validate_workbook(path)
    assert report["status"] == "fail"
    assert any(f["code"] == "missing_required_column" for f in report["findings"])


def test_snapshot_role_unresolved_is_blocked(tmp_path):
    path = _build_minimal_workbook(tmp_path, blank_snapshot_columns=True)
    report = validate_workbook(path)
    assert report["status"] == "fail"
    assert any(f["code"] == "snapshot_role_unresolved" for f in report["findings"])
