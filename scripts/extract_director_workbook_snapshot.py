#!/usr/bin/env python3
"""Normalize Sales Director workbooks into deck-friendly JSON snapshots."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from monthly_platform.quarterly_pipeline import (
        quarterly_pipeline_display_from_snapshot,
    )
    from monthly_platform.period import resolve_period_context, sheet_names
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.quarterly_pipeline import (
        quarterly_pipeline_display_from_snapshot,
    )
    from scripts.monthly_platform.period import resolve_period_context


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_ROOT = REPO_ROOT / "output" / "director_live_workbooks"
DEFAULT_OUTPUT_ROOT = REPO_ROOT / "output" / "director_workbook_snapshots"


def slugify(value: str) -> str:
    token = re.sub(r"[^0-9A-Za-z]+", "-", (value or "").strip().lower()).strip("-")
    return token or "snapshot"


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def compact_money(value: float) -> str:
    n = float(value or 0)
    if abs(n) >= 1_000_000:
        return f"EUR {n / 1_000_000:.1f}M"
    if abs(n) >= 1_000:
        return f"EUR {n / 1_000:.0f}K"
    return f"EUR {n:,.0f}"


def quarter_window(snapshot_date: str) -> tuple[str, str]:
    dt = datetime.fromisoformat(as_text(snapshot_date)[:10]).date()
    quarter_start_month = ((dt.month - 1) // 3) * 3 + 1
    quarter_start = date(dt.year, quarter_start_month, 1)
    if quarter_start_month == 10:
        next_quarter = date(dt.year + 1, 1, 1)
    else:
        next_quarter = date(dt.year, quarter_start_month + 3, 1)
    quarter_end = next_quarter - timedelta(days=1)
    return quarter_start.isoformat(), quarter_end.isoformat()


def is_within_window(date_text: Any, start: str, end: str) -> bool:
    token = as_text(date_text)[:10]
    return bool(token) and start <= token <= end


def cache_slugify(name: str) -> str:
    parts = [part for part in (name or "").strip().lower().split() if part]
    if not parts:
        return "snapshot"
    if len(parts) == 1:
        return parts[0]
    return f"{parts[-1]}-{parts[0]}"


def row_values(ws, row_idx: int, max_col: int | None = None) -> list[Any]:
    limit = max_col or ws.max_column
    return [ws.cell(row_idx, col_idx).value for col_idx in range(1, limit + 1)]


def first_value(ws, row_idx: int, col_idx: int = 1) -> Any:
    return ws.cell(row_idx, col_idx).value


def is_blank_row(values: list[Any]) -> bool:
    return all(value in (None, "") for value in values)


def table_from_header(
    ws, header_row: int, *, max_col: int | None = None
) -> list[dict[str, Any]]:
    headers = [as_text(value) for value in row_values(ws, header_row, max_col)]
    headers = [header for header in headers if header]
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = row_values(ws, row_idx, len(headers))
        if is_blank_row(values):
            break
        rows.append({headers[i]: values[i] for i in range(len(headers))})
    return rows


def find_row(ws, text: str) -> int | None:
    target = text.strip().lower()
    for row_idx in range(1, ws.max_row + 1):
        if as_text(first_value(ws, row_idx)).strip().lower() == target:
            return row_idx
    return None


def load_json(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def parse_scorecard(ws) -> dict[str, Any]:
    sections: dict[str, dict[str, Any]] = {}
    current = ""
    for row_idx in range(1, ws.max_row + 1):
        label = as_text(first_value(ws, row_idx, 1))
        value = ws.cell(row_idx, 2).value
        if not label:
            continue
        if value in (None, "") and (
            label == label.upper() or label.startswith("CRO FORECAST TIE-OUT")
        ):
            current = slugify(label)
            sections[current] = {"title": label, "metrics": {}}
            continue
        if current and value not in (None, ""):
            sections[current]["metrics"][label] = value
    title = as_text(ws["A1"].value)
    generated = as_text(ws["A2"].value).replace("Generated:", "").strip()
    return {
        "title": title,
        "generated": generated,
        "sections": sections,
    }


def summarize_transition_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        old_value = as_text(row.get("Old Category") or row.get("OldValue")) or "Unknown"
        new_value = as_text(row.get("New Category") or row.get("NewValue")) or "Unknown"
        key = (old_value, new_value)
        if key not in grouped:
            grouped[key] = {
                "from": old_value,
                "to": new_value,
                "count": 0,
                "arr": 0.0,
            }
        grouped[key]["count"] += 1
        grouped[key]["arr"] += as_number(
            row.get("ARR (€ converted)") or row.get("ConvertedARR")
        )
    summary = list(grouped.values())
    summary.sort(
        key=lambda item: (-item["count"], -item["arr"], item["from"], item["to"])
    )
    for item in summary:
        item["arr"] = round(item["arr"], 2)
    return summary


def filtered_q1_cache_section(cache_dir: Path, territory: str) -> dict[str, Any]:
    won_rows = load_json(cache_dir / "soql_won_q1.json")
    lost_rows = load_json(cache_dir / "soql_lost_q1.json")
    open_rows = load_json(cache_dir / "soql_open_pipeline.json")
    fact_rows = load_json(cache_dir / "forecast_fact_Q1_2026.json")
    close_history = load_json(cache_dir / "field_history_CloseDate.json")
    forecast_history = load_json(cache_dir / "field_history_ForecastCategoryName.json")

    opportunity_lookup = {
        row["Id"]: row
        for row in (won_rows + lost_rows + open_rows)
        if isinstance(row, dict) and row.get("Id")
    }

    promise_rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"Category": "", "Count": 0, "ARR (€ converted)": 0.0}
    )
    for fact in fact_rows:
        opp = opportunity_lookup.get(fact.get("OpportunityId"))
        if not opp:
            continue
        category = as_text(fact.get("ForecastCategoryName")) or "Unknown"
        item = promise_rollup[category]
        item["Category"] = category
        item["Count"] += 1
        item["ARR (€ converted)"] += as_number(
            opp.get("ConvertedARR") or opp.get("APTS_Opportunity_ARR__c")
        )
    promise_baseline = list(promise_rollup.values())
    promise_baseline.sort(
        key=lambda item: (-item["ARR (€ converted)"], item["Category"])
    )
    for item in promise_baseline:
        item["ARR (€ converted)"] = round(item["ARR (€ converted)"], 2)

    pushed_deals: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row in close_history:
        opp = row.get("Opportunity") or {}
        old_close = as_text(row.get("OldValue"))[:10]
        new_close = as_text(row.get("NewValue"))[:10]
        if not old_close.startswith(("2026-01", "2026-02", "2026-03")):
            continue
        if new_close < "2026-04-01":
            continue
        opp_id = as_text(row.get("OpportunityId"))
        if not opp_id or opp_id in seen_ids:
            continue
        seen_ids.add(opp_id)
        pushed_deals.append(
            {
                "Account": as_text((opp.get("Account") or {}).get("Name")),
                "Opportunity": as_text(opp.get("Name")),
                "Owner": as_text((opp.get("Owner") or {}).get("Name")),
                "ARR (€ converted)": round(
                    as_number(
                        opp.get("ConvertedARR") or opp.get("APTS_Opportunity_ARR__c")
                    ),
                    2,
                ),
                "Old Close": old_close,
                "New Close": new_close,
                "Stage": as_text(opp.get("StageName")),
            }
        )

    forecast_movements: list[dict[str, Any]] = []
    for row in forecast_history:
        opp = row.get("Opportunity") or {}
        forecast_movements.append(
            {
                "Opportunity": as_text(opp.get("Name")),
                "Owner": as_text((opp.get("Owner") or {}).get("Name")),
                "ARR (€ converted)": round(
                    as_number(
                        opp.get("ConvertedARR") or opp.get("APTS_Opportunity_ARR__c")
                    ),
                    2,
                ),
                "Old Category": as_text(row.get("OldValue")),
                "New Category": as_text(row.get("NewValue")),
                "Date": as_text(row.get("CreatedDate"))[:10],
            }
        )

    won_arr = round(
        sum(
            as_number(row.get("ConvertedARR") or row.get("APTS_Opportunity_ARR__c"))
            for row in won_rows
        ),
        2,
    )
    lost_arr = round(
        sum(
            as_number(row.get("ConvertedARR") or row.get("APTS_Opportunity_ARR__c"))
            for row in lost_rows
        ),
        2,
    )
    slipped_arr = round(
        sum(as_number(row.get("ARR (€ converted)")) for row in pushed_deals),
        2,
    )

    return {
        "territory_scope": territory,
        "promise_baseline": promise_baseline,
        "promise_baseline_note": (
            "Derived from ForecastingFact Q1 joined to director-scoped open and Q1 closed opportunities. "
            "Use as the Pipeline Inspection population, not as a quarter-start snapshot."
        ),
        "actuals": {
            "won_count": len(won_rows),
            "won_arr": won_arr,
            "lost_count": len(lost_rows),
            "lost_arr": lost_arr,
            "slipped_count": len(pushed_deals),
            "slipped_arr": slipped_arr,
        },
        "pushed_deals": pushed_deals,
        "forecast_movements": forecast_movements,
        "forecast_movement_summary": summarize_transition_rows(forecast_movements),
    }


def parse_q1_review(
    ws, *, cache_dir: Path | None = None, territory: str = ""
) -> dict[str, Any]:
    header_a = find_row(ws, "Category")
    forecast_vs_actual = table_from_header(ws, header_a) if header_a else []
    summary_metrics = {}
    for row_idx in range(11, min(15, ws.max_row + 1)):
        label = as_text(first_value(ws, row_idx, 1))
        value = ws.cell(row_idx, 2).value
        if label:
            summary_metrics[label] = value
    pushed_row = find_row(ws, "Account")
    pushed_deals = table_from_header(ws, pushed_row) if pushed_row else []
    payload = {
        "title": as_text(ws["A1"].value),
        "workbook_forecast_vs_actual": forecast_vs_actual,
        "workbook_summary_metrics": summary_metrics,
        "workbook_pushed_deals": pushed_deals,
        "scope_warning": (
            "The workbook Q1 forecast and pushed-deals blocks are sourced from global extracts. "
            "Director-scoped Q1 promise and slippage should come from the hidden cache."
        ),
    }
    if cache_dir is not None:
        cache_section = filtered_q1_cache_section(cache_dir, territory)
        payload.update(cache_section)
    else:
        payload.update(
            {
                "promise_baseline": [],
                "actuals": {},
                "pushed_deals": pushed_deals,
                "forecast_movements": [],
                "forecast_movement_summary": [],
            }
        )
    payload["summary_metrics"] = {
        "Won Q1 (count)": payload.get("actuals", {}).get("won_count"),
        "Won Q1 (ARR € converted)": payload.get("actuals", {}).get("won_arr"),
        "Lost Q1 (count)": payload.get("actuals", {}).get("lost_count"),
        "Lost Q1 (ARR € converted)": payload.get("actuals", {}).get("lost_arr"),
        "Slipped out of Q1 (count)": payload.get("actuals", {}).get("slipped_count"),
        "Slipped out of Q1 (ARR € converted)": payload.get("actuals", {}).get(
            "slipped_arr"
        ),
    }
    return payload


def parse_pipeline_detail(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    stage_rollup: dict[str, dict[str, float]] = defaultdict(
        lambda: {"arr": 0.0, "count": 0.0}
    )
    for record in records:
        stage = as_text(record.get("Stage")) or "Unknown"
        stage_rollup[stage]["arr"] += as_number(record.get("ARR (€ converted)"))
        stage_rollup[stage]["count"] += 1
    stage_breakdown = [
        {"stage": stage, "arr": round(values["arr"], 2), "count": int(values["count"])}
        for stage, values in sorted(
            stage_rollup.items(), key=lambda item: (-item[1]["arr"], item[0])
        )
    ]
    top_opportunities = sorted(
        records,
        key=lambda record: as_number(record.get("ARR (€ converted)")),
        reverse=True,
    )[:10]
    return {
        "records": records,
        "stage_breakdown": stage_breakdown,
        "top_opportunities": top_opportunities,
    }


def top_q2_active_opportunities(
    records: list[dict[str, Any]], snapshot_date: str, limit: int = 10
) -> list[dict[str, Any]]:
    start, end = quarter_window(snapshot_date)
    filtered = [
        record
        for record in records
        if is_within_window(record.get("Close Date"), start, end)
        and as_text(record.get("Forecast Category")).lower() != "omitted"
    ]
    filtered.sort(
        key=lambda record: (
            -as_number(record.get("ARR (€ converted)")),
            as_text(record.get("Close Date")),
            as_text(record.get("Opportunity")),
        )
    )
    return filtered[:limit]


def parse_q2_outlook(ws) -> dict[str, Any]:
    breakdown_header = find_row(ws, "Forecast Category")
    breakdown = table_from_header(ws, breakdown_header) if breakdown_header else []

    commit_heading = find_row(ws, "B  Commit Deals (Q2 Close Date)")
    commit_deals = table_from_header(ws, commit_heading + 1) if commit_heading else []

    best_heading = find_row(ws, "C  Best Case Deals (Q2 Close Date)")
    best_case_deals = table_from_header(ws, best_heading + 1) if best_heading else []

    coverage_heading = find_row(ws, "Metric")
    coverage_rows = table_from_header(ws, coverage_heading) if coverage_heading else []
    coverage = {as_text(row.get("Metric")): row.get("Value") for row in coverage_rows}

    by_category = {as_text(row.get("Forecast Category")): row for row in breakdown}
    return {
        "title": as_text(ws["A1"].value),
        "note": as_text(ws["A2"].value),
        "breakdown": breakdown,
        "by_category": by_category,
        "commit_deals": commit_deals,
        "best_case_deals": best_case_deals,
        "coverage": coverage,
    }


def parse_commercial_approval(ws) -> dict[str, Any]:
    summary_header = find_row(ws, "Category")
    summary = table_from_header(ws, summary_header) if summary_header else []

    missing_heading = find_row(
        ws, "B  Missing Approval Candidates (Stage 3+, Land, Not Approved)"
    )
    missing_candidates = (
        table_from_header(ws, missing_heading + 1) if missing_heading else []
    )

    approved_heading = find_row(ws, "C  Approved Deals YTD (2026)")
    approved_ytd = (
        table_from_header(ws, approved_heading + 1) if approved_heading else []
    )

    return {
        "title": as_text(ws["A1"].value),
        "summary": summary,
        "missing_candidates": missing_candidates,
        "approved_ytd": approved_ytd,
    }


def parse_renewals(ws) -> dict[str, Any]:
    kpi_heading = find_row(ws, "KPI")
    kpi_rows = table_from_header(ws, kpi_heading) if kpi_heading else []
    kpis = {as_text(row.get("KPI")): row.get("Value") for row in kpi_rows}

    risk_heading = find_row(ws, "Risk Level")
    risk_levels = table_from_header(ws, risk_heading) if risk_heading else []

    open_renewals_heading = find_row(ws, "Account")
    open_renewals: list[dict[str, Any]] = []
    if open_renewals_heading:
        row_after = row_values(ws, open_renewals_heading + 1, 2)
        if as_text(row_after[0]) != "No open renewals found.":
            open_renewals = table_from_header(ws, open_renewals_heading)

    return {
        "title": as_text(ws["A1"].value),
        "kpis": kpis,
        "risk_levels": risk_levels,
        "open_renewals": open_renewals,
    }


def enrich_renewals(renewals: dict[str, Any], snapshot_date: str) -> dict[str, Any]:
    rows = list((renewals.get("open_renewals") or []))
    start, end = quarter_window(snapshot_date)
    q2_rows = [
        row for row in rows if is_within_window(row.get("Close Date"), start, end)
    ]
    q2_rows.sort(
        key=lambda row: (
            -as_number(row.get("Renewal ACV (€ converted)")),
            as_text(row.get("Close Date")),
            as_text(row.get("Opportunity")),
        )
    )
    renewals["summary_metrics"] = {
        "open_deal_count": len(rows),
        "open_acv": round(
            sum(as_number(row.get("Renewal ACV (€ converted)")) for row in rows),
            2,
        ),
        "q2_open_deal_count": len(q2_rows),
        "q2_open_acv": round(
            sum(as_number(row.get("Renewal ACV (€ converted)")) for row in q2_rows),
            2,
        ),
    }
    renewals["q2_open_renewals"] = q2_rows[:10]
    return renewals


def parse_rep_performance(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    top_reps = sorted(
        records,
        key=lambda record: as_number(record.get("Open Pipeline ARR (€ converted)")),
        reverse=True,
    )[:8]
    return {"records": records, "top_reps": top_reps}


def parse_won_lost(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    won = [
        record for record in records if as_text(record.get("Status")).lower() == "won"
    ]
    lost = [
        record for record in records if as_text(record.get("Status")).lower() == "lost"
    ]
    return {"records": records, "won": won, "lost": lost}


def parse_risk_register(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    top_arr = sorted(
        records,
        key=lambda record: as_number(record.get("ARR (€ converted)")),
        reverse=True,
    )[:10]
    return {"records": records, "top_arr": top_arr}


def parse_data_quality(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    total = next(
        (record for record in records if as_text(record.get("Rep")).upper() == "TOTAL"),
        None,
    )
    if total is None:
        total = {
            "Rep": "TOTAL",
            "Stale 30d": sum(as_number(record.get("Stale 30d")) for record in records),
            "No Activity": sum(
                as_number(record.get("No Activity")) for record in records
            ),
            "Overdue Close": sum(
                as_number(record.get("Overdue Close")) for record in records
            ),
            "Missing Amount": sum(
                as_number(record.get("Missing Amount")) for record in records
            ),
            "Missing Next Step": sum(
                as_number(record.get("Missing Next Step")) for record in records
            ),
            "Missing Approval": sum(
                as_number(record.get("Missing Approval")) for record in records
            ),
            "Aging 365+": sum(
                as_number(record.get("Aging 365+")) for record in records
            ),
        }
    ranked = [
        record for record in records if as_text(record.get("Rep")).upper() != "TOTAL"
    ]
    ranked.sort(
        key=lambda record: (
            as_number(record.get("Stale 30d")) + as_number(record.get("No Activity")),
            as_number(record.get("Aging 365+")),
        ),
        reverse=True,
    )
    return {"records": records, "total": total, "top_issues": ranked[:8]}


def parse_sources(ws) -> dict[str, Any]:
    records = table_from_header(ws, 1)
    return {"records": records, "count": len(records)}


def build_factual_bullets(snapshot: dict[str, Any]) -> list[str]:
    scorecard = snapshot["scorecard"]["sections"]
    pipeline = scorecard.get("pipeline-health", {}).get("metrics", {})
    risk = scorecard.get("risk", {}).get("metrics", {})
    process = scorecard.get("process-compliance", {}).get("metrics", {})
    q2 = snapshot["q2_outlook"]["by_category"]
    top_open = snapshot["pipeline_detail"]["top_opportunities"][:1]
    top_risk = snapshot["risk_register"]["top_arr"][:1]

    bullets = [
        (
            f"All-open ARR {pipeline.get('Pipeline ARR — All Open (any close date)', '—')}; "
            f"FY26 close-date ARR {pipeline.get('Pipeline ARR — FY26 Close Dates Only (excl. Omitted)', '—')}; "
            f"Q2 active ARR {pipeline.get('Pipeline ARR — Q2 2026 Close Dates Only (excl. Omitted)', '—')}."
        ),
        (
            f"Q2 mix: Commit {compact_money(as_number(q2.get('Commit', {}).get('ARR (€ converted)')))}, "
            f"Best Case {compact_money(as_number(q2.get('Best Case', {}).get('ARR (€ converted)')))}, "
            f"Pipeline {compact_money(as_number(q2.get('Pipeline', {}).get('ARR (€ converted)')))}, "
            f"Omitted {compact_money(as_number(q2.get('Omitted', {}).get('ARR (€ converted)')))}."
        ),
        (
            f"Risk controls: stale ARR {risk.get('Stale 30d+ (ARR)', '—')}, "
            f"pushed 5+ ARR {risk.get('Pushed 5+ (ARR)', '—')}, "
            f"approval rate {process.get('Approval Rate (stage 3+)', '—')}."
        ),
    ]
    if top_open:
        record = top_open[0]
        bullets.append(
            f"Largest open opportunity: {as_text(record.get('Opportunity'))} at "
            f"{compact_money(as_number(record.get('ARR (€ converted)')))} "
            f"({as_text(record.get('Stage'))}, close {as_text(record.get('Close Date'))})."
        )
    if top_risk:
        record = top_risk[0]
        bullets.append(
            f"Highest-value risk-register deal: {as_text(record.get('Opportunity'))} at "
            f"{compact_money(as_number(record.get('ARR (€ converted)')))} with "
            f"{int(as_number(record.get('Push Count')))} pushes and "
            f"{int(as_number(record.get('Activity Days Ago')))} days since activity."
        )
    return bullets[:4]


def stage_number(stage_value: Any) -> int:
    match = re.match(r"^\s*(\d+)", as_text(stage_value))
    return int(match.group(1)) if match else 0


def iso_date(value: Any) -> str:
    token = as_text(value)[:10]
    return token if re.fullmatch(r"\d{4}-\d{2}-\d{2}", token) else ""


def days_since(snapshot_date: str, date_value: Any) -> int:
    token = iso_date(date_value)
    if not token:
        return 0
    try:
        return (date.fromisoformat(snapshot_date) - date.fromisoformat(token)).days
    except ValueError:
        return 0


def rows_from_sheet(ws) -> list[dict[str, Any]]:
    return table_from_header(ws, 1)


def parse_live_summary(ws) -> dict[str, Any]:
    title = as_text(ws["A1"].value)
    generated = (
        as_text(ws["A3"].value).replace("Snapshot date:", "").split("—", 1)[0].strip()
    )
    kpi_row = find_row(ws, "KPI")
    kpi_rows = table_from_header(ws, kpi_row, max_col=2) if kpi_row else []
    source_row = find_row(ws, "Sheet")
    source_rows = table_from_header(ws, source_row, max_col=3) if source_row else []
    return {
        "title": title,
        "generated": generated,
        "kpis": {as_text(row.get("KPI")): row.get("Value") for row in kpi_rows},
        "sources": source_rows,
    }


def normalize_live_pipeline_rows(
    rows: list[dict[str, Any]],
    *,
    snapshot_date: str,
    territory: str,
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        arr = round(as_number(row.get("ARR Unweighted (EUR)")), 2)
        forecast_arr = round(as_number(row.get("ARR Weighted (EUR)")), 2)
        probability = as_number(row.get("Probability %"))
        push_count = int(round(as_number(row.get("Push Count"))))
        last_activity = iso_date(row.get("Last Activity"))
        normalized.append(
            {
                "Account": as_text(row.get("Account")),
                "Opportunity": as_text(row.get("Opportunity")),
                "Owner": as_text(row.get("Owner")),
                "Stage": as_text(row.get("Stage")),
                "Close Date": iso_date(row.get("Close Date")),
                "ARR (€ converted)": arr,
                "ACV (€ converted)": arr,
                "Forecast ARR (€ converted)": forecast_arr,
                "Forecast Category": as_text(row.get("Forecast Category")),
                "Probability (%)": probability,
                "Type": as_text(row.get("Type")),
                "Sub-Type": "",
                "Push Count": push_count,
                "Age (Days)": 0,
                "Days In Stage": 0,
                "Last Activity": last_activity,
                "Activity Days Ago": days_since(snapshot_date, last_activity),
                "Risk Level": "",
                "Approval": as_text(row.get("Approved")) or "No",
                "Approval Status": "",
                "Next Step": as_text(row.get("Next Step")),
                "Director Book": territory,
                "Region": as_text(row.get("Sales Region")) or territory,
                "Industry": as_text(row.get("Industry")),
                "Lead Scope": as_text(row.get("Lead Scope")),
                "Tier": as_text(row.get("Tier")),
                "Created": iso_date(row.get("Created")),
                "Last Modified": iso_date(row.get("Last Modified")),
                "Competitor": as_text(row.get("Competitor")),
            }
        )
    return normalized


def parse_pipeline_detail_rows(records: list[dict[str, Any]]) -> dict[str, Any]:
    stage_rollup: dict[str, dict[str, float]] = defaultdict(
        lambda: {"arr": 0.0, "count": 0.0}
    )
    for record in records:
        stage = as_text(record.get("Stage")) or "Unknown"
        stage_rollup[stage]["arr"] += as_number(record.get("ARR (€ converted)"))
        stage_rollup[stage]["count"] += 1
    stage_breakdown = [
        {
            "stage": stage,
            "arr": round(values["arr"], 2),
            "count": int(values["count"]),
        }
        for stage, values in sorted(
            stage_rollup.items(), key=lambda item: (-item[1]["arr"], item[0])
        )
    ]
    top_opportunities = sorted(
        records,
        key=lambda record: as_number(record.get("ARR (€ converted)")),
        reverse=True,
    )[:10]
    return {
        "records": records,
        "stage_breakdown": stage_breakdown,
        "top_opportunities": top_opportunities,
    }


def build_live_q2_outlook(
    records: list[dict[str, Any]],
    snapshot_date: str,
) -> dict[str, Any]:
    period = resolve_period_context(
        as_of_date=snapshot_date,
        snapshot_date=snapshot_date,
        deck_date=snapshot_date,
    )
    start = period.current_quarter.start_date
    end = period.current_quarter.end_date
    window_rows = [
        record
        for record in records
        if is_within_window(record.get("Close Date"), start, end)
    ]
    breakdown_rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"Forecast Category": "", "Deal Count": 0, "ARR (€ converted)": 0.0}
    )
    for record in window_rows:
        category = as_text(record.get("Forecast Category")) or "Unspecified"
        bucket = breakdown_rollup[category]
        bucket["Forecast Category"] = category
        bucket["Deal Count"] += 1
        bucket["ARR (€ converted)"] += as_number(record.get("ARR (€ converted)"))
    breakdown = list(breakdown_rollup.values())
    breakdown.sort(
        key=lambda row: (
            -as_number(row.get("ARR (€ converted)")),
            row["Forecast Category"],
        )
    )
    for row in breakdown:
        row["ARR (€ converted)"] = round(as_number(row.get("ARR (€ converted)")), 2)
    return {
        "title": f"{period.current_quarter.title} Outlook",
        "note": (
            f"{period.current_quarter.title} = {start} to {end}. "
            "Excludes opportunities in Omitted forecast category."
        ),
        "breakdown": breakdown,
        "by_category": {
            as_text(row.get("Forecast Category")): row for row in breakdown
        },
        "commit_deals": [
            record
            for record in window_rows
            if as_text(record.get("Forecast Category")).lower() == "commit"
        ],
        "best_case_deals": [
            record
            for record in window_rows
            if as_text(record.get("Forecast Category")).lower()
            in {"best case", "bestcase"}
        ],
        "coverage": {},
        "top_q2_active_opportunities": top_q2_active_opportunities(
            records, snapshot_date
        ),
    }


def build_live_scorecard_sections(
    *,
    snapshot_date: str,
    territory: str,
    pipeline_rows: list[dict[str, Any]],
    commercial_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    period = resolve_period_context(
        as_of_date=snapshot_date,
        snapshot_date=snapshot_date,
        deck_date=snapshot_date,
    )
    fy_active_rows = [
        row
        for row in pipeline_rows
        if as_text(row.get("Forecast Category")).lower() != "omitted"
    ]
    q_rows = [
        row
        for row in fy_active_rows
        if is_within_window(
            row.get("Close Date"),
            period.current_quarter.start_date,
            period.current_quarter.end_date,
        )
    ]
    stage3plus = [row for row in commercial_rows if stage_number(row.get("Stage")) >= 3]
    approved = [
        row
        for row in stage3plus
        if as_text(row.get("Status")).lower().startswith("approved")
    ]
    missing_or_pending = [
        row
        for row in stage3plus
        if any(
            token in as_text(row.get("Status")).lower()
            for token in ("missing", "pending")
        )
    ]
    stale_rows = [
        row for row in pipeline_rows if as_number(row.get("Activity Days Ago")) >= 30
    ]
    pushed_rows = [
        row for row in pipeline_rows if as_number(row.get("Push Count")) >= 5
    ]
    return {
        "pipeline-health": {
            "title": "Pipeline Health",
            "metrics": {
                "Pipeline ARR — All Open (any close date)": compact_money(
                    sum(
                        as_number(row.get("ARR (€ converted)")) for row in pipeline_rows
                    )
                ),
                "Pipeline ARR — FY26 Close Dates Only (excl. Omitted)": compact_money(
                    sum(
                        as_number(row.get("ARR (€ converted)"))
                        for row in fy_active_rows
                    )
                ),
                (
                    f"Pipeline ARR — {period.current_quarter.label} {period.current_quarter.year} "
                    "Close Dates Only (excl. Omitted)"
                ): compact_money(
                    sum(as_number(row.get("ARR (€ converted)")) for row in q_rows)
                ),
                "Territory": territory,
            },
        },
        "process-compliance": {
            "title": "Process Compliance",
            "metrics": {
                "Approval Rate (stage 3+)": (
                    f"{(len(approved) / len(stage3plus) * 100):.0f}%"
                    if stage3plus
                    else "0%"
                ),
                "Missing Approval (stage 3+)": len(missing_or_pending),
                "Approved Deals (stage 3+)": len(approved),
            },
        },
        "risk": {
            "title": "Risk",
            "metrics": {
                "Stale 30d+ (ARR)": compact_money(
                    sum(as_number(row.get("ARR (€ converted)")) for row in stale_rows)
                ),
                "Pushed 5+ (ARR)": compact_money(
                    sum(as_number(row.get("ARR (€ converted)")) for row in pushed_rows)
                ),
                "Stale 30d+ (count)": len(stale_rows),
                "Pushed 5+ (count)": len(pushed_rows),
            },
        },
    }


def parse_live_commercial_approval(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    summary_rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"Category": "", "Deal Count": 0, "ARR (€ converted)": 0.0}
    )
    missing_candidates: list[dict[str, Any]] = []
    approved_ytd: list[dict[str, Any]] = []
    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        normalized = {
            "Account": as_text(row.get("Account")),
            "Opportunity": as_text(row.get("Opportunity")),
            "Owner": as_text(row.get("Owner")),
            "Stage": as_text(row.get("Stage")),
            "Close Date": iso_date(row.get("Close Date")),
            "ARR (€ converted)": round(as_number(row.get("ARR Unweighted (EUR)")), 2),
            "Status": as_text(row.get("Status")),
            "Approval Date": iso_date(row.get("Approval Date")),
            "Next Step": as_text(row.get("Next Step")),
            "Lead Scope": as_text(row.get("Lead Scope")),
        }
        normalized_rows.append(normalized)
        status = normalized["Status"].lower()
        if status.startswith("approved"):
            category = "Approved"
            approved_ytd.append(normalized)
        elif "pending" in status or "missing" in status:
            category = "Pending / Missing Approval"
            missing_candidates.append(normalized)
        else:
            category = "No Approval Needed"
        bucket = summary_rollup[category]
        bucket["Category"] = category
        bucket["Deal Count"] += 1
        bucket["ARR (€ converted)"] += as_number(normalized["ARR (€ converted)"])
    summary = list(summary_rollup.values())
    summary.sort(
        key=lambda row: (-as_number(row.get("ARR (€ converted)")), row["Category"])
    )
    for row in summary:
        row["ARR (€ converted)"] = round(as_number(row.get("ARR (€ converted)")), 2)
    return {
        "title": "Commercial Approval",
        "summary": summary,
        "missing_candidates": missing_candidates,
        "approved_ytd": approved_ytd,
        "records": normalized_rows,
    }


def parse_live_renewals(
    rows: list[dict[str, Any]],
    snapshot_date: str,
) -> dict[str, Any]:
    normalized_rows = [
        {
            "Close Date": iso_date(row.get("Close Date")),
            "Account": as_text(row.get("Account")),
            "Opportunity": as_text(row.get("Opportunity")),
            "Owner": as_text(row.get("Owner")),
            "Stage": as_text(row.get("Stage")),
            "Renewal ACV (€ converted)": round(
                as_number(row.get("ACV Unweighted (EUR)")), 2
            ),
            "Probability (%)": as_number(row.get("Probability %")),
            "Comments": as_text(row.get("Comments")),
        }
        for row in rows
    ]
    renewals = {
        "title": "Renewals & Retention",
        "kpis": {},
        "risk_levels": [],
        "open_renewals": normalized_rows,
    }
    return enrich_renewals(renewals, snapshot_date)


def parse_live_won_lost(rows: list[dict[str, Any]]) -> dict[str, Any]:
    records = [
        {
            "Account": as_text(row.get("Account")),
            "Opportunity": as_text(row.get("Opportunity")),
            "Owner": as_text(row.get("Owner")),
            "Stage": as_text(row.get("Stage")),
            "Close Date": iso_date(row.get("Close Date")),
            "ARR (€ converted)": round(as_number(row.get("ARR Unweighted (EUR)")), 2),
            "Type": as_text(row.get("Type")),
            "Reason Won/Lost": as_text(row.get("Reason")),
            "Lost To Competitor": as_text(row.get("Lost To Competitor")),
            "Industry": as_text(row.get("Industry")),
            "Sales Region": as_text(row.get("Sales Region")),
            "Created": iso_date(row.get("Created")),
            "Status": "won"
            if stage_number(row.get("Stage")) == 8
            else "lost"
            if stage_number(row.get("Stage")) in {0, 7}
            else "",
        }
        for row in rows
    ]
    won = [record for record in records if record.get("Status") == "won"]
    lost = [record for record in records if record.get("Status") == "lost"]
    return {"records": records, "won": won, "lost": lost}


def parse_live_rep_performance(
    pipeline_rows: list[dict[str, Any]],
    won_lost: dict[str, Any],
) -> dict[str, Any]:
    rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "Rep": "",
            "Open Pipeline ARR (€ converted)": 0.0,
            "Deal Count": 0,
            "Avg Deal Size (€)": 0.0,
            "Won ARR Q (€ converted)": 0.0,
            "Lost ARR Q (€ converted)": 0.0,
            "Win Rate %": 0.0,
            "Stale Deals": 0,
            "Pushed Deals": 0,
            "Missing Approvals": 0,
        }
    )
    for row in pipeline_rows:
        owner = as_text(row.get("Owner")) or "Unknown"
        item = rollup[owner]
        item["Rep"] = owner
        item["Open Pipeline ARR (€ converted)"] += as_number(
            row.get("ARR (€ converted)")
        )
        item["Deal Count"] += 1
        item["Stale Deals"] += int(as_number(row.get("Activity Days Ago")) >= 30)
        item["Pushed Deals"] += int(as_number(row.get("Push Count")) > 0)
        item["Missing Approvals"] += int(
            stage_number(row.get("Stage")) >= 3
            and as_text(row.get("Approval")).lower() != "yes"
        )
    for row in won_lost.get("won", []):
        owner = as_text(row.get("Owner")) or "Unknown"
        item = rollup[owner]
        item["Rep"] = owner
        item["Won ARR Q (€ converted)"] += as_number(row.get("ARR (€ converted)"))
    for row in won_lost.get("lost", []):
        owner = as_text(row.get("Owner")) or "Unknown"
        item = rollup[owner]
        item["Rep"] = owner
        item["Lost ARR Q (€ converted)"] += as_number(row.get("ARR (€ converted)"))
    records = list(rollup.values())
    for row in records:
        count = int(row.get("Deal Count") or 0)
        row["Open Pipeline ARR (€ converted)"] = round(
            as_number(row.get("Open Pipeline ARR (€ converted)")), 2
        )
        row["Avg Deal Size (€)"] = (
            round(row["Open Pipeline ARR (€ converted)"] / count, 2) if count else 0.0
        )
        won_count = sum(
            1
            for item in won_lost.get("won", [])
            if as_text(item.get("Owner")) == row["Rep"]
        )
        lost_count = sum(
            1
            for item in won_lost.get("lost", [])
            if as_text(item.get("Owner")) == row["Rep"]
        )
        total_closed = won_count + lost_count
        row["Win Rate %"] = (
            round((won_count / total_closed * 100.0), 2) if total_closed else 0.0
        )
        row["Won ARR Q (€ converted)"] = round(
            as_number(row.get("Won ARR Q (€ converted)")), 2
        )
        row["Lost ARR Q (€ converted)"] = round(
            as_number(row.get("Lost ARR Q (€ converted)")), 2
        )
    top_reps = sorted(
        records,
        key=lambda record: as_number(record.get("Open Pipeline ARR (€ converted)")),
        reverse=True,
    )[:8]
    return {"records": records, "top_reps": top_reps}


def parse_live_data_quality(rows: list[dict[str, Any]]) -> dict[str, Any]:
    records_rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "Rep": "",
            "Stale 30d": 0,
            "No Activity": 0,
            "Overdue Close": 0,
            "Missing Amount": 0,
            "Missing Next Step": 0,
            "Missing Approval": 0,
            "Aging 365+": 0,
            "Total Issues": 0,
        }
    )
    for row in rows:
        owner = as_text(row.get("Owner")) or "Unknown"
        item = records_rollup[owner]
        item["Rep"] = owner
        flag = as_text(row.get("Flag")).lower()
        if "no touch" in flag or "no activity" in flag:
            item["No Activity"] += 1
            item["Total Issues"] += 1
    records = list(records_rollup.values())
    total = {
        "Rep": "TOTAL",
        "Stale 30d": sum(as_number(record.get("Stale 30d")) for record in records),
        "No Activity": sum(as_number(record.get("No Activity")) for record in records),
        "Overdue Close": sum(
            as_number(record.get("Overdue Close")) for record in records
        ),
        "Missing Amount": sum(
            as_number(record.get("Missing Amount")) for record in records
        ),
        "Missing Next Step": sum(
            as_number(record.get("Missing Next Step")) for record in records
        ),
        "Missing Approval": sum(
            as_number(record.get("Missing Approval")) for record in records
        ),
        "Aging 365+": sum(as_number(record.get("Aging 365+")) for record in records),
        "Total Issues": sum(
            as_number(record.get("Total Issues")) for record in records
        ),
    }
    records.sort(key=lambda record: as_number(record.get("Total Issues")), reverse=True)
    return {"records": records, "total": total, "top_issues": records[:8]}


def parse_live_q1_review(
    *,
    q1_rows: list[dict[str, Any]],
    forecast_history_rows: list[dict[str, Any]],
    won_lost: dict[str, Any],
) -> dict[str, Any]:
    pushed_deals = [
        {
            "Account": as_text(row.get("Account")),
            "Opportunity": as_text(row.get("Opportunity")),
            "Owner": as_text(row.get("Owner")),
            "ARR (€ converted)": round(as_number(row.get("ARR Unweighted (EUR)")), 2),
            "Old Close": iso_date(row.get("Old Close")),
            "New Close": iso_date(row.get("New Close")),
            "Stage": as_text(row.get("Stage")),
        }
        for row in q1_rows
        if "q1 slipped" in as_text(row.get("Movement")).lower()
    ]
    forecast_movements = [
        {
            "Opportunity": as_text(row.get("Opportunity")),
            "Owner": as_text(row.get("Owner")),
            "ARR (€ converted)": round(as_number(row.get("ARR Unweighted (EUR)")), 2),
            "Old Category": as_text(row.get("From Category")),
            "New Category": as_text(row.get("To Category")),
            "Date": iso_date(row.get("Changed On")),
        }
        for row in forecast_history_rows
    ]
    won_q1 = [
        row
        for row in won_lost.get("won", [])
        if row.get("Close Date", "").startswith(("2026-01", "2026-02", "2026-03"))
    ]
    lost_q1 = [
        row
        for row in won_lost.get("lost", [])
        if row.get("Close Date", "").startswith(("2026-01", "2026-02", "2026-03"))
    ]
    actuals = {
        "won_count": len(won_q1),
        "won_arr": round(
            sum(as_number(row.get("ARR (€ converted)")) for row in won_q1), 2
        ),
        "lost_count": len(lost_q1),
        "lost_arr": round(
            sum(as_number(row.get("ARR (€ converted)")) for row in lost_q1), 2
        ),
        "slipped_count": len(pushed_deals),
        "slipped_arr": round(
            sum(as_number(row.get("ARR (€ converted)")) for row in pushed_deals), 2
        ),
    }
    return {
        "title": "Q1 Review",
        "workbook_forecast_vs_actual": [],
        "workbook_summary_metrics": {},
        "workbook_pushed_deals": pushed_deals,
        "scope_warning": (
            "Live workbook contract uses movement tabs rather than the retired "
            "Q1 review workbook blocks; promise baseline is unavailable here."
        ),
        "territory_scope": "",
        "promise_baseline": [],
        "promise_baseline_note": "Not available in the live workbook contract.",
        "actuals": actuals,
        "pushed_deals": pushed_deals,
        "forecast_movements": forecast_movements,
        "forecast_movement_summary": summarize_transition_rows(forecast_movements),
        "summary_metrics": {
            "Won Q1 (count)": actuals["won_count"],
            "Won Q1 (ARR € converted)": actuals["won_arr"],
            "Lost Q1 (count)": actuals["lost_count"],
            "Lost Q1 (ARR € converted)": actuals["lost_arr"],
            "Slipped out of Q1 (count)": actuals["slipped_count"],
            "Slipped out of Q1 (ARR € converted)": actuals["slipped_arr"],
        },
    }


def parse_live_sources(summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "records": summary.get("sources") or [],
        "count": len(summary.get("sources") or []),
    }


def parse_live_risk_register(records: list[dict[str, Any]]) -> dict[str, Any]:
    ranked = sorted(
        records,
        key=lambda record: (
            as_number(record.get("Push Count")),
            as_number(record.get("Activity Days Ago")),
            as_number(record.get("ARR (€ converted)")),
        ),
        reverse=True,
    )
    return {"records": ranked, "top_arr": ranked[:10]}


def extract_live_workbook(workbook_path: Path, wb) -> dict[str, Any]:
    SN = sheet_names()
    summary = parse_live_summary(wb["Summary"])
    title = summary["title"]
    match = re.match(r"(.+) \((.+)\)", title)
    director_name = match.group(1) if match else workbook_path.stem
    territory = match.group(2) if match else ""
    snapshot_date = summary["generated"]

    pipeline_rows = normalize_live_pipeline_rows(
        rows_from_sheet(wb[SN["pipeline_open"]]),
        snapshot_date=snapshot_date,
        territory=territory,
    )
    pipeline_detail = parse_pipeline_detail_rows(pipeline_rows)
    q2_outlook = build_live_q2_outlook(pipeline_rows, snapshot_date)
    commercial_approval = parse_live_commercial_approval(
        rows_from_sheet(wb["Commercial Approval"])
    )
    renewals = parse_live_renewals(rows_from_sheet(wb[SN["renewals"]]), snapshot_date)
    won_lost = parse_live_won_lost(rows_from_sheet(wb[SN["won_lost"]]))

    snapshot = {
        "director_name": director_name,
        "territory": territory,
        "snapshot_date": snapshot_date,
        "workbook_path": str(workbook_path),
        "workbook_file": workbook_path.name,
        "generated_at": datetime.now(UTC).isoformat(),
        "scorecard": {
            "title": f"Sales Director Scorecard — {title}",
            "generated": snapshot_date,
            "sections": build_live_scorecard_sections(
                snapshot_date=snapshot_date,
                territory=territory,
                pipeline_rows=pipeline_rows,
                commercial_rows=commercial_approval["records"],
            ),
        },
        "pipeline_detail": pipeline_detail,
        "q1_review": parse_live_q1_review(
            q1_rows=rows_from_sheet(wb["Q1 Movement"]),
            forecast_history_rows=rows_from_sheet(wb["Forecast Category History"]),
            won_lost=won_lost,
        ),
        "rep_performance": parse_live_rep_performance(pipeline_rows, won_lost),
        "won_lost": won_lost,
        "sources": parse_live_sources(summary),
        "q2_outlook": q2_outlook,
        "commercial_approval": commercial_approval,
        "renewals": renewals,
        "risk_register": parse_live_risk_register(pipeline_rows),
        "data_quality": parse_live_data_quality(rows_from_sheet(wb["Activity Volume"])),
        "quota_targets_note": "",
    }
    snapshot["quarterly_pipeline_display"] = quarterly_pipeline_display_from_snapshot(
        snapshot
    )
    snapshot["factual_bullets"] = build_factual_bullets(snapshot)
    return snapshot


def extract_workbook(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)
    if "Scorecard" not in wb.sheetnames and "Summary" in wb.sheetnames:
        return extract_live_workbook(workbook_path, wb)
    scorecard = parse_scorecard(wb["Scorecard"])
    title = scorecard["title"]
    match = re.match(r"Sales Director Scorecard — (.+) \((.+)\)", title)
    director_name = match.group(1) if match else workbook_path.stem
    territory = match.group(2) if match else ""

    cache_dir = workbook_path.parent / ".cache" / cache_slugify(director_name)
    pipeline_detail = parse_pipeline_detail(wb["Pipeline Detail"])
    q2_outlook = parse_q2_outlook(wb["Q2 Outlook"])
    q2_outlook["top_q2_active_opportunities"] = top_q2_active_opportunities(
        pipeline_detail["records"], scorecard["generated"]
    )
    renewals = enrich_renewals(
        parse_renewals(wb["Renewals & Retention"]), scorecard["generated"]
    )

    snapshot = {
        "director_name": director_name,
        "territory": territory,
        "snapshot_date": scorecard["generated"],
        "workbook_path": str(workbook_path),
        "workbook_file": workbook_path.name,
        "generated_at": datetime.now(UTC).isoformat(),
        "scorecard": scorecard,
        "pipeline_detail": pipeline_detail,
        "q1_review": parse_q1_review(
            wb["Q1 Review"],
            cache_dir=cache_dir if cache_dir.exists() else None,
            territory=territory,
        ),
        "rep_performance": parse_rep_performance(wb["Rep Performance"]),
        "won_lost": parse_won_lost(wb["Won-Lost"]),
        "sources": parse_sources(wb["Sources & Lineage"]),
        "q2_outlook": q2_outlook,
        "commercial_approval": parse_commercial_approval(wb["Commercial Approval"]),
        "renewals": renewals,
        "risk_register": parse_risk_register(wb["Risk Register"]),
        "data_quality": parse_data_quality(wb["Data Quality"]),
        "quota_targets_note": as_text(wb["Quota & Targets"]["A3"].value),
    }
    snapshot["quarterly_pipeline_display"] = quarterly_pipeline_display_from_snapshot(
        snapshot
    )
    snapshot["factual_bullets"] = build_factual_bullets(snapshot)
    return snapshot


def workbook_matches_director(path: Path, director: str) -> bool:
    legacy_prefix = f"Sales Director Data - {director} "
    if legacy_prefix in path.name:
        return True
    return slugify(path.stem) == slugify(director)


def workbook_paths(root: Path, snapshot_date: str, director: str | None) -> list[Path]:
    base = root / snapshot_date
    paths_by_name: dict[str, Path] = {}
    for pattern in ("Sales Director Data - *.xlsx", "*.xlsx"):
        for path in base.glob(pattern):
            if path.name.startswith("~$"):
                continue
            paths_by_name[path.name] = path
    paths = sorted(paths_by_name.values(), key=lambda path: path.name.lower())
    if director:
        paths = [path for path in paths if workbook_matches_director(path, director)]
    return paths


def run(
    snapshot_date: str, director: str | None, workbook_root: Path, output_root: Path
) -> list[Path]:
    out_dir = output_root / snapshot_date
    out_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    for workbook_path in workbook_paths(workbook_root, snapshot_date, director):
        snapshot = extract_workbook(workbook_path)
        out_path = out_dir / f"{slugify(snapshot['director_name'])}.json"
        out_path.write_text(
            json.dumps(snapshot, indent=2, ensure_ascii=True), encoding="utf-8"
        )
        outputs.append(out_path)
        print(out_path)
    return outputs


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--snapshot-date", default="2026-04-10")
    parser.add_argument("--director")
    parser.add_argument("--workbook-root", type=Path, default=DEFAULT_WORKBOOK_ROOT)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    outputs = run(
        args.snapshot_date, args.director, args.workbook_root, args.output_root
    )
    if not outputs:
        raise SystemExit("No workbooks matched the requested scope.")


if __name__ == "__main__":
    main()
