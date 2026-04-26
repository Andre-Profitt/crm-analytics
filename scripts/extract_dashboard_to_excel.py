#!/usr/bin/env python3
"""Extract live Salesforce data via SAQL + SOQL into an Excel workbook.

Mirrors the Pipeline Reporting & Insights dashboard queries with a
hardcoded territory filter. The Excel workbook becomes the editable
source of truth for the SimCorp director deck.

Usage:
    python3 scripts/extract_dashboard_to_excel.py \
        --territory APAC \
        --snapshot-date 2026-04-10 \
        [--output output/director_data_workbooks/2026-04-10/jesper-tyrer.xlsx]
"""

from __future__ import annotations

import argparse
import json
import subprocess
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]

# Territory definitions: map territory name → SAQL filter expressions
TERRITORIES = {
    "APAC": {
        "director": "Jesper Tyrer",
        "unit_group_filter": 'AccountUnitGroup == "SC Asia"',
        "soql_filter": "Account_Unit_Group__c = 'SC Asia'",
        "pi_list_view_id": "00BTb00000Ic7kTMAR",
    },
    "Central Europe": {
        "director": "Sarah Pittroff",
        "unit_group_filter": 'AccountUnitGroup == "SC EMEA" && SalesRegion == "Central Europe"',
        "soql_filter": "Account_Unit_Group__c = 'SC EMEA' AND Sales_Region__c = 'Central Europe'",
        "pi_list_view_id": "00BTb00000Kr3YvMAJ",
    },
    "UK & Ireland": {
        "director": "Dan Peppett",
        "unit_group_filter": 'AccountUnitGroup == "SC EMEA" && SalesRegion == "United Kingdom & Ireland"',
        "soql_filter": "Account_Unit_Group__c = 'SC EMEA' AND Sales_Region__c = 'United Kingdom & Ireland'",
        "pi_list_view_id": "00BTb00000Kr3yjMAB",
    },
    "Southern Europe": {
        "director": "Francois Thaury",
        "unit_group_filter": 'AccountUnitGroup == "SC EMEA" && SalesRegion == "Southwestern Europe"',
        "soql_filter": "Account_Unit_Group__c = 'SC EMEA' AND Sales_Region__c = 'Southwestern Europe'",
        "pi_list_view_id": "00BTb00000Kr3sHMAR",
    },
    "NL & Nordics": {
        "director": "Christian Ebbesen",
        "unit_group_filter": 'AccountUnitGroup == "SC EMEA" && SalesRegion == "Northern Europe"',
        "soql_filter": "Account_Unit_Group__c = 'SC EMEA' AND Sales_Region__c = 'Northern Europe'",
        "pi_list_view_id": "00BTb00000Kr4DFMAZ",
    },
    "Middle East & Africa": {
        "director": "Mourad Essofi",
        "unit_group_filter": 'AccountUnitGroup == "SC EMEA" && SalesRegion == "Middle East & Africa"',
        "soql_filter": "Account_Unit_Group__c = 'SC EMEA' AND Sales_Region__c = 'Middle East & Africa'",
        "pi_list_view_id": "00BQA00000GXOf32AH",
    },
    "Canada": {
        "director": "Megan Miceli",
        "unit_group_filter": 'AccountUnitGroup == "SC North America" && AccountUnit == "SC Canada"',
        "soql_filter": "Account_Unit_Group__c = 'SC North America' AND Account.Unit__c = 'SC Canada'",
        "pi_list_view_id": "00BTb00000Kr4ErMAJ",
    },
    "NA Asset Management": {
        "director": "Patrick Gaughan",
        "unit_group_filter": 'AccountUnitGroup == "SC North America" && AccountUnit == "SC USA"',
        "soql_filter": "Account_Unit_Group__c = 'SC North America' AND Account.Unit__c = 'SC USA'",
        "pi_list_view_id": "00BTb00000Kr4JhMAJ",
    },
    "Pension & Insurance": {
        "director": "Adam Steinhaus",
        "unit_group_filter": 'AccountUnitGroup == "SC North America"',
        "soql_filter": "Account_Unit_Group__c = 'SC North America'",
        "pi_list_view_id": "00BTb00000Kr4OXMAZ",
    },
}

HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=9)

# Internal account patterns to exclude
_INTERNAL = ["SimCorp", "Test Account", "Sandbox", "Demo"]


def _is_internal(name: str) -> bool:
    name_lower = (name or "").lower()
    return any(p.lower() in name_lower for p in _INTERNAL)


def get_auth() -> tuple[str, str]:
    result = subprocess.run(
        ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
        capture_output=True,
        text=True,
        check=True,
    )
    data = json.loads(result.stdout)["result"]
    return data["accessToken"], data["instanceUrl"]


def run_saql(token: str, instance_url: str, query: str) -> list[dict]:
    resp = requests.post(
        f"{instance_url}/services/data/v66.0/wave/query",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json={"query": query},
        timeout=60,
    )
    resp.raise_for_status()
    return resp.json().get("results", {}).get("records", [])


def run_soql(token: str, instance_url: str, query: str) -> list[dict]:
    resp = requests.get(
        f"{instance_url}/services/data/v66.0/query",
        headers={"Authorization": f"Bearer {token}"},
        params={"q": query},
        timeout=60,
    )
    resp.raise_for_status()
    records = resp.json().get("records", [])
    # Handle pagination
    next_url = resp.json().get("nextRecordsUrl")
    while next_url:
        resp = requests.get(
            f"{instance_url}{next_url}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=60,
        )
        resp.raise_for_status()
        records.extend(resp.json().get("records", []))
        next_url = resp.json().get("nextRecordsUrl")
    return records


def _add_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(title=name[:31])
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(headers[ci - 1])),
            *(len(str(r[ci - 1])) for r in rows[:50]) if rows else [0],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
    if rows:
        end_col = get_column_letter(len(headers))
        table_name = name.replace(" ", "_").replace("-", "_").replace("&", "And")[:30]
        table = Table(displayName=table_name, ref=f"A1:{end_col}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
        )
        ws.add_table(table)
    ws.freeze_panes = "A2"
    return ws


def extract_and_build(territory: str, snapshot_date: str, output_path: Path):
    config = TERRITORIES[territory]
    director = config["director"]
    soql_filter = config["soql_filter"]
    pi_lv_id = config["pi_list_view_id"]

    print(f"Extracting: {territory} ({director})")
    token, instance_url = get_auth()

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Pipeline Detail (SOQL — deal-level, FY26 only) ──
    print("  Pulling pipeline detail...", end=" ", flush=True)
    soql = f"""
        SELECT Account.Name, Name, StageName, CloseDate, ForecastCategoryName,
               APTS_Opportunity_ARR__c, APTS_Forecast_ARR__c, Owner.Name,
               Probability, PushCount, Type, CreatedDate,
               LastActivityDate, NextStep
        FROM Opportunity
        WHERE {soql_filter}
          AND IsClosed = false
          AND CloseDate <= 2026-12-31
          AND (NOT Account.Name LIKE '%SimCorp%')
          AND (NOT Account.Name LIKE '%Test Account%')
        ORDER BY APTS_Opportunity_ARR__c DESC NULLS LAST
    """
    pipeline = run_soql(token, instance_url, soql)
    print(f"{len(pipeline)} deals")

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "Close Date",
        "ARR",
        "Forecast ARR",
        "Probability %",
        "Push Count",
        "Type",
        "Last Activity",
        "Next Step",
    ]
    rows = []
    for r in pipeline:
        rows.append(
            [
                (r.get("Account") or {}).get("Name", ""),
                r.get("Name", ""),
                (r.get("Owner") or {}).get("Name", ""),
                r.get("StageName", ""),
                r.get("ForecastCategoryName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                r.get("APTS_Forecast_ARR__c") or 0,
                r.get("Probability") or 0,
                r.get("PushCount") or 0,
                r.get("Type", ""),
                r.get("LastActivityDate", ""),
                r.get("NextStep", ""),
            ]
        )
    _add_sheet(wb, "Pipeline Detail", headers, rows)

    # ── 2. Won/Lost (SOQL — FY26 closed) ──
    print("  Pulling won/lost...", end=" ", flush=True)
    soql_wl = f"""
        SELECT Account.Name, Name, StageName, CloseDate,
               APTS_Opportunity_ARR__c, Owner.Name, IsWon,
               Reason_Won_Lost__c, Lost_to_Competitor__c, Type
        FROM Opportunity
        WHERE {soql_filter}
          AND IsClosed = true
          AND CloseDate >= 2026-01-01 AND CloseDate <= 2026-12-31
          AND (NOT Account.Name LIKE '%SimCorp%')
          AND (NOT Account.Name LIKE '%Test Account%')
        ORDER BY CloseDate DESC
    """
    won_lost = run_soql(token, instance_url, soql_wl)
    print(f"{len(won_lost)} deals")

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR",
        "Won",
        "Reason",
        "Competitor",
        "Type",
    ]
    rows = []
    for r in won_lost:
        rows.append(
            [
                (r.get("Account") or {}).get("Name", ""),
                r.get("Name", ""),
                (r.get("Owner") or {}).get("Name", ""),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Won" if r.get("IsWon") else "Lost",
                r.get("Reason_Won_Lost__c", ""),
                r.get("Lost_to_Competitor__c", ""),
                r.get("Type", ""),
            ]
        )
    _add_sheet(wb, "Won Lost", headers, rows)

    # ── 3. Commercial Approval (SOQL) ──
    print("  Pulling commercial approvals...", end=" ", flush=True)
    soql_ca = f"""
        SELECT Account.Name, Name, StageName, CloseDate,
               APTS_Opportunity_ARR__c, Owner.Name,
               Stage_20_Approval__c, Stage_20_Approval_Date__c, Type, NextStep
        FROM Opportunity
        WHERE {soql_filter}
          AND IsClosed = false
          AND Type = 'Land'
          AND CloseDate <= 2026-12-31
          AND (NOT Account.Name LIKE '%SimCorp%')
          AND (NOT Account.Name LIKE '%Test Account%')
        ORDER BY APTS_Opportunity_ARR__c DESC NULLS LAST
    """
    approvals = run_soql(token, instance_url, soql_ca)
    print(f"{len(approvals)} land deals")

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR",
        "Approved",
        "Approval Date",
        "Next Step",
    ]
    rows = []
    for r in approvals:
        rows.append(
            [
                (r.get("Account") or {}).get("Name", ""),
                r.get("Name", ""),
                (r.get("Owner") or {}).get("Name", ""),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Yes" if r.get("Stage_20_Approval__c") else "No",
                r.get("Stage_20_Approval_Date__c", ""),
                r.get("NextStep", ""),
            ]
        )
    _add_sheet(wb, "Commercial Approval", headers, rows)

    # ── 4. Renewals (SOQL) ──
    print("  Pulling renewals...", end=" ", flush=True)
    soql_ren = f"""
        SELECT Account.Name, Name, StageName, CloseDate,
               Amount, Owner.Name, Probability, Type, NextStep
        FROM Opportunity
        WHERE {soql_filter}
          AND IsClosed = false
          AND Type = 'Renewal'
          AND CloseDate <= 2026-12-31
          AND (NOT Account.Name LIKE '%SimCorp%')
          AND (NOT Account.Name LIKE '%Test Account%')
        ORDER BY CloseDate ASC
    """
    renewals = run_soql(token, instance_url, soql_ren)
    print(f"{len(renewals)} renewals")

    headers = [
        "Close Date",
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ACV",
        "Probability %",
        "Comments",
    ]
    rows = []
    for r in renewals:
        rows.append(
            [
                r.get("CloseDate", ""),
                (r.get("Account") or {}).get("Name", ""),
                r.get("Name", ""),
                (r.get("Owner") or {}).get("Name", ""),
                r.get("StageName", ""),
                r.get("Amount") or 0,
                r.get("Probability") or 0,
                "",  # Comments column for director to fill
            ]
        )
    _add_sheet(wb, "Renewals", headers, rows)

    # ── 5. Pipeline Inspection (UI API) ──
    print("  Pulling pipeline inspection...", end=" ", flush=True)
    pi_headers = {"Authorization": f"Bearer {token}"}
    url = f"{instance_url}/services/data/v66.0/ui-api/list-records/{pi_lv_id}?pageSize=200"
    pi_records = []
    while url and len(pi_records) < 2000:
        resp = requests.get(url, headers=pi_headers, timeout=30)
        if resp.status_code != 200:
            break
        d = resp.json()
        pi_records.extend(d.get("records", []))
        next_url = d.get("nextPageUrl")
        url = f"{instance_url}{next_url}" if next_url else None

    # Parse PI records, filter FY26 + internal
    headers = [
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "Forecast ARR",
        "Close Date",
        "Push Count",
        "Score",
        "Priority",
    ]
    rows = []
    for rec in pi_records:
        f = rec.get("fields", {})
        name = str(f.get("Name", {}).get("value", ""))
        close = str(f.get("CloseDate", {}).get("value", ""))
        if _is_internal(name):
            continue
        if close and close[:4] > "2026":
            continue
        is_closed = f.get("IsClosed", {}).get("value", False)
        fc = str(f.get("ForecastCategoryName", {}).get("value", ""))
        if is_closed or fc in ("Omitted", "Closed"):
            continue

        owner_obj = f.get("Owner", {}).get("value")
        owner = ""
        if isinstance(owner_obj, dict):
            owner = owner_obj.get("fields", {}).get("Name", {}).get("value", "")
        score_obj = f.get("OpportunityScore", {}).get("value")
        score = None
        if isinstance(score_obj, dict):
            score = score_obj.get("fields", {}).get("Score", {}).get("value")

        rows.append(
            [
                name,
                owner,
                str(f.get("StageName", {}).get("value", "")),
                fc,
                f.get("APTS_Forecast_ARR__c", {}).get("value") or 0,
                close,
                f.get("PushCount", {}).get("value") or 0,
                score,
                "Yes" if f.get("IsPriorityRecord", {}).get("value") else "",
            ]
        )
    rows.sort(key=lambda x: -(x[4] or 0))
    print(f"{len(rows)} open FY26 deals")
    _add_sheet(wb, "Pipeline Inspection", headers, rows)

    # ── Summary sheet ──
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = f"{director} ({territory})"
    ws["A1"].font = Font(bold=True, size=14, color="083EA7")
    ws["A2"] = f"Snapshot: {snapshot_date}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A3"] = (
        f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')} (live from Salesforce)"
    )
    ws["A3"].font = Font(size=10, color="666666")
    ws["A4"] = "Source: Pipeline Reporting & Insights dashboard + SOQL + PI views"
    ws["A4"].font = Font(size=9, italic=True, color="999999")

    ws["A6"] = "Sheet"
    ws["B6"] = "Records"
    ws["C6"] = "Source"
    for col in ("A", "B", "C"):
        ws[f"{col}6"].font = HEADER_FONT
        ws[f"{col}6"].fill = HEADER_FILL

    sheet_info = [
        ("Pipeline Detail", len(pipeline), "SOQL — open FY26 opps"),
        ("Won Lost", len(won_lost), "SOQL — closed FY26"),
        ("Commercial Approval", len(approvals), "SOQL — open Land deals"),
        ("Renewals", len(renewals), "SOQL — open Renewal deals"),
        ("Pipeline Inspection", len(rows), "PI list view — open FY26"),
    ]
    for i, (sname, count, source) in enumerate(sheet_info, 7):
        ws[f"A{i}"] = sname
        ws[f"B{i}"] = count
        ws[f"C{i}"] = source
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40

    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")
    for sname, count, source in sheet_info:
        print(f"  {sname:25s}  {count:>5d}  ({source})")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--territory", required=True, choices=list(TERRITORIES.keys()))
    parser.add_argument("--snapshot-date", default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    config = TERRITORIES[args.territory]
    if args.output:
        out = args.output
    else:
        import re

        slug = re.sub(r"[^a-z0-9]+", "-", config["director"].lower()).strip("-")
        out = (
            REPO_ROOT
            / "output"
            / "director_data_workbooks"
            / args.snapshot_date
            / f"{slug}-live.xlsx"
        )

    out.parent.mkdir(parents=True, exist_ok=True)
    extract_and_build(args.territory, args.snapshot_date, out)


if __name__ == "__main__":
    main()
