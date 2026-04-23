"""
Monthly Sales Directors Review, end-to-end ETL.

Steps, in sequence:
  1. Extract live Salesforce data into nine per-director workbooks.
  2. Render a Land-only deck for each director.
  3. Build the consolidated SharePoint analysis workbook.
  4. Build the dashboard and Q1 analysis workbook.
  5. Write a manifest.json listing every output file with row counts
     and deck slide counts so you can verify what ran.

Idempotent for a given run date. Re-running the same day overwrites that
day's outputs. Pass --date YYYY-MM-DD to backdate a run or recompute
a prior snapshot from the extract already on disk.

Skip flags let you rerun a single stage without redoing the rest:
  --skip-extract
  --skip-decks
  --skip-analysis

Exit code is 0 if every step succeeded, 1 otherwise.
"""

import argparse
import json
import os
import shutil
import socket
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "output"
WORKBOOKS_ROOT = OUTPUT_ROOT / "director_live_workbooks"
DECKS_ROOT = OUTPUT_ROOT / "simcorp_director_decks"
SHAREPOINT_ROOT = OUTPUT_ROOT / "sharepoint"
LOGS_ROOT = OUTPUT_ROOT / "pipeline_logs"

# Territory config key -> professional deck territory label.
# "Pension & Insurance" is branded "NA Pension & Insurance" in the deck name.
_TERRITORY_DECK_LABELS: dict[str, str] = {
    "APAC": "APAC",
    "Central Europe": "Central Europe",
    "UK & Ireland": "UK & Ireland",
    "Southern Europe": "Southern Europe",
    "NL & Nordics": "NL & Nordics",
    "Middle East & Africa": "Middle East & Africa",
    "Canada": "Canada",
    "NA Asset Management": "NA Asset Management",
    "Pension & Insurance": "NA Pension & Insurance",
}

# Reverse lookup: lowered director last-name fragments -> territory config key.
# Deck filenames are "{director-slug}-LAND.pptx" where the slug is derived from
# the workbook stem (e.g. "jesper-tyrer").  We match by lower-cased stem prefix.
_DIRECTOR_STEM_TO_TERRITORY: dict[str, str] = {}


def _load_territory_config() -> dict:
    """Load sd_monthly_territories.json and populate director-stem lookup."""
    cfg_path = ROOT / "config" / "sd_monthly_territories.json"
    with open(cfg_path) as f:
        cfg = json.load(f)
    territories = cfg.get("territories", {})
    for territory_key, entry in territories.items():
        director = entry.get("director", "")
        # Build a slug that matches what the workbook stem looks like:
        # "Jesper Tyrer" -> "jesper-tyrer"
        slug = director.lower().replace(" ", "-")
        _DIRECTOR_STEM_TO_TERRITORY[slug] = territory_key
    return territories


def _month_year_label(date_stamp: str) -> str:
    """'2026-04-20' -> 'April 2026'."""
    dt = datetime.strptime(date_stamp[:10], "%Y-%m-%d")
    return dt.strftime("%B %Y")


def _rename_deliverables(decks_dir: Path, date_stamp: str) -> dict:
    """Rename deck files and sharepoint workbooks to professional names.

    Returns a step-like dict with rename results for the manifest.
    """
    month_year = _month_year_label(date_stamp)
    renames: list[dict[str, str]] = []
    errors: list[str] = []

    # --- Rename decks ---
    if decks_dir.exists():
        for pptx in sorted(decks_dir.glob("*.pptx")):
            if pptx.name.startswith("~"):
                continue
            old_name = pptx.name

            # Exec Rollup
            if old_name == "Exec Rollup.pptx":
                new_name = f"Sales Director Monthly - Exec Rollup - {month_year}.pptx"
            else:
                # Strip the "-LAND" suffix to get the director slug
                stem = pptx.stem
                if stem.endswith("-LAND"):
                    stem = stem[: -len("-LAND")]
                slug = stem.lower()
                territory_key = _DIRECTOR_STEM_TO_TERRITORY.get(slug)
                if territory_key is None:
                    errors.append(f"No territory mapping for deck stem '{stem}'")
                    continue
                label = _TERRITORY_DECK_LABELS.get(territory_key, territory_key)
                new_name = f"Sales Director Monthly - {label} - {month_year}.pptx"

            if old_name != new_name:
                new_path = pptx.parent / new_name
                shutil.move(str(pptx), str(new_path))
                renames.append({"from": old_name, "to": new_name})

    # --- Rename sharepoint workbooks ---
    if SHAREPOINT_ROOT.exists():
        for xlsx in sorted(SHAREPOINT_ROOT.glob("FY26 Pipeline Review, *.xlsx")):
            if xlsx.name.startswith("~"):
                continue
            old_name = xlsx.name
            # "FY26 Pipeline Review, APAC.xlsx" -> region = "APAC"
            region = old_name.replace("FY26 Pipeline Review, ", "").replace(".xlsx", "")
            new_name = f"FY26 Pipeline Review - {region} - {month_year}.xlsx"
            if old_name != new_name:
                new_path = xlsx.parent / new_name
                shutil.move(str(xlsx), str(new_path))
                renames.append({"from": old_name, "to": new_name})

    status = "ok" if not errors else "failed"
    print(f"  [3c_rename_deliverables] {status}: {len(renames)} files renamed")
    if errors:
        for e in errors:
            print(f"    ERROR: {e}")

    return {
        "name": "3c_rename_deliverables",
        "command": "rename_deliverables",
        "exit_code": 0 if not errors else 1,
        "duration_seconds": 0.0,
        "log_path": "",
        "status": status,
        "renames": renames,
        "errors": errors,
    }


def run_step(name, cmd, log_path):
    start = time.time()
    print(f"  [{name}] starting")
    with open(log_path, "w") as f:
        f.write(f"# {name}\n# {' '.join(str(c) for c in cmd)}\n\n")
        f.flush()
        result = subprocess.run(
            cmd,
            stdout=f,
            stderr=subprocess.STDOUT,
            cwd=ROOT,
            text=True,
        )
    duration = time.time() - start
    status = "ok" if result.returncode == 0 else "failed"
    print(f"  [{name}] {status} in {duration:.1f}s")
    return {
        "name": name,
        "command": " ".join(str(c) for c in cmd),
        "exit_code": result.returncode,
        "duration_seconds": round(duration, 1),
        "log_path": str(log_path.relative_to(ROOT)),
        "status": status,
    }


def inventory_xlsx(path):
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": str(exc)}
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        out[sn] = ws.max_row
    return out


def inventory_pptx(path):
    try:
        from pptx import Presentation

        prs = Presentation(str(path))
        return {"slides": len(prs.slides)}
    except Exception as exc:
        return {"error": str(exc)}


def _manifest_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _append_step_artifact(step: dict, artifact_type: str, path: Path) -> None:
    if not path.exists():
        return
    artifacts = step.setdefault("artifacts", [])
    artifact = {"type": artifact_type, "path": _manifest_path(path)}
    if artifact not in artifacts:
        artifacts.append(artifact)


def _single_flight_lock_path() -> Path:
    return LOGS_ROOT / ".run_monthly_director_review.lock"


def _pid_is_running(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    return True


def _read_lock_payload(path: Path) -> dict:
    try:
        payload = json.loads(path.read_text())
    except (OSError, ValueError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _acquire_single_flight_lock(run_date: str) -> dict:
    lock_path = _single_flight_lock_path()
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "pid": os.getpid(),
        "run_date": str(run_date)[:10],
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "hostname": socket.gethostname(),
        "command": " ".join(sys.argv),
    }
    for _ in range(2):
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            holder = _read_lock_payload(lock_path)
            holder_pid = int(holder.get("pid") or 0)
            if holder and _pid_is_running(holder_pid):
                return {
                    "acquired": False,
                    "reason": "already_running",
                    "path": lock_path,
                    "holder": holder,
                }
            try:
                lock_path.unlink()
            except FileNotFoundError:
                continue
            except OSError:
                return {
                    "acquired": False,
                    "reason": "stale_lock_cleanup_failed",
                    "path": lock_path,
                    "holder": holder,
                }
            continue
        else:
            with os.fdopen(fd, "w") as handle:
                json.dump(payload, handle, indent=2)
                handle.write("\n")
            return {
                "acquired": True,
                "reason": "acquired",
                "path": lock_path,
                "holder": payload,
            }
    return {
        "acquired": False,
        "reason": "lock_acquisition_failed",
        "path": lock_path,
        "holder": _read_lock_payload(lock_path),
    }


def _release_single_flight_lock(lock: dict | None) -> None:
    if not lock or not lock.get("acquired"):
        return
    path = Path(lock["path"])
    holder = _read_lock_payload(path)
    if holder and int(holder.get("pid") or 0) not in (0, os.getpid()):
        return
    try:
        path.unlink()
    except FileNotFoundError:
        return
    except OSError:
        return


def _write_lock_conflict(log_dir: Path, run_date: str, lock: dict) -> Path:
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    path = log_dir / f"single_flight_lock_conflict-{timestamp}.json"
    payload = {
        "run_date": str(run_date)[:10],
        "status": "blocked",
        "reason": str(lock.get("reason") or "already_running"),
        "lock_path": _manifest_path(Path(lock["path"])),
        "holder": lock.get("holder") or {},
        "logged_at": datetime.now().isoformat(timespec="seconds"),
    }
    path.write_text(json.dumps(payload, indent=2) + "\n")
    return path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--date",
        default=datetime.now().strftime("%Y-%m-%d"),
        help="Run date stamp, YYYY-MM-DD. Defaults to today.",
    )
    parser.add_argument("--skip-extract", action="store_true")
    parser.add_argument("--skip-decks", action="store_true")
    parser.add_argument("--skip-analysis", action="store_true")
    args = parser.parse_args()

    date_stamp = args.date
    log_dir = LOGS_ROOT / date_stamp
    log_dir.mkdir(parents=True, exist_ok=True)

    single_flight_lock = _acquire_single_flight_lock(date_stamp)
    if not single_flight_lock["acquired"]:
        conflict_path = _write_lock_conflict(log_dir, date_stamp, single_flight_lock)
        blocked_step = {
            "name": "0_single_flight_lock",
            "command": " ".join(sys.argv),
            "exit_code": 1,
            "duration_seconds": 0.0,
            "log_path": "",
            "status": "blocked",
        }
        _append_step_artifact(
            blocked_step,
            "single_flight_lock_conflict",
            conflict_path,
        )
        blocked_manifest = {
            "run_date": date_stamp,
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "single_flight_lock": {
                "status": "blocked",
                "reason": str(single_flight_lock.get("reason") or "already_running"),
                "path": _manifest_path(Path(single_flight_lock["path"])),
                "holder": single_flight_lock.get("holder") or {},
                "conflict_artifact": _manifest_path(conflict_path),
            },
            "steps": [blocked_step],
            "outputs": {
                "extracts": [],
                "decks": [],
                "reports": [],
            },
        }
        _write_manifest_with_release_packet(blocked_manifest, log_dir)
        holder = single_flight_lock.get("holder") or {}
        print(
            "Another monthly pipeline run is already active "
            f"(pid {holder.get('pid')}, run date {holder.get('run_date')})."
        )
        print(f"Lock conflict: {_manifest_path(conflict_path)}")
        return 1

    _load_territory_config()

    print(f"Monthly Sales Directors Review pipeline, date {date_stamp}")
    print(f"Logs: {log_dir.relative_to(ROOT)}")
    print()

    manifest = {
        "run_date": date_stamp,
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "single_flight_lock": {
            "status": "acquired",
            "path": _manifest_path(Path(single_flight_lock["path"])),
            "pid": single_flight_lock["holder"]["pid"],
            "started_at": single_flight_lock["holder"]["started_at"],
            "hostname": single_flight_lock["holder"]["hostname"],
        },
        "steps": [],
        "outputs": {
            "extracts": [],
            "decks": [],
            "reports": [],
        },
    }

    try:
        step = run_step(
            "0_source_contract_preflight",
            [
                sys.executable,
                "scripts/audit_sales_director_source_contract.py",
                "--date",
                date_stamp,
            ],
            log_dir / "0_source_contract_preflight.log",
        )
        _append_step_artifact(
            step,
            "source_contract_audit",
            OUTPUT_ROOT
            / "source_contract_audit"
            / date_stamp
            / "source_contract_audit.json",
        )
        _append_step_artifact(
            step,
            "source_contract_summary",
            OUTPUT_ROOT / "source_contract_audit" / date_stamp / "summary.md",
        )
        manifest["steps"].append(step)
        diff_step = run_step(
            "0b_source_contract_snapshot_diff",
            [
                sys.executable,
                "scripts/diff_source_contract_snapshots.py",
                "--current-date",
                date_stamp,
            ],
            log_dir / "0b_source_contract_snapshot_diff.log",
        )
        _append_step_artifact(
            diff_step,
            "source_contract_snapshot_diff",
            OUTPUT_ROOT
            / "source_contract_snapshot_diff"
            / date_stamp
            / "source_contract_snapshot_diff.json",
        )
        _append_step_artifact(
            diff_step,
            "source_contract_snapshot_diff_summary",
            OUTPUT_ROOT / "source_contract_snapshot_diff" / date_stamp / "summary.md",
        )
        manifest["steps"].append(diff_step)
        refresh_step = run_step(
            "0c_forward_quarter_registry_refresh",
            [
                sys.executable,
                "scripts/refresh_forward_quarter_registry.py",
                "--date",
                date_stamp,
            ],
            log_dir / "0c_forward_quarter_registry_refresh.log",
        )
        _append_step_artifact(
            refresh_step,
            "forward_quarter_registry_refresh",
            OUTPUT_ROOT
            / "source_contract_registry_refresh"
            / date_stamp
            / "registry_refresh.json",
        )
        _append_step_artifact(
            refresh_step,
            "forward_quarter_registry_refresh_summary",
            OUTPUT_ROOT
            / "source_contract_registry_refresh"
            / date_stamp
            / "summary.md",
        )
        _append_step_artifact(
            refresh_step,
            "forward_quarter_registry_proposed_config",
            OUTPUT_ROOT
            / "source_contract_registry_refresh"
            / date_stamp
            / "proposed_sd_monthly_territories.json",
        )
        manifest["steps"].append(refresh_step)
        if step["exit_code"] != 0:
            print("Source contract preflight failed. Aborting downstream steps.")
            _write_manifest_with_release_packet(manifest, log_dir)
            return 1

        # Stage 1: Extract live Salesforce data
        if not args.skip_extract:
            step = run_step(
                "1a_extract_salesforce",
                [
                    sys.executable,
                    "scripts/extract_director_live.py",
                    "--all",
                    "--snapshot-date",
                    date_stamp,
                ],
                log_dir / "1a_extract_salesforce.log",
            )
            _append_step_artifact(
                step,
                "director_live_extract_audit",
                OUTPUT_ROOT
                / "director_live_extract"
                / date_stamp
                / "director_live_extract_audit.json",
            )
            _append_step_artifact(
                step,
                "director_live_extract_summary",
                OUTPUT_ROOT / "director_live_extract" / date_stamp / "summary.md",
            )
            manifest["steps"].append(step)
            live_extract_audit_path = (
                OUTPUT_ROOT
                / "director_live_extract"
                / date_stamp
                / "director_live_extract_audit.json"
            )
            if live_extract_audit_path.exists():
                diff_step = run_step(
                    "1a2_director_live_extract_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_director_live_extract_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "1a2_director_live_extract_snapshot_diff.log",
                )
                _append_step_artifact(
                    diff_step,
                    "director_live_extract_snapshot_diff",
                    OUTPUT_ROOT
                    / "director_live_extract_snapshot_diff"
                    / date_stamp
                    / "director_live_extract_snapshot_diff.json",
                )
                _append_step_artifact(
                    diff_step,
                    "director_live_extract_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "director_live_extract_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(diff_step)
            if step["exit_code"] != 0:
                print("Extract failed. Aborting downstream steps.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1

            # Historical Trending snapshots, appended to the per-director workbooks
            step = run_step(
                "1b_extract_historical_trending",
                [
                    sys.executable,
                    "scripts/extract_historical_trending.py",
                    "--snapshot-date",
                    date_stamp,
                    "--workbooks-dir",
                    str(WORKBOOKS_ROOT / date_stamp),
                ],
                log_dir / "1b_extract_historical_trending.log",
            )
            _append_step_artifact(
                step,
                "historical_trending_extract_audit",
                OUTPUT_ROOT
                / "historical_trending_extract"
                / date_stamp
                / "historical_trending_extract_audit.json",
            )
            _append_step_artifact(
                step,
                "historical_trending_extract_summary",
                OUTPUT_ROOT / "historical_trending_extract" / date_stamp / "summary.md",
            )
            manifest["steps"].append(step)
            historical_audit_path = (
                OUTPUT_ROOT
                / "historical_trending_extract"
                / date_stamp
                / "historical_trending_extract_audit.json"
            )
            if historical_audit_path.exists():
                diff_step = run_step(
                    "1b2_historical_trending_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_historical_trending_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "1b2_historical_trending_snapshot_diff.log",
                )
                _append_step_artifact(
                    diff_step,
                    "historical_trending_snapshot_diff",
                    OUTPUT_ROOT
                    / "historical_trending_snapshot_diff"
                    / date_stamp
                    / "historical_trending_snapshot_diff.json",
                )
                _append_step_artifact(
                    diff_step,
                    "historical_trending_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "historical_trending_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(diff_step)
            if step["exit_code"] != 0:
                print("Historical trending extract failed. Aborting downstream steps.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1

            step = run_step(
                "1b3_validate_director_workbook_contract",
                [
                    sys.executable,
                    "scripts/validate_director_workbook_contract.py",
                    "--snapshot-date",
                    date_stamp,
                    "--workbooks-dir",
                    str(WORKBOOKS_ROOT / date_stamp),
                    "--require-historical",
                ],
                log_dir / "1b3_validate_director_workbook_contract.log",
            )
            _append_step_artifact(
                step,
                "director_workbook_contract_audit",
                OUTPUT_ROOT
                / "director_workbook_contract"
                / date_stamp
                / "director_workbook_contract_audit.json",
            )
            _append_step_artifact(
                step,
                "director_workbook_contract_summary",
                OUTPUT_ROOT / "director_workbook_contract" / date_stamp / "summary.md",
            )
            manifest["steps"].append(step)
            workbook_contract_audit_path = (
                OUTPUT_ROOT
                / "director_workbook_contract"
                / date_stamp
                / "director_workbook_contract_audit.json"
            )
            if workbook_contract_audit_path.exists():
                diff_step = run_step(
                    "1b4_director_workbook_contract_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_director_workbook_contract_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "1b4_director_workbook_contract_snapshot_diff.log",
                )
                _append_step_artifact(
                    diff_step,
                    "director_workbook_contract_snapshot_diff",
                    OUTPUT_ROOT
                    / "director_workbook_contract_snapshot_diff"
                    / date_stamp
                    / "director_workbook_contract_snapshot_diff.json",
                )
                _append_step_artifact(
                    diff_step,
                    "director_workbook_contract_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "director_workbook_contract_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(diff_step)
            if step["exit_code"] != 0:
                print("Workbook contract validation failed. Aborting downstream steps.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1

            # Data-quality audit — scans SF pipeline + account hygiene gaps,
            # writes per-run JSON + an Obsidian-ready summary + appends to the
            # rolling history ledger for month-over-month deltas. Runs regardless
            # of skip-analysis because data-quality is an extract-time concern.
            step = run_step(
                "1c_data_quality_audit",
                [
                    sys.executable,
                    "scripts/audit_data_quality.py",
                    "--date",
                    date_stamp,
                ],
                log_dir / "1c_data_quality_audit.log",
            )
            _append_step_artifact(
                step,
                "data_quality_flags",
                OUTPUT_ROOT / "data_quality" / date_stamp / "flags.json",
            )
            _append_step_artifact(
                step,
                "data_quality_summary",
                OUTPUT_ROOT / "data_quality" / date_stamp / "summary.md",
            )
            manifest["steps"].append(step)
            data_quality_flags_path = (
                OUTPUT_ROOT / "data_quality" / date_stamp / "flags.json"
            )
            if data_quality_flags_path.exists():
                data_quality_diff_step = run_step(
                    "1c2_data_quality_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_data_quality_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "1c2_data_quality_snapshot_diff.log",
                )
                _append_step_artifact(
                    data_quality_diff_step,
                    "data_quality_snapshot_diff",
                    OUTPUT_ROOT
                    / "data_quality_snapshot_diff"
                    / date_stamp
                    / "data_quality_snapshot_diff.json",
                )
                _append_step_artifact(
                    data_quality_diff_step,
                    "data_quality_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "data_quality_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(data_quality_diff_step)
        else:
            print("  skipping extract")

        # Stage 2: Analyze in Excel (SharePoint-ready analysis workbooks)
        if not args.skip_analysis:
            SHAREPOINT_ROOT.mkdir(parents=True, exist_ok=True)
            step = run_step(
                "2a_analyze_consolidated_review",
                [
                    sys.executable,
                    "scripts/build_sharepoint_analysis.py",
                    "--workbooks-dir",
                    str(WORKBOOKS_ROOT / date_stamp),
                    "--date",
                    date_stamp,
                ],
                log_dir / "2a_analyze_consolidated_review.log",
            )
            manifest["steps"].append(step)
            # Per-director regional workbooks: same tab structure as the master
            # but scoped to one territory. Directors receive their regional
            # workbook alongside their deck. Ops keeps the master above.
            sys.path.insert(0, str(ROOT / "scripts"))
            from build_sharepoint_analysis import DIRECTORS as _SD_DIRECTORS

            for _name, territory, _fname, *_ in _SD_DIRECTORS:
                safe = territory.replace(" ", "_").replace("&", "and").replace("/", "-")
                step = run_step(
                    f"2a2_regional_{safe}",
                    [
                        sys.executable,
                        "scripts/build_sharepoint_analysis.py",
                        "--workbooks-dir",
                        str(WORKBOOKS_ROOT / date_stamp),
                        "--date",
                        date_stamp,
                        "--territory",
                        territory,
                    ],
                    log_dir / f"2a2_regional_{safe}.log",
                )
                manifest["steps"].append(step)
            step = run_step(
                "2b_analyze_dashboard_q1",
                [sys.executable, "scripts/build_dashboard_analysis_excel.py"],
                log_dir / "2b_analyze_dashboard_q1.log",
            )
            manifest["steps"].append(step)
            analysis_contract_step = run_step(
                "2b2_validate_sharepoint_analysis_contract",
                [
                    sys.executable,
                    "scripts/validate_sharepoint_analysis_contract.py",
                    "--date",
                    date_stamp,
                    "--sharepoint-root",
                    str(SHAREPOINT_ROOT),
                ],
                log_dir / "2b2_validate_sharepoint_analysis_contract.log",
            )
            _append_step_artifact(
                analysis_contract_step,
                "sharepoint_analysis_contract_audit",
                OUTPUT_ROOT
                / "sharepoint_analysis_contract"
                / date_stamp
                / "sharepoint_analysis_contract_audit.json",
            )
            _append_step_artifact(
                analysis_contract_step,
                "sharepoint_analysis_contract_summary",
                OUTPUT_ROOT
                / "sharepoint_analysis_contract"
                / date_stamp
                / "summary.md",
            )
            manifest["steps"].append(analysis_contract_step)
            analysis_contract_audit_path = (
                OUTPUT_ROOT
                / "sharepoint_analysis_contract"
                / date_stamp
                / "sharepoint_analysis_contract_audit.json"
            )
            if analysis_contract_audit_path.exists():
                diff_step = run_step(
                    "2b3_sharepoint_analysis_contract_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_sharepoint_analysis_contract_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "2b3_sharepoint_analysis_contract_snapshot_diff.log",
                )
                _append_step_artifact(
                    diff_step,
                    "sharepoint_analysis_contract_snapshot_diff",
                    OUTPUT_ROOT
                    / "sharepoint_analysis_contract_snapshot_diff"
                    / date_stamp
                    / "sharepoint_analysis_contract_snapshot_diff.json",
                )
                _append_step_artifact(
                    diff_step,
                    "sharepoint_analysis_contract_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "sharepoint_analysis_contract_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(diff_step)
            if analysis_contract_step["exit_code"] != 0:
                print(
                    "SharePoint analysis contract validation failed. Aborting downstream steps."
                )
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1
        else:
            print("  skipping analysis")

        workbooks_dir = WORKBOOKS_ROOT / date_stamp

        # Stage 3: Ship decks (the final deliverable)
        if not args.skip_decks:
            if not workbooks_dir.exists():
                print(f"  workbook directory missing: {workbooks_dir}")
            else:
                decks_dir = DECKS_ROOT / date_stamp / "land-only"
                decks_dir.mkdir(parents=True, exist_ok=True)
                for wb_path in sorted(workbooks_dir.glob("*.xlsx")):
                    if wb_path.name.startswith("~"):
                        continue
                    name = wb_path.stem
                    out_path = decks_dir / f"{name}-LAND.pptx"
                    step = run_step(
                        f"3_ship_deck_{name}",
                        [
                            sys.executable,
                            "scripts/build_deck_from_excel.py",
                            "--workbook",
                            str(wb_path),
                            "--date",
                            date_stamp,
                            "--output",
                            str(out_path),
                            "--land-only",
                        ],
                        log_dir / f"3_ship_deck_{name}.log",
                    )
                    manifest["steps"].append(step)
                # Executive rollup deck, one across all directors
                step = run_step(
                    "3_ship_exec_rollup",
                    [
                        sys.executable,
                        "scripts/build_exec_rollup_deck.py",
                        "--workbooks-dir",
                        str(workbooks_dir),
                        "--output",
                        str(decks_dir / "Exec Rollup.pptx"),
                    ],
                    log_dir / "3_ship_exec_rollup.log",
                )
                manifest["steps"].append(step)

            font_normalization_step = run_step(
                "3a_normalize_deck_fonts",
                [
                    sys.executable,
                    "scripts/normalize_deck_fonts.py",
                    "--date",
                    date_stamp,
                    "--decks-dir",
                    str(DECKS_ROOT / date_stamp / "land-only"),
                ],
                log_dir / "3a_normalize_deck_fonts.log",
            )
            _append_step_artifact(
                font_normalization_step,
                "deck_font_normalization",
                OUTPUT_ROOT
                / "deck_font_normalization"
                / date_stamp
                / "deck_font_normalization.json",
            )
            _append_step_artifact(
                font_normalization_step,
                "deck_font_normalization_summary",
                OUTPUT_ROOT / "deck_font_normalization" / date_stamp / "summary.md",
            )
            manifest["steps"].append(font_normalization_step)
            if font_normalization_step["exit_code"] != 0:
                print("Deck font normalization failed. Aborting downstream steps.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1
            deck_delivery_step = run_step(
                "3b_validate_deck_delivery_contract",
                [
                    sys.executable,
                    "scripts/validate_deck_delivery_contract.py",
                    "--date",
                    date_stamp,
                    "--workbooks-dir",
                    str(workbooks_dir),
                    "--decks-dir",
                    str(DECKS_ROOT / date_stamp / "land-only"),
                ],
                log_dir / "3b_validate_deck_delivery_contract.log",
            )
            _append_step_artifact(
                deck_delivery_step,
                "deck_delivery_contract_audit",
                OUTPUT_ROOT
                / "deck_delivery_contract"
                / date_stamp
                / "deck_delivery_contract_audit.json",
            )
            _append_step_artifact(
                deck_delivery_step,
                "deck_delivery_contract_summary",
                OUTPUT_ROOT / "deck_delivery_contract" / date_stamp / "summary.md",
            )
            manifest["steps"].append(deck_delivery_step)
            deck_delivery_audit_path = (
                OUTPUT_ROOT
                / "deck_delivery_contract"
                / date_stamp
                / "deck_delivery_contract_audit.json"
            )
            if deck_delivery_audit_path.exists():
                diff_step = run_step(
                    "3b2_deck_delivery_contract_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_deck_delivery_contract_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "3b2_deck_delivery_contract_snapshot_diff.log",
                )
                _append_step_artifact(
                    diff_step,
                    "deck_delivery_contract_snapshot_diff",
                    OUTPUT_ROOT
                    / "deck_delivery_contract_snapshot_diff"
                    / date_stamp
                    / "deck_delivery_contract_snapshot_diff.json",
                )
                _append_step_artifact(
                    diff_step,
                    "deck_delivery_contract_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "deck_delivery_contract_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(diff_step)
            decks_payload_dir = DECKS_ROOT / date_stamp / "land-only"
            if decks_payload_dir.exists():
                fill_payload_diff_step = run_step(
                    "3b3_deck_fill_payload_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_deck_fill_payload_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "3b3_deck_fill_payload_snapshot_diff.log",
                )
                _append_step_artifact(
                    fill_payload_diff_step,
                    "deck_fill_payload_snapshot_diff",
                    OUTPUT_ROOT
                    / "deck_fill_payload_snapshot_diff"
                    / date_stamp
                    / "deck_fill_payload_snapshot_diff.json",
                )
                _append_step_artifact(
                    fill_payload_diff_step,
                    "deck_fill_payload_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "deck_fill_payload_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(fill_payload_diff_step)
                visual_diff_step = run_step(
                    "3b4_deck_visual_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_deck_visual_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "3b4_deck_visual_snapshot_diff.log",
                )
                _append_step_artifact(
                    visual_diff_step,
                    "deck_visual_snapshot_diff",
                    OUTPUT_ROOT
                    / "deck_visual_snapshot_diff"
                    / date_stamp
                    / "deck_visual_snapshot_diff.json",
                )
                _append_step_artifact(
                    visual_diff_step,
                    "deck_visual_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "deck_visual_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(visual_diff_step)
                golden_pack_step = run_step(
                    "3b5_assemble_golden_deck_regression_pack",
                    [
                        sys.executable,
                        "scripts/assemble_golden_deck_regression_pack.py",
                        "--date",
                        date_stamp,
                    ],
                    log_dir / "3b5_assemble_golden_deck_regression_pack.log",
                )
                _append_step_artifact(
                    golden_pack_step,
                    "golden_deck_regression_pack",
                    OUTPUT_ROOT
                    / "golden_deck_regression_pack"
                    / date_stamp
                    / "golden_deck_regression_pack.json",
                )
                _append_step_artifact(
                    golden_pack_step,
                    "golden_deck_regression_pack_summary",
                    OUTPUT_ROOT
                    / "golden_deck_regression_pack"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(golden_pack_step)
                font_audit_step = run_step(
                    "3b6_audit_deck_fonts",
                    [
                        sys.executable,
                        "scripts/audit_deck_fonts.py",
                        "--date",
                        date_stamp,
                    ],
                    log_dir / "3b6_audit_deck_fonts.log",
                )
                _append_step_artifact(
                    font_audit_step,
                    "deck_font_audit",
                    OUTPUT_ROOT
                    / "deck_font_audit"
                    / date_stamp
                    / "deck_font_audit.json",
                )
                _append_step_artifact(
                    font_audit_step,
                    "deck_font_audit_summary",
                    OUTPUT_ROOT / "deck_font_audit" / date_stamp / "summary.md",
                )
                manifest["steps"].append(font_audit_step)
                deck_font_audit_path = (
                    OUTPUT_ROOT
                    / "deck_font_audit"
                    / date_stamp
                    / "deck_font_audit.json"
                )
                if deck_font_audit_path.exists():
                    font_diff_step = run_step(
                        "3b7_deck_font_audit_snapshot_diff",
                        [
                            sys.executable,
                            "scripts/diff_deck_font_audit_snapshots.py",
                            "--current-date",
                            date_stamp,
                        ],
                        log_dir / "3b7_deck_font_audit_snapshot_diff.log",
                    )
                    _append_step_artifact(
                        font_diff_step,
                        "deck_font_audit_snapshot_diff",
                        OUTPUT_ROOT
                        / "deck_font_audit_snapshot_diff"
                        / date_stamp
                        / "deck_font_audit_snapshot_diff.json",
                    )
                    _append_step_artifact(
                        font_diff_step,
                        "deck_font_audit_snapshot_diff_summary",
                        OUTPUT_ROOT
                        / "deck_font_audit_snapshot_diff"
                        / date_stamp
                        / "summary.md",
                    )
                manifest["steps"].append(font_diff_step)
            if deck_delivery_step["exit_code"] != 0:
                print(
                    "Deck delivery contract validation failed. Aborting downstream steps."
                )
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1
        else:
            print("  skipping decks")

        # Inventory outputs for the manifest
        if workbooks_dir.exists():
            for wb_path in sorted(workbooks_dir.glob("*.xlsx")):
                if wb_path.name.startswith("~"):
                    continue
                manifest["outputs"]["extracts"].append(
                    {
                        "file": str(wb_path.relative_to(ROOT)),
                        "sheets": inventory_xlsx(wb_path),
                    }
                )
        decks_dir = DECKS_ROOT / date_stamp / "land-only"
        if decks_dir.exists():
            for pptx in sorted(decks_dir.glob("*.pptx")):
                if pptx.name.startswith("~"):
                    continue  # Office lockfiles from an open viewer
                manifest["outputs"]["decks"].append(
                    {
                        "file": str(pptx.relative_to(ROOT)),
                        **inventory_pptx(pptx),
                    }
                )
        for report_name in [
            "FY26 Pipeline Review, All Territories.xlsx",
            "Dashboard and Q1 Analysis.xlsx",
        ]:
            report_path = SHAREPOINT_ROOT / report_name
            if report_path.exists():
                manifest["outputs"]["reports"].append(
                    {
                        "file": str(report_path.relative_to(ROOT)),
                        "sheets": inventory_xlsx(report_path),
                    }
                )

        _write_manifest(manifest, log_dir)

        # Stage 4: Update Obsidian knowledge base
        if not args.skip_decks:
            step = run_step(
                "4_validate_tie_out",
                [
                    sys.executable,
                    "scripts/validate_tie_out.py",
                    "--date",
                    date_stamp,
                ],
                log_dir / "4_validate_tie_out.log",
            )
            _append_step_artifact(
                step,
                "tie_out_note",
                ROOT / "obsidian" / "Monthly" / date_stamp[:7] / "tie-out.md",
            )
            _append_step_artifact(
                step,
                "tie_out_audit",
                OUTPUT_ROOT / "tie_out" / date_stamp / "tie_out_audit.json",
            )
            _append_step_artifact(
                step,
                "tie_out_summary",
                OUTPUT_ROOT / "tie_out" / date_stamp / "summary.md",
            )
            manifest["steps"].append(step)
            tie_out_audit_path = (
                OUTPUT_ROOT / "tie_out" / date_stamp / "tie_out_audit.json"
            )
            tie_out_diff_step = None
            if tie_out_audit_path.exists():
                tie_out_diff_step = run_step(
                    "4a2_tie_out_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_tie_out_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "4a2_tie_out_snapshot_diff.log",
                )
                _append_step_artifact(
                    tie_out_diff_step,
                    "tie_out_snapshot_diff",
                    OUTPUT_ROOT
                    / "tie_out_snapshot_diff"
                    / date_stamp
                    / "tie_out_snapshot_diff.json",
                )
                _append_step_artifact(
                    tie_out_diff_step,
                    "tie_out_snapshot_diff_summary",
                    OUTPUT_ROOT / "tie_out_snapshot_diff" / date_stamp / "summary.md",
                )
                manifest["steps"].append(tie_out_diff_step)
            scope_step = run_step(
                "4b_audit_deck_scope",
                [
                    sys.executable,
                    "scripts/audit_deck_scope.py",
                    "--date",
                    date_stamp,
                ],
                log_dir / "4b_audit_deck_scope.log",
            )
            _append_step_artifact(
                scope_step,
                "deck_scope_audit",
                OUTPUT_ROOT / "deck_scope_audit" / date_stamp / "deck_scope_audit.json",
            )
            _append_step_artifact(
                scope_step,
                "deck_scope_summary",
                OUTPUT_ROOT / "deck_scope_audit" / date_stamp / "summary.md",
            )
            manifest["steps"].append(scope_step)
            _write_manifest(manifest, log_dir)
            if (
                step["exit_code"] != 0
                or (
                    tie_out_diff_step is not None
                    and tie_out_diff_step["exit_code"] != 0
                )
                or scope_step["exit_code"] != 0
            ):
                print("Validation failed. Aborting knowledge-base update.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1

        if not args.skip_analysis and not args.skip_decks:
            step = run_step(
                "5_update_obsidian_notes",
                [
                    sys.executable,
                    "scripts/generate_obsidian_notes.py",
                    "--date",
                    date_stamp,
                ],
                log_dir / "5_update_obsidian_notes.log",
            )
            _append_step_artifact(
                step,
                "obsidian_monthly_readme",
                ROOT / "obsidian" / "Monthly" / date_stamp[:7] / "README.md",
            )
            _append_step_artifact(
                step,
                "obsidian_snapshot_history",
                ROOT / "obsidian" / "snapshot_history.json",
            )
            manifest["steps"].append(step)
            notes_contract_step = run_step(
                "5a_validate_obsidian_notes_contract",
                [
                    sys.executable,
                    "scripts/validate_obsidian_notes_contract.py",
                    "--date",
                    date_stamp,
                ],
                log_dir / "5a_validate_obsidian_notes_contract.log",
            )
            _append_step_artifact(
                notes_contract_step,
                "obsidian_notes_contract_audit",
                OUTPUT_ROOT
                / "obsidian_notes_contract"
                / date_stamp
                / "obsidian_notes_contract_audit.json",
            )
            _append_step_artifact(
                notes_contract_step,
                "obsidian_notes_contract_summary",
                OUTPUT_ROOT / "obsidian_notes_contract" / date_stamp / "summary.md",
            )
            manifest["steps"].append(notes_contract_step)
            notes_contract_audit_path = (
                OUTPUT_ROOT
                / "obsidian_notes_contract"
                / date_stamp
                / "obsidian_notes_contract_audit.json"
            )
            notes_contract_diff_step = None
            if notes_contract_audit_path.exists():
                notes_contract_diff_step = run_step(
                    "5b_obsidian_notes_contract_snapshot_diff",
                    [
                        sys.executable,
                        "scripts/diff_obsidian_notes_contract_snapshots.py",
                        "--current-date",
                        date_stamp,
                    ],
                    log_dir / "5b_obsidian_notes_contract_snapshot_diff.log",
                )
                _append_step_artifact(
                    notes_contract_diff_step,
                    "obsidian_notes_contract_snapshot_diff",
                    OUTPUT_ROOT
                    / "obsidian_notes_contract_snapshot_diff"
                    / date_stamp
                    / "obsidian_notes_contract_snapshot_diff.json",
                )
                _append_step_artifact(
                    notes_contract_diff_step,
                    "obsidian_notes_contract_snapshot_diff_summary",
                    OUTPUT_ROOT
                    / "obsidian_notes_contract_snapshot_diff"
                    / date_stamp
                    / "summary.md",
                )
                manifest["steps"].append(notes_contract_diff_step)
            _write_manifest(manifest, log_dir)
            if notes_contract_step["exit_code"] != 0 or (
                notes_contract_diff_step is not None
                and notes_contract_diff_step["exit_code"] != 0
            ):
                print("Obsidian notes validation failed.")
                _write_manifest_with_release_packet(manifest, log_dir)
                return 1

        if not args.skip_decks:
            rename_result = _rename_deliverables(
                DECKS_ROOT / date_stamp / "land-only", date_stamp
            )
            manifest["steps"].append(rename_result)

        _write_manifest_with_release_packet(manifest, log_dir)
        _print_summary(manifest)

        return 0 if all(s["exit_code"] == 0 for s in manifest["steps"]) else 1
    finally:
        _release_single_flight_lock(single_flight_lock)


def _write_manifest(manifest, log_dir):
    manifest["finished_at"] = datetime.now().isoformat(timespec="seconds")
    path = log_dir / "manifest.json"
    with open(path, "w") as f:
        json.dump(manifest, f, indent=2)


def _write_manifest_with_release_packet(manifest, log_dir):
    _write_manifest(manifest, log_dir)
    try:
        try:
            from build_monthly_review_release_packet import (
                build_monthly_review_release_packet,
                build_release_packet_manifest_payload,
                write_monthly_review_release_packet_bundle,
            )
            from diff_monthly_review_release_packets import (
                build_snapshot_diff_bundle,
            )
            from build_monthly_review_release_packet_history import (
                refresh_release_packet_history,
            )
        except ModuleNotFoundError:  # pragma: no cover
            from scripts.build_monthly_review_release_packet import (
                build_monthly_review_release_packet,
                build_release_packet_manifest_payload,
                write_monthly_review_release_packet_bundle,
            )
            from scripts.diff_monthly_review_release_packets import (
                build_snapshot_diff_bundle,
            )
            from scripts.build_monthly_review_release_packet_history import (
                refresh_release_packet_history,
            )

        manifest_path = log_dir / "manifest.json"
        packet = build_monthly_review_release_packet(
            manifest=manifest,
            manifest_path=manifest_path,
            repo_root=ROOT,
        )
        packet_dir = write_monthly_review_release_packet_bundle(
            output_root=OUTPUT_ROOT / "monthly_review_release_packets",
            packet=packet,
        )
        release_packet_payload = build_release_packet_manifest_payload(
            repo_root=ROOT,
            packet=packet,
            packet_dir=packet_dir,
        )
        try:
            packet_diff, packet_diff_dir = build_snapshot_diff_bundle(
                current_payload=packet,
                packet_root=OUTPUT_ROOT / "monthly_review_release_packets",
                output_root=OUTPUT_ROOT / "monthly_review_release_packet_snapshot_diff",
            )
            release_packet_payload = build_release_packet_manifest_payload(
                repo_root=ROOT,
                packet=packet,
                packet_dir=packet_dir,
                packet_diff=packet_diff,
                packet_diff_dir=packet_diff_dir,
            )
        except Exception as exc:  # pragma: no cover
            release_packet_payload.update(
                {
                    "snapshot_diff_status": "error",
                    "snapshot_diff_error": str(exc),
                }
            )
        try:
            history_payload, history_dir = refresh_release_packet_history(
                packet_root=OUTPUT_ROOT / "monthly_review_release_packets",
                packet_diff_root=OUTPUT_ROOT
                / "monthly_review_release_packet_snapshot_diff",
                output_root=OUTPUT_ROOT / "monthly_review_release_packet_history",
            )
            latest_packet_diff = history_payload.get("latest_packet_diff") or {}
            latest_core_state_transition = (
                history_payload.get("latest_core_state_transition") or {}
            )
            release_packet_payload.update(
                {
                    "history_generated_at": history_payload.get("generated_at"),
                    "history_run_count": history_payload.get("run_count"),
                    "history_green_run_count": history_payload.get("green_run_count"),
                    "history_blocked_run_count": history_payload.get(
                        "blocked_run_count"
                    ),
                    "history_current_green_streak": history_payload.get(
                        "current_green_streak"
                    ),
                    "history_latest_core_state_transition_baseline_run_date": latest_core_state_transition.get(
                        "baseline_run_date"
                    ),
                    "history_latest_core_state_transition_run_date": latest_core_state_transition.get(
                        "current_run_date"
                    ),
                    "history_latest_core_state_transition_changes": list(
                        latest_core_state_transition.get("core_state_changes") or []
                    ),
                    "history_latest_core_state_transition_publish_blockers_added": list(
                        latest_core_state_transition.get("publish_blockers_added") or []
                    ),
                    "history_latest_core_state_transition_publish_blockers_resolved": list(
                        latest_core_state_transition.get("publish_blockers_resolved")
                        or []
                    ),
                    "history_latest_core_state_transition_pipeline_blockers_added": list(
                        latest_core_state_transition.get("pipeline_blockers_added")
                        or []
                    ),
                    "history_latest_core_state_transition_pipeline_blockers_resolved": list(
                        latest_core_state_transition.get("pipeline_blockers_resolved")
                        or []
                    ),
                    "history_latest_blocked_run_date": history_payload.get(
                        "latest_blocked_run_date"
                    ),
                    "history_latest_blocked_publish_blockers": list(
                        history_payload.get("latest_blocked_publish_blockers") or []
                    ),
                    "history_latest_blocked_pipeline_blockers": list(
                        history_payload.get("latest_blocked_pipeline_blockers") or []
                    ),
                    "history_latest_drift_baseline_run_date": latest_packet_diff.get(
                        "baseline_run_date"
                    ),
                    "history_latest_drift_run_date": latest_packet_diff.get(
                        "current_run_date"
                    ),
                    "history_latest_drift_changed_gates": list(
                        latest_packet_diff.get("changed_gates") or []
                    ),
                    "history_latest_drift_change_summaries": list(
                        latest_packet_diff.get("gate_change_summaries") or []
                    ),
                    "history_dir": _manifest_path(history_dir),
                    "history_json_path": _manifest_path(history_dir / "history.json"),
                    "history_summary_path": _manifest_path(history_dir / "summary.md"),
                }
            )
        except Exception as exc:  # pragma: no cover
            release_packet_payload.update(
                {
                    "history_status": "error",
                    "history_error": str(exc),
                }
            )
        manifest["release_packet"] = release_packet_payload
    except Exception as exc:  # pragma: no cover
        manifest["release_packet"] = {
            "status": "error",
            "error": str(exc),
        }
    _write_manifest(manifest, log_dir)


def _print_summary(manifest):
    print()
    print("=" * 70)
    print("Monthly Sales Directors Review, pipeline summary")
    print("=" * 70)
    print("Flow: Salesforce extract  ->  Excel analysis  ->  PowerPoint decks")
    print()
    for step in manifest["steps"]:
        print(
            f"  [{step['status']:6s}]  {step['name']:36s}"
            f"  {step['duration_seconds']:>6.1f}s"
        )
    print()
    run_date = manifest["run_date"]
    print("Outputs:")
    print(
        f"  Salesforce extracts:  {len(manifest['outputs']['extracts'])} workbooks"
        f"  (output/director_live_workbooks/{run_date}/)"
    )
    print(
        f"  Analysis reports:     {len(manifest['outputs']['reports'])} workbooks"
        f"  (output/sharepoint/)"
    )
    print(
        f"  PowerPoint decks:     {len(manifest['outputs']['decks'])} decks"
        f"  (output/simcorp_director_decks/{run_date}/land-only/)"
    )
    print()
    print(f"Manifest: output/pipeline_logs/{run_date}/manifest.json")
    if manifest["outputs"]["decks"]:
        print()
        print("Decks ready to ship:")
        for d in manifest["outputs"]["decks"]:
            print(f"  {d['file']}")


if __name__ == "__main__":
    sys.exit(main())
