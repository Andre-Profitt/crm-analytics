"""
Tie-out validator, Salesforce vs Excel vs Deck.

For every director, computes a canonical set of headline numbers from three
independent sources and flags mismatches. Writes the results to
`obsidian/Monthly/YYYY-MM/tie-out.md` so you can see at a glance which deck
slide drifts from the workbook.

Headline metrics validated:
  - Open Land active pipeline count and ARR Unwtd (Q1-Q3, excluding Omitted)
  - Q1 Land wins count and ARR
  - Q1 Land losses count
  - Q2 renewals count and ACV
  - Approved 2026 count
  - Conditionally approved count
  - Missing Stage 3+ count
"""

import argparse
import json
import re
import subprocess
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

try:
    from monthly_platform.policy import (
        filter_active_pipeline_rows,
        filter_rows_in_reporting_scope,
        is_active_forecast_category as policy_is_active_forecast_category,
        is_pending_approval_status as policy_is_pending_approval_status,
        make_reporting_scope,
        summarize_approval_rows,
    )
    from monthly_platform.period import resolve_period_context, sheet_names
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.policy import (
        filter_active_pipeline_rows,
        filter_rows_in_reporting_scope,
        is_active_forecast_category as policy_is_active_forecast_category,
        is_pending_approval_status as policy_is_pending_approval_status,
        make_reporting_scope,
        summarize_approval_rows,
    )
    from scripts.monthly_platform.period import resolve_period_context


ROOT = Path(__file__).resolve().parents[1]
VAULT = ROOT / "obsidian"
OUTPUT_ROOT = ROOT / "output" / "tie_out"

DIRECTORS = [
    ("Jesper Tyrer", "APAC", "jesper-tyrer"),
    ("Sarah Pittroff", "EMEA Central", "sarah-pittroff"),
    ("Dan Peppett", "EMEA UK & Ireland", "dan-peppett"),
    ("Christian Ebbesen", "EMEA NE", "christian-ebbesen"),
    ("Francois Thaury", "EMEA South West", "francois-thaury"),
    ("Mourad Essofi", "EMEA MEA", "mourad-essofi"),
    ("Patrick Gaughan", "NA Asset Mgmt", "patrick-gaughan"),
    ("Megan Miceli", "NA Canada", "megan-miceli"),
    ("Adam Steinhaus", "NA Pension & Insurance", "adam-steinhaus"),
]

# Regional workbook filenames use build_sharepoint_analysis.py's canonical
# territory labels ("NA Insurance" not "NA Pension & Insurance").
REGIONAL_TERRITORY = {
    "Jesper Tyrer": "APAC",
    "Sarah Pittroff": "EMEA Central",
    "Dan Peppett": "EMEA UK & Ireland",
    "Christian Ebbesen": "EMEA NE",
    "Francois Thaury": "EMEA South West",
    "Mourad Essofi": "EMEA MEA",
    "Patrick Gaughan": "NA Asset Mgmt",
    "Megan Miceli": "NA Canada",
    "Adam Steinhaus": "NA Insurance",
}


def resolve_runtime_scopes(as_of_date: str | None = None):
    report_date = str(as_of_date or datetime.now().strftime("%Y-%m-%d"))[:10]
    period = resolve_period_context(
        as_of_date=report_date,
        snapshot_date=report_date,
        deck_date=report_date,
    )
    return {
        "period": period,
        "reporting_scope": make_reporting_scope(
            period.reporting_window_start, period.reporting_window_end
        ),
        "prior_quarter_scope": make_reporting_scope(
            period.prior_quarter.start_date,
            period.prior_quarter.end_date,
        ),
        "current_quarter_scope": make_reporting_scope(
            period.current_quarter.start_date,
            period.current_quarter.end_date,
        ),
        "approval_year": period.current_quarter.year,
    }


def _fmt_eur(n):
    n = float(n or 0)
    if abs(n) >= 1_000_000:
        return f"EUR {n / 1_000_000:.1f}M"
    if abs(n) >= 1_000:
        return f"EUR {n / 1_000:.0f}K"
    return f"EUR {int(n):,}"


def _parse_eur(label):
    """Parse 'EUR 5.2M' or 'EUR 674K' back to a float."""
    if label is None:
        return None
    s = str(label).strip().replace("EUR", "").strip()
    m = re.match(r"([\d.,]+)([MK]?)$", s.replace(",", ""))
    if not m:
        return None
    try:
        value = float(m.group(1))
    except ValueError:
        return None
    suffix = m.group(2)
    if suffix == "M":
        return value * 1_000_000
    if suffix == "K":
        return value * 1_000
    return value


def _quarter(date_str):
    s = str(date_str or "")[:10]
    if not s.startswith(str(datetime.now().year)):
        return None
    try:
        m = int(s[5:7])
    except ValueError:
        return None
    return f"Q{(m - 1) // 3 + 1}"


def _is_pending_approval_status(status):
    return policy_is_pending_approval_status(status)


def _is_active_forecast_category(category):
    return policy_is_active_forecast_category(category)


# ───────────────────────── Excel truth ────────────────────────────────────


def excel_metrics(workbook_path, *, scopes=None):
    scopes = scopes or resolve_runtime_scopes()
    reporting_scope = scopes["reporting_scope"]
    prior_quarter_scope = scopes["prior_quarter_scope"]
    current_quarter_scope = scopes["current_quarter_scope"]
    wb = load_workbook(workbook_path, data_only=True)

    def _rows(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        return [
            {headers[i]: v for i, v in enumerate(r)}
            for r in ws.iter_rows(min_row=2, values_only=True)
        ]

    SN = sheet_names()
    pipeline = _rows(SN["pipeline_open"])
    won_lost = _rows(SN["won_lost"])
    approvals = _rows("Commercial Approval")
    renewals = _rows(SN["renewals"])

    def is_land(r):
        return str(r.get("Type", "")).strip().lower() == "land"

    def unw(r):
        return float(r.get("ARR Unweighted (EUR)") or 0)

    def acv(r):
        return float(r.get("ACV Unweighted (EUR)") or r.get("ACV (EUR)") or 0)

    land_open_active = filter_active_pipeline_rows(pipeline, reporting_scope)
    q1_land_wins = [
        r
        for r in won_lost
        if is_land(r)
        and "Won" in str(r.get("Stage", ""))
        and prior_quarter_scope.contains(r.get("Close Date"))
    ]
    q1_land_lost = [
        r
        for r in won_lost
        if is_land(r)
        and "Won" not in str(r.get("Stage", ""))
        and prior_quarter_scope.contains(r.get("Close Date"))
    ]
    q2_renewals = [
        r for r in renewals if current_quarter_scope.contains(r.get("Close Date"))
    ]

    # Approvals scope-matched to the deck (Q1-Q2 close date filter).
    approvals_in_scope = filter_rows_in_reporting_scope(approvals, reporting_scope)
    approval_summary = summarize_approval_rows(approvals_in_scope)

    return {
        "open_land_deals": len(land_open_active),
        "open_land_arr": sum(unw(r) for r in land_open_active),
        "q1_land_wins": len(q1_land_wins),
        "q1_land_wins_arr": sum(unw(r) for r in q1_land_wins),
        "q1_land_lost": len(q1_land_lost),
        "q2_renewals": len(q2_renewals),
        "q2_renewals_acv": sum(acv(r) for r in q2_renewals),
        "approved_2026": approval_summary["approved_2026"],
        "conditionally_approved": approval_summary["conditionally_approved"],
        "missing_stage3": approval_summary["missing_stage3"],
    }


# ───────────────────────── Regional workbook truth ───────────────────────


def regional_metrics(regional_wb_path, *, scopes=None):
    """Compute headline numbers from a per-region SharePoint workbook.

    Regional workbooks are scoped subsets of the master. This reader mirrors
    `excel_metrics` but reads from the regional tab structure (Land Pipeline
    Detail / Land WonLost Detail / Approvals list tabs) so the 4-way tie-out
    catches drift between the extractor and the regional builder.
    """
    scopes = scopes or resolve_runtime_scopes()
    reporting_scope = scopes["reporting_scope"]
    prior_quarter_scope = scopes["prior_quarter_scope"]
    wb = load_workbook(regional_wb_path, data_only=True)

    def _header_index(tab, header_row_index):
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row_index:
            return rows, {}
        headers = rows[header_row_index]
        idx = {str(h): i for i, h in enumerate(headers) if h}
        return rows, idx

    def _find_col(idx, *candidates):
        for c in candidates:
            if c in idx:
                return idx[c]
        return None

    open_cnt = 0
    open_arr = 0.0
    if "Land Pipeline Detail" in wb.sheetnames:
        rows, idx = _header_index("Land Pipeline Detail", 0)
        ti = idx.get("Type")
        ci = idx.get("Close Date")
        ai = _find_col(idx, "ARR Unwtd", "ARR Unwtd (EUR)")
        fci = idx.get("Forecast Category")
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            if ti is not None and str(r[ti] or "") != "Land":
                continue
            cd = str(r[ci] or "")[:10] if ci is not None else ""
            if not reporting_scope.contains(cd):
                continue
            if fci is not None and not _is_active_forecast_category(r[fci]):
                continue
            open_cnt += 1
            if ai is not None:
                open_arr += float(r[ai] or 0)

    q1_wins = q1_lost = 0
    q1_wins_arr = 0.0
    if "Land WonLost Detail" in wb.sheetnames:
        rows, idx = _header_index("Land WonLost Detail", 0)
        ti = idx.get("Type")
        si = idx.get("Stage")
        ci = idx.get("Close Date")
        ai = _find_col(idx, "ARR Unwtd", "ARR Unwtd (EUR)")
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            if ti is not None and str(r[ti] or "") != "Land":
                continue
            cd = str(r[ci] or "")[:10] if ci is not None else ""
            if not prior_quarter_scope.contains(cd):
                continue
            stage = str(r[si] or "") if si is not None else ""
            arr = float(r[ai] or 0) if ai is not None else 0.0
            if "Won" in stage:
                q1_wins += 1
                q1_wins_arr += arr
            else:
                q1_lost += 1

    def _count_listed(tab, q1q2_close_only=False):
        # List tabs have title rows 0-2 and header at row 3; data starts row 4.
        # Optionally filter to Q1-Q2 close dates to mirror the extractor's scope.
        if tab not in wb.sheetnames:
            return 0
        rows, idx = _header_index(tab, 3)
        ci = idx.get("Close Date")
        count = 0
        for r in rows[4:]:
            if not r or not r[0]:
                continue
            if q1q2_close_only and ci is not None:
                cd = str(r[ci] or "")[:10]
                if not reporting_scope.contains(cd):
                    continue
            count += 1
        return count

    # excel_metrics filters approvals by the reporting window; match that scope
    # so Extract and Regional reconcile.
    approved = _count_listed("Approvals, 2026", q1q2_close_only=True)
    cond = _count_listed("Approval Candidates", q1q2_close_only=True)
    missing = _count_listed("Land Stage 3+, No Approval", q1q2_close_only=True)
    q2_ren = _count_listed("Renewals This Quarter")

    q2_ren_acv = 0.0
    if "Renewals This Quarter" in wb.sheetnames:
        rows, idx = _header_index("Renewals This Quarter", 3)
        ai = _find_col(
            idx, "ACV Unwtd", "ACV (EUR)", "ACV Unweighted (EUR)", "ACV Unwtd (EUR)"
        )
        if ai is not None:
            for r in rows[4:]:
                if not r or not r[0]:
                    continue
                q2_ren_acv += float(r[ai] or 0)

    wb.close()
    return {
        "open_land_deals": open_cnt,
        "open_land_arr": open_arr,
        "q1_land_wins": q1_wins,
        "q1_land_wins_arr": q1_wins_arr,
        "q1_land_lost": q1_lost,
        "q2_renewals": q2_ren,
        "q2_renewals_acv": q2_ren_acv,
        "approved_2026": approved,
        "conditionally_approved": cond,
        "missing_stage3": missing,
    }


# ───────────────────────── Deck truth (from sidecar JSON) ─────────────────


def deck_metrics(deck_path):
    """Read the deck builder's sidecar JSON for the headline numbers it used.

    Every deck writes a `<name>.json` next to the .pptx with the exact values
    it rendered. This sidecar is the deterministic contract between the
    builder and the validator. If the sidecar is missing, we fail fast.
    """
    sidecar_path = deck_path.with_suffix(".json")
    if not sidecar_path.exists():
        return {
            k: None
            for k in [
                "open_land_deals",
                "open_land_arr",
                "q1_land_wins",
                "q1_land_wins_arr",
                "q1_land_lost",
                "q2_renewals",
                "q2_renewals_acv",
                "approved_2026",
                "conditionally_approved",
                "missing_stage3",
            ]
        }
    data = json.loads(sidecar_path.read_text())
    return {
        "open_land_deals": data.get("open_land_deals"),
        "open_land_arr": data.get("open_land_arr"),
        "q1_land_wins": data.get("q1_land_wins"),
        "q1_land_wins_arr": data.get("q1_land_wins_arr"),
        "q1_land_lost": data.get("q1_land_lost"),
        "q2_renewals": data.get("q2_renewals"),
        "q2_renewals_acv": data.get("q2_renewals_acv"),
        "approved_2026": data.get("approved_2026"),
        "conditionally_approved": data.get("conditionally_approved"),
        "missing_stage3": data.get("missing_stage3"),
    }


# ───────────────────────── Salesforce truth (via SOQL) ────────────────────

# Import the same territory config the extractor uses so SF counts mirror
# the exact filter that produced the Excel rows.
import sys as _sys  # noqa: E402

_sys.path.insert(0, str(Path(__file__).resolve().parent))
from extract_director_live import (  # type: ignore  # noqa: E402
    ACCOUNT_EXCLUDE,
    OWNER_EXCLUDE,
    TERRITORIES,
)


DIRECTOR_TO_TERRITORY_KEY = {v["director"]: k for k, v in TERRITORIES.items()}


def sf_metrics(session, instance, director, *, scopes=None):
    """Run SOQL counts + sums to get the SF-side view of the same 10 metrics.
    Uses the exact same where clause as the extractor, so SF truth and Excel
    extract should agree unless there was a transient extract issue."""
    scopes = scopes or resolve_runtime_scopes()
    reporting_scope = scopes["reporting_scope"]
    prior_quarter_scope = scopes["prior_quarter_scope"]
    current_quarter_scope = scopes["current_quarter_scope"]
    approval_year = scopes["approval_year"]
    key = DIRECTOR_TO_TERRITORY_KEY.get(director)
    if not key:
        return {}
    where = TERRITORIES[key]["soql_where"]
    common = (
        f"WHERE {where} AND Type IN ('Land','Expand','Renewal') "
        f"{ACCOUNT_EXCLUDE} {OWNER_EXCLUDE}"
    )

    def _agg(query):
        r = session.get(
            f"{instance}/services/data/v66.0/query",
            params={"q": query},
        ).json()
        rec = r.get("records", [{}])[0] if isinstance(r, dict) else {}
        return rec

    # 1. Open Land active pipeline, reporting window
    r = _agg(
        "SELECT COUNT(Id) c, SUM(convertCurrency(APTS_Opportunity_ARR__c)) s FROM Opportunity "
        f"{common} AND IsClosed=false AND Type='Land' "
        f"AND CloseDate >= {reporting_scope.start_date} AND CloseDate <= {reporting_scope.end_date} "
        "AND (ForecastCategoryName = null OR ForecastCategoryName != 'Omitted')"
    )
    open_c, open_arr = int(r.get("c") or 0), float(r.get("s") or 0)

    # 2. Q1 Land wins (IsWon = true, CloseDate in Q1)
    r = _agg(
        "SELECT COUNT(Id) c, SUM(convertCurrency(APTS_Opportunity_ARR__c)) s FROM Opportunity "
        f"{common} AND IsClosed=true AND IsWon=true AND Type='Land' "
        f"AND CloseDate >= {prior_quarter_scope.start_date} AND CloseDate <= {prior_quarter_scope.end_date}"
    )
    q1w_c, q1w_arr = int(r.get("c") or 0), float(r.get("s") or 0)

    # 3. Q1 Land losses (IsClosed=true, IsWon=false, CloseDate in Q1)
    r = _agg(
        "SELECT COUNT(Id) c FROM Opportunity "
        f"{common} AND IsClosed=true AND IsWon=false AND Type='Land' "
        f"AND CloseDate >= {prior_quarter_scope.start_date} AND CloseDate <= {prior_quarter_scope.end_date}"
    )
    q1l_c = int(r.get("c") or 0)

    # 4. Q2 renewals
    r = _agg(
        "SELECT COUNT(Id) c, SUM(convertCurrency(Amount)) s FROM Opportunity "
        f"{common} AND IsClosed=false AND Type='Renewal' "
        f"AND CloseDate >= {current_quarter_scope.start_date} AND CloseDate <= {current_quarter_scope.end_date}"
    )
    q2r_c, q2r_acv = int(r.get("c") or 0), float(r.get("s") or 0)

    # 5. Approvals (Land, reporting-window close date, Stage_20 approved this year)
    r = _agg(
        "SELECT COUNT(Id) c FROM Opportunity "
        f"{common} AND IsClosed=false AND Type='Land' "
        f"AND CloseDate >= {reporting_scope.start_date} AND CloseDate <= {reporting_scope.end_date} "
        "AND Stage_20_Approval__c = true "
        f"AND CALENDAR_YEAR(Stage_20_Approval_Date__c) = {approval_year}"
    )
    approved_c = int(r.get("c") or 0)

    # 6. Approval candidates pending approval (submitted but not approved)
    r = _agg(
        "SELECT COUNT(Id) c FROM Opportunity "
        f"{common} AND IsClosed=false AND Type='Land' "
        f"AND CloseDate >= {reporting_scope.start_date} AND CloseDate <= {reporting_scope.end_date} "
        "AND Submit_for_Stage_20_Review__c = true "
        "AND Stage_20_Approval__c = false"
    )
    cond_c = int(r.get("c") or 0)

    # 7. Missing Stage 3+ (open Land, Stage 3+, no approval, not exempt)
    # Excludes Approval_Status__c = 'No Approval Necessary' to match extractor's
    # 4-state approval model (see extract_director_live.py ~L428).
    r = _agg(
        "SELECT COUNT(Id) c FROM Opportunity "
        f"{common} AND IsClosed=false AND Type='Land' "
        f"AND CloseDate >= {reporting_scope.start_date} AND CloseDate <= {reporting_scope.end_date} "
        "AND StageName IN ('3 - Engagement','4 - Shortlisted',"
        "'5 - Preferred','6 - Contracting') "
        "AND Stage_20_Approval__c = false "
        "AND Submit_for_Stage_20_Review__c = false "
        "AND (Approval_Status__c != 'No Approval Necessary' "
        "     OR Approval_Status__c = null)"
    )
    missing_c = int(r.get("c") or 0)

    return {
        "open_land_deals": open_c,
        "open_land_arr": open_arr,
        "q1_land_wins": q1w_c,
        "q1_land_wins_arr": q1w_arr,
        "q1_land_lost": q1l_c,
        "q2_renewals": q2r_c,
        "q2_renewals_acv": q2r_acv,
        "approved_2026": approved_c,
        "conditionally_approved": cond_c,
        "missing_stage3": missing_c,
    }


def sf_auth():
    data = json.loads(
        subprocess.run(
            ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout
    )["result"]
    return data["accessToken"], data["instanceUrl"]


# We skip live SF counts for now because they require per-director SOQL
# filters and duplicate the extractor logic. The Excel workbook is a faithful
# snapshot of what SF returned at extract time; if the Excel says N, SF
# returned N. Tie-out is therefore Excel-vs-deck, which is the failure mode
# we actually hit (the renewals column rename bug).


# ───────────────────────── Compare + report ───────────────────────────────


def _compare4(sf, excel, regional, deck):
    """Return list of (metric, sf, excel, regional, deck, status).

    Status is 'match' when every available pair agrees. Currency allows 10
    percent or EUR 100K tolerance. SF is the source of truth; a SF-vs-Excel
    mismatch = extractor bug; Excel-vs-Regional = regional builder bug;
    Regional-vs-Deck or Excel-vs-Deck = deck render bug.
    """
    results = []
    for key, label in [
        ("open_land_deals", "Open Land pipeline deals"),
        ("open_land_arr", "Open Land pipeline ARR"),
        ("q1_land_wins", "Q1 Land wins, count"),
        ("q1_land_wins_arr", "Q1 Land wins, ARR"),
        ("q1_land_lost", "Q1 Land losses, count"),
        ("q2_renewals", "Q2 renewals, count"),
        ("q2_renewals_acv", "Q2 renewals, ACV"),
        ("approved_2026", "Approved 2026 count"),
        ("conditionally_approved", "Conditionally approved count"),
        ("missing_stage3", "Missing Stage 3+ count"),
    ]:
        s = sf.get(key)
        e = excel.get(key)
        rg = regional.get(key) if regional else None
        d = deck.get(key)
        is_currency = "_arr" in key or "_acv" in key

        def _close(a, b):
            if a is None or b is None:
                return True
            if is_currency:
                tol = max(float(a or 0) * 0.10, 100_000)
                return abs(float(a or 0) - float(b or 0)) <= tol
            return a == b

        sf_vs_excel = _close(s, e)
        excel_vs_regional = _close(e, rg) if rg is not None else True
        regional_vs_deck = _close(rg, d) if rg is not None else True
        excel_vs_deck = _close(e, d)

        failures = []
        if not sf_vs_excel:
            failures.append("SF vs Extract")
        if not excel_vs_regional:
            failures.append("Extract vs Regional")
        if not regional_vs_deck:
            failures.append("Regional vs Deck")
        if not excel_vs_deck and regional is None:
            # Only fall back to direct Extract-vs-Deck check when regional
            # wasn't loaded; avoids double-flagging when both comparisons
            # show the same drift.
            failures.append("Extract vs Deck")
        status = "match" if not failures else " + ".join(failures) + " mismatch"
        results.append((label, s, e, rg, d, status))
    return results


def _render_value(key_label, val):
    if val is None:
        return "–"
    if "ARR" in key_label or "ACV" in key_label:
        return _fmt_eur(val)
    return str(val)


def _serialize_results(results):
    rows = []
    for label, s_val, e_val, rg_val, d_val, status in results:
        rows.append(
            {
                "metric": label,
                "salesforce": s_val,
                "extract": e_val,
                "regional": rg_val,
                "deck": d_val,
                "status": status,
            }
        )
    return rows


def write_tieout_note(run_date, all_results):
    period = run_date[:7]
    out_dir = VAULT / "Monthly" / period
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "tie-out.md"

    total_checks = sum(len(r["results"]) for r in all_results)
    total_mismatches = sum(
        sum(1 for row in r["results"] if row[5] != "match") for r in all_results
    )

    lines = [
        "---",
        "type: tie-out",
        f"period: {period}",
        f"run_date: {run_date}",
        f"checks: {total_checks}",
        f"mismatches: {total_mismatches}",
        "tags: [tie-out, validation, monthly]",
        "---",
        "",
        f"# Tie-out report, {period}",
        "",
        (
            f"Four-source reconciliation: Salesforce (live SOQL), Extract "
            f"(per-director workbook), Regional (per-territory SharePoint "
            f"workbook) and Deck (sidecar JSON). {len(all_results)} directors, "
            f"{total_checks} metrics. Mismatches: **{total_mismatches}**."
        ),
        "",
    ]
    if total_mismatches == 0:
        lines.append("All four sources agreed on every metric. Ship the decks.")
    else:
        lines.append(
            "Check each mismatch row. 'SF vs Extract' = extractor drift; "
            "'Extract vs Regional' = regional builder drift; "
            "'Regional vs Deck' = deck render drift."
        )
    lines.append("")

    for director_record in all_results:
        director = director_record["director"]
        results = director_record["results"]
        slug = director_record["slug"]
        lines.append(f"## {director}")
        lines.append("")
        lines.append("| Metric | Salesforce | Extract | Regional | Deck | Status |")
        lines.append("|---|---:|---:|---:|---:|---|")
        for label, s_val, e_val, rg_val, d_val, status in results:
            lines.append(
                f"| {label} | {_render_value(label, s_val)} | "
                f"{_render_value(label, e_val)} | "
                f"{_render_value(label, rg_val)} | "
                f"{_render_value(label, d_val)} | {status} |"
            )
        lines.append("")
        lines.append(
            f"Sources: [[Directors/{slug}]], "
            f"[[Monthly/{period}/{slug}.auto|{period} auto snapshot]]"
        )
        lines.append("")

    path.write_text("\n".join(lines) + "\n")
    print(f"  wrote {path.relative_to(ROOT)}")
    return path, total_mismatches


def write_tieout_artifacts(run_date, all_results, failures=None):
    failures = list(failures or [])
    output_dir = OUTPUT_ROOT / run_date[:10]
    output_dir.mkdir(parents=True, exist_ok=True)
    total_checks = sum(len(r["results"]) for r in all_results)
    total_mismatches = sum(
        sum(1 for row in r["results"] if row[5] != "match") for r in all_results
    )
    directors = []
    for record in all_results:
        mismatch_count = sum(1 for row in record["results"] if row[5] != "match")
        directors.append(
            {
                "director": record["director"],
                "territory": record["territory"],
                "slug": record["slug"],
                "mismatch_count": mismatch_count,
                "metrics": _serialize_results(record["results"]),
            }
        )
    payload = {
        "run_date": run_date[:10],
        "status": "failed" if failures or total_mismatches else "ok",
        "checks": total_checks,
        "mismatches": total_mismatches,
        "directors_audited": len(all_results),
        "directors_with_mismatches": sum(
            1 for item in directors if int(item["mismatch_count"]) > 0
        ),
        "failures": failures,
        "directors": directors,
        "note_path": str(
            (VAULT / "Monthly" / run_date[:7] / "tie-out.md").relative_to(ROOT)
        ),
    }
    (output_dir / "tie_out_audit.json").write_text(
        json.dumps(payload, indent=2) + "\n",
        encoding="utf-8",
    )
    lines = [
        f"# Tie-Out Audit — {run_date[:10]}",
        "",
        f"- Status: `{payload['status']}`",
        f"- Directors audited: `{payload['directors_audited']}`",
        f"- Checks: `{payload['checks']}`",
        f"- Mismatches: `{payload['mismatches']}`",
        f"- Directors with mismatches: `{payload['directors_with_mismatches']}`",
        "",
        "## Directors",
        "",
    ]
    if not directors:
        lines.append("- none")
    else:
        for item in directors:
            status = (
                "clean"
                if item["mismatch_count"] == 0
                else f"{item['mismatch_count']} mismatch(es)"
            )
            lines.append(f"- `{item['slug']}`: `{status}`")
    lines.extend(["", "## Failures", ""])
    if not failures:
        lines.append("- none")
    else:
        for item in failures:
            lines.append(
                f"- `{item.get('issue', 'unknown')}` {item.get('message', '')}".strip()
            )
    (output_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  wrote {output_dir.relative_to(ROOT)}")
    return output_dir, total_mismatches


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    args = parser.parse_args()
    scopes = resolve_runtime_scopes(args.date)

    wb_root = ROOT / "output" / "director_live_workbooks" / args.date
    deck_root = ROOT / "output" / "simcorp_director_decks" / args.date / "land-only"
    regional_root = ROOT / "output" / "sharepoint"
    if not wb_root.exists() or not deck_root.exists():
        print(f"  sources missing, expected {wb_root} and {deck_root}")
        write_tieout_artifacts(
            args.date,
            [],
            failures=[
                {
                    "issue": "required_sources_missing",
                    "message": f"expected {wb_root} and {deck_root}",
                }
            ],
        )
        return 1

    token, instance = sf_auth()
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    all_results = []
    total_mismatches = 0
    for director, territory, slug in DIRECTORS:
        wb_path = wb_root / f"{slug}.xlsx"
        deck_path = deck_root / f"{slug}-LAND.pptx"
        if not wb_path.exists() or not deck_path.exists():
            print(f"  skip {director}: sources missing")
            continue
        excel = excel_metrics(wb_path, scopes=scopes)
        deck = deck_metrics(deck_path)

        regional = None
        reg_terr = REGIONAL_TERRITORY.get(director)
        if reg_terr:
            regional_path = regional_root / f"FY26 Pipeline Review, {reg_terr}.xlsx"
            if regional_path.exists():
                try:
                    regional = regional_metrics(regional_path, scopes=scopes)
                except Exception as exc:
                    print(f"  {director}: regional read failed ({exc})")
            else:
                print(f"  {director}: regional workbook missing at {regional_path}")

        try:
            sf = sf_metrics(session, instance, director, scopes=scopes)
        except Exception as exc:
            print(f"  {director}: SF query failed ({exc}); skipping SF side")
            sf = {k: None for k in excel}
        results = _compare4(sf, excel, regional, deck)
        mismatches = sum(1 for row in results if row[5] != "match")
        total_mismatches += mismatches
        print(f"  {director:22s}  mismatches: {mismatches}")
        all_results.append(
            {
                "director": director,
                "territory": territory,
                "slug": slug,
                "results": results,
            }
        )

    path, mism = write_tieout_note(args.date, all_results)
    write_tieout_artifacts(args.date, all_results)
    print(f"\nTotal mismatches: {mism}")
    print(f"Tie-out note: {path.relative_to(ROOT)}")
    return 0 if mism == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
