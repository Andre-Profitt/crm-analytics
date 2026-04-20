"""
Data-quality audit — scans SF pipeline, account, installation, and quote data
for hygiene gaps, scored and tiered by severity.

Concept extracted from Hooman Hashemi's Sales Ops Commercial Health &
Governance Dashboard: every check names a *breach* and a *consequence*, each
has a plain-English SF logic rule for self-documentation, and checks are
tiered into Critical / Important / Domain so exec attention cascades.

Run standalone or as pipeline stage 1c_data_quality_audit. Outputs:

  output/data_quality/YYYY-MM-DD/flags.json        — every check + count + severity
  output/data_quality/YYYY-MM-DD/summary.md        — exec-facing tiered summary
  output/data_quality/history.json                 — rolling MoM ledger

Ledger enables month-over-month deltas: "Stage 3+ no NextStep went 327 → 280."
"""

from __future__ import annotations

import argparse
import json
import subprocess
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

import requests


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "output" / "data_quality"
LEDGER_PATH = OUTPUT_ROOT / "history.json"

FY_START = "2026-01-01"
FY_END = "2026-12-31"
Q1_END = "2026-03-31"


# ────────────────────────── Check registry ──────────────────────────────────


@dataclass
class HygieneCheck:
    """One named data-quality rule that returns a count from SF.

    Attributes:
        key: stable identifier used in ledger + JSON output (never renamed)
        label: exec-facing name shown on dashboards
        grain: object grain this runs against (deal/account/installation/quote/closed_deal/whitespace)
        severity: Critical / Important / Domain — drives ordering + color
        category: loose grouping (hygiene/approval/activity/slip/finance/renewal/…)
        soql: full SOQL query that returns a single COUNT()
        sf_logic: plain-English rule description for the self-documenting alerts table
        action: what an owner should do if this fires (for runbook automation later)
        baseline: True for total-count rows that aren't themselves gaps
    """

    key: str
    label: str
    grain: str
    severity: str
    category: str
    soql: str
    sf_logic: str
    action: str = ""
    baseline: bool = False


# Scope constants (reused inside the SOQL strings)
OPP_OPEN = (
    f"IsClosed = false AND Type IN ('Land','Expand','Renewal') "
    f"AND CloseDate >= {FY_START} AND CloseDate <= {FY_END}"
)
OPP_CLOSED_Q1 = (
    f"IsClosed = true AND Type IN ('Land','Expand') "
    f"AND CloseDate >= {FY_START} AND CloseDate <= {Q1_END}"
)
ACCT_ACTIVE = (
    "Id IN (SELECT AccountId FROM Opportunity "
    f"WHERE IsClosed = false AND CloseDate >= {FY_START} AND CloseDate <= {FY_END} "
    "AND Type IN ('Land','Expand','Renewal'))"
)

CHECKS: list[HygieneCheck] = [
    # ─── BASELINE counts (not gaps) ──────────────────────────────────────
    HygieneCheck(
        key="total_open",
        label="Total open FY26 pipeline",
        grain="deal",
        severity="baseline",
        category="baseline",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN}",
        sf_logic=f"COUNT(Opportunity) WHERE {OPP_OPEN}",
        baseline=True,
    ),
    HygieneCheck(
        key="total_q1_closed",
        label="Total closed Q1 2026 (Land+Expand)",
        grain="closed_deal",
        severity="baseline",
        category="baseline",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_CLOSED_Q1}",
        sf_logic=f"COUNT(Opportunity) WHERE {OPP_CLOSED_Q1}",
        baseline=True,
    ),
    HygieneCheck(
        key="total_active_accounts",
        label="Accounts with FY26 open pipeline",
        grain="account",
        severity="baseline",
        category="baseline",
        soql=f"SELECT COUNT() FROM Account WHERE {ACCT_ACTIVE}",
        sf_logic=f"COUNT(Account) WHERE {ACCT_ACTIVE}",
        baseline=True,
    ),
    # ─── CRITICAL · governance breaches ──────────────────────────────────
    HygieneCheck(
        key="stage3_plus_500k_no_approval",
        label="Stage 3+ ≥$500k, no commercial approval (excl. exempt)",
        grain="deal",
        severity="Critical",
        category="governance",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting') "
            "AND APTS_Opportunity_ARR__c >= 500000 "
            "AND Stage_20_Approval__c = NULL "
            "AND Approval_Status__c != 'No Approval Necessary'"
        ),
        sf_logic="Stage ≥ 3 AND ARR ≥ 500k AND not approved AND not exempt (Approval_Status ≠ 'No Approval Necessary')",
        action="Escalate to approval committee. Deal past governance checkpoint without sign-off.",
    ),
    HygieneCheck(
        key="land_stage3_no_approval_any",
        label="Land Stage 3+ needs approval but not submitted",
        grain="deal",
        severity="Critical",
        category="approval",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} AND Type = 'Land' "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting') "
            "AND Stage_20_Approval__c = NULL AND Submit_for_Stage_20_Review__c = NULL "
            "AND Approval_Status__c != 'No Approval Necessary'"
        ),
        sf_logic="Type='Land' AND Stage ≥ 3 AND not approved AND not submitted AND not exempt",
        action="Submit for Stage 20 review or record existing approval.",
    ),
    HygieneCheck(
        key="land_stage3_pending_approval",
        label="Land Stage 3+ submitted, awaiting approval",
        grain="deal",
        severity="Important",
        category="approval",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} AND Type = 'Land' "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting') "
            "AND Stage_20_Approval__c = NULL "
            "AND (Submit_for_Stage_20_Review__c != NULL OR Approval_Status__c = 'Needs Approval')"
        ),
        sf_logic="Type='Land' AND Stage ≥ 3 AND submitted but not yet approved",
        action="Track turnaround time. Escalate if pending >48h with close date <30d.",
    ),
    HygieneCheck(
        key="closed_won_oi_blank",
        label="Closed Won · OI (Order Inflow) blank",
        grain="closed_deal",
        severity="Critical",
        category="finance",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"StageName = '8 - Won' AND CloseDate >= {FY_START} "
            "AND (APTS_Forecast_NPP_UWT__c = NULL OR APTS_Forecast_NPP_UWT__c = 0)"
        ),
        sf_logic="StageName='8 - Won' AND FY26 close AND APTS_Forecast_NPP_UWT__c IS NULL OR 0",
        action="Finance cannot recognise revenue. Backfill OI on the deal.",
    ),
    HygieneCheck(
        key="closed_won_quota_retirement_blank",
        label="Closed Won · Quota Retirement blank",
        grain="closed_deal",
        severity="Critical",
        category="finance",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"StageName = '8 - Won' AND CloseDate >= {FY_START} "
            "AND (APTS_Forecast_Quota_Retirement__c = NULL OR APTS_Forecast_Quota_Retirement__c = 0)"
        ),
        sf_logic="StageName='8 - Won' AND APTS_Forecast_Quota_Retirement__c IS NULL OR 0",
        action="Quota credit will not apply. Populate Quota Retirement.",
    ),
    HygieneCheck(
        key="closedate_in_past_open",
        label="Close date in past · opportunity still open",
        grain="deal",
        severity="Critical",
        category="forecasting",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            "IsClosed = false AND CloseDate < TODAY "
            "AND Type IN ('Land','Expand','Renewal')"
        ),
        sf_logic="IsClosed=false AND CloseDate < TODAY AND Type IN (Land,Expand,Renewal)",
        action="Distorts pipeline and forecast. Reset close date or mark closed.",
    ),
    HygieneCheck(
        key="contract_expired_active_pipeline",
        label="Contract end date in past · account has open pipeline",
        grain="account",
        severity="Critical",
        category="renewal",
        soql=(
            "SELECT COUNT() FROM Account WHERE "
            f"{ACCT_ACTIVE} "
            "AND APTS_Contract_End_Date__c != NULL "
            "AND APTS_Contract_End_Date__c < TODAY"
        ),
        sf_logic="APTS_Contract_End_Date__c < TODAY AND account has active Opportunity",
        action="Renewal backlog. Route to Renewals team.",
    ),
    HygieneCheck(
        key="ghost_installation",
        label="Active Installation · ExtendedToDate in past",
        grain="installation",
        severity="Critical",
        category="asset-integrity",
        soql=(
            "SELECT COUNT() FROM Installation__c WHERE "
            "Status__c = 'Active' AND ExtendedToDate__c != NULL "
            "AND ExtendedToDate__c < TODAY"
        ),
        sf_logic="Installation__c.Status='Active' AND ExtendedToDate__c < TODAY",
        action="Ghost installation overstating book of business. Deactivate or extend.",
    ),
    HygieneCheck(
        key="no_short_code_stage3",
        label="Account with no Short Code · active Stage 3+ opp",
        grain="account",
        severity="Critical",
        category="finance-blocker",
        soql=(
            "SELECT COUNT() FROM Account WHERE "
            "(Short_Code__c = NULL OR Short_Code__c = '') "
            "AND Id IN (SELECT AccountId FROM Opportunity "
            "WHERE IsClosed = false "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting'))"
        ),
        sf_logic="Short_Code__c = NULL AND has Opportunity at Stage ≥ 3",
        action="Finance blocked. INS file cannot be updated. Assign Short Code.",
    ),
    HygieneCheck(
        key="kyc_not_approved_stage4",
        label="KYC not Approved · account has Stage 4+ opp",
        grain="account",
        severity="Critical",
        category="compliance",
        soql=(
            "SELECT COUNT() FROM Account WHERE "
            "(KYC_Approval_Status__c != 'Approved' OR KYC_Approval_Status__c = NULL) "
            "AND Id IN (SELECT AccountId FROM Opportunity "
            "WHERE IsClosed = false "
            "AND StageName IN ('4 - Shortlisted','5 - Preferred','6 - Contracting'))"
        ),
        sf_logic="KYC_Approval_Status__c ≠ 'Approved' AND has Stage ≥ 4 Opportunity",
        action="Deal cannot close until KYC cleared. Escalate to compliance.",
    ),
    HygieneCheck(
        key="owner_inactive_open_opp",
        label="Opportunity owner is inactive SF user",
        grain="deal",
        severity="Critical",
        category="assignment-rot",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} AND Owner.IsActive = false"
        ),
        sf_logic=f"{OPP_OPEN} AND Owner.IsActive = false",
        action="Person left. Alerts go nowhere. Reassign owner.",
    ),
    # ─── IMPORTANT · hygiene ─────────────────────────────────────────────
    HygieneCheck(
        key="mid_stage_no_next_step",
        label="Stage 3+ with no NextStep",
        grain="deal",
        severity="Important",
        category="hygiene",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} "
            "AND (NextStep = NULL OR NextStep = '') "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting')"
        ),
        sf_logic="Stage ≥ 3 AND NextStep IS NULL OR empty",
        action="Deal un-coachable. Owner must log next action.",
    ),
    HygieneCheck(
        key="no_activity_ever",
        label="LastActivityDate NULL (never logged)",
        grain="deal",
        severity="Important",
        category="activity",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND LastActivityDate = NULL",
        sf_logic=f"{OPP_OPEN} AND LastActivityDate = NULL",
        action="Integration audit — email/calendar sync may be broken.",
    ),
    HygieneCheck(
        key="no_activity_60d",
        label="LastActivityDate > 60 days ago",
        grain="deal",
        severity="Important",
        category="activity",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND LastActivityDate < LAST_N_DAYS:60",
        sf_logic=f"{OPP_OPEN} AND LastActivityDate < LAST_N_DAYS:60",
        action="Deal stale. Review and either progress or disqualify.",
    ),
    HygieneCheck(
        key="no_activity_90d",
        label="LastActivityDate > 90 days ago",
        grain="deal",
        severity="Important",
        category="activity",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND LastActivityDate < LAST_N_DAYS:90",
        sf_logic=f"{OPP_OPEN} AND LastActivityDate < LAST_N_DAYS:90",
        action="Very stale. Dead-deal candidate.",
    ),
    HygieneCheck(
        key="push_5_plus",
        label="PushCount ≥ 5 (chronic slip)",
        grain="deal",
        severity="Important",
        category="slip",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND PushCount >= 5",
        sf_logic=f"{OPP_OPEN} AND PushCount >= 5",
        action="Commit integrity coaching — repeat slip pattern.",
    ),
    HygieneCheck(
        key="push_3_plus",
        label="PushCount ≥ 3",
        grain="deal",
        severity="Important",
        category="slip",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND PushCount >= 3",
        sf_logic=f"{OPP_OPEN} AND PushCount >= 3",
        action="Push pattern review with owner.",
    ),
    HygieneCheck(
        key="aging_365_plus",
        label="Pipeline aged 365+ days",
        grain="deal",
        severity="Important",
        category="lifecycle",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND CreatedDate < LAST_N_DAYS:365",
        sf_logic=f"{OPP_OPEN} AND CreatedDate < LAST_N_DAYS:365",
        action="Monthly auto-review queue.",
    ),
    HygieneCheck(
        key="early_stage_imminent_close",
        label="Stage 1-2 with close date ≤30 days (unrealistic)",
        grain="deal",
        severity="Important",
        category="forecasting",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} "
            "AND StageName IN ('1 - Prospecting','2 - Discovery') "
            "AND CloseDate <= NEXT_N_DAYS:30"
        ),
        sf_logic="Stage IN (1,2) AND CloseDate <= TODAY+30",
        action="Sandbagging or misforecast. Verify close date.",
    ),
    HygieneCheck(
        key="close_date_dec31_placeholder",
        label="Close date = Dec 31 (placeholder) · Stage 3+",
        grain="deal",
        severity="Important",
        category="forecasting",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_OPEN} "
            "AND CALENDAR_MONTH(CloseDate) = 12 AND DAY_IN_MONTH(CloseDate) = 31 "
            "AND StageName IN ('3 - Engagement','4 - Shortlisted','5 - Preferred','6 - Contracting')"
        ),
        sf_logic="Stage ≥ 3 AND CloseDate month=12 AND day=31",
        action="Placeholder date. Inflates Q4 forecast. Set real close date.",
    ),
    HygieneCheck(
        key="missing_forecast_arr",
        label="Missing APTS_Forecast_ARR__c",
        grain="deal",
        severity="Important",
        category="forecasting",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND APTS_Forecast_ARR__c = NULL",
        sf_logic=f"{OPP_OPEN} AND APTS_Forecast_ARR__c = NULL",
        action="Forecast rollup impaired. Set Forecast ARR.",
    ),
    HygieneCheck(
        key="missing_lead_scope",
        label="Missing Lead_Scope__c",
        grain="deal",
        severity="Important",
        category="routing",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND Lead_Scope__c = NULL",
        sf_logic=f"{OPP_OPEN} AND Lead_Scope__c = NULL",
    ),
    HygieneCheck(
        key="missing_primary_quote",
        label="No APTS_Primary_Quote attached",
        grain="deal",
        severity="Important",
        category="quoting",
        soql=f"SELECT COUNT() FROM Opportunity WHERE {OPP_OPEN} AND APTS_Primary_Quote__c = NULL",
        sf_logic=f"{OPP_OPEN} AND APTS_Primary_Quote__c = NULL",
        action="Deal cannot ship quote. Stage-gate or process misalignment.",
    ),
    HygieneCheck(
        key="quote_over_12mo_active",
        label="Apttus quote >12 months old · opp still active",
        grain="quote",
        severity="Important",
        category="quote-stall",
        soql=(
            "SELECT COUNT() FROM Apttus_Proposal__Proposal__c WHERE "
            "CreatedDate < LAST_N_DAYS:365 "
            "AND Apttus_Proposal__Opportunity__r.IsClosed = false"
        ),
        sf_logic="Proposal.CreatedDate < TODAY-365 AND Opportunity.IsClosed = false",
        action="Pricing, AUM, product config may all be stale. Regenerate quote.",
    ),
    HygieneCheck(
        key="quote_in_review_5d",
        label="Apttus quote in review/approval >5 days",
        grain="quote",
        severity="Important",
        category="quote-stall",
        soql=(
            "SELECT COUNT() FROM Apttus_Proposal__Proposal__c WHERE "
            "Apttus_Proposal__Approval_Stage__c IN ('In Review','Approval Required') "
            "AND LastModifiedDate < LAST_N_DAYS:5"
        ),
        sf_logic="Proposal.Approval_Stage IN (In Review, Approval Required) AND LastModifiedDate < TODAY-5",
        action="Internally blocked — Legal, Finance, or Sales Ops.",
    ),
    HygieneCheck(
        key="nda_not_signed",
        label="NDA not signed · account has open pipeline",
        grain="account",
        severity="Important",
        category="legal",
        soql=f"SELECT COUNT() FROM Account WHERE {ACCT_ACTIVE} AND APTS_NDA_Signed__c = FALSE",
        sf_logic=f"APTS_NDA_Signed__c = FALSE AND {ACCT_ACTIVE}",
        action="Legal ops audit — either the flag is stale or NDA is genuinely outstanding.",
    ),
    HygieneCheck(
        key="missing_account_source",
        label="Account: missing AccountSource",
        grain="account",
        severity="Important",
        category="attribution",
        soql=(
            "SELECT COUNT() FROM Account WHERE "
            f"{ACCT_ACTIVE} AND (AccountSource = NULL OR AccountSource = '')"
        ),
        sf_logic=f"{ACCT_ACTIVE} AND AccountSource IS NULL or empty",
        action="Marketing attribution gap.",
    ),
    # ─── IMPORTANT · closed-deal hygiene ────────────────────────────────
    HygieneCheck(
        key="no_opp_no_reason",
        label="Q1 'No Opportunity' disqualified without Reason",
        grain="closed_deal",
        severity="Important",
        category="hygiene",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_CLOSED_Q1} "
            "AND StageName = '0 - No Opportunity' "
            "AND (Reason_Won_Lost__c = NULL OR Reason_Won_Lost__c = '')"
        ),
        sf_logic="Q1 Closed AND StageName='0 - No Opportunity' AND Reason_Won_Lost__c empty",
        action="Can't analyse disqualification pattern without reason.",
    ),
    HygieneCheck(
        key="lost_no_competitor",
        label="Q1 Lost without competitor attribution",
        grain="closed_deal",
        severity="Important",
        category="attribution",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_CLOSED_Q1} "
            "AND StageName = '0 - Lost' "
            "AND (Lost_to_Competitor__c = NULL OR Lost_to_Competitor__c = '')"
        ),
        sf_logic="Q1 Closed AND StageName='0 - Lost' AND Lost_to_Competitor__c empty",
        action="Can't build competitive playbook without attribution.",
    ),
    HygieneCheck(
        key="won_zero_arr",
        label="Q1 Won with zero ARR",
        grain="closed_deal",
        severity="Critical",
        category="hygiene",
        soql=(
            "SELECT COUNT() FROM Opportunity WHERE "
            f"{OPP_CLOSED_Q1} "
            "AND StageName = '8 - Won' "
            "AND (APTS_Opportunity_ARR__c = 0 OR APTS_Opportunity_ARR__c = NULL)"
        ),
        sf_logic="Q1 Closed AND StageName='8 - Won' AND ARR IS NULL OR 0",
        action="Revenue tracking gap. Audit each one — may be renewal like-for-like.",
    ),
    # ─── DOMAIN · whitespace signals ─────────────────────────────────────
    HygieneCheck(
        key="competitor_installed_no_opp",
        label="Installed competitor product · no open opp (winback)",
        grain="whitespace",
        severity="Domain",
        category="winback",
        soql=(
            "SELECT COUNT() FROM Account WHERE "
            "Id IN (SELECT Account__c FROM Installed_Competitor_Product__c) "
            "AND Id NOT IN (SELECT AccountId FROM Opportunity "
            "WHERE IsClosed = false AND Type IN ('Land','Expand','Renewal'))"
        ),
        sf_logic="Account has Installed_Competitor_Product AND no open Opp",
        action="Route to BDR for outreach campaign.",
    ),
    HygieneCheck(
        key="whales_no_pipeline",
        label="Whale accounts (AuM >10B) with zero open pipeline",
        grain="whitespace",
        severity="Domain",
        category="winback",
        soql=(
            "SELECT COUNT() FROM Account WHERE AuM_m__c > 10000 "
            "AND Id NOT IN (SELECT AccountId FROM Opportunity WHERE IsClosed = false)"
        ),
        sf_logic="AuM_m__c > 10000 AND no open Opportunity",
        action="Critical whitespace. Executive assignment.",
    ),
    HygieneCheck(
        key="axioma_clients_all",
        label="Axioma clients (total — for whitespace sizing)",
        grain="whitespace",
        severity="Domain",
        category="crosssell",
        soql="SELECT COUNT() FROM Account WHERE Axioma_Client__c = true",
        sf_logic="Account.Axioma_Client__c = true",
        baseline=True,
    ),
]


# ────────────────────────── Execute a check ──────────────────────────────


def _auth():
    out = subprocess.check_output(
        ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"]
    )
    data = json.loads(out)["result"]
    return data["accessToken"], data["instanceUrl"]


def _session(token: str):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


def _run_check(session, instance: str, check: HygieneCheck) -> dict:
    start = time.monotonic()
    try:
        r = session.get(
            f"{instance}/services/data/v66.0/query",
            params={"q": check.soql},
            timeout=90,
        )
        if r.status_code >= 400:
            return {
                **asdict(check),
                "count": None,
                "error": f"{r.status_code}: {r.text[:200]}",
                "duration_ms": int((time.monotonic() - start) * 1000),
            }
        data = r.json()
        if isinstance(data, dict) and "totalSize" in data and data.get("done"):
            return {
                **asdict(check),
                "count": data["totalSize"],
                "duration_ms": int((time.monotonic() - start) * 1000),
            }
        return {
            **asdict(check),
            "count": None,
            "error": f"unexpected response: {str(data)[:200]}",
            "duration_ms": int((time.monotonic() - start) * 1000),
        }
    except (requests.HTTPError, requests.ConnectionError, requests.Timeout) as exc:
        return {
            **asdict(check),
            "count": None,
            "error": str(exc)[:200],
            "duration_ms": int((time.monotonic() - start) * 1000),
        }


# ────────────────────────── SF writeback (Phase 4) ──────────────────────


def _sf_writeback(
    session, instance: str, run_date: str, results: list[dict], ledger: dict
) -> dict:
    """Push this run's results to the Hygiene_* custom objects.

    Writes only aggregate Snapshot rows — per-record Flag writeback is a
    bigger lift (needs per-check sample queries to enumerate affected IDs)
    and is deferred to a later phase. Aggregates alone drive the KPI tiles,
    trend charts, and the hero-alert banner on the Sales Ops dashboard.

    Silently no-ops if the objects don't exist yet. Returns a dict with
    status + counts so the caller can log.
    """
    # Find prior counts from ledger for delta computation
    prior_counts = {}
    snaps = ledger.get("snapshots", [])
    prior_candidates = [s for s in snaps if s.get("run_date", "") < run_date]
    if prior_candidates:
        prior_counts = prior_candidates[-1].get("counts", {})

    # Hero alert: the Critical check with the highest current count (same
    # rule as the markdown summary — stays consistent across surfaces).
    crit_with_count = [
        r
        for r in results
        if r.get("severity") == "Critical" and isinstance(r.get("count"), int)
    ]
    hero_key = None
    if crit_with_count:
        hero_key = max(crit_with_count, key=lambda r: r["count"])["key"]

    snapshot_records = []
    for r in results:
        if r.get("count") is None:
            continue  # skip errored checks
        curr = r["count"]
        prior = prior_counts.get(r["key"])
        delta = (curr - prior) if isinstance(prior, int) else None
        snapshot_records.append(
            {
                "attributes": {"type": "Hygiene_Snapshot__c"},
                "Run_Date__c": run_date,
                "Metric_Key__c": r["key"],
                "Metric_Label__c": r["label"][:255],
                "Severity__c": r.get("severity"),
                "Category__c": r.get("category"),
                "Grain__c": r.get("grain"),
                "Count__c": curr,
                "Prior_Count__c": prior,
                "Delta__c": delta,
                "SF_Logic__c": r.get("sf_logic", "")[:2000],
                "Action__c": r.get("action", "")[:2000],
                "Is_Hero_Alert__c": (r["key"] == hero_key),
            }
        )

    # Salesforce composite/sobjects caps at 200 records per call
    status = {"sent": 0, "failed": 0, "errors": []}
    url = f"{instance}/services/data/v66.0/composite/sobjects"
    for i in range(0, len(snapshot_records), 200):
        batch = snapshot_records[i : i + 200]
        try:
            r = session.post(
                url,
                headers={"Content-Type": "application/json"},
                json={"allOrNone": False, "records": batch},
                timeout=60,
            )
        except (requests.ConnectionError, requests.Timeout) as exc:
            status["failed"] += len(batch)
            status["errors"].append(f"connection: {exc}")
            continue
        if r.status_code == 404 or (
            r.status_code == 400
            and "sObject type" in r.text
            and "not supported" in r.text
        ):
            # Objects not deployed yet — graceful no-op
            status["schema_missing"] = True
            status["errors"].append("Hygiene_Snapshot__c not deployed yet — skipped")
            return status
        if r.status_code >= 400:
            status["failed"] += len(batch)
            status["errors"].append(f"{r.status_code}: {r.text[:300]}")
            continue
        # Composite API can return 200 with per-record INVALID_TYPE errors
        # when the object doesn't exist. Detect that and bail once, cleanly.
        batch_results = r.json()
        all_invalid_type = batch_results and all(
            (not res.get("success"))
            and any(
                e.get("statusCode") == "INVALID_TYPE" for e in (res.get("errors") or [])
            )
            for res in batch_results
        )
        if all_invalid_type:
            status["schema_missing"] = True
            status["errors"] = [
                "Hygiene_Snapshot__c not deployed yet — deploy the objects first "
                "(see docs/2026-04-16-hygiene-objects-deploy-runbook.md)"
            ]
            # Reset failed count — "schema missing" isn't a failure condition
            # to alarm on, it's a pre-deployment no-op.
            status["failed"] = 0
            return status
        for result in batch_results:
            if result.get("success"):
                status["sent"] += 1
            else:
                status["failed"] += 1
                errs = result.get("errors", [])
                for e in errs:
                    status["errors"].append(
                        f"{e.get('statusCode', '')}: {e.get('message', '')[:200]}"
                    )
    return status


# ────────────────────────── Ledger + summary ────────────────────────────


def _update_ledger(run_date: str, results: list[dict]) -> dict:
    LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    if LEDGER_PATH.exists():
        try:
            ledger = json.loads(LEDGER_PATH.read_text())
        except (json.JSONDecodeError, ValueError):
            ledger = {"snapshots": []}
    else:
        ledger = {"snapshots": []}
    ledger["snapshots"] = [
        s for s in ledger.get("snapshots", []) if s.get("run_date") != run_date
    ]
    ledger["snapshots"].append(
        {
            "run_date": run_date,
            "counts": {r["key"]: r.get("count") for r in results},
        }
    )
    ledger["snapshots"].sort(key=lambda s: s["run_date"])
    LEDGER_PATH.write_text(json.dumps(ledger, indent=2) + "\n")
    return ledger


def _delta(curr: int | None, prior: int | None) -> str:
    if curr is None:
        return "—"
    if prior is None:
        return f"{curr}"
    d = curr - prior
    if d == 0:
        return f"{curr} (no change)"
    sign = "+" if d > 0 else ""
    # ⚠ when the count went UP (worse); ✓ when it went DOWN (better)
    marker = "⚠" if d > 0 else "✓"
    return f"{curr} ({sign}{d} {marker})"


def _write_summary(run_date: str, results: list[dict], ledger: dict) -> Path:
    out_dir = OUTPUT_ROOT / run_date
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "summary.md"

    prior = None
    snaps = ledger.get("snapshots", [])
    prior_candidates = [s for s in snaps if s.get("run_date", "") < run_date]
    if prior_candidates:
        prior = prior_candidates[-1]
    prior_counts = (prior or {}).get("counts", {}) if prior else {}

    # Bucket by severity
    by_sev: dict[str, list[dict]] = {"Critical": [], "Important": [], "Domain": []}
    baselines = []
    for r in results:
        if r.get("baseline"):
            baselines.append(r)
        elif r.get("severity") in by_sev:
            by_sev[r["severity"]].append(r)

    # Hero alert = the Critical check with the highest count
    crit_sorted = sorted(
        [r for r in by_sev["Critical"] if isinstance(r.get("count"), int)],
        key=lambda r: -r["count"],
    )
    hero = crit_sorted[0] if crit_sorted else None

    lines = [
        "---",
        "type: data-quality-audit",
        f"run_date: {run_date}",
        "tags: [data-quality, hygiene, governance, monthly]",
        "---",
        "",
        f"# Data Quality & Governance — {run_date}",
        "",
        (
            "Every check below is a *breach*, not just a metric — count named, "
            "severity tiered, SF logic self-documented, delta vs prior run_date "
            "in parentheses. ⚠ means getting worse; ✓ means improving. Concept "
            "extracted from Hooman's Sales Ops Commercial Health & Governance "
            "Dashboard."
        ),
        "",
    ]

    if hero:
        lines += [
            "> ## 🚨 Hero alert",
            f"> **{hero['label']}: {hero.get('count', '—')}** — {hero.get('action', '')}",
            f"> SF logic: `{hero['sf_logic']}`",
            "",
        ]

    # Baselines (context)
    lines += ["## Baseline scope", ""]
    lines += ["| Scope | Count |", "|---|---:|"]
    for r in baselines:
        lines.append(f"| {r['label']} | {r.get('count', '—')} |")
    lines.append("")

    for sev_label, emoji in [
        ("Critical", "🔴"),
        ("Important", "🟠"),
        ("Domain", "🏷️"),
    ]:
        rows = by_sev[sev_label]
        if not rows:
            continue
        lines += [
            f"## {emoji} {sev_label} ({len(rows)} checks)",
            "",
            "| Alert | Current | Grain | Category | SF logic |",
            "|---|---|---|---|---|",
        ]
        # Sort: higher count first, error rows last
        rows = sorted(
            rows,
            key=lambda r: (
                0 if isinstance(r.get("count"), int) else 1,
                -(r.get("count") or 0),
            ),
        )
        for r in rows:
            curr = r.get("count")
            prior_val = prior_counts.get(r["key"])
            delta_str = (
                _delta(curr, prior_val)
                if isinstance(curr, int)
                else r.get("error", "err")
            )
            lines.append(
                f"| {r['label']} | {delta_str} | {r['grain']} | {r['category']} | `{r['sf_logic']}` |"
            )
        lines.append("")

    lines += [
        "## Operational note",
        "",
        (
            "This report is generated by `scripts/audit_data_quality.py` as pipeline "
            "stage `1c_data_quality_audit`. Full per-check JSON is in "
            "`flags.json` alongside this file. The rolling MoM ledger is "
            "`output/data_quality/history.json`."
        ),
    ]
    path.write_text("\n".join(lines) + "\n")
    return path


# ────────────────────────── SharePoint Excel output ─────────────────────


def _write_sharepoint_excel(run_date: str, results: list[dict], ledger: dict) -> Path:
    """Write a formatted Excel workbook to output/sharepoint/ so it syncs to
    SharePoint alongside the other SD Monthly workbooks. Power BI or Excel
    Online can chart from it natively — no SF custom-object permissions needed.

    Structure mirrors Hooman's concept: severity-tiered alert table with
    SF logic column, MoM delta, hero alert highlight, and per-grain tabs.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    SHAREPOINT = ROOT / "output" / "sharepoint"
    SHAREPOINT.mkdir(parents=True, exist_ok=True)
    wb_path = SHAREPOINT / "Sales Ops Hygiene Dashboard.xlsx"

    NAVY = "083EA7"
    RED_BG = "F8D7DA"
    AMBER_BG = "FFF3CD"
    GREEN_BG = "D4EDDA"
    HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    BODY_FONT = Font(size=9, name="Calibri")
    BODY_BOLD = Font(bold=True, size=9, name="Calibri")
    TITLE_FONT = Font(bold=True, size=16, color=NAVY, name="Calibri")
    SUBTITLE_FONT = Font(italic=True, size=9, color="666666", name="Calibri")
    BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")

    # Prior counts for delta computation
    prior_counts = {}
    snaps = ledger.get("snapshots", [])
    prior_candidates = [s for s in snaps if s.get("run_date", "") < run_date]
    if prior_candidates:
        prior_counts = prior_candidates[-1].get("counts", {})

    # Hero alert
    crit_with_count = [
        r
        for r in results
        if r.get("severity") == "Critical" and isinstance(r.get("count"), int)
    ]
    hero_key = None
    if crit_with_count:
        hero_key = max(crit_with_count, key=lambda r: r["count"])["key"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Hygiene Dashboard"

    # Title block
    ws["A1"] = "Sales Ops — Data Quality & Governance"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Automated scan · {run_date} · {len(results)} checks · "
        f"Concept: Hooman Hashemi's Commercial Health & Governance Dashboard"
    )
    ws["A2"].font = SUBTITLE_FONT

    # Hero alert banner
    hero_row = [r for r in results if r.get("key") == hero_key]
    if hero_row:
        h = hero_row[0]
        ws["A4"] = (
            f"🚨 HERO ALERT: {h['label']}: {h.get('count', '—')} — {h.get('action', '')}"
        )
        ws["A4"].font = Font(bold=True, size=12, color="CC0000", name="Calibri")
        ws["A4"].fill = PatternFill(
            start_color=RED_BG, end_color=RED_BG, fill_type="solid"
        )
        ws.merge_cells("A4:H4")

    # Column headers
    row = 6
    headers = [
        "Severity",
        "Alert",
        "Current Count",
        "Prior Count",
        "Delta",
        "Trend",
        "Grain",
        "Category",
        "SF Logic (rule)",
        "Action",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    # Severity sort order
    sev_order = {"Critical": 0, "Important": 1, "Domain": 2, "baseline": 3}
    sorted_results = sorted(
        results,
        key=lambda r: (
            sev_order.get(r.get("severity", ""), 9),
            0 if isinstance(r.get("count"), int) else 1,
            -(r.get("count") or 0),
        ),
    )

    # Data rows
    row += 1
    first_data_row = row
    sev_fills = {
        "Critical": PatternFill(
            start_color=RED_BG, end_color=RED_BG, fill_type="solid"
        ),
        "Important": PatternFill(
            start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid"
        ),
        "Domain": PatternFill(
            start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"
        ),
        "baseline": PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        ),
    }

    for r in sorted_results:
        curr = r.get("count")
        prior = prior_counts.get(r["key"])
        delta = (
            (curr - prior) if isinstance(curr, int) and isinstance(prior, int) else None
        )
        trend = ""
        if delta is not None:
            if delta > 0:
                trend = "⚠ worse"
            elif delta < 0:
                trend = "✓ better"
            else:
                trend = "— same"

        vals = [
            r.get("severity", ""),
            r.get("label", ""),
            curr if isinstance(curr, int) else r.get("error", "err"),
            prior if isinstance(prior, int) else "",
            delta if delta is not None else "",
            trend,
            r.get("grain", ""),
            r.get("category", ""),
            r.get("sf_logic", ""),
            r.get("action", ""),
        ]
        sev = r.get("severity", "")
        fill = sev_fills.get(sev)

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (3, 4, 5):
                c.alignment = RIGHT
                c.number_format = "#,##0"
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_BOLD if r.get("key") == hero_key else BODY_FONT
            if fill and ci == 1:
                c.fill = fill
        row += 1
    last_data_row = row - 1

    # Conditional formatting on Count column
    if last_data_row >= first_data_row:
        ws.conditional_formatting.add(
            f"C{first_data_row}:C{last_data_row}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="D9534F",
                showValue=True,
            ),
        )

    # Column widths
    widths = [12, 40, 12, 12, 10, 10, 14, 16, 50, 40]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A6:{get_column_letter(len(headers))}{last_data_row}"

    # ── History tab (trend over time) ──
    if len(snaps) >= 2:
        ws2 = wb.create_sheet("Trend")
        ws2["A1"] = "MoM Trend by Metric"
        ws2["A1"].font = TITLE_FONT

        run_dates = [s["run_date"] for s in snaps]
        all_keys = []
        seen = set()
        for r in sorted_results:
            if r["key"] not in seen and not r.get("baseline"):
                all_keys.append(r["key"])
                seen.add(r["key"])

        # Header
        hrow = 3
        ws2.cell(row=hrow, column=1, value="Metric").font = HEADER_FONT
        ws2.cell(row=hrow, column=1).fill = HEADER_FILL
        ws2.cell(row=hrow, column=2, value="Severity").font = HEADER_FONT
        ws2.cell(row=hrow, column=2).fill = HEADER_FILL
        for di, d in enumerate(run_dates, 3):
            c = ws2.cell(row=hrow, column=di, value=d)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL

        for ri, key in enumerate(all_keys, hrow + 1):
            label_r = next((r for r in results if r["key"] == key), None)
            ws2.cell(
                row=ri, column=1, value=label_r["label"] if label_r else key
            ).font = BODY_FONT
            ws2.cell(
                row=ri, column=2, value=label_r.get("severity", "") if label_r else ""
            ).font = BODY_FONT
            for di, snap in enumerate(snaps, 3):
                val = snap.get("counts", {}).get(key)
                c = ws2.cell(row=ri, column=di, value=val if val is not None else "")
                c.alignment = RIGHT
                c.font = BODY_FONT

        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 12
        for di in range(3, 3 + len(run_dates)):
            ws2.column_dimensions[get_column_letter(di)].width = 14

    # ── Methodology tab ──
    ws3 = wb.create_sheet("Methodology")
    ws3["A1"] = "Methodology"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = (
        "Every check is a single SOQL COUNT() query against live Salesforce data. "
        "SF Logic column shows the exact rule. Severity tiers follow Hooman Hashemi's "
        "Commercial Health & Governance Dashboard concept: Critical = governance breach, "
        "Important = hygiene gap, Domain = whitespace signal."
    )
    ws3["A2"].font = SUBTITLE_FONT
    ws3.column_dimensions["A"].width = 120

    wb.save(str(wb_path))
    return wb_path


# ────────────────────────── main ────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument(
        "--filter-severity",
        default=None,
        choices=["Critical", "Important", "Domain"],
        help="Run only checks of this severity (default: all)",
    )
    parser.add_argument(
        "--write-to-sf",
        action="store_true",
        help=(
            "POST aggregate Snapshot rows to Hygiene_Snapshot__c via Composite "
            "REST. No-op if the object isn't deployed yet. Default: local only."
        ),
    )
    args = parser.parse_args()

    token, instance = _auth()
    session = _session(token)

    checks = CHECKS
    if args.filter_severity:
        checks = [c for c in checks if c.severity == args.filter_severity or c.baseline]

    out_dir = OUTPUT_ROOT / args.date
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"[data-quality] scanning {len(checks)} checks... ({args.date})")
    results = []
    total_ms = 0
    for c in checks:
        res = _run_check(session, instance, c)
        total_ms += res.get("duration_ms", 0)
        results.append(res)
        status = "ok" if "error" not in res else "err"
        tag = f"{c.severity[0]}" if c.severity != "baseline" else "·"
        print(f"  [{tag}] {status:3s} {c.key:40s} count={res.get('count')}")

    (out_dir / "flags.json").write_text(
        json.dumps({"results": results}, indent=2) + "\n"
    )
    ledger = _update_ledger(args.date, results)
    summary = _write_summary(args.date, results, ledger)

    # SharePoint Excel workbook — always written so it syncs alongside the
    # other SD Monthly workbooks. Power BI / Excel Online can chart from it.
    excel_path = _write_sharepoint_excel(args.date, results, ledger)
    print(f"  Excel: {excel_path.relative_to(ROOT)}")

    # Optional SF writeback — inert if Hygiene_Snapshot__c isn't deployed yet.
    sf_status = None
    if args.write_to_sf:
        print()
        print("[data-quality] attempting SF writeback...")
        sf_status = _sf_writeback(session, instance, args.date, results, ledger)
        (out_dir / "sf_writeback_status.json").write_text(
            json.dumps(sf_status, indent=2) + "\n"
        )

    # Recap
    by_sev_ct: dict[str, int] = {}
    for r in results:
        by_sev_ct[r["severity"]] = by_sev_ct.get(r["severity"], 0) + 1
    print()
    print(f"  Total queries: {len(results)} ({total_ms / 1000:.1f}s)")
    print(f"  Severity mix: {by_sev_ct}")
    print(f"  Wrote:   {out_dir.relative_to(ROOT)}")
    print(f"  Summary: {summary.relative_to(ROOT)}")
    print(
        f"  Ledger:  {LEDGER_PATH.relative_to(ROOT)}  ({len(ledger.get('snapshots', []))} runs recorded)"
    )
    if sf_status is not None:
        print(
            f"  SF writeback: sent={sf_status['sent']} failed={sf_status['failed']} "
            f"errors={len(sf_status['errors'])}"
        )
        for err in sf_status["errors"][:3]:
            print(f"    · {err[:150]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
