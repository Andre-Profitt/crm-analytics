#!/usr/bin/env python3
"""Extract live Salesforce data into a director Excel workbook.

Uses Alex P's methodology: Opportunity ARR (unweighted), FY26 only,
stages 1-6 for open pipeline, account/owner exclusions.

The Excel workbook is the single editable source of truth.
The SimCorp deck renderer reads from this Excel.

Usage:
    python3 scripts/extract_director_live.py --territory APAC
    python3 scripts/extract_director_live.py --territory "UK & Ireland"
    python3 scripts/extract_director_live.py --all
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_ROOT = REPO_ROOT / "output" / "director_live_workbooks"

# ── Territory config ──
# Loaded from config/sd_monthly_territories.json so that onboarding a new
# director is a config edit, not a code change. The keys below preserve the
# exact territory labels downstream builders expect (APAC, Central Europe, ...).
_TERRITORY_CONFIG_PATH = REPO_ROOT / "config" / "sd_monthly_territories.json"
try:
    _TERR_DATA = json.loads(_TERRITORY_CONFIG_PATH.read_text())
    TERRITORIES = _TERR_DATA.get("territories", {})
    if not TERRITORIES:
        raise ValueError(
            "config/sd_monthly_territories.json has no 'territories' block"
        )
except (OSError, ValueError, json.JSONDecodeError) as _exc:
    raise SystemExit(
        f"Unable to load territory config at {_TERRITORY_CONFIG_PATH}: {_exc}"
    )

# ── Alex P's filter methodology ──
ACCOUNT_EXCLUDE = "AND (NOT Account.Name LIKE '%simcorp%') AND (NOT Account.Name LIKE '%test%') AND (NOT Account.Name LIKE '%delete%')"
OWNER_EXCLUDE = (
    "AND (NOT Owner.Name LIKE '%Sabiniewicz%') AND (NOT Owner.Name LIKE '%Profit%')"
)
TYPE_FILTER = "AND Type IN ('Land', 'Expand', 'Renewal')"
FY26_CLOSE = "AND CloseDate >= 2026-01-01 AND CloseDate <= 2026-12-31"
OPEN_STAGES = "AND StageName IN ('1 - Prospecting', '2 - Discovery', '3 - Engagement', '4 - Shortlisted', '5 - Preferred', '6 - Contracting')"
CLOSED_STAGES = "AND IsClosed = true"

# ── Shared columns (matches Alex's report) ──
# ARR fields wrapped in convertCurrency() so multi-currency books (APAC USD,
# NA USD, EMEA CHF/SEK) report in the corporate currency (EUR) rather than
# the deal's native currency. Alias back to the original field name so the
# response-parsing code does not change.
PIPELINE_FIELDS = """
    Id, Account.Name, Name, Owner.Name, StageName, CloseDate,
    convertCurrency(APTS_Opportunity_ARR__c) APTS_Opportunity_ARR__c,
    convertCurrency(APTS_Forecast_ARR__c) APTS_Forecast_ARR__c,
    ForecastCategoryName, Probability, PushCount, Type,
    Lead_Scope__c, Account.Industry, Account.Tier_Calculation__c,
    Account_Unit_Group__c, Sales_Region__c,
    CreatedDate, LastActivityDate, NextStep, LastModifiedDate,
    Stage_20_Approval__c, Stage_20_Approval_Date__c,
    Submit_for_Stage_20_Review__c, Submit_for_Stage_20_Review_Date__c,
    Approval_Status__c, Lost_to_Competitor__c
""".strip()

# ── Styling ──
HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=9)
EUR_FMT = "#,##0"


def get_auth() -> tuple[str, str]:
    result = subprocess.run(
        ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
        capture_output=True,
        text=True,
        check=True,
    )
    data = json.loads(result.stdout)["result"]
    return data["accessToken"], data["instanceUrl"]


# ── SF HTTP: retry + backoff + telemetry ──
# Telemetry bucket shared across threads. Each query appends
# {label, rows, duration_ms, attempts, status}.
QUERY_TELEMETRY: list[dict] = []
_TELEMETRY_LOCK = __import__("threading").Lock()


def _sf_get_with_retry(session, url, params=None, timeout=120, attempts=3, label=""):
    """GET with exponential backoff on 429/5xx/connection errors.

    Salesforce returns 429 (rate-limit) or 503 (shed load) under sustained
    traffic. Transient network errors also happen. A 3-attempt retry with
    1s / 2s / 4s backoff covers all realistic cases without masking real
    server failures (which persist past 3 attempts).
    """
    import time

    last_exc: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            resp = session.get(url, params=params, timeout=timeout)
            if resp.status_code in (429, 500, 502, 503, 504) and attempt < attempts:
                time.sleep(2 ** (attempt - 1))
                continue
            resp.raise_for_status()
            return resp, attempt
        except (requests.ConnectionError, requests.Timeout) as exc:
            last_exc = exc
            if attempt < attempts:
                time.sleep(2 ** (attempt - 1))
                continue
            raise
    if last_exc:
        raise last_exc
    raise RuntimeError(f"SF GET exhausted retries: {label or url}")


def run_soql(session, instance_url: str, query: str, label: str = "") -> list[dict]:
    """SOQL with pagination, retry, and telemetry.

    Caller passes a prepared requests.Session (auth header already set).
    Records query timing + row count into QUERY_TELEMETRY for manifest export.
    """
    import time

    start = time.monotonic()
    records = []
    url = f"{instance_url}/services/data/v66.0/query"
    params: dict | None = {"q": query}
    total_attempts = 0
    while True:
        resp, attempts = _sf_get_with_retry(
            session, url, params=params, label=label or "soql"
        )
        total_attempts += attempts
        d = resp.json()
        records.extend(d.get("records", []))
        next_url = d.get("nextRecordsUrl")
        if not next_url:
            break
        url = f"{instance_url}{next_url}"
        params = None
    duration_ms = int((time.monotonic() - start) * 1000)
    with _TELEMETRY_LOCK:
        QUERY_TELEMETRY.append(
            {
                "label": label or "soql",
                "rows": len(records),
                "duration_ms": duration_ms,
                "attempts": total_attempts,
                "status": "ok",
            }
        )
    return records


def fetch_pi(session, instance_url: str, lv_id: str, label: str = "") -> list[dict]:
    """Pipeline Inspection list-view records with retry + telemetry."""
    import time

    start = time.monotonic()
    url = f"{instance_url}/services/data/v66.0/ui-api/list-records/{lv_id}?pageSize=200"
    records = []
    total_attempts = 0
    while url and len(records) < 2000:
        resp, attempts = _sf_get_with_retry(
            session, url, timeout=60, label=label or f"pi:{lv_id}"
        )
        total_attempts += attempts
        if resp.status_code != 200:
            break
        d = resp.json()
        records.extend(d.get("records", []))
        next_url = d.get("nextPageUrl")
        url = f"{instance_url}{next_url}" if next_url else None
    duration_ms = int((time.monotonic() - start) * 1000)
    with _TELEMETRY_LOCK:
        QUERY_TELEMETRY.append(
            {
                "label": label or f"pi:{lv_id}",
                "rows": len(records),
                "duration_ms": duration_ms,
                "attempts": total_attempts,
                "status": "ok",
            }
        )
    return records


def build_session(token: str) -> requests.Session:
    """Build a requests.Session with the SF bearer token set.

    Sessions are thread-safe for read-only use (our pipeline), so a single
    session can drive all 9 concurrent territory extractions.
    """
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


def _val(record, field):
    """Safely get a field value, handling relationship fields."""
    parts = field.split(".")
    obj = record
    for p in parts:
        if obj is None:
            return ""
        if isinstance(obj, dict):
            obj = obj.get(p)
        else:
            return ""
    return obj if obj is not None else ""


def _add_sheet(wb, name, headers, rows, eur_cols=None):
    ws = wb.create_sheet(title=name[:31])
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            if eur_cols and ci in eur_cols and isinstance(val, (int, float)):
                cell.number_format = EUR_FMT
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(headers[ci - 1])),
            *(len(str(r[ci - 1])) for r in rows[:50]) if rows else [0],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
    if rows:
        end_col = get_column_letter(len(headers))
        table_name = (
            name.replace(" ", "_")
            .replace("-", "_")
            .replace("&", "And")
            .replace("/", "")[:30]
        )
        try:
            table = Table(displayName=table_name, ref=f"A1:{end_col}{len(rows) + 1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
            )
            ws.add_table(table)
        except Exception:
            pass  # skip table if name collision
    ws.freeze_panes = "A2"


def extract_territory(
    territory: str,
    snapshot_date: str,
    output_path: Path,
    session=None,
    instance_url: str | None = None,
):
    """Extract one territory to a director workbook.

    Accepts a pre-built session + instance_url so the caller can share a
    single auth across concurrent extractions. Falls back to computing auth
    locally if not provided (backward-compatible for single-territory use).
    """
    config = TERRITORIES[territory]
    director = config["director"]
    where = config["soql_where"]
    pi_lv = config["pi_list_view_id"]

    print(f"\n{'=' * 60}")
    print(f"  {director} ({territory})")
    print(f"{'=' * 60}")

    if session is None or instance_url is None:
        token, instance_url = get_auth()
        session = build_session(token)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Open Pipeline (stages 1-6, FY26) ──
    print("  Pipeline...", end=" ", flush=True)
    q = f"SELECT {PIPELINE_FIELDS} FROM Opportunity WHERE {where} AND IsClosed = false {FY26_CLOSE} {OPEN_STAGES} {TYPE_FILTER} {ACCOUNT_EXCLUDE} {OWNER_EXCLUDE} ORDER BY APTS_Opportunity_ARR__c DESC NULLS LAST"
    pipeline = run_soql(session, instance_url, q, label=f"{territory}:pipeline_open")
    print(f"{len(pipeline)} deals")

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "Close Date",
        "ARR Unweighted (EUR)",
        "ARR Weighted (EUR)",
        "Probability %",
        "Push Count",
        "Type",
        "Lead Scope",
        "Industry",
        "Tier",
        "Sales Region",
        "Created",
        "Last Activity",
        "Next Step",
        "Last Modified",
        "Approved",
        "Approval Date",
        "Competitor",
    ]
    rows = []
    for r in pipeline:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("ForecastCategoryName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                r.get("APTS_Forecast_ARR__c") or 0,
                r.get("Probability") or 0,
                r.get("PushCount") or 0,
                r.get("Type", ""),
                r.get("Lead_Scope__c", ""),
                _val(r, "Account.Industry"),
                _val(r, "Account.Tier_Calculation__c"),
                r.get("Sales_Region__c", ""),
                (r.get("CreatedDate") or "")[:10],
                r.get("LastActivityDate", ""),
                r.get("NextStep", ""),
                (r.get("LastModifiedDate") or "")[:10],
                "Yes" if r.get("Stage_20_Approval__c") else "No",
                r.get("Stage_20_Approval_Date__c", ""),
                # Known competitor on an OPEN deal — a risk signal worth
                # flagging. Same field we already pull on closed deals.
                r.get("Lost_to_Competitor__c", "") or "",
            ]
        )
    _add_sheet(wb, "Pipeline Open FY26", headers, rows, eur_cols=[7, 8])

    # ── 2. Won/Lost (FY26 closed) ──
    print("  Won/Lost...", end=" ", flush=True)
    # Reason_Won_Lost__c is the only field we need that isn't already in
    # PIPELINE_FIELDS; Lost_to_Competitor__c is now pulled on both open and
    # closed deals, so it's already included.
    q = f"SELECT {PIPELINE_FIELDS}, Reason_Won_Lost__c FROM Opportunity WHERE {where} AND IsClosed = true {FY26_CLOSE} {CLOSED_STAGES} {TYPE_FILTER} {ACCOUNT_EXCLUDE} {OWNER_EXCLUDE} ORDER BY CloseDate DESC"
    won_lost = run_soql(session, instance_url, q, label=f"{territory}:won_lost")
    print(f"{len(won_lost)} deals")

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR Unweighted (EUR)",
        "Type",
        "Reason",
        "Lost To Competitor",
        "Industry",
        "Sales Region",
        "Created",
    ]
    rows = []
    for r in won_lost:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                r.get("Type", ""),
                r.get("Reason_Won_Lost__c", ""),
                r.get("Lost_to_Competitor__c", ""),
                _val(r, "Account.Industry"),
                r.get("Sales_Region__c", ""),
                (r.get("CreatedDate") or "")[:10],
            ]
        )
    _add_sheet(wb, "Won Lost FY26", headers, rows, eur_cols=[6])

    # ── 3. Commercial Approval (open Land, FY26) ──
    # Approvals reuse the same open pipeline records, filtered to Land type.
    # All approval fields are now in PIPELINE_FIELDS, so we skip the separate
    # SOQL call entirely — saves one SF round-trip per director (9/run).
    print("  Approvals... (filtering from pipeline)", end=" ", flush=True)
    approvals = [r for r in pipeline if str(r.get("Type") or "").strip() == "Land"]
    approved_2026, approved_prior, pending, missing = [], [], [], []
    for r in approvals:
        approval_status = str(r.get("Approval_Status__c") or "").strip()
        if r.get("Stage_20_Approval__c"):
            # Fully approved — split by year
            if str(r.get("Stage_20_Approval_Date__c", ""))[:4] == "2026":
                approved_2026.append(r)
            else:
                approved_prior.append(r)
        elif (
            r.get("Submit_for_Stage_20_Review__c")
            or approval_status == "Needs Approval"
        ):
            # Submitted for review OR flagged as needing approval but not yet approved
            pending.append(r)
        elif (
            r.get("StageName", "") >= "3" and approval_status != "No Approval Necessary"
        ):
            # Stage 3+ with no approval AND not exempt → truly missing
            missing.append(r)
        # Deals with "No Approval Necessary" at Stage 3+ are exempt — not counted as missing
    print(
        f"{len(approvals)} land ({len(approved_2026)} approved 2026, {len(approved_prior)} prior, {len(pending)} pending, {len(missing)} missing stage 3+)"
    )

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR Unweighted (EUR)",
        "Status",
        "Approval Date",
        "Next Step",
        "Lead Scope",
    ]
    rows = []
    for r in approved_2026:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Approved 2026",
                r.get("Stage_20_Approval_Date__c", ""),
                r.get("NextStep", ""),
                r.get("Lead_Scope__c", ""),
            ]
        )
    for r in approved_prior:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Approved (prior year)",
                r.get("Stage_20_Approval_Date__c", ""),
                r.get("NextStep", ""),
                r.get("Lead_Scope__c", ""),
            ]
        )
    for r in pending:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Pending Approval",
                r.get("Submit_for_Stage_20_Review_Date__c", ""),
                r.get("NextStep", ""),
                r.get("Lead_Scope__c", ""),
            ]
        )
    for r in missing:
        rows.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("CloseDate", ""),
                r.get("APTS_Opportunity_ARR__c") or 0,
                "Missing (Stage 3+)",
                "",
                r.get("NextStep", ""),
                r.get("Lead_Scope__c", ""),
            ]
        )
    _add_sheet(wb, "Commercial Approval", headers, rows, eur_cols=[6])

    # ── 4. Renewals (open, FY26, sorted by close date) ──
    print("  Renewals...", end=" ", flush=True)
    q = (
        f"SELECT Account.Name, Name, Owner.Name, StageName, CloseDate, "
        f"convertCurrency(Amount) Amount, Probability, NextStep FROM Opportunity "
        f"WHERE {where} AND IsClosed = false AND Type = 'Renewal' "
        f"{FY26_CLOSE} {ACCOUNT_EXCLUDE} {OWNER_EXCLUDE} ORDER BY CloseDate ASC"
    )
    renewals = run_soql(session, instance_url, q, label=f"{territory}:renewals")
    print(f"{len(renewals)} renewals")

    headers = [
        "Close Date",
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ACV Unweighted (EUR)",
        "Probability %",
        "Comments",
    ]
    rows = []
    for r in renewals:
        rows.append(
            [
                r.get("CloseDate", ""),
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                r.get("StageName", ""),
                r.get("Amount") or 0,
                r.get("Probability") or 0,
                "",  # Comments column for director
            ]
        )
    _add_sheet(wb, "Renewals FY26", headers, rows, eur_cols=[6])

    # ── 5. Pipeline Inspection (open FY26 from PI list view) ──
    print("  PI view...", end=" ", flush=True)
    pi_raw = fetch_pi(session, instance_url, pi_lv, label=f"{territory}:pi")
    headers = [
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "ARR Weighted (EUR)",
        "Close Date",
        "Push Count",
        "Score",
        "Priority",
    ]
    rows = []
    for rec in pi_raw:
        f = rec.get("fields", {})
        name = str(f.get("Name", {}).get("value", ""))
        close = str(f.get("CloseDate", {}).get("value", ""))
        is_closed = f.get("IsClosed", {}).get("value", False)
        fc = str(f.get("ForecastCategoryName", {}).get("value", ""))
        # Filter: FY26, not internal, not closed, not omitted
        if is_closed or fc in ("Omitted", "Closed"):
            continue
        if close and close[:4] > "2026":
            continue
        if any(p.lower() in name.lower() for p in ("simcorp", "test account")):
            continue
        owner_obj = f.get("Owner", {}).get("value")
        owner = (
            owner_obj.get("fields", {}).get("Name", {}).get("value", "")
            if isinstance(owner_obj, dict)
            else ""
        )
        score_obj = f.get("OpportunityScore", {}).get("value")
        score = (
            score_obj.get("fields", {}).get("Score", {}).get("value")
            if isinstance(score_obj, dict)
            else None
        )
        rows.append(
            [
                name,
                owner,
                str(f.get("StageName", {}).get("value", "")),
                fc,
                f.get("APTS_Forecast_ARR__c", {}).get("value") or 0,
                close,
                f.get("PushCount", {}).get("value") or 0,
                score,
                "Yes" if f.get("IsPriorityRecord", {}).get("value") else "",
            ]
        )
    rows.sort(key=lambda x: -(x[4] or 0))
    print(f"{len(rows)} open FY26 deals")
    _add_sheet(wb, "Pipeline Inspection", headers, rows, eur_cols=[5])

    # ── 5b. Activity volume per open deal ──
    # Group Task + Event count per Opportunity for the 30/60/90-day windows.
    # Surfaces "silent" deals that LastActivityDate alone doesn't reveal
    # (e.g. deal with one email 65 days ago has LastActivityDate = 65d but
    # only shows up as "stale" in Deal Risk Scoring — activity count adds
    # whether there's ANY recent touch).
    print("  Activity volume...", end=" ", flush=True)
    open_opp_ids = [r.get("Id") for r in pipeline if r.get("Id")]
    activity_rows = []
    if open_opp_ids:
        # Salesforce caps IN () lists at ~10k ids; our director books are <100,
        # so we batch in 200s defensively.
        from datetime import timedelta as _td

        today = datetime.now().date()
        # Single 90-day window for now. If we add 30/60-day breakdowns later,
        # compute them the same way and COUNT_DISTINCT by WhatId over the
        # appropriate ActivityDate range.
        d90 = (today - _td(days=90)).isoformat()

        def _batch(seq, n):
            for i in range(0, len(seq), n):
                yield seq[i : i + n]

        # Per-opp activity aggregate. Separate Task and Event streams so we
        # can render a breakdown (calls/emails in Task, meetings in Event).
        # SOQL note: ActivityDate does not support MAX() — we rely on the
        # Opportunity's own LastActivityDate (already in PIPELINE_FIELDS) for
        # "last touch" timing, and use these queries purely for count.
        act_by_opp: dict[str, dict] = {}
        for batch in _batch(open_opp_ids, 200):
            ids_sql = ", ".join(f"'{i}'" for i in batch)
            q_task = (
                "SELECT WhatId, COUNT(Id) n "
                f"FROM Task WHERE WhatId IN ({ids_sql}) "
                f"AND ActivityDate >= {d90} GROUP BY WhatId"
            )
            q_event = (
                "SELECT WhatId, COUNT(Id) n "
                f"FROM Event WHERE WhatId IN ({ids_sql}) "
                f"AND ActivityDate >= {d90} GROUP BY WhatId"
            )
            try:
                tasks = run_soql(
                    session, instance_url, q_task, label=f"{territory}:tasks_90d"
                )
            except requests.HTTPError:
                tasks = []
            try:
                events = run_soql(
                    session, instance_url, q_event, label=f"{territory}:events_90d"
                )
            except requests.HTTPError:
                events = []
            for rec in tasks:
                wid = rec.get("WhatId")
                if not wid:
                    continue
                a = act_by_opp.setdefault(
                    wid, {"tasks": 0, "events": 0, "last_date": None}
                )
                a["tasks"] = int(rec.get("n") or 0)
                d = rec.get("last_date")
                if d and (a["last_date"] is None or d > a["last_date"]):
                    a["last_date"] = d
            for rec in events:
                wid = rec.get("WhatId")
                if not wid:
                    continue
                a = act_by_opp.setdefault(
                    wid, {"tasks": 0, "events": 0, "last_date": None}
                )
                a["events"] = int(rec.get("n") or 0)
                d = rec.get("last_date")
                if d and (a["last_date"] is None or d > a["last_date"]):
                    a["last_date"] = d

        # Compose activity rows: counts from the aggregate dict, last-touch
        # date from the Opportunity's own LastActivityDate (which SOQL already
        # computes across Tasks+Events+EmailMessage so is authoritative).
        for rec in pipeline:
            oid = rec.get("Id")
            acct = _val(rec, "Account.Name")
            opp = rec.get("Name", "")
            owner = _val(rec, "Owner.Name")
            agg = act_by_opp.get(str(oid or ""), {"tasks": 0, "events": 0})
            total_90 = int(agg.get("tasks", 0)) + int(agg.get("events", 0))
            last_activity = rec.get("LastActivityDate") or ""
            activity_rows.append(
                [
                    acct,
                    opp,
                    owner,
                    agg.get("tasks", 0),
                    agg.get("events", 0),
                    total_90,
                    last_activity,
                    "No touch 90d" if total_90 == 0 else "",
                ]
            )
    activity_rows.sort(key=lambda x: (x[5], x[6] or ""))  # silent first
    print(f"{sum(1 for r in activity_rows if r[5] == 0)} silent deals (90d)")
    _add_sheet(
        wb,
        "Activity Volume",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Tasks 90d",
            "Events 90d",
            "Total Touches 90d",
            "Last Activity",
            "Flag",
        ],
        activity_rows,
    )

    # ── 5c. Per-deal commit breakdown ──
    # ForecastingItem rolls up per (owner, period, category) — it has no
    # OpportunityId column, so we can't join per-deal via that table. The
    # Opportunity's own ForecastCategoryName + APTS_Forecast_ARR__c
    # (already in PIPELINE_FIELDS) IS the per-deal commit view, so we
    # project that here as the "Commit Items" sheet — same shape as a
    # ForecastingItem breakdown but sourced authoritatively from the deal.
    print("  Commit items... ", end=" ", flush=True)
    fi_rows_out = []
    for r in pipeline:
        cat = str(r.get("ForecastCategoryName") or "").strip()
        # Skip Omitted + blanks; those aren't part of any commit.
        if not cat or cat == "Omitted":
            continue
        close = str(r.get("CloseDate") or "")[:10]
        # Derive quarter from close date for period grouping.
        qtr = ""
        if close.startswith("2026") and len(close) >= 7:
            try:
                m = int(close[5:7])
                qtr = f"Q{(m - 1) // 3 + 1} 2026"
            except ValueError:
                qtr = ""
        fi_rows_out.append(
            [
                _val(r, "Account.Name"),
                r.get("Name", ""),
                _val(r, "Owner.Name"),
                cat,
                r.get("APTS_Forecast_ARR__c") or 0,
                r.get("APTS_Opportunity_ARR__c") or 0,
                close,
                qtr,
                r.get("StageName", ""),
            ]
        )
    fi_rows_out.sort(key=lambda x: -(x[4] or 0))
    print(f"{len(fi_rows_out)} commit rows")
    _add_sheet(
        wb,
        "Commit Items",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Forecast Category",
            "Forecast ARR Wtd (EUR)",
            "ARR Unwtd (EUR)",
            "Close Date",
            "Period",
            "Stage",
        ],
        fi_rows_out,
        eur_cols=[5, 6],
    )

    # ── 6. Q1 Movement (field history analysis) ──
    print("  Q1 movement...", end=" ", flush=True)
    # Convert the Opportunity-root `where` clause into one that works from
    # OpportunityFieldHistory: every bare SimCorp field needs an "Opportunity." prefix.
    # Do a targeted replace on the known SimCorp relationships so it works for any territory.
    where_for_history = where
    for fld in (
        "Account_Unit_Group__c",
        "Sales_Region__c",
        "Account.Unit__c",
        "Account.Region__c",
        "Account.Industry",
        "Lead_Scope__c",
    ):
        where_for_history = where_for_history.replace(fld, f"Opportunity.{fld}")
    # Avoid double-prefixing (e.g. if `where` already had "Opportunity." before any of them)
    where_for_history = where_for_history.replace(
        "Opportunity.Opportunity.", "Opportunity."
    )
    # Extended capture: pull CloseDate + StageName + ForecastCategoryName
    # history in one query so Q1 slip classification, stage-at-loss, and
    # commit-drift analyses all share one source-of-truth event stream.
    close_history = run_soql(
        session,
        instance_url,
        f"""
        SELECT OpportunityId, Opportunity.Name, Opportunity.Account.Name,
               Opportunity.Owner.Name, Opportunity.StageName, Opportunity.CloseDate,
               convertCurrency(Opportunity.APTS_Opportunity_ARR__c) APTS_Opportunity_ARR__c,
               Opportunity.IsClosed, Opportunity.IsWon,
               Field, OldValue, NewValue, CreatedDate
        FROM OpportunityFieldHistory
        WHERE Field IN ('CloseDate', 'StageName', 'ForecastCategoryName')
          AND {where_for_history}
          AND (NOT Opportunity.Account.Name LIKE '%simcorp%')
          AND (NOT Opportunity.Account.Name LIKE '%test%')
        ORDER BY CreatedDate ASC
    """,
        label=f"{territory}:field_history",
    )
    # Filter to CloseDate events for the existing Q1 slip / post-Q1 push
    # classification below. StageName and ForecastCategoryName events are
    # written separately so downstream analytics can consume them.
    stage_history_events = [r for r in close_history if r.get("Field") == "StageName"]
    fcat_history_events = [
        r for r in close_history if r.get("Field") == "ForecastCategoryName"
    ]
    close_history = [r for r in close_history if r.get("Field") == "CloseDate"]

    # Classify: Q1 slipped, post-Q1 pushed
    q1_slipped_rows = []
    post_q1_pushed_rows = []
    seen_slip = set()
    seen_push = set()
    for r in close_history:
        opp = r.get("Opportunity") or {}
        oid = r.get("OpportunityId", "")
        name = opp.get("Name", "")
        old_val = str(r.get("OldValue", ""))
        new_val = str(r.get("NewValue", ""))
        change_date = str(r.get("CreatedDate", ""))[:10]
        arr = opp.get("APTS_Opportunity_ARR__c") or 0
        account = (opp.get("Account") or {}).get("Name", "")
        owner = (opp.get("Owner") or {}).get("Name", "")
        stage = opp.get("StageName", "")
        is_closed = opp.get("IsClosed", False)

        # Q1 slipped: close date WAS in Q1 2026, got pushed out of Q1,
        # AND the push event itself occurred in 2026. Excludes legacy deals
        # that had a Q1 2026 close date set in 2025 and got pushed before
        # FY26 even started; the review is about what slipped THIS year.
        if (
            old_val >= "2026-01-01"
            and old_val <= "2026-03-31"
            and new_val > "2026-03-31"
            and change_date >= "2026-01-01"
            and oid not in seen_slip
        ):
            seen_slip.add(oid)
            q1_slipped_rows.append(
                [
                    account,
                    name,
                    owner,
                    stage,
                    "Q1 Slipped",
                    old_val,
                    new_val,
                    change_date,
                    arr,
                ]
            )

        # Post-Q1 pushed: any push after March 31
        if change_date >= "2026-04-01" and not is_closed and oid not in seen_push:
            seen_push.add(oid)
            post_q1_pushed_rows.append(
                [
                    account,
                    name,
                    owner,
                    stage,
                    "Post-Q1 Push",
                    old_val,
                    new_val,
                    change_date,
                    arr,
                ]
            )

    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Movement",
        "Old Close",
        "New Close",
        "Changed On",
        "ARR Unweighted (EUR)",
    ]
    all_movement = q1_slipped_rows + post_q1_pushed_rows
    all_movement.sort(key=lambda x: -(x[8] or 0))
    print(f"{len(q1_slipped_rows)} Q1 slips, {len(post_q1_pushed_rows)} post-Q1 pushes")
    _add_sheet(wb, "Q1 Movement", headers, all_movement, eur_cols=[9])

    # ── 6b. Stage History + Forecast Category History (from extended query) ──
    # Raw event streams for downstream analytics (stage-at-loss, forecast
    # drift, commit accuracy). One row per field-change event.
    def _history_row(r):
        opp = r.get("Opportunity") or {}
        return [
            (opp.get("Account") or {}).get("Name", ""),
            opp.get("Name", ""),
            (opp.get("Owner") or {}).get("Name", ""),
            opp.get("StageName", ""),
            str(r.get("OldValue") or ""),
            str(r.get("NewValue") or ""),
            str(r.get("CreatedDate", ""))[:10],
            opp.get("APTS_Opportunity_ARR__c") or 0,
        ]

    stage_headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage (live)",
        "From Stage",
        "To Stage",
        "Changed On",
        "ARR Unweighted (EUR)",
    ]
    stage_rows = [_history_row(r) for r in stage_history_events]
    stage_rows.sort(key=lambda x: x[6], reverse=True)
    print(f"  Stage history: {len(stage_rows)} events")
    _add_sheet(wb, "Stage History", stage_headers, stage_rows, eur_cols=[8])

    fcat_headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage (live)",
        "From Category",
        "To Category",
        "Changed On",
        "ARR Unweighted (EUR)",
    ]
    fcat_rows = [_history_row(r) for r in fcat_history_events]
    fcat_rows.sort(key=lambda x: x[6], reverse=True)
    print(f"  Forecast category history: {len(fcat_rows)} events")
    _add_sheet(wb, "Forecast Category History", fcat_headers, fcat_rows, eur_cols=[8])

    # ── Summary sheet ──
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = f"{director} ({territory})"
    ws["A1"].font = Font(bold=True, size=14, color="083EA7")
    ws["A2"] = "Reporting period: FY2026 (Q1-Q4)"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A3"] = (
        f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')} — live from Salesforce"
    )
    ws["A3"].font = Font(size=10, color="666666")
    ws["A4"] = (
        "Methodology: Alex P — ARR Unweighted = APTS_Opportunity_ARR__c (full deal value); "
        "ARR Weighted = APTS_Forecast_ARR__c (probability-weighted). "
        "Excl simcorp/test/delete accounts, excl Sabiniewicz/Profit owners."
    )
    ws["A4"].font = Font(size=8, italic=True, color="999999")

    # KPI summary
    total_pipeline_arr = sum(r.get("APTS_Opportunity_ARR__c") or 0 for r in pipeline)
    won = [r for r in won_lost if "Won" in (r.get("StageName") or "")]
    lost = [
        r
        for r in won_lost
        if "Lost" in (r.get("StageName") or "")
        or "Opt Out" in (r.get("StageName") or "")
    ]
    won_arr = sum(r.get("APTS_Opportunity_ARR__c") or 0 for r in won)
    lost_arr = sum(r.get("APTS_Opportunity_ARR__c") or 0 for r in lost)
    renewal_acv = sum(r.get("Amount") or 0 for r in renewals)

    ws["A6"] = "KPI"
    ws["B6"] = "Value"
    ws["A6"].font = HEADER_FONT
    ws["A6"].fill = HEADER_FILL
    ws["B6"].font = HEADER_FONT
    ws["B6"].fill = HEADER_FILL
    kpis = [
        ("Open Pipeline Unweighted (stages 1-6)", f"EUR {total_pipeline_arr:,.0f}"),
        ("Open Deal Count", str(len(pipeline))),
        ("Won ARR Unweighted FY26", f"EUR {won_arr:,.0f}"),
        ("Won Deal Count", str(len(won))),
        ("Lost ARR Unweighted FY26", f"EUR {lost_arr:,.0f}"),
        ("Lost Deal Count", str(len(lost))),
        ("Approved 2026 (Land)", str(len(approved_2026))),
        ("Approved Prior Year", str(len(approved_prior))),
        ("Pending Approval", str(len(pending))),
        ("Missing Approval (Stage 3+)", str(len(missing))),
        ("Open Renewal ACV Unweighted", f"EUR {renewal_acv:,.0f}"),
        ("Open Renewals", str(len(renewals))),
        ("PI Open Deals (FY26)", str(len(rows))),
    ]
    for i, (label, val) in enumerate(kpis, 7):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val

    # Sheet index
    ws[f"A{len(kpis) + 8}"] = "Sheet"
    ws[f"B{len(kpis) + 8}"] = "Records"
    ws[f"C{len(kpis) + 8}"] = "Source"
    for col in ("A", "B", "C"):
        ws[f"{col}{len(kpis) + 8}"].font = HEADER_FONT
        ws[f"{col}{len(kpis) + 8}"].fill = HEADER_FILL
    sheets = [
        ("Pipeline Open FY26", len(pipeline), "SOQL — open, stages 1-6, FY26"),
        ("Won Lost FY26", len(won_lost), "SOQL — closed, stages 0/7/8, FY26"),
        ("Commercial Approval", len(approvals), "SOQL — open Land, FY26"),
        ("Renewals FY26", len(renewals), "SOQL — open Renewal, FY26"),
        ("Pipeline Inspection", len(rows), "PI list view — open, FY26"),
    ]
    for i, (sname, count, source) in enumerate(sheets, len(kpis) + 9):
        ws[f"A{i}"] = sname
        ws[f"B{i}"] = count
        ws[f"C{i}"] = source

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40

    # ── Save ──
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n  Saved: {output_path}")
    print(f"  Pipeline: {len(pipeline)} deals, EUR {total_pipeline_arr:,.0f}")
    print(
        f"  Won/Lost: {len(won)} won (EUR {won_arr:,.0f}) / {len(lost)} lost (EUR {lost_arr:,.0f})"
    )
    print(
        f"  Approvals: {len(approved_2026)} approved 2026 / {len(approved_prior)} prior / {len(pending)} pending / {len(missing)} missing"
    )
    print(f"  Renewals: {len(renewals)} (EUR {renewal_acv:,.0f} ACV)")
    print(f"  PI: {len(rows)} open FY26 deals")

    return output_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--territory", default=None, help="Territory name (e.g. APAC)")
    parser.add_argument("--all", action="store_true", help="Extract all territories")
    parser.add_argument("--snapshot-date", default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    args = parser.parse_args()

    if args.all:
        territories = list(TERRITORIES.keys())
    elif args.territory:
        if args.territory not in TERRITORIES:
            print(f"Unknown territory: {args.territory}", file=sys.stderr)
            print(f"Available: {', '.join(TERRITORIES.keys())}", file=sys.stderr)
            sys.exit(1)
        territories = [args.territory]
    else:
        print("Specify --territory or --all", file=sys.stderr)
        sys.exit(1)

    import re

    # Auth once; reuse across every territory. Saves ~1s × N sf-CLI subprocess
    # calls and removes N/A token-rotation windows mid-run.
    token, instance_url = get_auth()
    session = build_session(token)

    def _run_one(territory):
        config = TERRITORIES[territory]
        slug = re.sub(r"[^a-z0-9]+", "-", config["director"].lower()).strip("-")
        output_path = args.output_root / args.snapshot_date / f"{slug}.xlsx"
        return extract_territory(
            territory,
            args.snapshot_date,
            output_path,
            session=session,
            instance_url=instance_url,
        )

    # Parallelize when multiple territories — each extraction is ~80% I/O
    # against SF, so threads scale well. Sequential for a single-territory
    # run to keep logs readable.
    if len(territories) > 1:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        print(f"\nExtracting {len(territories)} territories in parallel...\n")
        with ThreadPoolExecutor(max_workers=min(9, len(territories))) as pool:
            futures = {pool.submit(_run_one, t): t for t in territories}
            failures = []
            for f in as_completed(futures):
                t = futures[f]
                try:
                    f.result()
                except Exception as exc:
                    failures.append((t, exc))
                    print(f"  [FAIL] {t}: {exc}")
            if failures:
                print(f"\n{len(failures)} failure(s):")
                for t, exc in failures:
                    print(f"  {t}: {exc}")
                sys.exit(1)
    else:
        for t in territories:
            _run_one(t)

    # Flush query telemetry to a run log so manifest builders can pick it up.
    if QUERY_TELEMETRY:
        from datetime import datetime as _dt

        tele_dir = REPO_ROOT / "output" / "pipeline_logs" / args.snapshot_date
        tele_dir.mkdir(parents=True, exist_ok=True)
        tele_path = tele_dir / "sf_query_telemetry.json"
        existing = []
        if tele_path.exists():
            try:
                existing = json.loads(tele_path.read_text())
            except (json.JSONDecodeError, ValueError):
                existing = []
        existing.append(
            {
                "stage": "extract_director_live",
                "run_at": _dt.now().isoformat(timespec="seconds"),
                "queries": QUERY_TELEMETRY,
                "totals": {
                    "queries": len(QUERY_TELEMETRY),
                    "rows": sum(q["rows"] for q in QUERY_TELEMETRY),
                    "duration_ms": sum(q["duration_ms"] for q in QUERY_TELEMETRY),
                },
            }
        )
        tele_path.write_text(json.dumps(existing, indent=2) + "\n")
        print(f"\nSF query telemetry: {tele_path.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
