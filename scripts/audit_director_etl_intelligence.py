#!/usr/bin/env python3
"""Audit DirectorBundle JSON against rendered workbook analytics coverage.

This is a read-only ETL intelligence gate:
- JSON bundle is the source contract.
- Excel workbook is a downstream render artifact.
- The audit detects where source facts are present but not exposed to analysts.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.monthly_platform.models import DirectorBundle  # noqa: E402
from scripts.monthly_platform.intelligence import (  # noqa: E402
    as_rows,
    build_deal_risk_table,
    churn_table,
    movement_summary,
    owner_metrics,
)


DATASET_RENDER_CONTRACT: dict[str, dict[str, Any]] = {
    "pipeline_open": {
        "sheet": "Pipeline Open {fy}",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "stage",
            "forecast_category",
            "close_date",
            "arr_unweighted",
            "arr_weighted",
            "probability",
            "push_count",
            "deal_type",
            "lead_scope",
            "industry",
            "tier",
            "sales_region",
            "created_date",
            "last_activity_date",
            "next_step",
            "last_modified_date",
            "approved",
            "approval_date",
            "competitor",
        },
    },
    "won_lost": {
        "sheet": "Won Lost {fy}",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "stage",
            "close_date",
            "arr_unweighted",
            "deal_type",
            "reason_won_lost",
            "competitor",
            "industry",
            "sales_region",
            "created_date",
        },
    },
    "renewals": {
        "sheet": "Renewals {fy}",
        "rendered_fields": {
            "close_date",
            "account",
            "opportunity",
            "owner",
            "stage",
            "acv_unweighted",
            "probability",
            "comments",
        },
    },
    "approvals": {
        "sheet": "Commercial Approval",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "stage",
            "close_date",
            "arr_unweighted",
            "status",
            "approval_date",
            "next_step",
            "lead_scope",
        },
    },
    "pi_current": {
        "sheet": "Pipeline Inspection",
        "rendered_fields": {
            "opportunity",
            "owner",
            "stage",
            "forecast_category",
            "arr_weighted",
            "currency",
            "close_date",
            "push_count",
            "score",
            "priority",
        },
    },
    "pi_forward": {
        "sheet": "Pipeline Inspection Forward",
        "rendered_fields": {
            "opportunity",
            "owner",
            "stage",
            "forecast_category",
            "arr_weighted",
            "currency",
            "close_date",
            "push_count",
            "score",
            "priority",
        },
    },
    "activity": {
        "sheet": "Activity Volume",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "tasks_90d",
            "events_90d",
            "total_touches_90d",
            "last_activity_date",
            "flag",
        },
    },
    "commit_items": {
        "sheet": "Commit Items",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "forecast_category",
            "arr_weighted",
            "arr_unweighted",
            "close_date",
            "period",
            "stage",
        },
    },
    "movement_prior": {
        "sheet": "Q1 Movement",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "stage",
            "movement_type",
            "old_close",
            "new_close",
            "changed_on",
            "arr_unweighted",
        },
    },
    "movement_current": {
        "sheet": "Q2 Movement",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "stage",
            "movement_type",
            "old_close",
            "new_close",
            "changed_on",
            "arr_unweighted",
        },
    },
    "stage_events": {
        "sheet": "Stage History",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "current_stage",
            "old_value",
            "new_value",
            "created_date",
            "arr_unweighted",
        },
    },
    "forecast_category_events": {
        "sheet": "Forecast Category History",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "current_stage",
            "old_value",
            "new_value",
            "created_date",
            "arr_unweighted",
        },
    },
    "close_date_events": {
        "sheet": "Close Date History",
        "rendered_fields": {
            "account",
            "opportunity",
            "owner",
            "current_stage",
            "old_value",
            "new_value",
            "created_date",
            "arr_unweighted",
            "is_closed",
        },
    },
    "snapshot_trend": {
        "sheet": None,
        "rendered_fields": set(),
        "expected_sheet": "Snapshot Trend",
    },
}


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "unknown"


def fy_from_snapshot(snapshot_date: str) -> str:
    year = int(snapshot_date[:4])
    return f"FY{year % 100:02d}"


def is_populated(value: Any) -> bool:
    return value is not None and value != ""


def parse_iso_date(value: Any) -> date | None:
    if not isinstance(value, str) or len(value) < 10:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def dataset_profile(rows: list[dict[str, Any]]) -> dict[str, Any]:
    fields = sorted({field for row in rows for field in row})
    field_stats: dict[str, dict[str, Any]] = {}
    for field in fields:
        values = [row.get(field) for row in rows]
        populated = [value for value in values if is_populated(value)]
        numeric = [
            float(value)
            for value in populated
            if isinstance(value, int | float) and not isinstance(value, bool)
        ]
        dates = [d for d in (parse_iso_date(value) for value in populated) if d]
        categorical = [
            str(value)
            for value in populated
            if not isinstance(value, int | float) or isinstance(value, bool)
        ]
        stat: dict[str, Any] = {
            "populated": len(populated),
            "populated_ratio": round(len(populated) / len(rows), 4) if rows else 0,
        }
        if numeric:
            stat["numeric_sum"] = round(sum(numeric), 2)
            stat["numeric_avg"] = round(mean(numeric), 2)
            stat["numeric_max"] = round(max(numeric), 2)
        if dates:
            stat["min_date"] = min(dates).isoformat()
            stat["max_date"] = max(dates).isoformat()
        if categorical:
            stat["top_values"] = Counter(categorical).most_common(5)
        field_stats[field] = stat
    return {
        "row_count": len(rows),
        "field_count": len(fields),
        "fields": fields,
        "field_stats": field_stats,
    }


def workbook_profile(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    profile: dict[str, Any] = {"path": str(workbook_path), "sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [
            ws.cell(row=1, column=col).value
            for col in range(1, (ws.max_column or 0) + 1)
        ]
        data_rows = max((ws.max_row or 1) - 1, 0)
        profile["sheets"][name] = {
            "rows": data_rows,
            "columns": len([h for h in headers if h is not None]),
            "headers": [h for h in headers if h is not None],
        }
    wb.close()
    return profile


def sheet_name_for(dataset: str, fy: str) -> str | None:
    template = DATASET_RENDER_CONTRACT[dataset].get("sheet")
    return template.format(fy=fy) if template else None


def classify_coverage_gaps(
    *,
    rows_by_dataset: dict[str, list[dict[str, Any]]],
    profiles: dict[str, dict[str, Any]],
    workbook: dict[str, Any],
    fy: str,
) -> list[dict[str, Any]]:
    gaps: list[dict[str, Any]] = []
    sheets = workbook["sheets"]
    for dataset, rows in rows_by_dataset.items():
        contract = DATASET_RENDER_CONTRACT[dataset]
        row_count = len(rows)
        sheet = sheet_name_for(dataset, fy)
        if row_count and sheet is None:
            gaps.append(
                {
                    "severity": "high",
                    "type": "dataset_not_rendered",
                    "dataset": dataset,
                    "rows": row_count,
                    "expected_sheet": contract.get("expected_sheet"),
                    "why_it_matters": (
                        "Source facts exist in JSON but are invisible in the analyst workbook."
                    ),
                }
            )
        if sheet and row_count and sheet not in sheets:
            gaps.append(
                {
                    "severity": "high",
                    "type": "sheet_missing",
                    "dataset": dataset,
                    "sheet": sheet,
                    "rows": row_count,
                    "why_it_matters": "Workbook row parity cannot be trusted.",
                }
            )
        if sheet and sheet in sheets:
            workbook_rows = sheets[sheet]["rows"]
            if workbook_rows != row_count:
                gaps.append(
                    {
                        "severity": "high",
                        "type": "row_count_mismatch",
                        "dataset": dataset,
                        "sheet": sheet,
                        "bundle_rows": row_count,
                        "workbook_rows": workbook_rows,
                    }
                )
        rendered = set(contract.get("rendered_fields") or set())
        missing_fields = []
        for field in profiles[dataset]["fields"]:
            if field in rendered:
                continue
            stats = profiles[dataset]["field_stats"][field]
            if stats["populated"] == 0:
                continue
            missing_fields.append(
                {
                    "field": field,
                    "populated": stats["populated"],
                    "populated_ratio": stats["populated_ratio"],
                    "top_values": stats.get("top_values"),
                    "numeric_sum": stats.get("numeric_sum"),
                }
            )
        if missing_fields:
            severity = "medium"
            if row_count and sheet is None:
                severity = "high"
            elif any(f["populated_ratio"] >= 0.8 for f in missing_fields):
                severity = "medium"
            gaps.append(
                {
                    "severity": severity,
                    "type": "populated_fields_not_rendered",
                    "dataset": dataset,
                    "rows": row_count,
                    "sheet": sheet,
                    "fields": missing_fields,
                }
            )
    return gaps


def build_recommendations(
    *,
    gaps: list[dict[str, Any]],
    risk_table: list[dict[str, Any]],
    rows: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    recommendations = []
    if any(
        g["type"] in {"dataset_not_rendered", "sheet_missing"}
        and g["dataset"] == "close_date_events"
        for g in gaps
    ):
        recommendations.append(
            {
                "priority": "P0",
                "recommendation": "Add a Close Date History workbook sheet and deck-ready churn summary.",
                "evidence": f"{len(rows['close_date_events'])} close-date events exist in JSON but are not rendered.",
            }
        )
    if risk_table:
        recommendations.append(
            {
                "priority": "P0",
                "recommendation": "Promote a Deal Risk Index from joined JSON facts into the workbook/deck fact pack.",
                "evidence": f"{len(risk_table)} scored risk rows generated from activity, approval, push, and history signals.",
            }
        )
    if rows["snapshot_trend"] == []:
        recommendations.append(
            {
                "priority": "P1",
                "recommendation": "Decide whether snapshot trend is mandatory; it is modeled but empty in this bundle.",
                "evidence": "The semantic layer has snapshot_trend, but this run has zero rows.",
            }
        )
    if any(g["type"] == "populated_fields_not_rendered" for g in gaps):
        recommendations.append(
            {
                "priority": "P1",
                "recommendation": "Add an ETL Coverage Appendix that lists high-population fields omitted from workbook tabs.",
                "evidence": "Some modeled JSON fields are populated but not exposed in Excel.",
            }
        )
    return recommendations


def build_etl_intelligence_audit(
    *, bundle_path: Path, workbook_path: Path
) -> dict[str, Any]:
    bundle = DirectorBundle.from_json(bundle_path.read_text(encoding="utf-8"))
    rows = as_rows(bundle)
    fy = fy_from_snapshot(bundle.snapshot_date)
    profiles = {name: dataset_profile(dataset_rows) for name, dataset_rows in rows.items()}
    workbook = workbook_profile(workbook_path)
    gaps = classify_coverage_gaps(
        rows_by_dataset=rows, profiles=profiles, workbook=workbook, fy=fy
    )
    risk_table = build_deal_risk_table(rows)
    owner_table = owner_metrics(rows)
    close_churn = churn_table(rows, "close_date_events", "close_date")
    forecast_churn = churn_table(rows, "forecast_category_events", "forecast_category")
    stage_churn = churn_table(rows, "stage_events", "stage")
    recommendations = build_recommendations(
        gaps=gaps, risk_table=risk_table, rows=rows
    )
    high_gaps = [g for g in gaps if g["severity"] == "high"]
    medium_gaps = [g for g in gaps if g["severity"] == "medium"]
    return {
        "schema_version": 1,
        "audit_type": "director_etl_intelligence_audit",
        "bundle_path": str(bundle_path),
        "workbook_path": str(workbook_path),
        "snapshot_date": bundle.snapshot_date,
        "director": bundle.director,
        "territory": bundle.territory,
        "summary": {
            "datasets": len(rows),
            "bundle_rows": sum(len(dataset_rows) for dataset_rows in rows.values()),
            "workbook_sheets": len(workbook["sheets"]),
            "coverage_gap_count": len(gaps),
            "high_gap_count": len(high_gaps),
            "medium_gap_count": len(medium_gaps),
            "deal_risk_rows": len(risk_table),
            "recommendation_count": len(recommendations),
        },
        "workbook_profile": workbook,
        "dataset_profiles": profiles,
        "coverage_gaps": gaps,
        "recommendations": recommendations,
        "analytics": {
            "deal_risk_index": risk_table,
            "owner_metrics": owner_table,
            "movement_summary": movement_summary(rows),
            "close_date_churn": close_churn,
            "forecast_category_churn": forecast_churn,
            "stage_churn": stage_churn,
        },
    }


def markdown_summary(audit: dict[str, Any]) -> str:
    summary = audit["summary"]
    lines = [
        f"# ETL Intelligence Audit - {audit['director']}",
        "",
        f"- Snapshot date: `{audit['snapshot_date']}`",
        f"- Territory: `{audit['territory']}`",
        f"- Bundle rows: `{summary['bundle_rows']}` across `{summary['datasets']}` datasets",
        f"- Workbook sheets: `{summary['workbook_sheets']}`",
        f"- Coverage gaps: `{summary['coverage_gap_count']}` (`{summary['high_gap_count']}` high)",
        f"- Deal risk rows: `{summary['deal_risk_rows']}`",
        "",
        "## Recommendations",
        "",
    ]
    for rec in audit["recommendations"]:
        lines.append(f"- **{rec['priority']}** {rec['recommendation']} Evidence: {rec['evidence']}")
    if not audit["recommendations"]:
        lines.append("- No recommendations generated.")
    lines.extend(["", "## High Coverage Gaps", ""])
    high_gaps = [g for g in audit["coverage_gaps"] if g["severity"] == "high"]
    for gap in high_gaps:
        if gap["type"] == "dataset_not_rendered":
            lines.append(
                f"- `{gap['dataset']}` has `{gap['rows']}` JSON rows but no workbook sheet."
            )
        elif gap["type"] == "sheet_missing":
            lines.append(
                f"- `{gap['dataset']}` expected `{gap['sheet']}` but sheet is missing."
            )
        elif gap["type"] == "row_count_mismatch":
            lines.append(
                f"- `{gap['dataset']}` row mismatch: bundle `{gap['bundle_rows']}`, workbook `{gap['workbook_rows']}`."
            )
        else:
            lines.append(f"- `{gap['dataset']}`: {gap['type']}")
    if not high_gaps:
        lines.append("- None.")
    lines.extend(["", "## Top Deal Risk Rows", ""])
    for row in audit["analytics"]["deal_risk_index"][:10]:
        reasons = "; ".join(row["risk_reasons"])
        lines.append(
            f"- `{row['opportunity']}` ({row['owner']}): score `{row['risk_score']}`, "
            f"ARR `{row['arr_unweighted']:,.0f}` - {reasons}"
        )
    if not audit["analytics"]["deal_risk_index"]:
        lines.append("- None.")
    lines.append("")
    return "\n".join(lines)


def default_output_dir(bundle_path: Path, audit: dict[str, Any]) -> Path:
    return (
        ROOT
        / "output"
        / "etl_intelligence_audit"
        / audit["snapshot_date"]
        / slugify(audit["director"])
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit DirectorBundle JSON against workbook analytics coverage."
    )
    parser.add_argument("--bundle", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument(
        "--fail-on-high",
        action="store_true",
        help="Exit non-zero if high-severity coverage gaps are found.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    audit = build_etl_intelligence_audit(
        bundle_path=args.bundle, workbook_path=args.workbook
    )
    output_dir = args.output_dir or default_output_dir(args.bundle, audit)
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "etl_intelligence_audit.json"
    md_path = output_dir / "summary.md"
    json_path.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    md_path.write_text(markdown_summary(audit), encoding="utf-8")
    summary = audit["summary"]
    print(
        json.dumps(
            {
                "status": "completed",
                "json": str(json_path),
                "summary_md": str(md_path),
                "high_gap_count": summary["high_gap_count"],
                "coverage_gap_count": summary["coverage_gap_count"],
                "deal_risk_rows": summary["deal_risk_rows"],
            },
            indent=2,
        )
    )
    if args.fail_on_high and summary["high_gap_count"]:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
