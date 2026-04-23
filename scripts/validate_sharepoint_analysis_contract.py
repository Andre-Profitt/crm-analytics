#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from monthly_platform.historical_trending import (
        resolve_historical_trending_contract,
    )
    from monthly_platform.period import resolve_period_context
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.historical_trending import (
        resolve_historical_trending_contract,
    )
    from scripts.monthly_platform.period import resolve_period_context


ROOT = Path(__file__).resolve().parents[1]
SHAREPOINT_ROOT = ROOT / "output" / "sharepoint"
OUTPUT_ROOT = ROOT / "output" / "sharepoint_analysis_contract"


def _resolve_workbook_names(snapshot_date: str | None = None):
    kw = {"snapshot_date": snapshot_date} if snapshot_date else {}
    period = resolve_period_context(**kw)
    fy = period.fiscal_year
    pq = period.prior_quarter
    return {
        "period": period,
        "fy": fy,
        "master": f"{fy} Pipeline Review, All Territories.xlsx",
        "dashboard": f"Dashboard and {pq.label} Analysis.xlsx",
        "approvals_sheet": f"Approvals, 20{fy[-2:]}",
    }


_WB = _resolve_workbook_names()
_FY = _WB["fy"]

MASTER_WORKBOOK = _WB["master"]
DASHBOARD_WORKBOOK = _WB["dashboard"]

_TERRITORY_TO_REGIONAL_LABEL: dict[str, str] = {
    "APAC": "APAC",
    "Central Europe": "EMEA Central",
    "UK & Ireland": "EMEA UK & Ireland",
    "Southern Europe": "EMEA South West",
    "NL & Nordics": "EMEA NE",
    "Middle East & Africa": "EMEA MEA",
    "Canada": "NA Canada",
    "NA Asset Management": "NA Asset Mgmt",
    "Pension & Insurance": "NA Insurance",
}

_REGIONAL_WORKBOOKS_FALLBACK = [
    ("APAC", f"{_FY} Pipeline Review, APAC.xlsx"),
    ("EMEA Central", f"{_FY} Pipeline Review, EMEA Central.xlsx"),
    ("EMEA UK & Ireland", f"{_FY} Pipeline Review, EMEA UK & Ireland.xlsx"),
    ("EMEA NE", f"{_FY} Pipeline Review, EMEA NE.xlsx"),
    ("EMEA South West", f"{_FY} Pipeline Review, EMEA South West.xlsx"),
    ("EMEA MEA", f"{_FY} Pipeline Review, EMEA MEA.xlsx"),
    ("NA Asset Mgmt", f"{_FY} Pipeline Review, NA Asset Mgmt.xlsx"),
    ("NA Canada", f"{_FY} Pipeline Review, NA Canada.xlsx"),
    ("NA Insurance", f"{_FY} Pipeline Review, NA Insurance.xlsx"),
]


def _load_regional_workbooks() -> list[tuple[str, str]]:
    cfg_path = ROOT / "config" / "sd_monthly_territories.json"
    if not cfg_path.exists():
        return _REGIONAL_WORKBOOKS_FALLBACK
    try:
        territories = json.loads(cfg_path.read_text(encoding="utf-8")).get(
            "territories", {}
        )
        return [
            (
                _TERRITORY_TO_REGIONAL_LABEL.get(key, key),
                f"{_FY} Pipeline Review, {_TERRITORY_TO_REGIONAL_LABEL.get(key, key)}.xlsx",
            )
            for key in territories
        ]
    except Exception:
        return _REGIONAL_WORKBOOKS_FALLBACK


REGIONAL_WORKBOOKS = _load_regional_workbooks()

MASTER_REQUIRED_SHEETS_BASE = [
    "Executive Insights",
    "Parameters",
    "Summary",
    "Forecast Reconciliation",
    "Land Pipeline Detail",
    "Land WonLost Detail",
    f"Approvals, 20{_FY[-2:]}",
    "Approval Candidates",
    "Land Stage 3+, No Approval",
    "Renewals This Quarter",
    "Pipeline Pivot",
    "ARR Concentration",
    "Pipeline Velocity",
    "Slip Risk by Owner",
    "Territory Scorecard",
    "Deal Risk Scoring",
    "Source Map",
    "Methodology",
]

REGIONAL_REQUIRED_SHEETS = [
    "Executive Insights",
    "Parameters",
    "Summary",
    "Forecast Reconciliation",
    "Land Pipeline Detail",
    "Land WonLost Detail",
    f"Approvals, 20{_FY[-2:]}",
    "Approval Candidates",
    "Land Stage 3+, No Approval",
    "Renewals This Quarter",
    "Pipeline Pivot",
    "ARR Concentration",
    "Pipeline Velocity",
    "Slip Risk by Owner",
    "Territory Scorecard",
    "Source Map",
    "Methodology",
]


def _dashboard_required_sheets(prior_q_label: str = "Q1") -> list[str]:
    return [
        "Dashboard Overview",
        "Pipeline Overview by Stage",
        "Business At Risk",
        "Stage Transition Matrix",
        f"{prior_q_label} History Raw",
        "Pipeline Inspection Raw",
        "PI Summary",
        "Methodology",
    ]


DASHBOARD_REQUIRED_SHEETS = _dashboard_required_sheets()

MIN_SHEET_COUNTS = {
    "master": 37,
    "regional": 32,
    "dashboard": 35,
}


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _infer_report_date(sharepoint_root: Path | None, report_date: str | None) -> str:
    token = str(report_date or "").strip()[:10]
    if token:
        return token
    if sharepoint_root is not None:
        for candidate in [Path(sharepoint_root), *Path(sharepoint_root).parents]:
            try:
                datetime.strptime(candidate.name, "%Y-%m-%d")
                return candidate.name
            except ValueError:
                continue
    return datetime.now().strftime("%Y-%m-%d")


def _historical_contract(report_date: str):
    period = resolve_period_context(
        as_of_date=report_date,
        snapshot_date=report_date,
        deck_date=report_date,
    )
    return resolve_historical_trending_contract(
        retrospective_label=period.prior_quarter.label,
        retrospective_title=period.prior_quarter.title,
        current_label=period.current_quarter.label,
        current_title=period.current_quarter.title,
    )


def _read_workbook_contract(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheetnames = list(wb.sheetnames)
        sheet_rows = {name: int(wb[name].max_row or 0) for name in sheetnames}
    finally:
        wb.close()
    return {
        "sheetnames": sheetnames,
        "sheet_count": len(sheetnames),
        "sheet_rows": sheet_rows,
        "file_size_bytes": int(path.stat().st_size),
    }


def _validate_workbook(
    *,
    workbook_id: str,
    workbook_type: str,
    territory: str | None,
    path: Path,
    required_sheets: list[str],
    min_sheet_count: int,
    validated: list[dict[str, Any]],
    failures: list[dict[str, Any]],
) -> None:
    if not path.exists():
        failures.append(
            {
                "workbook_id": workbook_id,
                "workbook_type": workbook_type,
                "territory": territory,
                "issue": "missing_workbook",
                "message": f"missing {path}",
            }
        )
        return

    try:
        info = _read_workbook_contract(path)
    except Exception as exc:
        failures.append(
            {
                "workbook_id": workbook_id,
                "workbook_type": workbook_type,
                "territory": territory,
                "issue": "workbook_load_failed",
                "message": str(exc),
            }
        )
        return

    if int(info["sheet_count"]) < min_sheet_count:
        failures.append(
            {
                "workbook_id": workbook_id,
                "workbook_type": workbook_type,
                "territory": territory,
                "issue": "sheet_count_below_minimum",
                "message": (
                    f"{info['sheet_count']} sheet(s); expected at least "
                    f"{min_sheet_count}"
                ),
            }
        )

    missing_sheets = [
        name for name in required_sheets if name not in info["sheetnames"]
    ]
    if missing_sheets:
        failures.append(
            {
                "workbook_id": workbook_id,
                "workbook_type": workbook_type,
                "territory": territory,
                "issue": "missing_required_sheets",
                "message": ", ".join(missing_sheets),
            }
        )
        return

    if int(info["sheet_count"]) < min_sheet_count:
        return

    validated.append(
        {
            "workbook_id": workbook_id,
            "workbook_type": workbook_type,
            "territory": territory,
            "workbook_path": _display_path(path),
            "sheet_count": int(info["sheet_count"]),
            "file_size_bytes": int(info["file_size_bytes"]),
            "required_sheets": list(required_sheets),
            "key_sheet_row_counts": {
                name: int(info["sheet_rows"].get(name) or 0) for name in required_sheets
            },
        }
    )


def _write_run_audit(output_dir: Path, payload: dict[str, Any]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "sharepoint_analysis_contract_audit.json").write_text(
        json.dumps(payload, indent=2) + "\n",
        encoding="utf-8",
    )
    lines = [
        f"# SharePoint Analysis Contract Audit — {payload['run_date']}",
        "",
        f"- Status: `{payload['status']}`",
        f"- SharePoint root: `{payload['sharepoint_root']}`",
        f"- Validated workbooks: `{len(payload.get('validated') or [])}`",
        f"- Failures: `{len(payload.get('failures') or [])}`",
        f"- Warnings: `{len(payload.get('warnings') or [])}`",
        "",
        "## Validated",
        "",
    ]
    validated = payload.get("validated") or []
    if not validated:
        lines.append("- none")
    else:
        for item in validated:
            territory = f" ({item['territory']})" if item.get("territory") else ""
            lines.append(
                f"- `{item['workbook_id']}`{territory}: "
                f"`{item['sheet_count']}` sheet(s)"
            )
    lines.extend(["", "## Failures", ""])
    failures = payload.get("failures") or []
    if not failures:
        lines.append("- none")
    else:
        for item in failures:
            lines.append(
                f"- `{item.get('workbook_id', '')}`: "
                f"`{item.get('issue', 'unknown')}` {item.get('message', '')}".strip()
            )
    lines.extend(["", "## Warnings", ""])
    warnings = payload.get("warnings") or []
    if not warnings:
        lines.append("- none")
    else:
        for item in warnings:
            lines.append(
                f"- `{item.get('workbook_id', '')}`: "
                f"`{item.get('issue', 'unknown')}` {item.get('message', '')}".strip()
            )
    (output_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--date",
        help="Explicit report date (YYYY-MM-DD) used for quarter-aware sheet checks.",
    )
    parser.add_argument(
        "--sharepoint-root",
        type=Path,
        help="Directory containing the analysis workbooks. Defaults to output/sharepoint.",
    )
    args = parser.parse_args()

    run_date = _infer_report_date(args.sharepoint_root, args.date)
    wb_names = _resolve_workbook_names(run_date)
    master_wb = wb_names["master"]
    dashboard_wb = wb_names["dashboard"]
    approvals_sheet = wb_names["approvals_sheet"]
    sharepoint_root = args.sharepoint_root or SHAREPOINT_ROOT
    output_dir = OUTPUT_ROOT / run_date
    validated: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    historical_contract = _historical_contract(run_date)
    req_approvals = approvals_sheet
    master_required = [
        s if s != f"Approvals, 20{_FY[-2:]}" else req_approvals
        for s in MASTER_REQUIRED_SHEETS_BASE
    ] + [
        historical_contract.retrospective_consolidated_sheet,
        historical_contract.current_consolidated_sheet,
    ]
    regional_required = [
        s if s != f"Approvals, 20{_FY[-2:]}" else req_approvals
        for s in REGIONAL_REQUIRED_SHEETS
    ]

    _validate_workbook(
        workbook_id="master",
        workbook_type="master",
        territory=None,
        path=sharepoint_root / master_wb,
        required_sheets=master_required,
        min_sheet_count=MIN_SHEET_COUNTS["master"],
        validated=validated,
        failures=failures,
    )

    regional_workbooks = _load_regional_workbooks()
    for territory, filename in regional_workbooks:
        _validate_workbook(
            workbook_id=f"regional:{territory}",
            workbook_type="regional",
            territory=territory,
            path=sharepoint_root / filename,
            required_sheets=regional_required,
            min_sheet_count=MIN_SHEET_COUNTS["regional"],
            validated=validated,
            failures=failures,
        )

    _validate_workbook(
        workbook_id="dashboard_q1",
        workbook_type="dashboard",
        territory=None,
        path=sharepoint_root / dashboard_wb,
        required_sheets=_dashboard_required_sheets(
            wb_names["period"].prior_quarter.label
        ),
        min_sheet_count=MIN_SHEET_COUNTS["dashboard"],
        validated=validated,
        failures=failures,
    )

    fy = wb_names["fy"]
    expected_regional_files = {filename for _, filename in regional_workbooks}
    actual_regional_files = {
        path.name
        for path in sharepoint_root.glob(f"{fy} Pipeline Review, *.xlsx")
        if path.name != master_wb
    }
    for filename in sorted(actual_regional_files - expected_regional_files):
        warnings.append(
            {
                "workbook_id": filename,
                "issue": "unexpected_regional_workbook",
                "message": filename,
            }
        )

    payload = {
        "run_date": run_date,
        "sharepoint_root": str(sharepoint_root),
        "status": "failed" if failures else "ok",
        "validated": validated,
        "failures": failures,
        "warnings": warnings,
    }
    _write_run_audit(output_dir, payload)
    print(f"SharePoint analysis contract audit: {_display_path(output_dir)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
