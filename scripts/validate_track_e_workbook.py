#!/usr/bin/env python3
"""Track E — real director workbook validator (E2).

Opens an actual director workbook (.xlsx) and validates it against
``config/director_workbook_contract.yaml``. Distinct from
``scripts/validate_track_e_workbook_contract.py`` (E1) which only
validates the YAML contract structurally; this script verifies that a
real workbook *file* matches what the contract declares.

Checks:
  - workbook opens
  - all 13 declared sheets exist
  - each sheet's required_columns are present at the declared
    header_row (default row 1; Q1/Q2 Snapshot Trend use row 2 because
    row 1 is a free-text snapshot manifest)
  - each pattern-based snapshot_role resolves to exactly one column
    when selection=earliest or latest is applied
  - runtime-based snapshot_roles are reported (no workbook check —
    they bind to run_metadata)
  - status enum values (where declared) match observed values
  - movement enum values (where declared) match observed values

Emits ``output/track_e/director_workbook_validation_report.json`` and
``.md`` summarising sheet status, missing columns, resolved snapshot
roles, and any enum drift.

Usage:
    python scripts/validate_track_e_workbook.py \\
        --workbook /Users/test/Downloads/jesper-tyrer-2026-04-20.xlsx \\
        --report-out output/track_e/director_workbook_validation_report.json
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl  # noqa: E402

from scripts.monthly_platform import director_workbook_contract  # noqa: E402


REPORT_SCHEMA_VERSION = "monthly_platform.director_workbook_validation.v1"


def _headers_for_sheet(ws, header_row: int) -> list[str]:
    return [
        str(ws.cell(header_row, c).value)
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value is not None
    ]


def _column_values(ws, header_row: int, column_name: str) -> list[Any]:
    """Return all data values under a named column (header_row + 1 onward)."""
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value) == column_name:
            col_idx = c
            break
    if col_idx is None:
        return []
    return [ws.cell(r, col_idx).value for r in range(header_row + 1, ws.max_row + 1)]


def validate_workbook(
    workbook_path: Path,
    *,
    contract: director_workbook_contract.DirectorWorkbookContract | None = None,
) -> dict[str, Any]:
    if contract is None:
        contract = director_workbook_contract.load()
    assert contract is not None
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    findings: list[dict[str, Any]] = []
    sheet_statuses: list[dict[str, Any]] = []
    actual_headers_by_sheet: dict[str, list[str]] = {}

    sheets_by_name = contract.sheets_by_name()

    # 1. sheets + required columns + enum drift.
    for sheet_decl in contract.raw.get("sheets", []):
        name = sheet_decl["name"]
        if name not in wb.sheetnames:
            findings.append(
                {
                    "severity": "blocker",
                    "code": "missing_sheet",
                    "path": f"sheets[{name!r}]",
                    "message": f"workbook is missing sheet {name!r}",
                }
            )
            sheet_statuses.append(
                {
                    "name": name,
                    "status": "missing",
                }
            )
            continue

        ws = wb[name]
        header_row = sheet_decl.get("header_row", 1)
        actual_headers = _headers_for_sheet(ws, header_row)
        actual_headers_by_sheet[name] = actual_headers
        actual_set = set(actual_headers)

        missing_cols = [
            c for c in (sheet_decl.get("required_columns") or []) if c not in actual_set
        ]
        for col in missing_cols:
            findings.append(
                {
                    "severity": "blocker",
                    "code": "missing_required_column",
                    "path": f"sheets[{name!r}].required_columns",
                    "message": f"sheet {name!r} missing required column {col!r}",
                }
            )

        # Enum drift checks (Status, Movement).
        enum_drift: dict[str, list[str]] = {}
        if sheet_decl.get("expected_status_values") and "Status" in actual_set:
            observed = {
                str(v).strip()
                for v in _column_values(ws, header_row, "Status")
                if v is not None and str(v).strip()
            }
            unexpected = sorted(observed - set(sheet_decl["expected_status_values"]))
            if unexpected:
                enum_drift["Status"] = unexpected
                findings.append(
                    {
                        "severity": "warning",
                        "code": "status_enum_drift",
                        "path": f"sheets[{name!r}].expected_status_values",
                        "message": f"sheet {name!r} has unexpected Status values: {unexpected}",
                    }
                )

        if sheet_decl.get("expected_movement_values") and "Movement" in actual_set:
            observed = {
                str(v).strip()
                for v in _column_values(ws, header_row, "Movement")
                if v is not None and str(v).strip()
            }
            unexpected = sorted(observed - set(sheet_decl["expected_movement_values"]))
            if unexpected:
                enum_drift["Movement"] = unexpected
                findings.append(
                    {
                        "severity": "warning",
                        "code": "movement_enum_drift",
                        "path": f"sheets[{name!r}].expected_movement_values",
                        "message": f"sheet {name!r} has unexpected Movement values: {unexpected}",
                    }
                )

        # Row count (informational).
        sheet_statuses.append(
            {
                "name": name,
                "status": "missing_columns" if missing_cols else "pass",
                "row_count": ws.max_row,
                "column_count": ws.max_column,
                "missing_columns": missing_cols,
                "enum_drift": enum_drift or None,
            }
        )

    # 2. snapshot_role resolution.
    resolved_roles: list[dict[str, Any]] = []
    for role_name in contract.snapshot_roles().keys():
        resolved = contract.resolve_pattern_role(role_name, actual_headers_by_sheet)
        resolved_roles.append(resolved.as_dict())
        if resolved.status == "missing":
            findings.append(
                {
                    "severity": "blocker",
                    "code": "snapshot_role_unresolved",
                    "path": f"snapshot_roles[{role_name!r}]",
                    "message": resolved.detail or f"role {role_name!r} did not resolve",
                }
            )

    blockers = [f for f in findings if f["severity"] == "blocker"]
    warnings = [f for f in findings if f["severity"] == "warning"]

    return {
        "schema_version": REPORT_SCHEMA_VERSION,
        "workbook_path": str(workbook_path),
        "contract_path": str(contract.path),
        "validated_at": datetime.now(timezone.utc).isoformat(),
        "status": "pass" if not blockers else "fail",
        "blocker_count": len(blockers),
        "warning_count": len(warnings),
        "sheet_count_declared": len(sheets_by_name),
        "sheet_count_present": sum(
            1 for s in sheet_statuses if s["status"] != "missing"
        ),
        "snapshot_role_count": len(resolved_roles),
        "resolved_snapshot_roles": resolved_roles,
        "sheet_statuses": sheet_statuses,
        "findings": findings,
    }


def render_markdown(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# Director workbook validation report\n")
    lines.append(f"- workbook: `{report['workbook_path']}`")
    lines.append(f"- contract: `{report['contract_path']}`")
    lines.append(f"- validated_at: {report['validated_at']}")
    lines.append(f"- **status: {report['status']}**")
    lines.append(
        f"- blockers: {report['blocker_count']} | warnings: {report['warning_count']}"
    )
    lines.append(
        f"- sheets: {report['sheet_count_present']}/{report['sheet_count_declared']} present"
    )
    lines.append(f"- snapshot_roles: {report['snapshot_role_count']} declared")
    lines.append("")

    lines.append("## Resolved snapshot roles\n")
    lines.append("| Role | Source | Sheet | Resolved column | Date | Status |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for r in report["resolved_snapshot_roles"]:
        lines.append(
            f"| `{r['role']}` | {r['source']} | {r.get('sheet') or '—'} | "
            f"`{r.get('physical_column') or '—'}` | {r.get('resolved_date') or '—'} | "
            f"{r['status']} |"
        )
    lines.append("")

    lines.append("## Sheets\n")
    lines.append("| Sheet | Status | Rows | Cols | Missing columns |")
    lines.append("| --- | --- | ---: | ---: | --- |")
    for s in report["sheet_statuses"]:
        missing = ", ".join(s.get("missing_columns") or []) or "—"
        lines.append(
            f"| {s['name']} | {s['status']} | {s.get('row_count', '—')} | "
            f"{s.get('column_count', '—')} | {missing} |"
        )
    lines.append("")

    if report["findings"]:
        lines.append("## Findings\n")
        for f in report["findings"]:
            lines.append(f"- **{f['severity']}** `{f['code']}` — {f['message']}")
        lines.append("")

    return "\n".join(lines) + "\n"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate a real director workbook against the Track E contract."
    )
    parser.add_argument("--workbook", required=True, help="Path to the .xlsx file")
    parser.add_argument("--contract", default=None)
    parser.add_argument("--report-out", default=None)
    parser.add_argument("--md-out", default=None)
    parser.add_argument("--show-findings", action="store_true")
    args = parser.parse_args(argv)

    contract = director_workbook_contract.load(args.contract)
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    report = validate_workbook(workbook_path, contract=contract)

    if args.report_out:
        out = Path(args.report_out)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
        print(f"report: {out}")

    if args.md_out:
        out = Path(args.md_out)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(render_markdown(report), encoding="utf-8")
        print(f"md: {out}")

    if args.show_findings or report["status"] == "fail":
        for f in report["findings"]:
            print(f"[{f['severity']}] {f['code']} {f.get('path', '')}: {f['message']}")

    print(
        f"director_workbook_validation: {report['status']} "
        f"(blockers={report['blocker_count']} warnings={report['warning_count']} "
        f"sheets={report['sheet_count_present']}/{report['sheet_count_declared']} "
        f"roles_resolved={sum(1 for r in report['resolved_snapshot_roles'] if r['status'] == 'pass')}/{report['snapshot_role_count']})"
    )
    return 0 if report["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
