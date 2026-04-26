#!/usr/bin/env python3
"""Phase 2 Workbook Builder — one Excel workbook per Sales Director.

Reads JSON cache files produced by extract_director_data.py and builds a
12-tab Excel workbook per director.

Usage:
    python3 scripts/build_director_workbooks.py --director "Dan Peppett" --snapshot-date 2026-04-10
    python3 scripts/build_director_workbooks.py --all --snapshot-date 2026-04-10
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

# Allow importing shared helpers from scripts/
_SCRIPTS = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCRIPTS))

import director_data_helpers as h  # noqa: E402
from md1_presets import find_md1_preset, load_md1_preset_config  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_BASE = REPO_ROOT / "output" / "director_data_dumps"

# Colours
TEAL_DARK = "003E52"
TEAL_TEXT = "003E52"
WHITE = "FFFFFF"
RED_FILL = "FDE8E8"
AMBER_FILL = "FFF3E0"
GREEN_FILL = "E8F5E9"

TODAY = date.today()


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------


def safe_num(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def safe_str(val, max_len=100):
    if val is None:
        return ""
    s = str(val)
    return s[:max_len] if len(s) > max_len else s


def nested_get(d, *keys, default=None):
    for k in keys:
        if isinstance(d, dict):
            d = d.get(k, default)
        else:
            return default
    return d


def parse_date(val) -> date | None:
    if not val:
        return None
    s = str(val)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def fmt_eur(val) -> str:
    n = safe_num(val)
    if n == 0:
        return "€0"
    if abs(n) >= 1_000_000:
        return f"€{n / 1_000_000:,.1f}M"
    if abs(n) >= 1_000:
        return f"€{n / 1_000:,.0f}K"
    return f"€{n:,.0f}"


def fmt_pct(val) -> str:
    return f"{safe_num(val):.1f}%"


def converted_value(record: dict, converted_key: str, raw_key: str) -> float:
    converted = record.get(converted_key)
    if converted is not None:
        return safe_num(converted)
    return safe_num(record.get(raw_key))


def opp_arr_eur(record: dict) -> float:
    return converted_value(record, "ConvertedARR", "APTS_Opportunity_ARR__c")


def opp_acv_eur(record: dict) -> float:
    return converted_value(record, "ConvertedACV", "Opportunity_Average_ACV__c")


def renewal_acv_eur(record: dict) -> float:
    if (
        record.get("ConvertedRenewalACV") is not None
        or record.get("APTS_Renewal_ACV__c") is not None
    ):
        return converted_value(record, "ConvertedRenewalACV", "APTS_Renewal_ACV__c")
    return opp_acv_eur(record)


def opp_forecast_arr_eur(record: dict) -> float:
    return converted_value(record, "ConvertedForecastARR", "APTS_Forecast_ARR__c")


def forecast_amount_eur(item: dict) -> float:
    """Hierarchical rollup forecast amount for a single ForecastingItem row.

    WARNING: Do NOT sum this across rows of a ForecastingItem result set —
    each opportunity appears once per level of the forecast hierarchy
    (rep -> manager -> regional -> CRO), so summing double-counts every deal
    4-5x. Use forecast_owner_only_eur() to aggregate org-wide totals.
    """
    return converted_value(item, "ConvertedForecastAmount", "ForecastAmount")


def forecast_adjusted_amount_eur(item: dict) -> float:
    return converted_value(item, "ConvertedAdjAmount", "AmountWithoutAdjustments")


def forecast_owner_only_eur(item: dict) -> float:
    """Direct (owner-only) portion of a ForecastingItem row.

    Safe to sum across rows: this is the slice of the forecast that belongs
    directly to the row's owner, with no contribution from subordinates'
    rollups, so it does not double-count across hierarchy levels. Sum of
    OwnerOnlyAmount across all rows of a ForecastingItem query equals the
    org-wide forecast for the relevant period (matches CRO forecast page).
    """
    return converted_value(item, "ConvertedOwnerOnlyAmount", "OwnerOnlyAmount")


def forecast_no_mgr_adj_eur(item: dict) -> float:
    """ForecastingItem amount excluding manager adjustments (still hierarchical)."""
    return converted_value(
        item, "ConvertedNoMgrAdjAmount", "AmountWithoutManagerAdjustment"
    )


def is_omitted(record: dict) -> bool:
    return safe_str(record.get("ForecastCategoryName")).strip().lower() == "omitted"


def active_pipeline_rows(records: list[dict]) -> list[dict]:
    return [record for record in records if not is_omitted(record)]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(bold=True, size=11, color=WHITE) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _data_font(bold=False, size=10, color="000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _title_font() -> Font:
    return Font(name="Calibri", bold=True, size=14, color=TEAL_TEXT)


def _section_font() -> Font:
    return Font(name="Calibri", bold=True, size=12, color=TEAL_TEXT)


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    fill = _make_fill(TEAL_DARK)
    font = _header_font()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def _auto_width(ws, min_w=10, max_w=40) -> None:
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        width = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def _apply_row_fill(ws, row: int, n_cols: int, hex_color: str) -> None:
    fill = _make_fill(hex_color)
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=row, column=col_idx).fill = fill


# ---------------------------------------------------------------------------
# Cache loader
# ---------------------------------------------------------------------------


class DirectorCache:
    def __init__(self, cache_dir: Path):
        self._dir = cache_dir

    def load(self, filename: str, default=None):
        p = self._dir / filename
        if not p.exists():
            return default if default is not None else []
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"  WARN: could not load {filename}: {exc}", file=sys.stderr)
            return default if default is not None else []

    @property
    def open_pipeline(self) -> list[dict]:
        return self.load("soql_open_pipeline.json")

    @property
    def won_this_quarter(self) -> list[dict]:
        return self.load("soql_won_this_quarter.json")

    @property
    def lost_this_quarter(self) -> list[dict]:
        return self.load("soql_lost_this_quarter.json")

    @property
    def won_q1(self) -> list[dict]:
        return self.load("soql_won_q1.json")

    @property
    def lost_q1(self) -> list[dict]:
        return self.load("soql_lost_q1.json")

    @property
    def pushed_deals(self) -> list[dict]:
        return self.load("soql_pushed_deals.json")

    @property
    def new_pipeline(self) -> list[dict]:
        return self.load("soql_new_pipeline.json")

    @property
    def forecast_categories(self) -> list[dict]:
        return self.load("soql_forecast_categories.json")

    @property
    def sources(self) -> list[dict]:
        return self.load("_sources.json")

    def forecast_items(self, period: str, type_: str) -> list[dict]:
        return self.load(f"forecast_item_{period}_{type_}.json")

    def field_history(self, field: str) -> list[dict]:
        return self.load(f"field_history_{field}.json")


# ---------------------------------------------------------------------------
# Tab 1: Scorecard
# ---------------------------------------------------------------------------


def build_scorecard(ws, cache: DirectorCache, director_name: str, territory: str):
    ws.title = "Scorecard"

    # Title
    ws["A1"].value = f"Sales Director Scorecard — {director_name} ({territory})"
    ws["A1"].font = _title_font()
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:B1")

    ws["A2"].value = f"Generated: {TODAY.isoformat()}"
    ws["A2"].font = _data_font(size=9, color="666666")

    current_row = 4

    def section(label: str):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=label).font = _section_font()
        ws.cell(row=current_row, column=1).fill = _make_fill("E8F4F8")
        ws.merge_cells(f"A{current_row}:B{current_row}")
        current_row += 1

    def kpi(name: str, value):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=name).font = _data_font(bold=True)
        ws.cell(row=current_row, column=2, value=value).font = _data_font()
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        current_row += 1

    pipeline = active_pipeline_rows(cache.open_pipeline)
    won_q = cache.won_this_quarter
    lost_q = cache.lost_this_quarter
    new_pipe = active_pipeline_rows(cache.new_pipeline)

    # — PIPELINE HEALTH —
    section("PIPELINE HEALTH")
    total_arr = sum(opp_arr_eur(r) for r in pipeline)
    deal_count = len(pipeline)
    avg_deal = (total_arr / deal_count) if deal_count else 0.0
    weighted = sum(
        opp_arr_eur(r) * safe_num(r.get("Probability")) / 100 for r in pipeline
    )
    new_arr = sum(opp_arr_eur(r) for r in new_pipe)

    # Horizon-scoped pipeline ARR — explicit close-date filters so the
    # "all open" number is never confused with the forecast number.
    fy26_start = date(2026, 1, 1)
    fy26_end = date(2026, 12, 31)

    def _close_in(record, start, end):
        cd = parse_date(record.get("CloseDate"))
        return cd is not None and start <= cd <= end

    fy26_pipeline = [r for r in pipeline if _close_in(r, fy26_start, fy26_end)]
    fy26_pipe_arr = sum(opp_arr_eur(r) for r in fy26_pipeline)
    q2_pipeline = [r for r in pipeline if _close_in(r, Q2_START, Q2_END)]
    q2_pipe_arr = sum(opp_arr_eur(r) for r in q2_pipeline)

    kpi("Pipeline ARR — All Open (any close date)", fmt_eur(total_arr))
    kpi(
        "Pipeline ARR — FY26 Close Dates Only (excl. Omitted)",
        fmt_eur(fy26_pipe_arr),
    )
    kpi(
        "Pipeline ARR — Q2 2026 Close Dates Only (excl. Omitted)",
        fmt_eur(q2_pipe_arr),
    )
    kpi("Deal Count", deal_count)
    kpi("Avg Deal Size", fmt_eur(avg_deal))
    kpi("Weighted Pipeline (probability-adj)", fmt_eur(weighted))
    kpi("Coverage Ratio", "—")  # placeholder: needs quota target
    kpi("New Pipeline This Quarter (excl. Omitted)", fmt_eur(new_arr))

    current_row += 1

    # — EXECUTION —
    section("EXECUTION")
    won_count = len(won_q)
    won_arr = sum(opp_arr_eur(r) for r in won_q)
    lost_count = len(lost_q)
    lost_arr = sum(opp_arr_eur(r) for r in lost_q)
    total_closed_count = won_count + lost_count
    wr_count = (won_count / total_closed_count * 100) if total_closed_count else 0.0
    total_closed_arr = won_arr + lost_arr
    wr_arr = (won_arr / total_closed_arr * 100) if total_closed_arr else 0.0

    kpi("Won This Quarter (count)", won_count)
    kpi("Won This Quarter (ARR)", fmt_eur(won_arr))
    kpi("Lost This Quarter (count)", lost_count)
    kpi("Lost This Quarter (ARR)", fmt_eur(lost_arr))
    kpi("Win Rate by Count", fmt_pct(wr_count))
    kpi("Win Rate by ARR", fmt_pct(wr_arr))

    current_row += 1

    # — RISK —
    section("RISK")
    stale_30 = [r for r in pipeline if safe_num(r.get("LastActivityInDays")) > 30]
    stale_30_arr = sum(opp_arr_eur(r) for r in stale_30)
    high_stale = [
        r
        for r in pipeline
        if safe_num(r.get("LastActivityInDays")) > 60
        and opp_forecast_arr_eur(r) >= 1_000_000
    ]
    high_stale_arr = sum(opp_forecast_arr_eur(r) for r in high_stale)
    pushed_5 = [r for r in pipeline if safe_num(r.get("PushCount")) >= 5]
    pushed_5_arr = sum(opp_arr_eur(r) for r in pushed_5)
    overdue = [
        r
        for r in pipeline
        if parse_date(r.get("CloseDate")) and parse_date(r.get("CloseDate")) < TODAY
    ]
    overdue_arr = sum(opp_arr_eur(r) for r in overdue)
    aging_365 = [r for r in pipeline if safe_num(r.get("AgeInDays")) > 365]
    aging_365_arr = sum(opp_arr_eur(r) for r in aging_365)

    kpi("Stale 30d+ (count)", len(stale_30))
    kpi("Stale 30d+ (ARR)", fmt_eur(stale_30_arr))
    kpi("High-Value Stale 60d / €1M+ (count)", len(high_stale))
    kpi("High-Value Stale 60d / €1M+ (ARR)", fmt_eur(high_stale_arr))
    kpi("Pushed 5+ (count)", len(pushed_5))
    kpi("Pushed 5+ (ARR)", fmt_eur(pushed_5_arr))
    kpi("Overdue Close (count)", len(overdue))
    kpi("Overdue Close (ARR)", fmt_eur(overdue_arr))
    kpi("Aging 365+ (count)", len(aging_365))
    kpi("Aging 365+ (ARR)", fmt_eur(aging_365_arr))

    current_row += 1

    # — PROCESS COMPLIANCE —
    section("PROCESS COMPLIANCE")

    # Stage >= "3" means stage number prefix >= 3
    def stage_ge3(r):
        stage = safe_str(r.get("StageName"))
        try:
            prefix = stage.split(" ")[0].replace("-", "").strip()
            return float(prefix) >= 3
        except (ValueError, IndexError):
            return False

    at_stage3_plus = [r for r in pipeline if stage_ge3(r)]
    approved = [r for r in at_stage3_plus if r.get("Stage_20_Approval__c") is True]
    approval_rate = (
        (len(approved) / len(at_stage3_plus) * 100) if at_stage3_plus else 0.0
    )

    missing_approval = [
        r
        for r in pipeline
        if r.get("Type") == "Land"
        and stage_ge3(r)
        and not r.get("Stage_20_Approval__c")
    ]

    kpi("Approval Rate (stage 3+)", fmt_pct(approval_rate))
    kpi("Missing Approval (Land, stage 3+)", len(missing_approval))

    current_row += 1

    # — CRO FORECAST TIE-OUT —
    # Sums OwnerOnlyAmount across the Q2 ForecastingItem result set so the
    # totals match the live CRO forecast page exactly. ForecastAmount cannot
    # be summed because each opportunity appears at every level of the
    # forecast hierarchy (rep -> manager -> regional -> CRO).
    section("CRO FORECAST TIE-OUT (Q2 FY2026, ARR, org-wide, EUR converted)")

    fi_q2 = cache.forecast_items("Q2_2026", "ARR")
    cro_totals = {"Pipeline": 0.0, "Best Case": 0.0, "Commit": 0.0, "Closed": 0.0}
    for item in fi_q2:
        cat = (
            item.get("ForecastCategoryName")
            or item.get("ForecastingItemCategory")
            or ""
        ).strip()
        # Normalise hierarchy variants like "BestCaseOnly" / "PipelineOnly"
        if cat in cro_totals:
            cro_totals[cat] += forecast_owner_only_eur(item)
        else:
            lc = cat.lower()
            if "pipeline" in lc:
                cro_totals["Pipeline"] += forecast_owner_only_eur(item)
            elif "bestcase" in lc or "best case" in lc:
                cro_totals["Best Case"] += forecast_owner_only_eur(item)
            elif "commit" in lc:
                cro_totals["Commit"] += forecast_owner_only_eur(item)
            elif "closed" in lc:
                cro_totals["Closed"] += forecast_owner_only_eur(item)

    open_total = cro_totals["Pipeline"] + cro_totals["Best Case"] + cro_totals["Commit"]

    kpi("Pipeline (OwnerOnly sum)", fmt_eur(cro_totals["Pipeline"]))
    kpi("Best Case (OwnerOnly sum)", fmt_eur(cro_totals["Best Case"]))
    kpi("Commit (OwnerOnly sum)", fmt_eur(cro_totals["Commit"]))
    kpi("Closed (OwnerOnly sum)", fmt_eur(cro_totals["Closed"]))
    kpi("Open Total (Pipeline + Best Case + Commit)", fmt_eur(open_total))

    ws.cell(
        row=current_row,
        column=1,
        value="Reconciles to: CRO forecast page (forecastingOwnerId=0055700000747OgAAI)",
    ).font = _data_font(size=9, color="666666")
    ws.merge_cells(f"A{current_row}:B{current_row}")
    current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22


# ---------------------------------------------------------------------------
# Tab 2: Pipeline Detail
# ---------------------------------------------------------------------------

PIPELINE_HEADERS = [
    "Account",
    "Opportunity",
    "Owner",
    "Stage",
    "Close Date",
    "ARR (€ converted)",
    "ACV (€ converted)",
    "Forecast ARR (€ converted)",
    "Forecast Category",
    "Probability (%)",
    "Type",
    "Sub-Type",
    "Push Count",
    "Age (Days)",
    "Days In Stage",
    "Last Activity",
    "Activity Days Ago",
    "Risk Level",
    "Approval",
    "Approval Status",
    "Next Step",
    "Territory Scope",
    "Region",
    "Industry",
]


def build_pipeline_detail(ws, cache: DirectorCache):
    ws.title = "Pipeline Detail"
    pipeline = active_pipeline_rows(cache.open_pipeline)

    _write_header_row(ws, 1, PIPELINE_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PIPELINE_HEADERS))}1"

    n_cols = len(PIPELINE_HEADERS)
    for row_idx, r in enumerate(pipeline, start=2):
        arr = opp_arr_eur(r)
        activity_days = safe_num(r.get("LastActivityInDays"))
        close_date = parse_date(r.get("CloseDate"))
        push_count = safe_num(r.get("PushCount"))

        row_data = [
            nested_get(r, "Account", "Name", default=""),
            safe_str(r.get("Name")),
            nested_get(r, "Owner", "Name", default=""),
            safe_str(r.get("StageName")),
            safe_str(r.get("CloseDate", ""))[:10],
            arr,
            opp_acv_eur(r),
            opp_forecast_arr_eur(r),
            safe_str(r.get("ForecastCategoryName")),
            safe_num(r.get("Probability")),
            safe_str(r.get("Type")),
            safe_str(r.get("APTS_Opportunity_Sub_Type__c")),
            int(push_count),
            int(safe_num(r.get("AgeInDays"))),
            int(safe_num(r.get("LastStageChangeInDays"))),
            safe_str(r.get("LastActivityDate", ""))[:10],
            int(activity_days),
            safe_str(r.get("Risk_Assessment_Level__c")),
            "Yes" if r.get("Stage_20_Approval__c") else "No",
            safe_str(r.get("Approval_Status__c")),
            safe_str(r.get("NextStep"), max_len=200),
            safe_str(r.get("Account_Unit__c") or r.get("Account_Unit_Group__c")),
            safe_str(r.get("Sales_Region__c")),
            nested_get(r, "Account", "Industry", default=""),
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _data_font()

        # Conditional formatting
        is_red = (
            push_count >= 5
            or (activity_days > 60 and arr >= 1_000_000)
            or (close_date is not None and close_date < TODAY)
        )
        is_amber = activity_days > 30

        if is_red:
            _apply_row_fill(ws, row_idx, n_cols, RED_FILL)
        elif is_amber:
            _apply_row_fill(ws, row_idx, n_cols, AMBER_FILL)

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 3: Q1 Review
# ---------------------------------------------------------------------------


def build_q1_review(ws, cache: DirectorCache, territory: str):
    ws.title = "Q1 Review"

    row = 1

    def title(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _title_font()
        row += 1

    def subheader(text, cols=6):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _section_font()
        ws.cell(row=row, column=1).fill = _make_fill("E8F4F8")
        row += 1

    def blank():
        nonlocal row
        row += 1

    title(f"Q1 2026 Review — {territory}")
    blank()

    # — Section A: Forecast vs Actual (org-wide, OwnerOnly to avoid hierarchy double-count) —
    subheader("A  Q1 2026 Forecast vs Actual (ARR, org-wide, EUR converted)")
    ws.cell(
        row=row,
        column=1,
        value=(
            "Note: Org-wide ForecastingItem totals. To match the CRO forecast page exactly, "
            "values use OwnerOnlyAmount to avoid hierarchical double-counting."
        ),
    ).font = _data_font(size=9, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    headers_a = ["Category", "Forecast Amount (€, OwnerOnly)"]
    _write_header_row(ws, row, headers_a)
    row += 1

    fi = cache.forecast_items("Q1_2026", "ARR")
    CATEGORIES = ["Commit", "BestCase", "Pipeline", "Closed"]
    totals = {cat: 0.0 for cat in CATEGORIES}
    for item in fi:
        cat = item.get("ForecastCategoryName") or item.get(
            "ForecastingItemCategory", ""
        )
        for c in CATEGORIES:
            if c.lower() in cat.lower():
                totals[c] += forecast_owner_only_eur(item)
                break

    for cat in CATEGORIES:
        ws.cell(row=row, column=1, value=cat).font = _data_font()
        ws.cell(row=row, column=2, value=round(totals[cat], 2)).font = _data_font()
        row += 1

    # Won / Lost actuals
    won_q1 = cache.won_q1
    lost_q1 = cache.lost_q1
    won_arr = sum(opp_arr_eur(r) for r in won_q1)
    lost_arr = sum(opp_arr_eur(r) for r in lost_q1)
    blank()
    ws.cell(row=row, column=1, value="Won Q1 (count)").font = _data_font(bold=True)
    ws.cell(row=row, column=2, value=len(won_q1)).font = _data_font()
    row += 1
    ws.cell(row=row, column=1, value="Won Q1 (ARR € converted)").font = _data_font(
        bold=True
    )
    ws.cell(row=row, column=2, value=round(won_arr, 2)).font = _data_font()
    row += 1
    ws.cell(row=row, column=1, value="Lost Q1 (count)").font = _data_font(bold=True)
    ws.cell(row=row, column=2, value=len(lost_q1)).font = _data_font()
    row += 1
    ws.cell(row=row, column=1, value="Lost Q1 (ARR € converted)").font = _data_font(
        bold=True
    )
    ws.cell(row=row, column=2, value=round(lost_arr, 2)).font = _data_font()
    row += 1
    blank()

    # — Section B: Deals Pushed Out of Q1 —
    subheader("B  Deals Pushed Out of Q1 (CloseDate moved from Q1 → Q2+)")
    headers_b = [
        "Account",
        "Opportunity",
        "Owner",
        "ARR (€ converted)",
        "Old Close",
        "New Close",
        "Stage",
    ]
    _write_header_row(ws, row, headers_b)
    row += 1

    fh_cd = cache.field_history("CloseDate")
    q1_start = date(2026, 1, 1)
    q1_end = date(2026, 3, 31)
    q2_start = date(2026, 4, 1)

    pushed_out = []
    seen_ids = set()
    for fh in fh_cd:
        old_date = parse_date(fh.get("OldValue"))
        new_date = parse_date(fh.get("NewValue"))
        if not old_date or not new_date:
            continue
        if not (q1_start <= old_date <= q1_end):
            continue
        if new_date < q2_start:
            continue
        opp_id = fh.get("OpportunityId", "")
        if opp_id in seen_ids:
            continue
        seen_ids.add(opp_id)
        pushed_out.append(fh)

    for fh in pushed_out:
        opp = fh.get("Opportunity") or {}
        arr = opp_arr_eur(opp)
        ws.cell(
            row=row, column=1, value=nested_get(opp, "Account", "Name", default="")
        ).font = _data_font()
        ws.cell(row=row, column=2, value=safe_str(opp.get("Name"))).font = _data_font()
        ws.cell(
            row=row, column=3, value=nested_get(opp, "Owner", "Name", default="")
        ).font = _data_font()
        ws.cell(row=row, column=4, value=round(arr, 2)).font = _data_font()
        ws.cell(
            row=row, column=5, value=safe_str(fh.get("OldValue", ""))[:10]
        ).font = _data_font()
        ws.cell(
            row=row, column=6, value=safe_str(fh.get("NewValue", ""))[:10]
        ).font = _data_font()
        ws.cell(
            row=row, column=7, value=safe_str(opp.get("StageName"))
        ).font = _data_font()
        row += 1

    if not pushed_out:
        ws.cell(
            row=row, column=1, value="No deals pushed out of Q1 found."
        ).font = _data_font(color="888888")
        row += 1

    blank()

    # — Section C: Forecast Category Movement —
    subheader("C  Forecast Category Movement")
    headers_c = [
        "Opportunity",
        "Owner",
        "ARR (€ converted)",
        "Old Category",
        "New Category",
        "Date",
    ]
    _write_header_row(ws, row, headers_c)
    row += 1

    fh_fcn = cache.field_history("ForecastCategoryName")
    for fh in fh_fcn:
        opp = fh.get("Opportunity") or {}
        arr = opp_arr_eur(opp)
        ws.cell(row=row, column=1, value=safe_str(opp.get("Name"))).font = _data_font()
        ws.cell(
            row=row, column=2, value=nested_get(opp, "Owner", "Name", default="")
        ).font = _data_font()
        ws.cell(row=row, column=3, value=round(arr, 2)).font = _data_font()
        ws.cell(
            row=row, column=4, value=safe_str(fh.get("OldValue"))
        ).font = _data_font()
        ws.cell(
            row=row, column=5, value=safe_str(fh.get("NewValue"))
        ).font = _data_font()
        ws.cell(
            row=row, column=6, value=safe_str(fh.get("CreatedDate", ""))[:10]
        ).font = _data_font()
        row += 1

    if not fh_fcn:
        ws.cell(
            row=row, column=1, value="No forecast category movements found."
        ).font = _data_font(color="888888")

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 4: Rep Performance
# ---------------------------------------------------------------------------

REP_HEADERS = [
    "Rep",
    "Open Pipeline ARR (€ converted)",
    "Deal Count",
    "Avg Deal Size (€)",
    "Won ARR Q (€ converted)",
    "Lost ARR Q (€ converted)",
    "Win Rate %",
    "Stale Deals",
    "Pushed Deals",
    "Missing Approvals",
]


def build_rep_performance(ws, cache: DirectorCache):
    ws.title = "Rep Performance"

    pipeline = active_pipeline_rows(cache.open_pipeline)
    won_q = cache.won_this_quarter
    lost_q = cache.lost_this_quarter

    reps: dict[str, dict] = {}

    def get_rep(name: str) -> dict:
        if name not in reps:
            reps[name] = {
                "pipeline_arr": 0.0,
                "deal_count": 0,
                "won_arr": 0.0,
                "lost_arr": 0.0,
                "stale": 0,
                "pushed": 0,
                "missing_approval": 0,
            }
        return reps[name]

    def stage_ge3(r):
        stage = safe_str(r.get("StageName"))
        try:
            prefix = stage.split(" ")[0].replace("-", "").strip()
            return float(prefix) >= 3
        except (ValueError, IndexError):
            return False

    for r in pipeline:
        rep = nested_get(r, "Owner", "Name", default="Unknown")
        d = get_rep(rep)
        arr = opp_arr_eur(r)
        d["pipeline_arr"] += arr
        d["deal_count"] += 1
        if safe_num(r.get("LastActivityInDays")) > 30:
            d["stale"] += 1
        if safe_num(r.get("PushCount")) >= 5:
            d["pushed"] += 1
        if (
            r.get("Type") == "Land"
            and stage_ge3(r)
            and not r.get("Stage_20_Approval__c")
        ):
            d["missing_approval"] += 1

    for r in won_q:
        rep = nested_get(r, "Owner", "Name", default="Unknown")
        get_rep(rep)["won_arr"] += opp_arr_eur(r)

    for r in lost_q:
        rep = nested_get(r, "Owner", "Name", default="Unknown")
        get_rep(rep)["lost_arr"] += opp_arr_eur(r)

    sorted_reps = sorted(reps.items(), key=lambda x: x[1]["pipeline_arr"], reverse=True)

    _write_header_row(ws, 1, REP_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REP_HEADERS))}1"

    for row_idx, (rep_name, d) in enumerate(sorted_reps, start=2):
        total_closed = d["won_arr"] + d["lost_arr"]
        wr = (d["won_arr"] / total_closed * 100) if total_closed else 0.0
        avg = (d["pipeline_arr"] / d["deal_count"]) if d["deal_count"] else 0.0

        row_data = [
            rep_name,
            round(d["pipeline_arr"], 2),
            d["deal_count"],
            round(avg, 2),
            round(d["won_arr"], 2),
            round(d["lost_arr"], 2),
            round(wr, 1),
            d["stale"],
            d["pushed"],
            d["missing_approval"],
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _data_font()

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 5: Won-Lost
# ---------------------------------------------------------------------------

WON_LOST_HEADERS = [
    "Status",
    "Account",
    "Opportunity",
    "Owner",
    "Type",
    "Stage",
    "ARR (€ converted)",
    "ACV (€ converted)",
    "Close Date",
    "Created Date",
    "Sales Cycle (Days)",
    "Reason Won/Lost",
    "Sub-Reason",
    "Competitor",
    "Lost Comments",
]


def build_won_lost(ws, cache: DirectorCache):
    ws.title = "Won-Lost"

    all_won = {r["Id"]: r for r in cache.won_this_quarter}
    all_won.update({r["Id"]: r for r in cache.won_q1})
    all_lost = {r["Id"]: r for r in cache.lost_this_quarter}
    all_lost.update({r["Id"]: r for r in cache.lost_q1})

    _write_header_row(ws, 1, WON_LOST_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(WON_LOST_HEADERS))}1"

    def _write_opp(row_idx: int, r: dict, status: str):
        close_date = parse_date(r.get("CloseDate"))
        created_date = parse_date(
            r.get("CreatedDate", "")[:10] if r.get("CreatedDate") else None
        )
        cycle_days = (
            (close_date - created_date).days if close_date and created_date else None
        )

        row_data = [
            status,
            nested_get(r, "Account", "Name", default=""),
            safe_str(r.get("Name")),
            nested_get(r, "Owner", "Name", default=""),
            safe_str(r.get("Type")),
            safe_str(r.get("StageName")),
            opp_arr_eur(r),
            opp_acv_eur(r),
            safe_str(r.get("CloseDate", ""))[:10],
            safe_str(r.get("CreatedDate", ""))[:10],
            cycle_days,
            safe_str(r.get("Reason_Won_Lost__c")),
            safe_str(r.get("Sub_Reason__c")),
            safe_str(r.get("Lost_to_Competitor__c")),
            safe_str(r.get("Lost_Comments__c"), max_len=200),
        ]
        n_cols = len(WON_LOST_HEADERS)
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _data_font()

        fill_color = GREEN_FILL if status == "Won" else RED_FILL
        _apply_row_fill(ws, row_idx, n_cols, fill_color)

    current_row = 2
    for r in all_won.values():
        _write_opp(current_row, r, "Won")
        current_row += 1

    for r in all_lost.values():
        _write_opp(current_row, r, "Lost")
        current_row += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 6: Sources & Lineage
# ---------------------------------------------------------------------------

SOURCES_HEADERS = [
    "Source ID",
    "Source Type",
    "Name",
    "Query / Endpoint",
    "Record Count",
    "Extracted At",
]


def build_sources(ws, cache: DirectorCache):
    ws.title = "Sources & Lineage"

    sources = cache.sources
    _write_header_row(ws, 1, SOURCES_HEADERS)

    for row_idx, s in enumerate(sources, start=2):
        row_data = [
            safe_str(s.get("source_id")),
            safe_str(s.get("source_type")),
            safe_str(s.get("name")),
            safe_str(s.get("query_or_endpoint"), max_len=500),
            s.get("record_count", 0),
            safe_str(s.get("extracted_at")),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _data_font()

    # Wide query column
    ws.column_dimensions["D"].width = 60
    _auto_width(ws)
    ws.column_dimensions["D"].width = 60  # re-apply after auto


# ---------------------------------------------------------------------------
# Tab 7: Q2 Outlook
# ---------------------------------------------------------------------------

Q2_START = date(2026, 4, 1)
Q2_END = date(2026, 6, 30)


def build_q2_outlook(ws, cache: DirectorCache, territory: str = ""):
    ws.title = "Q2 Outlook"

    row = 1

    def title(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _title_font()
        row += 1

    def subheader(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _section_font()
        ws.cell(row=row, column=1).fill = _make_fill("E8F4F8")
        row += 1

    def blank():
        nonlocal row
        row += 1

    title(f"Q2 2026 Outlook — {territory}")
    ws.cell(
        row=row,
        column=1,
        value=(
            "Q2 2026 = April 1 – June 30, 2026. "
            "Excludes opportunities in Omitted forecast category."
        ),
    ).font = _data_font(size=9, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    blank()

    # --- Section A: Q2 Forecast Breakdown ---
    subheader("A  Q2 Forecast Breakdown")
    headers_a = [
        "Forecast Category",
        "Deal Count",
        "ARR (€ converted)",
        "ACV (€ converted)",
    ]
    _write_header_row(ws, row, headers_a)
    row += 1

    forecast_cats = cache.forecast_categories
    for fc in forecast_cats:
        ws.cell(
            row=row, column=1, value=safe_str(fc.get("ForecastCategoryName"))
        ).font = _data_font()
        ws.cell(
            row=row, column=2, value=int(safe_num(fc.get("ct")))
        ).font = _data_font()
        ws.cell(
            row=row, column=3, value=round(safe_num(fc.get("arr")), 2)
        ).font = _data_font()
        ws.cell(
            row=row, column=4, value=round(safe_num(fc.get("acv")), 2)
        ).font = _data_font()
        row += 1

    if not forecast_cats:
        ws.cell(
            row=row, column=1, value="No forecast category data found."
        ).font = _data_font(color="888888")
        row += 1

    blank()

    pipeline = active_pipeline_rows(cache.open_pipeline)

    def is_q2_close(r):
        cd = parse_date(r.get("CloseDate"))
        return cd is not None and Q2_START <= cd <= Q2_END

    commit_deals = [
        r
        for r in pipeline
        if r.get("ForecastCategoryName") == "Commit" and is_q2_close(r)
    ]
    bestcase_deals = [
        r
        for r in pipeline
        if r.get("ForecastCategoryName") == "Best Case" and is_q2_close(r)
    ]

    deal_headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR (€ converted)",
        "Probability (%)",
    ]

    def write_deals(deals):
        nonlocal row
        _write_header_row(ws, row, deal_headers)
        row += 1
        if not deals:
            ws.cell(row=row, column=1, value="No deals found.").font = _data_font(
                color="888888"
            )
            row += 1
            return
        for r in sorted(
            deals,
            key=lambda x: opp_arr_eur(x),
            reverse=True,
        ):
            arr = opp_arr_eur(r)
            row_data = [
                nested_get(r, "Account", "Name", default=""),
                safe_str(r.get("Name")),
                nested_get(r, "Owner", "Name", default=""),
                safe_str(r.get("StageName")),
                safe_str(r.get("CloseDate", ""))[:10],
                round(arr, 2),
                safe_num(r.get("Probability")),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row, column=col_idx, value=val).font = _data_font()
            row += 1

    # --- Section B: Commit Deals ---
    subheader("B  Commit Deals (Q2 Close Date)")
    write_deals(commit_deals)
    blank()

    # --- Section C: Best Case Deals ---
    subheader("C  Best Case Deals (Q2 Close Date)")
    write_deals(bestcase_deals)
    blank()

    # --- Section D: Pipeline Coverage ---
    subheader("D  Pipeline Coverage")
    q2_active_pipeline = [r for r in pipeline if is_q2_close(r)]
    total_pipe_arr = sum(opp_arr_eur(r) for r in q2_active_pipeline)
    commit_arr = sum(opp_arr_eur(r) for r in commit_deals)
    bestcase_arr = sum(opp_arr_eur(r) for r in bestcase_deals)

    coverage_headers = ["Metric", "Value"]
    _write_header_row(ws, row, coverage_headers)
    row += 1
    for label, val in [
        ("Active Pipeline ARR (Q2 close, excl. Omitted)", fmt_eur(total_pipe_arr)),
        ("Commit ARR (Q2 close)", fmt_eur(commit_arr)),
        ("Best Case ARR (Q2 close)", fmt_eur(bestcase_arr)),
        ("Gap to Quota", "Awaiting quota targets"),
    ]:
        ws.cell(row=row, column=1, value=label).font = _data_font(bold=True)
        ws.cell(row=row, column=2, value=val).font = _data_font()
        row += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 8: Commercial Approval
# ---------------------------------------------------------------------------


def _stage_num(r) -> float:
    """Extract numeric prefix from StageName, returns -1 on failure."""
    stage = safe_str(r.get("StageName"))
    try:
        prefix = stage.split(" ")[0].replace("-", "").strip()
        return float(prefix)
    except (ValueError, IndexError):
        return -1.0


def build_commercial_approval(ws, cache: DirectorCache, territory: str = ""):
    ws.title = "Commercial Approval"

    row = 1

    def title(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _title_font()
        row += 1

    def subheader(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _section_font()
        ws.cell(row=row, column=1).fill = _make_fill("E8F4F8")
        row += 1

    def blank():
        nonlocal row
        row += 1

    title(f"Commercial Approval — {territory}")
    blank()

    pipeline = active_pipeline_rows(cache.open_pipeline)

    def is_approved(r):
        return r.get("Stage_20_Approval__c") is True

    def approval_not_needed(r):
        sn = _stage_num(r)
        status = safe_str(r.get("Approval_Status__c")).lower()
        return sn < 3 or "no approval" in status

    approved_opps = [r for r in pipeline if is_approved(r)]
    no_approval_opps = [
        r for r in pipeline if approval_not_needed(r) and not is_approved(r)
    ]
    pending_opps = [
        r for r in pipeline if not is_approved(r) and not approval_not_needed(r)
    ]

    # --- Section A: Approval Summary ---
    subheader("A  Approval Summary")
    summary_headers = ["Category", "Deal Count", "ARR (€ converted)"]
    _write_header_row(ws, row, summary_headers)
    row += 1

    for label, opps in [
        ("Approved", approved_opps),
        ("Pending / Missing Approval", pending_opps),
        ("No Approval Needed", no_approval_opps),
    ]:
        arr = sum(opp_arr_eur(r) for r in opps)
        ws.cell(row=row, column=1, value=label).font = _data_font(bold=True)
        ws.cell(row=row, column=2, value=len(opps)).font = _data_font()
        ws.cell(row=row, column=3, value=round(arr, 2)).font = _data_font()
        row += 1

    blank()

    # --- Section B: Missing Approval Candidates ---
    subheader("B  Missing Approval Candidates (Stage 3+, Land, Not Approved)")
    missing_hdrs = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR (€ converted)",
        "Close Date",
        "Next Step",
    ]
    _write_header_row(ws, row, missing_hdrs)
    row += 1

    missing = [
        r
        for r in pipeline
        if _stage_num(r) >= 3 and r.get("Type") == "Land" and not is_approved(r)
    ]
    missing_sorted = sorted(missing, key=lambda r: opp_arr_eur(r), reverse=True)

    if not missing_sorted:
        ws.cell(
            row=row, column=1, value="No missing approval candidates found."
        ).font = _data_font(color="888888")
        row += 1
    else:
        for r in missing_sorted:
            arr = opp_arr_eur(r)
            row_data = [
                nested_get(r, "Account", "Name", default=""),
                safe_str(r.get("Name")),
                nested_get(r, "Owner", "Name", default=""),
                safe_str(r.get("StageName")),
                round(arr, 2),
                safe_str(r.get("CloseDate", ""))[:10],
                safe_str(r.get("NextStep"), max_len=150),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row, column=col_idx, value=val).font = _data_font()
            _apply_row_fill(ws, row, len(missing_hdrs), AMBER_FILL)
            row += 1

    blank()

    # --- Section C: Approved Deals YTD ---
    subheader("C  Approved Deals YTD (2026)")
    approved_hdrs = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR (€ converted)",
        "Approval Date",
    ]
    _write_header_row(ws, row, approved_hdrs)
    row += 1

    approved_ytd = [
        r
        for r in pipeline
        if is_approved(r)
        and parse_date(r.get("Stage_20_Approval_Date__c")) is not None
        and parse_date(r.get("Stage_20_Approval_Date__c")).year == 2026
    ]

    if not approved_ytd:
        ws.cell(
            row=row, column=1, value="No approved deals YTD found."
        ).font = _data_font(color="888888")
        row += 1
    else:
        for r in sorted(
            approved_ytd,
            key=lambda r: opp_arr_eur(r),
            reverse=True,
        ):
            arr = opp_arr_eur(r)
            row_data = [
                nested_get(r, "Account", "Name", default=""),
                safe_str(r.get("Name")),
                nested_get(r, "Owner", "Name", default=""),
                safe_str(r.get("StageName")),
                round(arr, 2),
                safe_str(r.get("Stage_20_Approval_Date__c", ""))[:10],
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row, column=col_idx, value=val).font = _data_font()
            _apply_row_fill(ws, row, len(approved_hdrs), GREEN_FILL)
            row += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 9: Renewals & Retention
# ---------------------------------------------------------------------------


def build_renewals_retention(ws, cache: DirectorCache, territory: str = ""):
    ws.title = "Renewals & Retention"

    row = 1

    def title(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _title_font()
        row += 1

    def subheader(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _section_font()
        ws.cell(row=row, column=1).fill = _make_fill("E8F4F8")
        row += 1

    def blank():
        nonlocal row
        row += 1

    title(f"Renewals & Retention — {territory}")
    blank()

    # --- Section A: Retention KPIs ---
    subheader("A  Retention KPIs (Most Recent Year)")
    retention = cache.load("crma_retention_metrics.json")

    if retention:
        # Use most recent year (last item, or sort by YearLabel desc)
        latest = sorted(
            retention, key=lambda r: safe_str(r.get("YearLabel")), reverse=True
        )[0]
        starting = safe_num(latest.get("StartingARR"))
        renewal_won = safe_num(latest.get("RenewalWonARR"))
        expansion = safe_num(latest.get("ExpansionARR"))
        churn = safe_num(latest.get("ChurnARR"))
        ending = safe_num(latest.get("EndingARR"))
        new_logo = safe_num(latest.get("NewLogoARR"))

        grr = ((starting - churn) / starting * 100) if starting else 0.0
        nrr = ((starting + expansion - churn) / starting * 100) if starting else 0.0

        kpi_headers = ["KPI", "Value"]
        _write_header_row(ws, row, kpi_headers)
        row += 1

        for label, val in [
            ("Year", safe_str(latest.get("YearLabel"))),
            ("Starting ARR", fmt_eur(starting)),
            ("Renewal Won ARR", fmt_eur(renewal_won)),
            ("Expansion ARR", fmt_eur(expansion)),
            ("Churn ARR", fmt_eur(churn)),
            ("Ending ARR", fmt_eur(ending)),
            ("New Logo ARR", fmt_eur(new_logo)),
            ("GRR (Gross Retention Rate)", fmt_pct(grr)),
            ("NRR (Net Retention Rate)", fmt_pct(nrr)),
        ]:
            ws.cell(row=row, column=1, value=label).font = _data_font(bold=True)
            ws.cell(row=row, column=2, value=val).font = _data_font()
            row += 1
    else:
        ws.cell(
            row=row, column=1, value="No retention metrics found."
        ).font = _data_font(color="888888")
        row += 1

    blank()

    pipeline = active_pipeline_rows(cache.open_pipeline)
    open_renewals = [r for r in pipeline if r.get("Type") == "Renewal"]

    # --- Section B: Renewal Pipeline by Risk Level ---
    subheader("B  Renewal Pipeline by Risk Level")
    risk_headers = ["Risk Level", "Deal Count", "ACV (€ converted)"]
    _write_header_row(ws, row, risk_headers)
    row += 1

    renewals_by_risk: dict[str, dict[str, float]] = {}
    for renewal in open_renewals:
        risk_level = safe_str(renewal.get("Risk_Assessment_Level__c")) or "Unspecified"
        bucket = renewals_by_risk.setdefault(risk_level, {"ct": 0, "acv": 0.0})
        bucket["ct"] += 1
        bucket["acv"] += renewal_acv_eur(renewal)

    if renewals_by_risk:
        sorted_risks = sorted(
            renewals_by_risk.items(), key=lambda item: item[1]["acv"], reverse=True
        )
        for risk_level, stats in sorted_risks:
            ct = int(stats["ct"])
            acv = float(stats["acv"])

            fill = None
            if "high" in risk_level.lower():
                fill = RED_FILL
            elif "medium" in risk_level.lower():
                fill = AMBER_FILL
            elif "low" in risk_level.lower():
                fill = GREEN_FILL

            ws.cell(row=row, column=1, value=risk_level).font = _data_font()
            ws.cell(row=row, column=2, value=ct).font = _data_font()
            ws.cell(row=row, column=3, value=round(acv, 2)).font = _data_font()
            if fill:
                _apply_row_fill(ws, row, 3, fill)
            row += 1
    else:
        ws.cell(
            row=row, column=1, value="No renewal risk data found."
        ).font = _data_font(color="888888")
        row += 1

    blank()

    # --- Section C: Open Renewals ---
    subheader("C  Open Renewals")
    renewal_hdrs = [
        "Account",
        "Opportunity",
        "Owner",
        "Renewal ACV (€ converted)",
        "Probability (%)",
        "Close Date",
        "Stage",
    ]
    _write_header_row(ws, row, renewal_hdrs)
    row += 1

    open_renewals_sorted = sorted(
        open_renewals, key=lambda r: parse_date(r.get("CloseDate")) or date.max
    )

    if not open_renewals_sorted:
        ws.cell(row=row, column=1, value="No open renewals found.").font = _data_font(
            color="888888"
        )
        row += 1
    else:
        for r in open_renewals_sorted:
            acv = renewal_acv_eur(r)
            row_data = [
                nested_get(r, "Account", "Name", default=""),
                safe_str(r.get("Name")),
                nested_get(r, "Owner", "Name", default=""),
                round(acv, 2),
                safe_num(r.get("Probability")),
                safe_str(r.get("CloseDate", ""))[:10],
                safe_str(r.get("StageName")),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row, column=col_idx, value=val).font = _data_font()
            row += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 10: Risk Register
# ---------------------------------------------------------------------------


def build_risk_register(ws, cache: DirectorCache, territory: str = ""):
    ws.title = "Risk Register"

    pipeline = active_pipeline_rows(cache.open_pipeline)
    crma_ops = cache.load("crma_pipeline_ops.json")

    # Build CRMA lookup by OpportunityId
    crma_by_id: dict[str, dict] = {}
    for op in crma_ops:
        oid = op.get("OpportunityId")
        if oid:
            crma_by_id[oid] = op

    RISK_HEADERS = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR (€ converted)",
        "Close Date",
        "Push Count",
        "Activity Days Ago",
        "Age (Days)",
        "Risk Score",
        "Risk Flags",
        "Days In Stage",
        "Backward Moves",
        "Stale Count",
        "Past Due Count",
    ]

    _write_header_row(ws, 1, RISK_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RISK_HEADERS))}1"

    scored_rows = []
    for r in pipeline:
        opp_id = r.get("Id", "")
        crma = crma_by_id.get(opp_id, {})

        push_count = safe_num(r.get("PushCount"))
        activity_days = safe_num(r.get("LastActivityInDays"))
        arr = opp_arr_eur(r)
        close_date = parse_date(r.get("CloseDate"))
        age_days = safe_num(r.get("AgeInDays"))
        prob = safe_num(r.get("Probability"))
        stage_n = _stage_num(r)
        last_activity_date = r.get("LastActivityDate")
        opp_type = safe_str(r.get("Type"))
        approved = r.get("Stage_20_Approval__c") is True
        backward_moves = safe_num(crma.get("BackwardMoveCount", 0))

        # Determine current quarter end
        q_end = date(2026, 6, 30)

        score = 0
        flags = []

        if push_count >= 5:
            score += 3
            flags.append("PushCount≥5")
        if activity_days > 60 and arr >= 1_000_000:
            score += 3
            flags.append("HighValStale60d")
        if close_date is not None and close_date < TODAY:
            score += 2
            flags.append("Overdue")
        if activity_days > 30:
            score += 1
            flags.append("Stale30d")
        if age_days > 365:
            score += 2
            flags.append("Aging365+")
        if not last_activity_date:
            score += 3
            flags.append("NoActivity")
        if stage_n >= 3 and not approved and opp_type == "Land":
            score += 1
            flags.append("MissingApproval")
        if (
            prob < 50
            and close_date is not None
            and close_date <= q_end
            and close_date >= Q2_START
        ):
            score += 1
            flags.append("LowProbThisQ")
        if backward_moves > 0:
            score += 2
            flags.append("BackwardMove")

        if score <= 0:
            continue

        scored_rows.append((score, r, crma, flags))

    scored_rows.sort(key=lambda x: x[0], reverse=True)

    n_cols = len(RISK_HEADERS)
    for row_idx, (score, r, crma, flags) in enumerate(scored_rows, start=2):
        arr = opp_arr_eur(r)
        row_data = [
            nested_get(r, "Account", "Name", default=""),
            safe_str(r.get("Name")),
            nested_get(r, "Owner", "Name", default=""),
            safe_str(r.get("StageName")),
            round(arr, 2),
            safe_str(r.get("CloseDate", ""))[:10],
            int(safe_num(r.get("PushCount"))),
            int(safe_num(r.get("LastActivityInDays"))),
            int(safe_num(r.get("AgeInDays"))),
            score,
            ", ".join(flags),
            int(safe_num(crma.get("DaysInCurrentStage", 0))),
            int(safe_num(crma.get("BackwardMoveCount", 0))),
            int(safe_num(crma.get("StaleCount", 0))),
            int(safe_num(crma.get("PastDueCount", 0))),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _data_font()

        if score >= 6:
            _apply_row_fill(ws, row_idx, n_cols, RED_FILL)
        elif score >= 3:
            _apply_row_fill(ws, row_idx, n_cols, AMBER_FILL)

    if not scored_rows:
        ws.cell(row=2, column=1, value="No risky deals found.").font = _data_font(
            color="888888"
        )

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Tab 11: Data Quality
# ---------------------------------------------------------------------------


def build_data_quality(ws, cache: DirectorCache, territory: str = ""):
    ws.title = "Data Quality"

    pipeline = active_pipeline_rows(cache.open_pipeline)

    DQ_HEADERS = [
        "Rep",
        "Stale 30d",
        "No Activity",
        "Overdue Close",
        "Missing Amount",
        "Missing Next Step",
        "Missing Approval",
        "Aging 365+",
        "Total Issues",
    ]

    reps: dict[str, dict] = {}

    def get_rep(name: str) -> dict:
        if name not in reps:
            reps[name] = {
                "stale_30": 0,
                "no_activity": 0,
                "overdue_close": 0,
                "missing_amount": 0,
                "missing_next_step": 0,
                "missing_approval": 0,
                "aging_365": 0,
            }
        return reps[name]

    for r in pipeline:
        rep = nested_get(r, "Owner", "Name", default="Unknown")
        d = get_rep(rep)

        stage_n = _stage_num(r)
        arr = opp_arr_eur(r)
        activity_days = safe_num(r.get("LastActivityInDays"))
        last_activity_date = r.get("LastActivityDate")
        close_date = parse_date(r.get("CloseDate"))
        next_step = r.get("NextStep")
        approved = r.get("Stage_20_Approval__c") is True
        opp_type = safe_str(r.get("Type"))
        age_days = safe_num(r.get("AgeInDays"))

        if activity_days > 30:
            d["stale_30"] += 1
        if not last_activity_date:
            d["no_activity"] += 1
        if close_date is not None and close_date < TODAY:
            d["overdue_close"] += 1
        if arr is None or arr == 0:
            d["missing_amount"] += 1
        if (not next_step or str(next_step).strip() == "") and 3 <= stage_n <= 5:
            d["missing_next_step"] += 1
        if stage_n >= 3 and opp_type == "Land" and not approved:
            d["missing_approval"] += 1
        if age_days > 365:
            d["aging_365"] += 1

    _write_header_row(ws, 1, DQ_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DQ_HEADERS))}1"

    totals = {
        k: 0
        for k in [
            "stale_30",
            "no_activity",
            "overdue_close",
            "missing_amount",
            "missing_next_step",
            "missing_approval",
            "aging_365",
        ]
    }

    sorted_reps = sorted(reps.items(), key=lambda x: sum(x[1].values()), reverse=True)
    n_cols = len(DQ_HEADERS)

    for row_idx, (rep_name, d) in enumerate(sorted_reps, start=2):
        total_issues = sum(d.values())
        row_data = [
            rep_name,
            d["stale_30"],
            d["no_activity"],
            d["overdue_close"],
            d["missing_amount"],
            d["missing_next_step"],
            d["missing_approval"],
            d["aging_365"],
            total_issues,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _data_font()

        if total_issues >= 5:
            _apply_row_fill(ws, row_idx, n_cols, RED_FILL)
        elif total_issues >= 2:
            _apply_row_fill(ws, row_idx, n_cols, AMBER_FILL)

        for k in totals:
            totals[k] += d[k]

    # TOTAL row
    total_row = len(sorted_reps) + 2
    grand_total = sum(totals.values())
    total_data = [
        "TOTAL",
        totals["stale_30"],
        totals["no_activity"],
        totals["overdue_close"],
        totals["missing_amount"],
        totals["missing_next_step"],
        totals["missing_approval"],
        totals["aging_365"],
        grand_total,
    ]
    for col_idx, val in enumerate(total_data, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = _data_font(bold=True)
    _apply_row_fill(ws, total_row, n_cols, "E8F4F8")

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Placeholder tabs
# ---------------------------------------------------------------------------

PLACEHOLDERS = [
    (
        "Quota & Targets",
        "Placeholder — populated when Finance delivers regional targets",
    ),
]


def build_placeholder(ws, tab_name: str, message: str):
    ws.title = tab_name
    ws["A1"].value = tab_name
    ws["A1"].font = _title_font()
    ws["A3"].value = message
    ws["A3"].font = _data_font(color="888888")


# ---------------------------------------------------------------------------
# Main workbook assembler
# ---------------------------------------------------------------------------


def build_workbook(director_name: str, territory: str, cache_dir: Path, out_path: Path):
    cache = DirectorCache(cache_dir)

    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    print(f"  Building: {director_name} ({territory})")

    print("    Tab 1: Scorecard")
    ws1 = wb.create_sheet("Scorecard")
    build_scorecard(ws1, cache, director_name, territory)

    print("    Tab 2: Pipeline Detail")
    ws2 = wb.create_sheet("Pipeline Detail")
    build_pipeline_detail(ws2, cache)

    print("    Tab 3: Q1 Review")
    ws3 = wb.create_sheet("Q1 Review")
    build_q1_review(ws3, cache, territory)

    print("    Tab 4: Rep Performance")
    ws4 = wb.create_sheet("Rep Performance")
    build_rep_performance(ws4, cache)

    print("    Tab 5: Won-Lost")
    ws5 = wb.create_sheet("Won-Lost")
    build_won_lost(ws5, cache)

    print("    Tab 6: Sources & Lineage")
    ws6 = wb.create_sheet("Sources & Lineage")
    build_sources(ws6, cache)

    print("    Tab 7: Q2 Outlook")
    ws7 = wb.create_sheet("Q2 Outlook")
    build_q2_outlook(ws7, cache, territory)

    print("    Tab 8: Commercial Approval")
    ws8 = wb.create_sheet("Commercial Approval")
    build_commercial_approval(ws8, cache, territory)

    print("    Tab 9: Renewals & Retention")
    ws9 = wb.create_sheet("Renewals & Retention")
    build_renewals_retention(ws9, cache, territory)

    print("    Tab 10: Risk Register")
    ws10 = wb.create_sheet("Risk Register")
    build_risk_register(ws10, cache, territory)

    print("    Tab 11: Data Quality")
    ws11 = wb.create_sheet("Data Quality")
    build_data_quality(ws11, cache, territory)

    print("    Tab 12: Quota & Targets (placeholder)")
    for tab_name, message in PLACEHOLDERS:
        ws_ph = wb.create_sheet(tab_name)
        build_placeholder(ws_ph, tab_name, message)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"  Saved: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Build Sales Director Excel workbooks from cache."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--director", metavar="NAME", help='Director name e.g. "Dan Peppett"'
    )
    group.add_argument(
        "--all", action="store_true", help="Build for all directors in preset config"
    )
    parser.add_argument(
        "--snapshot-date",
        metavar="YYYY-MM-DD",
        default=TODAY.isoformat(),
        help="Snapshot date (default: today)",
    )
    args = parser.parse_args()

    snapshot_date = args.snapshot_date
    dump_dir = OUTPUT_BASE / snapshot_date
    cache_base = dump_dir / ".cache"

    config = load_md1_preset_config(REPO_ROOT / h.CONFIG_PATH)

    if args.all:
        presets = list(config.presets)
    else:
        preset = find_md1_preset(config, args.director)
        if preset is None:
            print(
                f"ERROR: Director '{args.director}' not found in preset config.",
                file=sys.stderr,
            )
            print(
                "Available directors:",
                [p.name for p in config.presets],
                file=sys.stderr,
            )
            sys.exit(1)
        presets = [preset]

    built = []
    for preset in presets:
        slug = h.slugify(preset.name)
        cache_dir = cache_base / slug
        if not cache_dir.exists():
            print(f"  SKIP: cache dir not found: {cache_dir}", file=sys.stderr)
            continue

        safe_name = preset.name.replace("/", "-").replace("\\", "-")
        out_filename = f"Sales Director Data - {safe_name} ({preset.territory}).xlsx"
        out_path = dump_dir / out_filename

        build_workbook(preset.name, preset.territory, cache_dir, out_path)
        built.append(out_path)

    print(f"\nDone. Built {len(built)} workbook(s).")
    for p in built:
        print(f"  {p}")


if __name__ == "__main__":
    main()
