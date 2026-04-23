"""
Pull the Historical Trending reports and append retrospective/current-quarter
snapshot sheets to each director's workbook.

Each report holds 4 snapshots with per-deal ARR and stage at each date,
plus Change columns. This is the authoritative SF view of what moved
quarter-to-quarter, higher fidelity than our OpportunityFieldHistory
reconstruction.

Output per director workbook:
  - Sheet "<prior quarter> Snapshot Trend"
  - Sheet "<current quarter> Snapshot Trend"

Re-runs replace those two sheets. Other sheets are untouched.
"""

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from monthly_platform.period import resolve_period_context
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.period import resolve_period_context

try:
    from monthly_platform.historical_trending import resolve_historical_trending_contract
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.historical_trending import (
        resolve_historical_trending_contract,
    )


ROOT = Path(__file__).resolve().parents[1]
TERRITORY_CONFIG_PATH = ROOT / "config" / "sd_monthly_territories.json"
AUDIT_OUTPUT_ROOT = ROOT / "output" / "historical_trending_extract"
SOURCE_CONTRACT_AUDIT_ROOT = ROOT / "output" / "source_contract_audit"


def _slugify(value: str) -> str:
    return str(value).strip().lower().replace(" ", "-")


def _load_report_registry() -> dict[str, dict[str, str]]:
    payload = json.loads(TERRITORY_CONFIG_PATH.read_text())
    territories = payload.get("territories") or {}
    registry: dict[str, dict[str, str]] = {}
    for territory, config in territories.items():
        director = str(config.get("director") or "").strip()
        quarter_ids: dict[str, str] = {}
        for quarter_label, report_id in (
            config.get("historical_trending_report_ids") or {}
        ).items():
            report_id = str(report_id or "").strip()
            if report_id:
                quarter_ids[str(quarter_label).strip().upper()] = report_id
        for quarter_label, report_id in (
            config.get("forward_quarter_historical_trending_report_ids") or {}
        ).items():
            report_id = str(report_id or "").strip()
            if report_id:
                quarter_ids[str(quarter_label).strip().upper()] = report_id
        if not director or not quarter_ids.get("Q1") or not quarter_ids.get("Q2"):
            raise ValueError(
                f"{TERRITORY_CONFIG_PATH}: territory {territory!r} is missing "
                "director or Q1/Q2 historical_trending_report_ids"
            )
        registry[_slugify(director)] = quarter_ids
    return registry


def _load_director_territory_map() -> dict[str, str]:
    payload = json.loads(TERRITORY_CONFIG_PATH.read_text())
    territories = payload.get("territories") or {}
    mapping: dict[str, str] = {}
    for territory, config in territories.items():
        director = str(config.get("director") or "").strip()
        if director:
            mapping[_slugify(director)] = str(territory)
    return mapping


# director workbook slug -> quarter label -> report id
QUARTER_REPORTS = _load_report_registry()
DIRECTOR_SLUG_TO_TERRITORY = _load_director_territory_map()

# Backward-compatible legacy shape used by existing tests/audits.
REPORTS = {
    slug: (quarter_ids["Q1"], quarter_ids["Q2"])
    for slug, quarter_ids in QUARTER_REPORTS.items()
}


HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=9)
EUR_FMT = "#,##0"


def _infer_report_date_from_workbooks_dir(workbooks_dir: Path | None) -> str | None:
    if workbooks_dir is None:
        return None
    path = Path(workbooks_dir).resolve()
    for candidate in [path, *path.parents]:
        try:
            datetime.strptime(candidate.name, "%Y-%m-%d")
            return candidate.name
        except ValueError:
            continue
    return None


def _resolve_runtime_context(
    *,
    snapshot_date: str | None = None,
    workbooks_dir: Path | None = None,
) -> dict[str, str | int]:
    report_date = (
        str(snapshot_date or "").strip()[:10]
        or _infer_report_date_from_workbooks_dir(workbooks_dir)
        or datetime.now().strftime("%Y-%m-%d")
    )
    period = resolve_period_context(
        as_of_date=report_date,
        snapshot_date=report_date,
        deck_date=report_date,
    )
    return {
        "report_date": report_date,
        "analysis_year": period.current_quarter.year,
        "retrospective_quarter_label": period.prior_quarter.label,
        "retrospective_quarter_title": period.prior_quarter.title,
        "current_quarter_label": period.current_quarter.label,
        "current_quarter_title": period.current_quarter.title,
    }


def _resolve_report_plan(
    *,
    snapshot_date: str | None = None,
    workbooks_dir: Path | None = None,
) -> dict[str, list[tuple[str, str]]]:
    runtime = _resolve_runtime_context(
        snapshot_date=snapshot_date,
        workbooks_dir=workbooks_dir,
    )
    contract = resolve_historical_trending_contract(
        retrospective_label=str(runtime["retrospective_quarter_label"]),
        retrospective_title=str(runtime["retrospective_quarter_title"]),
        current_label=str(runtime["current_quarter_label"]),
        current_title=str(runtime["current_quarter_title"]),
    )
    fallback_reports = _load_historical_report_audit_fallback(
        target_quarter_label=str(runtime["current_quarter_label"]),
        target_quarter_title=str(runtime["current_quarter_title"]),
        run_date=str(runtime["report_date"]),
    )
    missing = []
    plan: dict[str, list[tuple[str, str]]] = {}
    for slug, quarter_ids in QUARTER_REPORTS.items():
        retrospective_id = str(
            quarter_ids.get(str(runtime["retrospective_quarter_label"])) or ""
        ).strip()
        current_id = str(
            quarter_ids.get(str(runtime["current_quarter_label"])) or ""
        ).strip()
        if not current_id:
            current_id = str(fallback_reports.get(slug) or "").strip()
        missing_quarters = []
        if not retrospective_id:
            missing_quarters.append(str(runtime["retrospective_quarter_label"]))
        if not current_id:
            missing_quarters.append(str(runtime["current_quarter_label"]))
        if missing_quarters:
            missing.append(f"{slug} missing {', '.join(missing_quarters)}")
            continue
        plan[slug] = [
            (contract.retrospective_snapshot_sheet, retrospective_id),
            (contract.current_snapshot_sheet, current_id),
        ]
    if missing:
        raise ValueError(
            "Historical-trending report registry is incomplete for "
            f"{runtime['report_date']} ({runtime['retrospective_quarter_title']} / "
            f"{runtime['current_quarter_title']}). "
            + "; ".join(missing[:3])
        )
    return plan


def _load_historical_report_audit_fallback(
    *,
    target_quarter_label: str,
    target_quarter_title: str,
    run_date: str,
) -> dict[str, str]:
    fallback: dict[str, str] = {}
    audit_paths = sorted(
        SOURCE_CONTRACT_AUDIT_ROOT.glob("*/source_contract_audit.json"),
        reverse=True,
    )
    for audit_path in audit_paths:
        audit_date = audit_path.parent.name
        if audit_date > str(run_date)[:10]:
            continue
        try:
            payload = json.loads(audit_path.read_text())
        except (OSError, ValueError, json.JSONDecodeError):
            continue
        candidate = (
            payload.get("candidate_forward_quarter")
            or payload.get("candidate_q3")
            or {}
        )
        if (
            str(candidate.get("quarter_label") or "").strip() != str(target_quarter_label)
            or str(candidate.get("quarter_title") or "").strip() != str(target_quarter_title)
        ):
            continue
        for item in candidate.get("historical_reports") or []:
            territory = str(item.get("director_slug") or "").strip()
            report_id = str(item.get("report_id") or "").strip()
            if (
                not territory
                or not report_id
                or str(item.get("status") or "").strip() != "ok"
            ):
                continue
            for slug, mapped_territory in DIRECTOR_SLUG_TO_TERRITORY.items():
                if mapped_territory == territory and slug not in fallback:
                    fallback[slug] = report_id
        if fallback:
            break
    return fallback


def _auth():
    data = json.loads(
        subprocess.run(
            ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout
    )["result"]
    return data["accessToken"], data["instanceUrl"]


def _parse_column(raw):
    """Rewrite historical-trending column tokens to readable headers.

    Incoming:
        'Opportunity__hd.APTS_Forecast_ARR__c_hst.CONVERT.2026-04-12.Change'
        'Opportunity__hd.StageName__hst.CONVERT.2026-04-12'
        'Opportunity.Account.Name'
    Outgoing:
        'ARR Change 2026-04-12'
        'Stage 2026-04-12'
        'Account'
    """
    # Historical snapshot columns
    m = re.match(
        r"Opportunity__hd\.([A-Za-z0-9_]+)_hst(?:\.CONVERT)?\.(\d{4}-\d{2}-\d{2})(\.Change)?",
        raw,
    )
    if m:
        field, date, change = m.groups()
        friendly = {
            "APTS_Forecast_ARR__c": "ARR",
            "APTS_Opportunity_ARR__c": "Opp ARR",
            "StageName": "Stage",
            "CloseDate": "Close",
            "ForecastCategoryName": "ForecastCat",
        }.get(field, field)
        suffix = " Change" if change else ""
        return f"{friendly}{suffix} {date}"
    # Live (non-snapshot) columns
    short = {
        "Opportunity.Account.Name": "Account",
        "Opportunity.Name": "Opportunity",
        "Opportunity.CloseDate": "Close Date (live)",
        "Opportunity.StageName": "Stage (live)",
        "Opportunity.Owner.Name": "Owner",
        "Opportunity.APTS_Forecast_ARR__c": "ARR Wtd (live)",
        "Opportunity.APTS_Opportunity_ARR__c": "ARR Unwtd (live)",
    }.get(raw)
    return short or raw


def _run_report(session, instance, report_id):
    """Run a Historical Trending report and return (labels, dtypes, rows)."""
    r = session.post(
        f"{instance}/services/data/v66.0/analytics/reports/{report_id}"
        "?includeDetails=true",
        headers={"Content-Type": "application/json"},
    ).json()
    md = r.get("reportMetadata", {})
    cols = md.get("detailColumns", [])
    ext = r.get("reportExtendedMetadata", {}).get("detailColumnInfo", {})
    labels = [_parse_column(c) for c in cols]
    dtypes = [ext.get(c, {}).get("dataType", "string") for c in cols]
    rows = []
    for row in r.get("factMap", {}).get("T!T", {}).get("rows", []):
        rows.append([c.get("label", "") for c in row.get("dataCells", [])])
    return labels, dtypes, rows, md.get("historicalSnapshotDates", [])


def _validate_snapshot_freshness(
    snapshots: list[str] | tuple[str, ...],
    *,
    report_date: str,
) -> dict[str, object]:
    snapshot_dates = sorted(
        {
            str(snapshot)[:10]
            for snapshot in (snapshots or [])
            if str(snapshot or "").strip()
        }
    )
    issues: list[str] = []
    latest_snapshot = snapshot_dates[-1] if snapshot_dates else ""
    run_date = str(report_date)[:10]
    run_month = run_date[:7]
    if not snapshot_dates:
        issues.append("snapshot_dates_missing")
    else:
        if latest_snapshot[:7] != run_month:
            issues.append("snapshot_review_month_mismatch")
        if latest_snapshot > run_date:
            issues.append("snapshot_after_run_date")
    return {
        "snapshot_dates": snapshot_dates,
        "latest_snapshot_date": latest_snapshot,
        "run_date": run_date,
        "run_month": run_month,
        "issues": issues,
        "ok": not issues,
    }


def _write_run_audit(output_dir: Path, payload: dict[str, object]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "historical_trending_extract_audit.json").write_text(
        json.dumps(payload, indent=2) + "\n"
    )
    lines = [
        f"# Historical Trending Extract Audit — {payload['run_date']}",
        "",
        f"- Status: `{payload['status']}`",
        f"- Workbooks dir: `{payload['workbooks_dir']}`",
        f"- Scope: `{payload['scope']}`",
        f"- Retrospective quarter: `{payload['retrospective_quarter_title']}`",
        f"- Current quarter: `{payload['current_quarter_title']}`",
        "",
        "## Processed",
        "",
    ]
    processed = payload.get("processed") or []
    if not processed:
        lines.append("- none")
    else:
        for item in processed:
            lines.append(
                f"- `{item['slug']}` saved with {len(item.get('sheets') or [])} sheet(s)"
            )
    lines.extend(["", "## Failures", ""])
    failures = payload.get("failures") or []
    if not failures:
        lines.append("- none")
    else:
        for item in failures:
            issue_text = ", ".join(item.get("issues") or []) or "unknown"
            lines.append(
                f"- `{item.get('slug')}` / `{item.get('sheet_name')}`: {issue_text}"
            )
    (output_dir / "summary.md").write_text("\n".join(lines) + "\n")


def _write_sheet(wb, sheet_name, labels, dtypes, rows, snapshots):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Subtitle row with snapshot dates
    if snapshots:
        ws.cell(
            row=1,
            column=1,
            value=f"Historical Trending, snapshots: {', '.join(snapshots)}",
        ).font = Font(italic=True, size=9, color="595959")

    # Headers on row 2
    for i, label in enumerate(labels, 1):
        cell = ws.cell(row=2, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[2].height = 32

    # Data rows from row 3
    for r_i, row in enumerate(rows, 3):
        for c_i, (val, dt) in enumerate(zip(row, dtypes), 1):
            cell = ws.cell(row=r_i, column=c_i)
            if val is None or val == "":
                cell.value = None
            elif dt == "currency":
                # Values come as strings like 'EUR 1.234.567,89' or '-'
                try:
                    # Strip EUR / spaces / parse European format
                    s = str(val).replace("EUR", "").strip()
                    if s in ("-", ""):
                        cell.value = None
                    else:
                        s = (
                            s.replace(".", "").replace(",", ".")
                            if "," in s
                            else s.replace(",", "")
                        )
                        cell.value = float(s)
                        cell.number_format = EUR_FMT
                except (ValueError, TypeError):
                    cell.value = val
            else:
                cell.value = val
            cell.font = BODY_FONT

    # Freeze header and first 2 columns (Account, Opportunity)
    ws.freeze_panes = "C3"
    # Auto-width-ish
    for col_idx in range(1, len(labels) + 1):
        letter = ws.cell(row=2, column=col_idx).column_letter
        header_len = len(str(labels[col_idx - 1] or ""))
        ws.column_dimensions[letter].width = max(12, min(header_len + 2, 22))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbooks-dir",
        type=Path,
    )
    parser.add_argument(
        "--snapshot-date",
        help="Explicit report date (YYYY-MM-DD) used for workbook path resolution and period guards.",
    )
    parser.add_argument(
        "--director",
        help="Only process one director slug (e.g. jesper-tyrer)",
    )
    args = parser.parse_args()

    runtime = _resolve_runtime_context(
        snapshot_date=args.snapshot_date,
        workbooks_dir=args.workbooks_dir,
    )
    audit_dir = AUDIT_OUTPUT_ROOT / str(runtime["report_date"])
    workbooks_dir = args.workbooks_dir or (
        Path("output/director_live_workbooks") / runtime["report_date"]
    )
    try:
        report_plan = _resolve_report_plan(
            snapshot_date=args.snapshot_date,
            workbooks_dir=workbooks_dir,
        )
    except ValueError as exc:
        _write_run_audit(
            audit_dir,
            {
                "run_date": str(runtime["report_date"]),
                "workbooks_dir": str(workbooks_dir),
                "scope": args.director or "all",
                "retrospective_quarter_title": str(
                    runtime["retrospective_quarter_title"]
                ),
                "current_quarter_title": str(runtime["current_quarter_title"]),
                "status": "failed",
                "processed": [],
                "failures": [{"issues": ["report_plan_resolution_failed"], "message": str(exc)}],
            },
        )
        print(f"  {exc}")
        print(
            "  Audit: "
            f"{(audit_dir / 'historical_trending_extract_audit.json').relative_to(ROOT)}"
        )
        return 1

    if not workbooks_dir.exists():
        _write_run_audit(
            audit_dir,
            {
                "run_date": str(runtime["report_date"]),
                "workbooks_dir": str(workbooks_dir),
                "scope": args.director or "all",
                "retrospective_quarter_title": str(
                    runtime["retrospective_quarter_title"]
                ),
                "current_quarter_title": str(runtime["current_quarter_title"]),
                "status": "failed",
                "processed": [],
                "failures": [
                    {
                        "issues": ["workbooks_dir_missing"],
                        "message": f"workbooks dir missing: {workbooks_dir}",
                    }
                ],
            },
        )
        print(f"  workbooks dir missing: {workbooks_dir}")
        print(
            "  Audit: "
            f"{(audit_dir / 'historical_trending_extract_audit.json').relative_to(ROOT)}"
        )
        return 1

    token, instance = _auth()
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    targets = [args.director] if args.director else list(QUARTER_REPORTS.keys())
    processed = []
    failures = []
    for slug in targets:
        if slug not in report_plan:
            print(f"  skip {slug}: no report mapping")
            continue
        wb_path = workbooks_dir / f"{slug}.xlsx"
        if not wb_path.exists():
            print(f"  skip {slug}: workbook missing")
            continue

        print(f"  {slug}...")
        wb = load_workbook(wb_path)
        workbook_failed = False
        processed_sheets = []

        for label, rid in report_plan[slug]:
            try:
                labels, dtypes, rows, snapshots = _run_report(session, instance, rid)
                validation = _validate_snapshot_freshness(
                    snapshots,
                    report_date=str(runtime["report_date"]),
                )
                if not validation["ok"]:
                    workbook_failed = True
                    failures.append(
                        {
                            "slug": slug,
                            "sheet_name": label,
                            "report_id": rid,
                            "issues": validation["issues"],
                            "latest_snapshot_date": validation["latest_snapshot_date"],
                            "run_date": validation["run_date"],
                        }
                    )
                    print(
                        "    "
                        f"{label}: blocked ({', '.join(validation['issues'])}; "
                        f"latest snapshot {validation['latest_snapshot_date'] or 'missing'} "
                        f"for run date {validation['run_date']})"
                    )
                    break
                _write_sheet(wb, label, labels, dtypes, rows, snapshots)
                processed_sheets.append(
                    {
                        "sheet_name": label,
                        "report_id": rid,
                        "row_count": len(rows),
                        "snapshot_dates": list(validation["snapshot_dates"]),
                        "latest_snapshot_date": validation["latest_snapshot_date"],
                    }
                )
                print(f"    {label}: {len(rows)} rows, snapshots {len(snapshots)}")
            except Exception as exc:
                workbook_failed = True
                failures.append(
                    {
                        "slug": slug,
                        "sheet_name": label,
                        "report_id": rid,
                        "issues": ["exception"],
                        "message": str(exc),
                    }
                )
                print(f"    {label}: failed ({exc})")
                break

        if workbook_failed:
            print("    workbook not saved")
            continue

        wb.save(wb_path)
        processed.append(
            {
                "slug": slug,
                "workbook_path": str(wb_path),
                "sheets": processed_sheets,
            }
        )

    status = "failed" if failures else "ok"
    _write_run_audit(
        audit_dir,
        {
            "run_date": str(runtime["report_date"]),
            "workbooks_dir": str(workbooks_dir),
            "scope": args.director or "all",
            "retrospective_quarter_title": str(runtime["retrospective_quarter_title"]),
            "current_quarter_title": str(runtime["current_quarter_title"]),
            "status": status,
            "processed": processed,
            "failures": failures,
        },
    )
    print("\nDone.")
    print(
        "Audit: "
        f"{(audit_dir / 'historical_trending_extract_audit.json').relative_to(ROOT)}"
    )
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
