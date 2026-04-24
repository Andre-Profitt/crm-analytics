#!/usr/bin/env python3
"""Build a grounded truth packet for Sales Director deck production.

The packet bridges Gold analytics to PowerPoint automation:
- claim registry with stable IDs and source paths
- RAG-ready JSONL corpus for grounded narrative generation
- think-cell source workbook for Excel links
- .ppttc JSON payload for think-cell JSON automation templates
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_director_gold_analytics import region_for_territory, slugify  # noqa: E402
from scripts.monthly_platform.bundle_validation import validate_bundle  # noqa: E402
from scripts.monthly_platform.models import DirectorBundle  # noqa: E402

DEFAULT_OUTPUT_ROOT = ROOT / "output" / "deck_truth_packets"
IMMATERIAL_NEGATIVE_ARR_THRESHOLD = 10_000
SIDECAR_CLAIM_FIELDS = [
    "open_land_deals",
    "open_land_arr",
    "open_land_arr_wtd",
    "q2_open_deals",
    "q2_open_arr",
    "q3_open_deals",
    "q3_open_arr",
    "q1_land_wins",
    "q1_land_wins_arr",
    "q1_land_lost",
    "q1_land_lost_arr",
    "q2_renewals",
    "q2_renewals_acv",
    "q3_renewals",
    "q3_renewals_acv",
    "approved_2026",
    "conditionally_approved",
    "missing_stage3",
]


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"{path}: expected JSON object")
    return payload


def save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def save_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def resolve_path(value: str | Path | None) -> Path | None:
    if not value:
        return None
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def tc_cell(value: Any, *, percentage: bool = False) -> dict[str, Any] | None:
    if value is None:
        return None
    if percentage:
        return {"percentage": float(value)}
    if isinstance(value, bool):
        return {"string": "true" if value else "false"}
    if isinstance(value, int | float):
        return {"number": value}
    return {"string": str(value)}


def tc_table(rows: list[dict[str, Any]], columns: list[str]) -> list[list[dict[str, Any] | None]]:
    table = [[tc_cell(column) for column in columns]]
    for row in rows:
        table.append([tc_cell(row.get(column)) for column in columns])
    return table


def tc_text(name: str, value: Any) -> dict[str, Any]:
    return {"name": name, "table": [[tc_cell(value)]]}


def metric(
    *,
    metric_id: str,
    scope: str,
    label: str,
    value: Any,
    unit: str,
    formula: str,
    source_artifact: Path,
    source_json_path: str,
    grain: str,
    horizon: str,
    currency: str | None,
    omitted_stage_policy: str,
    owner: str,
) -> dict[str, Any]:
    return {
        "metric_id": metric_id,
        "scope": scope,
        "label": label,
        "value": value,
        "unit": unit,
        "formula": formula,
        "source_artifact": display_path(source_artifact),
        "source_json_path": source_json_path,
        "grain": grain,
        "horizon": horizon,
        "currency": currency,
        "omitted_stage_policy": omitted_stage_policy,
        "owner": owner,
    }


def read_tieout_by_slug(tieout_path: Path | None) -> dict[str, dict[str, Any]]:
    if not tieout_path or not tieout_path.exists():
        return {}
    payload = load_json(tieout_path)
    return {
        str(item.get("slug") or ""): item
        for item in payload.get("directors") or []
        if item.get("slug")
    }


def tieout_mismatches(record: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not record:
        return []
    return [
        metric
        for metric in record.get("metrics") or []
        if str(metric.get("status") or "") != "match"
    ]


def bundle_validation_issues(bundle_path: Path | None) -> list[str]:
    if not bundle_path or not bundle_path.exists():
        return ["DirectorBundle JSON is missing."]
    try:
        bundle = DirectorBundle.from_json(bundle_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"DirectorBundle JSON failed to load: {exc}"]
    return validate_bundle(bundle)


def classify_bundle_issue(issue: str) -> dict[str, Any]:
    amount_match = re.search(r"negative arr_unweighted \((-?\d+(?:\.\d+)?)\)", issue)
    if amount_match and abs(float(amount_match.group(1))) < IMMATERIAL_NEGATIVE_ARR_THRESHOLD:
        return {
            "severity": "medium",
            "issue": "bundle_validation_warning",
            "evidence": issue,
        }
    return {
        "severity": "high",
        "issue": "bundle_validation_issue",
        "evidence": issue,
    }


def load_etl_summary(snapshot_date: str, slug: str) -> dict[str, Any] | None:
    path = (
        ROOT
        / "output"
        / "etl_intelligence_audit"
        / snapshot_date
        / slug
        / "etl_intelligence_audit.json"
    )
    if not path.exists():
        return None
    return load_json(path).get("summary") or {}


def claim(
    *,
    claim_id: str,
    metric_id: str,
    scope: str,
    label: str,
    value: Any,
    unit: str,
    source_artifact: Path,
    source_json_path: str,
    deck_element_name: str,
    narrative: str,
) -> dict[str, Any]:
    return {
        "claim_id": claim_id,
        "metric_id": metric_id,
        "scope": scope,
        "label": label,
        "value": value,
        "unit": unit,
        "source_artifact": display_path(source_artifact),
        "source_json_path": source_json_path,
        "deck_element_name": deck_element_name,
        "narrative": narrative,
    }


def compile_claim(
    metric_row: dict[str, Any],
    *,
    claim_id: str,
    deck_element_name: str,
    narrative: str,
) -> dict[str, Any]:
    return claim(
        claim_id=claim_id,
        metric_id=metric_row["metric_id"],
        scope=metric_row["scope"],
        label=metric_row["label"],
        value=metric_row["value"],
        unit=metric_row["unit"],
        source_artifact=Path(metric_row["source_artifact"]),
        source_json_path=metric_row["source_json_path"],
        deck_element_name=deck_element_name,
        narrative=narrative,
    )


def director_metric_claims(
    pack: dict[str, Any], pack_path: Path
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    slug = slugify(pack["director"])
    summary = pack["summary"]
    director = pack["director"]
    common = {
        "scope": "director",
        "source_artifact": pack_path,
        "grain": "director_book",
        "owner": "sales_director_monthly_gold_analytics",
    }
    metric_rows = [
        metric(
            metric_id=f"gold.director.{slug}.open_deals",
            label="Open deals",
            value=summary["open_deals"],
            unit="deals",
            formula="count(DirectorBundle.datasets.pipeline_open rows)",
            source_json_path="summary.open_deals",
            horizon="all_open_any_close_date",
            currency=None,
            omitted_stage_policy=(
                "included; this is an all-open pipeline count and must not be relabeled "
                "as active forecast pipeline without an omitted-stage exclusion."
            ),
            **common,
        ),
        metric(
            metric_id=f"gold.director.{slug}.open_arr",
            label="Open ARR",
            value=summary["open_arr"],
            unit="EUR ARR",
            formula="sum(DirectorBundle.datasets.pipeline_open.arr_unweighted)",
            source_json_path="summary.open_arr",
            horizon="all_open_any_close_date",
            currency="EUR",
            omitted_stage_policy=(
                "included; this all-open ARR includes Omitted forecast-category rows and "
                "must be labeled separately from active ex-Omitted pipeline."
            ),
            **common,
        ),
        metric(
            metric_id=f"gold.director.{slug}.deal_risk_rows",
            label="Deal risk rows",
            value=summary["deal_risk_rows"],
            unit="risk rows",
            formula="count(analytics.deal_risk_index rows with risk_score > 0)",
            source_json_path="summary.deal_risk_rows",
            horizon="all_open_any_close_date",
            currency=None,
            omitted_stage_policy=(
                "inherits all-open pipeline scope; source rows retain forecast_category "
                "so omitted exposure can be isolated."
            ),
            **common,
        ),
        metric(
            metric_id=f"gold.director.{slug}.close_date_events",
            label="Close-date events",
            value=summary["close_date_event_count"],
            unit="events",
            formula="count(DirectorBundle.datasets.close_date_events rows)",
            source_json_path="summary.close_date_event_count",
            horizon="full_snapshot_history",
            currency=None,
            omitted_stage_policy="not_applicable",
            **common,
        ),
        metric(
            metric_id=f"gold.director.{slug}.top20_concentration_pct",
            label="Top 20 pipeline concentration",
            value=summary["top_20_pipeline_concentration_pct"],
            unit="percent",
            formula="sum(top 20 open ARR) / sum(all open ARR) * 100",
            source_json_path="summary.top_20_pipeline_concentration_pct",
            horizon="all_open_any_close_date",
            currency=None,
            omitted_stage_policy=(
                "included; concentration is measured against all-open ARR including Omitted rows."
            ),
            **common,
        ),
        metric(
            metric_id=f"gold.director.{slug}.high_stage_zero_arr_count",
            label="Stage 3+ zero ARR count",
            value=summary["high_stage_zero_arr_count"],
            unit="deals",
            formula="count(pipeline_open rows where stage_number >= 3 and arr_unweighted == 0)",
            source_json_path="summary.high_stage_zero_arr_count",
            horizon="all_open_any_close_date",
            currency=None,
            omitted_stage_policy=(
                "included for hygiene detection only; zero-ARR rows must stay out of revenue "
                "headline claims."
            ),
            **common,
        ),
    ]
    narratives = {
        "open_deals": f"{director} has {summary['open_deals']} open deals.",
        "open_arr": f"{director} has EUR {summary['open_arr']:,.0f} open ARR.",
        "deal_risk_rows": (
            f"{director} has {summary['deal_risk_rows']} risk-scored open deals."
        ),
        "close_date_events": (
            f"{director} has {summary['close_date_event_count']} close-date history events."
        ),
        "top20_concentration_pct": (
            f"{director}'s top 20 deals carry "
            f"{summary['top_20_pipeline_concentration_pct']}% of open ARR."
        ),
        "high_stage_zero_arr_count": (
            f"{director} has {summary['high_stage_zero_arr_count']} "
            "stage 3+ open deals with zero ARR."
        ),
    }
    claim_rows = [
        compile_claim(
            metric_row,
            claim_id=metric_row["metric_id"].replace("gold.", "", 1),
            deck_element_name=f"{slug}_{metric_row['metric_id'].split('.')[-1]}",
            narrative=narratives[metric_row["metric_id"].split(".")[-1]],
        )
        for metric_row in metric_rows
    ]
    return metric_rows, claim_rows


def regional_metric_claims(
    gold_manifest: dict[str, Any], manifest_path: Path
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    metrics: list[dict[str, Any]] = []
    claims: list[dict[str, Any]] = []
    for rollup in gold_manifest.get("regional_rollups") or []:
        region_slug = slugify(rollup["region"])
        totals = rollup["totals"]
        common = {
            "scope": "region",
            "source_artifact": manifest_path,
            "grain": "regional_director_book_rollup",
            "horizon": "all_open_any_close_date",
            "owner": "sales_director_monthly_gold_analytics",
        }
        region_metrics = [
            metric(
                metric_id=f"gold.region.{region_slug}.open_deals",
                label="Open deals",
                value=totals["open_deals"],
                unit="deals",
                formula="sum(gold.director.*.open_deals for directors in region)",
                source_json_path=f"regional_rollups.{rollup['region']}.totals.open_deals",
                currency=None,
                omitted_stage_policy=(
                    "included; this regional all-open count must not be relabeled as "
                    "active forecast pipeline without omitted-stage exclusion."
                ),
                **common,
            ),
            metric(
                metric_id=f"gold.region.{region_slug}.open_arr",
                label="Open ARR",
                value=totals["open_arr"],
                unit="EUR ARR",
                formula="sum(gold.director.*.open_arr for directors in region)",
                source_json_path=f"regional_rollups.{rollup['region']}.totals.open_arr",
                currency="EUR",
                omitted_stage_policy=(
                    "included; regional all-open ARR includes Omitted forecast-category rows "
                    "and must be labeled separately from active ex-Omitted pipeline."
                ),
                **common,
            ),
        ]
        metrics.extend(region_metrics)
        claims.extend(
            [
                compile_claim(
                    region_metrics[0],
                    claim_id=f"region.{region_slug}.open_deals",
                    deck_element_name=f"{region_slug}_open_deals",
                    narrative=f"{rollup['region']} has {totals['open_deals']} open deals.",
                ),
                compile_claim(
                    region_metrics[1],
                    claim_id=f"region.{region_slug}.open_arr",
                    deck_element_name=f"{region_slug}_open_arr",
                    narrative=f"{rollup['region']} has EUR {totals['open_arr']:,.0f} open ARR.",
                ),
            ]
        )
    return metrics, claims


def sidecar_claim_id(slug: str, field: str) -> str:
    return f"deck.{slug}.{field}"


def sidecar_metric_specs() -> dict[str, dict[str, str | None]]:
    return {
        "open_land_deals": {
            "label": "Open Land pipeline deals",
            "unit": "deals",
            "formula": "count(active Land pipeline rows in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": None,
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "open_land_arr": {
            "label": "Open Land pipeline ARR",
            "unit": "EUR ARR",
            "formula": "sum(active Land pipeline ARR Unweighted EUR in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "open_land_arr_wtd": {
            "label": "Weighted open Land pipeline ARR",
            "unit": "EUR weighted ARR",
            "formula": "sum(active Land pipeline ARR Weighted EUR in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "q2_open_deals": {
            "label": "Q2 open Land deals",
            "unit": "deals",
            "formula": "count(active Land pipeline rows with current-quarter close dates)",
            "horizon": "current_quarter",
            "currency": None,
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "q2_open_arr": {
            "label": "Q2 open Land ARR",
            "unit": "EUR ARR",
            "formula": "sum(active Land pipeline ARR Unweighted EUR with current-quarter close dates)",
            "horizon": "current_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "q3_open_deals": {
            "label": "Forward-quarter open Land deals",
            "unit": "deals",
            "formula": "count(active Land pipeline rows with forward-quarter close dates)",
            "horizon": "forward_quarter",
            "currency": None,
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "q3_open_arr": {
            "label": "Forward-quarter open Land ARR",
            "unit": "EUR ARR",
            "formula": "sum(active Land pipeline ARR Unweighted EUR with forward-quarter close dates)",
            "horizon": "forward_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "excluded via active forecast-category policy",
        },
        "q1_land_wins": {
            "label": "Q1 Land wins",
            "unit": "deals",
            "formula": "count(Land won rows with prior-quarter close dates)",
            "horizon": "prior_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "q1_land_wins_arr": {
            "label": "Q1 Land won ARR",
            "unit": "EUR ARR",
            "formula": "sum(Land won ARR Unweighted EUR with prior-quarter close dates)",
            "horizon": "prior_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "not_applicable",
        },
        "q1_land_lost": {
            "label": "Q1 Land losses",
            "unit": "deals",
            "formula": "count(Land lost rows with prior-quarter close dates)",
            "horizon": "prior_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "q1_land_lost_arr": {
            "label": "Q1 Land lost ARR",
            "unit": "EUR ARR",
            "formula": "sum(Land lost ARR Unweighted EUR with prior-quarter close dates)",
            "horizon": "prior_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "not_applicable",
        },
        "q2_renewals": {
            "label": "Q2 renewals",
            "unit": "renewals",
            "formula": "count(renewal rows with current-quarter close dates)",
            "horizon": "current_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "q2_renewals_acv": {
            "label": "Q2 renewal ACV",
            "unit": "EUR ACV",
            "formula": "sum(renewal ACV EUR with current-quarter close dates)",
            "horizon": "current_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "not_applicable",
        },
        "q3_renewals": {
            "label": "Forward-quarter renewals",
            "unit": "renewals",
            "formula": "count(renewal rows with forward-quarter close dates)",
            "horizon": "forward_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "q3_renewals_acv": {
            "label": "Forward-quarter renewal ACV",
            "unit": "EUR ACV",
            "formula": "sum(renewal ACV EUR with forward-quarter close dates)",
            "horizon": "forward_quarter",
            "currency": "EUR",
            "omitted_stage_policy": "not_applicable",
        },
        "approved_2026": {
            "label": "Approved 2026 count",
            "unit": "deals",
            "formula": "count(commercial approval rows classified as approved in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "conditionally_approved": {
            "label": "Conditionally approved count",
            "unit": "deals",
            "formula": "count(commercial approval rows classified as conditionally approved in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
        "missing_stage3": {
            "label": "Missing Stage 3+ approval count",
            "unit": "deals",
            "formula": "count(stage 3+ commercial approval rows classified as missing in reporting scope)",
            "horizon": "reporting_window_q1_to_current_quarter",
            "currency": None,
            "omitted_stage_policy": "not_applicable",
        },
    }


def sidecar_metric_claims(
    *,
    slug: str,
    sidecar_path: Path | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any] | None]:
    if not sidecar_path or not sidecar_path.exists():
        return [], [], None
    payload = load_json(sidecar_path)
    specs = sidecar_metric_specs()
    metric_rows: list[dict[str, Any]] = []
    claim_rows: list[dict[str, Any]] = []
    required_claim_ids: dict[str, str] = {}
    embedded_claim_ids = payload.get("claim_ids") or {}
    for field in SIDECAR_CLAIM_FIELDS:
        if field not in payload or field not in specs:
            continue
        spec = specs[field]
        claim_id = sidecar_claim_id(slug, field)
        required_claim_ids[field] = claim_id
        metric_row = metric(
            metric_id=f"deck_sidecar.{slug}.{field}",
            scope="deck_sidecar",
            label=str(spec["label"]),
            value=payload.get(field),
            unit=str(spec["unit"]),
            formula=str(spec["formula"]),
            source_artifact=sidecar_path,
            source_json_path=field,
            grain="director_deck_sidecar",
            horizon=str(spec["horizon"]),
            currency=None if spec["currency"] is None else str(spec["currency"]),
            omitted_stage_policy=str(spec["omitted_stage_policy"]),
            owner="sales_director_monthly_deck_sidecar",
        )
        metric_rows.append(metric_row)
        claim_rows.append(
            compile_claim(
                metric_row,
                claim_id=claim_id,
                deck_element_name=f"{slug}_{field}",
                narrative=(
                    f"{payload.get('director') or slug} rendered {spec['label']} "
                    f"as {payload.get(field)}."
                ),
            )
        )
    missing_embedded = [
        field
        for field, claim_id in required_claim_ids.items()
        if embedded_claim_ids.get(field) != claim_id
    ]
    reference = {
        "slug": slug,
        "sidecar_path": display_path(sidecar_path),
        "claim_ids": required_claim_ids,
        "embedded": not missing_embedded,
        "missing_embedded_fields": missing_embedded,
    }
    return metric_rows, claim_rows, reference


def build_packet(
    *,
    snapshot_date: str,
    gold_root: Path,
    workbook_dir: Path,
    bundle_dir: Path,
    decks_dir: Path | None,
    tieout_path: Path | None,
    template_path: str,
    analyst_workbook_path: Path | None = None,
    source_backed_publish_gate_path: Path | None = None,
) -> dict[str, Any]:
    manifest_path = gold_root / snapshot_date / "manifest.json"
    gold_manifest = load_json(manifest_path)
    tieout = read_tieout_by_slug(tieout_path)
    source_backed_publish_gate = (
        load_json(source_backed_publish_gate_path)
        if source_backed_publish_gate_path and source_backed_publish_gate_path.exists()
        else None
    )
    source_backed_publish_ok = (
        source_backed_publish_gate is not None
        and source_backed_publish_gate.get("status") == "ok"
    )

    directors: list[dict[str, Any]] = []
    metrics, claims = regional_metric_claims(gold_manifest, manifest_path)
    sidecar_claim_refs: list[dict[str, Any]] = []
    blockers: list[dict[str, Any]] = []

    for entry in gold_manifest.get("directors") or []:
        director = str(entry["director"])
        slug = slugify(director)
        territory = str(entry["territory"])
        region = region_for_territory(territory)
        pack_path = resolve_path(entry.get("json_path"))
        if not pack_path or not pack_path.exists():
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "missing_gold_pack",
                    "evidence": str(entry.get("json_path") or ""),
                }
            )
            continue

        pack = load_json(pack_path)
        director_metrics, director_claim_rows = director_metric_claims(pack, pack_path)
        metrics.extend(director_metrics)
        claims.extend(director_claim_rows)
        bundle_path = resolve_path(entry.get("bundle_path")) or (bundle_dir / f"{slug}.json")
        workbook_path = analyst_workbook_path or workbook_dir / f"{slug}.xlsx"
        sidecar_path = decks_dir / f"{slug}-LAND.json" if decks_dir else None
        sidecar_metrics, sidecar_claim_rows, sidecar_ref = sidecar_metric_claims(
            slug=slug,
            sidecar_path=sidecar_path,
        )
        metrics.extend(sidecar_metrics)
        claims.extend(sidecar_claim_rows)
        if sidecar_ref:
            sidecar_claim_refs.append(sidecar_ref)
        etl_summary = load_etl_summary(snapshot_date, slug)
        bundle_issues = bundle_validation_issues(bundle_path)
        mismatches = tieout_mismatches(tieout.get(slug))

        if not workbook_path.exists():
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "missing_workbook",
                    "evidence": display_path(workbook_path),
                }
            )
        if sidecar_path and not sidecar_path.exists():
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "missing_deck_sidecar",
                    "evidence": display_path(sidecar_path),
                }
            )
        if sidecar_ref and not sidecar_ref["embedded"]:
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "sidecar_claim_ids_missing",
                    "evidence": ", ".join(sidecar_ref["missing_embedded_fields"]),
                }
            )
        for issue in bundle_issues:
            finding = classify_bundle_issue(issue)
            blockers.append(
                {
                    "severity": finding["severity"],
                    "scope": slug,
                    "issue": finding["issue"],
                    "evidence": finding["evidence"],
                }
            )
        if etl_summary is None:
            if not source_backed_publish_ok:
                blockers.append(
                    {
                        "severity": "medium",
                        "scope": slug,
                        "issue": "missing_etl_intelligence_audit",
                        "evidence": f"output/etl_intelligence_audit/{snapshot_date}/{slug}",
                    }
                )
        elif int(etl_summary.get("high_gap_count") or 0) > 0:
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "etl_high_coverage_gap",
                    "evidence": f"{etl_summary['high_gap_count']} high gap(s)",
                }
            )
        for mismatch in mismatches:
            blockers.append(
                {
                    "severity": "high",
                    "scope": slug,
                    "issue": "deck_tieout_mismatch",
                    "evidence": (
                        f"{mismatch.get('metric')}: status={mismatch.get('status')}; "
                        f"extract={mismatch.get('extract')}; deck={mismatch.get('deck')}"
                    ),
                }
            )

        summary = pack["summary"]
        directors.append(
            {
                "director": director,
                "slug": slug,
                "territory": territory,
                "region": region,
                "open_deals": summary["open_deals"],
                "open_arr": summary["open_arr"],
                "deal_risk_rows": summary["deal_risk_rows"],
                "close_date_event_count": summary["close_date_event_count"],
                "top_20_pipeline_concentration_pct": summary[
                    "top_20_pipeline_concentration_pct"
                ],
                "high_stage_zero_arr_count": summary["high_stage_zero_arr_count"],
                "bundle_issue_count": len(bundle_issues),
                "etl_high_gap_count": 0
                if etl_summary is None
                else int(etl_summary.get("high_gap_count") or 0),
                "tieout_mismatch_count": len(mismatches),
                "gold_pack": display_path(pack_path),
                "bundle_path": display_path(bundle_path) if bundle_path else "",
                "workbook_path": display_path(workbook_path),
                "deck_sidecar_path": display_path(sidecar_path)
                if sidecar_path
                else "",
            }
        )

    high_blockers = [item for item in blockers if item["severity"] == "high"]
    status = "blocked" if high_blockers else "ok"
    return {
        "artifact_type": "deck_truth_packet",
        "schema_version": "1",
        "generated_at": datetime.now(UTC).isoformat(),
        "snapshot_date": snapshot_date,
        "status": status,
        "truth_policy": {
            "numeric_claim_rule": "Every numeric deck claim must resolve to a claim_id in this packet.",
            "ai_rule": "AI narrative may retrieve from rag_corpus.jsonl but must cite claim_id values for numbers.",
            "publish_rule": "High-severity blockers prevent leadership publish.",
        },
        "sources": {
            "gold_manifest": display_path(manifest_path),
            "bundle_dir": display_path(bundle_dir),
            "workbook_dir": display_path(workbook_dir),
            "analyst_workbook_path": display_path(analyst_workbook_path)
            if analyst_workbook_path
            else "",
            "decks_dir": display_path(decks_dir) if decks_dir else "",
            "tieout_path": display_path(tieout_path) if tieout_path else "",
            "source_backed_publish_gate": display_path(source_backed_publish_gate_path)
            if source_backed_publish_gate_path
            else "",
        },
        "thinkcell": {
            "template_path": template_path,
            "requires_named_template_elements": True,
            "recommended_element_names": [
                "TruthStatus",
                "RegionalRollupsTable",
                "DirectorKpiTable",
                "PublishBlockersTable",
                "MetricStoreTable",
                "ClaimRegistryTable",
            ],
        },
        "summary": {
            "director_count": len(directors),
            "metric_count": len(metrics),
            "claim_count": len(claims),
            "blocker_count": len(blockers),
            "high_blocker_count": len(high_blockers),
            "tieout_mismatch_count": sum(
                row["tieout_mismatch_count"] for row in directors
            ),
            "bundle_issue_count": sum(row["bundle_issue_count"] for row in directors),
            "source_backed_publish_gate_status": (
                source_backed_publish_gate.get("status")
                if source_backed_publish_gate
                else None
            ),
        },
        "regional_rollups": gold_manifest.get("regional_rollups") or [],
        "directors": directors,
        "metrics": metrics,
        "claims": claims,
        "deck_sidecar_claim_refs": sidecar_claim_refs,
        "blockers": blockers,
    }


def write_rag_corpus(path: Path, packet: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for claim_row in packet["claims"]:
            record = {
                "id": claim_row["claim_id"],
                "kind": "validated_numeric_claim",
                "text": claim_row["narrative"],
                "metadata": {
                    "metric_id": claim_row["metric_id"],
                    "scope": claim_row["scope"],
                    "label": claim_row["label"],
                    "value": claim_row["value"],
                    "unit": claim_row["unit"],
                    "source_artifact": claim_row["source_artifact"],
                    "source_json_path": claim_row["source_json_path"],
                    "deck_element_name": claim_row["deck_element_name"],
                },
            }
            handle.write(json.dumps(record, ensure_ascii=True) + "\n")


def write_workbook(path: Path, packet: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Control"
    add_rows(
        ws,
        [
            ["Field", "Value"],
            ["snapshot_date", packet["snapshot_date"]],
            ["status", packet["status"]],
            ["director_count", packet["summary"]["director_count"]],
            ["metric_count", packet["summary"]["metric_count"]],
            ["claim_count", packet["summary"]["claim_count"]],
            ["high_blocker_count", packet["summary"]["high_blocker_count"]],
            ["tieout_mismatch_count", packet["summary"]["tieout_mismatch_count"]],
        ],
    )

    regional = wb.create_sheet("Regional Rollups")
    add_dict_rows(
        regional,
        packet["regional_rollups"],
        [
            "region",
            "director_count",
            "territories",
            "open_deals",
            "open_arr",
            "deal_risk_rows",
            "close_date_event_count",
        ],
        transform=lambda row: {
            **row,
            "territories": ", ".join(row.get("territories") or []),
            **(row.get("totals") or {}),
        },
    )

    directors = wb.create_sheet("Director KPI")
    add_dict_rows(
        directors,
        packet["directors"],
        [
            "region",
            "territory",
            "director",
            "open_deals",
            "open_arr",
            "deal_risk_rows",
            "close_date_event_count",
            "top_20_pipeline_concentration_pct",
            "high_stage_zero_arr_count",
            "bundle_issue_count",
            "etl_high_gap_count",
            "tieout_mismatch_count",
        ],
    )

    fields = wb.create_sheet("TC Text Fields")
    add_dict_rows(
        fields,
        packet["claims"],
        [
            "deck_element_name",
            "claim_id",
            "metric_id",
            "label",
            "value",
            "unit",
            "source_artifact",
            "source_json_path",
        ],
    )

    metric_store = wb.create_sheet("Metric Store")
    add_dict_rows(
        metric_store,
        packet["metrics"],
        [
            "metric_id",
            "scope",
            "label",
            "value",
            "unit",
            "formula",
            "source_artifact",
            "source_json_path",
            "grain",
            "horizon",
            "currency",
            "omitted_stage_policy",
            "owner",
        ],
    )

    claim_registry = wb.create_sheet("Claim Registry")
    add_dict_rows(
        claim_registry,
        packet["claims"],
        [
            "claim_id",
            "metric_id",
            "deck_element_name",
            "scope",
            "label",
            "value",
            "unit",
            "source_artifact",
            "source_json_path",
        ],
    )

    sidecar_claim_refs = wb.create_sheet("Sidecar Claim Refs")
    add_dict_rows(
        sidecar_claim_refs,
        packet["deck_sidecar_claim_refs"],
        [
            "slug",
            "sidecar_path",
            "claim_ids",
            "embedded",
            "missing_embedded_fields",
        ],
        transform=lambda row: {
            **row,
            "claim_ids": json.dumps(row.get("claim_ids") or {}, sort_keys=True),
            "missing_embedded_fields": ", ".join(
                row.get("missing_embedded_fields") or []
            ),
        },
    )

    tc_table_sheet = wb.create_sheet("TC Director Table")
    add_dict_rows(
        tc_table_sheet,
        packet["directors"],
        ["director", "territory", "open_deals", "open_arr", "deal_risk_rows"],
    )

    blockers = wb.create_sheet("Publish Blockers")
    add_dict_rows(
        blockers,
        packet["blockers"],
        ["severity", "scope", "issue", "evidence"],
    )

    for sheet in wb.worksheets:
        style_sheet(sheet)
    wb.save(path)


def add_rows(ws: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def add_dict_rows(
    ws: Any,
    rows: list[dict[str, Any]],
    columns: list[str],
    *,
    transform: Any | None = None,
) -> None:
    ws.append(columns)
    for row in rows:
        value = transform(row) if transform else row
        ws.append([value.get(column) for column in columns])


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(
            max(max_len + 2, 12), 48
        )


def write_ppttc(path: Path, packet: dict[str, Any]) -> None:
    regional_rows = []
    for row in packet["regional_rollups"]:
        totals = row.get("totals") or {}
        regional_rows.append(
            {
                "Region": row["region"],
                "Directors": row["director_count"],
                "Open Deals": totals.get("open_deals"),
                "Open ARR": totals.get("open_arr"),
                "Risk Rows": totals.get("deal_risk_rows"),
            }
        )
    director_rows = [
        {
            "Director": row["director"],
            "Territory": row["territory"],
            "Open Deals": row["open_deals"],
            "Open ARR": row["open_arr"],
            "Risk Rows": row["deal_risk_rows"],
            "Tie-Out Gaps": row["tieout_mismatch_count"],
        }
        for row in packet["directors"]
    ]
    blocker_rows = [
        {
            "Severity": row["severity"],
            "Scope": row["scope"],
            "Issue": row["issue"],
            "Evidence": row["evidence"],
        }
        for row in packet["blockers"][:20]
    ]
    metric_rows = [
        {
            "Metric ID": row["metric_id"],
            "Label": row["label"],
            "Value": row["value"],
            "Unit": row["unit"],
            "Horizon": row["horizon"],
            "Omitted Policy": row["omitted_stage_policy"],
        }
        for row in packet["metrics"][:25]
    ]
    claim_rows = [
        {
            "Claim ID": row["claim_id"],
            "Metric ID": row["metric_id"],
            "Deck Element": row["deck_element_name"],
            "Value": row["value"],
            "Unit": row["unit"],
        }
        for row in packet["claims"][:25]
    ]
    data = [
        tc_text("TruthStatus", packet["status"]),
        tc_text("TruthBlockerCount", packet["summary"]["high_blocker_count"]),
        {
            "name": "RegionalRollupsTable",
            "table": tc_table(
                regional_rows,
                ["Region", "Directors", "Open Deals", "Open ARR", "Risk Rows"],
            ),
        },
        {
            "name": "DirectorKpiTable",
            "table": tc_table(
                director_rows,
                [
                    "Director",
                    "Territory",
                    "Open Deals",
                    "Open ARR",
                    "Risk Rows",
                    "Tie-Out Gaps",
                ],
            ),
        },
        {
            "name": "PublishBlockersTable",
            "table": tc_table(
                blocker_rows,
                ["Severity", "Scope", "Issue", "Evidence"],
            ),
        },
        {
            "name": "MetricStoreTable",
            "table": tc_table(
                metric_rows,
                ["Metric ID", "Label", "Value", "Unit", "Horizon", "Omitted Policy"],
            ),
        },
        {
            "name": "ClaimRegistryTable",
            "table": tc_table(
                claim_rows,
                ["Claim ID", "Metric ID", "Deck Element", "Value", "Unit"],
            ),
        },
    ]
    payload = [{"template": packet["thinkcell"]["template_path"], "data": data}]
    save_json(path, payload)


def markdown_summary(packet: dict[str, Any]) -> str:
    lines = [
        f"# Deck Truth Packet - {packet['snapshot_date']}",
        "",
        f"- Status: `{packet['status']}`",
        f"- Directors: `{packet['summary']['director_count']}`",
        f"- Metrics: `{packet['summary']['metric_count']}`",
        f"- Claims: `{packet['summary']['claim_count']}`",
        f"- High blockers: `{packet['summary']['high_blocker_count']}`",
        f"- Tie-out mismatches: `{packet['summary']['tieout_mismatch_count']}`",
        "",
        "## Truth Policy",
        "",
        f"- {packet['truth_policy']['numeric_claim_rule']}",
        f"- {packet['truth_policy']['ai_rule']}",
        f"- {packet['truth_policy']['publish_rule']}",
        "",
        "## Blockers",
        "",
    ]
    if not packet["blockers"]:
        lines.append("- None.")
    else:
        for blocker in packet["blockers"][:30]:
            lines.append(
                f"- `{blocker['severity']}` `{blocker['scope']}` "
                f"{blocker['issue']}: {blocker['evidence']}"
            )
    lines.append("")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--snapshot-date", required=True)
    parser.add_argument(
        "--gold-root",
        type=Path,
        default=ROOT / "output" / "director_gold_analytics",
    )
    parser.add_argument("--workbook-dir", type=Path)
    parser.add_argument(
        "--analyst-workbook",
        type=Path,
        help=(
            "Single source-backed analyst workbook to use as the workbook evidence "
            "for every director in the packet."
        ),
    )
    parser.add_argument("--source-backed-publish-gate", type=Path)
    parser.add_argument("--bundle-dir", type=Path)
    parser.add_argument("--decks-dir", type=Path)
    parser.add_argument("--tieout-path", type=Path)
    parser.add_argument(
        "--template-path",
        default="sales_director_thinkcell_template.pptx",
        help="Template path to put in the .ppttc file.",
    )
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--json", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    snapshot_date = args.snapshot_date[:10]
    workbook_dir = args.workbook_dir or (
        ROOT / "output" / "director_live_workbooks" / snapshot_date
    )
    bundle_dir = args.bundle_dir or (ROOT / "output" / "director_bundles" / snapshot_date)
    decks_dir = args.decks_dir
    if decks_dir is None:
        candidate = ROOT / "output" / "simcorp_director_decks" / snapshot_date / "land-only"
        decks_dir = candidate if candidate.exists() else None
    tieout_path = args.tieout_path
    if tieout_path is None:
        candidate = ROOT / "output" / "tie_out" / snapshot_date / "tie_out_audit.json"
        tieout_path = candidate if candidate.exists() else None

    packet = build_packet(
        snapshot_date=snapshot_date,
        gold_root=args.gold_root,
        workbook_dir=workbook_dir,
        bundle_dir=bundle_dir,
        decks_dir=decks_dir,
        tieout_path=tieout_path,
        template_path=args.template_path,
        analyst_workbook_path=args.analyst_workbook,
        source_backed_publish_gate_path=args.source_backed_publish_gate,
    )
    output_dir = args.output_root / snapshot_date
    manifest_path = output_dir / "deck_truth_packet.json"
    workbook_path = output_dir / "thinkcell_source.xlsx"
    ppttc_path = output_dir / "thinkcell_data.ppttc"
    rag_path = output_dir / "rag_corpus.jsonl"
    summary_path = output_dir / "summary.md"

    save_json(manifest_path, packet)
    write_workbook(workbook_path, packet)
    write_ppttc(ppttc_path, packet)
    write_rag_corpus(rag_path, packet)
    save_text(summary_path, markdown_summary(packet))

    result = {
        "status": packet["status"],
        "snapshot_date": snapshot_date,
        "manifest_path": str(manifest_path),
        "thinkcell_source_workbook": str(workbook_path),
        "thinkcell_ppttc": str(ppttc_path),
        "rag_corpus": str(rag_path),
        "summary_path": str(summary_path),
        "high_blocker_count": packet["summary"]["high_blocker_count"],
        "tieout_mismatch_count": packet["summary"]["tieout_mismatch_count"],
    }
    print(json.dumps(result, indent=2) if args.json else result)
    return 2 if packet["status"] == "blocked" else 0


if __name__ == "__main__":
    raise SystemExit(main())
