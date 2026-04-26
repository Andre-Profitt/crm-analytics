"""Build source-backed analyst workbooks from DirectorBundle artifacts.

This module only reads DirectorBundle JSON and its manifest. It does not read
legacy workbook tabs, call Salesforce, or generate AI prose.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from scripts.monthly_platform.models import DirectorBundle

REPO_ROOT = Path(__file__).resolve().parents[2]
ANALYST_WORKBOOK_SCHEMA_VERSION = "monthly_platform.source_backed_analyst_workbook.v1"
SHEET_NAMES = [
    "Executive Summary",
    "Source Coverage",
    "Metric Store",
    "Deal Exceptions",
    "Region Narrative Inputs",
    "Analyst Notes Seed",
]

HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=9)
MONEY_FORMAT = '#,##0'
COUNT_FORMAT = '#,##0'


@dataclass(frozen=True)
class AnalystWorkbookResult:
    workbook_path: Path
    sheet_row_counts: dict[str, int]
    bundle_count: int
    status: str


@dataclass(frozen=True)
class LoadedBundle:
    bundle: DirectorBundle
    path: Path
    coverage: dict[str, Any]


def build_source_backed_analyst_workbook(
    *,
    manifest_path: Path,
    output_path: Path | None = None,
) -> AnalystWorkbookResult:
    manifest_path = Path(manifest_path)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    output_path = output_path or manifest_path.parent / "source_backed_analyst_workbook.xlsx"
    loaded_bundles = load_manifest_bundles(manifest_path)

    wb = Workbook()
    wb.remove(wb.active)
    _set_stable_properties(wb, manifest)

    sheet_rows = {
        "Executive Summary": executive_summary_rows(manifest, loaded_bundles),
        "Source Coverage": source_coverage_rows(manifest, loaded_bundles),
        "Metric Store": metric_store_rows(manifest, loaded_bundles),
        "Deal Exceptions": deal_exception_rows(manifest, loaded_bundles),
    }
    sheet_rows["Region Narrative Inputs"] = region_narrative_input_rows(
        manifest,
        loaded_bundles,
        sheet_rows["Deal Exceptions"],
    )
    sheet_rows["Analyst Notes Seed"] = analyst_notes_seed_rows(
        manifest,
        loaded_bundles,
        sheet_rows["Deal Exceptions"],
    )

    headers_by_sheet = _headers_by_sheet()
    for sheet_name in SHEET_NAMES:
        _add_sheet(
            wb,
            sheet_name,
            headers_by_sheet[sheet_name],
            sheet_rows[sheet_name],
            numeric_columns=_numeric_columns(sheet_name, headers_by_sheet[sheet_name]),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return AnalystWorkbookResult(
        workbook_path=output_path,
        sheet_row_counts={name: len(rows) for name, rows in sheet_rows.items()},
        bundle_count=len(loaded_bundles),
        status=str(manifest.get("status") or "unknown"),
    )


def load_manifest_bundles(manifest_path: Path) -> list[LoadedBundle]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    coverage_by_territory = _coverage_by_territory(manifest)
    loaded: list[LoadedBundle] = []
    for bundle_path_text in manifest.get("bundle_paths", []):
        bundle_path = _resolve_manifest_path(bundle_path_text, manifest_path)
        bundle = DirectorBundle.from_json(bundle_path.read_text(encoding="utf-8"))
        loaded.append(
            LoadedBundle(
                bundle=bundle,
                path=bundle_path,
                coverage=coverage_by_territory.get(bundle.territory, {}),
            )
        )
    return sorted(loaded, key=lambda item: (item.bundle.territory, item.bundle.director))


def source_coverage_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        source_backed = set(loaded.coverage.get("source_backed") or [])
        optional_empty = set(loaded.coverage.get("optional_empty") or [])
        publish_required = set(loaded.coverage.get("publish_required") or [])
        for dataset in sorted(bundle.dataset_counts):
            source = bundle.source_contract.sources.get(dataset)
            policy = _coverage_policy(dataset, source_backed, optional_empty, source)
            source_row_count = source.row_count if source else None
            dataset_row_count = bundle.dataset_counts[dataset]
            rows.append(
                [
                    bundle.snapshot_date,
                    manifest.get("source_run_id", ""),
                    bundle.director,
                    bundle.territory,
                    dataset,
                    dataset_row_count,
                    policy,
                    dataset if source else "",
                    source.source_type if source else "",
                    source.source_id if source else "",
                    source.query_label if source else "",
                    source_row_count,
                    "Yes" if dataset in publish_required else "No",
                    _coverage_evidence(policy, dataset_row_count, source_row_count),
                ]
            )
    return rows


def metric_store_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        source_backed = set(loaded.coverage.get("source_backed") or [])
        optional_empty = set(loaded.coverage.get("optional_empty") or [])
        metric_rows = [
            *_dataset_count_metrics(bundle),
            (
                "pipeline_open_arr_unweighted",
                _sum_dataset(bundle.datasets.pipeline_open, "arr_unweighted"),
                bundle.corp_ccy,
                "pipeline_open",
                "Sum of pipeline_open.arr_unweighted",
            ),
            (
                "pipeline_open_arr_weighted",
                _sum_dataset(bundle.datasets.pipeline_open, "arr_weighted"),
                bundle.corp_ccy,
                "pipeline_open",
                "Sum of pipeline_open.arr_weighted",
            ),
            (
                "pi_current_arr_weighted",
                _sum_dataset(bundle.datasets.pi_current, "arr_weighted"),
                bundle.corp_ccy,
                "pi_current",
                "Sum of pi_current.arr_weighted",
            ),
            (
                "pi_current_priority_count",
                sum(1 for deal in bundle.datasets.pi_current if deal.priority),
                "count",
                "pi_current",
                "Count where pi_current.priority is true",
            ),
            (
                "pi_forward_arr_weighted",
                _sum_dataset(bundle.datasets.pi_forward, "arr_weighted"),
                bundle.corp_ccy,
                "pi_forward",
                "Sum of pi_forward.arr_weighted",
            ),
            (
                "snapshot_trend_latest_arr",
                _latest_snapshot_arr(bundle),
                bundle.corp_ccy,
                "snapshot_trend",
                "Sum of latest snapshot ARR by opportunity",
            ),
            (
                "source_backed_dataset_count",
                len(source_backed),
                "count",
                "source_coverage",
                "Count from manifest dataset_coverage.source_backed",
            ),
            (
                "optional_empty_dataset_count",
                len(optional_empty),
                "count",
                "source_coverage",
                "Count from manifest dataset_coverage.optional_empty",
            ),
        ]
        for metric_key, metric_value, unit, source_dataset, evidence in metric_rows:
            rows.append(
                [
                    bundle.snapshot_date,
                    manifest.get("source_run_id", ""),
                    bundle.director,
                    bundle.territory,
                    metric_key,
                    metric_value,
                    unit,
                    source_dataset,
                    evidence,
                ]
            )
    return rows


def executive_summary_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        source_backed_count = len(loaded.coverage.get("source_backed") or [])
        optional_empty_count = len(loaded.coverage.get("optional_empty") or [])
        unknown_count = sum(
            1
            for dataset in bundle.dataset_counts
            if _coverage_policy(
                dataset,
                set(loaded.coverage.get("source_backed") or []),
                set(loaded.coverage.get("optional_empty") or []),
                bundle.source_contract.sources.get(dataset),
            )
            == "unknown"
        )
        exception_counts = _exception_counts_for_bundle(loaded)
        total_exceptions = sum(exception_counts.values())
        rows.append(
            [
                bundle.snapshot_date,
                manifest.get("source_run_id", ""),
                bundle.territory,
                bundle.director,
                _source_status(unknown_count, total_exceptions),
                (
                    f"{source_backed_count} source-backed datasets; "
                    f"{optional_empty_count} optional-empty datasets; "
                    f"{unknown_count} unknown policies."
                ),
                (
                    f"{_plural_count(bundle.dataset_counts.get('pipeline_open', 0), 'open pipeline deal')}; "
                    f"{_money(_sum_dataset(bundle.datasets.pipeline_open, 'arr_weighted'), bundle.corp_ccy)} "
                    "weighted ARR."
                ),
                (
                    f"{_plural_count(bundle.dataset_counts.get('pi_current', 0), 'current-quarter PI deal')}; "
                    f"{_plural_count(bundle.dataset_counts.get('pi_forward', 0), 'forward-quarter PI deal')}."
                ),
                _format_exception_summary(exception_counts),
                (
                    "Use the rows in this workbook to confirm coverage, size the pipeline, "
                    "and decide which flagged deals need director follow-up."
                ),
                "Deterministic source-backed summary; no generated judgment or unsupported claim.",
            ]
        )
    return rows


def deal_exception_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        for deal in bundle.datasets.pipeline_open:
            common = [
                bundle.snapshot_date,
                manifest.get("source_run_id", ""),
                bundle.director,
                bundle.territory,
                "pipeline_open",
                deal.opportunity,
                deal.account,
                deal.owner,
                deal.stage,
                deal.forecast_category,
                deal.close_date,
                deal.arr_weighted,
                bundle.corp_ccy,
            ]
            if deal.push_count > 0:
                rows.append(
                    [
                        *common,
                        "Close-date movement",
                        "Deal has pushed at least once",
                        "medium",
                        deal.push_count,
                        "pipeline_open.push_count > 0",
                        "Check close-date confidence and next confirmed milestone.",
                    ]
                )
            if not deal.next_step.strip():
                rows.append(
                    [
                        *common,
                        "Execution hygiene",
                        "Missing next step",
                        "medium",
                        "",
                        "pipeline_open.next_step is blank",
                        "Ask owner for the next customer-facing action and due date.",
                    ]
                )
        for source_dataset, deals in [
            ("pi_current", bundle.datasets.pi_current),
            ("pi_forward", bundle.datasets.pi_forward),
        ]:
            for deal in deals:
                common = [
                    bundle.snapshot_date,
                    manifest.get("source_run_id", ""),
                    bundle.director,
                    bundle.territory,
                    source_dataset,
                    deal.opportunity,
                    "",
                    deal.owner,
                    deal.stage,
                    deal.forecast_category,
                    deal.close_date,
                    deal.arr_weighted,
                    deal.currency or bundle.corp_ccy,
                ]
                if deal.priority:
                    rows.append(
                        [
                            *common,
                            "Pipeline Inspection",
                            "PI priority flag",
                            "medium",
                            "TRUE",
                            f"{source_dataset}.priority is true",
                            "Review PI priority rationale before leadership readout.",
                        ]
                    )
                if deal.push_count > 0:
                    rows.append(
                        [
                            *common,
                            "Close-date movement",
                            "Deal has pushed at least once",
                            "medium",
                            deal.push_count,
                            f"{source_dataset}.push_count > 0",
                            "Check close-date confidence and next confirmed milestone.",
                        ]
                    )
    return sorted(
        rows,
        key=lambda row: (
            str(row[3]),
            str(row[4]),
            str(row[5]),
            str(row[13]),
            str(row[14]),
        ),
    )


def region_narrative_input_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
    exception_rows: list[list[Any]],
) -> list[list[Any]]:
    exception_count_by_territory: dict[str, int] = {}
    for row in exception_rows:
        exception_count_by_territory[row[3]] = exception_count_by_territory.get(row[3], 0) + 1

    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        source_backed = set(loaded.coverage.get("source_backed") or [])
        optional_empty = set(loaded.coverage.get("optional_empty") or [])
        inputs = [
            (
                "source_backed_dataset_count",
                len(source_backed),
                "count",
                "Source Coverage",
                "Count from manifest dataset_coverage.source_backed",
            ),
            (
                "optional_empty_dataset_count",
                len(optional_empty),
                "count",
                "Source Coverage",
                "Count from manifest dataset_coverage.optional_empty",
            ),
            (
                "pipeline_open_deal_count",
                bundle.dataset_counts.get("pipeline_open", 0),
                "count",
                "Metric Store",
                "Count of pipeline_open rows",
            ),
            (
                "pipeline_open_arr_weighted",
                _sum_dataset(bundle.datasets.pipeline_open, "arr_weighted"),
                bundle.corp_ccy,
                "Metric Store",
                "Sum of pipeline_open.arr_weighted",
            ),
            (
                "pipeline_open_arr_unweighted",
                _sum_dataset(bundle.datasets.pipeline_open, "arr_unweighted"),
                bundle.corp_ccy,
                "Metric Store",
                "Sum of pipeline_open.arr_unweighted",
            ),
            (
                "pi_current_deal_count",
                bundle.dataset_counts.get("pi_current", 0),
                "count",
                "Metric Store",
                "Count of pi_current rows",
            ),
            (
                "pi_current_arr_weighted",
                _sum_dataset(bundle.datasets.pi_current, "arr_weighted"),
                bundle.corp_ccy,
                "Metric Store",
                "Sum of pi_current.arr_weighted",
            ),
            (
                "pi_forward_deal_count",
                bundle.dataset_counts.get("pi_forward", 0),
                "count",
                "Metric Store",
                "Count of pi_forward rows",
            ),
            (
                "snapshot_trend_point_count",
                bundle.dataset_counts.get("snapshot_trend", 0),
                "count",
                "Metric Store",
                "Count of snapshot_trend rows",
            ),
            (
                "deal_exception_count",
                exception_count_by_territory.get(bundle.territory, 0),
                "count",
                "Deal Exceptions",
                "Count of deterministic deal exception rows",
            ),
        ]
        for input_key, input_value, unit, source_sheet, evidence in inputs:
            rows.append(
                [
                    bundle.snapshot_date,
                    manifest.get("source_run_id", ""),
                    bundle.territory,
                    bundle.director,
                    input_key,
                    input_value,
                    unit,
                    source_sheet,
                    evidence,
                ]
            )
    return rows


def analyst_notes_seed_rows(
    manifest: dict[str, Any],
    loaded_bundles: list[LoadedBundle],
    exception_rows: list[list[Any]],
) -> list[list[Any]]:
    del exception_rows

    rows: list[list[Any]] = []
    for loaded in loaded_bundles:
        bundle = loaded.bundle
        source_backed_count = len(loaded.coverage.get("source_backed") or [])
        optional_empty_count = len(loaded.coverage.get("optional_empty") or [])
        exception_counts = _exception_counts_for_bundle(loaded)
        notes = [
            (
                "coverage",
                (
                    f"Coverage is explicit: {source_backed_count} source-backed datasets, "
                    f"{optional_empty_count} optional-empty datasets, and no inferred sources."
                ),
                "Source Coverage",
            ),
            (
                "pipeline_open",
                (
                    f"Open pipeline shows {_plural_count(bundle.dataset_counts.get('pipeline_open', 0), 'deal')} "
                    f"and {_money(_sum_dataset(bundle.datasets.pipeline_open, 'arr_weighted'), bundle.corp_ccy)} "
                    "weighted ARR."
                ),
                "Metric Store",
            ),
            (
                "pipeline_inspection",
                (
                    f"Pipeline Inspection includes "
                    f"{_plural_count(bundle.dataset_counts.get('pi_current', 0), 'current-quarter deal')} "
                    f"and {_plural_count(bundle.dataset_counts.get('pi_forward', 0), 'forward-quarter deal')}."
                ),
                "Metric Store",
            ),
            (
                "review_flags",
                _format_exception_summary(exception_counts),
                "Deal Exceptions",
            ),
        ]
        for note_type, note_seed, source_sheet in notes:
            rows.append(
                [
                    bundle.snapshot_date,
                    manifest.get("source_run_id", ""),
                    bundle.territory,
                    bundle.director,
                    note_type,
                    note_seed,
                    source_sheet,
                    "Deterministic seed from source-backed rows; analyst must verify before publishing.",
                ]
            )
    return rows


def _headers_by_sheet() -> dict[str, list[str]]:
    return {
        "Executive Summary": [
            "Snapshot Date",
            "Source Run ID",
            "Territory",
            "Director",
            "Source Status",
            "Coverage Summary",
            "Pipeline Summary",
            "PI Summary",
            "Exception Summary",
            "Analyst Focus",
            "Guardrail",
        ],
        "Source Coverage": [
            "Snapshot Date",
            "Source Run ID",
            "Director",
            "Territory",
            "Dataset",
            "Dataset Row Count",
            "Coverage Policy",
            "Source Key",
            "Source Type",
            "Source ID",
            "Source Label",
            "Source Row Count",
            "Publish Required",
            "Evidence",
        ],
        "Metric Store": [
            "Snapshot Date",
            "Source Run ID",
            "Director",
            "Territory",
            "Metric Key",
            "Metric Value",
            "Unit",
            "Source Dataset",
            "Evidence",
        ],
        "Deal Exceptions": [
            "Snapshot Date",
            "Source Run ID",
            "Director",
            "Territory",
            "Source Dataset",
            "Opportunity",
            "Account",
            "Owner",
            "Stage",
            "Forecast Category",
            "Close Date",
            "ARR Weighted",
            "Currency",
            "Exception Category",
            "Exception Type",
            "Severity",
            "Evidence Value",
            "Deterministic Rule",
            "Analyst Action",
        ],
        "Region Narrative Inputs": [
            "Snapshot Date",
            "Source Run ID",
            "Territory",
            "Director",
            "Input Key",
            "Input Value",
            "Unit",
            "Source Sheet",
            "Evidence",
        ],
        "Analyst Notes Seed": [
            "Snapshot Date",
            "Source Run ID",
            "Territory",
            "Director",
            "Note Type",
            "Note Seed",
            "Source Sheet",
            "Guardrail",
        ],
    }


def _numeric_columns(sheet_name: str, headers: list[str]) -> set[int]:
    numeric_headers = {
        "Dataset Row Count",
        "Source Row Count",
        "Metric Value",
        "ARR Weighted",
        "Input Value",
    }
    if sheet_name == "Deal Exceptions":
        numeric_headers.add("Evidence Value")
    return {
        index
        for index, header in enumerate(headers, start=1)
        if header in numeric_headers
    }


def _add_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    numeric_columns: set[int],
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(rows, 2):
        for column_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = DATA_FONT
            if column_index in numeric_columns and isinstance(value, int | float):
                cell.number_format = _number_format_for_cell(
                    headers[column_index - 1],
                    headers,
                    row,
                )
            if isinstance(value, str) and len(value) > 45:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22
    if rows:
        end_col = get_column_letter(len(headers))
        table = Table(
            displayName=_safe_table_name(sheet_name),
            ref=f"A1:{end_col}{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
        )
        ws.add_table(table)
    for column_index, header in enumerate(headers, 1):
        values = [header, *[row[column_index - 1] for row in rows[:100]]]
        width = min(max(len(str(value)) for value in values) + 3, 64)
        ws.column_dimensions[get_column_letter(column_index)].width = max(width, 12)
    for row_index in range(2, len(rows) + 2):
        ws.row_dimensions[row_index].height = 32 if sheet_name in {
            "Executive Summary",
            "Analyst Notes Seed",
        } else 18


def _coverage_by_territory(manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    coverage: dict[str, dict[str, Any]] = {}
    for director_summary in manifest.get("summary", {}).get("directors", []):
        territory = director_summary.get("territory")
        if territory:
            coverage[str(territory)] = dict(director_summary.get("dataset_coverage") or {})
    return coverage


def _coverage_policy(
    dataset: str,
    source_backed: set[str],
    optional_empty: set[str],
    source: Any | None,
) -> str:
    if dataset in source_backed:
        return "source_backed"
    if dataset in optional_empty:
        return "optional_empty"
    if source:
        return "source_backed"
    return "unknown"


def _coverage_evidence(
    policy: str,
    dataset_row_count: int,
    source_row_count: int | None,
) -> str:
    if source_row_count is None:
        if policy == "optional_empty" and dataset_row_count == 0:
            return "Optional-empty dataset has 0 rows."
        return "No source metadata in bundle source_contract."
    if source_row_count == dataset_row_count:
        return "Dataset row count matches source row count."
    return "Dataset row count differs from source row count."


def _exception_counts_for_bundle(loaded: LoadedBundle) -> dict[str, int]:
    counts = {
        "Close-date movement": 0,
        "Execution hygiene": 0,
        "Pipeline Inspection": 0,
    }
    bundle = loaded.bundle
    for deal in bundle.datasets.pipeline_open:
        if deal.push_count > 0:
            counts["Close-date movement"] += 1
        if not deal.next_step.strip():
            counts["Execution hygiene"] += 1
    for deals in (bundle.datasets.pi_current, bundle.datasets.pi_forward):
        for deal in deals:
            if deal.priority:
                counts["Pipeline Inspection"] += 1
            if deal.push_count > 0:
                counts["Close-date movement"] += 1
    return counts


def _format_exception_summary(counts: dict[str, int]) -> str:
    total = sum(counts.values())
    return (
        f"{_plural_count(total, 'review flag')}: "
        f"{counts.get('Close-date movement', 0)} close-date movement, "
        f"{counts.get('Execution hygiene', 0)} execution hygiene, "
        f"{counts.get('Pipeline Inspection', 0)} Pipeline Inspection."
    )


def _source_status(unknown_count: int, total_exceptions: int) -> str:
    if unknown_count:
        return "Needs source coverage review"
    if total_exceptions:
        return "Ready with deal review flags"
    return "Ready for analyst review"


def _dataset_count_metrics(bundle: DirectorBundle) -> list[tuple[str, int, str, str, str]]:
    return [
        (
            f"dataset_count.{dataset}",
            int(bundle.dataset_counts[dataset]),
            "count",
            dataset,
            f"Row count from DirectorBundle.dataset_counts.{dataset}",
        )
        for dataset in sorted(bundle.dataset_counts)
    ]


def _sum_dataset(rows: list[Any], attribute: str) -> float:
    return float(sum(float(getattr(row, attribute, 0) or 0) for row in rows))


def _latest_snapshot_arr(bundle: DirectorBundle) -> float:
    latest_by_opportunity: dict[str, tuple[str, float]] = {}
    for snapshot in bundle.datasets.snapshot_trend:
        prior = latest_by_opportunity.get(snapshot.opportunity)
        if prior is None or snapshot.snapshot_date > prior[0]:
            latest_by_opportunity[snapshot.opportunity] = (
                snapshot.snapshot_date,
                float(snapshot.arr_at_snapshot or 0),
            )
    return float(sum(value for _, value in latest_by_opportunity.values()))


def _resolve_manifest_path(path_text: str, manifest_path: Path) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    if path.exists():
        return path
    repo_path = REPO_ROOT / path
    if repo_path.exists():
        return repo_path
    return manifest_path.parent / path


def _safe_table_name(sheet_name: str) -> str:
    return "".join(char for char in sheet_name.title() if char.isalnum())[:30]


def _number_format_for_cell(header: str, headers: list[str], row: list[Any]) -> str:
    if "ARR" in header:
        return MONEY_FORMAT
    if header in {"Metric Value", "Input Value"}:
        try:
            unit = str(row[headers.index("Unit")])
        except ValueError:
            unit = ""
        return COUNT_FORMAT if unit == "count" else MONEY_FORMAT
    return COUNT_FORMAT


def _plural_count(count: int, singular: str) -> str:
    noun = singular if count == 1 else f"{singular}s"
    return f"{count:,} {noun}"


def _money(value: float, currency: str) -> str:
    return f"{currency} {value:,.0f}"


def _set_stable_properties(wb: Workbook, manifest: dict[str, Any]) -> None:
    snapshot_date = str(manifest.get("snapshot_date") or "2000-01-01")
    created = datetime.fromisoformat(f"{snapshot_date}T00:00:00+00:00")
    wb.properties.creator = "monthly_platform.analyst_workbook"
    wb.properties.lastModifiedBy = "monthly_platform.analyst_workbook"
    wb.properties.title = "Source-backed Analyst Workbook"
    wb.properties.subject = ANALYST_WORKBOOK_SCHEMA_VERSION
    wb.properties.created = created
    wb.properties.modified = created.astimezone(timezone.utc).replace(tzinfo=None)


def result_to_json(result: AnalystWorkbookResult) -> str:
    payload = asdict(result)
    payload["workbook_path"] = str(result.workbook_path)
    return json.dumps(payload, indent=2, sort_keys=True)
