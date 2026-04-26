# scripts/monthly_platform/excel_renderer.py
"""Render a DirectorBundle to an Excel workbook.

Pure function: DirectorBundle in, Excel workbook out.
No Salesforce calls. No territory resolution. No sidecar logic.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

if TYPE_CHECKING:
    from .models import DirectorBundle

try:
    from .intelligence import as_rows, build_deal_risk_table
except ImportError:  # pragma: no cover
    from scripts.monthly_platform.intelligence import as_rows, build_deal_risk_table

HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=9)
EUR_FMT = "#,##0"


def _add_sheet(wb, name, headers, rows, eur_cols=None):
    ws = wb.create_sheet(title=name[:31])
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            if eur_cols and ci in eur_cols and isinstance(val, (int, float)):
                cell.number_format = EUR_FMT
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(headers[ci - 1])),
            *(len(str(r[ci - 1])) for r in rows[:50]) if rows else [0],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
    if rows:
        end_col = get_column_letter(len(headers))
        table_name = (
            name.replace(" ", "_")
            .replace("-", "_")
            .replace("&", "And")
            .replace("/", "")[:30]
        )
        try:
            table = Table(displayName=table_name, ref=f"A1:{end_col}{len(rows) + 1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
            )
            ws.add_table(table)
        except Exception:
            pass
    ws.freeze_panes = "A2"


def render_bundle_to_excel(bundle: DirectorBundle, output_path: Path) -> None:
    analysis_year = int(bundle.snapshot_date[:4])
    fy = f"FY{analysis_year % 100:02d}"
    ccy = bundle.corp_ccy
    ds = bundle.datasets

    wb = Workbook()
    wb.remove(wb.active)

    # -- Pipeline Open --
    _add_sheet(
        wb,
        f"Pipeline Open {fy}",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Forecast Category",
            "Close Date",
            f"ARR Unweighted ({ccy})",
            f"ARR Weighted ({ccy})",
            "Probability %",
            "Push Count",
            "Type",
            "Lead Scope",
            "Industry",
            "Tier",
            "Sales Region",
            "Created",
            "Last Activity",
            "Next Step",
            "Last Modified",
            "Approved",
            "Approval Date",
            "Competitor",
        ],
        [
            [
                d.account,
                d.opportunity,
                d.owner,
                d.stage,
                d.forecast_category,
                d.close_date,
                d.arr_unweighted,
                d.arr_weighted,
                d.probability,
                d.push_count,
                d.deal_type,
                d.lead_scope,
                d.industry,
                d.tier,
                d.sales_region,
                d.created_date,
                d.last_activity_date or "",
                d.next_step,
                d.last_modified_date,
                "Yes" if d.approved else "No",
                d.approval_date or "",
                d.competitor,
            ]
            for d in ds.pipeline_open
        ],
        eur_cols={7, 8},
    )

    # -- Won Lost --
    _add_sheet(
        wb,
        f"Won Lost {fy}",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            f"ARR Unweighted ({ccy})",
            "Type",
            "Reason",
            "Lost To Competitor",
            "Industry",
            "Sales Region",
            "Created",
        ],
        [
            [
                d.account,
                d.opportunity,
                d.owner,
                d.stage,
                d.close_date,
                d.arr_unweighted,
                d.deal_type,
                d.reason_won_lost,
                d.competitor,
                d.industry,
                d.sales_region,
                d.created_date,
            ]
            for d in ds.won_lost
        ],
        eur_cols={6},
    )

    # -- Commercial Approval --
    _add_sheet(
        wb,
        "Commercial Approval",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            f"ARR Unweighted ({ccy})",
            "Status",
            "Approval Date",
            "Next Step",
            "Lead Scope",
        ],
        [
            [
                d.account,
                d.opportunity,
                d.owner,
                d.stage,
                d.close_date,
                d.arr_unweighted,
                d.status,
                d.approval_date or "",
                d.next_step,
                d.lead_scope,
            ]
            for d in ds.approvals
        ],
        eur_cols={6},
    )

    # -- Renewals --
    _add_sheet(
        wb,
        f"Renewals {fy}",
        [
            "Close Date",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            f"ACV Unweighted ({ccy})",
            "Probability %",
            "Comments",
        ],
        [
            [
                d.close_date,
                d.account,
                d.opportunity,
                d.owner,
                d.stage,
                d.acv_unweighted,
                d.probability,
                d.comments,
            ]
            for d in ds.renewals
        ],
        eur_cols={6},
    )

    # -- Pipeline Inspection --
    pi_headers = [
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "ARR Weighted (native ccy)",
        "Currency",
        "Close Date",
        "Push Count",
        "Score",
        "Priority",
    ]
    _add_sheet(
        wb,
        "Pipeline Inspection",
        pi_headers,
        [
            [
                d.opportunity,
                d.owner,
                d.stage,
                d.forecast_category,
                d.arr_weighted,
                d.currency,
                d.close_date,
                d.push_count,
                d.score,
                "Yes" if d.priority else "",
            ]
            for d in ds.pi_current
        ],
        eur_cols={5},
    )

    # -- Pipeline Inspection Forward --
    if ds.pi_forward:
        _add_sheet(
            wb,
            "Pipeline Inspection Forward",
            pi_headers,
            [
                [
                    d.opportunity,
                    d.owner,
                    d.stage,
                    d.forecast_category,
                    d.arr_weighted,
                    d.currency,
                    d.close_date,
                    d.push_count,
                    d.score,
                    "Yes" if d.priority else "",
                ]
                for d in ds.pi_forward
            ],
            eur_cols={5},
        )

    # -- Activity Volume --
    activity_sorted = sorted(
        ds.activity, key=lambda a: (a.total_touches_90d, a.last_activity_date or "")
    )
    _add_sheet(
        wb,
        "Activity Volume",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Tasks 90d",
            "Events 90d",
            "Total Touches 90d",
            "Last Activity",
            "Flag",
        ],
        [
            [
                a.account,
                a.opportunity,
                a.owner,
                a.tasks_90d,
                a.events_90d,
                a.total_touches_90d,
                a.last_activity_date or "",
                a.flag,
            ]
            for a in activity_sorted
        ],
    )

    # -- Commit Items --
    commit_sorted = sorted(ds.commit_items, key=lambda c: -(c.arr_weighted or 0))
    _add_sheet(
        wb,
        "Commit Items",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Forecast Category",
            f"Forecast ARR Wtd ({ccy})",
            f"ARR Unwtd ({ccy})",
            "Close Date",
            "Period",
            "Stage",
        ],
        [
            [
                c.account,
                c.opportunity,
                c.owner,
                c.forecast_category,
                c.arr_weighted,
                c.arr_unweighted,
                c.close_date,
                c.period,
                c.stage,
            ]
            for c in commit_sorted
        ],
        eur_cols={5, 6},
    )

    # -- Q1 Movement --
    movement_headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Movement",
        "Old Close",
        "New Close",
        "Changed On",
        f"ARR Unweighted ({ccy})",
    ]
    q1_sorted = sorted(ds.movement_prior, key=lambda m: -(m.arr_unweighted or 0))
    _add_sheet(
        wb,
        "Q1 Movement",
        movement_headers,
        [
            [
                m.account,
                m.opportunity,
                m.owner,
                m.stage,
                m.movement_type,
                m.old_close,
                m.new_close,
                m.changed_on,
                m.arr_unweighted,
            ]
            for m in q1_sorted
        ],
        eur_cols={9},
    )

    # -- Q2 Movement --
    q2_sorted = sorted(ds.movement_current, key=lambda m: -(m.arr_unweighted or 0))
    _add_sheet(
        wb,
        "Q2 Movement",
        movement_headers,
        [
            [
                m.account,
                m.opportunity,
                m.owner,
                m.stage,
                m.movement_type,
                m.old_close,
                m.new_close,
                m.changed_on,
                m.arr_unweighted,
            ]
            for m in q2_sorted
        ],
        eur_cols={9},
    )

    # -- Stage History --
    stage_sorted = sorted(ds.stage_events, key=lambda s: s.created_date, reverse=True)
    _add_sheet(
        wb,
        "Stage History",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage (live)",
            "From Stage",
            "To Stage",
            "Changed On",
            f"ARR Unweighted ({ccy})",
        ],
        [
            [
                s.account,
                s.opportunity,
                s.owner,
                s.current_stage,
                s.old_value,
                s.new_value,
                s.created_date,
                s.arr_unweighted,
            ]
            for s in stage_sorted
        ],
        eur_cols={8},
    )

    # -- Forecast Category History --
    fcat_sorted = sorted(
        ds.forecast_category_events, key=lambda f: f.created_date, reverse=True
    )
    _add_sheet(
        wb,
        "Forecast Category History",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage (live)",
            "From Category",
            "To Category",
            "Changed On",
            f"ARR Unweighted ({ccy})",
        ],
        [
            [
                f.account,
                f.opportunity,
                f.owner,
                f.current_stage,
                f.old_value,
                f.new_value,
                f.created_date,
                f.arr_unweighted,
            ]
            for f in fcat_sorted
        ],
        eur_cols={8},
    )

    # -- Close Date History --
    close_date_sorted = sorted(
        ds.close_date_events, key=lambda c: c.created_date, reverse=True
    )
    _add_sheet(
        wb,
        "Close Date History",
        [
            "Account",
            "Opportunity",
            "Owner",
            "Stage (live)",
            "Old Close",
            "New Close",
            "Changed On",
            f"ARR Unweighted ({ccy})",
            "Closed",
        ],
        [
            [
                c.account,
                c.opportunity,
                c.owner,
                c.current_stage,
                c.old_value,
                c.new_value,
                c.created_date,
                c.arr_unweighted,
                "Yes" if c.is_closed else "No",
            ]
            for c in close_date_sorted
        ],
        eur_cols={8},
    )

    # -- Deal Risk Index --
    risk_rows = build_deal_risk_table(as_rows(bundle), limit=None)
    _add_sheet(
        wb,
        "Deal Risk Index",
        [
            "Risk Score",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Forecast Category",
            f"ARR Unweighted ({ccy})",
            "Close Date",
            "Risk Reasons",
        ],
        [
            [
                r["risk_score"],
                r["account"],
                r["opportunity"],
                r["owner"],
                r["stage"],
                r["forecast_category"],
                r["arr_unweighted"],
                r["close_date"],
                "; ".join(r["risk_reasons"]),
            ]
            for r in risk_rows
        ],
        eur_cols={7},
    )

    # -- Summary (first tab, built last) --
    _build_summary(wb, bundle, fy, ccy, analysis_year)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def _build_summary(wb, bundle, fy, ccy, analysis_year):
    ds = bundle.datasets
    ws = wb.create_sheet(title="Summary", index=0)

    ws["A1"] = f"{bundle.director} ({bundle.territory})"
    ws["A1"].font = Font(bold=True, size=14, color="083EA7")
    ws["A2"] = f"Reporting period: {fy} (Q1-Q4)"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A3"] = f"Snapshot date: {bundle.snapshot_date} -- live pull from Salesforce"
    ws["A3"].font = Font(size=10, color="666666")
    ws["A4"] = (
        "Methodology: Alex P -- ARR Unweighted = APTS_Opportunity_ARR__c (full deal value); "
        "ARR Weighted = APTS_Forecast_ARR__c (probability-weighted). "
        "Excl simcorp/test/delete accounts, excl Sabiniewicz/Profit owners."
    )
    ws["A4"].font = Font(size=8, italic=True, color="999999")

    total_pipeline_arr = sum(d.arr_unweighted for d in ds.pipeline_open)
    won = [d for d in ds.won_lost if "Won" in d.stage]
    lost = [d for d in ds.won_lost if "Lost" in d.stage or "Opt Out" in d.stage]
    won_arr = sum(d.arr_unweighted for d in won)
    lost_arr = sum(d.arr_unweighted for d in lost)
    renewal_acv = sum(d.acv_unweighted for d in ds.renewals)

    approved_current = [
        d for d in ds.approvals if d.status.startswith(f"Approved {analysis_year}")
    ]
    approved_prior = [d for d in ds.approvals if d.status == "Approved (prior year)"]
    pending = [d for d in ds.approvals if d.status == "Pending Approval"]
    missing = [d for d in ds.approvals if d.status == "Missing (Stage 3+)"]

    ws["A6"] = "KPI"
    ws["B6"] = "Value"
    ws["A6"].font = HEADER_FONT
    ws["A6"].fill = HEADER_FILL
    ws["B6"].font = HEADER_FONT
    ws["B6"].fill = HEADER_FILL

    kpis = [
        ("Open Pipeline Unweighted (stages 1-6)", f"{ccy} {total_pipeline_arr:,.0f}"),
        ("Open Deal Count", str(len(ds.pipeline_open))),
        (f"Won ARR Unweighted {fy}", f"{ccy} {won_arr:,.0f}"),
        ("Won Deal Count", str(len(won))),
        (f"Lost ARR Unweighted {fy}", f"{ccy} {lost_arr:,.0f}"),
        ("Lost Deal Count", str(len(lost))),
        (f"Approved {analysis_year} (Land)", str(len(approved_current))),
        ("Approved Prior Year", str(len(approved_prior))),
        ("Pending Approval", str(len(pending))),
        ("Missing Approval (Stage 3+)", str(len(missing))),
        ("Open Renewal ACV Unweighted", f"{ccy} {renewal_acv:,.0f}"),
        ("Open Renewals", str(len(ds.renewals))),
        (f"PI Open Deals ({fy})", str(len(ds.pi_current))),
    ]
    if ds.pi_forward:
        kpis.append(("PI Forward Deals", str(len(ds.pi_forward))))

    for i, (label, val) in enumerate(kpis, 7):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val

    sheet_row = len(kpis) + 8
    ws[f"A{sheet_row}"] = "Sheet"
    ws[f"B{sheet_row}"] = "Records"
    ws[f"C{sheet_row}"] = "Source"
    for col in ("A", "B", "C"):
        ws[f"{col}{sheet_row}"].font = HEADER_FONT
        ws[f"{col}{sheet_row}"].fill = HEADER_FILL

    sheets_info = [
        (
            f"Pipeline Open {fy}",
            len(ds.pipeline_open),
            f"SOQL — open, stages 1-6, {fy}",
        ),
        (f"Won Lost {fy}", len(ds.won_lost), f"SOQL — closed, stages 0/7/8, {fy}"),
        ("Commercial Approval", len(ds.approvals), f"SOQL — open Land, {fy}"),
        (f"Renewals {fy}", len(ds.renewals), f"SOQL — open Renewal, {fy}"),
        (
            "Pipeline Inspection",
            len(ds.pi_current),
            f"PI list view — broad coaching population, {fy}",
        ),
    ]
    if ds.pi_forward:
        sheets_info.append(
            (
                "Pipeline Inspection Forward",
                len(ds.pi_forward),
                "PI list view — forward quarter",
            )
        )
    sheets_info.extend(
        [
            ("Activity Volume", len(ds.activity), "SOQL — tasks/events 90d"),
            ("Commit Items", len(ds.commit_items), "SOQL — forecast category open items"),
            ("Q1 Movement", len(ds.movement_prior), "SOQL — prior-quarter movement"),
            ("Q2 Movement", len(ds.movement_current), "SOQL — current-quarter movement"),
            ("Stage History", len(ds.stage_events), "SOQL — OpportunityFieldHistory StageName"),
            (
                "Forecast Category History",
                len(ds.forecast_category_events),
                "SOQL — OpportunityFieldHistory ForecastCategoryName",
            ),
            (
                "Close Date History",
                len(ds.close_date_events),
                "SOQL — OpportunityFieldHistory CloseDate",
            ),
            (
                "Deal Risk Index",
                len(build_deal_risk_table(as_rows(bundle), limit=None)),
                "Gold analytics — joined pipeline/activity/history risk score",
            ),
        ]
    )

    for i, (sname, count, source) in enumerate(sheets_info, sheet_row + 1):
        ws[f"A{i}"] = sname
        ws[f"B{i}"] = count
        ws[f"C{i}"] = source

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
