"""Build think-cell handoff artifacts from source-backed DirectorBundle manifests.

This module writes only local Excel/JSON artifacts. It does not require
Windows, PowerPoint, think-cell, Salesforce access, or CRM Analytics APIs.
"""

from __future__ import annotations

import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCHEMA_VERSION = "monthly_platform.thinkcell_source.v1"
DEFAULT_TEMPLATE_PATH = "sales_director_thinkcell_template.pptx"

CONTROL_COLUMNS = ["field", "value"]
SOURCE_COUNT_COLUMNS = [
    "metric_id",
    "director",
    "territory",
    "dataset",
    "row_count",
    "source_contract_row_count",
    "source_type",
    "source_id",
    "query_label",
    "bundle_path",
    "source_json_path",
]
COVERAGE_COLUMNS = [
    "director",
    "territory",
    "dataset",
    "policy",
    "required_for_publish",
    "row_count",
    "source_contract_present",
    "covered",
]
METRIC_STORE_COLUMNS = [
    "metric_id",
    "scope",
    "label",
    "value",
    "unit",
    "director",
    "territory",
    "dataset",
    "source_artifact",
    "source_json_path",
]
PPTTC_MAP_COLUMNS = [
    "element_name",
    "element_type",
    "sheet",
    "table_name",
    "description",
]

TABLE_NAMES = {
    "Control": "TC_Control",
    "Source Counts": "TC_SourceCounts",
    "Coverage": "TC_Coverage",
    "Metric Store": "TC_MetricStore",
    "PPTTC Map": "TC_PpttcMap",
}


def build_thinkcell_source_from_manifest(
    *,
    manifest_path: Path,
    output_dir: Path,
    template_path: str = DEFAULT_TEMPLATE_PATH,
) -> dict[str, Any]:
    manifest_path = Path(manifest_path)
    output_dir = Path(output_dir)
    manifest = _load_json(manifest_path)
    bundles = _load_manifest_bundles(manifest_path, manifest)
    payload = build_thinkcell_payload(
        manifest=manifest,
        manifest_path=manifest_path,
        bundles=bundles,
        template_path=template_path,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "thinkcell_source.xlsx"
    ppttc_path = output_dir / "thinkcell_data.ppttc"
    write_workbook(workbook_path, payload)
    write_ppttc(ppttc_path, payload)

    return {
        "schema_version": SCHEMA_VERSION,
        "snapshot_date": payload["snapshot_date"],
        "source_run_id": payload["source_run_id"],
        "status": payload["status"],
        "workbook_path": str(workbook_path),
        "ppttc_path": str(ppttc_path),
        "summary": payload["summary"],
    }


def build_thinkcell_payload(
    *,
    manifest: dict[str, Any],
    manifest_path: Path,
    bundles: list[dict[str, Any]],
    template_path: str = DEFAULT_TEMPLATE_PATH,
) -> dict[str, Any]:
    source_counts: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []

    for bundle_entry in bundles:
        bundle = bundle_entry["bundle"]
        bundle_path = bundle_entry["bundle_path"]
        coverage = bundle_entry["coverage"]
        source_counts.extend(
            _source_count_rows(bundle=bundle, bundle_path=bundle_path, coverage=coverage)
        )
        coverage_rows.extend(_coverage_rows(bundle=bundle, coverage=coverage))

    metrics = [_metric_row(row) for row in source_counts]
    summary = {
        "director_count": len(bundles),
        "source_backed_dataset_count": len(source_counts),
        "source_backed_row_count": sum(row["row_count"] for row in source_counts),
        "coverage_row_count": len(coverage_rows),
        "metric_count": len(metrics),
    }
    control_rows = _control_rows(
        manifest=manifest,
        manifest_path=manifest_path,
        summary=summary,
    )
    ppttc_map_rows = _ppttc_map_rows()

    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(UTC).isoformat(),
        "snapshot_date": manifest.get("snapshot_date", ""),
        "source_run_id": manifest.get("source_run_id", ""),
        "status": manifest.get("status", ""),
        "template_path": template_path,
        "control": control_rows,
        "source_counts": source_counts,
        "coverage": coverage_rows,
        "metrics": metrics,
        "ppttc_map": ppttc_map_rows,
        "summary": summary,
    }


def write_workbook(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _add_table_sheet(
        wb,
        name="Control",
        table_name=TABLE_NAMES["Control"],
        columns=CONTROL_COLUMNS,
        rows=payload["control"],
    )
    _add_table_sheet(
        wb,
        name="Source Counts",
        table_name=TABLE_NAMES["Source Counts"],
        columns=SOURCE_COUNT_COLUMNS,
        rows=payload["source_counts"],
    )
    _add_table_sheet(
        wb,
        name="Coverage",
        table_name=TABLE_NAMES["Coverage"],
        columns=COVERAGE_COLUMNS,
        rows=payload["coverage"],
    )
    _add_table_sheet(
        wb,
        name="Metric Store",
        table_name=TABLE_NAMES["Metric Store"],
        columns=METRIC_STORE_COLUMNS,
        rows=payload["metrics"],
    )
    _add_table_sheet(
        wb,
        name="PPTTC Map",
        table_name=TABLE_NAMES["PPTTC Map"],
        columns=PPTTC_MAP_COLUMNS,
        rows=payload["ppttc_map"],
    )
    wb.save(path)


def write_ppttc(path: Path, payload: dict[str, Any]) -> None:
    data = [
        _tc_text("ThinkCellSourceStatus", payload["status"]),
        _tc_text("ThinkCellSourceMetricCount", payload["summary"]["metric_count"]),
        _tc_text(
            "ThinkCellSourceBackedRows",
            payload["summary"]["source_backed_row_count"],
        ),
        {
            "name": "SourceBackedCountsTable",
            "table": _tc_table(payload["source_counts"], SOURCE_COUNT_COLUMNS),
        },
        {
            "name": "CoverageTable",
            "table": _tc_table(payload["coverage"], COVERAGE_COLUMNS),
        },
        {
            "name": "MetricStoreTable",
            "table": _tc_table(payload["metrics"], METRIC_STORE_COLUMNS),
        },
    ]
    ppttc_payload = [
        {
            "template": payload["template_path"],
            "data": data,
            "metadata": {
                "schema_version": SCHEMA_VERSION,
                "snapshot_date": payload["snapshot_date"],
                "source_run_id": payload["source_run_id"],
                "requires_think_cell_to_generate": False,
            },
        }
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(ppttc_payload, indent=2) + "\n", encoding="utf-8")


def latest_manifest_path(source_root: Path, snapshot_date: str, run_id: str | None) -> Path:
    snapshot_root = Path(source_root) / snapshot_date
    if run_id:
        candidate = snapshot_root / run_id / "director_bundle_manifest.json"
        if not candidate.exists():
            raise FileNotFoundError(f"DirectorBundle manifest not found: {candidate}")
        return candidate
    candidates = [
        path / "director_bundle_manifest.json"
        for path in snapshot_root.iterdir()
        if path.is_dir() and (path / "director_bundle_manifest.json").exists()
    ]
    if not candidates:
        raise FileNotFoundError(f"No DirectorBundle manifests found under {snapshot_root}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def default_output_dir(output_root: Path, manifest: dict[str, Any]) -> Path:
    snapshot_date = str(manifest.get("snapshot_date") or "unknown-snapshot")
    source_run_id = str(manifest.get("source_run_id") or "unknown-run")
    return Path(output_root) / snapshot_date / source_run_id


def _load_manifest_bundles(
    manifest_path: Path,
    manifest: dict[str, Any],
) -> list[dict[str, Any]]:
    summaries = manifest.get("summary", {}).get("directors", []) or []
    bundles: list[dict[str, Any]] = []
    for index, path_text in enumerate(manifest.get("bundle_paths") or []):
        bundle_path = _resolve_manifest_ref(path_text, manifest_path)
        bundle = _load_json(bundle_path)
        bundles.append(
            {
                "bundle_path": str(bundle_path),
                "bundle": bundle,
                "coverage": _coverage_for_bundle(bundle, summaries, index),
            }
        )
    return bundles


def _coverage_for_bundle(
    bundle: dict[str, Any],
    summaries: list[dict[str, Any]],
    index: int,
) -> dict[str, Any]:
    director = bundle.get("director")
    territory = bundle.get("territory")
    for summary in summaries:
        if summary.get("director") == director and summary.get("territory") == territory:
            return dict(summary.get("dataset_coverage") or {})
    for summary in summaries:
        if summary.get("territory") == territory:
            return dict(summary.get("dataset_coverage") or {})
    if index < len(summaries):
        return dict(summaries[index].get("dataset_coverage") or {})
    source_keys = set((bundle.get("source_contract", {}).get("sources") or {}).keys())
    dataset_names = set((bundle.get("dataset_counts") or {}).keys())
    source_backed = sorted(dataset_names & (source_keys - {"source_bundle"}))
    return {
        "source_backed": source_backed,
        "optional_empty": sorted(dataset_names - set(source_backed)),
        "publish_required": [],
        "source_requirement_ids": [],
    }


def _source_count_rows(
    *,
    bundle: dict[str, Any],
    bundle_path: str,
    coverage: dict[str, Any],
) -> list[dict[str, Any]]:
    director = str(bundle.get("director") or "")
    territory = str(bundle.get("territory") or "")
    dataset_counts = bundle.get("dataset_counts") or {}
    sources = bundle.get("source_contract", {}).get("sources") or {}
    rows: list[dict[str, Any]] = []
    for dataset in sorted(coverage.get("source_backed") or []):
        source = sources.get(dataset) or {}
        rows.append(
            {
                "metric_id": _metric_id(territory, director, dataset),
                "director": director,
                "territory": territory,
                "dataset": dataset,
                "row_count": int(dataset_counts.get(dataset) or 0),
                "source_contract_row_count": (
                    int(source["row_count"]) if source.get("row_count") is not None else None
                ),
                "source_type": source.get("source_type"),
                "source_id": source.get("source_id"),
                "query_label": source.get("query_label"),
                "bundle_path": bundle_path,
                "source_json_path": f"$.dataset_counts.{dataset}",
            }
        )
    return rows


def _coverage_rows(
    *,
    bundle: dict[str, Any],
    coverage: dict[str, Any],
) -> list[dict[str, Any]]:
    director = str(bundle.get("director") or "")
    territory = str(bundle.get("territory") or "")
    dataset_counts = bundle.get("dataset_counts") or {}
    sources = bundle.get("source_contract", {}).get("sources") or {}
    source_backed = set(coverage.get("source_backed") or [])
    optional_empty = set(coverage.get("optional_empty") or [])
    publish_required = set(coverage.get("publish_required") or [])
    known_datasets = sorted(set(dataset_counts) | source_backed | optional_empty)
    rows: list[dict[str, Any]] = []
    for dataset in known_datasets:
        if dataset in source_backed:
            policy = "source_backed"
        elif dataset in optional_empty:
            policy = "optional_empty"
        else:
            policy = "unclassified"
        row_count = int(dataset_counts.get(dataset) or 0)
        source_contract_present = dataset in sources
        if policy == "source_backed":
            covered = source_contract_present
        elif policy == "optional_empty":
            covered = row_count == 0
        else:
            covered = False
        rows.append(
            {
                "director": director,
                "territory": territory,
                "dataset": dataset,
                "policy": policy,
                "required_for_publish": dataset in publish_required,
                "row_count": row_count,
                "source_contract_present": source_contract_present,
                "covered": covered,
            }
        )
    return rows


def _metric_row(source_count_row: dict[str, Any]) -> dict[str, Any]:
    territory = source_count_row["territory"]
    dataset = source_count_row["dataset"]
    return {
        "metric_id": source_count_row["metric_id"],
        "scope": "source_backed_director_bundle",
        "label": f"{territory} {dataset} row count",
        "value": source_count_row["row_count"],
        "unit": "rows",
        "director": source_count_row["director"],
        "territory": territory,
        "dataset": dataset,
        "source_artifact": source_count_row["bundle_path"],
        "source_json_path": source_count_row["source_json_path"],
    }


def _control_rows(
    *,
    manifest: dict[str, Any],
    manifest_path: Path,
    summary: dict[str, Any],
) -> list[dict[str, Any]]:
    source_requirement_ids = sorted(
        {
            str(requirement_id)
            for director in manifest.get("summary", {}).get("directors", []) or []
            for requirement_id in (
                director.get("dataset_coverage", {}).get("source_requirement_ids") or []
            )
        }
    )
    rows = [
        ("schema_version", SCHEMA_VERSION),
        ("snapshot_date", manifest.get("snapshot_date", "")),
        ("source_run_id", manifest.get("source_run_id", "")),
        ("manifest_status", manifest.get("status", "")),
        ("director_count", summary["director_count"]),
        ("source_backed_dataset_count", summary["source_backed_dataset_count"]),
        ("source_backed_row_count", summary["source_backed_row_count"]),
        ("coverage_row_count", summary["coverage_row_count"]),
        ("metric_count", summary["metric_count"]),
        ("manifest_path", str(manifest_path)),
        ("source_requirement_ids", ", ".join(source_requirement_ids)),
    ]
    return [{"field": field, "value": value} for field, value in rows]


def _ppttc_map_rows() -> list[dict[str, Any]]:
    return [
        {
            "element_name": "ThinkCellSourceStatus",
            "element_type": "text",
            "sheet": "Control",
            "table_name": TABLE_NAMES["Control"],
            "description": "Manifest status from the source-backed DirectorBundle run.",
        },
        {
            "element_name": "SourceBackedCountsTable",
            "element_type": "table",
            "sheet": "Source Counts",
            "table_name": TABLE_NAMES["Source Counts"],
            "description": "One row per source-backed dataset count with deterministic metric IDs.",
        },
        {
            "element_name": "CoverageTable",
            "element_type": "table",
            "sheet": "Coverage",
            "table_name": TABLE_NAMES["Coverage"],
            "description": "Dataset coverage policy and source-contract presence from the manifest.",
        },
        {
            "element_name": "MetricStoreTable",
            "element_type": "table",
            "sheet": "Metric Store",
            "table_name": TABLE_NAMES["Metric Store"],
            "description": "Deterministic row-count metrics only; no derived ARR or risk metrics.",
        },
    ]


def _add_table_sheet(
    wb: Workbook,
    *,
    name: str,
    table_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column) for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for column_index in range(1, len(columns) + 1):
        values = [ws.cell(row=row_index, column=column_index).value for row_index in range(1, ws.max_row + 1)]
        max_len = max(len(str(value or "")) for value in values)
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 60)
    ws.freeze_panes = "A2"

    end_col = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A1:{end_col}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _tc_cell(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return {"string": "true" if value else "false"}
    if isinstance(value, int | float):
        return {"number": value}
    return {"string": str(value)}


def _tc_table(
    rows: list[dict[str, Any]],
    columns: list[str],
) -> list[list[dict[str, Any] | None]]:
    table = [[_tc_cell(column) for column in columns]]
    for row in rows:
        table.append([_tc_cell(row.get(column)) for column in columns])
    return table


def _tc_text(name: str, value: Any) -> dict[str, Any]:
    return {"name": name, "table": [[_tc_cell(value)]]}


def _metric_id(territory: str, director: str, dataset: str) -> str:
    return f"source_backed.{_slug(territory)}.{_slug(director)}.{dataset}.row_count"


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "unknown"


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _resolve_manifest_ref(value: str, manifest_path: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    manifest_relative = manifest_path.parent / path
    if manifest_relative.exists():
        return manifest_relative
    cwd_relative = Path.cwd() / path
    if cwd_relative.exists():
        return cwd_relative
    return manifest_relative
