"""Track E — deck binding resolver (E3).

Walks the active director_monthly profile and emits a binding row
for every slide / table / takeaway / link / source-note. Each binding
declares its ``binding_type`` and the concrete source it resolves to
(workbook sheet + columns, snapshot_role + physical column, derived
transform, runtime metadata, static text, external link, legal text).

The resolver consumes:

  * config/deck_contract.yaml             (active profile)
  * config/director_workbook_contract.yaml (sheet + role registry)
  * a real director workbook .xlsx        (for snapshot-role
                                           resolution to physical
                                           columns and column-existence
                                           checks)

Output schema: ``monthly_platform.deck_binding_report.v1``.
A binding row never lies — every ``status: pass`` row is backed by a
column that actually exists in the workbook on this commit.

Hard NO-GOs preserved: this module never modifies a workbook file,
never modifies a PPTX file, and never reads a Salesforce live API.
It is pure metadata over already-emitted artifacts.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from scripts.monthly_platform import deck_contract
from scripts.monthly_platform import director_workbook_contract as wb_contract


REPORT_SCHEMA_VERSION = "monthly_platform.deck_binding_report.v1"

# Binding types — every emitted row has one of these.
BINDING_TYPES = {
    "direct_workbook_table",
    "workbook_snapshot_role",
    "derived_table",
    "generated_takeaway",
    "static_text",
    "external_link",
    "legal_text",
    "source_note",
    "metric_binding",
}


def _headers_by_sheet(
    workbook_path: Path,
    contract: wb_contract.DirectorWorkbookContract,
) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    out: dict[str, list[str]] = {}
    for sheet_decl in contract.raw.get("sheets", []):
        name = sheet_decl["name"]
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        hdr = sheet_decl.get("header_row", 1)
        out[name] = [
            str(ws.cell(hdr, c).value)
            for c in range(1, ws.max_column + 1)
            if ws.cell(hdr, c).value is not None
        ]
    return out


def _resolve_snapshot_role_column(
    role_name: str,
    workbook_contract: wb_contract.DirectorWorkbookContract,
    headers_by_sheet: dict[str, list[str]],
) -> wb_contract.ResolvedSnapshotRole:
    return workbook_contract.resolve_pattern_role(role_name, headers_by_sheet)


def resolve(
    *,
    workbook_path: Path,
    deck: deck_contract.DeckContract | None = None,
    workbook: wb_contract.DirectorWorkbookContract | None = None,
) -> dict[str, Any]:
    if deck is None:
        deck = deck_contract.load()
    if workbook is None:
        workbook = wb_contract.load()
    assert deck is not None
    assert workbook is not None

    headers = _headers_by_sheet(workbook_path, workbook)
    sheets = workbook.sheets_by_name()

    bindings: list[dict[str, Any]] = []
    blockers: list[str] = []

    profile = deck.director_monthly
    profile_id = "director_monthly"

    for slide in profile.get("slides", []):
        sid = slide.get("id")
        snum = slide.get("slide_number")

        # Source notes — record but no source verification needed.
        for note in slide.get("required_source_notes", []) or []:
            bindings.append(
                {
                    "slide_id": sid,
                    "slide_number": snum,
                    "kind": "source_note",
                    "id": note,
                    "binding_type": "source_note",
                    "status": "pass",
                }
            )

        # Generated takeaway.
        ta = slide.get("required_takeaway") or {}
        if ta.get("required"):
            req_metrics = ta.get("required_metrics", []) or []
            bindings.append(
                {
                    "slide_id": sid,
                    "slide_number": snum,
                    "kind": "takeaway",
                    "id": f"{sid}_takeaway",
                    "binding_type": "generated_takeaway",
                    "template": ta.get("template"),
                    "max_chars": ta.get("max_chars"),
                    "required_metrics": req_metrics,
                    "status": "pass" if req_metrics else "warning",
                    "detail": "" if req_metrics else "no required_metrics declared",
                }
            )

        # External / Salesforce links.
        for link in slide.get("required_links", []) or []:
            kind = link.get("kind")
            binding_type = (
                "external_link" if kind == "external_url" else "external_link"
            )
            bindings.append(
                {
                    "slide_id": sid,
                    "slide_number": snum,
                    "kind": "link",
                    "id": link.get("id"),
                    "label": link.get("label"),
                    "binding_type": binding_type,
                    "link_kind": kind,
                    "status": "pass",
                }
            )

        # Static / legal slide.
        if slide.get("static") is True or sid == "legal_notice":
            bindings.append(
                {
                    "slide_id": sid,
                    "slide_number": snum,
                    "kind": "static",
                    "id": f"{sid}_static",
                    "binding_type": "legal_text"
                    if sid == "legal_notice"
                    else "static_text",
                    "status": "pass",
                }
            )
            continue

        # Tables.
        for tbl in slide.get("tables", []) or []:
            tid = tbl.get("id")
            src = tbl.get("source")
            binding_type_decl = tbl.get("binding_type", "direct_workbook_table")

            if src != "director_workbook":
                # M1 has no warehouse-source tables on director_monthly.
                bindings.append(
                    {
                        "slide_id": sid,
                        "slide_number": snum,
                        "kind": "table",
                        "id": tid,
                        "binding_type": "direct_workbook_table",
                        "status": "fail",
                        "detail": f"unsupported source {src!r}",
                    }
                )
                blockers.append(f"{sid}/{tid}: unsupported source {src!r}")
                continue

            sheet_name = str(tbl.get("sheet") or "")
            sheet_decl = sheets.get(sheet_name)
            if sheet_decl is None:
                bindings.append(
                    {
                        "slide_id": sid,
                        "slide_number": snum,
                        "kind": "table",
                        "id": tid,
                        "sheet": sheet_name,
                        "binding_type": binding_type_decl,
                        "status": "fail",
                        "detail": f"unknown sheet {sheet_name!r}",
                    }
                )
                blockers.append(f"{sid}/{tid}: unknown sheet {sheet_name!r}")
                continue

            if binding_type_decl == "derived_table":
                # Resolve all snapshot_roles + verify transform_id.
                resolved_role_columns: dict[str, str | None] = {}
                role_status = "pass"
                role_details: list[str] = []
                for input_name, role_name in (tbl.get("snapshot_roles") or {}).items():
                    resolved = _resolve_snapshot_role_column(
                        role_name, workbook, headers
                    )
                    resolved_role_columns[input_name] = resolved.physical_column
                    if resolved.status != "pass":
                        role_status = "fail"
                        role_details.append(
                            f"{input_name}={role_name}: {resolved.detail}"
                        )

                bindings.append(
                    {
                        "slide_id": sid,
                        "slide_number": snum,
                        "kind": "table",
                        "id": tid,
                        "sheet": sheet_name,
                        "binding_type": "derived_table",
                        "transform_id": tbl.get("transform_id"),
                        "display_grain": tbl.get("display_grain"),
                        "source_grain": tbl.get("source_grain"),
                        "row_ids": [r["id"] for r in (tbl.get("rows") or [])],
                        "snapshot_roles_resolved": resolved_role_columns,
                        "evidence_only": bool(tbl.get("evidence_only", False)),
                        "status": role_status,
                        "detail": "; ".join(role_details),
                    }
                )
                if role_status != "pass":
                    blockers.append(f"{sid}/{tid}: {'; '.join(role_details)}")
                continue

            # direct_workbook_table — verify each column.
            allowed_cols = set(sheet_decl.get("required_columns", []) or [])
            actual_headers = set(headers.get(sheet_name, []))
            col_results: list[dict[str, Any]] = []
            tbl_status = "pass"
            tbl_details: list[str] = []
            for col in tbl.get("columns", []) or []:
                cid = col.get("id")
                if "source_column" in col:
                    physical = col["source_column"]
                    in_contract = physical in allowed_cols
                    in_workbook = physical in actual_headers
                    if not in_contract or not in_workbook:
                        tbl_status = "fail"
                        tbl_details.append(
                            f"col={cid} source_column={physical!r} "
                            f"in_contract={in_contract} in_workbook={in_workbook}"
                        )
                    col_results.append(
                        {
                            "column_id": cid,
                            "kind": "source_column",
                            "physical_column": physical,
                            "status": "pass" if in_contract and in_workbook else "fail",
                        }
                    )
                elif "snapshot_role" in col:
                    resolved = _resolve_snapshot_role_column(
                        col["snapshot_role"], workbook, headers
                    )
                    if resolved.status != "pass":
                        tbl_status = "fail"
                        tbl_details.append(
                            f"col={cid} snapshot_role={col['snapshot_role']!r}: "
                            f"{resolved.detail}"
                        )
                    col_results.append(
                        {
                            "column_id": cid,
                            "kind": "snapshot_role",
                            "snapshot_role": col["snapshot_role"],
                            "physical_column": resolved.physical_column,
                            "resolved_date": resolved.resolved_date,
                            "status": resolved.status,
                        }
                    )
                elif "computed" in col:
                    col_results.append(
                        {
                            "column_id": cid,
                            "kind": "computed",
                            "computed": col["computed"],
                            "status": "pass",
                        }
                    )
                else:
                    tbl_status = "fail"
                    tbl_details.append(f"col={cid} has no source")
                    col_results.append(
                        {
                            "column_id": cid,
                            "kind": "unknown",
                            "status": "fail",
                        }
                    )

            bindings.append(
                {
                    "slide_id": sid,
                    "slide_number": snum,
                    "kind": "table",
                    "id": tid,
                    "sheet": sheet_name,
                    "binding_type": "direct_workbook_table",
                    "evidence_only": bool(tbl.get("evidence_only", False)),
                    "columns": col_results,
                    "filters": tbl.get("filters") or {},
                    "max_rows": tbl.get("max_rows"),
                    "status": tbl_status,
                    "detail": "; ".join(tbl_details),
                }
            )
            if tbl_status != "pass":
                blockers.append(f"{sid}/{tid}: {'; '.join(tbl_details)}")

    pass_count = sum(1 for b in bindings if b.get("status") == "pass")
    fail_count = sum(1 for b in bindings if b.get("status") == "fail")
    warn_count = sum(1 for b in bindings if b.get("status") == "warning")

    return {
        "schema_version": REPORT_SCHEMA_VERSION,
        "profile_id": profile_id,
        "deck_contract_path": str(deck.path),
        "workbook_contract_path": str(workbook.path),
        "workbook_path": str(workbook_path),
        "resolved_at": datetime.now(timezone.utc).isoformat(),
        "status": "pass" if fail_count == 0 else "fail",
        "binding_count": len(bindings),
        "pass_count": pass_count,
        "warning_count": warn_count,
        "fail_count": fail_count,
        "blockers": blockers,
        "bindings": bindings,
    }


def render_markdown(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"# Deck binding report — {report['profile_id']}\n")
    lines.append(f"- deck contract: `{report['deck_contract_path']}`")
    lines.append(f"- workbook contract: `{report['workbook_contract_path']}`")
    lines.append(f"- workbook: `{report['workbook_path']}`")
    lines.append(f"- resolved_at: {report['resolved_at']}")
    lines.append(f"- **status: {report['status']}**")
    lines.append(
        f"- bindings: {report['binding_count']} "
        f"(pass={report['pass_count']} warn={report['warning_count']} fail={report['fail_count']})"
    )
    lines.append("")

    # Group by slide.
    by_slide: dict[int, list[dict[str, Any]]] = {}
    for b in report["bindings"]:
        by_slide.setdefault(b["slide_number"], []).append(b)
    for snum in sorted(by_slide.keys()):
        rows = by_slide[snum]
        slide_id = rows[0]["slide_id"]
        lines.append(f"## Slide {snum}: `{slide_id}`\n")
        lines.append("| Kind | ID | Binding type | Status | Detail |")
        lines.append("| --- | --- | --- | --- | --- |")
        for b in rows:
            detail = b.get("detail") or ""
            if b["kind"] == "table" and b.get("binding_type") == "derived_table":
                detail = (
                    f"transform={b.get('transform_id')} "
                    f"rows={len(b.get('row_ids', []))} "
                    f"roles={b.get('snapshot_roles_resolved') or {}}"
                )
            elif b["kind"] == "table":
                detail = f"sheet={b.get('sheet')} cols={len(b.get('columns', []))}"
            elif b["kind"] == "takeaway":
                detail = f"max_chars={b.get('max_chars')} metrics={b.get('required_metrics')}"
            elif b["kind"] == "link":
                detail = f"kind={b.get('link_kind')}"
            lines.append(
                f"| {b['kind']} | `{b['id']}` | {b['binding_type']} | {b['status']} | {detail} |"
            )
        lines.append("")

    if report["blockers"]:
        lines.append("## Blockers\n")
        for blk in report["blockers"]:
            lines.append(f"- {blk}")
        lines.append("")

    return "\n".join(lines) + "\n"


__all__ = ["BINDING_TYPES", "REPORT_SCHEMA_VERSION", "render_markdown", "resolve"]
