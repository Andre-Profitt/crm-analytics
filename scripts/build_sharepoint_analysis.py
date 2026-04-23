"""
Build a consolidated pipeline review workbook for SharePoint upload.

Pulls from the 9 director workbooks and the Salesforce forecast API.
Output is a single .xlsx formatted for executive review.
"""

import argparse
import json
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from monthly_platform.historical_trending import (
        resolve_historical_trending_contract,
    )
    from monthly_platform.period import resolve_period_context
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.historical_trending import (
        resolve_historical_trending_contract,
    )
    from scripts.monthly_platform.period import resolve_period_context

# JSON territory key -> workbook regional label used in tabs and CLI --territory
_TERRITORY_KEY_TO_LABEL: dict[str, str] = {
    "APAC": "APAC",
    "Central Europe": "EMEA Central",
    "UK & Ireland": "EMEA UK & Ireland",
    "Southern Europe": "EMEA South West",
    "NL & Nordics": "EMEA NE",
    "Middle East & Africa": "EMEA MEA",
    "Canada": "NA Canada",
    "NA Asset Management": "NA Asset Mgmt",
    "Pension & Insurance": "NA Insurance",
}

# SF OwnerId + Territory2Id per director (not in the territory config JSON).
_DIRECTOR_SF_IDS: dict[str, tuple[str | None, str | None]] = {
    "Jesper Tyrer": ("005Tb00000PY6SpIAL", "0MI7S0000008XKyWAM"),
    "Patrick Gaughan": ("005Tb00000XYMJI", "0MITb0000000JM1"),
    "Megan Miceli": ("005Tb00000MlZXC", "0MIQA00000004wT"),
    "Adam Steinhaus": ("005QA000006WqOD", "0MITb0000000dvp"),
    "Christian Ebbesen": ("0052o00000BeANW", "0MIQA00000005ZB"),
    "Dan Peppett": ("00557000006VpU9", "0MIQA00000005fd"),
    "Sarah Pittroff": ("005Tb00000WVuoK", "0MIQA00000005Vx"),
    "Francois Thaury": ("005D000000272No", "0MIQA00000005cP"),
    "Mourad Essofi": (None, None),
}


def _load_directors() -> list[tuple[str, str, str, str | None, str | None]]:
    """Build DIRECTORS list from config/sd_monthly_territories.json.

    Returns list of (name, territory_label, filename, owner_id, territory2_id).
    """
    cfg_path = (
        Path(__file__).resolve().parents[1] / "config" / "sd_monthly_territories.json"
    )
    with open(cfg_path) as f:
        cfg = json.load(f)

    directors: list[tuple[str, str, str, str | None, str | None]] = []
    for territory_key, entry in cfg["territories"].items():
        name: str = entry["director"]
        label = _TERRITORY_KEY_TO_LABEL.get(territory_key, territory_key)
        filename = f"{name.lower().replace(' ', '-')}.xlsx"
        owner_id, territory2_id = _DIRECTOR_SF_IDS.get(name, (None, None))
        directors.append((name, label, filename, owner_id, territory2_id))
    return directors


DIRECTORS = _load_directors()

FTYPE = "0Db7S000000zDaMSAU"  # Opportunity ARR

# Scope label used in titles and narrative. Set by main() to "All Territories"
# (master) or the territory name (regional). Module-level so every tab builder
# can consult it without threading a parameter through 27 functions.
SCOPE_LABEL = "All Territories"


def _month_span_title(start_date: str, end_date: str) -> str:
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)
    return f"{start.strftime('%B')} through {end.strftime('%B %Y')}"


def _infer_report_date_from_workbooks_dir(workbooks_dir: Path | None) -> str | None:
    if workbooks_dir is None:
        return None
    path = Path(workbooks_dir).resolve()
    for candidate in [path, *path.parents]:
        try:
            datetime.strptime(candidate.name, "%Y-%m-%d")
            return candidate.name
        except ValueError:
            continue
    return None


def _resolve_runtime_period_context(
    *,
    as_of_date: str | None = None,
    workbooks_dir: Path | None = None,
) -> dict[str, str | int]:
    report_date = (
        str(as_of_date or "").strip()[:10]
        or _infer_report_date_from_workbooks_dir(workbooks_dir)
        or datetime.now().strftime("%Y-%m-%d")
    )
    period = resolve_period_context(
        as_of_date=report_date,
        snapshot_date=report_date,
        deck_date=report_date,
    )
    analysis_year = period.current_quarter.year
    return {
        "report_date": report_date,
        "report_date_display": datetime.fromisoformat(report_date).strftime("%d %B %Y"),
        "analysis_year": analysis_year,
        "fy_label": period.fiscal_year,
        "fy_short": str(analysis_year)[-2:],
        "forecast_year_start": f"{analysis_year}-01-01",
        "forecast_year_end": f"{analysis_year}-12-31",
        "prior_quarter_label": period.prior_quarter.label,
        "prior_quarter_title": period.prior_quarter.title,
        "prior_quarter_start": period.prior_quarter.start_date,
        "prior_quarter_end": period.prior_quarter.end_date,
        "prior_quarter_range_label": period.prior_quarter.range_label,
        "current_quarter_label": period.current_quarter.label,
        "current_quarter_title": period.current_quarter.title,
        "current_quarter_fy_title": f"{period.current_quarter.label} {period.fiscal_year}",
        "current_quarter_start": period.current_quarter.start_date,
        "current_quarter_end": period.current_quarter.end_date,
        "current_quarter_month_start": period.current_quarter.month_start,
        "current_quarter_month_end": period.current_quarter.month_end,
        "current_quarter_range_label": period.current_quarter.range_label,
        "current_quarter_months_title": _month_span_title(
            period.current_quarter.start_date,
            period.current_quarter.end_date,
        ),
    }


RUNTIME_PERIOD = _resolve_runtime_period_context()


def _configure_runtime_period(
    *,
    as_of_date: str | None = None,
    workbooks_dir: Path | None = None,
) -> dict[str, str | int]:
    global RUNTIME_PERIOD
    RUNTIME_PERIOD = _resolve_runtime_period_context(
        as_of_date=as_of_date,
        workbooks_dir=workbooks_dir,
    )
    return RUNTIME_PERIOD


def _historical_trending_contract() -> object:
    return resolve_historical_trending_contract(
        retrospective_label=str(RUNTIME_PERIOD["prior_quarter_label"]),
        retrospective_title=str(RUNTIME_PERIOD["prior_quarter_title"]),
        current_label=str(RUNTIME_PERIOD["current_quarter_label"]),
        current_title=str(RUNTIME_PERIOD["current_quarter_title"]),
    )


def _is_q1_of_analysis_year(value) -> bool:
    token = str(value or "")[:10]
    return (
        bool(token)
        and RUNTIME_PERIOD["prior_quarter_start"]
        <= token
        <= RUNTIME_PERIOD["prior_quarter_end"]
    )


def _is_in_current_quarter(value) -> bool:
    token = str(value or "")[:10]
    return bool(token) and (
        RUNTIME_PERIOD["current_quarter_start"]
        <= token
        <= RUNTIME_PERIOD["current_quarter_end"]
    )


# Canonical territory label used in every output tab. Maps from the raw
# Salesforce Sales_Region__c / Account_Unit_Group__c value to the short name
# the directors use.
TERRITORY_MAP = {
    # APAC
    ("SC Asia", ""): "APAC",
    ("SC Asia", "APAC"): "APAC",
    # EMEA
    ("SC EMEA", "Northern Europe"): "EMEA NE",
    ("SC EMEA", "Central Europe"): "EMEA Central",
    ("SC EMEA", "United Kingdom & Ireland"): "EMEA UK & Ireland",
    ("SC EMEA", "Southwestern Europe"): "EMEA South West",
    ("SC EMEA", "Middle East & Africa"): "EMEA MEA",
    # NA: use unit-based split so AM / Canada / P&I line up with directors
    ("SC North America", ""): "NA",
}


def canonical_territory(account_unit_group, sales_region, unit=None):
    """Return the canonical short-form territory label or the raw SR fallback."""
    key = (str(account_unit_group or ""), str(sales_region or ""))
    if key in TERRITORY_MAP:
        return TERRITORY_MAP[key]
    # NA split requires Account.Unit__c which we do not always have
    if key[0] == "SC North America":
        if unit == "SC Canada":
            return "NA Canada"
        if unit == "SC USA":
            return "NA Asset Mgmt"
        return "NA"
    return sales_region or account_unit_group or "Unspecified"


# Corporate styling, kept plain so it does not read as machine generated.
NAVY = "1F3864"
LIGHT_GRAY = "F2F2F2"
BORDER_GRAY = "BFBFBF"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
CAPTION_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
THIN = Side(style="thin", color=BORDER_GRAY)
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def get_sf_auth():
    data = json.loads(
        subprocess.run(
            ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
            capture_output=True,
            text=True,
            check=True,
            timeout=30,
        ).stdout
    )["result"]
    return data["accessToken"], data["instanceUrl"]


def forecast_page_fy26(session, instance, oid, tid):
    """Return forecast-page buckets for a director territory."""
    out = {"Commit": 0, "Best Case": 0, "Pipeline": 0, "Closed": 0}
    if not oid or not tid:
        return out
    pq = (
        f"SELECT Id FROM Period WHERE StartDate >= {RUNTIME_PERIOD['forecast_year_start']} "
        f"AND StartDate <= {RUNTIME_PERIOD['forecast_year_end']} AND Type='Quarter'"
    )
    pids = [
        p["Id"]
        for p in session.get(f"{instance}/services/data/v66.0/query", params={"q": pq})
        .json()
        .get("records", [])
    ]
    if not pids:
        return out
    q = (
        "SELECT ForecastCategoryName, SUM(ForecastAmount) s "
        "FROM ForecastingItem "
        f"WHERE OwnerId='{oid}' AND Territory2Id='{tid}' "
        f"AND ForecastingTypeId='{FTYPE}' AND CurrencyIsoCode='EUR' "
        f"AND PeriodId IN ('{chr(39).join(pids).join([chr(39), chr(39)])[1:-1]}') "
        "GROUP BY ForecastCategoryName"
    )
    # Clean separator fix (avoid the quote-trick above)
    ids = "','".join(pids)
    q = (
        "SELECT ForecastCategoryName, SUM(ForecastAmount) s "
        "FROM ForecastingItem "
        f"WHERE OwnerId='{oid}' AND Territory2Id='{tid}' "
        f"AND ForecastingTypeId='{FTYPE}' AND CurrencyIsoCode='EUR' "
        f"AND PeriodId IN ('{ids}') "
        "GROUP BY ForecastCategoryName"
    )
    r = session.get(f"{instance}/services/data/v66.0/query", params={"q": q}).json()
    for row in r.get("records", []):
        cat = row.get("ForecastCategoryName")
        if cat in out:
            out[cat] = float(row.get("s") or 0)
    return out


def _load(path):
    """Load a director workbook into a simple list of dicts per sheet."""
    wb = load_workbook(path, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: v for i, v in enumerate(r)})
        sheets[sn] = rows
    wb.close()
    return sheets


def _quarter(s):
    s = str(s or "")
    if len(s) < 7:
        return None
    try:
        m = int(s[5:7])
    except ValueError:
        return None
    return f"Q{(m - 1) // 3 + 1}"


def _apply_header(ws, row, cols):
    for i, _ in enumerate(cols, 1):
        c = ws.cell(row=row, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def _zebra(ws, start_row, end_row, n_cols):
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_exec_dashboard(wb, data, overdue_rows, kyc_rows):
    """Single-screen headline view that goes first in the workbook.

    Six KPI cards across the top, then two small tables below.
    """
    ws = wb.create_sheet("Exec Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{RUNTIME_PERIOD['fy_label']} Pipeline Review"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    ws["A2"] = f"Data as of {RUNTIME_PERIOD['report_date_display']}"
    ws["A2"].font = CAPTION_FONT

    # Aggregate headline numbers
    total_open_deals = 0
    total_open_unwtd = 0.0
    total_open_wtd = 0.0
    total_q1_won = 0
    total_q1_won_arr = 0.0
    total_q1_lost_arr = 0.0
    total_q1_slip_arr = 0.0
    total_approved_n = 0
    total_approved_arr = 0.0
    total_cond_n = 0
    total_cond_arr = 0.0
    total_missing_n = 0

    for director_name, _territory, *_rest in data:
        d = data_store[director_name]
        total_open_deals += d["open_count"]
        total_open_unwtd += d["open_unwtd"]
        total_open_wtd += d["open_wtd"]
        total_q1_won += d["q1_won_count"]
        total_q1_won_arr += d["q1_won_arr"]
        total_q1_lost_arr += d["q1_lost_arr"]
        total_q1_slip_arr += d["slip_still_open_arr"]
        total_approved_n += len(d["approved_2026"])
        total_approved_arr += sum(a["arr_unwtd"] for a in d["approved_2026"])
        total_cond_n += len(d["approval_candidates"])
        total_cond_arr += sum(a["arr_unwtd"] for a in d["approval_candidates"])
        total_missing_n += len(d["approval_missing"])

    # KPI cards: label + value, arranged in a 2 x 3 grid
    cards = [
        (
            "Open Land pipeline",
            f"{total_open_deals} deals",
            f"EUR {total_open_unwtd:,.0f} unwtd / EUR {total_open_wtd:,.0f} wtd",
        ),
        (
            f"{RUNTIME_PERIOD['prior_quarter_label']} wins",
            f"{total_q1_won} deals",
            f"EUR {total_q1_won_arr:,.0f} unwtd",
        ),
        (
            f"{RUNTIME_PERIOD['prior_quarter_label']} losses at risk",
            f"EUR {total_q1_lost_arr:,.0f}",
            f"Slips still open: EUR {total_q1_slip_arr:,.0f}",
        ),
        (
            f"Approved {RUNTIME_PERIOD['analysis_year']}",
            f"{total_approved_n} deals",
            f"EUR {total_approved_arr:,.0f} unwtd",
        ),
        (
            "Conditionally approved (pending)",
            f"{total_cond_n} deals",
            f"EUR {total_cond_arr:,.0f} unwtd, largest approval pipeline item",
        ),
        (
            "Data hygiene",
            f"{total_missing_n} missing Stage 3+ approvals",
            f"{len(overdue_rows)} overdue open opps, {len(kyc_rows)} KYC missing",
        ),
    ]

    start_row = 4
    card_h = 4
    for idx, (label, value, note) in enumerate(cards):
        r = start_row + (idx // 3) * card_h
        c = 1 + (idx % 3) * 4
        # Label
        ws.cell(row=r, column=c, value=label).font = Font(
            name="Calibri", size=10, bold=True, color=NAVY
        )
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = LEFT
        # Value
        ws.cell(row=r + 1, column=c, value=value).font = Font(
            name="Calibri", size=16, bold=True
        )
        ws.cell(row=r + 1, column=c).border = BORDER
        ws.cell(row=r + 1, column=c).alignment = LEFT
        # Note
        ws.cell(row=r + 2, column=c, value=note).font = Font(
            name="Calibri", size=9, color="595959"
        )
        ws.cell(row=r + 2, column=c).border = BORDER
        ws.cell(row=r + 2, column=c).alignment = LEFT
        # Merge the three cells under the card for border cohesion
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 2)
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 2)
        ws.merge_cells(start_row=r + 2, start_column=c, end_row=r + 2, end_column=c + 2)
        ws.row_dimensions[r].height = 18
        ws.row_dimensions[r + 1].height = 28
        ws.row_dimensions[r + 2].height = 18

    # Per-director roll-up table below the cards
    tbl_start = start_row + 2 * card_h + 2
    ws.cell(row=tbl_start, column=1, value="Per-director Summary").font = BODY_BOLD
    headers = [
        "Director",
        "Territory",
        "Open Land Deals",
        "Open ARR Unwtd",
        "Open ARR Wtd",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost",
        f"Approved {RUNTIME_PERIOD['analysis_year']}",
        "Cond Approved",
    ]
    h_row = tbl_start + 1
    for i, h in enumerate(headers, 1):
        ws.cell(row=h_row, column=i, value=h)
    _apply_header(ws, h_row, headers)

    r = h_row + 1
    for director_name, territory, *_rest in data:
        d = data_store[director_name]
        app_arr = sum(a["arr_unwtd"] for a in d["approved_2026"])
        cond_arr = sum(a["arr_unwtd"] for a in d["approval_candidates"])
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(row=r, column=2, value=territory).alignment = LEFT
        ws.cell(row=r, column=3, value=d["open_count"]).alignment = CENTER
        for ci, val in enumerate(
            [
                d["open_unwtd"],
                d["open_wtd"],
                d["q1_won_count"],
                d["q1_won_arr"],
                d["q1_lost_count"],
                app_arr,
                cond_arr,
            ],
            4,
        ):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci in (6, 8):
                cell.alignment = CENTER
            else:
                cell.alignment = RIGHT
                cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    _zebra(ws, h_row + 1, r - 1, len(headers))
    _set_widths(ws, [20, 20, 12, 16, 16, 10, 16, 10, 14, 14])
    ws.freeze_panes = f"A{h_row + 1}"


def build_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{RUNTIME_PERIOD['fy_label']} Pipeline Review, {SCOPE_LABEL}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = f"Data pulled from Salesforce on {RUNTIME_PERIOD['report_date_display']}"
    ws["A2"].font = CAPTION_FONT
    ws["A3"] = (
        "Open pipeline is Land only (matches the forecast page scope). "
        "ARR is in EUR. Unweighted is the full deal value. "
        "Weighted is the probability-weighted forecast."
    )
    ws["A3"].font = CAPTION_FONT
    ws["A3"].alignment = LEFT

    headers = [
        "Director",
        "Territory",
        "Open Land Deals",
        "Open Land ARR Unwtd",
        "Open Land ARR Wtd",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Slips Still Open",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Slip ARR at Risk",
    ]
    start = 5
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    totals = defaultdict(float)
    totals_int = defaultdict(int)
    for director_name, territory, *_rest in data:
        d = data_store[director_name]
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(row=r, column=2, value=territory).alignment = LEFT
        ws.cell(row=r, column=3, value=d["open_count"]).alignment = RIGHT
        ws.cell(row=r, column=4, value=d["open_unwtd"]).alignment = RIGHT
        ws.cell(row=r, column=5, value=d["open_wtd"]).alignment = RIGHT
        ws.cell(row=r, column=6, value=d["q1_won_count"]).alignment = RIGHT
        ws.cell(row=r, column=7, value=d["q1_won_arr"]).alignment = RIGHT
        ws.cell(row=r, column=8, value=d["q1_lost_count"]).alignment = RIGHT
        ws.cell(row=r, column=9, value=d["q1_lost_arr"]).alignment = RIGHT
        ws.cell(row=r, column=10, value=d["slip_still_open_count"]).alignment = RIGHT
        ws.cell(row=r, column=11, value=d["slip_still_open_arr"]).alignment = RIGHT
        for cname, key in [
            ("D", "open_unwtd"),
            ("E", "open_wtd"),
            ("G", "q1_won_arr"),
            ("I", "q1_lost_arr"),
            ("K", "slip_still_open_arr"),
        ]:
            ws[f"{cname}{r}"].number_format = "#,##0"
        for cname in ("D", "E", "G", "I", "K"):
            totals[cname] += ws[f"{cname}{r}"].value or 0
        for cname, key in [
            ("C", "open_count"),
            ("F", "q1_won_count"),
            ("H", "q1_lost_count"),
            ("J", "slip_still_open_count"),
        ]:
            totals_int[cname] += ws[f"{cname}{r}"].value or 0
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    # Totals row
    ws.cell(row=r, column=1, value="Total").font = BODY_BOLD
    for cname in ("C", "F", "H", "J"):
        ws[f"{cname}{r}"] = totals_int[cname]
        ws[f"{cname}{r}"].alignment = RIGHT
        ws[f"{cname}{r}"].font = BODY_BOLD
    for cname in ("D", "E", "G", "I", "K"):
        ws[f"{cname}{r}"] = totals[cname]
        ws[f"{cname}{r}"].alignment = RIGHT
        ws[f"{cname}{r}"].number_format = "#,##0"
        ws[f"{cname}{r}"].font = BODY_BOLD
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="E7E6E6")

    _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, [20, 22, 14, 18, 18, 10, 16, 10, 16, 16, 18])
    ws.freeze_panes = "A6"


def build_forecast_recon_sheet(wb, data):
    ws = wb.create_sheet("Forecast Reconciliation")
    ws["A1"] = "Deck Land Book vs Forecast Page"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "The Salesforce forecast page only surfaces deals a rep has "
        "categorised into the forecast. Our deck includes every open "
        "Land deal in the territory. The columns below let you see "
        "both and judge coverage."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 28

    headers = [
        "Director",
        "Territory",
        "FP Closed",
        "Deck Won",
        "FP Commit",
        "Deck Commit (Unwtd)",
        "FP Best Case",
        "Deck Best Case (Unwtd)",
        "FP Pipeline",
        "Deck Pipeline (Unwtd)",
        "FP Total",
        "Deck Total",
    ]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    for director_name, territory, *_rest in data:
        d = data_store[director_name]
        fp = d["forecast_page"]
        dk = d["deck_buckets"]
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(row=r, column=2, value=territory).alignment = LEFT
        ws.cell(row=r, column=3, value=fp["Closed"]).alignment = RIGHT
        ws.cell(row=r, column=4, value=dk["Closed"]).alignment = RIGHT
        ws.cell(row=r, column=5, value=fp["Commit"]).alignment = RIGHT
        ws.cell(row=r, column=6, value=dk["Commit"]).alignment = RIGHT
        ws.cell(row=r, column=7, value=fp["Best Case"]).alignment = RIGHT
        ws.cell(row=r, column=8, value=dk["Best Case"]).alignment = RIGHT
        ws.cell(row=r, column=9, value=fp["Pipeline"]).alignment = RIGHT
        ws.cell(row=r, column=10, value=dk["Pipeline"]).alignment = RIGHT
        fp_total = sum(fp.values())
        dk_total = sum(dk.values())
        ws.cell(row=r, column=11, value=fp_total).alignment = RIGHT
        ws.cell(row=r, column=12, value=dk_total).alignment = RIGHT
        for col_letter in "CDEFGHIJKL":
            ws[f"{col_letter}{r}"].number_format = "#,##0"
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, [20, 22, 12, 12, 12, 16, 12, 16, 12, 16, 12, 14])
    ws.freeze_panes = "A5"


def build_q1_slips_sheet(wb, data):
    ws = wb.create_sheet(f"{RUNTIME_PERIOD['prior_quarter_label']} Slips, Still Open")
    ws["A1"] = (
        f"{RUNTIME_PERIOD['prior_quarter_title']} Deals That Slipped and Are Still Open"
    )
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"These deals had a {RUNTIME_PERIOD['prior_quarter_label']} close date at some point, got pushed, "
        "and are still on the pipeline today. Filter by territory and "
        "by type. Deals already won or lost are excluded so the ARR "
        "reflects current exposure."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 32

    headers = [
        "Territory",
        "Type",
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Current Close Date",
        "ARR Unwtd (EUR)",
        "ARR Wtd (EUR)",
        "Push Count",
        "Root Cause / Commentary",
    ]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    all_slips = []
    for director_name, territory, *_rest in data:
        for deal in data_store[director_name]["open_slips"]:
            all_slips.append({**deal, "territory": territory})
    all_slips.sort(key=lambda x: -(x.get("arr_unwtd") or 0))

    for s in all_slips:
        ws.cell(row=r, column=1, value=s["territory"]).alignment = LEFT
        ws.cell(row=r, column=2, value=s.get("type", "")).alignment = LEFT
        ws.cell(row=r, column=3, value=s.get("account", "")).alignment = LEFT
        ws.cell(row=r, column=4, value=s.get("opportunity", "")).alignment = LEFT
        ws.cell(row=r, column=5, value=s.get("owner", "")).alignment = LEFT
        ws.cell(row=r, column=6, value=s.get("stage", "")).alignment = LEFT
        ws.cell(row=r, column=7, value=s.get("close_date", "")).alignment = CENTER
        ws.cell(row=r, column=8, value=s.get("arr_unwtd") or 0).alignment = RIGHT
        ws.cell(row=r, column=9, value=s.get("arr_wtd") or 0).alignment = RIGHT
        ws.cell(row=r, column=10, value=s.get("push_count") or 0).alignment = CENTER
        ws.cell(
            row=r, column=11, value=""
        ).alignment = LEFT  # blank for director to fill
        ws[f"H{r}"].number_format = "#,##0"
        ws[f"I{r}"].number_format = "#,##0"
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, [20, 10, 32, 42, 20, 20, 14, 18, 18, 10, 45])
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{r - 1}"
    ws.freeze_panes = f"A{start + 1}"


def build_closed_won_sheet(wb, data):
    ws = wb.create_sheet("Closed Won YTD")
    ws["A1"] = f"Closed Won Deals, {RUNTIME_PERIOD['fy_label']} YTD"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"Every closed-won opportunity across {SCOPE_LABEL} for {RUNTIME_PERIOD['fy_label']}. "
        "Renewals are typically zero-ARR by design (like-for-like). "
        "ARR figures are unweighted."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 28

    headers = [
        "Territory",
        "Close Date",
        "Account",
        "Opportunity",
        "Owner",
        "Type",
        "Sales Region",
        "ARR Unwtd (EUR)",
    ]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    rows = []
    for director_name, territory, *_rest in data:
        for w in data_store[director_name]["won_any"]:
            rows.append({**w, "territory": territory})
    rows.sort(key=lambda x: -(x.get("arr_unwtd") or 0))

    for w in rows:
        ws.cell(row=r, column=1, value=w["territory"]).alignment = LEFT
        ws.cell(row=r, column=2, value=w.get("close_date", "")).alignment = CENTER
        ws.cell(row=r, column=3, value=w.get("account", "")).alignment = LEFT
        ws.cell(row=r, column=4, value=w.get("opportunity", "")).alignment = LEFT
        ws.cell(row=r, column=5, value=w.get("owner", "")).alignment = LEFT
        ws.cell(row=r, column=6, value=w.get("type", "")).alignment = CENTER
        ws.cell(row=r, column=7, value=w.get("sales_region", "")).alignment = LEFT
        ws.cell(row=r, column=8, value=w.get("arr_unwtd") or 0).alignment = RIGHT
        ws[f"H{r}"].number_format = "#,##0"
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, [20, 14, 32, 42, 20, 10, 18, 18])
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{r - 1}"
    ws.freeze_panes = f"A{start + 1}"


def _render_list_sheet(
    ws,
    title,
    subtitle,
    headers,
    rows_data,
    widths,
    numeric_cols=None,
    wrap_cols=None,
    start=4,
):
    """Generic rendering for a list tab with title/subtitle/headers/rows."""
    numeric_cols = numeric_cols or []
    wrap_cols = wrap_cols or []
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 32

    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    for row in rows_data:
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci in numeric_cols:
                cell.alignment = RIGHT
                cell.number_format = "#,##0"
            elif ci in wrap_cols:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            else:
                cell.alignment = LEFT
            cell.border = BORDER
            cell.font = BODY_FONT
        r += 1
    if rows_data:
        _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, widths)
    if rows_data:
        ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{r - 1}"
    ws.freeze_panes = f"A{start + 1}"


def build_approvals_ytd_sheet(wb, data):
    ws = wb.create_sheet(f"Approvals, {RUNTIME_PERIOD['analysis_year']}")
    rows = []
    for director_name, territory, *_rest in data:
        for a in data_store[director_name]["approved_2026"]:
            rows.append(
                [
                    territory,
                    director_name,
                    a.get("account", ""),
                    a.get("opportunity", ""),
                    a.get("owner", ""),
                    str(a.get("approval_date", "") or "")[:10],
                    str(a.get("close_date", "") or "")[:10],
                    a.get("stage", ""),
                    a.get("arr_unwtd") or 0,
                ]
            )
    rows.sort(key=lambda x: -(x[-1] or 0))
    _render_list_sheet(
        ws,
        f"Commercial Approvals, Approved in {RUNTIME_PERIOD['analysis_year']}",
        (
            f"Land deals with commercial approval signed off in {RUNTIME_PERIOD['analysis_year']}. "
            "One row per approved opportunity. ARR is unweighted."
        ),
        [
            "Territory",
            "Director",
            "Account",
            "Opportunity",
            "Owner",
            "Approval Date",
            "Close Date",
            "Stage",
            "ARR Unwtd (EUR)",
        ],
        rows,
        widths=[20, 20, 32, 42, 20, 14, 14, 18, 18],
        numeric_cols=[9],
    )


def build_approvals_overview_sheet(wb, data):
    """Global + regional approval overview. Matches Rebekka's slide 6 layout.

    Block A: Global totals across all territories (count + ARR).
    Block B: Regional breakout rows, one per territory.
    """
    ws = wb.create_sheet("Approvals Overview")
    ws["A1"] = "Commercial Approvals Overview"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    # Aggregate totals
    total_approved_n = 0
    total_approved_arr = 0.0
    total_cond_n = 0
    total_cond_arr = 0.0
    total_missing_n = 0
    total_missing_arr = 0.0
    regional = {}
    for director_name, territory, *_rest in data:
        d = data_store[director_name]
        app = d["approved_2026"]
        cond = d["approval_candidates"]
        miss = d["approval_missing"]
        app_arr = sum(a["arr_unwtd"] for a in app)
        cond_arr = sum(a["arr_unwtd"] for a in cond)
        miss_arr = sum(a["arr_unwtd"] for a in miss)
        total_approved_n += len(app)
        total_approved_arr += app_arr
        total_cond_n += len(cond)
        total_cond_arr += cond_arr
        total_missing_n += len(miss)
        total_missing_arr += miss_arr
        regional[(director_name, territory)] = {
            "approved_n": len(app),
            "approved_arr": app_arr,
            "cond_n": len(cond),
            "cond_arr": cond_arr,
            "miss_n": len(miss),
            "miss_arr": miss_arr,
        }

    # Block A: Global totals
    ws["A3"] = "Global Totals"
    ws["A3"].font = BODY_BOLD
    ws["A4"] = "Status"
    ws["B4"] = "Deal Count"
    ws["C4"] = "ARR Unwtd (EUR)"
    _apply_header(ws, 4, ["Status", "Deal Count", "ARR Unwtd (EUR)"])
    totals_rows = [
        (
            f"Approved {RUNTIME_PERIOD['analysis_year']} YTD",
            total_approved_n,
            total_approved_arr,
        ),
        ("Conditionally Approved", total_cond_n, total_cond_arr),
        ("Missing Approval (Land Stage 3+)", total_missing_n, total_missing_arr),
    ]
    r = 5
    for label, n, arr in totals_rows:
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        ws.cell(row=r, column=2, value=n).alignment = CENTER
        cell = ws.cell(row=r, column=3, value=arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        r += 1

    # Block B: Regional breakout
    r += 2
    ws.cell(row=r, column=1, value="Regional Breakdown").font = BODY_BOLD
    r += 1
    headers = [
        "Director",
        "Territory",
        "Approved YTD #",
        "Approved YTD ARR",
        "Conditionally Approved #",
        "Conditionally Approved ARR",
        "Missing Approval #",
        "Missing Approval ARR",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    _apply_header(ws, r, headers)
    start_regional = r + 1
    r = start_regional
    for (director_name, territory), v in regional.items():
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(row=r, column=2, value=territory).alignment = LEFT
        ws.cell(row=r, column=3, value=v["approved_n"]).alignment = CENTER
        cell = ws.cell(row=r, column=4, value=v["approved_arr"])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        ws.cell(row=r, column=5, value=v["cond_n"]).alignment = CENTER
        cell = ws.cell(row=r, column=6, value=v["cond_arr"])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        ws.cell(row=r, column=7, value=v["miss_n"]).alignment = CENTER
        cell = ws.cell(row=r, column=8, value=v["miss_arr"])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    _zebra(ws, start_regional, r - 1, len(headers))
    _set_widths(ws, [22, 24, 14, 18, 20, 22, 14, 20])
    ws.freeze_panes = f"A{start_regional}"


def build_approvals_candidates_sheet(wb, data):
    ws = wb.create_sheet("Approval Candidates")
    rows = []
    for director_name, territory, *_rest in data:
        for a in data_store[director_name]["approval_candidates"]:
            rows.append(
                [
                    territory,
                    director_name,
                    a.get("account", ""),
                    a.get("opportunity", ""),
                    a.get("owner", ""),
                    a.get("stage", ""),
                    str(a.get("close_date", "") or "")[:10],
                    a.get("status", ""),
                    a.get("next_step", ""),
                    a.get("arr_unwtd") or 0,
                ]
            )
    rows.sort(key=lambda x: -(x[-1] or 0))
    _render_list_sheet(
        ws,
        "Commercial Approval Candidates",
        (
            "Land deals pending approval or submitted for Stage 20 review. "
            "Use this list to track what is moving through the approval pipeline."
        ),
        [
            "Territory",
            "Director",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            "Status",
            "Next Step",
            "ARR Unwtd (EUR)",
        ],
        rows,
        widths=[20, 20, 32, 42, 20, 18, 14, 18, 45, 18],
        numeric_cols=[10],
        wrap_cols=[9],
    )


def build_approvals_missing_sheet(wb, data):
    ws = wb.create_sheet("Land Stage 3+, No Approval")
    rows = []
    for director_name, territory, *_rest in data:
        for a in data_store[director_name]["approval_missing"]:
            rows.append(
                [
                    territory,
                    director_name,
                    a.get("account", ""),
                    a.get("opportunity", ""),
                    a.get("owner", ""),
                    a.get("stage", ""),
                    str(a.get("close_date", "") or "")[:10],
                    a.get("next_step", ""),
                    a.get("arr_unwtd") or 0,
                ]
            )
    rows.sort(key=lambda x: -(x[-1] or 0))
    _render_list_sheet(
        ws,
        "Land Stage 3+, Missing Commercial Approval",
        (
            "Land deals in Stage 3 (Engagement) or later that do not have "
            "commercial approval. These need to be escalated to the "
            "approval committee."
        ),
        [
            "Territory",
            "Director",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            "Next Step",
            "ARR Unwtd (EUR)",
        ],
        rows,
        widths=[20, 20, 32, 42, 20, 18, 14, 45, 18],
        numeric_cols=[9],
        wrap_cols=[8],
    )


def build_overdue_open_opps_sheet(wb, overdue_rows):
    ws = wb.create_sheet("Overdue Open Opps")
    # First: summary grouped by Owner with record count, sorted by count desc
    owner_counts = defaultdict(
        lambda: {"count": 0, "arr": 0.0, "territory": "", "director": ""}
    )
    for r in overdue_rows:
        owner = r.get("owner", "(unknown)")
        owner_counts[owner]["count"] += 1
        owner_counts[owner]["arr"] += r.get("arr_unwtd") or 0
    summary_rows = sorted(
        [[owner, d["count"], d["arr"]] for owner, d in owner_counts.items()],
        key=lambda x: -x[1],
    )

    ws["A1"] = "Overdue Open Opportunities"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Open opportunities with a close date in the past. "
        "Top block groups by owner, sorted by deal count. "
        "Detail block lists every overdue deal below."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 32

    # Summary block
    summary_headers = ["Owner", "Overdue Deal Count", "Overdue ARR Unwtd (EUR)"]
    start = 4
    for i, h in enumerate(summary_headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, summary_headers)
    r = start + 1
    for row in summary_rows:
        ws.cell(row=r, column=1, value=row[0]).alignment = LEFT
        ws.cell(row=r, column=2, value=row[1]).alignment = CENTER
        ws.cell(row=r, column=3, value=row[2]).alignment = RIGHT
        ws[f"C{r}"].number_format = "#,##0"
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    _zebra(ws, start + 1, r - 1, 3)

    # Detail block
    ws.cell(row=r + 1, column=1, value="Detail, every overdue deal").font = BODY_BOLD
    detail_start = r + 2
    detail_headers = [
        "Territory",
        "Director",
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR Unwtd (EUR)",
        "Push Count",
    ]
    for i, h in enumerate(detail_headers, 1):
        ws.cell(row=detail_start, column=i, value=h)
    _apply_header(ws, detail_start, detail_headers)

    r = detail_start + 1
    overdue_rows.sort(key=lambda x: -(x.get("arr_unwtd") or 0))
    for o in overdue_rows:
        ws.cell(row=r, column=1, value=o.get("territory", "")).alignment = LEFT
        ws.cell(row=r, column=2, value=o.get("director", "")).alignment = LEFT
        ws.cell(row=r, column=3, value=o.get("account", "")).alignment = LEFT
        ws.cell(row=r, column=4, value=o.get("opportunity", "")).alignment = LEFT
        ws.cell(row=r, column=5, value=o.get("owner", "")).alignment = LEFT
        ws.cell(row=r, column=6, value=o.get("stage", "")).alignment = LEFT
        ws.cell(row=r, column=7, value=o.get("close_date", "")).alignment = CENTER
        ws.cell(row=r, column=8, value=o.get("arr_unwtd") or 0).alignment = RIGHT
        ws.cell(row=r, column=9, value=o.get("push_count") or 0).alignment = CENTER
        ws[f"H{r}"].number_format = "#,##0"
        for col in range(1, len(detail_headers) + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = BODY_FONT
        r += 1
    if overdue_rows:
        _zebra(ws, detail_start + 1, r - 1, len(detail_headers))
        ws.auto_filter.ref = (
            f"A{detail_start}:{get_column_letter(len(detail_headers))}{r - 1}"
        )
    _set_widths(ws, [22, 22, 32, 42, 22, 18, 14, 18, 10])
    ws.freeze_panes = "A5"


def build_kyc_missing_sheet(wb, kyc_rows):
    ws = wb.create_sheet("KYC Missing")
    rows = sorted(
        [
            [
                r.get("account", ""),
                r.get("region", ""),
                r.get("industry", ""),
                r.get("kyc_status", ""),
                r.get("open_opps", 0),
                r.get("open_arr") or 0,
            ]
            for r in kyc_rows
        ],
        key=lambda x: -(x[5] or 0),
    )
    _render_list_sheet(
        ws,
        "Accounts With Open Pipeline and No KYC Approval",
        (
            "Accounts with at least one open opportunity that are not "
            "KYC-approved. These block the deal from progressing to "
            "contracting and should be escalated to the owning rep."
        ),
        [
            "Account",
            "Sales Region",
            "Industry",
            "KYC Status",
            "Open Opps",
            "Open ARR Unwtd (EUR)",
        ],
        rows,
        widths=[40, 22, 22, 20, 12, 22],
        numeric_cols=[6],
    )


def build_renewals_sheet(wb, data):
    ws = wb.create_sheet("Renewals This Quarter")
    rows = []
    for director_name, territory, *_rest in data:
        for ren in data_store[director_name]["renewals_q2"]:
            rows.append(
                [
                    territory,
                    director_name,
                    ren.get("account", ""),
                    ren.get("opportunity", ""),
                    ren.get("owner", ""),
                    ren.get("stage", ""),
                    str(ren.get("close_date", "") or "")[:10],
                    ren.get("probability") or 0,
                    ren.get("acv") or 0,
                ]
            )
    rows.sort(key=lambda x: -(x[-1] or 0))
    _render_list_sheet(
        ws,
        f"Renewals Due This Quarter ({RUNTIME_PERIOD['current_quarter_fy_title']})",
        (
            "Open Type=Renewal opportunities with a close date in "
            f"{RUNTIME_PERIOD['current_quarter_months_title']}. Use Probability and Stage to gauge "
            "likelihood. ACV is in EUR."
        ),
        [
            "Territory",
            "Director",
            "Account",
            "Opportunity",
            "Owner",
            "Stage",
            "Close Date",
            "Probability %",
            "ACV (EUR)",
        ],
        rows,
        widths=[20, 20, 32, 42, 20, 18, 14, 14, 18],
        numeric_cols=[9],
    )


def build_win_rate_sheet(wb, session, instance):
    """Pulls the SD Win Rate by Stage report and flattens it into a table."""
    RID = "00OTb000008gUrVMAU"
    r = session.post(
        f"{instance}/services/data/v66.0/analytics/reports/{RID}?includeDetails=false",
        headers={"Content-Type": "application/json"},
    ).json()

    ws = wb.create_sheet("Win Rate by Stage")
    ws["A1"] = f"Win Rate by Stage, {RUNTIME_PERIOD['fy_label']} YTD"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    headers = ["Stage", "Deal Count", "Won ARR (EUR)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header(ws, 3, headers)
    row = 4
    total_won_n = 0
    total_lost_n = 0
    total_won_arr = 0.0
    total_lost_arr = 0.0
    for g in r.get("groupingsDown", {}).get("groupings", []):
        label = g.get("label", "")
        fm = r.get("factMap", {}).get(f"{g['key']}!T", {})
        aggs = fm.get("aggregates", [])
        count = aggs[1].get("value", 0) if len(aggs) > 1 else 0
        arr = aggs[0].get("value", 0) if aggs else 0
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        ws.cell(row=row, column=2, value=count).alignment = CENTER
        cell = ws.cell(row=row, column=3, value=arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).font = BODY_FONT
            ws.cell(row=row, column=ci).border = BORDER
        if "Won" in label:
            total_won_n += count
            total_won_arr += arr
        else:
            total_lost_n += count
            total_lost_arr += arr
        row += 1

    # Totals block below
    row += 1
    ws.cell(row=row, column=1, value="Summary").font = BODY_BOLD
    row += 1
    summary_rows = [
        ("Total closed deals", total_won_n + total_lost_n, ""),
        ("Won", total_won_n, total_won_arr),
        ("Lost", total_lost_n, total_lost_arr),
        (
            "Win rate, by deal count",
            f"{(total_won_n / max(total_won_n + total_lost_n, 1) * 100):.1f}%",
            "",
        ),
        (
            "Win rate, by ARR",
            f"{(total_won_arr / max(total_won_arr + total_lost_arr, 1) * 100):.1f}%",
            "",
        ),
    ]
    for label, n, arr in summary_rows:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=2, value=n).alignment = RIGHT
        if isinstance(arr, (int, float)) and arr:
            cell = ws.cell(row=row, column=3, value=arr)
            cell.number_format = "#,##0"
            cell.alignment = RIGHT
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).border = BORDER
            if ci != 1:
                ws.cell(row=row, column=ci).font = BODY_FONT
        row += 1
    _set_widths(ws, [26, 16, 18])
    ws.freeze_panes = "A4"


def build_days_in_stage_sheet(wb, session, instance):
    """Pulls the SD Days in Stage report."""
    RID = "00OTb000008gUt7MAE"
    r = session.post(
        f"{instance}/services/data/v66.0/analytics/reports/{RID}?includeDetails=false",
        headers={"Content-Type": "application/json"},
    ).json()

    ws = wb.create_sheet("Days in Stage")
    ws["A1"] = "Days in Stage, Open Pipeline"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = ["Stage", "Open Deals", "Avg Days", "Max Days"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header(ws, 3, headers)
    row = 4
    for g in r.get("groupingsDown", {}).get("groupings", []):
        label = g.get("label", "")
        fm = r.get("factMap", {}).get(f"{g['key']}!T", {})
        aggs = fm.get("aggregates", [])
        avg = aggs[0].get("value", 0) if aggs else 0
        mx = aggs[1].get("value", 0) if len(aggs) > 1 else 0
        count = aggs[2].get("value", 0) if len(aggs) > 2 else 0
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        ws.cell(row=row, column=2, value=int(count)).alignment = CENTER
        ws.cell(row=row, column=3, value=round(avg, 1)).alignment = CENTER
        ws.cell(row=row, column=4, value=int(mx)).alignment = CENTER
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).font = BODY_FONT
            ws.cell(row=row, column=ci).border = BORDER
        row += 1
    _set_widths(ws, [22, 14, 14, 14])
    ws.freeze_panes = "A4"


def build_churn_sheet(wb):
    ws = wb.create_sheet("Churn Risk")
    ws["A1"] = "Churn Risk, Pending Finance Data Feed"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Churn data is held by Finance. Alex Profit receives it from "
        "a named Finance contact each month. Feed is not yet wired "
        "into Salesforce."
    )
    ws["A2"].font = BODY_FONT
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 40

    ws["A4"] = "Next step"
    ws["A4"].font = BODY_BOLD
    ws["B4"] = (
        "Confirm Finance contact with Alex Profit, agree refresh "
        "cadence, add the feed to the weekly extract."
    )
    ws["B4"].font = BODY_FONT
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 32

    # Proxy: top 10 at-risk renewals from Business At Risk (dashboard report)
    ws["A6"] = "Interim proxy"
    ws["A6"].font = BODY_BOLD
    ws["B6"] = (
        "Until the Finance feed lands, see Business At Risk tab in "
        f"the Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis workbook for renewals with a "
        "Termination Risk flag."
    )
    ws["B6"].font = BODY_FONT
    ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[6].height = 32
    _set_widths(ws, [20, 90])


def build_snapshot_trend_consolidated(wb, period_label, sheet_source, workbooks_dir):
    """Consolidate one historical-trending snapshot sheet across all directors.

    `sheet_source` is the quarter-specific source tab name, such as
    `Q2 Snapshot Trend`.
    Writes a single tab with a leading Territory column so managers can
    filter across the whole org.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws = wb.create_sheet(f"{period_label} Trend Consolidated")

    header_written = False
    all_rows = []
    columns = None
    snapshots = None
    for director_name, territory, fname, *_ in DIRECTORS:
        wb_path = workbooks_dir / fname
        if not wb_path.exists():
            continue
        try:
            dwb = load_workbook(wb_path, data_only=True, read_only=True)
        except Exception as exc:
            print(f"  [WARN] skipping {wb_path.name}: {exc}")
            continue
        if sheet_source not in dwb.sheetnames:
            continue
        sws = dwb[sheet_source]
        # Row 1 is the subtitle with snapshot dates; row 2 is headers; data
        # starts row 3.
        if not header_written:
            snapshots = str(sws.cell(row=1, column=1).value or "")
            columns = [
                sws.cell(row=2, column=c).value for c in range(1, sws.max_column + 1)
            ]
            header_written = True
        for r_idx in range(3, sws.max_row + 1):
            row = [
                sws.cell(row=r_idx, column=c).value
                for c in range(1, (sws.max_column or 0) + 1)
            ]
            if not any(v is not None for v in row):
                continue
            all_rows.append([territory] + row)

    if not header_written or not columns:
        ws.cell(
            row=1, column=1, value=f"No {sheet_source} data across directors."
        ).font = CAPTION_FONT
        return

    if snapshots:
        ws.cell(row=1, column=1, value=snapshots).font = Font(
            italic=True, size=9, color="595959"
        )

    # Identify the ARR and StageName columns so we can emit helper columns
    # (Initial ARR, Final ARR, Initial Stage, Final Stage, Bucket) that the
    # Forecast Variance tab uses as SUMIFS source.
    import re as _re

    arr_col_idxs = []  # position in columns list
    stage_col_idxs = []
    for ci, h in enumerate(columns or []):
        s = str(h or "")
        if _re.match(r"^ARR \d{4}-\d{2}-\d{2}$", s):
            arr_col_idxs.append(ci)
        elif _re.match(r"^StageName_ \d{4}-\d{2}-\d{2}$", s):
            stage_col_idxs.append(ci)

    helper_headers = (
        [
            "Initial ARR",
            "Final ARR",
            "Initial Stage",
            "Final Stage",
            "Bucket",
        ]
        if arr_col_idxs
        else []
    )

    headers = (
        ["Territory"]
        + [str(c) if c is not None else "" for c in columns]
        + helper_headers
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _apply_header(ws, 2, headers)
    ws.row_dimensions[2].height = 30

    n_src_cols = 1 + len(columns)  # Territory + original columns
    TERMINAL_STAGES = {"8 - Won", "0 - Lost", "0 - No Opportunity"}

    for r_idx, row in enumerate(all_rows, 3):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_i, value=val)
            cell.font = BODY_FONT
            if isinstance(val, (int, float)) and c_i > 4:
                cell.number_format = "#,##0"
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

        if not arr_col_idxs:
            continue
        # Helpers live AFTER the source columns. Column index in the SHEET is
        # 1 + source_idx (the +1 is because the source columns were written
        # starting at column 1 of the `row` array, which already includes the
        # leading Territory).
        row_offset = 1  # cells written 1-indexed
        initial_arr_val = row[arr_col_idxs[0] + row_offset]
        final_arr_val = row[arr_col_idxs[-1] + row_offset]
        initial_stage_val = (
            row[stage_col_idxs[0] + row_offset] if stage_col_idxs else None
        )
        final_stage_val = (
            row[stage_col_idxs[-1] + row_offset] if stage_col_idxs else None
        )
        try:
            initial_arr = float(initial_arr_val or 0)
        except (TypeError, ValueError):
            initial_arr = 0.0
        try:
            final_arr = float(final_arr_val or 0)
        except (TypeError, ValueError):
            final_arr = 0.0
        initial_stage = str(initial_stage_val or "").strip()
        final_stage = str(final_stage_val or "").strip()

        # Classify. Matches the Python logic in build_forecast_variance so
        # SUMIFS totals on Forecast Variance tab reconcile.
        started_terminal = initial_stage in TERMINAL_STAGES
        ended_won = final_stage == "8 - Won"
        ended_lost = final_stage in ("0 - Lost", "0 - No Opportunity")
        if started_terminal:
            bucket = "AlreadyClosed"
        elif ended_won:
            bucket = "Won"
        elif ended_lost:
            bucket = "Lost"
        elif initial_arr == 0 and final_arr > 0:
            bucket = "Added"
        else:
            delta = final_arr - initial_arr
            if delta > 0:
                bucket = "RevisedUp"
            elif delta < 0:
                bucket = "RevisedDown"
            else:
                bucket = "Unchanged"

        helper_start = n_src_cols + 1
        for off, val in enumerate(
            [initial_arr, final_arr, initial_stage, final_stage, bucket]
        ):
            cell = ws.cell(row=r_idx, column=helper_start + off, value=val)
            cell.font = BODY_FONT
            if off < 2:
                cell.number_format = "#,##0"
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

    # Named table for easy pivoting
    if all_rows:
        last_col = get_column_letter(len(headers))
        last_row = 2 + len(all_rows)
        table_name = f"{period_label}TrendConsolidated".replace(" ", "").replace(
            "-", ""
        )
        try:
            tbl = Table(displayName=table_name, ref=f"A2:{last_col}{last_row}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
            )
            ws.add_table(tbl)
        except ValueError:
            pass
    ws.freeze_panes = "C3"
    widths = [18] + [
        min(max(len(str(columns[i - 1] or "")) + 2, 12), 22)
        for i in range(1, len(columns) + 1)
    ]
    _set_widths(ws, widths)


def build_land_detail_sheet(wb):
    """Consolidate every director's Land open pipeline into one table so
    downstream tabs can SUMIFS against it."""
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws = wb.create_sheet("Land Pipeline Detail")
    headers = [
        "Director",
        "Territory",
        "Type",
        "Stage",
        "Forecast Category",
        "Account",
        "Opportunity",
        "Owner",
        "Close Date",
        "ARR Unwtd",
        "ARR Wtd",
        "Push Count",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header(ws, 1, headers)

    r = 2
    for director in data_store:
        d = data_store[director]
        # pull every Land open row from the raw workbook
        for row in d.get("all_land_open_rows", []):
            values = [
                director,
                row.get("Sales Region") or row.get("territory", ""),
                row.get("Type", ""),
                row.get("Stage", ""),
                row.get("Forecast Category", ""),
                row.get("Account", ""),
                row.get("Opportunity", ""),
                row.get("Owner", ""),
                str(row.get("Close Date", "") or "")[:10],
                float(row.get("ARR Unweighted (EUR)") or 0),
                float(row.get("ARR Weighted (EUR)") or 0),
                int(row.get("Push Count") or 0),
            ]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = BODY_FONT
                if ci in (10, 11):
                    cell.alignment = RIGHT
                    cell.number_format = "#,##0"
                elif ci == 12:
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT
            r += 1

    if r > 2:
        last_col = get_column_letter(len(headers))
        tbl = Table(displayName="LandPipelineDetail", ref=f"A1:{last_col}{r - 1}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)
    ws.freeze_panes = "A2"
    _set_widths(ws, [20, 22, 10, 20, 16, 30, 40, 22, 12, 14, 14, 10])
    return "LandPipelineDetail"


def build_land_won_lost_detail_sheet(wb):
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws = wb.create_sheet("Land WonLost Detail")
    headers = [
        "Director",
        "Territory",
        "Type",
        "Stage",
        "Account",
        "Opportunity",
        "Owner",
        "Close Date",
        "ARR Unwtd",
        "Reason",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header(ws, 1, headers)

    r = 2
    for director in data_store:
        d = data_store[director]
        for row in d.get("all_land_won_lost_rows", []):
            values = [
                director,
                row.get("Sales Region", ""),
                row.get("Type", ""),
                row.get("Stage", ""),
                row.get("Account", ""),
                row.get("Opportunity", ""),
                row.get("Owner", ""),
                str(row.get("Close Date", "") or "")[:10],
                float(row.get("ARR Unweighted (EUR)") or 0),
                row.get("Reason", ""),
            ]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = BODY_FONT
                if ci == 9:
                    cell.alignment = RIGHT
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = LEFT
            r += 1

    if r > 2:
        last_col = get_column_letter(len(headers))
        tbl = Table(displayName="LandWonLostDetail", ref=f"A1:{last_col}{r - 1}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)
    ws.freeze_panes = "A2"
    _set_widths(ws, [20, 22, 10, 20, 30, 40, 22, 12, 14, 24])


def build_summary_live(wb, data):
    """Summary tab where every number is a SUMIFS/COUNTIFS formula pointing
    to the detail tabs. Edit the detail row, Summary updates."""
    ws = wb.create_sheet("Summary Live")
    ws["A1"] = f"{RUNTIME_PERIOD['fy_label']} Pipeline Review, Live Formulas"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"Every cell below is a SUMIFS/COUNTIFS formula against the "
        f"Land Pipeline Detail and Land WonLost Detail tabs. Updated "
        f"{RUNTIME_PERIOD['report_date_display']}."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT

    headers = [
        "Director",
        "Open Land Deals",
        "Open Land ARR Unwtd",
        "Open Land ARR Wtd",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won Deals",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost Deals",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost ARR",
    ]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    _apply_header(ws, start, headers)

    r = start + 1
    for director_name, territory, *_rest in data:
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        # COUNTIFS open Land deals
        ws.cell(
            row=r,
            column=2,
            value=f'=COUNTIFS(LandPipelineDetail[Director],A{r},LandPipelineDetail[Type],"Land")',
        ).alignment = CENTER
        ws.cell(
            row=r,
            column=3,
            value=f'=SUMIFS(LandPipelineDetail[ARR Unwtd],LandPipelineDetail[Director],A{r},LandPipelineDetail[Type],"Land")',
        )
        ws.cell(
            row=r,
            column=4,
            value=f'=SUMIFS(LandPipelineDetail[ARR Wtd],LandPipelineDetail[Director],A{r},LandPipelineDetail[Type],"Land")',
        )
        ws.cell(
            row=r,
            column=5,
            value=(
                f"=COUNTIFS(LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Won*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        ).alignment = CENTER
        ws.cell(
            row=r,
            column=6,
            value=(
                f"=SUMIFS(LandWonLostDetail[ARR Unwtd],"
                f"LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Won*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        )
        ws.cell(
            row=r,
            column=7,
            value=(
                f"=COUNTIFS(LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Lost*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        ).alignment = CENTER
        ws.cell(
            row=r,
            column=8,
            value=(
                f"=SUMIFS(LandWonLostDetail[ARR Unwtd],"
                f"LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Lost*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        )
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        for col_letter in ("C", "D", "F", "H"):
            ws[f"{col_letter}{r}"].number_format = "#,##0"
            ws[f"{col_letter}{r}"].alignment = RIGHT
        r += 1
    _zebra(ws, start + 1, r - 1, len(headers))
    _set_widths(ws, [22, 14, 18, 18, 12, 16, 12, 16])
    ws.freeze_panes = "A5"


def build_pipeline_pivot(wb):
    """Stage x Director pivot of open Land pipeline with stacked bar chart.

    Answers: Where is pipeline concentrated by stage, per director? Which stages
    are thin vs thick? Which directors have bookable-stage (5+) weight?
    """
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Pipeline Pivot")
    ws["A1"] = "Land Pipeline by Stage and Director"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Open Land deals, ARR Unweighted (EUR). Stacked bar shows how each "
        "director's pipeline composes by stage. Source: Land Pipeline Detail."
    )
    ws["A2"].font = CAPTION_FONT

    agg = defaultdict(lambda: defaultdict(float))
    stages_seen = set()
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            stage = str(r.get("Stage") or "(unknown)")
            arr = float(r.get("ARR Unweighted (EUR)") or 0)
            agg[director_name][stage] += arr
            stages_seen.add(stage)

    stages = sorted(stages_seen)
    headers = ["Director"] + stages + ["Total"]

    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28
    header_row = row

    row += 1
    first_data_row = row
    col_totals = [0.0] * len(stages)
    for director_name in data_store:
        ws.cell(row=row, column=1, value=director_name).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER
        row_total = 0.0
        for si, stage in enumerate(stages, 2):
            v = agg[director_name].get(stage, 0.0)
            c = ws.cell(row=row, column=si, value=v if v else None)
            c.font = BODY_FONT
            c.number_format = "#,##0"
            c.alignment = RIGHT
            c.border = BORDER
            row_total += v
            col_totals[si - 2] += v
        tc = ws.cell(row=row, column=len(stages) + 2, value=row_total)
        tc.font = BODY_BOLD
        tc.number_format = "#,##0"
        tc.alignment = RIGHT
        tc.border = BORDER
        row += 1
    last_data_row = row - 1

    ws.cell(row=row, column=1, value="TOTAL").font = BODY_BOLD
    ws.cell(row=row, column=1).border = BORDER
    for si, t in enumerate(col_totals, 2):
        c = ws.cell(row=row, column=si, value=t)
        c.font = BODY_BOLD
        c.number_format = "#,##0"
        c.border = BORDER
    gtc = ws.cell(row=row, column=len(stages) + 2, value=sum(col_totals))
    gtc.font = BODY_BOLD
    gtc.number_format = "#,##0"
    gtc.border = BORDER

    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Open Pipeline by Stage, stacked per Director"
    chart.y_axis.title = "ARR (EUR)"
    chart.x_axis.title = "Director"
    data_ref = Reference(
        ws,
        min_col=2,
        max_col=len(stages) + 1,
        min_row=header_row,
        max_row=last_data_row,
    )
    cats_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 12
    chart.width = 24
    anchor = get_column_letter(len(stages) + 4)
    ws.add_chart(chart, f"{anchor}{header_row}")

    ws.column_dimensions["A"].width = 22
    for ci in range(2, len(stages) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.freeze_panes = f"B{first_data_row}"


def build_arr_concentration(wb):
    """Top 20 open deals + Pareto. Highlights single-deal dependency."""
    from openpyxl.chart import BarChart, LineChart, Reference

    ws = wb.create_sheet("ARR Concentration")
    ws["A1"] = "ARR Concentration, Top 20 Open Land Deals"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Ranked by ARR Unweighted. Cumulative share shows how much of total "
        "pipeline sits in the top deals (single-deal exposure risk)."
    )
    ws["A2"].font = CAPTION_FONT

    deals = []
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            deals.append(
                {
                    "director": director_name,
                    "account": r.get("Account"),
                    "opportunity": r.get("Opportunity"),
                    "stage": r.get("Stage"),
                    "owner": r.get("Owner"),
                    "arr": float(r.get("ARR Unweighted (EUR)") or 0),
                }
            )
    deals.sort(key=lambda x: x["arr"], reverse=True)
    total = sum(d["arr"] for d in deals) or 1.0
    top20 = deals[:20]

    headers = [
        "Rank",
        "Director",
        "Account",
        "Opportunity",
        "Stage",
        "Owner",
        "ARR (EUR)",
        "% of Total",
        "Cumulative %",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28
    header_row = row

    row += 1
    first_data_row = row
    cum = 0.0
    for rank, d in enumerate(top20, 1):
        cum += d["arr"]
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=d["director"])
        ws.cell(row=row, column=3, value=d["account"])
        ws.cell(row=row, column=4, value=d["opportunity"])
        ws.cell(row=row, column=5, value=d["stage"])
        ws.cell(row=row, column=6, value=d["owner"])
        arr_c = ws.cell(row=row, column=7, value=d["arr"])
        arr_c.number_format = "#,##0"
        arr_c.alignment = RIGHT
        pct_c = ws.cell(row=row, column=8, value=d["arr"] / total)
        pct_c.number_format = "0.0%"
        pct_c.alignment = RIGHT
        cum_c = ws.cell(row=row, column=9, value=cum / total)
        cum_c.number_format = "0.0%"
        cum_c.alignment = RIGHT
        for ci in range(1, 10):
            ws.cell(row=row, column=ci).border = BORDER
            ws.cell(row=row, column=ci).font = BODY_FONT
        row += 1
    last_data_row = row - 1

    # Data bar on ARR column (column 7) for instant visual ranking + gradient
    # on cumulative share column (column 9) so Pareto is obvious at a glance.
    if last_data_row >= first_data_row:
        from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

        ws.conditional_formatting.add(
            f"G{first_data_row}:G{last_data_row}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="4472C4",
                showValue=True,
            ),
        )
        ws.conditional_formatting.add(
            f"I{first_data_row}:I{last_data_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F7F9FC",
                mid_type="num",
                mid_value=0.5,
                mid_color="9AC0EA",
                end_type="num",
                end_value=1,
                end_color="2E5AAC",
            ),
        )

    row += 1
    summary = (
        f"Top 20 account for {cum / total:.0%} of {total / 1e6:.1f}M EUR total "
        f"pipeline ({len(deals)} open Land deals)."
    )
    ws.cell(row=row, column=1, value=summary).font = BODY_BOLD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Top 20 Deals ARR (bars) and Cumulative Share (line)"
    chart.y_axis.title = "ARR (EUR)"
    chart.x_axis.title = "Rank"
    data_ref = Reference(ws, min_col=7, min_row=header_row, max_row=last_data_row)
    cats_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 12
    chart.width = 22

    line = LineChart()
    line_ref = Reference(ws, min_col=9, min_row=header_row, max_row=last_data_row)
    line.add_data(line_ref, titles_from_data=True)
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    line.y_axis.number_format = "0%"
    chart += line
    ws.add_chart(chart, "K4")

    widths = [6, 18, 28, 34, 18, 22, 14, 12, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_pipeline_velocity(wb, workbooks_dir):
    """Time-series of prior/current quarter ARR at each snapshot date.

    Answers: Is pipeline growing or slipping between snapshots? Which territories
    are losing ARR quarter-over-quarter?
    """
    import re
    from openpyxl.chart import LineChart, Reference

    contract = _historical_trending_contract()
    ws = wb.create_sheet("Pipeline Velocity")
    ws["A1"] = "Pipeline Velocity, Historical Trending Snapshots"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Sum of ARR (EUR) at each Historical Trending snapshot date per director. "
        "Rising line = pipeline growing. Falling = slipping or being closed. "
        f"Source: {contract.retrospective_snapshot_sheet} and "
        f"{contract.current_snapshot_sheet} sheets in per-director workbooks."
    )
    ws["A2"].font = CAPTION_FONT

    per_director_retrospective = defaultdict(dict)
    per_director_current = defaultdict(dict)

    for director_name, _territory, fname, _oid, _tid in DIRECTORS:
        wb_path = workbooks_dir / fname
        if not wb_path.exists():
            continue
        sub = load_workbook(wb_path, read_only=True, data_only=True)
        for source_sheet, bucket in [
            (
                contract.retrospective_snapshot_sheet,
                per_director_retrospective,
            ),
            (contract.current_snapshot_sheet, per_director_current),
        ]:
            if source_sheet not in sub.sheetnames:
                continue
            ws_src = sub[source_sheet]
            rows_iter = list(ws_src.iter_rows(values_only=True))
            if len(rows_iter) < 3:
                continue
            header = rows_iter[1]
            arr_cols = []
            for ci, h in enumerate(header):
                if h and isinstance(h, str) and re.match(r"^ARR \d{4}-\d{2}-\d{2}$", h):
                    arr_cols.append((ci, h.split(" ", 1)[1]))
            for r in rows_iter[2:]:
                for ci, date in arr_cols:
                    v = r[ci] if ci < len(r) else None
                    try:
                        bucket[director_name][date] = bucket[director_name].get(
                            date, 0.0
                        ) + (float(v) if v is not None else 0.0)
                    except (TypeError, ValueError):
                        pass
        sub.close()

    row = 4
    for period, per_director in [
        (
            f"{contract.retrospective_title} (Historical)",
            per_director_retrospective,
        ),
        (f"{contract.current_title} (Current)", per_director_current),
    ]:
        if not per_director:
            continue
        dates = sorted({d for dm in per_director.values() for d in dm})
        ws.cell(
            row=row, column=1, value=f"{period}, ARR by Snapshot (EUR)"
        ).font = BODY_BOLD
        row += 1

        header_cells = ["Director"] + dates
        header_row = row
        for ci, h in enumerate(header_cells, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
        ws.row_dimensions[row].height = 24
        row += 1

        for director_name in per_director:
            ws.cell(row=row, column=1, value=director_name).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=1).border = BORDER
            for ci, d in enumerate(dates, 2):
                v = per_director[director_name].get(d, 0.0)
                c = ws.cell(row=row, column=ci, value=v if v else None)
                c.font = BODY_FONT
                c.number_format = "#,##0"
                c.border = BORDER
                c.alignment = RIGHT
            row += 1
        last_data_row = row - 1

        chart = LineChart()
        chart.title = f"{period} Pipeline ARR over Time"
        chart.y_axis.title = "ARR (EUR)"
        chart.x_axis.title = "Snapshot Date"
        data_ref = Reference(
            ws,
            min_col=1,
            max_col=len(dates) + 1,
            min_row=header_row + 1,
            max_row=last_data_row,
        )
        chart.add_data(data_ref, titles_from_data=True, from_rows=True)
        cats_ref = Reference(
            ws,
            min_col=2,
            max_col=len(dates) + 1,
            min_row=header_row,
            max_row=header_row,
        )
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 22
        anchor = get_column_letter(len(dates) + 3)
        ws.add_chart(chart, f"{anchor}{header_row}")

        row += 3

    # ── Full-FY current snapshot (complements HT trending for Q1-Q2) ──
    # Historical Trending reports only cover Q1+Q2. For Q3+Q4 context we can't
    # show a trend, but we CAN show the current open ARR per director × quarter.
    # This gives directors the full-year book at a glance alongside the trend.
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"{RUNTIME_PERIOD['fy_label']} Current Open Pipeline, by Quarter",
    ).font = BODY_BOLD
    row += 1
    snap_header_row = row
    q_labels = [
        "Q1 (Jan-Mar)",
        "Q2 (Apr-Jun)",
        "Q3 (Jul-Sep)",
        "Q4 (Oct-Dec)",
        f"{RUNTIME_PERIOD['fy_label']} Total",
    ]
    snap_headers = ["Director"] + q_labels
    for ci, h in enumerate(snap_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 24
    row += 1
    snap_first = row

    def _qtr_of(date_str: str) -> int:
        s = str(date_str or "")[:10]
        if len(s) < 7 or not s.startswith(f"{RUNTIME_PERIOD['analysis_year']}"):
            return 0
        try:
            m = int(s[5:7])
        except ValueError:
            return 0
        return (m - 1) // 3 + 1

    total_by_q = [0.0, 0.0, 0.0, 0.0]
    for director_name, d in data_store.items():
        per_q = [0.0, 0.0, 0.0, 0.0]
        for r in d.get("all_land_open_rows", []):
            q = _qtr_of(r.get("Close Date"))
            if 1 <= q <= 4:
                per_q[q - 1] += float(r.get("ARR Unweighted (EUR)") or 0)
        row_total = sum(per_q)
        for i in range(4):
            total_by_q[i] += per_q[i]
        vals = [director_name] + per_q + [row_total]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci > 1:
                c.number_format = "#,##0"
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1
    # TOTAL row
    total_vals = ["TOTAL"] + total_by_q + [sum(total_by_q)]
    for ci, v in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        if ci > 1:
            c.number_format = "#,##0"
            c.alignment = RIGHT
        else:
            c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_BOLD
    snap_last = row
    if snap_last >= snap_first:
        from openpyxl.formatting.rule import DataBarRule

        for ci in range(2, 6):  # Q1-Q4 columns
            letter = get_column_letter(ci)
            ws.conditional_formatting.add(
                f"{letter}{snap_first}:{letter}{snap_last - 1}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="4472C4",
                    showValue=True,
                ),
            )

    ws.column_dimensions["A"].width = 22
    for ci in range(2, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 15


def build_slip_risk(wb):
    """Owner-level push intensity and open ARR at risk."""
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Slip Risk by Owner")
    ws["A1"] = "Slip Risk by Owner"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Open Land deals aggregated per owner. High Total Pushes plus high Open "
        "ARR = highest slip risk exposure. Source: Land Pipeline Detail."
    )
    ws["A2"].font = CAPTION_FONT

    owners = defaultdict(
        lambda: {
            "director": "",
            "deal_count": 0,
            "arr": 0.0,
            "total_push": 0,
            "max_push": 0,
        }
    )
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            owner = str(r.get("Owner") or "(unassigned)")
            arr = float(r.get("ARR Unweighted (EUR)") or 0)
            push = int(r.get("Push Count") or 0)
            o = owners[owner]
            o["director"] = director_name
            o["deal_count"] += 1
            o["arr"] += arr
            o["total_push"] += push
            o["max_push"] = max(o["max_push"], push)

    sorted_owners = sorted(
        owners.items(), key=lambda x: (-x[1]["total_push"], -x[1]["arr"])
    )
    top = sorted_owners[:25]

    headers = [
        "Owner",
        "Director",
        "# Deals",
        "Open ARR (EUR)",
        "Total Pushes",
        "Avg Push/Deal",
        "Max Pushes on 1 Deal",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28
    header_row = row

    row += 1
    first_data_row = row
    for owner, stats in top:
        ws.cell(row=row, column=1, value=owner)
        ws.cell(row=row, column=2, value=stats["director"])
        ws.cell(row=row, column=3, value=stats["deal_count"])
        arr_c = ws.cell(row=row, column=4, value=stats["arr"])
        arr_c.number_format = "#,##0"
        arr_c.alignment = RIGHT
        ws.cell(row=row, column=5, value=stats["total_push"])
        avg = stats["total_push"] / stats["deal_count"] if stats["deal_count"] else 0
        avg_c = ws.cell(row=row, column=6, value=round(avg, 1))
        avg_c.alignment = RIGHT
        ws.cell(row=row, column=7, value=stats["max_push"])
        for ci in range(1, 8):
            ws.cell(row=row, column=ci).border = BORDER
            ws.cell(row=row, column=ci).font = BODY_FONT
        row += 1
    last_data_row = row - 1

    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Top 25 Owners by Total Push Count"
    chart.y_axis.title = "Owner"
    chart.x_axis.title = "Total Pushes (open deals)"
    data_ref = Reference(ws, min_col=5, min_row=header_row, max_row=last_data_row)
    cats_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 16
    chart.width = 18
    ws.add_chart(chart, "I4")

    widths = [28, 20, 10, 16, 13, 14, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_territory_scorecard(wb, overdue_rows, kyc_rows):
    """Cross-KPI heatmap: one row per director, color-coded columns."""
    from collections import Counter
    from openpyxl.formatting.rule import ColorScaleRule

    ws = wb.create_sheet("Territory Scorecard")
    ws["A1"] = "Territory Scorecard"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "One row per director. Each numeric column has an independent heatmap: "
        "green-to-red for bad-when-high (overdue, slips, lost), red-to-green "
        "for good-when-high (coverage, win rate, won ARR). All ARR in EUR."
    )
    ws["A2"].font = CAPTION_FONT

    terr_to_dir = {t: n for n, t, *_ in DIRECTORS}
    overdue_by_dir = Counter(
        terr_to_dir.get(r.get("territory", ""), "") for r in overdue_rows
    )
    kyc_by_dir = Counter(
        terr_to_dir.get(r.get("territory", ""), r.get("region", "")) for r in kyc_rows
    )

    headers = [
        "Director",
        "Open Deals",
        "Open ARR Unwtd",
        "Open ARR Wtd",
        "Avg Probability %",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Win Rate %",
        "Slips Still Open",
        "Slip ARR",
        "Overdue Opps",
        "KYC Missing",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 32

    row += 1
    first_data_row = row
    for name, d in data_store.items():
        q1w = d["q1_won_arr"]
        q1l = d["q1_lost_arr"]
        q1_total = q1w + q1l
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        ws.cell(row=row, column=2, value=d["open_count"])
        c = ws.cell(row=row, column=3, value=d["open_unwtd"])
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=4, value=d["open_wtd"])
        c.number_format = "#,##0"
        cov = d["open_wtd"] / d["open_unwtd"] if d["open_unwtd"] else 0
        c = ws.cell(row=row, column=5, value=cov)
        c.number_format = "0.0%"
        c = ws.cell(row=row, column=6, value=q1w)
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=7, value=q1l)
        c.number_format = "#,##0"
        win_rate = q1w / q1_total if q1_total else 0
        c = ws.cell(row=row, column=8, value=win_rate)
        c.number_format = "0.0%"
        ws.cell(row=row, column=9, value=d["slip_still_open_count"])
        c = ws.cell(row=row, column=10, value=d["slip_still_open_arr"])
        c.number_format = "#,##0"
        ws.cell(row=row, column=11, value=overdue_by_dir.get(name, 0))
        ws.cell(row=row, column=12, value=kyc_by_dir.get(name, 0))
        for ci in range(1, 13):
            ws.cell(row=row, column=ci).border = BORDER
            if ci > 1:
                ws.cell(row=row, column=ci).alignment = RIGHT
        row += 1
    last_data_row = row - 1

    # Good-when-high columns (green at top): coverage %, Q1 Won ARR, Win Rate %
    good_when_high = {5, 6, 8}
    # Everything else numeric is bad-when-high
    for col_idx in range(2, 13):
        col_letter = get_column_letter(col_idx)
        if col_idx in good_when_high:
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            )
        ws.conditional_formatting.add(
            f"{col_letter}{first_data_row}:{col_letter}{last_data_row}", rule
        )

    widths = [22, 10, 16, 16, 12, 14, 14, 12, 12, 14, 12, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"B{first_data_row}"


STAGE_ORDER = {
    "0 - No Opportunity": -2,
    "0 - Lost": -1,
    "1 - Prospecting": 1,
    "2 - Discovery": 2,
    "3 - Engagement": 3,
    "4 - Shortlisted": 4,
    "5 - Preferred": 5,
    "6 - Contracting": 6,
    "8 - Won": 8,
}


# Audit-surfaced thresholds. These are the ONLY place in the codebase where
# risk scoring cutoffs and insight thresholds live. The Parameters tab
# (build_parameters_sheet) writes each value out to a named cell in the
# workbook so an analyst can inspect them without opening the code. The
# Methodology tab documents the rationale for each.
RISK_RULES = [
    # (code, description, weight, trigger)
    ("PUSH_HIGH", "Push count >= 5 (repeat slip pattern)", 40, "push >= 5"),
    ("PUSH_MED", "Push count 3-4", 20, "push >= 3"),
    ("OVERDUE", "Close date in the past, still open", 50, "days_to_close < 0"),
    (
        "CLOSE_SOON",
        "Close <30 days away and stage below 4-Shortlisted",
        30,
        "days_to_close < 30 and stage_num < 4",
    ),
    ("STALE", "No activity for 60+ days", 15, "days_since_activity > 60"),
    ("NO_NEXT_STEP", "Next Step field empty or '-'", 10, "next_step empty"),
    (
        "LOW_FCST",
        "Weighted/unweighted coverage below 20% on deal above 500K",
        15,
        "fcst/arr < 0.2 and arr > 500K",
    ),
    (
        "HIGH_VALUE_PUSH",
        "ARR above 1M and has been pushed at least twice",
        10,
        "arr > 1M and push >= 2",
    ),
]

INSIGHT_THRESHOLDS = {
    "concentration_top3_pct_min": 0.3,
    "velocity_material_change_eur": 100_000,
    "loss_driven_ratio": 2.0,
    "risk_score_triage_floor": 60,
    "slip_owner_pushes_min": 20,
}


def _stage_num(stage):
    return STAGE_ORDER.get(str(stage or "").strip(), 0)


def _parse_date(v):
    import datetime as _dt

    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v)[:10]
    try:
        return _dt.date.fromisoformat(s)
    except ValueError:
        return None


def build_parameters_sheet(wb):
    """Surface every threshold and rule as a named cell.

    The Parameters tab is the single visible audit document. Each risk
    rule, each insight threshold, and stage ordering is written here so
    an analyst can see exactly what drives the scoring without reading
    Python. Named cells (RiskWeight_*, Thresh_*) are referenced by
    other tabs via defined names so changing a number in-workbook
    changes the downstream formulas.
    """
    ws = wb.create_sheet("Parameters", 1)
    ws["A1"] = "Analysis Parameters and Thresholds"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Every rule the analysis applies. Each value is addressable by name "
        "(see Defined Names in the workbook) so formulas on other tabs "
        "reference these cells directly. Change a number here to see where "
        "it propagates, then rebuild to recompute. Rationale for each rule "
        "is on the Methodology tab."
    )
    ws["A2"].font = CAPTION_FONT

    # Risk scoring rules
    row = 4
    ws.cell(row=row, column=1, value="Risk scoring rules").font = BODY_BOLD
    row += 1
    headers = ["Code", "Trigger description", "Weight", "Rule in code"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    row += 1
    for code, desc, weight, trigger in RISK_RULES:
        ws.cell(row=row, column=1, value=code).font = Font(
            name="Calibri", size=10, bold=True
        )
        ws.cell(row=row, column=2, value=desc).alignment = LEFT
        weight_cell = ws.cell(row=row, column=3, value=weight)
        weight_cell.alignment = RIGHT
        weight_cell.font = BODY_BOLD
        ws.cell(row=row, column=4, value=trigger).font = Font(
            name="Consolas", size=9, color="595959"
        )
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).border = BORDER
        # Defined name: RiskWeight_PUSH_HIGH, etc.
        try:
            from openpyxl.workbook.defined_name import DefinedName

            dn = DefinedName(
                f"RiskWeight_{code}",
                attr_text=f"Parameters!$C${row}",
            )
            wb.defined_names[f"RiskWeight_{code}"] = dn
        except Exception:
            pass
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Insight thresholds").font = BODY_BOLD
    row += 1
    ws.cell(row=row, column=1, value="Name")
    ws.cell(row=row, column=2, value="Value")
    ws.cell(row=row, column=3, value="What it controls")
    for ci in range(1, 4):
        c = ws.cell(row=row, column=ci)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    row += 1
    thresh_rows = [
        (
            "Thresh_ConcentrationTop3Pct",
            INSIGHT_THRESHOLDS["concentration_top3_pct_min"],
            "Minimum top-3-deal share before the concentration finding "
            "appears in Executive Insights.",
            "0.0%",
        ),
        (
            "Thresh_VelocityMaterialEUR",
            INSIGHT_THRESHOLDS["velocity_material_change_eur"],
            f"Minimum absolute {RUNTIME_PERIOD['prior_quarter_label']} snapshot change (EUR) to call out as "
            "material in Executive Insights.",
            "#,##0",
        ),
        (
            "Thresh_LossDrivenRatio",
            INSIGHT_THRESHOLDS["loss_driven_ratio"],
            "Ratio of losses to wins (or vice versa) that triggers the "
            "loss-driven/win-driven narrative in Forecast Variance slides.",
            "0.00",
        ),
        (
            "Thresh_RiskScoreTriage",
            INSIGHT_THRESHOLDS["risk_score_triage_floor"],
            "Risk score at or above which deals are called out in the "
            "Executive Insights triage bullet.",
            "0",
        ),
        (
            "Thresh_SlipOwnerPushes",
            INSIGHT_THRESHOLDS["slip_owner_pushes_min"],
            "Total pushes across open deals before an owner is flagged in "
            "the Slip Risk summary.",
            "0",
        ),
    ]
    for name, val, desc, fmt in thresh_rows:
        ws.cell(row=row, column=1, value=name).font = Font(
            name="Consolas", size=10, bold=True
        )
        val_cell = ws.cell(row=row, column=2, value=val)
        val_cell.font = BODY_BOLD
        val_cell.alignment = RIGHT
        val_cell.number_format = fmt
        ws.cell(row=row, column=3, value=desc).alignment = LEFT
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).border = BORDER
        try:
            from openpyxl.workbook.defined_name import DefinedName

            dn = DefinedName(name, attr_text=f"Parameters!$B${row}")
            wb.defined_names[name] = dn
        except Exception:
            pass
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Stage numbering").font = BODY_BOLD
    row += 1
    ws.cell(row=row, column=1, value="Stage name").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Stage number").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    for stage, num in sorted(STAGE_ORDER.items(), key=lambda x: x[1]):
        ws.cell(row=row, column=1, value=stage)
        ws.cell(row=row, column=2, value=num).alignment = RIGHT
        for ci in range(1, 3):
            ws.cell(row=row, column=ci).border = BORDER
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 36


def build_deal_risk_scoring(wb, run_date):
    """Composite risk score per open Land deal. Ranked top 30.

    Signals: push count, close-date slip risk, stage staleness, missing
    next step, low weighted coverage. Each contributes to a 0-100-ish
    score with reason codes so sales ops can triage.
    """
    import datetime as _dt

    from openpyxl.formatting.rule import ColorScaleRule

    ws = wb.create_sheet("Deal Risk Scoring")
    ws["A1"] = "Deal Risk Scoring, Open Land Pipeline"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Composite risk score per open Land deal. Higher = more at risk. "
        "Reason codes: PUSH (>=3 pushes), OVERDUE (close date past), "
        "CLOSE_SOON (close <30d and stage<4), STALE (no activity 60d+), "
        "NO_NEXT_STEP, LOW_FCST (weighted coverage <20% on >500K deal), "
        "HIGH_VALUE_PUSH (>1M and pushed)."
    )
    ws["A2"].font = CAPTION_FONT

    try:
        today = _dt.date.fromisoformat(run_date)
    except (ValueError, TypeError):
        today = _dt.date.today()

    # Build weight lookup from the RISK_RULES constant so Parameters tab and
    # scoring share a single source.
    rw = {code: weight for code, _desc, weight, _trigger in RISK_RULES}

    scored = []
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            arr = float(r.get("ARR Unweighted (EUR)") or 0)
            fcst = float(r.get("ARR Weighted (EUR)") or 0)
            push = int(r.get("Push Count") or 0)
            stage = str(r.get("Stage") or "")
            stage_n = _stage_num(stage)
            close_d = _parse_date(r.get("Close Date"))
            last_act = _parse_date(r.get("Last Activity"))
            next_step = str(r.get("Next Step") or "").strip()

            days_to_close = (close_d - today).days if close_d else None
            days_since_act = (today - last_act).days if last_act else None

            score = 0
            proof_parts = []  # e.g. ["PUSH_HIGH(+40)", "STALE(+15)"]

            def _add(code):
                nonlocal score
                w = rw[code]
                score += w
                proof_parts.append(f"{code}(+{w})")

            if push >= 5:
                _add("PUSH_HIGH")
            elif push >= 3:
                _add("PUSH_MED")
            if days_to_close is not None and days_to_close < 0:
                _add("OVERDUE")
            elif (
                days_to_close is not None
                and days_to_close < 30
                and stage_n > 0
                and stage_n < 4
            ):
                _add("CLOSE_SOON")
            if days_since_act is not None and days_since_act > 60:
                _add("STALE")
            if not next_step or next_step in ("-", "–"):
                _add("NO_NEXT_STEP")
            coverage = fcst / arr if arr else 1
            if arr > 500_000 and coverage < 0.2:
                _add("LOW_FCST")
            if arr > 1_000_000 and push >= 2:
                _add("HIGH_VALUE_PUSH")

            if score == 0:
                continue
            proof = " + ".join(proof_parts) + f" = {score}"
            # Short reason codes column remains for brevity (without weights)
            reasons = [p.split("(")[0] for p in proof_parts]
            scored.append(
                {
                    "score": score,
                    "director": director_name,
                    "account": r.get("Account", ""),
                    "opportunity": r.get("Opportunity", ""),
                    "stage": stage,
                    "owner": r.get("Owner", ""),
                    "close_date": str(close_d) if close_d else "",
                    "days_to_close": days_to_close if days_to_close is not None else "",
                    "days_since_act": (
                        days_since_act if days_since_act is not None else ""
                    ),
                    "arr": arr,
                    "push": push,
                    "reasons": ", ".join(reasons),
                    "proof": proof,
                }
            )

    scored.sort(key=lambda x: (-x["score"], -x["arr"]))
    top = scored[:30]

    headers = [
        "Rank",
        "Risk Score",
        "Director",
        "Account",
        "Opportunity",
        "Stage",
        "Owner",
        "Close Date",
        "Days to Close",
        "Days Since Activity",
        "ARR (EUR)",
        "Pushes",
        "Reason Codes",
        "Proof (rules triggered, each weight from Parameters)",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 30

    row += 1
    first_data_row = row
    for rank, s in enumerate(top, 1):
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=s["score"])
        ws.cell(row=row, column=3, value=s["director"])
        ws.cell(row=row, column=4, value=s["account"])
        ws.cell(row=row, column=5, value=s["opportunity"])
        ws.cell(row=row, column=6, value=s["stage"])
        ws.cell(row=row, column=7, value=s["owner"])
        ws.cell(row=row, column=8, value=s["close_date"])
        ws.cell(row=row, column=9, value=s["days_to_close"])
        ws.cell(row=row, column=10, value=s["days_since_act"])
        arr_c = ws.cell(row=row, column=11, value=s["arr"])
        arr_c.number_format = "#,##0"
        arr_c.alignment = RIGHT
        ws.cell(row=row, column=12, value=s["push"])
        ws.cell(row=row, column=13, value=s["reasons"])
        proof_cell = ws.cell(row=row, column=14, value=s["proof"])
        proof_cell.font = Font(name="Consolas", size=9, color="595959")
        for ci in range(1, 15):
            ws.cell(row=row, column=ci).border = BORDER
            if ci != 14:
                ws.cell(row=row, column=ci).font = BODY_FONT
        row += 1
    last_data_row = row - 1

    if last_data_row >= first_data_row:
        rule = ColorScaleRule(
            start_type="min",
            start_color="FFEB84",
            mid_type="percentile",
            mid_value=50,
            mid_color="F8696B",
            end_type="max",
            end_color="9C0006",
        )
        ws.conditional_formatting.add(f"B{first_data_row}:B{last_data_row}", rule)

    # Footer note with trail back to Parameters
    ws.cell(
        row=last_data_row + 2,
        column=1,
        value=(
            "Weights shown in Proof column are defined on the Parameters tab "
            "(RiskWeight_<CODE>). Trigger conditions and rationale on "
            "Methodology. Raw deal data is on Land Pipeline Detail."
        ),
    ).font = CAPTION_FONT
    ws.merge_cells(
        start_row=last_data_row + 2,
        start_column=1,
        end_row=last_data_row + 2,
        end_column=14,
    )

    widths = [5, 10, 18, 26, 30, 16, 20, 12, 12, 14, 14, 8, 28, 45]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"C{first_data_row}"


def build_commit_accuracy(wb):
    """Prior-quarter commit accuracy per owner from the trend workbook.

    Operational definition (no ForecastingItem history required):
      commit book = deals that sat in Stage 4+ (Shortlisted+) at the first
        snapshot of the quarter (Initial Stage column).
      commit won = of those, deals now in Bucket=Won at end of quarter.
      accuracy = commit_won_arr / commit_book_arr.

    This is intentionally stage-based rather than ForecastCategory-based so
    it works from the data we already extract. When OpportunityFieldHistory
    capture for ForecastCategoryName is added, swap the stage check for a
    ForecastCategory == 'Commit' check on the starting snapshot and rerun.
    """
    from openpyxl.formatting.rule import ColorScaleRule

    contract = _historical_trending_contract()
    src_sheet_name = contract.retrospective_consolidated_sheet
    if src_sheet_name not in wb.sheetnames:
        return

    src = wb[src_sheet_name]
    rows = list(src.iter_rows(values_only=True))
    if len(rows) < 3:
        return
    headers = [str(h) if h else "" for h in rows[1]]
    col = {h: i for i, h in enumerate(headers) if h}
    required = ("Territory", "Owner", "Initial Stage", "Initial ARR", "Bucket")
    if any(c not in col for c in required):
        return

    # Stage >= 4 counts as "commit-like" at start. Matches Parameters stage
    # numbering but inline since Parameters is just a constant here.
    def _stage_rank(s: str) -> int:
        s = str(s or "")
        if not s or len(s) < 1 or s[0] == "0":
            return 0
        try:
            return int(s[0])
        except (ValueError, IndexError):
            return 0

    by_owner = {}
    for r in rows[2:]:
        if not r or r[col["Territory"]] is None:
            continue
        owner = str(r[col["Owner"]] or "")
        if not owner:
            continue
        territory = str(r[col["Territory"]] or "")
        init_stage = str(r[col["Initial Stage"]] or "")
        initial_arr = float(r[col["Initial ARR"]] or 0)
        bucket = str(r[col["Bucket"]] or "")
        rank = _stage_rank(init_stage)

        agg = by_owner.setdefault(
            owner,
            {
                "territory": territory,
                "book_deals": 0,
                "book_arr": 0.0,
                "won_deals": 0,
                "won_arr": 0.0,
                "lost_deals": 0,
                "lost_arr": 0.0,
                "still_open_deals": 0,
                "still_open_arr": 0.0,
            },
        )
        if rank >= 4:
            agg["book_deals"] += 1
            agg["book_arr"] += initial_arr
            if bucket == "Won":
                agg["won_deals"] += 1
                agg["won_arr"] += initial_arr
            elif bucket == "Lost":
                agg["lost_deals"] += 1
                agg["lost_arr"] += initial_arr
            else:
                agg["still_open_deals"] += 1
                agg["still_open_arr"] += initial_arr

    ws = wb.create_sheet("Commit Accuracy")
    ws["A1"] = f"{contract.retrospective_title} Commit Accuracy by Owner"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Owner-level commit integrity. Commit book = deals in Stage 4+ "
        f"(Shortlisted or later) at the first snapshot of {contract.retrospective_label}. "
        "Accuracy = won "
        "ARR / commit book ARR. Low scores signal owners who committed but "
        "didn't deliver. Stage-based proxy until ForecastCategoryName history "
        "is wired into the extract."
    )
    ws["A2"].font = CAPTION_FONT

    header_row = 4
    hdrs = [
        "Owner",
        "Territory",
        f"Commit Book (Stage 4+, {contract.retrospective_label} Start)",
        "Commit Book ARR",
        "Won ARR",
        "Lost ARR",
        "Still Open ARR",
        "Accuracy (%)",
    ]
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=header_row, column=i, value=h)
    _apply_header(ws, header_row, hdrs)

    # Filter to owners with non-zero commit book so the tab is useful; sort
    # by ARR so the biggest books appear first.
    ranked = sorted(
        [(o, a) for o, a in by_owner.items() if a["book_arr"] > 0],
        key=lambda x: -x[1]["book_arr"],
    )

    first_data_row = header_row + 1
    for i, (owner, a) in enumerate(ranked):
        row = first_data_row + i
        accuracy = a["won_arr"] / a["book_arr"] if a["book_arr"] else 0
        values = [
            owner,
            a["territory"],
            a["book_deals"],
            a["book_arr"],
            a["won_arr"],
            a["lost_arr"],
            a["still_open_arr"],
            accuracy,
        ]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (4, 5, 6, 7):
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci == 8:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            elif ci == 3:
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
    last_data_row = first_data_row + len(ranked) - 1

    # Color scale on Accuracy column (red at 0, green at 1).
    if ranked:
        ws.conditional_formatting.add(
            f"H{first_data_row}:H{last_data_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8D7DA",
                mid_type="num",
                mid_value=0.5,
                mid_color="FFF3CD",
                end_type="num",
                end_value=1,
                end_color="D4EDDA",
            ),
        )

    widths = [24, 18, 14, 16, 14, 14, 16, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_forecast_variance(wb, workbooks_dir):
    """Decompose prior-quarter pipeline change by bucket via SUMIFS formulas.

    Every numeric cell is a SUMIFS against the prior-quarter trend tab
    (which carries per-row Initial ARR, Final ARR, Bucket helper columns).
    Click any cell in this tab to see the exact formula and drill into
    the source rows. Totals reconcile to (Final - Initial) because buckets
    are mutually exclusive by construction.
    """
    from openpyxl.chart import BarChart, Reference

    contract = _historical_trending_contract()
    ws = wb.create_sheet("Forecast Variance")
    ws["A1"] = f"Forecast Variance, {contract.retrospective_title} Decomposition"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"Every number below is a SUMIFS against the "
        f"'{contract.retrospective_consolidated_sheet}' "
        "tab. Click a cell to see the formula. The Bucket helper column on "
        "that tab assigns each deal to exactly one of {AlreadyClosed, Won, "
        "Lost, Added, RevisedUp, RevisedDown, Unchanged}, so the buckets are "
        "mutually exclusive and reconcile to (Final - Initial). Rules for "
        "bucket assignment are on Methodology."
    )
    ws["A2"].font = CAPTION_FONT

    # Find the helper column letters on the prior-quarter trend sheet.
    src_sheet_name = contract.retrospective_consolidated_sheet
    if src_sheet_name not in wb.sheetnames:
        ws.cell(row=4, column=1, value=f"{src_sheet_name} missing").font = BODY_BOLD
        return
    src = wb[src_sheet_name]
    # Header is row 2 on that tab
    headers = [src.cell(row=2, column=c).value for c in range(1, src.max_column + 1)]
    col_map = {}
    for ci, h in enumerate(headers, 1):
        if h:
            col_map[str(h)] = get_column_letter(ci)
    for needed in ("Territory", "Initial ARR", "Final ARR", "Bucket"):
        if needed not in col_map:
            ws.cell(
                row=4,
                column=1,
                value=f"Missing helper column '{needed}' on {src_sheet_name}.",
            )
            return
    terr_col = col_map["Territory"]
    init_col = col_map["Initial ARR"]
    final_col = col_map["Final ARR"]
    bucket_col = col_map["Bucket"]
    # Last source data row
    last_src_row = src.max_row

    # Reference range strings
    RT = f"'{src_sheet_name}'!${terr_col}$3:${terr_col}${last_src_row}"
    RI = f"'{src_sheet_name}'!${init_col}$3:${init_col}${last_src_row}"
    RF = f"'{src_sheet_name}'!${final_col}$3:${final_col}${last_src_row}"
    RB = f"'{src_sheet_name}'!${bucket_col}$3:${bucket_col}${last_src_row}"

    headers = [
        "Director",
        "Territory",
        "Initial ARR",
        "Final ARR",
        "Net Delta",
        "Won (-)",
        "Lost (-)",
        "Added (+)",
        "Revised Up (+)",
        "Revised Down (-)",
        "Check (should be 0)",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 30

    row += 1
    first_data_row = row

    # DIRECTORS entries are (director_name, territory, fname, oid, tid)
    # where territory in the tuple matches the Territory column value on
    # the prior-quarter consolidated trend tab.
    for director_name, territory, _fname, *_ in DIRECTORS:
        terr_literal = territory
        ws.cell(row=row, column=1, value=director_name).alignment = LEFT
        ws.cell(row=row, column=2, value=terr_literal).alignment = LEFT
        # Initial ARR: SUMIFS(InitialARR, Territory, T, Bucket, "<>AlreadyClosed")
        # Final ARR: SUMIFS(FinalARR, Territory, T, Bucket, "<>Won", Bucket, "<>Lost", Bucket, "<>AlreadyClosed")
        f_init = f'=SUMIFS({RI},{RT},B{row},{RB},"<>AlreadyClosed")'
        f_final = (
            f'=SUMIFS({RF},{RT},B{row},{RB},"<>Won",{RB},"<>Lost",'
            f'{RB},"<>AlreadyClosed")'
        )
        f_net = f"=D{row}-C{row}"
        f_won = f'=SUMIFS({RI},{RT},B{row},{RB},"Won")'
        f_lost = f'=SUMIFS({RI},{RT},B{row},{RB},"Lost")'
        f_added = f'=SUMIFS({RF},{RT},B{row},{RB},"Added")'
        f_up = (
            f'=SUMIFS({RF},{RT},B{row},{RB},"RevisedUp")'
            f'-SUMIFS({RI},{RT},B{row},{RB},"RevisedUp")'
        )
        f_down = (
            f'=SUMIFS({RI},{RT},B{row},{RB},"RevisedDown")'
            f'-SUMIFS({RF},{RT},B{row},{RB},"RevisedDown")'
        )
        f_check = f"=E{row}-(-F{row}-G{row}+H{row}+I{row}-J{row})"

        for ci, formula in enumerate(
            [f_init, f_final, f_net, f_won, f_lost, f_added, f_up, f_down, f_check],
            3,
        ):
            c = ws.cell(row=row, column=ci, value=formula)
            c.number_format = "#,##0"
            c.alignment = RIGHT
        for ci in range(1, 12):
            ws.cell(row=row, column=ci).border = BORDER
            ws.cell(row=row, column=ci).font = BODY_FONT
        row += 1
    last_data_row = row - 1

    # Totals row, SUM of the per-director formulas above
    ws.cell(row=row, column=1, value="TOTAL").font = BODY_BOLD
    ws.cell(row=row, column=2, value="All").font = BODY_BOLD
    for ci in range(3, 12):
        letter = get_column_letter(ci)
        c = ws.cell(
            row=row,
            column=ci,
            value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})",
        )
        c.number_format = "#,##0"
        c.alignment = RIGHT
        c.font = BODY_BOLD
    for ci in range(1, 12):
        ws.cell(row=row, column=ci).border = BORDER
    totals_row = row

    # Color scale on Net Delta (col E): red when negative, green when positive.
    # Applied on the per-director rows only (exclude TOTAL since its scale is
    # dominated by aggregate magnitude).
    if last_data_row >= first_data_row:
        from openpyxl.formatting.rule import ColorScaleRule

        ws.conditional_formatting.add(
            f"E{first_data_row}:E{last_data_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8D7DA",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="D4EDDA",
            ),
        )

    # Waterfall chart. Pulls from the TOTAL row so it stays in sync with
    # any future edits to the formula bar.
    chart_start = row + 3
    ws.cell(
        row=chart_start, column=1, value="Net Variance Breakdown, all directors"
    ).font = BODY_BOLD
    chart_rows = [
        ("Initial", f"=C{totals_row}"),
        ("Won (-)", f"=-F{totals_row}"),
        ("Lost (-)", f"=-G{totals_row}"),
        ("Added (+)", f"=H{totals_row}"),
        ("Revised Up (+)", f"=I{totals_row}"),
        ("Revised Down (-)", f"=-J{totals_row}"),
        ("Final", f"=D{totals_row}"),
    ]
    for i, (label, formula) in enumerate(chart_rows, 1):
        ws.cell(row=chart_start + i, column=1, value=label)
        c = ws.cell(row=chart_start + i, column=2, value=formula)
        c.number_format = "#,##0"

    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = f"{contract.retrospective_label} Pipeline Variance, Bucket Breakdown"
    chart.y_axis.title = "EUR"
    chart.x_axis.title = "Bucket"
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=chart_start + 1,
        max_row=chart_start + len(chart_rows),
    )
    cats_ref = Reference(
        ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(chart_rows)
    )
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, "D" + str(chart_start))

    # Footer: how to audit
    note_row = last_data_row + 2
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Audit: click any cell to see the SUMIFS formula. Filter the "
            f"{src_sheet_name} tab by 'Bucket' to see the raw deals that "
            "contribute to a bucket."
        ),
    ).font = CAPTION_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)

    widths = [20, 18, 14, 14, 14, 14, 14, 14, 14, 14, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"C{first_data_row}"


def build_executive_insights(wb, overdue_rows, kyc_rows, run_date):
    """Analyst briefing, table-structured. Every metric is a live formula.

    Four columns: # / Finding / Metric / Source. Metric cells are formulas
    against the upstream tabs so clicking shows the derivation. Source
    cells are HYPERLINK formulas that jump to the evidence. No hardcoded
    numbers in the table body.
    """

    ws = wb.create_sheet("Executive Insights", 0)
    ws["A1"] = "Executive Insights"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    ws["A2"] = (
        f"Review dated {run_date}. Metrics are live formulas against the "
        "upstream tabs (click a cell to see the formula bar). Source column "
        "links jump to the evidence tab. All thresholds live on Parameters; "
        "definitions on Methodology."
    )
    ws["A2"].font = CAPTION_FONT
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    # Compute aggregates from data_store
    total_open_count = sum(d["open_count"] for d in data_store.values())
    total_open_unwtd = sum(d["open_unwtd"] for d in data_store.values())
    total_open_wtd = sum(d["open_wtd"] for d in data_store.values())
    total_coverage = total_open_wtd / total_open_unwtd if total_open_unwtd else 0
    total_q1_won = sum(d["q1_won_arr"] for d in data_store.values())
    total_q1_lost = sum(d["q1_lost_arr"] for d in data_store.values())
    total_q1_outcome = total_q1_won + total_q1_lost

    # Top-deal concentration (recomputes the same aggregate as ARR Concentration)
    all_deals = []
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            all_deals.append(
                (
                    director_name,
                    r.get("Account", ""),
                    r.get("Opportunity", ""),
                    float(r.get("ARR Unweighted (EUR)") or 0),
                )
            )
    all_deals.sort(key=lambda x: -x[3])
    top5_sum = sum(d[3] for d in all_deals[:5])
    top20_sum = sum(d[3] for d in all_deals[:20])

    # Pipeline velocity: read prior-quarter totals from the already-built tab.
    contract = _historical_trending_contract()
    q1_initial = q1_final = 0.0
    if "Pipeline Velocity" in wb.sheetnames:
        vws = wb["Pipeline Velocity"]
        vrows = list(vws.iter_rows(values_only=True))
        block_header_idx = None
        for i, r in enumerate(vrows):
            if (
                r
                and isinstance(r[0], str)
                and r[0].startswith(contract.retrospective_title)
                and "ARR by Snapshot" in r[0]
            ):
                block_header_idx = i
                break
        if block_header_idx is not None:
            dates_row = vrows[block_header_idx + 1]
            dates = [c for c in dates_row[1:] if c]
            if dates:
                pass  # dates range available if needed
            director_rows = []
            i = block_header_idx + 2
            while i < len(vrows) and vrows[i] and vrows[i][0]:
                label = vrows[i][0]
                if "ARR by Snapshot" in str(label):
                    break
                series = [vrows[i][j + 1] or 0 for j in range(len(dates))]
                director_rows.append((label, series))
                q1_initial += series[0] if series else 0
                q1_final += series[-1] if series else 0
                i += 1
            if director_rows:
                pass  # movers available if needed

    # Top-stage backlog from Pipeline Pivot
    biggest_stage = ""
    biggest_stage_arr = 0.0
    if "Pipeline Pivot" in wb.sheetnames:
        pws = wb["Pipeline Pivot"]
        prows = list(pws.iter_rows(values_only=True))
        if len(prows) > 4:
            header = prows[3]
            stages = [h for h in header[1:-1] if h]
            totals_per_stage = {s: 0.0 for s in stages}
            for r in prows[4:]:
                if not r or r[0] is None or r[0] == "TOTAL":
                    continue
                for si, stage in enumerate(stages, 1):
                    v = r[si] if si < len(r) else None
                    if isinstance(v, (int, float)):
                        totals_per_stage[stage] += float(v)
            if totals_per_stage:
                biggest_stage, biggest_stage_arr = max(
                    totals_per_stage.items(), key=lambda x: x[1]
                )

    # Top slip-risk owners
    slip_top = []
    if "Slip Risk by Owner" in wb.sheetnames:
        sws = wb["Slip Risk by Owner"]
        srows = list(sws.iter_rows(values_only=True))
        for r in srows[4:9]:
            if not r or r[0] is None:
                continue
            slip_top.append(
                {
                    "owner": r[0],
                    "director": r[1],
                    "deals": r[2] or 0,
                    "arr": r[3] or 0,
                    "pushes": r[4] or 0,
                }
            )

    # Deal Risk top
    risk_top = []
    if "Deal Risk Scoring" in wb.sheetnames:
        rws = wb["Deal Risk Scoring"]
        rrows = list(rws.iter_rows(values_only=True))
        for r in rrows[4:9]:
            if not r or r[0] is None or not isinstance(r[0], int):
                continue
            risk_top.append(
                {
                    "score": r[1],
                    "director": r[2],
                    "account": r[3],
                    "opportunity": r[4],
                    "arr": r[10] or 0,
                    "reasons": r[12],
                }
            )

    # Best and worst territory by Q1 win rate
    scored_territories = []
    for name, d in data_store.items():
        q1w = d["q1_won_arr"]
        q1l = d["q1_lost_arr"]
        tot = q1w + q1l
        wr = q1w / tot if tot else None
        scored_territories.append({"name": name, "wr": wr, "won": q1w, "lost": q1l})
    with_wr = [t for t in scored_territories if t["wr"] is not None]
    # (dead expressions for best/worst territory, concentration, overdue/kyc
    # totals were removed — values are available from scored_territories,
    # data_store, overdue_rows, kyc_rows if narrative needs them later)

    # Build narrative
    def _m(n):
        if n is None:
            return "-"
        if abs(n) >= 1_000_000:
            return f"{n / 1_000_000:.1f}M EUR"
        if abs(n) >= 1_000:
            return f"{n / 1_000:.0f}K EUR"
        return f"{n:,.0f} EUR"

    row = 4
    # Headline block — unchanged prose, but the underlying numbers here are
    # aggregates, not claims; the live formula counterparts live below.
    ws.cell(
        row=row,
        column=1,
        value=f"Headline: {total_open_count} open Land deals, "
        f"{_m(total_open_unwtd)} unweighted ({_m(total_open_wtd)} weighted, "
        f"{total_coverage * 100:.0f}% coverage).",
    ).font = BODY_BOLD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Table header
    headers = ["#", "Finding", "Metric (live formula)", "Detail", "Source"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28
    row += 1

    # Locate source cells dynamically for formulas
    variance_total_row = None
    if "Forecast Variance" in wb.sheetnames:
        vws = wb["Forecast Variance"]
        for r in range(4, vws.max_row + 1):
            if str(vws.cell(row=r, column=1).value or "") == "TOTAL":
                variance_total_row = r
                break

    concentration_rows = []  # (rank, director, arr_cell_ref)
    if "ARR Concentration" in wb.sheetnames:
        aws = wb["ARR Concentration"]
        for r in range(5, min(aws.max_row, 25) + 1):
            v = aws.cell(row=r, column=1).value
            if isinstance(v, int):
                concentration_rows.append((v, r))

    slip_top_row = None
    if "Slip Risk by Owner" in wb.sheetnames:
        slip_top_row = 5  # first data row below the sheet's header

    scorecard_first_data = None
    if "Territory Scorecard" in wb.sheetnames:
        scorecard_first_data = 5

    def _hyperlink(tab_name, cell="A1", label=None):
        if label is None:
            label = f"Open '{tab_name}'"
        return f'=HYPERLINK("#\'{tab_name}\'!{cell}", "{label}")'

    def _write_row(num, finding, formula_or_value, detail, source_formula):
        nonlocal row
        ws.cell(row=row, column=1, value=num).alignment = CENTER
        c = ws.cell(row=row, column=2, value=finding)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.font = BODY_FONT
        mc = ws.cell(row=row, column=3, value=formula_or_value)
        mc.alignment = RIGHT
        mc.font = BODY_BOLD
        dc = ws.cell(row=row, column=4, value=detail)
        dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        dc.font = BODY_FONT
        sc = ws.cell(row=row, column=5, value=source_formula)
        sc.font = Font(name="Calibri", size=10, color="083EA7", underline="single")
        sc.alignment = CENTER
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = BORDER
        ws.row_dimensions[row].height = 40
        row += 1

    num = 1

    # 1. Pipeline velocity, as the Forecast Variance net delta (live formula)
    if variance_total_row:
        _write_row(
            num,
            f"{_historical_trending_contract().retrospective_label} pipeline net change, Final minus Initial (bucket-decomposed, "
            "not just a snapshot delta).",
            f"='Forecast Variance'!E{variance_total_row}",
            "Positive = pipeline grew; negative = shrank. Sum of Won/Lost/"
            "Added/Revised reconciles exactly to this number.",
            _hyperlink(
                "Forecast Variance",
                f"A{variance_total_row}",
                "Forecast Variance TOTAL",
            ),
        )
        num += 1

    # 2. Top 5 concentration %. Divides sum of top-5 ARR by the director
    # total book from Territory Scorecard TOTAL (or LandPipelineDetail table).
    if concentration_rows and len(concentration_rows) >= 5:
        # top 5 rows on ARR Concentration are ranks 1-5 at rows 5-9 typically
        top5_cells = "'ARR Concentration'!G5:G9"
        formula = f"=SUM({top5_cells})/SUM(LandPipelineDetail[ARR Unwtd])"
        _write_row(
            num,
            "Share of global open Land pipeline held in the top 5 individual "
            "deals. Single-deal dependency.",
            formula,
            "Denominator is the sum of every deal on Land Pipeline Detail. "
            "Numerator sums the top 5 on ARR Concentration (see threshold "
            "Thresh_ConcentrationTop3Pct on Parameters).",
            _hyperlink("ARR Concentration", "A5", "ARR Concentration"),
        )
        c = ws.cell(row=row - 1, column=3)
        c.number_format = "0.0%"
        num += 1

    # 3. Slip risk, top owner total pushes (pushes column is E on Slip Risk tab)
    if slip_top_row:
        _write_row(
            num,
            "Highest-push owner's total open-deal push count. Push >= "
            "Thresh_SlipOwnerPushes is a flag.",
            f"='Slip Risk by Owner'!E{slip_top_row}",
            "E column on Slip Risk by Owner is Total Pushes, sorted "
            "descending. Pair with the ARR column (D) on the same row.",
            _hyperlink(
                "Slip Risk by Owner",
                f"A{slip_top_row}",
                "Slip Risk by Owner",
            ),
        )
        num += 1

    # 4. Stage backlog concentration (use MAX across the Pipeline Pivot
    # TOTAL row). Pipeline Pivot's TOTAL row sits at the bottom.
    if "Pipeline Pivot" in wb.sheetnames:
        pws = wb["Pipeline Pivot"]
        total_row_p = None
        for r in range(5, pws.max_row + 1):
            if str(pws.cell(row=r, column=1).value or "") == "TOTAL":
                total_row_p = r
                break
        if total_row_p:
            max_col = pws.max_column - 1  # exclude Total column
            last_stage_letter = get_column_letter(max_col)
            _write_row(
                num,
                f"Largest open-ARR stage in {SCOPE_LABEL} (where pipeline is bunched).",
                f"=MAX('Pipeline Pivot'!B{total_row_p}:{last_stage_letter}{total_row_p})",
                "TOTAL row on Pipeline Pivot sums each stage across "
                "directors. MAX picks the bottleneck stage.",
                _hyperlink("Pipeline Pivot", f"A{total_row_p}", "Pipeline Pivot"),
            )
            c = ws.cell(row=row - 1, column=3)
            c.number_format = "#,##0"
            num += 1

    pq_slips_sheet = f"{RUNTIME_PERIOD['prior_quarter_label']} Slips, Still Open"
    if pq_slips_sheet in wb.sheetnames:
        _write_row(
            num,
            f"Count of {RUNTIME_PERIOD['prior_quarter_label']} deals that slipped out of {RUNTIME_PERIOD['prior_quarter_label']} and are still open.",
            f"=COUNTA('{pq_slips_sheet}'!A:A)-4",
            "Subtracts the 4 header rows on that tab. Each row is an "
            f"opportunity that had a {RUNTIME_PERIOD['prior_quarter_title']} close date at some point and "
            "has not yet been Won or Lost.",
            _hyperlink(pq_slips_sheet, "A1", pq_slips_sheet),
        )
        num += 1

    # 6. Deal Risk count >= triage threshold (uses Parameters named cell)
    if "Deal Risk Scoring" in wb.sheetnames:
        _write_row(
            num,
            "Open deals scoring at or above the triage floor (Thresh_"
            "RiskScoreTriage on Parameters).",
            "=COUNTIF('Deal Risk Scoring'!B5:B34,\">=\"&Thresh_RiskScoreTriage)",
            "B column on Deal Risk Scoring is Risk Score. Threshold is the "
            "named cell Thresh_RiskScoreTriage on Parameters.",
            _hyperlink("Deal Risk Scoring", "A5", "Deal Risk Scoring"),
        )
        num += 1

    # 7. Territory coverage range. Shows how unevenly weighted coverage is
    # across directors; Territory Scorecard column E is Coverage %.
    if scorecard_first_data:
        _write_row(
            num,
            "Weighted coverage spread across territories (max minus min). "
            "Wide = inconsistent forecasting discipline.",
            f"=MAX('Territory Scorecard'!E{scorecard_first_data}:E13)"
            f"-MIN('Territory Scorecard'!E{scorecard_first_data}:E13)",
            "Coverage = Forecast ARR / Open ARR per director. Wide range "
            "points to uneven stage weightings or probability quality.",
            _hyperlink(
                "Territory Scorecard",
                f"A{scorecard_first_data}",
                "Territory Scorecard",
            ),
        )
        c = ws.cell(row=row - 1, column=3)
        c.number_format = "0.0%"
        num += 1

    # 8. Data hygiene
    _write_row(
        num,
        "Open opportunities with past close dates (overdue).",
        len(overdue_rows),
        "Python-computed count of overdue open opps (avoids COUNTA "
        "double-count from summary block rows).",
        _hyperlink("Overdue Open Opps", "A1", "Overdue Open Opps"),
    )
    num += 1

    _write_row(
        num,
        "Accounts in the active pipeline with no KYC approval on file.",
        len(kyc_rows),
        "Python-computed count of KYC-missing accounts (avoids COUNTA "
        "double-count from summary block rows).",
        _hyperlink("KYC Missing", "A1", "KYC Missing"),
    )
    num += 1

    # Trailer: narrative summary for an exec who wants one paragraph
    row += 1
    ws.cell(row=row, column=1, value="Where to go next").font = BODY_BOLD
    row += 1
    next_steps = [
        (
            "Drill Deal Risk Scoring, triage every row with score at or above "
            "Thresh_RiskScoreTriage.",
            "Deal Risk Scoring",
            "A5",
        ),
        (
            "Review Top 5 on ARR Concentration with Sales Ops before "
            "quarter-end close.",
            "ARR Concentration",
            "A5",
        ),
        (
            "For each owner on Slip Risk by Owner confirm close-date "
            "integrity and push reasons.",
            "Slip Risk by Owner",
            "A5",
        ),
        (
            "Use Forecast Variance to explain the week-over-week move to the CRO.",
            "Forecast Variance",
            "A4",
        ),
        (
            "Cross-check Territory Scorecard against commitments; flag any "
            "red column as action item.",
            "Territory Scorecard",
            "A5",
        ),
    ]
    for text, tab, cell_ref in next_steps:
        ws.cell(row=row, column=1, value="-").alignment = CENTER
        ws.cell(row=row, column=2, value=text).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        sc = ws.cell(
            row=row,
            column=5,
            value=_hyperlink(tab, cell_ref, f"Open '{tab}'"),
        )
        sc.font = Font(name="Calibri", size=10, color="083EA7", underline="single")
        sc.alignment = CENTER
        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 24


# ──────────────────────── Advanced analytics tabs ───────────────────────
# New tabs added during the Excel analytics audit. Each answers a specific
# sales-director question that the older tabs didn't cover. Ordered so the
# most-asked questions come first.


def build_owner_scorecard(wb):
    """Per-owner view across all metrics a director needs for 1:1 prep.

    Pulls from data_store so every territory is covered. Combines deals,
    open ARR, win rate, avg deal size, avg cycle time, push intensity, and
    last activity date into one line per owner. Sort: open ARR descending
    so the biggest books surface first.
    """
    from openpyxl.formatting.rule import ColorScaleRule

    ws = wb.create_sheet("Owner Scorecard")
    ws["A1"] = "Owner Scorecard"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Per-owner view across the full monthly review KPI set. Source: Land "
        "Pipeline Detail + Land WonLost Detail. Use for 1:1 coaching prep — one "
        "row per owner shows the full story."
    )
    ws["A2"].font = CAPTION_FONT

    from collections import defaultdict as _dd

    owners = _dd(
        lambda: {
            "director": "",
            "open_deals": 0,
            "open_arr": 0.0,
            "q1_won": 0,
            "q1_won_arr": 0.0,
            "q1_lost": 0,
            "q1_lost_arr": 0.0,
            "total_push": 0,
            "max_push": 0,
            "cycle_days_sum": 0.0,
            "cycle_days_n": 0,
            "last_activity": "",
        }
    )
    for director_name, d in data_store.items():
        for r in d.get("all_land_open_rows", []):
            owner = str(r.get("Owner") or "(unassigned)")
            o = owners[owner]
            o["director"] = director_name
            o["open_deals"] += 1
            o["open_arr"] += float(r.get("ARR Unweighted (EUR)") or 0)
            o["total_push"] += int(r.get("Push Count") or 0)
            o["max_push"] = max(o["max_push"], int(r.get("Push Count") or 0))
            la = str(r.get("Last Activity") or "")[:10]
            if la and la > (o["last_activity"] or ""):
                o["last_activity"] = la
        for r in d.get("all_land_won_lost_rows", []):
            owner = str(r.get("Owner") or "(unassigned)")
            stage = str(r.get("Stage") or "")
            arr = float(r.get("ARR Unweighted (EUR)") or 0)
            created = str(r.get("Created") or "")[:10]
            closed = str(r.get("Close Date") or "")[:10]
            o = owners[owner]
            if not o["director"]:
                o["director"] = director_name
            if "Won" in stage:
                o["q1_won"] += 1
                o["q1_won_arr"] += arr
            elif "Lost" in stage:
                o["q1_lost"] += 1
                o["q1_lost_arr"] += arr
            # Cycle time: only counted for resolved (won/lost) deals with
            # both dates populated. Anything closed in <=0 or >1500 days is
            # treated as garbage (rounded-date imports, historical imports).
            if created and closed and created >= "2020-01-01" and closed >= created:
                try:
                    from datetime import date as _date

                    c1 = _date.fromisoformat(created)
                    c2 = _date.fromisoformat(closed)
                    days = (c2 - c1).days
                    if 0 < days <= 1500:
                        o["cycle_days_sum"] += days
                        o["cycle_days_n"] += 1
                except ValueError:
                    pass

    ranked = sorted(owners.items(), key=lambda x: -x[1]["open_arr"])

    headers = [
        "Owner",
        "Director",
        "Open Deals",
        "Open ARR (EUR)",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR",
        f"{RUNTIME_PERIOD['prior_quarter_label']} Lost",
        "Win Rate (ARR)",
        "Avg Deal Size",
        "Avg Cycle Days",
        "Total Pushes",
        "Max Push on 1 Deal",
        "Last Activity",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28
    header_row = row

    row += 1
    first_data_row = row
    for owner, s in ranked:
        total_closed_arr = s["q1_won_arr"] + s["q1_lost_arr"]
        win_rate = s["q1_won_arr"] / total_closed_arr if total_closed_arr else 0
        total_closed_n = s["q1_won"] + s["q1_lost"]
        avg_size = (
            (s["q1_won_arr"] + s["q1_lost_arr"]) / total_closed_n
            if total_closed_n
            else (s["open_arr"] / s["open_deals"] if s["open_deals"] else 0)
        )
        avg_cycle = s["cycle_days_sum"] / s["cycle_days_n"] if s["cycle_days_n"] else 0
        vals = [
            owner,
            s["director"],
            s["open_deals"],
            s["open_arr"],
            s["q1_won"],
            s["q1_won_arr"],
            s["q1_lost"],
            win_rate,
            avg_size,
            avg_cycle,
            s["total_push"],
            s["max_push"],
            s["last_activity"] or "-",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (4, 6, 9):
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci == 8:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            elif ci in (3, 5, 7, 10, 11, 12):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1
    last_data_row = row - 1

    if last_data_row >= first_data_row:
        # Color-scale win rate (column H) green-good, red-bad.
        ws.conditional_formatting.add(
            f"H{first_data_row}:H{last_data_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8D7DA",
                mid_type="num",
                mid_value=0.5,
                mid_color="FFF3CD",
                end_type="num",
                end_value=1,
                end_color="D4EDDA",
            ),
        )

    widths = [22, 18, 9, 14, 8, 14, 8, 12, 14, 12, 12, 14, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_competitive_win_loss(wb):
    """Competitor × count × lost ARR aggregate from closed Land deals."""
    from openpyxl.formatting.rule import DataBarRule

    ws = wb.create_sheet("Competitive Win Loss")
    ws["A1"] = "Competitive Win / Loss"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Closed Land deals aggregated by Lost_to_Competitor__c. Surfaces "
        "repeat losses to a specific competitor — the '$10M we lost to X' "
        "signal that's immediately actionable. Source: Land WonLost Detail."
    )
    ws["A2"].font = CAPTION_FONT

    from collections import defaultdict as _dd

    by_comp = _dd(lambda: {"lost_n": 0, "lost_arr": 0.0, "won_n": 0, "won_arr": 0.0})
    total_lost_arr = 0.0
    for d in data_store.values():
        for r in d.get("all_land_won_lost_rows", []):
            comp = str(r.get("Lost To Competitor") or "").strip()
            stage = str(r.get("Stage") or "")
            arr = float(r.get("ARR Unweighted (EUR)") or 0)
            if "Won" in stage:
                if comp:
                    by_comp[comp]["won_n"] += 1
                    by_comp[comp]["won_arr"] += arr
            elif "Lost" in stage:
                if comp:
                    by_comp[comp]["lost_n"] += 1
                    by_comp[comp]["lost_arr"] += arr
                    total_lost_arr += arr

    ranked = sorted(by_comp.items(), key=lambda x: -x[1]["lost_arr"])
    headers = [
        "Competitor",
        "Losses",
        "Lost ARR (EUR)",
        "% of Lost ARR",
        "Wins against",
        "Win ARR against (EUR)",
        "Net (Wins - Losses ARR)",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    for comp, s in ranked:
        pct = s["lost_arr"] / total_lost_arr if total_lost_arr else 0
        net = s["won_arr"] - s["lost_arr"]
        vals = [comp, s["lost_n"], s["lost_arr"], pct, s["won_n"], s["won_arr"], net]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (3, 6, 7):
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci == 4:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            elif ci in (2, 5):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1
    last_data_row = row - 1

    if last_data_row >= first_data_row:
        ws.conditional_formatting.add(
            f"C{first_data_row}:C{last_data_row}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="D9534F",
                showValue=True,
            ),
        )

    widths = [34, 10, 18, 14, 14, 20, 20]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_stage_conversion_per_territory(wb):
    """Per-territory stage entry → advance → win funnel.

    REMOVED: The original implementation fabricated conversion rates from
    current stage position (a deal at Stage 5 was counted as having entered
    stages 1-4, which is wrong -- deals can skip stages). Real implementation
    requires OpportunityFieldHistory stage_history_events in data_store, which
    gather_director_data does not currently extract.
    """
    print("  [SKIP] Stage Conversion: no stage history data available")


def build_sales_velocity(wb):
    """Classic sales velocity = (wins × avg deal size × win rate) / cycle days.

    Per-director + TOTAL. The velocity number is an approximate daily ARR
    production rate: higher = team is converting pipeline faster.
    """
    from openpyxl.formatting.rule import DataBarRule

    ws = wb.create_sheet("Sales Velocity")
    ws["A1"] = "Sales Velocity"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Velocity (EUR/day) = (# wins × avg deal size × win rate) / avg cycle "
        "days. Industry-standard productivity KPI. Higher is better. Compare "
        "to last month to see whether the team is speeding up or slowing down."
    )
    ws["A2"].font = CAPTION_FONT

    headers = [
        "Director",
        "Territory",
        f"Wins ({RUNTIME_PERIOD['prior_quarter_label']})",
        "Avg Deal Size",
        "Win Rate (ARR)",
        "Avg Cycle Days",
        "Velocity (EUR/day)",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    totals = {
        "wins": 0,
        "won_arr": 0.0,
        "lost_arr": 0.0,
        "cycle_sum": 0.0,
        "cycle_n": 0,
    }
    for director_name, d in data_store.items():
        territory = ""
        for _n, _terr, *_rest in DIRECTORS:
            if _n == director_name:
                territory = _terr
                break
        wins = d.get("q1_won_count", 0)
        won_arr = d.get("q1_won_arr", 0.0)
        lost_arr = d.get("q1_lost_arr", 0.0)
        total_arr = won_arr + lost_arr
        win_rate = won_arr / total_arr if total_arr else 0
        avg_size = won_arr / wins if wins else 0
        # Cycle time: approximate from won list if dates present.
        cycle_sum = 0.0
        cycle_n = 0
        for r in d.get("all_land_won_lost_rows", []):
            created = str(r.get("Created") or "")[:10]
            closed = str(r.get("Close Date") or "")[:10]
            if created and closed and created >= "2020-01-01" and closed >= created:
                try:
                    from datetime import date as _date

                    days = (
                        _date.fromisoformat(closed) - _date.fromisoformat(created)
                    ).days
                    if 0 < days <= 1500:
                        cycle_sum += days
                        cycle_n += 1
                except ValueError:
                    pass
        avg_cycle = cycle_sum / cycle_n if cycle_n else 0
        velocity = (wins * avg_size * win_rate) / avg_cycle if avg_cycle else 0

        totals["wins"] += wins
        totals["won_arr"] += won_arr
        totals["lost_arr"] += lost_arr
        totals["cycle_sum"] += cycle_sum
        totals["cycle_n"] += cycle_n

        vals = [director_name, territory, wins, avg_size, win_rate, avg_cycle, velocity]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (4, 7):
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci == 5:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            elif ci in (3, 6):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1

    total_arr = totals["won_arr"] + totals["lost_arr"]
    total_win_rate = totals["won_arr"] / total_arr if total_arr else 0
    total_wins = totals["wins"]
    total_closed = total_wins + int(
        sum(d.get("q1_lost_count", 0) for d in data_store.values())
    )
    total_avg_size = total_arr / total_closed if total_closed else 0
    total_avg_cycle = (
        totals["cycle_sum"] / totals["cycle_n"] if totals["cycle_n"] else 0
    )
    total_velocity = (
        (total_wins * total_avg_size * total_win_rate) / total_avg_cycle
        if total_avg_cycle
        else 0
    )
    vals = [
        "TOTAL",
        "",
        total_wins,
        total_avg_size,
        total_win_rate,
        total_avg_cycle,
        total_velocity,
    ]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        if ci in (4, 7):
            c.number_format = "#,##0"
            c.alignment = RIGHT
        elif ci == 5:
            c.number_format = "0.0%"
            c.alignment = RIGHT
        elif ci in (3, 6):
            c.alignment = RIGHT
        else:
            c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_BOLD
    last_data_row = row

    if last_data_row >= first_data_row + 1:
        # Data bar on velocity (exclude TOTAL).
        ws.conditional_formatting.add(
            f"G{first_data_row}:G{last_data_row - 1}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="4472C4",
                showValue=True,
            ),
        )

    widths = [20, 18, 10, 14, 14, 14, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_deal_age_distribution(wb):
    """P25/P50/P75/P90 age-in-stage per territory × stage."""
    ws = wb.create_sheet("Deal Age Distribution")
    ws["A1"] = "Deal Age Distribution by Stage"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Percentiles of days-in-pipeline (Created → Today) for open deals, "
        "grouped by territory and stage. Long right-tail (P90 >> P50) means "
        "pipeline is stuck at this stage for this territory."
    )
    ws["A2"].font = CAPTION_FONT

    from collections import defaultdict as _dd
    from datetime import date as _date

    today = _date.today()
    ages = _dd(lambda: _dd(list))  # ages[territory][stage] = [days,...]
    for director_name, d in data_store.items():
        territory = ""
        for _n, _terr, *_rest in DIRECTORS:
            if _n == director_name:
                territory = _terr
                break
        for r in d.get("all_land_open_rows", []):
            stage = str(r.get("Stage") or "")
            created = str(r.get("Created") or "")[:10]
            if not created or created < "2020-01-01":
                continue
            try:
                days = (today - _date.fromisoformat(created)).days
                if 0 < days <= 3000:
                    ages[territory][stage].append(days)
            except ValueError:
                continue

    def _pct(values, p):
        if not values:
            return 0
        values = sorted(values)
        k = (len(values) - 1) * (p / 100.0)
        f = int(k)
        c = min(f + 1, len(values) - 1)
        return values[f] + (values[c] - values[f]) * (k - f)

    headers = [
        "Territory",
        "Stage",
        "Count",
        "P25",
        "P50 (median)",
        "P75",
        "P90",
        "Max",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    stages_positive = [
        "1 - Prospecting",
        "2 - Discovery",
        "3 - Engagement",
        "4 - Shortlisted",
        "5 - Preferred",
        "6 - Contracting",
    ]
    for _n, territory, *_rest in DIRECTORS:
        for stage in stages_positive:
            vals_days = ages[territory].get(stage, [])
            if not vals_days:
                continue
            vals = [
                territory,
                stage,
                len(vals_days),
                int(_pct(vals_days, 25)),
                int(_pct(vals_days, 50)),
                int(_pct(vals_days, 75)),
                int(_pct(vals_days, 90)),
                max(vals_days),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                if ci >= 3:
                    c.alignment = RIGHT
                else:
                    c.alignment = LEFT
                c.border = BORDER
                c.font = BODY_FONT
            row += 1

    widths = [18, 20, 8, 8, 12, 8, 8, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_account_penetration(wb):
    """Accounts with multiple open opps (cross-sell) or single opp (upsell)."""
    ws = wb.create_sheet("Account Penetration")
    ws["A1"] = "Account Penetration"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Accounts ranked by count of open Land opportunities. Accounts with "
        "multiple opps = cross-sell motion in flight. Accounts with one opp "
        "and high ARR = land-expand candidates. Source: Land Pipeline Detail."
    )
    ws["A2"].font = CAPTION_FONT

    from collections import defaultdict as _dd

    accounts = _dd(
        lambda: {
            "territory": "",
            "deals": 0,
            "arr": 0.0,
            "owners": set(),
            "stages": set(),
        }
    )
    for director_name, d in data_store.items():
        territory = ""
        for _n, _terr, *_rest in DIRECTORS:
            if _n == director_name:
                territory = _terr
                break
        for r in d.get("all_land_open_rows", []):
            acct = str(r.get("Account") or "")
            if not acct:
                continue
            a = accounts[acct]
            a["territory"] = territory
            a["deals"] += 1
            a["arr"] += float(r.get("ARR Unweighted (EUR)") or 0)
            a["owners"].add(str(r.get("Owner") or ""))
            a["stages"].add(str(r.get("Stage") or ""))

    ranked = sorted(accounts.items(), key=lambda x: (-x[1]["deals"], -x[1]["arr"]))
    headers = [
        "Account",
        "Territory",
        "Open Deals",
        "Open ARR (EUR)",
        "Owners",
        "Distinct Stages",
        "Signal",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    for acct, s in ranked[:100]:
        signal = ""
        if s["deals"] >= 3:
            signal = "Deep penetration — multi-deal account"
        elif s["deals"] == 2:
            signal = "Cross-sell in flight"
        elif s["deals"] == 1 and s["arr"] >= 1_000_000:
            signal = "Single large opp — land-expand candidate"
        elif s["deals"] == 1:
            signal = "Single opp"
        vals = [
            acct,
            s["territory"],
            s["deals"],
            s["arr"],
            ", ".join(sorted(s["owners"]))[:80],
            len(s["stages"]),
            signal,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci == 4:
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci in (3, 6):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1

    widths = [40, 18, 10, 16, 40, 14, 34]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_forecast_bias(wb):
    """Owner commit vs delivered history via the snapshot_history ledger."""
    ws = wb.create_sheet("Forecast Bias")
    ws["A1"] = "Forecast Bias, Owner Commit vs Delivered"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Commit credibility over time. Needs ≥2 run_date entries in "
        "obsidian/snapshot_history.json to produce a delta. First run of a "
        "period shows the current commit total only; subsequent runs show "
        "movement. Over/under bias surfaces consistent forecasters."
    )
    ws["A2"].font = CAPTION_FONT

    import json as _json
    from pathlib import Path as _Path

    ledger_path = _Path("obsidian/snapshot_history.json")
    if not ledger_path.exists():
        ws.cell(row=4, column=1, value="Ledger not yet present").font = BODY_BOLD
        return
    try:
        history = _json.loads(ledger_path.read_text())
    except (_json.JSONDecodeError, ValueError):
        ws.cell(row=4, column=1, value="Ledger unreadable").font = BODY_BOLD
        return

    snapshots = history.get("snapshots") or []
    if len(snapshots) < 2:
        ws.cell(
            row=4,
            column=1,
            value=f"Need ≥2 snapshots (currently {len(snapshots)}). "
            "Re-run generate_obsidian_notes.py on a prior date to seed history.",
        ).font = BODY_BOLD
        return

    prior = snapshots[-2]
    current = snapshots[-1]
    headers = [
        "Director",
        "Territory",
        f"Prior ({prior.get('run_date')})",
        f"Current ({current.get('run_date')})",
        "Delta (EUR)",
        "Delta %",
        "Bias Signal",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    prior_dirs = prior.get("directors", {})
    curr_dirs = current.get("directors", {})
    for director_name in sorted(set(prior_dirs) | set(curr_dirs)):
        p = prior_dirs.get(director_name, {})
        c_entry = curr_dirs.get(director_name, {})
        prior_arr = float(p.get("open_land_arr_unwtd") or 0)
        curr_arr = float(c_entry.get("open_land_arr_unwtd") or 0)
        territory = c_entry.get("territory") or p.get("territory") or ""
        delta = curr_arr - prior_arr
        pct = delta / prior_arr if prior_arr else 0
        if abs(pct) < 0.05:
            signal = "Steady"
        elif pct <= -0.15:
            signal = "Commit shrank — watch for misses"
        elif pct >= 0.15:
            signal = "Commit grew — watch for sandbagging"
        elif pct > 0:
            signal = "Modestly growing"
        else:
            signal = "Modestly shrinking"
        vals = [director_name, territory, prior_arr, curr_arr, delta, pct, signal]
        for ci, v in enumerate(vals, 1):
            cc = ws.cell(row=row, column=ci, value=v)
            if ci in (3, 4, 5):
                cc.number_format = "#,##0"
                cc.alignment = RIGHT
            elif ci == 6:
                cc.number_format = "0.0%"
                cc.alignment = RIGHT
            else:
                cc.alignment = LEFT
            cc.border = BORDER
            cc.font = BODY_FONT
        row += 1

    widths = [20, 18, 16, 16, 14, 10, 38]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_approvals_dashboard(wb):
    print("  [SKIP] Approvals Dashboard: removed (redundant with Approvals Overview)")
    return
    ws["A1"] = "Approvals Dashboard"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "One-page view of the commercial approval pipeline — approved / "
        "pending / missing in a single scan. Source tabs kept as drills "
        f"(Approvals {RUNTIME_PERIOD['analysis_year']}, Approval Candidates, Land Stage 3+ No Approval)."
    )
    ws["A2"].font = CAPTION_FONT

    headers = [
        "Director",
        "Territory",
        f"Approved {RUNTIME_PERIOD['analysis_year']}",
        f"Approved {RUNTIME_PERIOD['analysis_year']} ARR",
        "Conditionally Approved",
        "Conditional ARR",
        "Missing Stage 3+",
        "Missing ARR",
    ]
    row = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    totals = {"a_n": 0, "a_arr": 0.0, "c_n": 0, "c_arr": 0.0, "m_n": 0, "m_arr": 0.0}
    for director_name, d in data_store.items():
        territory = ""
        for _n, _terr, *_rest in DIRECTORS:
            if _n == director_name:
                territory = _terr
                break
        approved = d.get("approved_2026") or []
        pending = d.get("approval_candidates") or []
        missing = d.get("approval_missing") or []
        a_arr = sum(float(r.get("arr_unwtd", 0) or 0) for r in approved)
        c_arr = sum(float(r.get("arr_unwtd", 0) or 0) for r in pending)
        m_arr = sum(float(r.get("arr_unwtd", 0) or 0) for r in missing)
        totals["a_n"] += len(approved)
        totals["a_arr"] += a_arr
        totals["c_n"] += len(pending)
        totals["c_arr"] += c_arr
        totals["m_n"] += len(missing)
        totals["m_arr"] += m_arr
        vals = [
            director_name,
            territory,
            len(approved),
            a_arr,
            len(pending),
            c_arr,
            len(missing),
            m_arr,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci in (4, 6, 8):
                c.number_format = "#,##0"
                c.alignment = RIGHT
            elif ci in (3, 5, 7):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1
    total_vals = [
        "TOTAL",
        "",
        totals["a_n"],
        totals["a_arr"],
        totals["c_n"],
        totals["c_arr"],
        totals["m_n"],
        totals["m_arr"],
    ]
    for ci, v in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        if ci in (4, 6, 8):
            c.number_format = "#,##0"
            c.alignment = RIGHT
        elif ci in (3, 5, 7):
            c.alignment = RIGHT
        else:
            c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_BOLD

    widths = [20, 18, 13, 18, 18, 18, 15, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_days_in_stage_by_director(wb):
    """Per-director matrix of avg days-in-stage across the funnel.

    REMOVED: The original implementation computed (today - CreatedDate).days,
    which is pipeline age, NOT days in the current stage. This contradicted the
    correct global Days in Stage tab (sourced from the SF report). Correct
    implementation requires OpportunityHistory or a DaysInStage field, which
    we don't extract per-director.
    """
    print(
        "  [SKIP] Days in Stage by Director: removed (used pipeline age, not actual stage duration)"
    )


def build_win_rate_trend(wb):
    """Win rate over snapshot history (needs ≥2 snapshots in ledger)."""
    ws = wb.create_sheet("Win Rate Trend")
    ws["A1"] = "Win Rate Trend, by Director"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Win rate (ARR basis) per run_date, pulled from the snapshot history "
        "ledger. Directors trending up are speeding up. Needs ≥2 snapshots."
    )
    ws["A2"].font = CAPTION_FONT

    import json as _json
    from pathlib import Path as _Path

    ledger_path = _Path("obsidian/snapshot_history.json")
    if not ledger_path.exists():
        ws.cell(row=4, column=1, value="Ledger not yet present").font = BODY_BOLD
        return
    try:
        history = _json.loads(ledger_path.read_text())
    except (_json.JSONDecodeError, ValueError):
        ws.cell(row=4, column=1, value="Ledger unreadable").font = BODY_BOLD
        return
    snapshots = history.get("snapshots") or []
    if not snapshots:
        ws.cell(row=4, column=1, value="No snapshots in ledger").font = BODY_BOLD
        return

    all_directors = sorted({dn for s in snapshots for dn in (s.get("directors") or {})})
    run_dates = [s.get("run_date") for s in snapshots]

    row = 4
    headers = ["Director"] + run_dates
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    first_data_row = row
    for director_name in all_directors:
        cells = [director_name]
        for s in snapshots:
            d = (s.get("directors") or {}).get(director_name, {})
            won_arr = float(d.get("q1_won_arr") or 0)
            lost_arr = float(d.get("q1_lost_arr") or 0)
            total = won_arr + lost_arr
            rate = won_arr / total if total else 0
            cells.append(rate)
        for ci, v in enumerate(cells, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci > 1:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            c.border = BORDER
            c.font = BODY_FONT
        row += 1

    widths = [22] + [14] * len(run_dates)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{first_data_row}"


def build_charts_sheet(wb):
    """Native Excel bar and pie charts driven by the detail tabs."""
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Charts")
    ws["A1"] = "Charts"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    # Build small summary block (director -> open ARR) using SUMIFS for chart source
    ws["A3"] = "Open Land Pipeline ARR, by Director"
    ws["A3"].font = BODY_BOLD
    ws["A4"] = "Director"
    ws["B4"] = "Open ARR Unwtd"
    ws["C4"] = "Open ARR Wtd"
    _apply_header(ws, 4, ["Director", "Open ARR Unwtd", "Open ARR Wtd"])

    r = 5
    for director_name in data_store:
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(
            row=r,
            column=2,
            value=f"=SUMIFS(LandPipelineDetail[ARR Unwtd],LandPipelineDetail[Director],A{r})",
        )
        ws.cell(
            row=r,
            column=3,
            value=f"=SUMIFS(LandPipelineDetail[ARR Wtd],LandPipelineDetail[Director],A{r})",
        )
        for ci in (2, 3):
            ws.cell(row=r, column=ci).number_format = "#,##0"
            ws.cell(row=r, column=ci).alignment = RIGHT
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    last_r = r - 1

    # Chart 1: Open ARR by Director (horizontal bar)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Open Land ARR, Unweighted vs Weighted"
    chart.y_axis.title = "Director"
    chart.x_axis.title = "EUR"
    data = Reference(ws, min_col=2, max_col=3, min_row=4, max_row=last_r)
    cats = Reference(ws, min_col=1, min_row=5, max_row=last_r)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, "E3")

    start2 = last_r + 3
    ws.cell(
        row=start2,
        column=1,
        value=f"{RUNTIME_PERIOD['prior_quarter_label']} Won and Lost ARR, by Director",
    ).font = BODY_BOLD
    start2 += 1
    pq_won = f"{RUNTIME_PERIOD['prior_quarter_label']} Won ARR"
    pq_lost = f"{RUNTIME_PERIOD['prior_quarter_label']} Lost ARR"
    ws.cell(row=start2, column=1, value="Director")
    ws.cell(row=start2, column=2, value=pq_won)
    ws.cell(row=start2, column=3, value=pq_lost)
    _apply_header(ws, start2, ["Director", pq_won, pq_lost])

    r = start2 + 1
    for director_name in data_store:
        ws.cell(row=r, column=1, value=director_name).alignment = LEFT
        ws.cell(
            row=r,
            column=2,
            value=(
                f"=SUMIFS(LandWonLostDetail[ARR Unwtd],"
                f"LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Won*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        )
        ws.cell(
            row=r,
            column=3,
            value=(
                f"=SUMIFS(LandWonLostDetail[ARR Unwtd],"
                f"LandWonLostDetail[Director],A{r},"
                f'LandWonLostDetail[Type],"Land",'
                f'LandWonLostDetail[Stage],"*Lost*",'
                f'LandWonLostDetail[Close Date],">={RUNTIME_PERIOD["prior_quarter_start"]}",'
                f'LandWonLostDetail[Close Date],"<={RUNTIME_PERIOD["prior_quarter_end"]}")'
            ),
        )
        for ci in (2, 3):
            ws.cell(row=r, column=ci).number_format = "#,##0"
            ws.cell(row=r, column=ci).alignment = RIGHT
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).font = BODY_FONT
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    last_r2 = r - 1

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.grouping = "clustered"
    chart2.title = f"{RUNTIME_PERIOD['prior_quarter_label']} Won vs Lost ARR, Land"
    chart2.y_axis.title = "EUR"
    chart2.x_axis.title = "Director"
    data2 = Reference(ws, min_col=2, max_col=3, min_row=start2, max_row=last_r2)
    cats2 = Reference(ws, min_col=1, min_row=start2 + 1, max_row=last_r2)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 12
    chart2.width = 22
    ws.add_chart(chart2, f"E{start2 - 1}")

    _set_widths(ws, [24, 18, 18])


def build_source_map_sheet(wb):
    """Deck slide to Excel tab mapping so every claim in the presentation
    can be traced back to its source here."""
    ws = wb.create_sheet("Source Map")
    ws["A1"] = "Deck Slide to Source Tab"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    headers = [
        "Deck Slide",
        "What the slide shows",
        "Source Workbook",
        "Source Tab",
        "Filter / Notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header(ws, 3, headers)

    mapping = [
        (
            "Executive Summary",
            "Open pipeline, won/lost, commit, pushed deals",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Summary",
            "Row for the director",
        ),
        (
            f"{RUNTIME_PERIOD['prior_quarter_label']} Promised vs Delivered",
            f"{RUNTIME_PERIOD['prior_quarter_label']} plan vs {RUNTIME_PERIOD['prior_quarter_label']} closed, top 5 wins and losses",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Closed Won YTD",
            f"Filter Territory; use {RUNTIME_PERIOD['prior_quarter_label']} Slips Still Open for what slipped",
        ),
        (
            f"{_historical_trending_contract().retrospective_label} Forecast Variance",
            (
                f"{_historical_trending_contract().retrospective_label} pipeline "
                "decomposition: Won / Lost / Added / Revised with net delta"
            ),
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Forecast Variance",
            (
                "SUMIFS against "
                f"{_historical_trending_contract().retrospective_consolidated_sheet} "
                "Bucket helper; TOTAL row reconciles"
            ),
        ),
        (
            f"{RUNTIME_PERIOD['current_quarter_label']} Outlook",
            f"{RUNTIME_PERIOD['current_quarter_label']} book, commit, best case, closed so far",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Pipeline Overview by Stage",
            f"Filter Close Date to {RUNTIME_PERIOD['current_quarter_title']}",
        ),
        (
            "Pipeline Overview by Stage",
            "Stage distribution with weighted and unweighted ARR",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Pipeline Overview by Stage",
            "Filter Territory; stage is a grouping column",
        ),
        (
            "Top Deals",
            "Largest open deals by ARR",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Pipeline Overview by Stage",
            "Sort ARR Unwtd descending",
        ),
        (
            f"Top {RUNTIME_PERIOD['current_quarter_label']} Deals at Risk",
            f"Open Land deals with highest composite risk score, {RUNTIME_PERIOD['current_quarter_label']} close",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Deal Risk Scoring",
            f"Filter Director; filter Close Date to {RUNTIME_PERIOD['current_quarter_title']}; Proof column shows each rule",
        ),
        (
            "Pushed Deals and PI",
            "Deals with 3+ pushes, top pushed list",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Pipeline Inspection Raw",
            "Sort Push Count descending; PI Summary has bucket view",
        ),
        (
            f"{RUNTIME_PERIOD['prior_quarter_label']} Movement",
            f"{RUNTIME_PERIOD['prior_quarter_label']} slipped deals with old and new close dates",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            f"{RUNTIME_PERIOD['prior_quarter_label']} Slips, Still Open",
            "Filter Territory; Root Cause column is for director input",
        ),
        (
            "Forecast Accuracy",
            "Win rate, forecast category mix",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Commercial Approval Current State + Pipeline Overview by Stage",
            "Use forecast category grouping",
        ),
        (
            "Commercial Approvals",
            "YTD actuals, FY targets, conditionally approved, candidates",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            f"Approvals, {RUNTIME_PERIOD['analysis_year']} + Approval Candidates",
            "Filter Territory; FY view includes prior year approvals",
        ),
        (
            "Missing Approval (Land Stage 3+)",
            "Land deals in Stage 3+ with no commercial approval",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Land Stage 3+, No Approval",
            "Filter Territory",
        ),
        (
            "Renewals",
            f"Renewals due in {RUNTIME_PERIOD['current_quarter_label']} with ACV and probability",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Renewals This Quarter",
            f"Already scoped to {RUNTIME_PERIOD['current_quarter_title']}",
        ),
        (
            "Churn Risk",
            "Placeholder; awaiting Finance feed from Alex P",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Churn Risk",
            "No data yet; shows the data-ownership note",
        ),
        (
            "Forecast Page Reconciliation",
            "Deck Land book vs SF forecast page numbers",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Forecast Reconciliation",
            "Filter Territory",
        ),
        (
            "Data quality, hygiene",
            "Overdue opps, KYC missing, loss reasons",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Overdue Open Opps + KYC Missing",
            "Both tabs have summary blocks grouped by owner or account",
        ),
        (
            "Conversion and velocity",
            "Stage conversion rates, time in stage, push intensity",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Stage Conversion + Time in Stage + Push Intensity",
            "Global; not yet filterable by director",
        ),
        (
            "Concentration and loss analysis",
            "Top 10 and top 25% share, loss reasons, stage at loss",
            f"Dashboard and {RUNTIME_PERIOD['prior_quarter_label']} Analysis",
            "Concentration + Loss Reasons + Stage at Loss",
            "Global; use Region column to filter",
        ),
        (
            "Win Rate",
            "Win rate by deal count and by ARR, by stage",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Win Rate by Stage",
            "Direct pull from the SD Win Rate by Stage report",
        ),
        (
            "Days in Stage",
            "Average and max days open deals sit in each stage",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Days in Stage",
            "Direct pull from the SD Days in Stage report",
        ),
        (
            "Headline summary",
            "Nine KPI cards plus per-director roll-up",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Exec Dashboard",
            "First tab of the workbook",
        ),
        (
            "Open Land pipeline, raw rows",
            f"Every open Land deal in {SCOPE_LABEL}",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Land Pipeline Detail",
            "Named Excel Table, ready for Insert > PivotTable",
        ),
        (
            "Land won or lost, raw rows",
            f"Every Land closed deal in {SCOPE_LABEL}",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Land WonLost Detail",
            "Named Excel Table, ready for Insert > PivotTable",
        ),
        (
            "Native charts",
            f"Open ARR by director plus {RUNTIME_PERIOD['prior_quarter_label']} Won vs Lost by director",
            f"{RUNTIME_PERIOD['fy_label']} Pipeline Review",
            "Charts",
            "Both driven by SUMIFS against the detail tables",
        ),
    ]

    r = 4
    for slide, shows, book, tab, notes in mapping:
        ws.cell(row=r, column=1, value=slide).font = BODY_BOLD
        ws.cell(row=r, column=2, value=shows).alignment = LEFT
        ws.cell(row=r, column=3, value=book).alignment = LEFT
        ws.cell(row=r, column=4, value=tab).alignment = LEFT
        ws.cell(row=r, column=5, value=notes).alignment = LEFT
        for ci in range(1, 6):
            c = ws.cell(row=r, column=ci)
            c.font = BODY_FONT if ci != 1 else BODY_BOLD
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.row_dimensions[r].height = 32
        r += 1
    _set_widths(ws, [34, 46, 24, 42, 50])
    ws.freeze_panes = "A4"


def build_notes_sheet(wb):
    ws = wb.create_sheet("Methodology")
    ws["A1"] = "Methodology"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"Data refreshed {RUNTIME_PERIOD['report_date_display']} "
        "from Salesforce. Scope is the nine MD-1 sales director "
        "territories."
    )
    ws["A2"].font = CAPTION_FONT
    ws["A2"].alignment = LEFT

    notes = [
        ("", ""),
        (
            "Data source",
            (
                "Salesforce REST API, v66.0. Live opportunity, account "
                "and forecast item records. No intermediate caching; "
                "every refresh re-reads the org."
            ),
        ),
        ("", ""),
        (
            "Territory scope",
            (
                "One director per territory, listed on the Summary tab. "
                "Accounts with simcorp, test or delete in the name are "
                "excluded. Owners named Sabiniewicz or Profit are "
                "excluded. Only Type values Land, Expand and Renewal "
                "are considered."
            ),
        ),
        ("", ""),
        (
            "Pipeline scope",
            (
                "Open pipeline and won/lost records on most tabs are "
                "filtered to Type equals Land so the figures align to "
                "the Salesforce forecast page. Expand and Renewal are "
                "shown separately on the Renewals and Closed Won tabs."
            ),
        ),
        ("", ""),
        (
            "ARR Unweighted",
            (
                "APTS_Opportunity_ARR__c. The full deal value that "
                "would book if the opportunity closes at 100 percent."
            ),
        ),
        (
            "ARR Weighted",
            (
                "APTS_Forecast_ARR__c. The probability weighted "
                "forecast value. This is what Salesforce rolls up into "
                "the Commit and Best Case columns on the forecast page."
            ),
        ),
        ("", ""),
        (
            "ACV for renewals",
            (
                "Opportunity.Amount on Type equals Renewal records. "
                "Labelled ACV on the Renewals tab because that is the "
                "convention used by the renewal team."
            ),
        ),
        ("", ""),
        (
            "Forecast reconciliation",
            (
                "Forecast Page values come from the ForecastingItem "
                "object, filtered by OwnerId, Territory2Id and the "
                "Opportunity ARR forecast type. Deck values come from "
                "the raw Land opportunity book. The page only includes "
                "deals a rep or manager has categorised into the "
                "forecast, which is why the totals diverge."
            ),
        ),
        ("", ""),
        (
            f"{RUNTIME_PERIOD['prior_quarter_label']} slips, still open",
            (
                f"Deals that had a {RUNTIME_PERIOD['prior_quarter_title']} close date at any point in "
                "OpportunityFieldHistory, have since been pushed to a "
                "later date, and are still open. Deals that closed "
                "won or lost after slipping are excluded so the ARR "
                "figure reflects current exposure."
            ),
        ),
        ("", ""),
        (
            "Overdue open opportunities",
            (
                "Open opportunities with CloseDate earlier than "
                "today, any type, global. The summary block groups "
                "by owner and sorts by record count so the worst "
                "offenders surface first."
            ),
        ),
        ("", ""),
        (
            "KYC missing",
            (
                "Accounts that have at least one open opportunity "
                "and a KYC_Approval_Status__c other than Approved. "
                "These block progression through contracting."
            ),
        ),
        ("", ""),
        (
            "Churn risk",
            (
                "Not yet automated. The source data sits with Finance "
                "and is received by Alex P. Next step is to formalise "
                "that feed so it appears here on each refresh."
            ),
        ),
        ("", ""),
        (
            "Risk score, purpose",
            (
                "Deal Risk Scoring ranks open Land deals by a composite "
                "0-100+ score. Weights are stored on Parameters (named "
                "cells RiskWeight_<CODE>) and documented below. A score "
                "at or above Thresh_RiskScoreTriage (also on Parameters) "
                "is the triage threshold used by Executive Insights."
            ),
        ),
        (
            "Risk code: PUSH_HIGH / PUSH_MED",
            (
                "Weight 40 if Push Count >= 5 (repeat slippage pattern), "
                "20 if 3-4. Push Count comes straight from the "
                "Opportunity.PushCount field; Salesforce increments it "
                "whenever close date moves to a later quarter."
            ),
        ),
        (
            "Risk code: OVERDUE",
            (
                "Weight 50 if close date is before today and the deal is "
                "still open. OVERDUE also appears on its own tab "
                "(Overdue Open Opps) sourced from a separate SOQL query."
            ),
        ),
        (
            "Risk code: CLOSE_SOON",
            (
                "Weight 30 if close date is within 30 days and stage is "
                "below 4-Shortlisted. Early-stage deals with imminent "
                "close dates almost always slip. Mutually exclusive with "
                "OVERDUE; a deal only gets one of the two."
            ),
        ),
        (
            "Risk code: STALE",
            (
                "Weight 15 if LastActivityDate is more than 60 days ago. "
                "Pulled from Opportunity.LastActivityDate. Lack of "
                "activity is the strongest leading indicator of loss."
            ),
        ),
        (
            "Risk code: NO_NEXT_STEP",
            (
                "Weight 10 if the Next Step field is blank or a dash. "
                "Rep hygiene signal; does not imply the deal is bad, but "
                "removes the manager's ability to coach."
            ),
        ),
        (
            "Risk code: LOW_FCST",
            (
                "Weight 15 when weighted ARR / unweighted ARR is below "
                "20% on a deal above 500K. The rep is carrying a large "
                "deal but forecasting little of it, signalling low "
                "confidence or a probability mismatch."
            ),
        ),
        (
            "Risk code: HIGH_VALUE_PUSH",
            (
                "Weight 10 when ARR is above 1M and the deal has been "
                "pushed at least twice. Amplifies PUSH_MED for large "
                "deals since a single high-value slip moves the quarter."
            ),
        ),
        ("", ""),
        (
            "Forecast Variance, bucket classification",
            (
                "On the "
                f"{_historical_trending_contract().retrospective_consolidated_sheet} "
                "tab each row is tagged "
                "with a Bucket (helper column at the end). Rules: "
                "AlreadyClosed if the initial snapshot stage was Won, "
                "Lost or No Opportunity. Won if the final stage is Won "
                "and started open. Lost if final stage is Lost or No "
                "Opportunity and started open. Added if initial ARR is "
                "0 and final ARR is positive. RevisedUp/RevisedDown if "
                "open at both ends with a non-zero ARR delta. Unchanged "
                "otherwise. Buckets are mutually exclusive; Forecast "
                "Variance uses SUMIFS(..., Bucket, <name>) to compute "
                "each column."
            ),
        ),
        ("", ""),
        (
            "Executive Insights, how to read",
            (
                "Four columns: Finding / Metric / Detail / Source. The "
                "Metric column is a live formula (SUMIFS, MAX, COUNTIF "
                "etc) over the upstream tabs. The Source column is a "
                "HYPERLINK that jumps to the evidence tab. No number is "
                "typed in; click any Metric cell to see the formula."
            ),
        ),
        ("", ""),
        (
            "Changing a threshold",
            (
                "Edit the cell on the Parameters tab. Defined names "
                "(RiskWeight_<CODE>, Thresh_<name>) update the formulas "
                "that reference them. Note: risk scores themselves are "
                "pre-computed in the build script, so changing a "
                "RiskWeight cell re-documents the rule but does NOT "
                "rescore deals; to rescore, edit the RISK_RULES block "
                "in scripts/build_sharepoint_analysis.py and rerun."
            ),
        ),
        ("", ""),
        (
            "Refresh",
            (
                "Workbook regenerated from scripts/extract_director_live.py "
                "and scripts/build_sharepoint_analysis.py. No manual "
                "edits; re-running the scripts reproduces the same "
                "file from live Salesforce data."
            ),
        ),
    ]

    r = 4
    for label, body in notes:
        if label:
            ws.cell(row=r, column=1, value=label).font = BODY_BOLD
            ws.cell(row=r, column=2, value=body).font = BODY_FONT
            ws.cell(row=r, column=2).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            ws.row_dimensions[r].height = max(28, 16 * (len(body) // 80 + 1))
        r += 1
    _set_widths(ws, [22, 90])


def gather_director_data(wb_path, oid, tid, session, instance):
    sheets = _load(wb_path)
    pipeline = sheets.get(f"Pipeline Open {RUNTIME_PERIOD['fy_label']}", [])
    won_lost = sheets.get(f"Won Lost {RUNTIME_PERIOD['fy_label']}", [])
    q1_movement = sheets.get(f"{RUNTIME_PERIOD['prior_quarter_label']} Movement", [])

    # Land open pipeline
    land_open = [
        r for r in pipeline if str(r.get("Type", "")).strip().lower() == "land"
    ]

    # Bucket the Land open pipeline by Forecast Category
    deck_buckets = {
        "Commit": 0,
        "Best Case": 0,
        "Pipeline": 0,
        "Closed": 0,
        "Omitted": 0,
    }
    for r in land_open:
        cat = r.get("Forecast Category") or "Pipeline"
        if cat not in deck_buckets:
            cat = "Pipeline"
        deck_buckets[cat] += float(r.get("ARR Unweighted (EUR)") or 0)

    # Won/lost: Summary-tab metrics use Land only (forecast page scope).
    # Closed Won YTD tab keeps every win so managers can see all bookings.
    q1_won_deals = []
    q1_lost_deals = []
    all_won_land = []
    all_won_any = []
    for r in won_lost:
        is_land = str(r.get("Type", "")).strip().lower() == "land"
        if "Won" in str(r.get("Stage", "")):
            rec = {
                "close_date": str(r.get("Close Date", "") or "")[:10],
                "account": r.get("Account", ""),
                "opportunity": r.get("Opportunity", ""),
                "owner": r.get("Owner", ""),
                "type": r.get("Type", ""),
                "sales_region": r.get("Sales Region", ""),
                "arr_unwtd": float(r.get("ARR Unweighted (EUR)") or 0),
            }
            all_won_any.append(rec)
            if is_land:
                all_won_land.append(rec)
                if _is_q1_of_analysis_year(rec["close_date"]):
                    q1_won_deals.append(rec)
        elif "Lost" in str(r.get("Stage", "")):
            if is_land and _is_q1_of_analysis_year(str(r.get("Close Date", ""))):
                q1_lost_deals.append(
                    {
                        "arr_unwtd": float(r.get("ARR Unweighted (EUR)") or 0),
                    }
                )
    deck_buckets["Closed"] = sum(d["arr_unwtd"] for d in all_won_land)

    q1_slipped_opps = set()
    for r in q1_movement:
        if r.get("Movement") == f"{RUNTIME_PERIOD['prior_quarter_label']} Slipped":
            q1_slipped_opps.add(r.get("Opportunity"))

    open_slips = []
    for r in land_open:
        opp = r.get("Opportunity")
        if opp in q1_slipped_opps:
            open_slips.append(
                {
                    "account": r.get("Account"),
                    "opportunity": opp,
                    "owner": r.get("Owner"),
                    "stage": r.get("Stage"),
                    "close_date": str(r.get("Close Date", "") or "")[:10],
                    "type": r.get("Type"),
                    "arr_unwtd": float(r.get("ARR Unweighted (EUR)") or 0),
                    "arr_wtd": float(r.get("ARR Weighted (EUR)") or 0),
                    "push_count": int(r.get("Push Count") or 0),
                }
            )

    # Commercial approvals: pull three lists from the workbook sheet.
    # The sheet is Land-only by design of the extractor.
    approvals = sheets.get("Commercial Approval", [])

    def _approval_row(r):
        return {
            "account": r.get("Account", ""),
            "opportunity": r.get("Opportunity", ""),
            "owner": r.get("Owner", ""),
            "stage": r.get("Stage", ""),
            "close_date": str(r.get("Close Date", "") or "")[:10],
            "status": r.get("Status", ""),
            "approval_date": str(r.get("Approval Date", "") or "")[:10],
            "next_step": r.get("Next Step", ""),
            "arr_unwtd": float(r.get("ARR Unweighted (EUR)") or 0),
        }

    approved_2026 = [
        _approval_row(r)
        for r in approvals
        if str(r.get("Status", "")).strip()
        == f"Approved {RUNTIME_PERIOD['analysis_year']}"
    ]
    approval_candidates = [
        _approval_row(r)
        for r in approvals
        if "Conditionally" in str(r.get("Status", ""))
        or "Pending" in str(r.get("Status", ""))
    ]
    approval_missing = [
        _approval_row(r) for r in approvals if "Missing" in str(r.get("Status", ""))
    ]

    # Renewals due this quarter, pinned to the explicit report-date quarter.
    renewals_all = sheets.get(f"Renewals {RUNTIME_PERIOD['fy_label']}", [])
    renewals_q2 = []
    for r in renewals_all:
        cd = str(r.get("Close Date", "") or "")[:10]
        if not _is_in_current_quarter(cd):
            continue
        renewals_q2.append(
            {
                "account": r.get("Account", ""),
                "opportunity": r.get("Opportunity", ""),
                "owner": r.get("Owner", ""),
                "stage": r.get("Stage", ""),
                "close_date": cd,
                "probability": r.get("Probability %") or 0,
                "acv": float(r.get("ACV Unweighted (EUR)") or r.get("ACV (EUR)") or 0),
            }
        )

    return {
        "open_count": len(land_open),
        "open_unwtd": sum(float(r.get("ARR Unweighted (EUR)") or 0) for r in land_open),
        "open_wtd": sum(float(r.get("ARR Weighted (EUR)") or 0) for r in land_open),
        "q1_won_count": len(q1_won_deals),
        "q1_won_arr": sum(d["arr_unwtd"] for d in q1_won_deals),
        "q1_lost_count": len(q1_lost_deals),
        "q1_lost_arr": sum(d["arr_unwtd"] for d in q1_lost_deals),
        "slip_still_open_count": len(open_slips),
        "slip_still_open_arr": sum(s["arr_unwtd"] for s in open_slips),
        "open_slips": open_slips,
        "won_any": all_won_any,
        "won": all_won_land,
        "deck_buckets": deck_buckets,
        "forecast_page": forecast_page_fy26(session, instance, oid, tid),
        "approved_2026": approved_2026,
        "approval_candidates": approval_candidates,
        "approval_missing": approval_missing,
        "renewals_q2": renewals_q2,
        # Raw rows used by the live detail tabs (Land Pipeline Detail + Land
        # WonLost Detail) so SUMIFS/COUNTIFS formulas can resolve against them.
        "all_land_open_rows": land_open,
        "all_land_won_lost_rows": [
            r for r in won_lost if str(r.get("Type", "")).strip().lower() == "land"
        ],
    }


def fetch_overdue_and_kyc(session, instance):
    """Pull overdue open opps and KYC-missing accounts globally via SOQL."""
    base_excl = (
        "(NOT Account.Name LIKE '%simcorp%') AND (NOT Account.Name LIKE '%test%') "
        "AND (NOT Account.Name LIKE '%delete%') "
        "AND (NOT Owner.Name LIKE '%Sabiniewicz%') "
        "AND (NOT Owner.Name LIKE '%Profit%')"
    )
    # 1) Overdue open opportunities
    fields = (
        "Id, Name, Account.Name, Owner.Name, StageName, CloseDate, "
        "APTS_Opportunity_ARR__c, PushCount, Type, "
        "Account_Unit_Group__c, Sales_Region__c"
    )
    q = (
        f"SELECT {fields} FROM Opportunity "
        f"WHERE IsClosed=false AND CloseDate < TODAY "
        f"AND Type IN ('Land','Expand','Renewal') AND {base_excl}"
    )
    url = f"{instance}/services/data/v66.0/query"
    records = []
    params = {"q": q}
    while True:
        r = session.get(url, params=params).json()
        records += r.get("records", [])
        if r.get("done"):
            break
        url = f"{instance}{r['nextRecordsUrl']}"
        params = {}

    overdue_rows = []
    for rec in records:
        terr = canonical_territory(
            rec.get("Account_Unit_Group__c"),
            rec.get("Sales_Region__c"),
        )
        overdue_rows.append(
            {
                "territory": terr,
                "director": "",
                "account": (rec.get("Account") or {}).get("Name", ""),
                "opportunity": rec.get("Name", ""),
                "owner": (rec.get("Owner") or {}).get("Name", ""),
                "stage": rec.get("StageName", ""),
                "close_date": rec.get("CloseDate", ""),
                "arr_unwtd": float(rec.get("APTS_Opportunity_ARR__c") or 0),
                "push_count": int(rec.get("PushCount") or 0),
            }
        )

    # 2) KYC missing: accounts with open pipeline + KYC not Approved
    url = f"{instance}/services/data/v66.0/query"
    q_kyc = (
        "SELECT Id, Name, Industry, Region__c, KYC_Approval_Status__c, "
        "(SELECT Id, APTS_Opportunity_ARR__c FROM Opportunities "
        "WHERE IsClosed=false AND Type IN ('Land','Expand','Renewal')) "
        "FROM Account "
        "WHERE Id IN (SELECT AccountId FROM Opportunity "
        "WHERE IsClosed=false AND Type IN ('Land','Expand','Renewal')) "
        "AND (KYC_Approval_Status__c = null "
        "OR KYC_Approval_Status__c != 'Approved') "
        "AND (NOT Name LIKE '%simcorp%') AND (NOT Name LIKE '%test%') "
        "AND (NOT Name LIKE '%delete%')"
    )
    kyc_records = []
    params = {"q": q_kyc}
    while True:
        r = session.get(url, params=params).json()
        kyc_records += r.get("records", [])
        if r.get("done"):
            break
        url = f"{instance}{r['nextRecordsUrl']}"
        params = {}
    kyc_rows = []
    for rec in kyc_records:
        opps = (rec.get("Opportunities") or {}).get("records") or []
        kyc_rows.append(
            {
                "account": rec.get("Name", ""),
                "region": rec.get("Region__c", ""),
                "industry": rec.get("Industry", ""),
                "kyc_status": rec.get("KYC_Approval_Status__c") or "Not Started",
                "open_opps": len(opps),
                "open_arr": sum(
                    float(o.get("APTS_Opportunity_ARR__c") or 0) for o in opps
                ),
            }
        )
    return overdue_rows, kyc_rows


# module-level so the sheet builders can read it
data_store: dict = {}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbooks-dir",
        type=Path,
        default=Path("output/director_live_workbooks")
        / datetime.now().strftime("%Y-%m-%d"),
    )
    parser.add_argument(
        "--date",
        default=None,
        help=(
            "Explicit report date YYYY-MM-DD. Defaults to the workbooks-dir "
            "folder date when present, otherwise today."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path. Default derives from --territory.",
    )
    parser.add_argument(
        "--territory",
        default=None,
        help=(
            "Scope the workbook to a single territory (e.g. 'APAC', "
            "'EMEA UK & Ireland'). Omit for the cross-territory master."
        ),
    )
    args = parser.parse_args()
    _configure_runtime_period(
        as_of_date=args.date,
        workbooks_dir=args.workbooks_dir,
    )

    # Filter DIRECTORS to the requested territory. In-place so downstream
    # module globals (data_store) only see the scoped set.
    global SCOPE_LABEL
    if args.territory:
        scoped = [d for d in DIRECTORS if d[1] == args.territory]
        if not scoped:
            known = sorted({d[1] for d in DIRECTORS})
            raise SystemExit(f"Unknown territory {args.territory!r}. Valid: {known}")
        DIRECTORS[:] = scoped
        SCOPE_LABEL = args.territory

    # Default output: master → "All Territories.xlsx"; regional → "{Territory}.xlsx"
    # (slashes in territory names replaced with hyphens for a legal filename).
    if args.output is None:
        safe = SCOPE_LABEL.replace("/", "-")
        args.output = (
            Path("output/sharepoint")
            / f"{RUNTIME_PERIOD['fy_label']} Pipeline Review, {safe}.xlsx"
        )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    token, instance = get_sf_auth()
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    # Parallel loader: gather_director_data is ~80% I/O (SF REST + openpyxl
    # read), so a thread pool with 9 workers cuts the 30-35s serial phase to
    # roughly max(single-director time) + thread overhead. Requests sessions
    # are thread-safe when used read-only, which is all we do here.
    from concurrent.futures import ThreadPoolExecutor, as_completed

    loadable = []
    for name, _territory, fname, oid, tid in DIRECTORS:
        wb_path = args.workbooks_dir / fname
        if not wb_path.exists():
            print(f"  skip {name}: workbook missing at {wb_path}")
            continue
        loadable.append((name, wb_path, oid, tid))

    if loadable:
        print(f"  loading {len(loadable)} director(s) in parallel...")
        with ThreadPoolExecutor(max_workers=min(9, len(loadable))) as pool:
            futures = {
                pool.submit(
                    gather_director_data, wb_path, oid, tid, session, instance
                ): name
                for (name, wb_path, oid, tid) in loadable
            }
            # Also kick off overdue + KYC in parallel with director loads.
            hygiene_future = pool.submit(fetch_overdue_and_kyc, session, instance)
            for fut in as_completed(futures):
                name = futures[fut]
                try:
                    data_store[name] = fut.result()
                except Exception as exc:
                    print(f"    [FAIL] {name}: {exc}")
                    continue
                print(f"    [ok] {name}")
            overdue_rows, kyc_rows = hygiene_future.result()
    else:
        print("  fetching global overdue and KYC lists...")
        overdue_rows, kyc_rows = fetch_overdue_and_kyc(session, instance)

    # Scope global hygiene rows to the selected territory. Overdue has a
    # canonical territory field; KYC rows only carry Account.Region__c which
    # is region-level (e.g. "Americas") so we include if region prefix matches
    # the territory's region family.
    if args.territory:
        overdue_rows = [r for r in overdue_rows if r.get("territory") == args.territory]
        kyc_rows = [
            r
            for r in kyc_rows
            if r.get("region") and args.territory.lower() in str(r["region"]).lower()
        ]

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Ordered for monthly review flow: headline, summary, approvals,
    # Q1 retrospective, open pipeline, data hygiene, reference.
    build_exec_dashboard(wb, DIRECTORS, overdue_rows, kyc_rows)
    build_summary_sheet(wb, DIRECTORS)
    build_forecast_recon_sheet(wb, DIRECTORS)
    build_approvals_overview_sheet(wb, DIRECTORS)
    build_approvals_ytd_sheet(wb, DIRECTORS)
    build_approvals_candidates_sheet(wb, DIRECTORS)
    build_approvals_missing_sheet(wb, DIRECTORS)
    build_q1_slips_sheet(wb, DIRECTORS)
    # Win Rate and Days in Stage hit global SF reports and aren't
    # territory-filterable. Master only; regional stays focused per-director.
    if not args.territory:
        build_win_rate_sheet(wb, session, instance)
        build_days_in_stage_sheet(wb, session, instance)
    build_closed_won_sheet(wb, DIRECTORS)
    build_renewals_sheet(wb, DIRECTORS)
    build_overdue_open_opps_sheet(wb, overdue_rows)
    build_kyc_missing_sheet(wb, kyc_rows)
    build_churn_sheet(wb)
    build_land_detail_sheet(wb)
    build_land_won_lost_detail_sheet(wb)
    contract = _historical_trending_contract()
    build_snapshot_trend_consolidated(
        wb,
        contract.retrospective_label,
        contract.retrospective_snapshot_sheet,
        args.workbooks_dir,
    )
    build_snapshot_trend_consolidated(
        wb,
        contract.current_label,
        contract.current_snapshot_sheet,
        args.workbooks_dir,
    )
    build_pipeline_pivot(wb)
    build_arr_concentration(wb)
    build_pipeline_velocity(wb, args.workbooks_dir)
    build_slip_risk(wb)
    build_territory_scorecard(wb, overdue_rows, kyc_rows)
    build_parameters_sheet(wb)
    build_deal_risk_scoring(wb, str(RUNTIME_PERIOD["report_date"]))
    build_forecast_variance(wb, args.workbooks_dir)
    build_commit_accuracy(wb)
    # Advanced analytics tabs (audit-driven additions)
    build_owner_scorecard(wb)
    build_competitive_win_loss(wb)
    build_stage_conversion_per_territory(wb)
    build_sales_velocity(wb)
    build_deal_age_distribution(wb)
    build_account_penetration(wb)
    build_forecast_bias(wb)
    # build_approvals_dashboard removed — redundant with build_approvals_overview_sheet
    build_days_in_stage_by_director(wb)
    build_win_rate_trend(wb)
    build_executive_insights(
        wb, overdue_rows, kyc_rows, str(RUNTIME_PERIOD["report_date"])
    )
    build_charts_sheet(wb)
    build_source_map_sheet(wb)
    build_notes_sheet(wb)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")


if __name__ == "__main__":
    main()
