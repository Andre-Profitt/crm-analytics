"""
Consolidated analytics workbook for the Sales Directors Monthly dashboard.

Pulls every report on the dashboard, plus Q1 field-history data, plus the
Pipeline Inspection extracts from all nine territory workbooks. Writes a
single .xlsx with raw tabs followed by analytical summaries.
"""

import argparse
import json
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from monthly_platform.period import resolve_period_context, sheet_names
except ModuleNotFoundError:  # pragma: no cover
    from scripts.monthly_platform.period import resolve_period_context, sheet_names

SN = sheet_names()
PERIOD = resolve_period_context()

DASHBOARD_ID = "01ZTb00000FSP7hMAH"  # Sales Directors Monthly
SALES_OPS_DASHBOARD_ID = "01ZTb00000FSP9JMAX"  # Sales Ops Quarterly KPI
WORKBOOKS_DIR = Path(f"output/director_live_workbooks/{PERIOD.snapshot_date}")

NAVY = "1F3864"
LIGHT = "F2F2F2"
BORDER_GRAY = "BFBFBF"
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
CAPTION = Font(name="Calibri", size=9, italic=True, color="595959")
THIN = Side(style="thin", color=BORDER_GRAY)
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def get_auth():
    data = json.loads(
        subprocess.run(
            ["sf", "org", "display", "--target-org", "apro@simcorp.com", "--json"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout
    )["result"]
    return data["accessToken"], data["instanceUrl"]


def run_report(session, instance, rid):
    """Return (cols, labels, dtypes, rows, totals, format) for a report.

    For SUMMARY reports the grouping levels (Tier, Risk, etc.) are promoted
    to leading columns so every detail row shows the grouping it belonged to.
    """
    r = session.post(
        f"{instance}/services/data/v66.0/analytics/reports/{rid}?includeDetails=true",
        headers={"Content-Type": "application/json"},
    ).json()
    md = r.get("reportMetadata", {})
    cols = md.get("detailColumns", [])
    ext = r.get("reportExtendedMetadata", {}).get("detailColumnInfo", {})
    detail_labels = [ext.get(c, {}).get("label", c) for c in cols]
    detail_dtypes = [ext.get(c, {}).get("dataType", "string") for c in cols]

    grp_down = md.get("groupingsDown") or []
    grp_ext = r.get("reportExtendedMetadata", {}).get("groupingColumnInfo", {})
    grp_labels = [
        grp_ext.get(g.get("name"), {}).get("label", g.get("name")) for g in grp_down
    ]

    fact = r.get("factMap", {})
    rows = []

    if md.get("reportFormat") == "TABULAR":
        for row in fact.get("T!T", {}).get("rows", []):
            rows.append([c.get("label", "") for c in row.get("dataCells", [])])
        labels = detail_labels
        dtypes = detail_dtypes
    else:
        # Walk the grouping tree to map each leaf key to its label chain
        key_to_labels = {}

        def walk(groupings, path):
            for g in groupings:
                key = g.get("key")
                lbl = g.get("label", "")
                this_path = path + [lbl]
                key_to_labels[key] = this_path
                children = g.get("groupings") or []
                if children:
                    walk(children, this_path)

        walk((r.get("groupingsDown") or {}).get("groupings", []), [])

        for key, fm in fact.items():
            if key == "T!T" or "!" not in key:
                continue
            group_key = key.split("!")[0]
            leaf = key_to_labels.get(group_key, [])
            # Pad to the full grouping depth
            pad = list(leaf) + [""] * (len(grp_labels) - len(leaf))
            for row in fm.get("rows", []):
                detail_vals = [c.get("label", "") for c in row.get("dataCells", [])]
                rows.append(pad + detail_vals)

        labels = list(grp_labels) + detail_labels
        dtypes = (["string"] * len(grp_labels)) + detail_dtypes

    total_agg = fact.get("T!T", {}).get("aggregates", [])
    return cols, labels, dtypes, rows, total_agg, md.get("reportFormat")


def apply_header(ws, row, n_cols):
    for i in range(1, n_cols + 1):
        c = ws.cell(row=row, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def zebra(ws, r_start, r_end, n_cols):
    for r in range(r_start, r_end + 1):
        if r % 2 == 0:
            fill = PatternFill("solid", fgColor=LIGHT)
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _dedupe_table_name(wb, base):
    name = base
    i = 1
    existing = {
        t.name for ws in wb.worksheets for t in getattr(ws, "tables", {}).values()
    }
    while name in existing:
        i += 1
        name = f"{base}{i}"
    return name


def _as_excel_table(ws, labels, rows, dtypes, table_name):
    """Render rows as a native Excel Table (ListObject) for one-click pivoting.

    Data starts at row 1 (headers) so Excel treats the entire used range as
    the table. Returns the Table object.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    for i, lbl in enumerate(labels, 1):
        ws.cell(row=1, column=i, value=lbl)
    apply_header(ws, 1, len(labels))

    r = 2
    for row in rows:
        for ci, (val, dt) in enumerate(zip(row, dtypes), 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = BODY
            if dt in ("currency", "double", "percent"):
                cell.alignment = RIGHT
                cell.number_format = "#,##0"
            elif dt == "int":
                cell.alignment = CENTER
            elif dt == "date":
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        r += 1

    last_row = max(r - 1, 2)  # need at least 1 data row for a valid table
    last_col_letter = get_column_letter(len(labels))
    ref = f"A1:{last_col_letter}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"

    widths = []
    for i, lbl in enumerate(labels):
        w = max(len(str(lbl)) + 4, 14)
        if rows:
            longest = max((len(str(row[i] or "")) for row in rows), default=0)
            w = max(w, min(longest + 2, 45))
        widths.append(w)
    set_widths(ws, widths)
    return tbl


def write_raw_report_tab(ws, title, subtitle, labels, dtypes, rows):
    """Legacy wrapper: just renders as an Excel Table. Title/subtitle ignored."""
    safe_name = "".join(c for c in title if c.isalnum())[:30] or "Data"
    _as_excel_table(ws, labels, rows, dtypes, _dedupe_table_name(ws.parent, safe_name))


def parse_eur(label):
    """Parse a Salesforce EUR label like 'EUR 1.234.567,89' or '1,234,567' into float."""
    if label is None:
        return 0.0
    s = str(label)
    # Strip EUR, spaces, currency prefixes
    for token in ("EUR", "USD", "CAD", "GBP", "AUD", "JPY", "CHF", "€", "$"):
        s = s.replace(token, "")
    s = s.strip()
    if not s or s == "-":
        return 0.0
    # Euro locale uses "." thousands and "," decimal; detect by last separator
    if "," in s and "." in s:
        # If last "," comes after last ".", that "," is decimal
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # Ambiguous; assume comma is decimal if only one "," and 2 digits after
        if s.count(",") == 1 and len(s.split(",")[-1]) == 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ──────────────── Q1 History ────────────────


def fetch_q1_history(session, instance):
    """Close date + stage changes for prior-quarter deals, global."""
    base_excl = (
        "(NOT Opportunity.Account.Name LIKE '%simcorp%') "
        "AND (NOT Opportunity.Account.Name LIKE '%test%') "
        "AND (NOT Opportunity.Account.Name LIKE '%delete%') "
        "AND (NOT Opportunity.Owner.Name LIKE '%Sabiniewicz%') "
        "AND (NOT Opportunity.Owner.Name LIKE '%Profit%')"
    )
    fields = (
        "OpportunityId, Opportunity.Name, Opportunity.Account.Name, "
        "Opportunity.Account_Unit_Group__c, "
        "Opportunity.Sales_Region__c, Opportunity.Type, "
        "Opportunity.Owner.Name, Opportunity.StageName, "
        "Opportunity.CloseDate, Opportunity.IsClosed, "
        "Opportunity.IsWon, Opportunity.APTS_Opportunity_ARR__c, "
        "Field, OldValue, NewValue, CreatedDate"
    )
    q = (
        f"SELECT {fields} FROM OpportunityFieldHistory "
        f"WHERE Field IN ('CloseDate', 'StageName') "
        f"AND Opportunity.Type IN ('Land','Expand','Renewal') "
        f"AND {base_excl} "
        f"AND CreatedDate >= 2025-10-01T00:00:00Z "
        f"ORDER BY CreatedDate ASC"
    )
    records = []
    url = f"{instance}/services/data/v66.0/query"
    params = {"q": q}
    while True:
        r = session.get(url, params=params).json()
        records += r.get("records", [])
        if r.get("done"):
            break
        url = f"{instance}{r['nextRecordsUrl']}"
        params = {}
    rows = []
    for rec in records:
        opp = rec.get("Opportunity") or {}
        acc = opp.get("Account") or {}
        own = opp.get("Owner") or {}
        aug = opp.get("Account_Unit_Group__c") or ""
        reg = opp.get("Sales_Region__c") or ""
        if aug == "SC Asia":
            territory = "APAC"
        elif aug == "SC North America":
            territory = f"NA {reg}".strip()
        elif aug == "SC EMEA":
            territory = f"EMEA {reg}".strip()
        else:
            territory = aug or "Unspecified"
        rows.append(
            {
                "territory": territory,
                "opp_id": rec.get("OpportunityId", ""),
                "opportunity": opp.get("Name", ""),
                "account": acc.get("Name", ""),
                "type": opp.get("Type", ""),
                "owner": own.get("Name", ""),
                "current_stage": opp.get("StageName", ""),
                "current_close": opp.get("CloseDate", ""),
                "is_closed": bool(opp.get("IsClosed")),
                "is_won": bool(opp.get("IsWon")),
                "arr": float(opp.get("APTS_Opportunity_ARR__c") or 0),
                "field": rec.get("Field", ""),
                "old_value": str(rec.get("OldValue") or "")[:10]
                if rec.get("Field") == "CloseDate"
                else str(rec.get("OldValue") or ""),
                "new_value": str(rec.get("NewValue") or "")[:10]
                if rec.get("Field") == "CloseDate"
                else str(rec.get("NewValue") or ""),
                "changed_on": str(rec.get("CreatedDate", ""))[:10],
            }
        )
    return rows


def build_q1_history_raw(wb, history_rows):
    pq = PERIOD.prior_quarter
    ws = wb.create_sheet(f"{pq.label} History Raw")
    write_raw_report_tab(
        ws,
        f"{pq.title} Field History, Raw Events",
        (
            "Every CloseDate and StageName change on a non-test, non-internal "
            "Land/Expand/Renewal opportunity since October 2025. One row per "
            "change event. Filter by Field to see only close-date slips or only "
            "stage transitions."
        ),
        [
            "Territory",
            "Opportunity",
            "Account",
            "Type",
            "Owner",
            "Current Stage",
            "Current Close",
            "Is Closed",
            "Is Won",
            "ARR Unwtd (EUR)",
            "Field Changed",
            "Old Value",
            "New Value",
            "Changed On",
        ],
        ["string"] * 6
        + [
            "date",
            "string",
            "string",
            "currency",
            "string",
            "string",
            "string",
            "date",
        ],
        [
            [
                r["territory"],
                r["opportunity"],
                r["account"],
                r["type"],
                r["owner"],
                r["current_stage"],
                r["current_close"],
                "Yes" if r["is_closed"] else "No",
                "Yes" if r["is_won"] else "No",
                r["arr"],
                r["field"],
                r["old_value"],
                r["new_value"],
                r["changed_on"],
            ]
            for r in history_rows
        ],
    )


def build_q1_slips_pivot(wb, history_rows):
    """Pivot: Territory x Type x Stage with count and ARR for prior-quarter slips (still-open)."""
    pq = PERIOD.prior_quarter
    # Identify slip events: CloseDate change where old is in prior quarter and new > it
    slips_by_opp = {}
    for r in history_rows:
        if r["field"] != "CloseDate":
            continue
        old = r["old_value"]
        new = r["new_value"]
        if not (old and new) or old == new:
            continue
        if pq.start_date <= old <= pq.end_date and new > pq.end_date:
            slips_by_opp[r["opp_id"]] = r  # overwrite keeps latest
    # Only keep those still open
    still_open = {k: v for k, v in slips_by_opp.items() if not v["is_closed"]}

    ws = wb.create_sheet(f"{pq.label} Slips Pivot")
    ws["A1"] = f"{pq.label} Slips, Still Open"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    # Pivot 1: Territory x Type (count + ARR)
    pivot1 = defaultdict(lambda: defaultdict(lambda: {"n": 0, "arr": 0.0}))
    for r in still_open.values():
        pivot1[r["territory"]][r["type"]]["n"] += 1
        pivot1[r["territory"]][r["type"]]["arr"] += r["arr"]
    types = ["Land", "Expand", "Renewal"]

    ws["A4"] = "By Territory and Type"
    ws["A4"].font = BODY_BOLD
    headers = (
        ["Territory"]
        + [f"{t} Count" for t in types]
        + [f"{t} ARR" for t in types]
        + ["Total Count", "Total ARR"]
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    apply_header(ws, 5, len(headers))

    r = 6
    totals = defaultdict(float)
    totals_int = defaultdict(int)
    for terr in sorted(pivot1.keys()):
        row_counts = [pivot1[terr][t]["n"] for t in types]
        row_arrs = [pivot1[terr][t]["arr"] for t in types]
        total_n = sum(row_counts)
        total_a = sum(row_arrs)
        ws.cell(row=r, column=1, value=terr).alignment = LEFT
        for i, n in enumerate(row_counts, 2):
            ws.cell(row=r, column=i, value=n).alignment = CENTER
        for i, a in enumerate(row_arrs, 2 + len(types)):
            cell = ws.cell(row=r, column=i, value=a)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
        ws.cell(row=r, column=2 + 2 * len(types), value=total_n).alignment = CENTER
        cell = ws.cell(row=r, column=3 + 2 * len(types), value=total_a)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        for i, n in enumerate(row_counts):
            totals_int[types[i]] += n
            totals[types[i]] += row_arrs[i]
        totals_int["total"] += total_n
        totals["total"] += total_a
        r += 1
    # Totals row
    ws.cell(row=r, column=1, value="Total").font = BODY_BOLD
    for i, t in enumerate(types, 2):
        ws.cell(row=r, column=i, value=totals_int[t]).alignment = CENTER
    for i, t in enumerate(types, 2 + len(types)):
        cell = ws.cell(row=r, column=i, value=totals[t])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
    ws.cell(
        row=r, column=2 + 2 * len(types), value=totals_int["total"]
    ).alignment = CENTER
    cell = ws.cell(row=r, column=3 + 2 * len(types), value=totals["total"])
    cell.alignment = RIGHT
    cell.number_format = "#,##0"
    for ci in range(1, len(headers) + 1):
        ws.cell(row=r, column=ci).font = BODY_BOLD
        ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor="E7E6E6")
        ws.cell(row=r, column=ci).border = BORDER
    zebra(ws, 6, r - 1, len(headers))
    set_widths(ws, [20] + [10] * 3 + [16] * 3 + [10, 18])
    ws.freeze_panes = "A6"


def build_q1_slips_by_owner(wb, history_rows):
    pq = PERIOD.prior_quarter
    ws = wb.create_sheet(f"{pq.label} Slips by Owner")
    # Latest slip per opp, still open
    slips = {}
    for r in history_rows:
        if r["field"] != "CloseDate":
            continue
        old, new = r["old_value"], r["new_value"]
        if not (old and new) or old == new:
            continue
        if pq.start_date <= old <= pq.end_date and new > pq.end_date:
            slips[r["opp_id"]] = r
    still_open = [r for r in slips.values() if not r["is_closed"]]

    by_owner = defaultdict(lambda: {"n": 0, "arr": 0.0, "territory": ""})
    for r in still_open:
        o = r["owner"] or "(unknown)"
        by_owner[o]["n"] += 1
        by_owner[o]["arr"] += r["arr"]
        by_owner[o]["territory"] = r["territory"]

    rows = sorted(
        [[o, d["territory"], d["n"], d["arr"]] for o, d in by_owner.items()],
        key=lambda x: (-x[2], -x[3]),
    )

    ws["A1"] = f"{pq.label} Slips, by Owner"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    headers = ["Owner", "Primary Territory", "Slip Count", "Slip ARR (EUR)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    apply_header(ws, 4, len(headers))
    r = 5
    for row in rows:
        ws.cell(row=r, column=1, value=row[0]).alignment = LEFT
        ws.cell(row=r, column=2, value=row[1]).alignment = LEFT
        ws.cell(row=r, column=3, value=row[2]).alignment = CENTER
        cell = ws.cell(row=r, column=4, value=row[3])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    if rows:
        zebra(ws, 5, r - 1, len(headers))
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{r - 1}"
    set_widths(ws, [26, 22, 12, 18])
    ws.freeze_panes = "A5"


def build_q1_changes_by_region(wb, history_rows):
    """Cross-tab: territory x change type with counts and ARR impact."""
    pq = PERIOD.prior_quarter
    cq = PERIOD.current_quarter
    fq = PERIOD.forward_quarter
    # Derive Q4 end from the reporting window (end of FY)
    fy_end = PERIOD.reporting_window_end
    next_fy = f"FY{int(PERIOD.fiscal_year[2:]) + 1:02d}"

    # Latest slip per opp
    slip_latest = {}
    for r in history_rows:
        if r["field"] != "CloseDate":
            continue
        old, new = r["old_value"], r["new_value"]
        if not (old and new) or old == new:
            continue
        slip_latest[r["opp_id"]] = r

    # Categorise each latest slip event
    regions = defaultdict(
        lambda: {
            "pq_to_cq": {"n": 0, "arr": 0.0},
            "pq_to_fq": {"n": 0, "arr": 0.0},
            "pq_to_q4": {"n": 0, "arr": 0.0},
            "pq_to_next_fy": {"n": 0, "arr": 0.0},
            "pulled_in": {"n": 0, "arr": 0.0},
            "closed_won": {"n": 0, "arr": 0.0},
            "closed_lost": {"n": 0, "arr": 0.0},
        }
    )
    for r in slip_latest.values():
        t = r["territory"] or "Unspecified"
        old, new = r["old_value"], r["new_value"]
        arr = r["arr"]
        in_pq = pq.start_date <= old <= pq.end_date
        if not in_pq:
            continue
        if r["is_closed"]:
            if r["is_won"]:
                bucket = "closed_won"
            else:
                bucket = "closed_lost"
        elif new < old:
            bucket = "pulled_in"
        elif new <= cq.end_date:
            bucket = "pq_to_cq"
        elif new <= fq.end_date:
            bucket = "pq_to_fq"
        elif new <= fy_end:
            bucket = "pq_to_q4"
        else:
            bucket = "pq_to_next_fy"
        regions[t][bucket]["n"] += 1
        regions[t][bucket]["arr"] += arr

    ws = wb.create_sheet(f"{pq.label} Changes by Region")
    ws["A1"] = f"{pq.title} Close-Date Changes, by Region"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    bucket_list = [
        ("pq_to_cq", f"Slipped to {cq.label}"),
        ("pq_to_fq", f"Slipped to {fq.label}"),
        ("pq_to_q4", "Slipped to Q4"),
        ("pq_to_next_fy", f"Slipped to {next_fy}+"),
        ("pulled_in", "Pulled In"),
        ("closed_won", "Closed Won"),
        ("closed_lost", "Closed Lost"),
    ]
    headers = (
        ["Region"]
        + [f"{lbl} Count" for _, lbl in bucket_list]
        + [f"{lbl} ARR" for _, lbl in bucket_list]
        + ["Total Count", "Total ARR"]
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))

    r = 4
    grand_counts = defaultdict(int)
    grand_arrs = defaultdict(float)
    for region in sorted(regions.keys()):
        data = regions[region]
        ws.cell(row=r, column=1, value=region).alignment = LEFT
        total_n = 0
        total_arr = 0.0
        for i, (key, _) in enumerate(bucket_list, 2):
            n = data[key]["n"]
            ws.cell(row=r, column=i, value=n).alignment = CENTER
            total_n += n
            grand_counts[key] += n
        for i, (key, _) in enumerate(bucket_list, 2 + len(bucket_list)):
            arr = data[key]["arr"]
            cell = ws.cell(row=r, column=i, value=arr)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
            total_arr += arr
            grand_arrs[key] += arr
        ws.cell(
            row=r, column=2 + 2 * len(bucket_list), value=total_n
        ).alignment = CENTER
        cell = ws.cell(row=r, column=3 + 2 * len(bucket_list), value=total_arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    # Grand total row
    ws.cell(row=r, column=1, value="Total").font = BODY_BOLD
    total_all_n = 0
    total_all_arr = 0.0
    for i, (key, _) in enumerate(bucket_list, 2):
        ws.cell(row=r, column=i, value=grand_counts[key]).alignment = CENTER
        total_all_n += grand_counts[key]
    for i, (key, _) in enumerate(bucket_list, 2 + len(bucket_list)):
        cell = ws.cell(row=r, column=i, value=grand_arrs[key])
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        total_all_arr += grand_arrs[key]
    ws.cell(
        row=r, column=2 + 2 * len(bucket_list), value=total_all_n
    ).alignment = CENTER
    cell = ws.cell(row=r, column=3 + 2 * len(bucket_list), value=total_all_arr)
    cell.alignment = RIGHT
    cell.number_format = "#,##0"
    for ci in range(1, len(headers) + 1):
        ws.cell(row=r, column=ci).font = BODY_BOLD
        ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor="E7E6E6")
        ws.cell(row=r, column=ci).border = BORDER
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22] + [10] * len(bucket_list) + [14] * len(bucket_list) + [10, 14])
    ws.freeze_panes = "B4"


def build_stage_progression(wb, history_rows):
    """Matrix of old-stage to new-stage transitions among prior-quarter opportunities."""
    pq = PERIOD.prior_quarter
    # Prior-quarter deals = any deal whose close date ever sat in that quarter
    q1_opp_ids = set()
    for r in history_rows:
        if r["field"] == "CloseDate":
            old, new = r["old_value"], r["new_value"]
            if old and pq.start_date <= old <= pq.end_date:
                q1_opp_ids.add(r["opp_id"])
            if new and pq.start_date <= new <= pq.end_date:
                q1_opp_ids.add(r["opp_id"])
    # Stage transitions on those deals
    matrix = defaultdict(lambda: defaultdict(int))
    stage_order = [
        "1 - Prospecting",
        "2 - Discovery",
        "3 - Engagement",
        "4 - Shortlisted",
        "5 - Preferred",
        "6 - Contracting",
        "7 - Opt Out",
        "8 - Won",
        "0 - Lost",
        "0 - No Opportunity",
    ]
    for r in history_rows:
        if r["field"] != "StageName":
            continue
        if r["opp_id"] not in q1_opp_ids:
            continue
        old = r["old_value"] or "(none)"
        new = r["new_value"] or "(none)"
        matrix[old][new] += 1
    stages = sorted(
        set(matrix.keys()) | {s for d in matrix.values() for s in d.keys()},
        key=lambda s: stage_order.index(s) if s in stage_order else 99,
    )

    ws = wb.create_sheet(f"{pq.label} Stage Transitions")
    ws["A1"] = f"{pq.label} Stage Transitions (From / To)"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    headers = ["From \\ To"] + stages
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    apply_header(ws, 4, len(headers))
    r = 5
    for old in stages:
        ws.cell(row=r, column=1, value=old).alignment = LEFT
        ws.cell(row=r, column=1).font = BODY_BOLD
        ws.cell(row=r, column=1).border = BORDER
        for i, new in enumerate(stages, 2):
            n = matrix[old].get(new, 0)
            cell = ws.cell(row=r, column=i, value=n or "")
            cell.alignment = CENTER
            cell.font = BODY
            cell.border = BORDER
            # Shade non-zero cells; red for backwards, green for forwards
            if n:
                try:
                    is_forward = stage_order.index(new) > stage_order.index(old)
                except ValueError:
                    is_forward = True
                color = "E2EFDA" if is_forward else "F8CBAD"
                cell.fill = PatternFill("solid", fgColor=color)
        r += 1
    set_widths(ws, [22] + [14] * len(stages))
    ws.freeze_panes = "B5"


# ──────────────── Pipeline Inspection consolidation ────────────────


def build_pi_consolidated(wb):
    """Consolidate PI extracts from all director workbooks."""
    all_rows = []
    for wb_file in sorted(WORKBOOKS_DIR.glob("*.xlsx")):
        name = wb_file.stem.replace("-", " ").title()
        try:
            wb_dir = load_workbook(wb_file, data_only=True)
            if "Pipeline Inspection" not in wb_dir.sheetnames:
                continue
            ws = wb_dir["Pipeline Inspection"]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {headers[i]: v for i, v in enumerate(row)}
                rec["Director"] = name
                all_rows.append(rec)
        except Exception as e:
            print(f"  PI load error {name}: {e}")

    ws = wb.create_sheet("Pipeline Inspection Raw")
    labels = [
        "Director",
        "Opportunity",
        "Owner",
        "Stage",
        "Forecast Category",
        "ARR Weighted (EUR)",
        "Close Date",
        "Push Count",
        "Score",
        "Priority",
    ]
    rows_data = [
        [
            r.get("Director", ""),
            r.get("Opportunity", ""),
            r.get("Owner", ""),
            r.get("Stage", ""),
            r.get("Forecast Category", ""),
            float(r.get("ARR Weighted (EUR)") or 0),
            r.get("Close Date", ""),
            r.get("Push Count") or 0,
            r.get("Score") or 0,
            r.get("Priority", ""),
        ]
        for r in all_rows
    ]
    rows_data.sort(key=lambda x: (-x[7], -x[5]))  # push count desc, then ARR
    write_raw_report_tab(
        ws,
        "Pipeline Inspection, All Territories",
        (
            "Consolidated PI list from each director workbook. Sorted by "
            "push count descending so the most stalled deals appear first."
        ),
        labels,
        [
            "string",
            "string",
            "string",
            "string",
            "string",
            "currency",
            "date",
            "int",
            "int",
            "string",
        ],
        rows_data,
    )
    return all_rows


def build_pi_summary(wb, pi_rows):
    """PI summary pivots: push-count buckets, forecast category, priority."""
    ws = wb.create_sheet("PI Summary")
    ws["A1"] = "Pipeline Inspection, Summary"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    directors = sorted({r.get("Director", "?") for r in pi_rows})
    # Push-count buckets
    buckets = [
        ("0", lambda n: n == 0),
        ("1-2", lambda n: 1 <= n <= 2),
        ("3-4", lambda n: 3 <= n <= 4),
        ("5+", lambda n: n >= 5),
    ]

    # Build director-by-bucket count matrix
    r_start = 4
    ws.cell(row=r_start - 1, column=1, value="Push-count buckets").font = BODY_BOLD
    headers = ["Director"] + [b[0] for b in buckets] + ["Total Deals", "Total ARR Wtd"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r_start, column=i, value=h)
    apply_header(ws, r_start, len(headers))
    r = r_start + 1
    for d in directors:
        rows_d = [x for x in pi_rows if x.get("Director") == d]
        counts = []
        for _, pred in buckets:
            n = sum(1 for x in rows_d if pred(int(x.get("Push Count") or 0)))
            counts.append(n)
        total = len(rows_d)
        arr = sum(float(x.get("ARR Weighted (EUR)") or 0) for x in rows_d)
        ws.cell(row=r, column=1, value=d).alignment = LEFT
        for i, n in enumerate(counts, 2):
            ws.cell(row=r, column=i, value=n).alignment = CENTER
        ws.cell(row=r, column=2 + len(buckets), value=total).alignment = CENTER
        cell = ws.cell(row=r, column=3 + len(buckets), value=arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, r_start + 1, r - 1, len(headers))

    # Forecast-category mix
    r_start = r + 2
    ws.cell(row=r_start - 1, column=1, value="Forecast category mix").font = BODY_BOLD
    cats = ["Commit", "Best Case", "Pipeline", "Omitted", "Closed"]
    headers = ["Director"] + cats + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r_start, column=i, value=h)
    apply_header(ws, r_start, len(headers))
    r = r_start + 1
    for d in directors:
        rows_d = [x for x in pi_rows if x.get("Director") == d]
        ws.cell(row=r, column=1, value=d).alignment = LEFT
        total = 0.0
        for i, cat in enumerate(cats, 2):
            arr = sum(
                float(x.get("ARR Weighted (EUR)") or 0)
                for x in rows_d
                if str(x.get("Forecast Category") or "") == cat
            )
            total += arr
            cell = ws.cell(row=r, column=i, value=arr)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=2 + len(cats), value=total)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, r_start + 1, r - 1, len(headers))

    # Priority tier
    r_start = r + 2
    ws.cell(
        row=r_start - 1, column=1, value="Priority tier (from PI list view)"
    ).font = BODY_BOLD
    priorities = sorted({str(x.get("Priority") or "") for x in pi_rows} - {""})
    headers = ["Director"] + list(priorities) + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r_start, column=i, value=h)
    apply_header(ws, r_start, len(headers))
    r = r_start + 1
    for d in directors:
        rows_d = [x for x in pi_rows if x.get("Director") == d]
        ws.cell(row=r, column=1, value=d).alignment = LEFT
        total = 0
        for i, p in enumerate(priorities, 2):
            n = sum(1 for x in rows_d if str(x.get("Priority") or "") == p)
            total += n
            ws.cell(row=r, column=i, value=n).alignment = CENTER
        ws.cell(row=r, column=2 + len(priorities), value=total).alignment = CENTER
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, r_start + 1, r - 1, len(headers))
    set_widths(ws, [20] + [14] * 8)


# ──────────────── Dashboard roll-up tab ────────────────


def build_dashboard_overview(wb, report_totals):
    ws = wb.create_sheet("Dashboard Overview", 0)
    ws["A1"] = "Dashboard Roll-up"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = f"Data as of {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = CAPTION
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 28

    headers = [
        "Source Dashboard",
        "Widget",
        "Report ID",
        "Format",
        "Record Count",
        "Headline Total",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    apply_header(ws, 4, len(headers))
    r = 5
    for widget, rid, fmt, count, total, source in report_totals:
        ws.cell(row=r, column=1, value=source).alignment = LEFT
        ws.cell(row=r, column=2, value=widget).alignment = LEFT
        ws.cell(row=r, column=3, value=rid).alignment = LEFT
        ws.cell(row=r, column=4, value=fmt).alignment = CENTER
        ws.cell(row=r, column=5, value=count).alignment = CENTER
        ws.cell(row=r, column=6, value=total).alignment = LEFT
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, 5, r - 1, len(headers))
    set_widths(ws, [24, 38, 22, 10, 14, 30])
    ws.freeze_panes = "A5"


def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    ws["A1"] = "Methodology"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        f"Refreshed {datetime.now().strftime('%d %B %Y')} from live Salesforce data."
    )
    ws["A2"].font = CAPTION
    ws["A2"].alignment = LEFT

    notes = [
        ("", ""),
        (
            "Dashboard source",
            (
                f"Sales Directors Monthly Pipeline and Insights "
                f"({DASHBOARD_ID}). Each widget tab holds the raw rows "
                "returned by the underlying report. The Dashboard Overview "
                "tab lists all widgets with headline counts and totals."
            ),
        ),
        ("", ""),
        (
            f"{PERIOD.prior_quarter.label} history",
            (
                "Pulled from OpportunityFieldHistory for the CloseDate and "
                "StageName fields, starting October 2025. One row per change "
                f"event. Raw events live on {PERIOD.prior_quarter.label} History "
                "Raw; analytical summaries follow on the next three tabs."
            ),
        ),
        ("", ""),
        (
            f"{PERIOD.prior_quarter.label} slip definition",
            (
                f"An opportunity that had a CloseDate in "
                f"{PERIOD.prior_quarter.title} at some point "
                "and has since been pushed to a later date. The "
                f"{PERIOD.prior_quarter.label} Slips Pivot "
                "counts only deals still open today; deals that have since "
                "closed won or lost are excluded."
            ),
        ),
        ("", ""),
        (
            "Stage transition matrix",
            (
                "Every StageName change on any opportunity whose close date "
                f"ever sat in {PERIOD.prior_quarter.title}. Green cells are "
                "forward motion, red are regressions. Use this to spot "
                "widespread slippage between specific stages."
            ),
        ),
        ("", ""),
        (
            "Pipeline Inspection",
            (
                "Consolidated from each director workbook. PI Summary shows "
                "three analytical cuts: push-count buckets, forecast category "
                "mix by director, and priority tier distribution. ARR on the "
                "PI tabs is weighted (APTS_Forecast_ARR__c)."
            ),
        ),
        ("", ""),
        (
            "Common filters",
            (
                "Accounts matching simcorp, test or delete are excluded. "
                "Owners named Sabiniewicz or Profit are excluded. Only Land, "
                "Expand and Renewal types are considered."
            ),
        ),
        ("", ""),
        (
            "Refresh",
            (
                "Reproduce this file with "
                "scripts/build_dashboard_analysis_excel.py. Re-running the "
                "script reads Salesforce live and rewrites every tab."
            ),
        ),
    ]
    r = 4
    for label, body in notes:
        if label:
            ws.cell(row=r, column=1, value=label).font = BODY_BOLD
            ws.cell(row=r, column=2, value=body).font = BODY
            ws.cell(row=r, column=2).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            ws.row_dimensions[r].height = max(28, 16 * (len(body) // 80 + 1))
        r += 1
    set_widths(ws, [22, 90])


def _load_open_pipeline_all():
    """Union of Pipeline Open rows across all director workbooks, tagged with director."""
    SN = sheet_names()
    out = []
    for wb_file in sorted(WORKBOOKS_DIR.glob("*.xlsx")):
        name = wb_file.stem.replace("-", " ").title()
        try:
            wb = load_workbook(wb_file, data_only=True)
            if SN["pipeline_open"] not in wb.sheetnames:
                continue
            ws = wb[SN["pipeline_open"]]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {headers[i]: v for i, v in enumerate(row)}
                rec["_director"] = name
                out.append(rec)
        except Exception as exc:
            print(f"  open-pipeline load error {name}: {exc}")
    return out


def _load_won_lost_all():
    SN = sheet_names()
    out = []
    for wb_file in sorted(WORKBOOKS_DIR.glob("*.xlsx")):
        name = wb_file.stem.replace("-", " ").title()
        try:
            wb = load_workbook(wb_file, data_only=True)
            if SN["won_lost"] not in wb.sheetnames:
                continue
            ws = wb[SN["won_lost"]]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {headers[i]: v for i, v in enumerate(row)}
                rec["_director"] = name
                out.append(rec)
        except Exception as exc:
            print(f"  won-lost load error {name}: {exc}")
    return out


def _territory_of(rec):
    """Infer Rebekka-style region label from Account Unit Group / Sales Region."""
    aug = rec.get("Account_Unit_Group__c") or rec.get("Account Unit Group") or ""
    sr = rec.get("Sales_Region__c") or rec.get("Sales Region") or ""
    if aug == "SC Asia" or sr == "APAC":
        return "APAC"
    if aug == "SC North America":
        return f"NA {sr}".strip()
    if aug == "SC EMEA":
        return f"EMEA {sr}".strip()
    return aug or sr or "Unspecified"


# #1: Stage-to-stage conversion rate
def build_stage_conversion(wb, history_rows):
    """For each stage, count deals that entered it and how many later reached
    a subsequent stage. Conversion rate = reached next stage / entered."""
    stages_fwd = [
        "1 - Prospecting",
        "2 - Discovery",
        "3 - Engagement",
        "4 - Shortlisted",
        "5 - Preferred",
        "6 - Contracting",
        "8 - Won",
    ]
    # Build stage history per opp (ordered by changed_on)
    per_opp = defaultdict(list)
    for r in history_rows:
        if r["field"] == "StageName":
            per_opp[r["opp_id"]].append(
                (r["changed_on"], r["old_value"], r["new_value"])
            )
    # Sort events per opp
    for oid in per_opp:
        per_opp[oid].sort(key=lambda x: x[0])

    # For each opp, list stages ever entered (ordered): start with earliest old_value
    # then each new_value. Deduplicate in order.
    entered_any = defaultdict(set)
    for oid, events in per_opp.items():
        seen = set()
        first_old = events[0][1] if events else None
        if first_old:
            seen.add(first_old)
        for _, _old, new in events:
            if new:
                seen.add(new)
        entered_any[oid] = seen

    # Tally: for each stage X, count opps that entered X, and how many reached X+1+
    rows = []
    for i, stage in enumerate(stages_fwd[:-1]):
        later = stages_fwd[i + 1 :]
        entered_n = sum(1 for s in entered_any.values() if stage in s)
        advanced_n = sum(
            1
            for s in entered_any.values()
            if stage in s and any(ns in s for ns in later)
        )
        to_won_n = sum(1 for s in entered_any.values() if stage in s and "8 - Won" in s)
        rate = (advanced_n / entered_n * 100) if entered_n else 0
        won_rate = (to_won_n / entered_n * 100) if entered_n else 0
        rows.append(
            [stage, entered_n, advanced_n, f"{rate:.1f}%", to_won_n, f"{won_rate:.1f}%"]
        )

    ws = wb.create_sheet("Stage Conversion")
    ws["A1"] = f"Stage Conversion Rates, {PERIOD.fiscal_year} to date"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = [
        "Stage",
        "Deals Entered",
        "Advanced to Next+",
        "Advance Rate",
        "Reached Won",
        "Win Rate",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for row in rows:
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = CENTER if ci > 1 else LEFT
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22, 14, 18, 14, 14, 12])
    ws.freeze_panes = "A4"


# #2: Time in stage (median days)
def build_time_in_stage(wb, history_rows):
    """Median and average days spent in each stage, computed from stage
    transitions. Only closed transitions are measured (stage X → stage Y
    gives the duration X occupied)."""
    from datetime import datetime as _dt

    stage_order = [
        "1 - Prospecting",
        "2 - Discovery",
        "3 - Engagement",
        "4 - Shortlisted",
        "5 - Preferred",
        "6 - Contracting",
    ]

    def _parse(d):
        try:
            return _dt.strptime(str(d)[:10], "%Y-%m-%d")
        except Exception:
            return None

    # Build per-opp event list
    per_opp = defaultdict(list)
    for r in history_rows:
        if r["field"] == "StageName":
            dt = _parse(r["changed_on"])
            if dt:
                per_opp[r["opp_id"]].append((dt, r["old_value"], r["new_value"]))
    durations = defaultdict(list)  # stage -> list of days
    for oid, events in per_opp.items():
        events.sort(key=lambda x: x[0])
        for i in range(len(events) - 1):
            # Stage old_value was occupied from events[i].dt to events[i+1].dt?
            # Actually: each event marks a transition from old to new.
            # So stage new_value starts at events[i].dt and ends at events[i+1].dt.
            start_dt = events[i][0]
            end_dt = events[i + 1][0]
            stage_occupied = events[i][2]  # new at this event
            if stage_occupied:
                days = (end_dt - start_dt).days
                if 0 <= days <= 800:
                    durations[stage_occupied].append(days)
    # Summary
    import statistics

    rows = []
    for s in stage_order:
        d = durations.get(s, [])
        if d:
            rows.append(
                [s, len(d), int(statistics.median(d)), int(sum(d) / len(d)), max(d)]
            )
        else:
            rows.append([s, 0, 0, 0, 0])

    ws = wb.create_sheet("Time in Stage")
    ws["A1"] = "Time in Stage, Days"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = [
        "Stage",
        "Completed Transitions",
        "Median Days",
        "Average Days",
        "Max Days",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for row in rows:
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = CENTER if ci > 1 else LEFT
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22, 20, 14, 14, 14])
    ws.freeze_panes = "A4"


# #6: Push intensity buckets by territory
def build_push_intensity(wb, open_rows):
    """Distribution of push count per open deal, by territory."""
    buckets = [
        ("0", lambda n: n == 0),
        ("1-2", lambda n: 1 <= n <= 2),
        ("3-4", lambda n: 3 <= n <= 4),
        ("5+", lambda n: n >= 5),
    ]
    by_terr = defaultdict(lambda: [0] * len(buckets))
    arr_by_terr = defaultdict(lambda: [0.0] * len(buckets))
    for r in open_rows:
        terr = r.get("Sales Region") or r.get("_director", "Unspecified")
        n = int(r.get("Push Count") or 0)
        arr = float(r.get("ARR Unweighted (EUR)") or 0)
        for i, (_, pred) in enumerate(buckets):
            if pred(n):
                by_terr[terr][i] += 1
                arr_by_terr[terr][i] += arr

    ws = wb.create_sheet("Push Intensity")
    ws["A1"] = "Push Count Distribution, Open Pipeline"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = (
        ["Region"]
        + [f"{b[0]} Count" for b in buckets]
        + [f"{b[0]} ARR Unwtd" for b in buckets]
        + ["Total Deals"]
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for terr in sorted(by_terr.keys()):
        counts = by_terr[terr]
        arrs = arr_by_terr[terr]
        ws.cell(row=r, column=1, value=terr).alignment = LEFT
        for i, n in enumerate(counts, 2):
            ws.cell(row=r, column=i, value=n).alignment = CENTER
        for i, a in enumerate(arrs, 2 + len(buckets)):
            cell = ws.cell(row=r, column=i, value=a)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
        ws.cell(
            row=r, column=2 + 2 * len(buckets), value=sum(counts)
        ).alignment = CENTER
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22] + [10] * len(buckets) + [14] * len(buckets) + [12])
    ws.freeze_panes = "B4"


# #7: Top deal concentration
def build_concentration(wb, open_rows):
    """Pipeline concentration: top-N deals, top owners, by territory."""

    # Sort all open deals by ARR desc
    def arr_of(r):
        return float(r.get("ARR Unweighted (EUR)") or 0)

    by_terr = defaultdict(list)
    for r in open_rows:
        terr = r.get("Sales Region") or r.get("_director", "Unspecified")
        by_terr[terr].append(r)

    ws = wb.create_sheet("Concentration")
    ws["A1"] = "Pipeline Concentration"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = [
        "Region",
        "Total Deals",
        "Total ARR",
        "Top 10 ARR",
        "Top 10 %",
        "Top 25% Deals ARR",
        "Top 25% %",
        "Max Owner ARR",
        "Max Owner",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for terr in sorted(by_terr.keys()):
        deals = sorted(by_terr[terr], key=arr_of, reverse=True)
        total = sum(arr_of(d) for d in deals)
        top10 = sum(arr_of(d) for d in deals[:10])
        top25pct_cnt = max(1, int(len(deals) * 0.25))
        top25 = sum(arr_of(d) for d in deals[:top25pct_cnt])
        owner_totals = defaultdict(float)
        for d in deals:
            owner_totals[d.get("Owner", "?")] += arr_of(d)
        max_owner = max(owner_totals, key=owner_totals.get, default="")
        max_owner_arr = owner_totals.get(max_owner, 0.0)
        ws.cell(row=r, column=1, value=terr).alignment = LEFT
        ws.cell(row=r, column=2, value=len(deals)).alignment = CENTER
        for i, val in enumerate([total, top10], 3):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
        ws.cell(
            row=r, column=5, value=f"{(top10 / total * 100 if total else 0):.0f}%"
        ).alignment = CENTER
        cell = ws.cell(row=r, column=6, value=top25)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        ws.cell(
            row=r, column=7, value=f"{(top25 / total * 100 if total else 0):.0f}%"
        ).alignment = CENTER
        cell = ws.cell(row=r, column=8, value=max_owner_arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        ws.cell(row=r, column=9, value=max_owner).alignment = LEFT
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22, 10, 16, 16, 10, 16, 10, 16, 24])
    ws.freeze_panes = "A4"


# #16: Loss reason mix
def build_loss_reasons(wb, won_lost_rows):
    """Crosstab of loss reasons by region. Only closed-lost, not won."""
    losses = [r for r in won_lost_rows if "Won" not in str(r.get("Stage") or "")]
    # Pivot
    reasons = defaultdict(lambda: defaultdict(lambda: {"n": 0, "arr": 0.0}))
    all_reasons = set()
    for r in losses:
        terr = r.get("Sales Region") or r.get("_director", "Unspecified")
        reason = str(r.get("Reason") or "(not recorded)")
        all_reasons.add(reason)
        reasons[terr][reason]["n"] += 1
        reasons[terr][reason]["arr"] += float(r.get("ARR Unweighted (EUR)") or 0)

    ws = wb.create_sheet("Loss Reasons")
    ws["A1"] = "Closed Lost by Reason and Region"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    reason_list = sorted(all_reasons)
    headers = ["Region"] + reason_list + ["Total Losses", "Total Lost ARR"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for terr in sorted(reasons.keys()):
        ws.cell(row=r, column=1, value=terr).alignment = LEFT
        total_n = 0
        total_arr = 0.0
        for i, rn in enumerate(reason_list, 2):
            n = reasons[terr][rn]["n"]
            total_n += n
            total_arr += reasons[terr][rn]["arr"]
            ws.cell(row=r, column=i, value=n or "").alignment = CENTER
        ws.cell(row=r, column=2 + len(reason_list), value=total_n).alignment = CENTER
        cell = ws.cell(row=r, column=3 + len(reason_list), value=total_arr)
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22] + [16] * len(reason_list) + [12, 18])
    ws.freeze_panes = "B4"


# #17: Stage at loss
def build_stage_at_loss(wb, history_rows, won_lost_rows):
    """For each lost deal, what was the highest stage it ever reached?"""
    lost_ids = {
        r.get("Opportunity")
        for r in won_lost_rows
        if "Won" not in str(r.get("Stage") or "")
    }
    # Map opp name -> highest stage from history
    stage_order = [
        "1 - Prospecting",
        "2 - Discovery",
        "3 - Engagement",
        "4 - Shortlisted",
        "5 - Preferred",
        "6 - Contracting",
    ]

    def _level(s):
        try:
            return stage_order.index(s)
        except ValueError:
            return -1

    highest = defaultdict(int)
    opp_to_terr = {}
    opp_to_arr = {}
    for r in won_lost_rows:
        if r.get("Opportunity") in lost_ids:
            opp_to_terr[r.get("Opportunity")] = r.get("Sales Region") or r.get(
                "_director", "Unspecified"
            )
            opp_to_arr[r.get("Opportunity")] = float(r.get("ARR Unweighted (EUR)") or 0)
    # Walk history to find each lost deal's peak stage
    opp_id_to_name = {}
    for r in history_rows:
        opp_id_to_name[r["opp_id"]] = r["opportunity"]
    for r in history_rows:
        if r["field"] != "StageName":
            continue
        name = r["opportunity"]
        if name not in lost_ids:
            continue
        for val in (r["old_value"], r["new_value"]):
            lvl = _level(val)
            if lvl > highest.get(name, -1):
                highest[name] = lvl
    # If no history, fall back to current stage in won_lost sheet
    for r in won_lost_rows:
        name = r.get("Opportunity")
        if name not in lost_ids:
            continue
        stage = str(r.get("Stage") or "")
        lvl = _level(stage)
        if name not in highest and lvl >= 0:
            highest[name] = lvl

    # Pivot: Region x Stage
    grid = defaultdict(lambda: [0] * len(stage_order))
    arr_grid = defaultdict(lambda: [0.0] * len(stage_order))
    for name, lvl in highest.items():
        if lvl < 0:
            continue
        terr = opp_to_terr.get(name, "Unspecified")
        grid[terr][lvl] += 1
        arr_grid[terr][lvl] += opp_to_arr.get(name, 0)

    ws = wb.create_sheet("Stage at Loss")
    ws["A1"] = "Stage Reached Before Losing, by Region"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    headers = (
        ["Region"]
        + [f"{s} #" for s in stage_order]
        + [f"{s} ARR" for s in stage_order]
        + ["Total Count", "Total Lost ARR"]
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    apply_header(ws, 3, len(headers))
    r = 4
    for terr in sorted(grid.keys()):
        ws.cell(row=r, column=1, value=terr).alignment = LEFT
        counts = grid[terr]
        arrs = arr_grid[terr]
        for i, n in enumerate(counts, 2):
            ws.cell(row=r, column=i, value=n or "").alignment = CENTER
        for i, a in enumerate(arrs, 2 + len(stage_order)):
            cell = ws.cell(row=r, column=i, value=a)
            cell.alignment = RIGHT
            cell.number_format = "#,##0"
        ws.cell(
            row=r, column=2 + 2 * len(stage_order), value=sum(counts)
        ).alignment = CENTER
        cell = ws.cell(row=r, column=3 + 2 * len(stage_order), value=sum(arrs))
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).font = BODY
            ws.cell(row=r, column=ci).border = BORDER
        r += 1
    zebra(ws, 4, r - 1, len(headers))
    set_widths(ws, [22] + [12] * len(stage_order) + [14] * len(stage_order) + [10, 16])
    ws.freeze_panes = "B4"


def build_stage_transition_matrix(wb, history_rows):
    """From-stage x to-stage heatmap of Q1 transitions.

    Answers: Where do deals get stuck (self-loops)? Which forward transitions
    are rare vs common? Which backward transitions (demotions) are happening?
    """
    from openpyxl.formatting.rule import ColorScaleRule

    ws = wb.create_sheet("Stage Transition Matrix")
    ws["A1"] = (
        f"Stage Transition Matrix, {PERIOD.prior_quarter.label} {PERIOD.fiscal_year}"
    )
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = (
        "Rows = From Stage, Columns = To Stage. Cell = count of unique opportunities that made "
        "that transition during Q1. Heatmap: white (none) -> yellow -> red (high volume). "
        "Backward transitions (right-to-left of diagonal) signal deal demotions."
    )
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")

    stage_events = [r for r in history_rows if r.get("field") == "StageName"]
    matrix = defaultdict(int)
    unique_pairs = defaultdict(set)
    stages = set()
    for e in stage_events:
        frm = str(e.get("old_value", "") or "").strip()
        to = str(e.get("new_value", "") or "").strip()
        if not frm or not to or frm == to:
            continue
        unique_pairs[(frm, to)].add(e.get("opp_id", ""))
        stages.add(frm)
        stages.add(to)
    for pair, opps in unique_pairs.items():
        matrix[pair] = len(opps)
    all_stages = sorted(stages)

    row = 4
    c = ws.cell(row=row, column=1, value="From \\ To")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    for ci, stage in enumerate(all_stages, 2):
        c = ws.cell(row=row, column=ci, value=stage)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c = ws.cell(row=row, column=len(all_stages) + 2, value="Total From")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 48

    row += 1
    first_data_row = row
    col_totals = [0] * len(all_stages)
    for frm in all_stages:
        c = ws.cell(row=row, column=1, value=frm)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = LEFT
        c.border = BORDER
        row_total = 0
        for ci, to in enumerate(all_stages, 2):
            v = matrix.get((frm, to), 0)
            cc = ws.cell(row=row, column=ci, value=v if v else None)
            cc.alignment = CENTER
            cc.border = BORDER
            cc.font = BODY
            row_total += v
            col_totals[ci - 2] += v
        tc = ws.cell(row=row, column=len(all_stages) + 2, value=row_total)
        tc.alignment = CENTER
        tc.border = BORDER
        tc.font = Font(name="Calibri", size=10, bold=True)
        row += 1
    last_data_row = row - 1

    # Total To row
    c = ws.cell(row=row, column=1, value="Total To")
    c.font = Font(name="Calibri", size=10, bold=True)
    c.border = BORDER
    for ci, t in enumerate(col_totals, 2):
        cc = ws.cell(row=row, column=ci, value=t if t else None)
        cc.font = Font(name="Calibri", size=10, bold=True)
        cc.alignment = CENTER
        cc.border = BORDER
    tt = ws.cell(row=row, column=len(all_stages) + 2, value=sum(col_totals))
    tt.font = Font(name="Calibri", size=10, bold=True)
    tt.alignment = CENTER
    tt.border = BORDER

    if all_stages and last_data_row >= first_data_row:
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(len(all_stages) + 1)
        rule = ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="F8696B",
        )
        ws.conditional_formatting.add(
            f"{first_col_letter}{first_data_row}:{last_col_letter}{last_data_row}",
            rule,
        )

    ws.column_dimensions["A"].width = 22
    for ci in range(2, len(all_stages) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = f"B{first_data_row}"


def main():
    global WORKBOOKS_DIR
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbooks-dir",
        type=Path,
        default=WORKBOOKS_DIR,
        help="Directory containing the extracted director workbooks used for PI consolidation.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output/sharepoint/Dashboard and Q1 Analysis.xlsx"),
    )
    args = parser.parse_args()

    WORKBOOKS_DIR = args.workbooks_dir
    args.output.parent.mkdir(parents=True, exist_ok=True)
    token, instance = get_auth()
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    # 1) Enumerate components from both dashboards
    reports_seen = {}
    ordered_widgets = []  # list of (widget, rid, source_dashboard_label)
    for dash_id, dash_label in [
        (DASHBOARD_ID, "Sales Directors Monthly"),
        (SALES_OPS_DASHBOARD_ID, "Sales Ops Quarterly KPI"),
    ]:
        dash = session.get(
            f"{instance}/services/data/v66.0/analytics/dashboards/{dash_id}/describe"
        ).json()
        for c in dash.get("components", []):
            rid = c.get("reportId")
            if not rid or rid in reports_seen:
                continue
            widget = c.get("header") or "(untitled)"
            reports_seen[rid] = widget
            ordered_widgets.append((widget, rid, dash_label))

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    # 2) Raw tabs per dashboard widget
    print(f"Pulling {len(ordered_widgets)} dashboard reports...")
    report_totals = []
    for widget, rid, source in ordered_widgets:
        print(f"  [{source[:17]}] {widget} ({rid})...")
        try:
            cols, labels, dtypes, rows, total_agg, fmt = run_report(
                session, instance, rid
            )
        except Exception as e:
            print(f"    error: {e}")
            continue
        # Build summary number
        headline = ""
        count = len(rows)
        if total_agg:
            headline = " | ".join(f"{a.get('label', '')}" for a in total_agg)
        report_totals.append((widget, rid, fmt, count, headline, source))

        # Sheet name: sanitise to <= 31 chars
        name = widget.replace("/", " ").replace(":", "")[:31]
        if name in wb.sheetnames:
            name = (name + " " + rid[-4:])[:31]
        ws = wb.create_sheet(name)
        subtitle = (
            f"Source report {rid}. Rows: {count}. {headline}"
            if headline
            else f"Source report {rid}. Rows: {count}."
        )
        write_raw_report_tab(ws, widget, subtitle, labels, dtypes, rows)

    # 3) Q1 History tabs
    print("Pulling Q1 field history...")
    history = fetch_q1_history(session, instance)
    print(f"  {len(history)} events")
    build_q1_history_raw(wb, history)
    build_q1_slips_pivot(wb, history)
    build_q1_slips_by_owner(wb, history)
    build_q1_changes_by_region(wb, history)
    build_stage_progression(wb, history)

    # 4) Deeper analytics
    print("Loading all director workbooks for cross-tab analyses...")
    open_rows = _load_open_pipeline_all()
    won_lost_rows = _load_won_lost_all()

    build_stage_conversion(wb, history)
    build_time_in_stage(wb, history)
    build_push_intensity(wb, open_rows)
    build_concentration(wb, open_rows)
    build_loss_reasons(wb, won_lost_rows)
    build_stage_at_loss(wb, history, won_lost_rows)
    build_stage_transition_matrix(wb, history)

    # 5) PI consolidation
    print("Consolidating Pipeline Inspection...")
    pi_rows = build_pi_consolidated(wb)
    build_pi_summary(wb, pi_rows)

    # 5) Overview + Methodology
    build_dashboard_overview(wb, report_totals)
    build_methodology(wb)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")


if __name__ == "__main__":
    main()
