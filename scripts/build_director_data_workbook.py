#!/usr/bin/env python3
"""Build an Excel data workbook with pivot-ready sheets for a director.

Reads the workbook snapshot (already territory-scoped), applies the internal
account filter, and writes clean data sheets. Pivot tables are configured
as Excel Table objects so the user can insert PivotTables/PivotCharts manually
in Excel — openpyxl cannot create native PivotTables, but the data is
structured for one-click pivot creation.

Usage:
    python3 scripts/build_director_data_workbook.py \
        --snapshot output/director_workbook_snapshots/2026-04-10/dan-peppett.json \
        --pi-snapshot output/pipeline_inspection_snapshots/2026-04-10/uk-ireland.json \
        --external-inputs-date 2026-04-10 \
        [--output output/director_data_workbooks/2026-04-10/dan-peppett.xlsx]
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_validated_director_brief import (
    _clean_snapshot,
    _is_internal_account,
    as_number,
    as_text,
    load_snapshot,
)

HEADER_FILL = PatternFill(start_color="083EA7", end_color="083EA7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=9)
EUR_FMT = "#,##0"


def _add_sheet(
    wb: Workbook,
    name: str,
    headers: list[str],
    rows: list[list],
    eur_cols: list[int] | None = None,
):
    """Add a sheet with headers, data, and an Excel Table."""
    ws = wb.create_sheet(title=name[:31])

    # Headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            if eur_cols and ci in eur_cols and isinstance(val, (int, float)):
                cell.number_format = EUR_FMT

    # Auto-width
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(headers[ci - 1])),
            *(len(str(r[ci - 1])) for r in rows[:50]) if rows else [0],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Add Excel Table for pivot-ready data
    if rows:
        end_col = get_column_letter(len(headers))
        end_row = len(rows) + 1
        table_name = name.replace(" ", "_").replace("-", "_").replace("&", "And")[:30]
        table = Table(
            displayName=table_name,
            ref=f"A1:{end_col}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
        )
        ws.add_table(table)

    # Freeze header row
    ws.freeze_panes = "A2"
    return ws


def build_workbook(
    snapshot_path: Path,
    output_path: Path,
    pi_path: Path | None = None,
    ext_date: str | None = None,
):
    snap = _clean_snapshot(load_snapshot(snapshot_path))
    director = snap.get("director_name", "Director")
    territory = snap.get("territory", "")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Pipeline Detail ---
    records = (snap.get("pipeline_detail") or {}).get("records") or []
    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "ARR",
        "Forecast ARR",
        "Forecast Category",
        "Probability %",
        "Type",
        "Push Count",
        "Age Days",
        "Days In Stage",
        "Next Step",
    ]
    rows = []
    for r in records:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_text(r.get("Close Date")),
                as_number(r.get("ARR (€ converted)")),
                as_number(r.get("Forecast ARR (€ converted)")),
                as_text(r.get("Forecast Category")),
                as_number(r.get("Probability (%)")),
                as_text(r.get("Type")),
                int(as_number(r.get("Push Count"))),
                int(as_number(r.get("Age (Days)"))),
                int(as_number(r.get("Days In Stage"))),
                as_text(r.get("Next Step")),
            ]
        )
    _add_sheet(wb, "Pipeline Detail", headers, rows, eur_cols=[6, 7])

    # --- Renewals ---
    ren_rows = (snap.get("renewals") or {}).get("open_renewals") or []
    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "Close Date",
        "Renewal ACV",
        "Risk Level",
    ]
    rows = []
    for r in ren_rows:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_text(r.get("Close Date")),
                as_number(r.get("Renewal ACV (€ converted)")),
                as_text(r.get("Risk Level")),
            ]
        )
    _add_sheet(wb, "Renewals", headers, rows, eur_cols=[6])

    # --- Commercial Approval ---
    ca = snap.get("commercial_approval") or {}
    missing = ca.get("missing_candidates") or []
    approved = ca.get("approved_ytd") or []
    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR",
        "Status",
        "Approval Date",
    ]
    rows = []
    for r in approved:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_number(r.get("ARR (€ converted)")),
                "Approved 2026",
                as_text(r.get("Approval Date")),
            ]
        )
    for r in missing:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_number(r.get("ARR (€ converted)")),
                "Missing Approval",
                "",
            ]
        )
    _add_sheet(wb, "Commercial Approval", headers, rows, eur_cols=[5])

    # --- Won Lost ---
    wl = snap.get("won_lost") or {}
    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR",
        "Outcome",
        "Reason",
        "Competitor",
        "Close Date",
    ]
    rows = []
    for r in wl.get("won") or []:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_number(r.get("ARR (€ converted)")),
                "Won",
                as_text(r.get("Reason Won/Lost")),
                as_text(r.get("Lost to Competitor")),
                as_text(r.get("Close Date")),
            ]
        )
    for r in wl.get("lost") or []:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_number(r.get("ARR (€ converted)")),
                "Lost",
                as_text(r.get("Reason Won/Lost")),
                as_text(r.get("Lost to Competitor")),
                as_text(r.get("Close Date")),
            ]
        )
    _add_sheet(wb, "Won Lost", headers, rows, eur_cols=[5])

    # --- Risk Register ---
    rr = (snap.get("risk_register") or {}).get("top_arr") or []
    headers = [
        "Account",
        "Opportunity",
        "Owner",
        "Stage",
        "ARR",
        "Push Count",
        "Activity Days Ago",
        "Days In Stage",
    ]
    rows = []
    for r in rr:
        rows.append(
            [
                as_text(r.get("Account")),
                as_text(r.get("Opportunity")),
                as_text(r.get("Owner")),
                as_text(r.get("Stage")),
                as_number(r.get("ARR (€ converted)")),
                int(as_number(r.get("Push Count"))),
                int(as_number(r.get("Activity Days Ago"))),
                int(as_number(r.get("Days In Stage"))),
            ]
        )
    _add_sheet(wb, "Risk Register", headers, rows, eur_cols=[5])

    # --- Data Quality ---
    dq_issues = (snap.get("data_quality") or {}).get("top_issues") or []
    headers = [
        "Rep",
        "Total Issues",
        "No Activity",
        "Overdue Close",
        "Missing Next Step",
        "Missing Amount",
        "Missing Approval",
    ]
    rows = []
    for r in dq_issues:
        rows.append(
            [
                as_text(r.get("Rep")),
                int(as_number(r.get("Total Issues"))),
                int(as_number(r.get("No Activity"))),
                int(as_number(r.get("Overdue Close"))),
                int(as_number(r.get("Missing Next Step"))),
                int(as_number(r.get("Missing Amount"))),
                int(as_number(r.get("Missing Approval"))),
            ]
        )
    _add_sheet(wb, "Data Quality", headers, rows)

    # --- Pipeline Inspection (if available) ---
    if pi_path and pi_path.exists():
        with open(pi_path) as f:
            pi = json.load(f)
        pi_records = pi.get("records") or []
        # Filter internal accounts from PI too
        pi_clean = [
            r
            for r in pi_records
            if not _is_internal_account(r.get("name", ""))
            and (r.get("close_date", "")[:4] <= "2026" if r.get("close_date") else True)
        ]
        pi_open = [
            r
            for r in pi_clean
            if not r.get("is_closed")
            and r.get("forecast_category") not in ("Omitted", "Closed")
        ]
        headers = [
            "Opportunity",
            "Owner",
            "Stage",
            "Forecast Category",
            "Forecast ARR",
            "Close Date",
            "Push Count",
            "Score",
            "Priority",
        ]
        rows = []
        for r in pi_open:
            rows.append(
                [
                    r.get("name", ""),
                    r.get("owner", ""),
                    r.get("stage", ""),
                    r.get("forecast_category", ""),
                    r.get("forecast_arr", 0),
                    r.get("close_date", ""),
                    r.get("push_count", 0),
                    r.get("score"),
                    "Yes" if r.get("is_priority") else "",
                ]
            )
        rows.sort(key=lambda x: -(x[4] or 0))
        _add_sheet(wb, "Pipeline Inspection", headers, rows, eur_cols=[5])

    # --- Summary sheet (first) ---
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = f"{director} ({territory})"
    ws["A1"].font = Font(bold=True, size=14, color="083EA7")
    ws["A2"] = f"Snapshot: {snap.get('snapshot_date', '')}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(size=10, color="666666")

    ws["A5"] = "Sheet"
    ws["B5"] = "Records"
    ws["A5"].font = HEADER_FONT
    ws["A5"].fill = HEADER_FILL
    ws["B5"].font = HEADER_FONT
    ws["B5"].fill = HEADER_FILL
    sheet_counts = [
        (s.title, s.max_row - 1 if s.max_row else 0)
        for s in wb.worksheets
        if s.title != "Summary"
    ]
    for i, (sname, count) in enumerate(sheet_counts, 6):
        ws[f"A{i}"] = sname
        ws[f"B{i}"] = count
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12

    wb.save(str(output_path))
    print(f"Saved: {output_path}")
    for sname, count in sheet_counts:
        print(f"  {sname:25s}  {count:>5d} rows")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--snapshot", type=Path, required=True)
    parser.add_argument("--pi-snapshot", type=Path, default=None)
    parser.add_argument("--external-inputs-date", default=None)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    if not args.snapshot.exists():
        print(f"ERROR: Snapshot not found: {args.snapshot}", file=sys.stderr)
        sys.exit(1)

    if args.output:
        out = args.output
    else:
        slug = args.snapshot.stem
        out = REPO_ROOT / "output" / "director_data_workbooks" / slug + ".xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(args.snapshot, out, args.pi_snapshot)


if __name__ == "__main__":
    main()
