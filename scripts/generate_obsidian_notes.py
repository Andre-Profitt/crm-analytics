"""
Generate Obsidian notes from a pipeline run.

Reads the manifest + director workbooks and writes:
  obsidian/Monthly/YYYY-MM/README.md        snapshot summary for the month
  obsidian/Monthly/YYYY-MM/<director>.md    per-director monthly note
  obsidian/Directors/<director>.md          standing director page (idempotent)

Called as the last stage of `run_monthly_director_review.py` or by hand.
"""

import argparse
import json
import re  # noqa: F401  -- used by _read_snapshot_trend
from collections import Counter, defaultdict  # noqa: F401
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
VAULT = ROOT / "obsidian"
SHAREPOINT = ROOT / "output" / "sharepoint"

DIRECTORS = [
    ("Jesper Tyrer", "APAC", "jesper-tyrer.xlsx"),
    ("Sarah Pittroff", "EMEA Central", "sarah-pittroff.xlsx"),
    ("Dan Peppett", "EMEA UK & Ireland", "dan-peppett.xlsx"),
    ("Christian Ebbesen", "EMEA NE", "christian-ebbesen.xlsx"),
    ("Francois Thaury", "EMEA South West", "francois-thaury.xlsx"),
    ("Mourad Essofi", "EMEA MEA", "mourad-essofi.xlsx"),
    ("Patrick Gaughan", "NA Asset Management", "patrick-gaughan.xlsx"),
    ("Megan Miceli", "NA Canada", "megan-miceli.xlsx"),
    ("Adam Steinhaus", "NA Pension & Insurance", "adam-steinhaus.xlsx"),
]

# Q1 Trend Consolidated uses short territory labels that differ from DIRECTORS.
TERRITORY_TO_DIRECTOR = {
    "APAC": "Jesper Tyrer",
    "EMEA Central": "Sarah Pittroff",
    "EMEA UK & Ireland": "Dan Peppett",
    "EMEA NE": "Christian Ebbesen",
    "EMEA South West": "Francois Thaury",
    "EMEA MEA": "Mourad Essofi",
    "NA Asset Mgmt": "Patrick Gaughan",
    "NA Canada": "Megan Miceli",
    "NA Insurance": "Adam Steinhaus",
}

# Stage rank for "highest stage reached before loss" lookups.
STAGE_RANK = {
    "1 - Prospecting": 1,
    "2 - Discovery": 2,
    "3 - Engagement": 3,
    "4 - Shortlisted": 4,
    "5 - Preferred": 5,
    "6 - Contracting": 6,
    "8 - Won": 8,
    "0 - Lost": -1,
    "0 - No Opportunity": -2,
}
STAGES_IN_ORDER = [
    "1 - Prospecting",
    "2 - Discovery",
    "3 - Engagement",
    "4 - Shortlisted",
    "5 - Preferred",
    "6 - Contracting",
]

Q2_RANGE = ("2026-04-01", "2026-06-30")


def _slug(name):
    """Convert a director display name into a kebab-case filename slug."""
    return name.lower().replace(" ", "-").replace(",", "").replace("'", "")


def _fmt(n):
    if n is None:
        return "-"
    if isinstance(n, (int, float)):
        if abs(n) >= 1_000_000:
            return f"EUR {n / 1_000_000:.1f}M"
        if abs(n) >= 1_000:
            return f"EUR {n / 1_000:.0f}K"
        return f"EUR {n:,.0f}"
    return str(n)


def _quarter(s):
    s = str(s or "")
    if not s.startswith("2026"):
        return None
    try:
        m = int(s[5:7])
    except ValueError:
        return None
    return f"Q{(m - 1) // 3 + 1}"


def _load_director(wb_path):
    wb = load_workbook(wb_path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: v for i, v in enumerate(r)})
        out[sn] = rows
    return out


def _director_stats(director, territory, wb_path):
    sheets = _load_director(wb_path)

    def is_land(r):
        return str(r.get("Type", "")).strip().lower() == "land"

    pipeline = sheets.get("Pipeline Open FY26", [])
    won_lost = sheets.get("Won Lost FY26", [])
    approvals = sheets.get("Commercial Approval", [])
    renewals = sheets.get("Renewals FY26", [])
    q1_mov = sheets.get("Q1 Movement", [])

    def in_q1q2(r):
        cd = str(r.get("Close Date", "") or "")[:10]
        return "2026-01-01" <= cd <= "2026-06-30"

    land_open = [r for r in pipeline if is_land(r)]
    # Deck scope is Land + Q1-Q2 close date; capture that slice separately so
    # the snapshot history ledger matches what the deck renders and MoM
    # deltas reconcile apples-to-apples.
    land_open_q1q2 = [r for r in land_open if in_q1q2(r)]
    # Approvals in deck scope also require Q1-Q2 close; keeps Missing Stage 3+
    # and Approved 2026 counts matching what the deck shows.
    approvals_in_scope = [r for r in approvals if in_q1q2(r)]
    land_wl = [r for r in won_lost if is_land(r)]
    q1_won = [
        r
        for r in land_wl
        if "Won" in str(r.get("Stage", "")) and _quarter(r.get("Close Date")) == "Q1"
    ]
    q1_lost = [
        r
        for r in land_wl
        if "Won" not in str(r.get("Stage", ""))
        and _quarter(r.get("Close Date")) == "Q1"
    ]

    stats = {
        "director": director,
        "territory": territory,
        "open_land_deals": len(land_open),
        "open_land_arr_unwtd": sum(
            float(r.get("ARR Unweighted (EUR)") or 0) for r in land_open
        ),
        "open_land_arr_wtd": sum(
            float(r.get("ARR Weighted (EUR)") or 0) for r in land_open
        ),
        # Q1-Q2 scoped slices matching the deck's filter.
        "open_land_q1q2_deals": len(land_open_q1q2),
        "open_land_q1q2_arr_unwtd": sum(
            float(r.get("ARR Unweighted (EUR)") or 0) for r in land_open_q1q2
        ),
        "open_land_q1q2_arr_wtd": sum(
            float(r.get("ARR Weighted (EUR)") or 0) for r in land_open_q1q2
        ),
        "q1_won_count": len(q1_won),
        "q1_won_arr": sum(float(r.get("ARR Unweighted (EUR)") or 0) for r in q1_won),
        "q1_lost_count": len(q1_lost),
        "q1_lost_arr": sum(float(r.get("ARR Unweighted (EUR)") or 0) for r in q1_lost),
        "approved_2026": sum(
            1
            for r in approvals_in_scope
            if str(r.get("Status", "")).strip() == "Approved 2026"
        ),
        "conditionally_approved": sum(
            1 for r in approvals_in_scope if "Pending" in str(r.get("Status", ""))
        ),
        "missing_approval": sum(
            1 for r in approvals_in_scope if "Missing" in str(r.get("Status", ""))
        ),
        "renewals_q2": sum(
            1
            for r in renewals
            if "2026-04" <= str(r.get("Close Date", ""))[:7] <= "2026-06"
        ),
        "q1_slip_events": len(q1_mov),
        "top_land_deals": sorted(
            land_open,
            key=lambda r: float(r.get("ARR Unweighted (EUR)") or 0),
            reverse=True,
        )[:5],
        "owner_push_concentration": Counter(
            r.get("Owner") for r in land_open if int(r.get("Push Count") or 0) >= 3
        ).most_common(3),
    }
    return stats


def _read_analytics_workbook():
    """Pull the analytics tabs from the consolidated FY26 Pipeline Review workbook.

    Returns a dict of findings (pivot, concentration, velocity, slip, scorecard,
    transitions). Missing tabs yield empty fields rather than errors so the notes
    still render when an earlier stage was skipped.
    """
    fy26 = SHAREPOINT / "FY26 Pipeline Review, All Territories.xlsx"
    dash = SHAREPOINT / "Dashboard and Q1 Analysis.xlsx"

    findings = {
        "top_deals": [],
        "concentration_summary": None,
        "pivot_stage_totals": {},
        "pivot_director_totals": [],
        "velocity_q1": {},
        "velocity_q2": {},
        "slip_top_owners": [],
        "scorecard": [],
        "transitions_top": [],
        "transitions_demotions": [],
        "deal_risk": [],
        "variance_by_director": {},
        "variance_totals": None,
        "thresholds": {},
    }

    if fy26.exists():
        wb = load_workbook(fy26, data_only=True, read_only=True)

        if "ARR Concentration" in wb.sheetnames:
            ws = wb["ARR Concentration"]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows[4:24]:
                if not r or r[0] is None or not isinstance(r[0], int):
                    continue
                findings["top_deals"].append(
                    {
                        "rank": r[0],
                        "director": r[1],
                        "account": r[2],
                        "opportunity": r[3],
                        "stage": r[4],
                        "owner": r[5],
                        "arr": r[6] or 0,
                        "pct": r[7] or 0,
                        "cum_pct": r[8] or 0,
                    }
                )
            for r in rows:
                if (
                    r
                    and isinstance(r[0], str)
                    and r[0].startswith("Top 20 account for")
                ):
                    findings["concentration_summary"] = r[0]
                    break

        if "Pipeline Pivot" in wb.sheetnames:
            ws = wb["Pipeline Pivot"]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[3] if len(rows) > 3 else []
            stages = [h for h in header[1:-1] if h]
            totals = {s: 0.0 for s in stages}
            dir_totals = []
            for r in rows[4:]:
                if not r or r[0] is None or r[0] == "TOTAL":
                    continue
                for si, stage in enumerate(stages, 1):
                    v = r[si] if si < len(r) else None
                    if isinstance(v, (int, float)):
                        totals[stage] += float(v)
                dir_totals.append((r[0], r[-1] or 0))
            findings["pivot_stage_totals"] = totals
            findings["pivot_director_totals"] = sorted(
                dir_totals, key=lambda x: -float(x[1] or 0)
            )

        if "Pipeline Velocity" in wb.sheetnames:
            ws = wb["Pipeline Velocity"]
            rows = list(ws.iter_rows(values_only=True))
            # Two blocks separated by a header "... ARR by Snapshot (EUR)"
            block_header_rows = []
            for i, r in enumerate(rows):
                if (
                    r
                    and isinstance(r[0], str)
                    and r[0].endswith("ARR by Snapshot (EUR)")
                ):
                    block_header_rows.append(i)
            for bi, header_row_idx in enumerate(block_header_rows):
                label = rows[header_row_idx][0]
                col_header = rows[header_row_idx + 1]
                dates = [c for c in col_header[1:] if c]
                data_rows = []
                i = header_row_idx + 2
                while i < len(rows):
                    r = rows[i]
                    if not r or r[0] is None:
                        break
                    if isinstance(r[0], str) and r[0].endswith("ARR by Snapshot (EUR)"):
                        break
                    director = r[0]
                    series = [r[j + 1] or 0 for j in range(len(dates))]
                    data_rows.append((director, series))
                    i += 1
                bucket = (
                    findings["velocity_q1"]
                    if "Q1" in label
                    else findings["velocity_q2"]
                )
                bucket["dates"] = dates
                bucket["directors"] = data_rows

        if "Slip Risk by Owner" in wb.sheetnames:
            ws = wb["Slip Risk by Owner"]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows[4:14]:
                if not r or r[0] is None:
                    continue
                findings["slip_top_owners"].append(
                    {
                        "owner": r[0],
                        "director": r[1],
                        "deals": r[2] or 0,
                        "arr": r[3] or 0,
                        "pushes": r[4] or 0,
                        "avg_push": r[5] or 0,
                        "max_push": r[6] or 0,
                    }
                )

        if "Parameters" in wb.sheetnames:
            ws = wb["Parameters"]
            for r in ws.iter_rows(values_only=True):
                if not r or r[0] is None:
                    continue
                name = str(r[0])
                if name.startswith("Thresh_") or name.startswith("RiskWeight_"):
                    findings["thresholds"][name] = r[1]

        if "Deal Risk Scoring" in wb.sheetnames:
            ws = wb["Deal Risk Scoring"]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows[4:]:
                if not r or r[0] is None or not isinstance(r[0], int):
                    continue
                findings["deal_risk"].append(
                    {
                        "rank": r[0],
                        "score": r[1] or 0,
                        "director": r[2] or "",
                        "account": r[3] or "",
                        "opportunity": r[4] or "",
                        "stage": r[5] or "",
                        "owner": r[6] or "",
                        "close_date": str(r[7] or "")[:10],
                        "days_to_close": r[8],
                        "days_since_activity": r[9],
                        "arr": r[10] or 0,
                        "pushes": r[11] or 0,
                        "reason_codes": r[12] or "",
                        "proof": r[13] or "",
                    }
                )

        # Forecast Variance cells are SUMIFS formulas that openpyxl cannot
        # evaluate (read_only data_only returns None). Aggregate directly from
        # Q1 Trend Consolidated's Bucket helper column instead.
        if "Q1 Trend Consolidated" in wb.sheetnames:
            ws = wb["Q1 Trend Consolidated"]
            rows = list(ws.iter_rows(values_only=True))
            # Row 1 is header (Territory, Account, ..., Initial ARR, Final ARR,
            # Initial Stage, Final Stage, Bucket).
            by_dir = {}
            total = {
                "initial": 0.0,
                "final": 0.0,
                "delta": 0.0,
                "Won": 0.0,
                "Lost": 0.0,
                "Added": 0.0,
                "RevisedUp": 0.0,
                "RevisedDown": 0.0,
            }
            for r in rows[2:]:
                if not r or r[0] is None:
                    continue
                territory = str(r[0])
                director = TERRITORY_TO_DIRECTOR.get(territory)
                if not director:
                    continue
                initial = float(r[20] or 0)  # type: ignore[arg-type]
                final = float(r[21] or 0)  # type: ignore[arg-type]
                bucket = str(r[24]) if r[24] is not None else ""
                agg = by_dir.setdefault(
                    director,
                    {
                        "territory": territory,
                        "initial": 0.0,
                        "final": 0.0,
                        "delta": 0.0,
                        "Won": 0.0,
                        "Lost": 0.0,
                        "Added": 0.0,
                        "RevisedUp": 0.0,
                        "RevisedDown": 0.0,
                    },
                )
                # Match build_sharepoint_analysis.py SUMIFS rules exactly:
                # - Initial: sum where Bucket != AlreadyClosed
                # - Final: sum where Bucket not in (Won, Lost, AlreadyClosed)
                # - Won/Lost: sum Initial for matching bucket
                # - Added: sum Final for matching bucket
                # - RevisedUp: Final - Initial for matching bucket
                # - RevisedDown: Initial - Final for matching bucket
                if bucket != "AlreadyClosed":
                    agg["initial"] += initial
                    total["initial"] += initial
                if bucket not in ("Won", "Lost", "AlreadyClosed"):
                    agg["final"] += final
                    total["final"] += final
                if bucket == "Won":
                    agg["Won"] += initial
                    total["Won"] += initial
                elif bucket == "Lost":
                    agg["Lost"] += initial
                    total["Lost"] += initial
                elif bucket == "Added":
                    agg["Added"] += final
                    total["Added"] += final
                elif bucket == "RevisedUp":
                    diff = final - initial
                    agg["RevisedUp"] += diff
                    total["RevisedUp"] += diff
                elif bucket == "RevisedDown":
                    diff = initial - final
                    agg["RevisedDown"] += diff
                    total["RevisedDown"] += diff
            for agg in by_dir.values():
                agg["delta"] = agg["final"] - agg["initial"]
            total["delta"] = total["final"] - total["initial"]
            findings["variance_by_director"] = by_dir
            findings["variance_totals"] = total

        if "Territory Scorecard" in wb.sheetnames:
            ws = wb["Territory Scorecard"]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows[4:]:
                if not r or r[0] is None:
                    continue
                findings["scorecard"].append(
                    {
                        "director": r[0],
                        "open_deals": r[1] or 0,
                        "open_unwtd": r[2] or 0,
                        "open_wtd": r[3] or 0,
                        "coverage": r[4] or 0,
                        "q1_won": r[5] or 0,
                        "q1_lost": r[6] or 0,
                        "win_rate": r[7] or 0,
                        "slips": r[8] or 0,
                        "slip_arr": r[9] or 0,
                        "overdue": r[10] or 0,
                        "kyc_missing": r[11] or 0,
                    }
                )
        wb.close()

    if dash.exists():
        wb = load_workbook(dash, data_only=True, read_only=True)
        if "Stage Transition Matrix" in wb.sheetnames:
            ws = wb["Stage Transition Matrix"]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[3] if len(rows) > 3 else []
            stages = [h for h in header[1:-1] if h]
            pairs = []
            stage_order = {s: i for i, s in enumerate(stages)}
            for r in rows[4:]:
                if not r or r[0] is None or r[0] == "Total To":
                    continue
                frm = r[0]
                for ci, to in enumerate(stages, 1):
                    v = r[ci] if ci < len(r) else None
                    if isinstance(v, (int, float)) and v > 0:
                        pairs.append((frm, to, int(v)))
            pairs.sort(key=lambda p: -p[2])
            findings["transitions_top"] = pairs[:8]
            demos = [
                p
                for p in pairs
                if p[0] in stage_order
                and p[1] in stage_order
                and stage_order[p[1]] < stage_order[p[0]]
            ]
            findings["transitions_demotions"] = demos[:5]
        wb.close()

    return findings


def _read_snapshot_trend(wb_path, sheet_name):
    """Sum ARR per snapshot date for a director. Returns (dates, totals)."""
    if not wb_path.exists():
        return [], []
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return [], []
    header = rows[1]
    arr_cols = []
    for ci, h in enumerate(header):
        if h and isinstance(h, str) and re.match(r"^ARR \d{4}-\d{2}-\d{2}$", h):
            arr_cols.append((ci, h.split(" ", 1)[1]))
    totals = defaultdict(float)
    for r in rows[2:]:
        for ci, date in arr_cols:
            v = r[ci] if ci < len(r) else None
            try:
                totals[date] += float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                pass
    dates = [d for _, d in arr_cols]
    return dates, [totals[d] for d in dates]


def _q1_losses_for_director(wb_path):
    """Pull Q1 2026 Land losses from a director workbook.

    Scope matches slide 4 (Q1 Promised vs Delivered) in the main deck: Land
    only, Q1 2026 close date, not Won. Each returned dict has account,
    opportunity, owner, close_date, arr, reason, competitor.
    """
    if not wb_path.exists():
        return []
    wb = load_workbook(str(wb_path), read_only=True, data_only=True)
    if "Won Lost FY26" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Won Lost FY26"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = rows[0]
    losses = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        if str(d.get("Type", "")).strip().lower() != "land":
            continue
        if "Won" in str(d.get("Stage", "")):
            continue
        cd = str(d.get("Close Date", "") or "")[:10]
        if not ("2026-01-01" <= cd <= "2026-03-31"):
            continue
        losses.append(
            {
                "account": d.get("Account", "") or "",
                "opportunity": d.get("Opportunity", "") or "",
                "owner": d.get("Owner", "") or "",
                "close_date": cd,
                "arr": float(d.get("ARR Unweighted (EUR)") or 0),  # type: ignore[arg-type]
                "reason": str(d.get("Reason") or "(not recorded)"),
                "competitor": str(d.get("Lost To Competitor") or ""),
            }
        )
    return losses


def _load_stage_at_loss(opp_names):
    """Cross-reference Dashboard's Q1 History Raw to find the highest stage
    each opp ever reached before going to Lost. Returns {opp_name: stage}.
    """
    dash = SHAREPOINT / "Dashboard and Q1 Analysis.xlsx"
    if not dash.exists() or not opp_names:
        return {}
    wb = load_workbook(dash, read_only=True, data_only=True)
    if "Q1 History Raw" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Q1 History Raw"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    headers = rows[0]
    opp_set = set(opp_names)
    best = {}
    for r in rows[1:]:
        d = dict(zip(headers, r))
        opp = d.get("Opportunity")
        if opp not in opp_set:
            continue
        if d.get("Field Changed") != "StageName":
            continue
        for fld in ("Old Value", "New Value"):
            stage = str(d.get(fld) or "")
            rank = STAGE_RANK.get(stage, 0)
            if rank >= 1:
                cur = best.get(opp)
                if cur is None or rank > STAGE_RANK.get(cur, 0):
                    best[opp] = stage
    return best


def _territory_slug(territory):
    """Match slide_churn() in build_deck_from_excel.py for asset lookups."""
    return (territory or "").lower().replace(" & ", "-").replace(" ", "-")


def _update_snapshot_history(run_date, all_stats):
    """Write this run's headline numbers to a cross-run ledger at
    `obsidian/snapshot_history.json`. Month-over-month consumers read this
    ledger to find a prior baseline without re-parsing auto.md files.
    """
    path = VAULT / "snapshot_history.json"
    if path.exists():
        try:
            history = json.loads(path.read_text())
        except (json.JSONDecodeError, ValueError):
            history = {"snapshots": []}
    else:
        history = {"snapshots": []}

    per_dir = {}
    for s in all_stats:
        per_dir[s["director"]] = {
            "territory": s["territory"],
            # Deck uses Q1-Q2 scoped numbers; ledger mirrors that so MoM
            # deltas reconcile to what the deck actually renders.
            "open_land_deals": s.get("open_land_q1q2_deals", s["open_land_deals"]),
            "open_land_arr_unwtd": s.get(
                "open_land_q1q2_arr_unwtd", s["open_land_arr_unwtd"]
            ),
            "open_land_arr_wtd": s.get(
                "open_land_q1q2_arr_wtd", s["open_land_arr_wtd"]
            ),
            # Keep the unscoped totals too, as "all_*" siblings, for summary
            # views that want the whole book.
            "all_open_land_deals": s["open_land_deals"],
            "all_open_land_arr_unwtd": s["open_land_arr_unwtd"],
            "q1_won_count": s["q1_won_count"],
            "q1_won_arr": s["q1_won_arr"],
            "q1_lost_count": s["q1_lost_count"],
            "q1_lost_arr": s["q1_lost_arr"],
            "approved_2026": s["approved_2026"],
            "conditionally_approved": s["conditionally_approved"],
            "missing_approval": s["missing_approval"],
            "renewals_q2": s["renewals_q2"],
        }

    totals = {
        "open_land_deals": sum(s["open_land_deals"] for s in all_stats),
        "open_land_arr_unwtd": sum(s["open_land_arr_unwtd"] for s in all_stats),
        "open_land_arr_wtd": sum(s["open_land_arr_wtd"] for s in all_stats),
        "q1_won_count": sum(s["q1_won_count"] for s in all_stats),
        "q1_won_arr": sum(s["q1_won_arr"] for s in all_stats),
        "q1_lost_count": sum(s["q1_lost_count"] for s in all_stats),
        "q1_lost_arr": sum(s["q1_lost_arr"] for s in all_stats),
        "approved_2026": sum(s["approved_2026"] for s in all_stats),
        "missing_approval": sum(s["missing_approval"] for s in all_stats),
    }

    # Replace the entry for this run_date if present (idempotent), else append.
    history["snapshots"] = [
        e for e in history.get("snapshots", []) if e.get("run_date") != run_date
    ]
    history["snapshots"].append(
        {
            "run_date": run_date,
            "period": run_date[:7],
            "directors": per_dir,
            "totals": totals,
        }
    )
    history["snapshots"].sort(key=lambda e: e["run_date"])
    path.write_text(json.dumps(history, indent=2) + "\n")
    return path


def _churn_asset_path(territory):
    """Path to the churn screenshot keyed by territory slug (if present)."""
    slug = _territory_slug(territory)
    return ROOT / "assets" / "rebekka-screenshots" / f"{slug}-churn.png"


def write_monthly_summary(
    month_dir, run_date, all_stats, manifest, findings=None, all_losses=None
):
    path = month_dir / "README.md"
    lines = [
        f"# {datetime.strptime(run_date, '%Y-%m-%d').strftime('%B %Y')} review",
        "",
        f"Snapshot: {run_date}. Full manifest at "
        f"`output/pipeline_logs/{run_date}/manifest.json`.",
        "",
        "## Headline numbers across all nine territories",
        "",
    ]
    total_open = sum(s["open_land_deals"] for s in all_stats)
    total_unwtd = sum(s["open_land_arr_unwtd"] for s in all_stats)
    total_wtd = sum(s["open_land_arr_wtd"] for s in all_stats)
    total_q1_won = sum(s["q1_won_count"] for s in all_stats)
    total_q1_won_arr = sum(s["q1_won_arr"] for s in all_stats)
    total_q1_lost = sum(s["q1_lost_count"] for s in all_stats)
    total_q1_lost_arr = sum(s["q1_lost_arr"] for s in all_stats)
    total_approved = sum(s["approved_2026"] for s in all_stats)
    total_cond = sum(s["conditionally_approved"] for s in all_stats)
    total_missing = sum(s["missing_approval"] for s in all_stats)

    lines.extend(
        [
            f"- Open Land pipeline: {total_open} deals, "
            f"{_fmt(total_unwtd)} unweighted, {_fmt(total_wtd)} weighted.",
            f"- Q1 Land wins: {total_q1_won} deals, {_fmt(total_q1_won_arr)}.",
            f"- Q1 Land losses: {total_q1_lost} deals, {_fmt(total_q1_lost_arr)}.",
            f"- Approvals: {total_approved} approved 2026, "
            f"{total_cond} pending approval, "
            f"{total_missing} missing Stage 3+.",
            "",
            "## By director",
            "",
            "| Director | Territory | Open Deals | Open ARR Unwtd | Q1 Won | "
            "Approved | Missing |",
            "|---|---|---:|---:|---:|---:|---:|",
        ]
    )
    for s in all_stats:
        link = f"[[{s['director']}|{s['director']}]]"
        lines.append(
            f"| {link} | {s['territory']} | "
            f"{s['open_land_deals']} | {_fmt(s['open_land_arr_unwtd'])} | "
            f"{s['q1_won_count']} | {s['approved_2026']} | "
            f"{s['missing_approval']} |"
        )
    # Analytics findings from the consolidated FY26 Pipeline Review workbook.
    # Caller may pre-load so per-director writers reuse the same read.
    if findings is None:
        findings = _read_analytics_workbook()
    if findings.get("concentration_summary") or findings.get("top_deals"):
        lines.extend(["", "## ARR concentration", ""])
        if findings.get("concentration_summary"):
            lines.append(findings["concentration_summary"])
            lines.append("")
        for d in findings["top_deals"][:5]:
            lines.append(
                f"{d['rank']}. {d['account']} - {d['opportunity']} "
                f"({d['director']}, {d['stage']}): {_fmt(d['arr'])} "
                f"({d['pct'] * 100:.1f}% of total; cum {d['cum_pct'] * 100:.1f}%)"
            )

    if findings.get("pivot_stage_totals"):
        lines.extend(["", "## Pipeline by stage, all directors", ""])
        lines.append("| Stage | ARR Unwtd |")
        lines.append("|---|---:|")
        for stage, total in sorted(findings["pivot_stage_totals"].items()):
            lines.append(f"| {stage} | {_fmt(total)} |")

    if findings.get("velocity_q1") or findings.get("velocity_q2"):
        lines.extend(["", "## Pipeline velocity", ""])
        for label, block in [
            ("Q1 2026", findings.get("velocity_q1") or {}),
            ("Q2 2026", findings.get("velocity_q2") or {}),
        ]:
            dates = block.get("dates") or []
            dirs = block.get("directors") or []
            if not dates or not dirs:
                continue
            first, last = dates[0], dates[-1]
            total_first = sum(s[1][0] for s in dirs if s[1])
            total_last = sum(s[1][-1] for s in dirs if s[1])
            delta = total_last - total_first
            sign = "+" if delta >= 0 else "-"
            lines.append(
                f"- {label}: {_fmt(total_first)} on {first} -> "
                f"{_fmt(total_last)} on {last} "
                f"({sign}{_fmt(abs(delta))})."
            )
            # Biggest mover
            movers = sorted(dirs, key=lambda s: (s[1][-1] - s[1][0]) if s[1] else 0)
            if movers:
                worst = movers[0]
                best = movers[-1]
                worst_delta = worst[1][-1] - worst[1][0]
                best_delta = best[1][-1] - best[1][0]
                lines.append(
                    f"  - Biggest drop: {worst[0]} "
                    f"({_fmt(worst_delta)}). Biggest gain: {best[0]} "
                    f"({_fmt(best_delta)})."
                )

    if findings.get("slip_top_owners"):
        lines.extend(["", "## Slip risk, top owners by push count", ""])
        lines.append(
            "| Owner | Director | Deals | Open ARR | Total Pushes | Max on 1 |"
        )
        lines.append("|---|---|---:|---:|---:|---:|")
        for o in findings["slip_top_owners"]:
            lines.append(
                f"| {o['owner']} | {o['director']} | {o['deals']} | "
                f"{_fmt(o['arr'])} | {o['pushes']} | {o['max_push']} |"
            )

    if findings.get("transitions_top") or findings.get("transitions_demotions"):
        lines.extend(["", "## Stage transitions, Q1", ""])
        if findings.get("transitions_top"):
            lines.append("Top forward flows:")
            for frm, to, n in findings["transitions_top"]:
                lines.append(f"- {frm} -> {to}: {n}")
        if findings.get("transitions_demotions"):
            lines.append("")
            lines.append("Backward (demotion) transitions:")
            for frm, to, n in findings["transitions_demotions"]:
                lines.append(f"- {frm} -> {to}: {n}")

    # Deal Risk top scored deals. Triage threshold is read from Parameters so
    # changes on that tab propagate straight into the README narrative.
    if findings.get("deal_risk"):
        triage = findings.get("thresholds", {}).get("Thresh_RiskScoreTriage", 60)
        try:
            triage_num = float(triage)
        except (TypeError, ValueError):
            triage_num = 60.0
        risk = findings["deal_risk"]
        over = [d for d in risk if (d.get("score") or 0) >= triage_num]
        lines.extend(
            [
                "",
                "## Deal Risk, top scored deals",
                "",
                f"Composite risk score per open Land deal, weights on Parameters "
                f"tab. Triage threshold is {int(triage_num)}; "
                f"{len(over)} deals at or above.",
                "",
                "| Rank | Score | Director | Account | Opportunity | ARR | Proof |",
                "|---:|---:|---|---|---|---:|---|",
            ]
        )
        for d in risk[:10]:
            lines.append(
                f"| {d['rank']} | {d['score']} | {d['director']} | "
                f"{d['account']} | {d['opportunity']} | {_fmt(d['arr'])} | "
                f"{d['proof']} |"
            )

    # Forecast Variance, Q1 2026 decomposition. Values computed from the
    # Bucket helper column on Q1 Trend Consolidated (SUMIFS on the tab).
    if findings.get("variance_totals"):
        totals = findings["variance_totals"]
        by_dir = findings.get("variance_by_director", {})
        loss_driven_ratio = findings.get("thresholds", {}).get(
            "Thresh_LossDrivenRatio", 2
        )
        try:
            ratio = float(loss_driven_ratio)
        except (TypeError, ValueError):
            ratio = 2.0
        won = totals.get("Won", 0)
        lost = totals.get("Lost", 0)
        if won > 0 and lost / max(won, 1) >= ratio:
            narrative = "loss-driven"
        elif lost > 0 and won / max(lost, 1) >= ratio:
            narrative = "win-driven"
        else:
            narrative = "balanced"
        delta = totals.get("delta", 0)
        sign = "shrank" if delta < 0 else "expanded"
        lines.extend(
            [
                "",
                "## Forecast Variance, Q1 2026",
                "",
                f"Initial ARR {_fmt(totals.get('initial', 0))} -> "
                f"final {_fmt(totals.get('final', 0))} "
                f"(net {_fmt(delta)}). Q1 book {sign} EUR "
                f"{abs(delta) / 1_000_000:.1f}M: {narrative} "
                f"(Won {_fmt(won)}, Lost {_fmt(lost)}, "
                f"Added {_fmt(totals.get('Added', 0))}, "
                f"RevisedUp {_fmt(totals.get('RevisedUp', 0))}, "
                f"RevisedDown {_fmt(totals.get('RevisedDown', 0))}).",
                "",
                "| Director | Initial | Final | Delta | Won | Lost | Added | Rev Up | Rev Down |",
                "|---|---:|---:|---:|---:|---:|---:|---:|---:|",
            ]
        )
        director_order = [name for name, _, _ in DIRECTORS]
        for name in director_order:
            v = by_dir.get(name)
            if not v:
                continue
            lines.append(
                f"| {name} | {_fmt(v['initial'])} | {_fmt(v['final'])} | "
                f"{_fmt(v['delta'])} | {_fmt(v['Won'])} | {_fmt(v['Lost'])} | "
                f"{_fmt(v['Added'])} | {_fmt(v['RevisedUp'])} | "
                f"{_fmt(v['RevisedDown'])} |"
            )
        lines.append(
            f"| TOTAL | {_fmt(totals.get('initial', 0))} | "
            f"{_fmt(totals.get('final', 0))} | {_fmt(delta)} | "
            f"{_fmt(won)} | {_fmt(lost)} | {_fmt(totals.get('Added', 0))} | "
            f"{_fmt(totals.get('RevisedUp', 0))} | "
            f"{_fmt(totals.get('RevisedDown', 0))} |"
        )

    # Q1 Land losses by reason, global view (aggregate across all 9 directors).
    if all_losses:
        from collections import defaultdict as _dd

        reason_agg = _dd(lambda: {"n": 0, "arr": 0.0})
        total_n = 0
        total_arr = 0.0
        for losses in all_losses.values():
            for loss in losses:
                reason_agg[loss["reason"]]["n"] += 1
                reason_agg[loss["reason"]]["arr"] += loss["arr"]
                total_n += 1
                total_arr += loss["arr"]
        if total_n:
            lines.extend(
                [
                    "",
                    "## Q1 Land losses by reason (global)",
                    "",
                    f"Aggregated across all 9 territories. Scope matches the "
                    f"per-director deck Q1 retrospective: Land, Q1 2026 close "
                    f"date, not Won. Total {total_n} deals, {_fmt(total_arr)}.",
                    "",
                    "| Reason | Count | Lost ARR |",
                    "|---|---:|---:|",
                ]
            )
            for reason, agg in sorted(reason_agg.items(), key=lambda x: -x[1]["n"]):
                lines.append(f"| {reason} | {agg['n']} | {_fmt(agg['arr'])} |")
            lines.append(f"| Total | {total_n} | {_fmt(total_arr)} |")

    # Audit parameters reference.
    lines.extend(
        [
            "",
            "## Audit parameters",
            "",
            "Risk weights and insight thresholds live in "
            "`output/sharepoint/FY26 Pipeline Review, All Territories.xlsx` on "
            "the Parameters tab. Values are addressable via Defined Names "
            "(`RiskWeight_*`, `Thresh_*`) and drive Executive Insights, "
            "Deal Risk Scoring, and Forecast Variance. Methodology tab "
            "explains each rule.",
        ]
    )

    lines.extend(
        [
            "",
            "## Analysis workbook references",
            "",
            "- `output/sharepoint/FY26 Pipeline Review, All Territories.xlsx` "
            "tabs: Pipeline Pivot, ARR Concentration, Pipeline Velocity, "
            "Slip Risk by Owner, Territory Scorecard",
            "- `output/sharepoint/Dashboard and Q1 Analysis.xlsx` "
            "tabs: Stage Transition Matrix, PI Summary, plus every widget "
            "from the Sales Directors Monthly and Sales Ops Quarterly KPI "
            "dashboards as raw data tabs",
            "",
            "## Pipeline run",
            "",
        ]
    )
    for step in manifest.get("steps", []):
        status = step.get("status", "?")
        name = step.get("name", "?")
        duration = step.get("duration_seconds", 0)
        lines.append(f"- `{status}` {name} ({duration:.1f}s)")
    lines.append("")
    lines.append("## Outputs")
    lines.append("")
    n_decks = len(manifest.get("outputs", {}).get("decks", []))
    n_extracts = len(manifest.get("outputs", {}).get("extracts", []))
    n_reports = len(manifest.get("outputs", {}).get("reports", []))
    lines.extend(
        [
            f"- Extracts: {n_extracts}",
            f"- Analysis reports: {n_reports}",
            f"- Decks shipped: {n_decks}",
            "",
            "See [[README]] for full vault index.",
        ]
    )
    path.write_text("\n".join(lines) + "\n")
    return path


def write_monthly_director(
    month_dir,
    stats,
    run_date,
    wb_path=None,
    findings=None,
    losses=None,
    stage_at_loss=None,
):
    """Auto-generated monthly snapshot. Overwritten every run. Never hand-edit."""
    slug = _slug(stats["director"])
    path = month_dir / f"{slug}.auto.md"
    top = stats["top_land_deals"]
    push = stats["owner_push_concentration"]
    period = run_date[:7]

    # Pull per-director Q1/Q2 snapshot trajectory if the workbook is available
    velocity_lines = []
    if wb_path is not None and wb_path.exists():
        for period_label, sheet in [
            ("Q1 2026", "Q1 Snapshot Trend"),
            ("Q2 2026", "Q2 Snapshot Trend"),
        ]:
            dates, totals = _read_snapshot_trend(wb_path, sheet)
            if not dates or not totals:
                continue
            first, last = totals[0], totals[-1]
            delta = last - first
            sign = "+" if delta >= 0 else "-"
            velocity_lines.append(
                f"- {period_label}: {_fmt(first)} on {dates[0]} -> "
                f"{_fmt(last)} on {dates[-1]} "
                f"({sign}{_fmt(abs(delta))})."
            )

    lines = [
        "---",
        "type: monthly-snapshot",
        "generated: true",
        f'director: "[[Directors/{slug}]]"',
        f"period: {period}",
        f"run_date: {run_date}",
        f"source: director_live_workbooks/{run_date}/{_slug(stats['director'])}.xlsx",
        "tags: [monthly, auto]",
        "---",
        "",
        "> Auto-generated by scripts/generate_obsidian_notes.py.",
        "> Do not hand-edit. Put commentary in the .notes.md sibling file.",
        "",
        f"# {stats['director']}, {stats['territory']}, {run_date}",
        "",
        "## Headline",
        "",
        f"- Open Land pipeline: {stats['open_land_deals']} deals, "
        f"{_fmt(stats['open_land_arr_unwtd'])} unweighted, "
        f"{_fmt(stats['open_land_arr_wtd'])} weighted.",
        f"- Q1 Land outcome: {stats['q1_won_count']} wins "
        f"({_fmt(stats['q1_won_arr'])}), {stats['q1_lost_count']} losses "
        f"({_fmt(stats['q1_lost_arr'])}).",
        f"- Commercial approvals: {stats['approved_2026']} approved 2026, "
        f"{stats['conditionally_approved']} pending approval, "
        f"{stats['missing_approval']} missing Stage 3+.",
        f"- Q2 renewals due: {stats['renewals_q2']}.",
        f"- Q1 movement events in scope: {stats['q1_slip_events']}.",
    ]
    if top:
        lines.extend(["", "## Top open Land deals", ""])
        for d in top:
            lines.append(
                f"- {d.get('Account', '')}: {d.get('Opportunity', '')}, "
                f"{d.get('Stage', '')}, "
                f"close {str(d.get('Close Date', ''))[:10]}, "
                f"{_fmt(float(d.get('ARR Unweighted (EUR)') or 0))}"
            )
    if push:
        lines.extend(["", "## Push concentration (owners with 3+ pushes)", ""])
        for owner, n in push:
            lines.append(f"- {owner}: {n} pushed deals")
    if velocity_lines:
        lines.extend(["", "## Pipeline velocity, per snapshot", ""])
        lines.extend(velocity_lines)

    # Top Q2 deals at risk, scoped to this director + Q2 2026 close dates.
    # Reconciles to deck slide 9 (Top Q2 Deals at Risk).
    if findings and findings.get("deal_risk"):
        director = stats["director"]
        q2_lo, q2_hi = Q2_RANGE
        mine = [
            d
            for d in findings["deal_risk"]
            if d.get("director") == director
            and q2_lo <= str(d.get("close_date", ""))[:10] <= q2_hi
        ]
        if mine:
            lines.extend(
                [
                    "",
                    "## Top Q2 deals at risk",
                    "",
                    "Composite risk score, this director, Q2 2026 close dates. "
                    "Reason codes and weights on Parameters tab.",
                    "",
                    "| Score | Account | Opportunity | Stage | ARR | Reasons | Proof |",
                    "|---:|---|---|---|---:|---|---|",
                ]
            )
            for d in mine[:10]:
                lines.append(
                    f"| {d['score']} | {d['account']} | {d['opportunity']} | "
                    f"{d['stage']} | {_fmt(d['arr'])} | {d['reason_codes']} | "
                    f"{d['proof']} |"
                )

    # Q1 Land losses by reason + stage at loss. Scope matches deck slide 4.
    if losses:
        from collections import defaultdict as _dd

        reason_agg = _dd(lambda: {"n": 0, "arr": 0.0})
        for loss in losses:
            reason_agg[loss["reason"]]["n"] += 1
            reason_agg[loss["reason"]]["arr"] += loss["arr"]
        total_n = sum(a["n"] for a in reason_agg.values())
        total_arr = sum(a["arr"] for a in reason_agg.values())
        lines.extend(
            [
                "",
                "## Q1 Land losses by reason",
                "",
                f"Scope matches deck slide 4 (Q1 Promised vs Delivered). "
                f"{total_n} deals, {_fmt(total_arr)}.",
                "",
                "| Reason | Count | Lost ARR |",
                "|---|---:|---:|",
            ]
        )
        for reason, agg in sorted(reason_agg.items(), key=lambda x: -x[1]["n"]):
            lines.append(f"| {reason} | {agg['n']} | {_fmt(agg['arr'])} |")
        lines.append(f"| Total | {total_n} | {_fmt(total_arr)} |")

        if stage_at_loss is not None:
            stage_agg = _dd(lambda: {"n": 0, "arr": 0.0})
            unclassified_n = 0
            unclassified_arr = 0.0
            for loss in losses:
                stage = stage_at_loss.get(loss["opportunity"])
                if stage:
                    stage_agg[stage]["n"] += 1
                    stage_agg[stage]["arr"] += loss["arr"]
                else:
                    unclassified_n += 1
                    unclassified_arr += loss["arr"]
            lines.extend(
                [
                    "",
                    "## Stage at loss, Q1",
                    "",
                    "Highest stage each lost opp ever reached before going to "
                    "Lost. Early-stage = qualification gap; late-stage = "
                    "execution gap. Source: OpportunityFieldHistory via "
                    "Dashboard workbook Q1 History Raw.",
                    "",
                    "| Stage reached | Count | Lost ARR |",
                    "|---|---:|---:|",
                ]
            )
            for stg in STAGES_IN_ORDER:
                if stage_agg[stg]["n"] == 0 and stage_agg[stg]["arr"] == 0:
                    continue
                lines.append(
                    f"| {stg} | {stage_agg[stg]['n']} | {_fmt(stage_agg[stg]['arr'])} |"
                )
            if unclassified_n:
                lines.append(
                    f"| (no stage history) | {unclassified_n} | "
                    f"{_fmt(unclassified_arr)} |"
                )

    # Churn screenshot reference, only if the asset is on disk. Keyed by
    # territory slug to match build_deck_from_excel.py slide_churn().
    churn_path = _churn_asset_path(stats["territory"])
    if churn_path.exists():
        lines.extend(
            [
                "",
                "## Churn reference",
                "",
                f"Churn view imported from prior monthly pack — see deck "
                f"slide 16. Asset: `{churn_path.relative_to(ROOT)}`.",
            ]
        )

    lines.extend(
        [
            "",
            "## Links",
            "",
            f"- Standing page: [[Directors/{slug}]]",
            f"- Commentary: [[Monthly/{period}/{slug}.notes]]",
            f"- Monthly index: [[Monthly/{period}/README|{period}]]",
        ]
    )
    path.write_text("\n".join(lines) + "\n")

    # Also seed a companion .notes.md for human commentary the first time only.
    notes_path = month_dir / f"{slug}.notes.md"
    if not notes_path.exists():
        notes_lines = [
            "---",
            "type: monthly-notes",
            f'director: "[[Directors/{slug}]]"',
            f"period: {period}",
            "tags: [monthly, commentary]",
            "---",
            "",
            f"# {stats['director']}, commentary for {period}",
            "",
            f"Auto snapshot: ![[Monthly/{period}/{slug}.auto]]",
            "",
            "## Commentary",
            "",
            "(Director commentary goes here. Safe to edit; generator will",
            "not overwrite.)",
            "",
            "## Action items",
            "",
            "- ",
        ]
        notes_path.write_text("\n".join(notes_lines) + "\n")
    return path


def write_standing_director(stats):
    slug = _slug(stats["director"])
    path = VAULT / "Directors" / f"{slug}.md"
    if path.exists():
        # Do not overwrite: standing pages are editable by humans.
        return path

    region_tag = stats["territory"].split()[0].lower()  # apac / emea / na
    lines = [
        "---",
        "type: director",
        f'name: "{stats["director"]}"',
        f'aliases: ["{stats["director"]}", "{stats["territory"]}"]',
        f'territory: "{stats["territory"]}"',
        f"region: {region_tag}",
        "status: active",
        f"tags: [director, region/{region_tag}]",
        "---",
        "",
        f"# {stats['director']}, {stats['territory']}",
        "",
        "Standing page. Safe to edit. Not regenerated.",
        "",
        "## Territory scope",
        "",
        f"See `TERRITORIES` in `scripts/extract_director_live.py` for the "
        f"exact SOQL filter used for {stats['director']}.",
        "",
        "## Known notes",
        "",
        "- ",
        "",
        "## Recent monthly snapshots",
        "",
        f"- [[Monthly/2026-04/{slug}.auto|April 2026 auto]]  "
        f"+ [[Monthly/2026-04/{slug}.notes|commentary]]",
    ]
    path.write_text("\n".join(lines) + "\n")
    return path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument(
        "--workbooks-dir",
        type=Path,
        default=None,
        help="Defaults to output/director_live_workbooks/<date>",
    )
    args = parser.parse_args()

    run_date = args.date
    wb_dir = args.workbooks_dir or (
        ROOT / "output" / "director_live_workbooks" / run_date
    )
    if not wb_dir.exists():
        print(f"  workbooks dir missing: {wb_dir}")
        return 1

    manifest_path = ROOT / "output" / "pipeline_logs" / run_date / "manifest.json"
    manifest = json.loads(manifest_path.read_text()) if manifest_path.exists() else {}

    month_key = run_date[:7]  # YYYY-MM
    month_dir = VAULT / "Monthly" / month_key
    month_dir.mkdir(parents=True, exist_ok=True)

    # Read analytics findings once so both writers share the same source data.
    findings = _read_analytics_workbook()

    # Collect Q1 Land losses per director up front so we can do a single
    # batched stage-at-loss lookup against the Dashboard workbook.
    losses_by_director = {}
    for director, _territory, fname in DIRECTORS:
        wb_path = wb_dir / fname
        if not wb_path.exists():
            continue
        losses_by_director[director] = _q1_losses_for_director(wb_path)
    all_lost_opps = [
        loss["opportunity"] for losses in losses_by_director.values() for loss in losses
    ]
    stage_at_loss = _load_stage_at_loss(all_lost_opps)

    all_stats = []
    for director, territory, fname in DIRECTORS:
        wb_path = wb_dir / fname
        if not wb_path.exists():
            print(f"  skip {director}: {wb_path} missing")
            continue
        stats = _director_stats(director, territory, wb_path)
        all_stats.append(stats)
        p = write_monthly_director(
            month_dir,
            stats,
            run_date,
            wb_path=wb_path,
            findings=findings,
            losses=losses_by_director.get(director, []),
            stage_at_loss=stage_at_loss,
        )
        write_standing_director(stats)
        print(f"  wrote {p.relative_to(ROOT)}")

    p = write_monthly_summary(
        month_dir,
        run_date,
        all_stats,
        manifest,
        findings=findings,
        all_losses=losses_by_director,
    )
    print(f"  wrote {p.relative_to(ROOT)}")

    # Update the cross-run snapshot history ledger so MoM deltas have a
    # baseline. Each entry is keyed by run_date; re-running the same date
    # overwrites that entry (idempotent). MoM consumers pick the closest
    # prior entry to compare against.
    hist_path = _update_snapshot_history(run_date, all_stats)
    print(f"  updated {hist_path.relative_to(ROOT)}")

    print(f"\nObsidian vault: {VAULT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
