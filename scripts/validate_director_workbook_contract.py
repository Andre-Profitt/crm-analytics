#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LIVE_AUDIT_ROOT = ROOT / "output" / "director_live_extract"
HISTORICAL_AUDIT_ROOT = ROOT / "output" / "historical_trending_extract"
OUTPUT_ROOT = ROOT / "output" / "director_workbook_contract"


LIVE_SHEET_SPECS = [
    {"label": "Summary", "type": "exact", "value": "Summary", "metric": None},
    {
        "label": "Pipeline Open",
        "type": "prefix",
        "value": "Pipeline Open FY",
        "metric": "pipeline_open",
    },
    {
        "label": "Won Lost",
        "type": "prefix",
        "value": "Won Lost FY",
        "metric": "won_lost",
    },
    {
        "label": "Commercial Approval",
        "type": "exact",
        "value": "Commercial Approval",
        "metric": "commercial_approval_land",
    },
    {
        "label": "Renewals",
        "type": "prefix",
        "value": "Renewals FY",
        "metric": "renewals",
    },
    {
        "label": "Pipeline Inspection",
        "type": "exact",
        "value": "Pipeline Inspection",
        "metric": "pipeline_inspection",
    },
    {
        "label": "Activity Volume",
        "type": "exact",
        "value": "Activity Volume",
        "metric": "activity_volume_rows",
    },
    {
        "label": "Commit Items",
        "type": "exact",
        "value": "Commit Items",
        "metric": "commit_items",
    },
    {
        "label": "Q1 Movement",
        "type": "exact",
        "value": "Q1 Movement",
        "metric": "q1_movement",
    },
    {
        "label": "Q2 Movement",
        "type": "exact",
        "value": "Q2 Movement",
        "metric": "q2_movement",
    },
    {
        "label": "Stage History",
        "type": "exact",
        "value": "Stage History",
        "metric": "stage_history_events",
    },
    {
        "label": "Forecast Category History",
        "type": "exact",
        "value": "Forecast Category History",
        "metric": "forecast_category_history_events",
    },
]


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _infer_report_date(workbooks_dir: Path | None, snapshot_date: str | None) -> str:
    token = str(snapshot_date or "").strip()[:10]
    if token:
        return token
    if workbooks_dir is not None:
        for candidate in [Path(workbooks_dir), *Path(workbooks_dir).parents]:
            try:
                datetime.strptime(candidate.name, "%Y-%m-%d")
                return candidate.name
            except ValueError:
                continue
    return datetime.now().strftime("%Y-%m-%d")


def _load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"{path}: expected JSON object")
    return payload


def _slug_from_workbook_path(value: str) -> str:
    return Path(str(value or "")).stem.strip()


def _resolve_workbook_path(workbook_value: str, workbooks_dir: Path) -> Path:
    raw = Path(str(workbook_value or ""))
    if raw.is_absolute():
        return raw
    if raw.parts:
        candidate = ROOT / raw
        if candidate.exists():
            return candidate
    return workbooks_dir / raw.name


def _sheet_row_count(ws, *, historical: bool) -> int:
    header_rows = 2 if historical else 1
    return max(int(ws.max_row or 0) - header_rows, 0)


def _expected_live_metric_count(counts: dict[str, Any], metric: str) -> int:
    if metric == "commercial_approval_land":
        explicit = counts.get("commercial_approval_sheet_rows")
        if explicit is not None:
            return int(explicit or 0)
        return sum(
            int(counts.get(key) or 0)
            for key in (
                "approved_current_year",
                "approved_prior_year",
                "pending_approval",
                "missing_approval",
            )
        )
    return int(counts.get(metric) or 0)


def _match_sheet_name(sheetnames: list[str], spec: dict[str, str]) -> str | None:
    if spec["type"] == "exact":
        return spec["value"] if spec["value"] in sheetnames else None
    matches = [name for name in sheetnames if name.startswith(spec["value"])]
    if len(matches) == 1:
        return matches[0]
    return None


def _write_run_audit(output_dir: Path, payload: dict[str, Any]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "director_workbook_contract_audit.json").write_text(
        json.dumps(payload, indent=2) + "\n",
        encoding="utf-8",
    )
    lines = [
        f"# Director Workbook Contract Audit — {payload['run_date']}",
        "",
        f"- Status: `{payload['status']}`",
        f"- Workbooks dir: `{payload['workbooks_dir']}`",
        f"- Scope: `{payload['scope']}`",
        f"- Workbooks validated: `{len(payload.get('validated') or [])}`",
        f"- Failures: `{len(payload.get('failures') or [])}`",
        f"- Warnings: `{len(payload.get('warnings') or [])}`",
        "",
        "## Validated",
        "",
    ]
    validated = payload.get("validated") or []
    if not validated:
        lines.append("- none")
    else:
        for item in validated:
            lines.append(
                f"- `{item['slug']}`: `{item['sheet_count']}` sheet(s), "
                f"`{len(item.get('historical_sheets') or [])}` historical sheet(s)"
            )
    lines.extend(["", "## Failures", ""])
    failures = payload.get("failures") or []
    if not failures:
        lines.append("- none")
    else:
        for item in failures:
            lines.append(
                f"- `{item.get('slug', '')}`: `{item.get('issue', 'unknown')}` "
                f"{item.get('message', '')}".strip()
            )
    lines.extend(["", "## Warnings", ""])
    warnings = payload.get("warnings") or []
    if not warnings:
        lines.append("- none")
    else:
        for item in warnings:
            lines.append(
                f"- `{item.get('slug', '')}`: `{item.get('issue', 'unknown')}` "
                f"{item.get('message', '')}".strip()
            )
    (output_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--snapshot-date",
        help="Explicit report date (YYYY-MM-DD) used for audit lookup and defaults.",
    )
    parser.add_argument(
        "--workbooks-dir",
        type=Path,
        help="Workbook directory to validate. Defaults to output/director_live_workbooks/<date>.",
    )
    parser.add_argument(
        "--director",
        help="Only validate one director slug (e.g. jesper-tyrer).",
    )
    parser.add_argument(
        "--require-historical",
        action="store_true",
        help="Fail if the historical-trending audit for the run date is missing.",
    )
    args = parser.parse_args()

    run_date = _infer_report_date(args.workbooks_dir, args.snapshot_date)
    workbooks_dir = args.workbooks_dir or (ROOT / "output" / "director_live_workbooks" / run_date)
    output_dir = OUTPUT_ROOT / run_date
    failures: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    validated: list[dict[str, Any]] = []

    live_audit_path = LIVE_AUDIT_ROOT / run_date / "director_live_extract_audit.json"
    if not live_audit_path.exists():
        payload = {
            "run_date": run_date,
            "workbooks_dir": str(workbooks_dir),
            "scope": args.director or "all",
            "status": "failed",
            "validated": [],
            "failures": [
                {
                    "issue": "live_extract_audit_missing",
                    "message": f"missing {live_audit_path}",
                }
            ],
            "warnings": [],
        }
        _write_run_audit(output_dir, payload)
        print(f"Workbook contract audit: {_display_path(output_dir)}")
        return 1

    live_audit = _load_json(live_audit_path)
    historical_audit_path = (
        HISTORICAL_AUDIT_ROOT / run_date / "historical_trending_extract_audit.json"
    )
    historical_audit = None
    if historical_audit_path.exists():
        historical_audit = _load_json(historical_audit_path)
    elif args.require_historical:
        failures.append(
            {
                "issue": "historical_extract_audit_missing",
                "message": f"missing {historical_audit_path}",
            }
        )

    if not workbooks_dir.exists():
        failures.append(
            {
                "issue": "workbooks_dir_missing",
                "message": f"missing {workbooks_dir}",
            }
        )

    live_items = list(live_audit.get("processed") or [])
    expected_slugs = set()
    live_index: dict[str, dict[str, Any]] = {}
    for item in live_items:
        slug = _slug_from_workbook_path(str(item.get("workbook_path") or ""))
        if not slug:
            continue
        if args.director and slug != args.director:
            continue
        expected_slugs.add(slug)
        live_index[slug] = item

    historical_index: dict[str, dict[str, Any]] = {}
    historical_failures_by_slug: dict[str, list[dict[str, Any]]] = {}
    if historical_audit:
        for item in historical_audit.get("processed") or []:
            slug = str(item.get("slug") or "").strip()
            if slug:
                historical_index[slug] = item
        for item in historical_audit.get("failures") or []:
            slug = str(item.get("slug") or "").strip()
            if slug:
                historical_failures_by_slug.setdefault(slug, []).append(item)

    if not args.director and workbooks_dir.exists():
        actual_slugs = {
            path.stem
            for path in workbooks_dir.glob("*.xlsx")
            if not path.name.startswith("~")
        }
        for slug in sorted(actual_slugs - expected_slugs):
            failures.append(
                {
                    "slug": slug,
                    "issue": "unexpected_workbook",
                    "message": f"unexpected workbook in {workbooks_dir}",
                }
            )

    for slug in sorted(expected_slugs):
        live_item = live_index[slug]
        workbook_path = _resolve_workbook_path(
            str(live_item.get("workbook_path") or ""),
            workbooks_dir,
        )
        if not workbook_path.exists():
            failures.append(
                {
                    "slug": slug,
                    "issue": "workbook_missing",
                    "message": str(workbook_path),
                }
            )
            continue

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        sheetnames = list(wb.sheetnames)
        slug_failures = []
        slug_warnings = []
        counts = dict(live_item.get("counts") or {})

        matched_sheet_names: list[str] = []
        for spec in LIVE_SHEET_SPECS:
            matched = _match_sheet_name(sheetnames, spec)
            if not matched:
                slug_failures.append(
                    {
                        "slug": slug,
                        "issue": "missing_sheet",
                        "message": spec["label"],
                    }
                )
                continue
            matched_sheet_names.append(matched)
            metric = spec["metric"]
            if metric:
                actual_count = _sheet_row_count(wb[matched], historical=False)
                expected_count = _expected_live_metric_count(counts, metric)
                if actual_count != expected_count:
                    slug_failures.append(
                        {
                            "slug": slug,
                            "issue": "row_count_mismatch",
                            "message": (
                                f"{matched}: expected {expected_count}, got {actual_count}"
                            ),
                        }
                    )

        forward_pi = dict(live_item.get("forward_quarter_pi") or {})
        forward_sheet_present = "Pipeline Inspection Forward" in sheetnames
        if str(forward_pi.get("status") or "unavailable") != "unavailable":
            if not forward_sheet_present:
                slug_failures.append(
                    {
                        "slug": slug,
                        "issue": "missing_sheet",
                        "message": "Pipeline Inspection Forward",
                    }
                )
            else:
                matched_sheet_names.append("Pipeline Inspection Forward")
                actual_count = _sheet_row_count(
                    wb["Pipeline Inspection Forward"], historical=False
                )
                expected_count = int(counts.get("pipeline_inspection_forward") or 0)
                if actual_count != expected_count:
                    slug_failures.append(
                        {
                            "slug": slug,
                            "issue": "row_count_mismatch",
                            "message": (
                                "Pipeline Inspection Forward: "
                                f"expected {expected_count}, got {actual_count}"
                            ),
                        }
                    )
        elif forward_sheet_present:
            slug_warnings.append(
                {
                    "slug": slug,
                    "issue": "unexpected_optional_sheet",
                    "message": "Pipeline Inspection Forward",
                }
            )

        historical_item = historical_index.get(slug)
        historical_sheet_names: list[str] = []
        if historical_item:
            for item in historical_item.get("sheets") or []:
                sheet_name = str(item.get("sheet_name") or "").strip()
                if not sheet_name:
                    continue
                if sheet_name not in sheetnames:
                    slug_failures.append(
                        {
                            "slug": slug,
                            "issue": "missing_sheet",
                            "message": sheet_name,
                        }
                    )
                    continue
                historical_sheet_names.append(sheet_name)
                actual_count = _sheet_row_count(wb[sheet_name], historical=True)
                expected_count = int(item.get("row_count") or 0)
                if actual_count != expected_count:
                    slug_failures.append(
                        {
                            "slug": slug,
                            "issue": "row_count_mismatch",
                            "message": (
                                f"{sheet_name}: expected {expected_count}, got {actual_count}"
                            ),
                        }
                    )
        elif args.require_historical:
            if slug in historical_failures_by_slug:
                slug_failures.append(
                    {
                        "slug": slug,
                        "issue": "historical_extract_failed",
                        "message": str(
                            historical_failures_by_slug[slug][0].get("issues") or []
                        ),
                    }
                )
            else:
                slug_failures.append(
                    {
                        "slug": slug,
                        "issue": "historical_contract_missing",
                        "message": "no processed historical entry",
                    }
                )

        if slug_failures:
            failures.extend(slug_failures)
        warnings.extend(slug_warnings)
        if not slug_failures:
            validated.append(
                {
                    "slug": slug,
                    "workbook_path": _display_path(workbook_path),
                    "sheet_count": len(sheetnames),
                    "historical_sheets": historical_sheet_names,
                }
            )

    payload = {
        "run_date": run_date,
        "workbooks_dir": str(workbooks_dir),
        "scope": args.director or "all",
        "status": "failed" if failures else "ok",
        "validated": validated,
        "failures": failures,
        "warnings": warnings,
    }
    _write_run_audit(output_dir, payload)
    print(f"Workbook contract audit: {_display_path(output_dir)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
