#!/usr/bin/env python3
"""Extract a normalized KPI catalog from the workbook provided by Ops."""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SECTION_RE = re.compile(r"^(?P<section_id>\d+(?:\.\d+)?)\s+(?P<section_name>.+?)\s+-\s+KPIs$")

HEADER_ALIASES = {
    "Timeline": "timeline",
    "ID": "kpi_id",
    "Initiative/KPI": "kpi_name",
    "Target": "target",
    "Definition": "definition",
    "Why it Matters": "why_it_matters",
    "Salesforce calculation": "salesforce_calculation",
    "Technical Salesforce calc (proposed framework)": "technical_salesforce_calc",
    "SF Link": "sf_link",
    "Impact": "impact",
    "Primary Bucket": "primary_bucket",
}

CANONICAL_SHEETS = ["Master", "RW - KPIs", "AP - KPIs"]

SECTION_TO_DASHBOARDS = {
    "1.1": ["Lead Funnel", "Lead Management KPIs", "BDR Manager", "BDR Rep Queue"],
    "1.2": ["Account Intelligence KPIs", "Customer Intelligence"],
    "1.3": ["Customer Intelligence"],
    "1.4": ["Executive Revenue Source Truth", "Forecast & Revenue Motions"],
    "1.5": ["Contract Operations"],
    "1.6": ["Contract Operations Renewals"],
    "1.7": ["Customer Intelligence", "Revenue Retention & Health"],
    "1.8": ["Revenue Retention & Health"],
    "1.9": ["Forecast & Revenue Motions"],
    "1.10": ["Revenue Retention & Health"],
    "1.14": ["Executive Product Mix & Industry"],
}


@dataclass
class SectionInfo:
    section_id: str
    section_name: str
    section_title: str


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    return str(value)


def _parse_section(text: str) -> SectionInfo | None:
    match = SECTION_RE.match(_clean(text))
    if not match:
        return None
    section_id = match.group("section_id")
    section_name = match.group("section_name")
    return SectionInfo(
        section_id=section_id,
        section_name=section_name,
        section_title=_clean(text),
    )


def _header_map(row_values: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, value in enumerate(row_values):
        key = HEADER_ALIASES.get(_clean(value))
        if key:
            result[key] = idx
    return result


def _extract_sheet(formula_ws, values_ws, sheet_name: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    current_section: SectionInfo | None = None
    current_header: dict[str, int] = {}

    for row_idx, row in enumerate(formula_ws.iter_rows(), start=1):
        formula_values = [cell.value for cell in row]
        cached_values = [cell.value for cell in values_ws[row_idx]]
        col_a = _clean(formula_values[0] if formula_values else "")

        section = _parse_section(col_a)
        if section:
            current_section = section
            current_header = {}
            continue

        if "Initiative/KPI" in [_clean(v) for v in formula_values]:
            current_header = _header_map([_clean(v) for v in formula_values])
            continue

        if not current_section or not current_header:
            continue

        kpi_idx = current_header.get("kpi_name")
        if kpi_idx is None:
            continue

        kpi_name = _clean(formula_values[kpi_idx] if kpi_idx < len(formula_values) else "")
        if not kpi_name or kpi_name == "Initiative/KPI":
            continue

        record: dict[str, Any] = {
            "source_sheet": sheet_name,
            "source_row": row_idx,
            "section_id": current_section.section_id,
            "section_name": current_section.section_name,
            "section_title": current_section.section_title,
            "dashboard_candidates": SECTION_TO_DASHBOARDS.get(current_section.section_id, []),
        }
        formulas: dict[str, str] = {}

        for key, idx in current_header.items():
            formula_value = formula_values[idx] if idx < len(formula_values) else None
            cached_value = cached_values[idx] if idx < len(cached_values) else None
            record[key] = _clean(cached_value if cached_value not in (None, "") else formula_value)
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formulas[key] = formula_value

        if formulas:
            record["formula_cells"] = formulas
        records.append(record)

    return records


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "source_sheet",
        "source_row",
        "section_id",
        "section_name",
        "section_title",
        "timeline",
        "kpi_id",
        "kpi_name",
        "target",
        "definition",
        "why_it_matters",
        "salesforce_calculation",
        "technical_salesforce_calc",
        "sf_link",
        "impact",
        "primary_bucket",
        "dashboard_candidates",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            serializable = dict(row)
            serializable["dashboard_candidates"] = " | ".join(row.get("dashboard_candidates", []))
            writer.writerow({key: serializable.get(key, "") for key in fieldnames})


def _section_summary(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = row["section_id"]
        bucket = summary.setdefault(
            key,
            {
                "section_title": row["section_title"],
                "count": 0,
                "high_impact": [],
                "dashboard_candidates": row.get("dashboard_candidates", []),
            },
        )
        bucket["count"] += 1
        if row.get("impact") == "HIGH" and len(bucket["high_impact"]) < 8:
            bucket["high_impact"].append(row["kpi_name"])
    return summary


def _write_markdown(path: Path, workbook_path: Path, all_rows: list[dict[str, Any]]) -> None:
    master_rows = [row for row in all_rows if row["source_sheet"] == "Master"]
    summary = _section_summary(master_rows)
    lines = [
        f"# KPI Workbook Summary ({date.today().isoformat()})",
        "",
        f"- Source workbook: `{workbook_path}`",
        f"- Canonical KPI rows extracted from `Master`: `{len(master_rows)}`",
        f"- Supporting sheet rows extracted: `{len(all_rows) - len(master_rows)}`",
        "",
        "## Section Summary",
        "",
        "| Section | KPI Rows | Dashboard Candidates |",
        "| --- | ---: | --- |",
    ]
    for key in sorted(summary, key=lambda value: [int(part) for part in value.split(".")]):
        bucket = summary[key]
        dashboards = ", ".join(bucket["dashboard_candidates"]) or "-"
        lines.append(f"| {bucket['section_title']} | {bucket['count']} | {dashboards} |")

    lines.extend(["", "## High-Impact KPI Samples", ""])
    for key in sorted(summary, key=lambda value: [int(part) for part in value.split(".")]):
        bucket = summary[key]
        lines.append(f"### {bucket['section_title']}")
        if bucket["high_impact"]:
            for kpi_name in bucket["high_impact"]:
                lines.append(f"- {kpi_name}")
        else:
            lines.append("- No `HIGH` rows identified in the extracted master catalog.")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the KPI workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/spreadsheet/kpi_catalog"),
        help="Directory for normalized KPI exports",
    )
    parser.add_argument(
        "--docs-path",
        type=Path,
        default=Path(f"docs/generated/KPI_WORKBOOK_SUMMARY_{date.today().isoformat()}.md"),
        help="Markdown summary output path",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    output_dir = args.output_dir.resolve()
    docs_path = args.docs_path.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    docs_path.parent.mkdir(parents=True, exist_ok=True)

    wb_formula = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)

    all_rows: list[dict[str, Any]] = []
    for sheet_name in CANONICAL_SHEETS:
        all_rows.extend(_extract_sheet(wb_formula[sheet_name], wb_values[sheet_name], sheet_name))

    catalog_json = output_dir / "kpi_catalog.json"
    catalog_csv = output_dir / "kpi_catalog.csv"
    summary_json = output_dir / "kpi_section_summary.json"

    catalog_json.write_text(json.dumps(all_rows, indent=2), encoding="utf-8")
    _write_csv(catalog_csv, all_rows)
    summary_json.write_text(json.dumps(_section_summary([r for r in all_rows if r["source_sheet"] == "Master"]), indent=2), encoding="utf-8")
    _write_markdown(docs_path, workbook_path, all_rows)

    print(f"Extracted {len(all_rows)} KPI rows")
    print(f"  JSON: {catalog_json}")
    print(f"  CSV:  {catalog_csv}")
    print(f"  MD:   {docs_path}")


if __name__ == "__main__":
    main()
